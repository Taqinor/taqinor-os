"""NTDMO22/23/24 — kit de démonstration (PDF/HTML/XLSX), jamais le moteur
premium `/proposal` (rule #4 CLAUDE.md), réservé aux sociétés `est_demo`."""
from unittest.mock import patch

from django.core.management import call_command
from django.test import TestCase, override_settings
from rest_framework.test import APIClient

from authentication.models import Company, CustomUser


@override_settings(DEBUG=True)
class DemoKitTest(TestCase):
    SLUG = 'taqinor-demo-full'

    @classmethod
    def setUpTestData(cls):
        call_command('seed_demo_company', verbosity=0)
        cls.demo = Company.objects.get(slug=cls.SLUG)
        cls.admin = CustomUser.objects.get(username='demo_admin_full')

    def setUp(self):
        self.client = APIClient()
        self.client.force_authenticate(self.admin)

    def test_forbidden_on_non_demo_company(self):
        real = Company.objects.create(nom='Réelle kit', slug='reelle-kit')
        resp = self.client.get(f'/api/django/companies/{real.id}/demo-kit/')
        self.assertEqual(resp.status_code, 403)
        resp = self.client.get(
            f'/api/django/companies/{real.id}/demo-kit/export.xlsx')
        self.assertEqual(resp.status_code, 403)

    def test_html_variant_lists_six_screens(self):
        resp = self.client.get(
            f'/api/django/companies/{self.demo.id}/demo-kit/?format=html')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'text/html; charset=utf-8')
        content = resp.content.decode('utf-8')
        for route in ('/ventes/devis/nouveau', '/crm/leads',
                      '/ventes/factures', '/chantiers', '/stock',
                      '/dashboard'):
            self.assertIn(route, content)

    def test_pdf_variant_generates_bytes(self):
        with patch('core.pdf._html_to_pdf_bytes',
                   return_value=b'%PDF-1.4 fake'):
            resp = self.client.get(
                f'/api/django/companies/{self.demo.id}/demo-kit/')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/pdf')
        self.assertTrue(resp.content.startswith(b'%PDF'))

    def test_xlsx_export_has_four_sheets(self):
        from openpyxl import load_workbook
        from io import BytesIO
        resp = self.client.get(
            f'/api/django/companies/{self.demo.id}/demo-kit/export.xlsx')
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(BytesIO(resp.content))
        self.assertEqual(
            set(wb.sheetnames), {'Leads', 'Devis', 'Chantiers', 'Factures'})
        for name in wb.sheetnames:
            ws = wb[name]
            self.assertGreater(ws.max_row, 1, f'{name} est vide')
