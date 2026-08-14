"""NTLOG31 — relevé mensuel des litiges transport (PDF/xlsx au choix),
filtrable `?periode=YYYY-MM`/`?transporteur=`. Le total « montant contesté »
des litiges OUVERTS du relevé correspond au total
`litiges_ouverts_montant_conteste` du dashboard NTLOG24 pour la même
période (critère d'acceptation)."""
import io
from datetime import datetime
from decimal import Decimal

from django.test import TestCase
from django.utils import timezone

from apps.transport.models import LitigeTransport, OrdreTransport

from ._helpers import auth, make_company, make_user

BASE = '/api/django/transport/litiges-transport/releve/'
DASHBOARD = '/api/django/transport/ordres-transport/tableau-bord-logistique/'


class ReleveLitigesTests(TestCase):
    def setUp(self):
        self.co_a = make_company('transport-rl-a', 'A')
        self.co_b = make_company('transport-rl-b', 'B')
        self.user_a = make_user(self.co_a, 'transport-rl-a')
        self.ordre = OrdreTransport.objects.create(
            company=self.co_a, numero='OT-RL-1')

    def _litige(self, montant, statut, *, created, montant_resolu=None):
        litige = LitigeTransport.objects.create(
            company=self.co_a, ordre_transport=self.ordre,
            statut=statut, montant_conteste=Decimal(montant),
            montant_resolu=(
                Decimal(montant_resolu) if montant_resolu is not None else None))
        litige.created_at = timezone.make_aware(datetime.fromisoformat(created))
        litige.save(update_fields=['created_at'])
        return litige

    def test_xlsx_entetes_et_lignes(self):
        import openpyxl

        self._litige(
            '500.00', LitigeTransport.Statut.OUVERT, created='2026-06-05T10:00:00')
        resp = auth(self.user_a).get(BASE)
        self.assertEqual(resp.status_code, 200)
        self.assertIn('spreadsheetml', resp['Content-Type'])
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        self.assertIn('Montant contesté', header)
        self.assertIn('Montant résolu', header)
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][0], 'OT-RL-1')

    def test_pdf_genere(self):
        self._litige(
            '500.00', LitigeTransport.Statut.OUVERT, created='2026-06-05T10:00:00')
        resp = auth(self.user_a).get(BASE, {'format': 'pdf'})
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp['Content-Type'], 'application/pdf')
        self.assertTrue(bytes(resp.content).startswith(b'%PDF'))

    def test_format_invalide_400(self):
        resp = auth(self.user_a).get(BASE, {'format': 'docx'})
        self.assertEqual(resp.status_code, 400)

    def test_total_conteste_ouverts_egale_dashboard_meme_periode(self):
        import openpyxl

        self._litige(
            '400.00', LitigeTransport.Statut.OUVERT, created='2026-06-05T10:00:00')
        self._litige(
            '120.50', LitigeTransport.Statut.OUVERT, created='2026-06-20T10:00:00')
        # Résolu dans la même période — figure dans le relevé mais PAS dans
        # le total "ouverts" du dashboard NTLOG24.
        self._litige(
            '999.00', LitigeTransport.Statut.RESOLU, created='2026-06-10T10:00:00',
            montant_resolu='950.00')

        resp_releve = auth(self.user_a).get(BASE, {'periode': '2026-06'})
        wb = openpyxl.load_workbook(io.BytesIO(resp_releve.content))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 3)  # les 3 litiges de la période, tous statuts.
        total_ouverts_releve = sum(
            Decimal(str(row[4])) for row in rows if row[3] == 'Ouvert')

        resp_dash = auth(self.user_a).get(DASHBOARD, {'periode': '2026-06'})
        total_dashboard = Decimal(str(resp_dash.data['litiges_ouverts_montant_conteste']))

        self.assertEqual(total_ouverts_releve, Decimal('520.50'))
        self.assertEqual(total_ouverts_releve, total_dashboard)

    def test_isolation_societe(self):
        import openpyxl

        autre_ordre = OrdreTransport.objects.create(company=self.co_b)
        LitigeTransport.objects.create(
            company=self.co_b, ordre_transport=autre_ordre,
            montant_conteste=Decimal('10.00'))
        resp = auth(self.user_a).get(BASE)
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 0)


class ResoudreMontantResoluTests(TestCase):
    def setUp(self):
        self.co_a = make_company('transport-rlr-a', 'A')
        self.user_a = make_user(self.co_a, 'transport-rlr-a')
        self.ordre = OrdreTransport.objects.create(company=self.co_a)
        self.litige = LitigeTransport.objects.create(
            company=self.co_a, ordre_transport=self.ordre,
            statut=LitigeTransport.Statut.EN_TRAITEMENT,
            montant_conteste=Decimal('1000.00'))

    def test_resoudre_avec_montant_resolu_optionnel(self):
        resp = auth(self.user_a).post(
            f'/api/django/transport/litiges-transport/{self.litige.id}/resoudre/',
            {'montant_resolu': '850.00'}, format='json')
        self.assertEqual(resp.status_code, 200, resp.data)
        self.litige.refresh_from_db()
        self.assertEqual(self.litige.montant_resolu, Decimal('850.00'))

    def test_resoudre_sans_montant_resolu_reste_none(self):
        resp = auth(self.user_a).post(
            f'/api/django/transport/litiges-transport/{self.litige.id}/resoudre/')
        self.assertEqual(resp.status_code, 200, resp.data)
        self.litige.refresh_from_db()
        self.assertIsNone(self.litige.montant_resolu)
