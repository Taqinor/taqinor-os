"""NTLOG27 — export comptable des coûts de fret (.xlsx), filtrable par
`?periode=YYYY-MM`. Le total exporté doit correspondre EXACTEMENT au total
`total_fret_ht` du dashboard NTLOG24 pour la même période (critère
d'acceptation) — les deux filtrent sur `created_at` via le même helper
`selectors._filtre_periode`."""
import io
from datetime import datetime
from decimal import Decimal

from django.test import TestCase
from django.utils import timezone

from apps.transport.models import CoutFretReel, OrdreTransport

from ._helpers import auth, make_company, make_user

BASE = '/api/django/transport/couts-fret/export/'
DASHBOARD = '/api/django/transport/ordres-transport/tableau-bord-logistique/'


class ExportCoutsFretTests(TestCase):
    def setUp(self):
        self.co_a = make_company('transport-ecf-a', 'A')
        self.co_b = make_company('transport-ecf-b', 'B')
        self.user_a = make_user(self.co_a, 'transport-ecf-a')
        self.ordre = OrdreTransport.objects.create(
            company=self.co_a, numero='OT-ECF-1')

    def _cout(self, montant, *, created):
        c = CoutFretReel.objects.create(
            company=self.co_a, ordre_transport=self.ordre,
            montant_ht=Decimal(montant))
        c.created_at = timezone.make_aware(datetime.fromisoformat(created))
        c.save(update_fields=['created_at'])
        return c

    def test_xlsx_entetes_et_ligne(self):
        import openpyxl

        self._cout('300.00', created='2026-06-05T10:00:00')
        resp = auth(self.user_a).get(BASE)
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertIn('spreadsheetml', resp['Content-Type'])
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        self.assertIn('Montant HT', header)
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][0], 'OT-ECF-1')

    def test_total_exporte_egale_total_dashboard_meme_periode(self):
        import openpyxl

        self._cout('300.00', created='2026-06-05T10:00:00')
        self._cout('120.50', created='2026-06-20T10:00:00')
        # Hors période — ne doit compter ni dans l'export ni dans le dashboard.
        self._cout('999.00', created='2026-07-01T10:00:00')

        resp_export = auth(self.user_a).get(BASE, {'periode': '2026-06'})
        wb = openpyxl.load_workbook(io.BytesIO(resp_export.content))
        ws = wb.active
        total_export = sum(
            Decimal(str(row[3])) for row in ws.iter_rows(min_row=2, values_only=True))

        resp_dash = auth(self.user_a).get(DASHBOARD, {'periode': '2026-06'})
        total_dashboard = Decimal(str(resp_dash.data['total_fret_ht']))

        self.assertEqual(total_export, Decimal('420.50'))
        self.assertEqual(total_export, total_dashboard)

    def test_isolation_societe(self):
        autre_ordre = OrdreTransport.objects.create(company=self.co_b)
        CoutFretReel.objects.create(
            company=self.co_b, ordre_transport=autre_ordre,
            montant_ht=Decimal('50.00'))
        resp = auth(self.user_a).get(BASE)
        wb_bytes = io.BytesIO(resp.content)
        import openpyxl
        wb = openpyxl.load_workbook(wb_bytes)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 0)
