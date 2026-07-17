"""NTFPA25 — export_synthese_fpa : produit un classeur .xlsx multi-onglets
(≥ 3 onglets) exploitable directement, sans retouche manuelle."""
import io
from datetime import date
from decimal import Decimal

from django.test import TestCase

from authentication.models import Company
from apps.fpa.models import (
    Categorie, CycleBudgetaire, Departement, LigneBudgetDepartement,
)
from apps.fpa.selectors import export_synthese_fpa


class TestExportSyntheseFpa(TestCase):
    def setUp(self):
        self.company, _ = Company.objects.get_or_create(
            slug='ntfpa25-co', defaults={'nom': 'NTFPA25 Co'})
        self.cycle = CycleBudgetaire.objects.create(
            company=self.company, nom='Budget 2027',
            date_debut=date(2027, 1, 1), date_fin=date(2027, 12, 31))
        self.dept = Departement.objects.create(
            company=self.company, code='IT', nom='IT')
        LigneBudgetDepartement.objects.create(
            company=self.company, cycle=self.cycle, departement=self.dept,
            categorie=Categorie.IT, mois=1, montant_prevu=Decimal('3000'))

    def test_export_produit_un_classeur_multi_onglets(self):
        contenu = export_synthese_fpa(self.company, self.cycle)
        self.assertIsInstance(contenu, bytes)
        self.assertGreater(len(contenu), 0)

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(contenu))
        self.assertGreaterEqual(len(wb.sheetnames), 3)
        self.assertIn('P&L prévisionnel', wb.sheetnames)
        self.assertIn('Variance', wb.sheetnames)
        self.assertIn('Scénarios', wb.sheetnames)

    def test_cycle_inconnu_retourne_none(self):
        self.assertIsNone(export_synthese_fpa(self.company, 999999))
