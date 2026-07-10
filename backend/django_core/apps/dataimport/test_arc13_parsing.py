"""ARC13 — tests unitaires du parseur générique ``apps.dataimport.parsing``.

Couvre : CSV (séparateur , et ;), encodages (utf-8 BOM, latin-1), en-têtes
(normalisation accents/casse/espaces), XLSX, robustesse (lignes vides,
en-tête manquant/dupliqué).
"""
import io

from django.test import SimpleTestCase

from .parsing import iter_rows, normalize_header


class NormalizeHeaderTests(SimpleTestCase):
    def test_lowercases_and_strips_accents(self):
        self.assertEqual(normalize_header('Prénom'), 'prenom')
        self.assertEqual(normalize_header('Numéro Série'), 'numero_serie')

    def test_spaces_and_dashes_become_underscore(self):
        self.assertEqual(normalize_header('Date de Début'), 'date_de_debut')
        self.assertEqual(normalize_header('id-externe'), 'id_externe')

    def test_none_is_empty_string(self):
        self.assertEqual(normalize_header(None), '')

    def test_already_normalized_is_unchanged(self):
        self.assertEqual(normalize_header('email'), 'email')


class IterRowsCsvTests(SimpleTestCase):
    def test_comma_delimiter(self):
        content = 'Nom,Email\nAlaoui,a@x.ma\nBennani,b@x.ma\n'
        headers, rows = iter_rows(content.encode('utf-8'), 'data.csv')
        self.assertEqual(headers, ['Nom', 'Email'])
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]['Nom'], 'Alaoui')
        self.assertEqual(rows[0]['Email'], 'a@x.ma')

    def test_semicolon_delimiter_autodetected(self):
        content = 'Nom;Email\nAlaoui;a@x.ma\n'
        headers, rows = iter_rows(content.encode('utf-8'), 'data.csv')
        self.assertEqual(headers, ['Nom', 'Email'])
        self.assertEqual(rows[0]['Nom'], 'Alaoui')

    def test_utf8_bom_stripped(self):
        content = '﻿Nom,Email\nAlaoui,a@x.ma\n'
        file_bytes = content.encode('utf-8-sig')
        headers, rows = iter_rows(file_bytes, 'data.csv')
        # Le BOM ne doit pas polluer le premier en-tête.
        self.assertEqual(headers, ['Nom', 'Email'])

    def test_latin1_fallback_on_invalid_utf8(self):
        # 'é' encodé en latin-1 (0xe9) n'est pas de l'utf-8 valide.
        content = 'Nom,Ville\nAlaoui,Créteil\n'
        file_bytes = content.encode('latin-1')
        headers, rows = iter_rows(file_bytes, 'data.csv')
        self.assertEqual(rows[0]['Ville'], 'Créteil')

    def test_blank_lines_skipped_but_comma_only_row_kept(self):
        # ``DictReader`` saute la ligne STRICTEMENT vide (le ``\n`` final), mais
        # la ligne ``,`` (valeurs vides mais séparateur tapé) est CONSERVÉE : la
        # couche appelante (import XPLT2) doit la voir pour la signaler « ligne
        # vide » au lieu de perdre silencieusement la donnée.
        content = 'Nom,Email\nAlaoui,a@x.ma\n,\n\n'
        headers, rows = iter_rows(content.encode('utf-8'), 'data.csv')
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1], {'Nom': '', 'Email': ''})

    def test_extension_case_insensitive_defaults_to_csv(self):
        content = 'Nom\nAlaoui\n'
        headers, rows = iter_rows(content.encode('utf-8'), 'DATA.CSV')
        self.assertEqual(headers, ['Nom'])
        self.assertEqual(len(rows), 1)


class IterRowsXlsxTests(SimpleTestCase):
    def _make_xlsx(self, headers, data_rows):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        for row in data_rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_reads_headers_and_rows(self):
        file_bytes = self._make_xlsx(
            ['Nom', 'Email'], [['Alaoui', 'a@x.ma'], ['Bennani', 'b@x.ma']])
        headers, rows = iter_rows(file_bytes, 'data.xlsx')
        self.assertEqual(headers, ['Nom', 'Email'])
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]['Nom'], 'Alaoui')
        self.assertEqual(rows[1]['Email'], 'b@x.ma')

    def test_uppercase_extension_detected(self):
        file_bytes = self._make_xlsx(['Nom'], [['Alaoui']])
        headers, rows = iter_rows(file_bytes, 'DATA.XLSX')
        self.assertEqual(headers, ['Nom'])
        self.assertEqual(len(rows), 1)

    def test_fully_empty_row_skipped(self):
        file_bytes = self._make_xlsx(
            ['Nom', 'Email'], [['Alaoui', 'a@x.ma'], [None, None]])
        headers, rows = iter_rows(file_bytes, 'data.xlsx')
        self.assertEqual(len(rows), 1)

    def test_missing_trailing_cells_do_not_crash(self):
        file_bytes = self._make_xlsx(
            ['Nom', 'Email', 'Ville'], [['Alaoui', 'a@x.ma']])
        headers, rows = iter_rows(file_bytes, 'data.xlsx')
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].get('Nom'), 'Alaoui')
        self.assertIsNone(rows[0].get('Ville'))
