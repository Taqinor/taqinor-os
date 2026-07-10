"""ARC13 — parseur d'import générique CSV/XLSX, partagé.

Constat (plan ARC13) : 31 fichiers hors ``dataimport`` re-codent leur propre
lecture ``csv.reader``/``csv.DictReader``/``openpyxl`` (détection d'encodage,
séparateur, en-têtes...) alors que ``apps.dataimport.services.parse_rows``
fait déjà ce travail pour les 6 cibles historiques (leads/clients/products/
fournisseurs/equipements/vehicules).

``iter_rows(file_bytes, filename)`` extrait cette logique dans une fonction
UNIQUE et réutilisable (CSV + XLSX, détection d'encodage utf-8/latin-1,
séparateur ``,``/``;``, en-têtes normalisés) à la fois pour les 6 cibles
``FIELD_MAPS`` existantes (``services.parse_rows`` délègue désormais ici, zéro
changement de comportement) et pour tout call-site hors ``dataimport`` qui
veut lire un fichier tabulaire sans dupliquer la lecture bas niveau
(ex. ``apps.rh.views``/``apps.contrats.views``, motif de migration
progressive — 3 pilotes migrés dans ce lot, les autres suivront).

``records/xlsx.py`` (export) n'est PAS concerné : ce module ne touche que la
LECTURE (import), jamais l'export.
"""
import csv
import io
import unicodedata


def normalize_header(s):
    """Normalise un en-tête : minuscules, sans accents, espaces/tirets → underscore.

    Robuste à ``None`` (en-tête vide dans un fichier mal formé).
    """
    s = (s or '').strip().lower()
    s = ''.join(c for c in unicodedata.normalize('NFD', s)
                if unicodedata.category(c) != 'Mn')
    return s.replace(' ', '_').replace('-', '_')


def _decode_csv_text(file_bytes):
    """Décode des octets CSV en texte : utf-8 (avec BOM) en priorité, repli
    latin-1 si le fichier n'est pas de l'utf-8 valide (fichiers exportés
    depuis Excel/Windows, courants côté clients).

    ``utf-8-sig`` retire le BOM d'encodage en tête ; on retire EN PLUS tout
    caractère BOM/ZWNBSP (U+FEFF) résiduel en début de texte — un fichier peut
    porter un BOM littéral au-delà de celui consommé par l'encodage (ou un BOM
    non consommé sur le chemin latin-1), qui polluerait sinon le premier
    en-tête (``'<BOM>Nom'`` au lieu de ``'Nom'``)."""
    try:
        text = file_bytes.decode('utf-8-sig')
    except UnicodeDecodeError:
        text = file_bytes.decode('latin-1')
    return text.lstrip('\ufeff')


def _iter_csv_rows(file_bytes):
    text = _decode_csv_text(file_bytes)
    sample = text[:2000]
    delim = ';' if sample.count(';') > sample.count(',') else ','
    reader = csv.DictReader(io.StringIO(text), delimiter=delim)
    headers = reader.fieldnames or []
    # ``csv.DictReader`` saute déjà les lignes STRICTEMENT vides (``[]`` — un
    # ``\n`` isolé). On NE filtre PAS en plus les lignes à valeurs vides (``,``) :
    # une ligne dont l'utilisateur a tapé les séparateurs est une DONNÉE (souvent
    # une erreur de saisie) que la couche appelante doit VOIR et signaler — le
    # flux d'import XPLT2 la compte en « ligne vide » plutôt que de la perdre en
    # silence. Comportement identique à l'ancien ``parse_rows`` (``list(reader)``).
    return headers, [dict(r) for r in reader]


def _iter_xlsx_rows(file_bytes):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    raw_headers = next(it, [])
    headers = [str(h) if h is not None else '' for h in raw_headers]
    rows = []
    for r in it:
        if r is None or all(v is None for v in r):
            continue
        row = {headers[i]: r[i] for i in range(len(headers)) if i < len(r)}
        rows.append(row)
    return headers, rows


def iter_rows(file_bytes, filename):
    """Lit un fichier CSV ou XLSX et renvoie ``(headers, rows)``.

    ``headers`` : liste des en-têtes BRUTS (ordre du fichier, non normalisés).
    ``rows`` : liste de dicts ``{en-tête brut: valeur}``, lignes vides omises.

    Détection du format par extension (``.xlsx`` → openpyxl lecture seule ;
    sinon CSV). Le CSV détecte automatiquement le séparateur (``,`` ou ``;``)
    et l'encodage (utf-8 avec BOM, repli latin-1). Robuste aux en-têtes
    manquants/dupliqués et aux lignes vides.
    """
    name = (filename or '').lower()
    if name.endswith('.xlsx'):
        return _iter_xlsx_rows(file_bytes)
    return _iter_csv_rows(file_bytes)
