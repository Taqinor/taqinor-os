"""T9 — import réutilisable CSV/XLSX (leads, clients, produits).

Flux en deux temps, multi-tenant :
  1. dry-run : on lit les 10 premières lignes, on mappe colonne → champ (par
     en-tête, insensible à la casse/accents), et on liste ce qui n'a PAS été
     mappé — pour validation AVANT le batch complet.
  2. commit : création UNIQUEMENT (jamais d'écrasement silencieux). Les doublons
     (email/téléphone pour leads/clients, SKU pour produits) sont signalés et
     ignorés. Les enregistrements créés sont marqués d'origine (import).

Séparé de la migration ponctuelle des 619 leads Odoo (gardée à part).
"""
import csv
import io
import unicodedata

from django.db import transaction

# Mapping en-tête (normalisé) → champ modèle, par cible.
FIELD_MAPS = {
    'leads': {
        'nom': 'nom', 'prenom': 'prenom', 'societe': 'societe',
        'email': 'email', 'telephone': 'telephone', 'tel': 'telephone',
        'ville': 'ville', 'whatsapp': 'whatsapp', 'adresse': 'adresse',
        # XPLT1 — identifiant externe optionnel (rapprochement upsert/maj).
        'external_id': 'external_id', 'id_externe': 'external_id',
    },
    'clients': {
        'nom': 'nom', 'prenom': 'prenom', 'email': 'email',
        'telephone': 'telephone', 'tel': 'telephone', 'adresse': 'adresse',
        'ice': 'ice',
        'external_id': 'external_id', 'id_externe': 'external_id',
    },
    'products': {
        'nom': 'nom', 'sku': 'sku', 'reference': 'sku', 'marque': 'marque',
        'prix_vente': 'prix_vente', 'prix': 'prix_vente',
        'prix_achat': 'prix_achat', 'quantite': 'quantite_stock',
        'quantite_stock': 'quantite_stock', 'stock': 'quantite_stock',
        'description': 'description',
    },
    # FG14 — Fournisseurs : import texte simple, pas de relation.
    'fournisseurs': {
        'nom': 'nom', 'contact': 'contact_personne',
        'contact_personne': 'contact_personne', 'email': 'email',
        'telephone': 'telephone', 'tel': 'telephone',
        'adresse': 'adresse',
    },
    # FG14 — Équipements : import avec résolution produit (par SKU) et
    # installation (par référence). Seuls les champs libres sont importables
    # directement ; produit/installation sont résolus côté commit().
    'equipements': {
        'numero_serie': 'numero_serie', 'serie': 'numero_serie',
        'sn': 'numero_serie', 'statut': 'statut', 'note': 'note',
        'produit_sku': 'produit_sku', 'sku': 'produit_sku',
        'installation_ref': 'installation_ref',
        'chantier': 'installation_ref', 'installation': 'installation_ref',
        'date_pose': 'date_pose',
    },
    # XFLT22 — Import initial du parc flotte. Écriture DÉLÉGUÉE à
    # ``apps.flotte.services.creer_vehicule_import`` (jamais les models
    # flotte directement, contrairement aux autres cibles ci-dessus —
    # règle explicite du plan flotte).
    'vehicules': {
        'immatriculation': 'immatriculation', 'immat': 'immatriculation',
        'marque': 'marque', 'modele': 'modele', 'modèle': 'modele',
        'energie': 'energie', 'énergie': 'energie',
        'kilometrage': 'kilometrage', 'km': 'kilometrage',
        'cv': 'cv', 'puissance_fiscale': 'cv',
    },
}

TARGETS = set(FIELD_MAPS)

# ERR53 — Plafond de lignes : au-delà, on refuse proprement (ValueError → 400
# clair côté vue) plutôt que de charger un fichier géant en mémoire et risquer
# un OOM. Doit rester aligné avec `views.MAX_ROWS`.
MAX_ROWS = 10000


class ImportTooLarge(ValueError):
    """Le fichier dépasse le plafond de lignes autorisé (ERR53)."""


def _norm(s):
    """Normalise un en-tête : minuscules, sans accents, espaces → underscore."""
    s = (s or '').strip().lower()
    s = ''.join(c for c in unicodedata.normalize('NFD', s)
                if unicodedata.category(c) != 'Mn')
    return s.replace(' ', '_').replace('-', '_')


def parse_rows(file_bytes, filename):
    """Renvoie (headers, rows[list[dict]]) depuis un CSV ou XLSX."""
    name = (filename or '').lower()
    if name.endswith('.xlsx'):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        headers = [str(h) if h is not None else '' for h in next(it, [])]
        rows = []
        for r in it:
            rows.append({headers[i]: r[i] for i in range(len(headers)) if i < len(r)})
        return headers, rows
    # CSV (utf-8, séparateur , ou ;)
    text = file_bytes.decode('utf-8-sig', errors='replace')
    sample = text[:2000]
    delim = ';' if sample.count(';') > sample.count(',') else ','
    reader = csv.DictReader(io.StringIO(text), delimiter=delim)
    headers = reader.fieldnames or []
    return headers, list(reader)


def _map_headers(headers, target):
    fmap = FIELD_MAPS[target]
    mapped, unmapped = {}, []
    for h in headers:
        field = fmap.get(_norm(h))
        if field:
            mapped[h] = field
        else:
            unmapped.append(h)
    return mapped, unmapped


def dry_run(file_bytes, filename, target):
    """Aperçu : mapping colonne→champ + 10 premières lignes mappées + non-mappés."""
    if target not in TARGETS:
        raise ValueError("Cible d'import inconnue.")
    headers, rows = parse_rows(file_bytes, filename)
    if len(rows) > MAX_ROWS:
        raise ImportTooLarge(
            f'Trop de lignes : {len(rows)} (max {MAX_ROWS}).')
    mapped, unmapped = _map_headers(headers, target)
    preview = []
    for row in rows[:10]:
        preview.append({field: row.get(col) for col, field in mapped.items()})
    return {
        'target': target,
        'colonnes': headers,
        'mapping': mapped,
        'non_mappees': unmapped,
        'apercu': preview,
        'total_lignes': len(rows),
    }


def _row_to_fields(row, mapped):
    return {field: row.get(col) for col, field in mapped.items()
            if row.get(col) not in (None, '')}


# XPLT1 — modes de commit. ``creer`` (défaut) reproduit exactement le
# comportement historique (création seule, doublons ignorés). ``maj``/
# ``upsert`` rapprochent d'abord par identifiant externe (ExternalRef) puis par
# contact normalisé (réutilise ``find_duplicates_by_contact``) : en cas de
# correspondance, seuls les champs FOURNIS par la ligne sont mis à jour
# (jamais d'écrasement silencieux par une valeur absente) ; ``maj`` n'importe
# JAMAIS de nouvelle fiche (une ligne sans correspondance est ignorée),
# ``upsert`` crée si aucune correspondance n'est trouvée.
MODES = {'creer', 'maj', 'upsert'}

DEFAULT_EXTERNAL_SYSTEM = 'import'


def _get_or_create_ref(company, external_system, external_id, obj):
    from .models import ExternalRef
    from django.contrib.contenttypes.models import ContentType
    ct = ContentType.objects.get_for_model(obj)
    ExternalRef.objects.get_or_create(
        company=company, external_system=external_system,
        external_id=external_id,
        defaults={'content_type': ct, 'object_id': obj.pk})


def _find_by_external_id(company, external_system, external_id, model):
    from .models import ExternalRef
    if not external_id:
        return None
    ref = ExternalRef.objects.filter(
        company=company, external_system=external_system,
        external_id=str(external_id)).first()
    if ref is None:
        return None
    return model.objects.filter(company=company, pk=ref.object_id).first()


def _apply_updates(instance, fields, skip_keys=()):
    """Met à jour uniquement les champs FOURNIS (non vides) — jamais d'écrasement
    par une valeur absente de la ligne importée."""
    changed = []
    for key, value in fields.items():
        if key in skip_keys:
            continue
        if not hasattr(instance, key):
            continue
        if value in (None, ''):
            continue
        if getattr(instance, key, None) != value:
            setattr(instance, key, value)
            changed.append(key)
    if changed:
        instance.save(update_fields=changed)
    return changed


def commit(file_bytes, filename, target, company, user, mode='creer',
           external_system=None):
    """Crée (mode=creer, défaut inchangé) ou rapproche+met à jour (maj/upsert)
    les enregistrements. Renvoie un récapitulatif."""
    if target not in TARGETS:
        raise ValueError("Cible d'import inconnue.")
    if mode not in MODES:
        raise ValueError("Mode d'import inconnu (creer, maj ou upsert).")
    # XPLT1 — le rapprochement maj/upsert n'est câblé que pour les cibles où un
    # contact (email/téléphone) permet un rapprochement fiable (leads, clients).
    # Les autres cibles gardent le comportement historique (création seule) et
    # refusent explicitement un mode qu'elles ne supportent pas encore, plutôt
    # que de l'ignorer silencieusement.
    if mode != 'creer' and target not in ('leads', 'clients'):
        raise ValueError(
            f"Le mode « {mode} » n'est pas supporté pour la cible « {target} » "
            "(seuls leads et clients supportent maj/upsert).")
    external_system = external_system or DEFAULT_EXTERNAL_SYSTEM
    headers, rows = parse_rows(file_bytes, filename)
    if len(rows) > MAX_ROWS:
        raise ImportTooLarge(
            f'Trop de lignes : {len(rows)} (max {MAX_ROWS}).')
    mapped, _ = _map_headers(headers, target)
    created, updated, skipped = 0, 0, []

    # ERR51 — Tout l'import est atomique : une erreur en cours de boucle annule
    # l'intégralité du lot (jamais de demi-import laissant le compteur perdu et
    # une partie des lignes déjà créées).
    with transaction.atomic():
        if target == 'leads':
            from apps.crm.models import Lead
            from apps.crm.services import find_duplicates_by_contact
            for i, row in enumerate(rows, 1):
                f = _row_to_fields(row, mapped)
                if not f.get('nom') and not f.get('email') and not f.get('telephone'):
                    skipped.append({'ligne': i, 'raison': 'ligne vide'})
                    continue
                ext_id = f.pop('external_id', None)

                existing = None
                if mode in ('maj', 'upsert'):
                    existing = _find_by_external_id(
                        company, external_system, ext_id, Lead)
                    if existing is None:
                        dupes = find_duplicates_by_contact(
                            company, phone=f.get('telephone'),
                            email=f.get('email'))
                        existing = dupes[0] if dupes else None

                if existing is not None:
                    _apply_updates(existing, f)
                    if ext_id:
                        _get_or_create_ref(
                            company, external_system, ext_id, existing)
                    updated += 1
                    continue

                if mode == 'maj':
                    skipped.append(
                        {'ligne': i, 'raison': 'aucune correspondance (maj seule)'})
                    continue

                # Création (mode=creer, ou mode=upsert sans correspondance).
                if mode == 'creer':
                    dup = Lead.objects.filter(company=company)
                    if f.get('email'):
                        dup = dup.filter(email__iexact=f['email'])
                    elif f.get('telephone'):
                        dup = dup.filter(telephone=f['telephone'])
                    else:
                        dup = Lead.objects.none()
                    if dup.exists():
                        skipped.append({'ligne': i, 'raison': 'doublon (existe déjà)'})
                        continue
                tags = (f.pop('tags', '') or '')
                f['tags'] = (tags + (', ' if tags else '') + 'Import').strip(', ')[:500]
                lead = Lead.objects.create(company=company, **f)
                if ext_id:
                    _get_or_create_ref(company, external_system, ext_id, lead)
                created += 1

        elif target == 'clients':
            from apps.crm.models import Client
            for i, row in enumerate(rows, 1):
                f = _row_to_fields(row, mapped)
                if not f.get('nom'):
                    skipped.append({'ligne': i, 'raison': 'nom manquant'})
                    continue
                ext_id = f.pop('external_id', None)

                existing = None
                if mode in ('maj', 'upsert'):
                    existing = _find_by_external_id(
                        company, external_system, ext_id, Client)
                    if existing is None and f.get('email'):
                        existing = Client.objects.filter(
                            company=company, email__iexact=f['email']).first()

                if existing is not None:
                    _apply_updates(existing, f)
                    if ext_id:
                        _get_or_create_ref(
                            company, external_system, ext_id, existing)
                    updated += 1
                    continue

                if mode == 'maj':
                    skipped.append(
                        {'ligne': i, 'raison': 'aucune correspondance (maj seule)'})
                    continue

                if mode == 'creer' and f.get('email') and Client.objects.filter(
                        company=company, email__iexact=f['email']).exists():
                    skipped.append({'ligne': i, 'raison': 'doublon (email existe)'})
                    continue
                client = Client.objects.create(company=company, **f)
                if ext_id:
                    _get_or_create_ref(company, external_system, ext_id, client)
                created += 1

        elif target == 'products':
            from decimal import Decimal, InvalidOperation
            from apps.stock.models import MouvementStock, Produit
            for i, row in enumerate(rows, 1):
                f = _row_to_fields(row, mapped)
                if not f.get('nom'):
                    skipped.append({'ligne': i, 'raison': 'nom manquant'})
                    continue
                if f.get('sku') and Produit.objects.filter(
                        company=company, sku=f['sku']).exists():
                    skipped.append({'ligne': i, 'raison': 'doublon (SKU existe)'})
                    continue
                for k in ('prix_vente', 'prix_achat'):
                    if k in f:
                        raw = (str(f[k]).replace('\xa0', '').replace(' ', '')
                               .replace(',', '.'))
                        try:
                            f[k] = Decimal(raw)
                        except (InvalidOperation, ValueError):
                            f.pop(k)
                # ERR52 — Le stock d'ouverture ne peut jamais être négatif et
                # passe par le registre des mouvements (audit) comme partout
                # ailleurs : on crée le produit à 0 puis on enregistre un
                # MouvementStock ENTREE pour la quantité importée.
                opening = 0
                if 'quantite_stock' in f:
                    try:
                        opening = int(float(f.pop('quantite_stock')))
                    except (ValueError, TypeError):
                        opening = 0
                    if opening < 0:
                        skipped.append(
                            {'ligne': i, 'raison': 'stock négatif refusé'})
                        continue
                f.setdefault('prix_vente', Decimal('0'))
                produit = Produit.objects.create(
                    company=company, quantite_stock=0, **f)
                if opening > 0:
                    produit.quantite_stock = opening
                    produit.save(update_fields=['quantite_stock'])
                    MouvementStock.objects.create(
                        company=company, produit=produit,
                        type_mouvement=MouvementStock.TypeMouvement.ENTREE,
                        quantite=opening, quantite_avant=0,
                        quantite_apres=opening, created_by=user,
                        note='Stock initial (import)')
                created += 1

        # FG14 — Fournisseurs.
        elif target == 'fournisseurs':
            from apps.stock.models import Fournisseur
            for i, row in enumerate(rows, 1):
                f = _row_to_fields(row, mapped)
                if not f.get('nom'):
                    skipped.append({'ligne': i, 'raison': 'nom manquant'})
                    continue
                if Fournisseur.objects.filter(
                        company=company, nom__iexact=f['nom']).exists():
                    skipped.append({'ligne': i, 'raison': 'doublon (nom existe)'})
                    continue
                Fournisseur.objects.create(company=company, **f)
                created += 1

        # FG14 — Équipements : résolution produit par SKU, installation par réf.
        elif target == 'equipements':
            import datetime
            from apps.sav.models import Equipement
            from apps.stock.models import Produit
            from apps.installations.models import Installation
            for i, row in enumerate(rows, 1):
                f = _row_to_fields(row, mapped)
                # Résolution produit (SKU obligatoire).
                produit_sku = f.pop('produit_sku', None)
                if not produit_sku:
                    skipped.append({'ligne': i, 'raison': 'produit_sku manquant'})
                    continue
                try:
                    produit = Produit.objects.get(company=company, sku=produit_sku)
                except Produit.DoesNotExist:
                    skipped.append({'ligne': i, 'raison': f'produit SKU inconnu : {produit_sku}'})
                    continue
                # Résolution installation (référence obligatoire).
                install_ref = f.pop('installation_ref', None)
                if not install_ref:
                    skipped.append({'ligne': i, 'raison': 'installation_ref manquant'})
                    continue
                try:
                    installation = Installation.objects.get(
                        company=company, reference=install_ref)
                except Installation.DoesNotExist:
                    skipped.append({'ligne': i, 'raison': f'installation inconnue : {install_ref}'})
                    continue
                # Numéro de série : doublon par (produit, installation, numero_serie).
                numero_serie = f.get('numero_serie')
                if numero_serie and Equipement.objects.filter(
                        company=company, produit=produit,
                        installation=installation,
                        numero_serie=numero_serie).exists():
                    skipped.append({'ligne': i, 'raison': 'doublon (série existe)'})
                    continue
                # Normalisation date_pose.
                if 'date_pose' in f:
                    raw_date = f['date_pose']
                    if isinstance(raw_date, datetime.datetime):
                        f['date_pose'] = raw_date.date()
                    elif isinstance(raw_date, str):
                        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
                            try:
                                f['date_pose'] = datetime.datetime.strptime(
                                    raw_date.strip(), fmt).date()
                                break
                            except (ValueError, AttributeError):
                                pass
                        else:
                            f.pop('date_pose')
                Equipement.objects.create(
                    company=company, produit=produit,
                    installation=installation, created_by=user, **f)
                created += 1

        # XFLT22 — Véhicules du parc flotte : écriture déléguée à
        # ``apps.flotte.services`` (jamais les models flotte directement).
        elif target == 'vehicules':
            from apps.flotte.services import creer_vehicule_import
            for i, row in enumerate(rows, 1):
                f = _row_to_fields(row, mapped)
                statut, message = creer_vehicule_import(company, f)
                if statut == 'cree':
                    created += 1
                elif statut == 'doublon':
                    skipped.append(
                        {'ligne': i, 'raison': 'doublon (immatriculation existe)'})
                else:
                    skipped.append({'ligne': i, 'raison': message or 'erreur'})

    return {'ok': True, 'target': target, 'mode': mode, 'created': created,
            'updated': updated, 'skipped': skipped, 'total': len(rows)}
