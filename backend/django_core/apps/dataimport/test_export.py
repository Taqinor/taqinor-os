"""N97 — tests de l'export configurable & sauvegarde (CSV/XLSX/JSON + ZIP).

Couvre : le périmètre société (on n'exporte que SA société), l'absence absolue
du prix d'achat dans CSV/XLSX/JSON et le ZIP, les trois formats, le bundle de
sauvegarde, et le verrou de permission (admin uniquement).

Distinct du module d'import : ne touche pas ``tests.py`` (import) ni son code.
"""
import io
import json
import zipfile
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import TestCase
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import AccessToken

from apps.crm.models import Client
from apps.stock.models import Produit
from authentication.models import Company

from . import exporters
from .export_registry import REGISTRY

User = get_user_model()

SECRET_BUY_PRICE = Decimal('1234.56')


class ExportBase(TestCase):
    def setUp(self):
        self.company = Company.objects.get_or_create(
            slug='exp-co', defaults={'nom': 'Export Co'})[0]
        self.other = Company.objects.get_or_create(
            slug='exp-other', defaults={'nom': 'Other Co'})[0]
        self.admin = User.objects.create_user(
            username='exp_admin', password='x', role_legacy='admin',
            company=self.company)
        self.api = APIClient()
        self.api.credentials(
            HTTP_AUTHORIZATION=f'Bearer {AccessToken.for_user(self.admin)}')

    def _auth(self, user):
        c = APIClient()
        c.credentials(HTTP_AUTHORIZATION=f'Bearer {AccessToken.for_user(user)}')
        return c

    def _body(self, resp):
        if getattr(resp, 'streaming', False):
            return b''.join(resp.streaming_content)
        return resp.content


class TestObjectsList(ExportBase):
    def test_catalogue_lists_objects_and_formats(self):
        resp = self.api.get('/api/django/imports/export-objects/')
        self.assertEqual(resp.status_code, 200, resp.data)
        keys = {o['key'] for o in resp.data['objects']}
        self.assertIn('clients', keys)
        self.assertIn('produits', keys)
        self.assertIn('devis', keys)
        fmt_keys = {f['key'] for f in resp.data['formats']}
        self.assertEqual(fmt_keys, {'csv', 'xlsx', 'json'})


class TestCompanyScoping(ExportBase):
    def test_export_only_own_company_rows(self):
        Client.objects.create(company=self.company, nom='Mien')
        Client.objects.create(company=self.other, nom='Autre')
        resp = self.api.post('/api/django/imports/export-object/',
                             {'object': 'clients', 'format': 'csv'},
                             format='json')
        self.assertEqual(resp.status_code, 200)
        text = self._body(resp).decode('utf-8')
        self.assertIn('Mien', text)
        self.assertNotIn('Autre', text)

    def test_registry_queryset_excludes_other_company(self):
        Produit.objects.create(company=self.company, nom='P-mien', prix_vente=10)
        Produit.objects.create(company=self.other, nom='P-autre', prix_vente=10)
        spec = REGISTRY['produits']
        noms = {r[spec.header().index('nom')] for r in spec.rows(self.company)}
        self.assertEqual(noms, {'P-mien'})


class TestNoPrixAchat(ExportBase):
    def setUp(self):
        super().setUp()
        # Produit avec un prix d'achat secret qui ne doit JAMAIS sortir.
        self.prod = Produit.objects.create(
            company=self.company, nom='Panneau Secret', prix_vente=Decimal('999'),
            prix_achat=SECRET_BUY_PRICE)

    def _export(self, fmt):
        resp = self.api.post('/api/django/imports/export-object/',
                             {'object': 'produits', 'format': fmt},
                             format='json')
        self.assertEqual(resp.status_code, 200, fmt)
        return self._body(resp)

    def test_header_never_contains_prix_achat(self):
        self.assertNotIn('prix_achat', REGISTRY['produits'].header())

    def test_prix_achat_absent_from_csv(self):
        body = self._export('csv').decode('utf-8')
        self.assertIn('Panneau Secret', body)  # produit bien exporté
        self.assertNotIn('prix_achat', body)
        self.assertNotIn(str(SECRET_BUY_PRICE), body)

    def test_prix_achat_absent_from_json(self):
        body = self._export('json').decode('utf-8')
        payload = json.loads(body)
        self.assertNotIn('prix_achat', payload['columns'])
        for rec in payload['records']:
            self.assertNotIn('prix_achat', rec)
        self.assertNotIn(str(SECRET_BUY_PRICE), body)

    def test_prix_achat_absent_from_xlsx(self):
        body = self._export('xlsx')
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(body), read_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                self.assertNotIn('prix_achat', str(cell))
                self.assertNotEqual(str(cell), str(SECRET_BUY_PRICE))
                self.assertNotEqual(str(cell), str(float(SECRET_BUY_PRICE)))

    def test_prix_achat_absent_from_zip(self):
        resp = self.api.post('/api/django/imports/sauvegarde/',
                             {'objects': ['produits'], 'format': 'csv'},
                             format='json')
        self.assertEqual(resp.status_code, 200)
        zf = zipfile.ZipFile(io.BytesIO(self._body(resp)))
        for name in zf.namelist():
            blob = zf.read(name)
            self.assertNotIn(b'prix_achat', blob)
            self.assertNotIn(str(SECRET_BUY_PRICE).encode(), blob)


class TestFormats(ExportBase):
    def setUp(self):
        super().setUp()
        Client.objects.create(company=self.company, nom='ClientA')

    def test_csv(self):
        resp = self.api.post('/api/django/imports/export-object/',
                             {'object': 'clients', 'format': 'csv'},
                             format='json')
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp['Content-Type'].startswith('text/csv'))
        self.assertIn('attachment', resp['Content-Disposition'])
        self.assertIn('ClientA', self._body(resp).decode('utf-8'))

    def test_xlsx(self):
        resp = self.api.post('/api/django/imports/export-object/',
                             {'object': 'clients', 'format': 'xlsx'},
                             format='json')
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(self._body(resp).startswith(b'PK'))  # zip-based xlsx

    def test_json(self):
        resp = self.api.post('/api/django/imports/export-object/',
                             {'object': 'clients', 'format': 'json'},
                             format='json')
        self.assertEqual(resp.status_code, 200)
        payload = json.loads(self._body(resp).decode('utf-8'))
        self.assertEqual(payload['object'], 'clients')
        self.assertEqual(payload['count'], 1)
        self.assertEqual(payload['records'][0]['nom'], 'ClientA')

    def test_unknown_object_400(self):
        resp = self.api.post('/api/django/imports/export-object/',
                             {'object': 'bogus', 'format': 'csv'},
                             format='json')
        self.assertEqual(resp.status_code, 400)

    def test_unknown_format_falls_back_to_csv(self):
        resp = self.api.post('/api/django/imports/export-object/',
                             {'object': 'clients', 'format': 'pdf'},
                             format='json')
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp['Content-Type'].startswith('text/csv'))


class TestSauvegardeBundle(ExportBase):
    def test_zip_bundle_contains_one_file_per_object_plus_manifest(self):
        Client.objects.create(company=self.company, nom='C1')
        Produit.objects.create(company=self.company, nom='P1', prix_vente=5)
        resp = self.api.post('/api/django/imports/sauvegarde/',
                             {'objects': ['clients', 'produits'],
                              'format': 'csv'}, format='json')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/zip')
        self.assertIn('.zip', resp['Content-Disposition'])
        zf = zipfile.ZipFile(io.BytesIO(self._body(resp)))
        names = zf.namelist()
        self.assertIn('MANIFEST.txt', names)
        data_files = [n for n in names if n != 'MANIFEST.txt']
        self.assertEqual(len(data_files), 2)
        self.assertTrue(any(n.startswith('clients_') for n in data_files))
        self.assertTrue(any(n.startswith('produits_') for n in data_files))

    def test_zip_bundle_default_is_all_objects(self):
        resp = self.api.post('/api/django/imports/sauvegarde/', {},
                             format='json')
        self.assertEqual(resp.status_code, 200)
        zf = zipfile.ZipFile(io.BytesIO(self._body(resp)))
        data_files = [n for n in zf.namelist() if n != 'MANIFEST.txt']
        self.assertEqual(len(data_files), len(REGISTRY))

    def test_build_backup_zip_helper_scopes_to_company(self):
        Client.objects.create(company=self.company, nom='Inclus')
        Client.objects.create(company=self.other, nom='Exclu')
        specs = [REGISTRY['clients']]
        data = exporters.build_backup_zip(specs, self.company, 'csv')
        zf = zipfile.ZipFile(io.BytesIO(data))
        blob = b''.join(zf.read(n) for n in zf.namelist())
        self.assertIn(b'Inclus', blob)
        self.assertNotIn(b'Exclu', blob)


class TestPermissionGating(ExportBase):
    def test_non_admin_blocked(self):
        plain = User.objects.create_user(
            username='exp_plain', password='x', role_legacy='normal',
            company=self.company)
        client = self._auth(plain)
        for url in ('/api/django/imports/export-objects/',):
            self.assertEqual(client.get(url).status_code, 403, url)
        post_urls = (
            ('/api/django/imports/export-object/',
             {'object': 'clients', 'format': 'csv'}),
            ('/api/django/imports/sauvegarde/', {}),
        )
        for url, body in post_urls:
            self.assertEqual(
                client.post(url, body, format='json').status_code, 403, url)

    def test_anonymous_blocked(self):
        anon = APIClient()
        self.assertEqual(
            anon.get('/api/django/imports/export-objects/').status_code, 401)

    def test_responsable_blocked(self):
        # Le palier Responsable n'est PAS admin : pas d'export/sauvegarde.
        resp_user = User.objects.create_user(
            username='exp_resp', password='x', role_legacy='responsable',
            company=self.company)
        client = self._auth(resp_user)
        self.assertEqual(
            client.get('/api/django/imports/export-objects/').status_code, 403)


class TestManagementCommand(ExportBase):
    def test_command_writes_one_file_per_object(self):
        import tempfile
        from django.core.management import call_command

        Client.objects.create(company=self.company, nom='CLI Client')
        with tempfile.TemporaryDirectory() as out:
            call_command('export_company_data', company_slug='exp-co',
                         out_dir=out, objects='clients', format='csv')
            import os
            files = os.listdir(out)
            self.assertEqual(len(files), 1)
            self.assertTrue(files[0].startswith('clients_'))
            with open(os.path.join(out, files[0]), encoding='utf-8') as fh:
                self.assertIn('CLI Client', fh.read())

    def test_command_zip_excludes_prix_achat(self):
        import os
        import tempfile
        from django.core.management import call_command

        Produit.objects.create(
            company=self.company, nom='Pompe', prix_vente=Decimal('500'),
            prix_achat=SECRET_BUY_PRICE)
        with tempfile.TemporaryDirectory() as out:
            call_command('export_company_data', company_slug='exp-co',
                         out_dir=out, objects='produits', format='csv',
                         zip=True)
            files = [f for f in os.listdir(out) if f.endswith('.zip')]
            self.assertEqual(len(files), 1)
            zf = zipfile.ZipFile(os.path.join(out, files[0]))
            blob = b''.join(zf.read(n) for n in zf.namelist())
            self.assertNotIn(b'prix_achat', blob)
            self.assertNotIn(str(SECRET_BUY_PRICE).encode(), blob)

    def test_command_requires_out_dir(self):
        from django.core.management import call_command
        from django.core.management.base import CommandError

        with self.assertRaises(CommandError):
            call_command('export_company_data', company_slug='exp-co')
