"""N97 — export / sauvegarde des données du tenant.

Une action ADMIN, en lecture seule, qui rassemble les données métier de la
société courante dans un SEUL classeur Excel (.xlsx), une feuille par objet
(leads, clients, devis, factures, produits, chantiers, équipements, tickets
SAV). C'est une sauvegarde lisible et ré-importable (les colonnes reprennent
celles de l'importateur réutilisable T9/N81), borné société, sans aucun prix
d'achat / marge (jamais de donnée interne sensible dans un export).

Aucun planificateur, aucun appel externe : l'admin déclenche la sauvegarde à la
demande et reçoit le fichier. Les vraies données client ne sont jamais
commitées dans le dépôt — le fichier est produit à la volée et téléchargé.
"""
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response

from authentication.permissions import IsAdminRole


def _co_qs(model, user):
    qs = model.objects.all()
    if user.company_id:
        return qs.filter(company=user.company)
    if user.is_superuser:
        return qs
    return qs.none()


def _fmt_dt(value, fmt='%Y-%m-%d %H:%M'):
    return value.strftime(fmt) if value else ''


def _fmt_d(value):
    return value.isoformat() if value else ''


def _leads_sheet(user):
    from apps.crm.models import Lead
    from apps.crm.exports import LEAD_EXPORT_HEADERS, lead_row
    rows = [lead_row(le) for le in _co_qs(Lead, user).order_by('-date_creation')]
    return 'Leads', LEAD_EXPORT_HEADERS, rows


def _clients_sheet(user):
    from apps.crm.models import Client
    headers = ['Nom', 'Prénom', 'Email', 'Téléphone', 'Adresse', 'ICE', 'Créé le']
    rows = [[
        c.nom or '', c.prenom or '', c.email or '', c.telephone or '',
        c.adresse or '', getattr(c, 'ice', '') or '',
        _fmt_dt(getattr(c, 'date_creation', None), '%Y-%m-%d'),
    ] for c in _co_qs(Client, user).order_by('nom')]
    return 'Clients', headers, rows


def _produits_sheet(user):
    from apps.stock.models import Produit
    # JAMAIS de prix d'achat / marge : champ interne (générateur uniquement).
    headers = ['SKU', 'Nom', 'Catégorie', 'Marque', 'Prix vente TTC',
               'Quantité', 'Garantie (mois)', 'Garantie production (mois)']
    rows = [[
        getattr(p, 'sku', '') or '', p.nom or '',
        getattr(getattr(p, 'categorie', None), 'nom', '') or '',
        getattr(p, 'marque', '') or '',
        str(getattr(p, 'prix_vente', '') or ''),
        str(getattr(p, 'quantite_stock', '') or ''),
        getattr(p, 'garantie_mois', '') or '',
        getattr(p, 'garantie_production_mois', '') or '',
    ] for p in _co_qs(Produit, user).order_by('nom')]
    return 'Produits', headers, rows


def _devis_sheet(user):
    from apps.ventes.models import Devis
    headers = ['Référence', 'Client', 'Statut', 'Total HT', 'Total TTC',
               'Créé le', 'Validité']
    rows = [[
        d.reference, getattr(d.client, 'nom', '') or '', d.get_statut_display(),
        str(d.total_ht or 0), str(d.total_ttc or 0),
        _fmt_dt(d.date_creation, '%Y-%m-%d'), _fmt_d(d.date_validite),
    ] for d in _co_qs(Devis, user).select_related('client').order_by('-id')]
    return 'Devis', headers, rows


def _factures_sheet(user):
    from apps.ventes.models import Facture
    headers = ['Référence', 'Client', 'Statut', 'Total TTC', 'Échéance', 'Créée le']
    rows = [[
        f.reference, getattr(f.client, 'nom', '') or '', f.get_statut_display(),
        str(getattr(f, 'total_ttc', '') or ''), _fmt_d(f.date_echeance),
        _fmt_dt(f.date_creation, '%Y-%m-%d'),
    ] for f in _co_qs(Facture, user).select_related('client').order_by('-id')]
    return 'Factures', headers, rows


def _chantiers_sheet(user):
    from apps.installations.models import Installation
    headers = ['Référence', 'Client', 'Statut', 'Type', 'kWc',
               'Pose prévue', 'Mise en service', 'Réception']
    rows = [[
        i.reference, getattr(i.client, 'nom', '') or '', i.get_statut_display(),
        i.get_type_installation_display() if i.type_installation else '',
        str(i.puissance_installee_kwc or ''),
        _fmt_d(i.date_pose_prevue), _fmt_d(i.date_mise_en_service),
        _fmt_d(i.date_reception),
    ] for i in _co_qs(Installation, user).select_related('client').order_by('-id')]
    return 'Chantiers', headers, rows


def _equipements_sheet(user):
    from apps.sav.models import Equipement
    headers = ['N° série', 'Produit', 'Pose', 'Fin garantie', 'Statut']
    rows = [[
        e.numero_serie or '', getattr(e.produit, 'nom', '') or '',
        _fmt_d(e.date_pose), _fmt_d(e.date_fin_garantie),
        e.get_statut_display() if hasattr(e, 'get_statut_display') else e.statut,
    ] for e in _co_qs(Equipement, user).select_related('produit').order_by('-id')]
    return 'Équipements', headers, rows


def _tickets_sheet(user):
    from apps.sav.models import Ticket
    headers = ['Référence', 'Client', 'Type', 'Statut', 'Ouverture', 'Résolution']
    rows = [[
        t.reference, getattr(t.client, 'nom', '') or '',
        t.get_type_display() if hasattr(t, 'get_type_display') else t.type,
        t.get_statut_display(), _fmt_d(t.date_ouverture), _fmt_d(t.date_resolution),
    ] for t in _co_qs(Ticket, user).select_related('client').order_by('-id')]
    return 'Tickets SAV', headers, rows


# Ordre des feuilles dans le classeur de sauvegarde.
_SHEET_BUILDERS = [
    _leads_sheet, _clients_sheet, _produits_sheet, _devis_sheet,
    _factures_sheet, _chantiers_sheet, _equipements_sheet, _tickets_sheet,
]


def build_backup_workbook(user):
    """Construit le classeur openpyxl multi-feuilles de la société de `user`.

    Renvoie (workbook, résumé{feuille: nb_lignes}). Séparé de la vue HTTP pour
    être testable sans la couche requête.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    # On retire la feuille par défaut puis on en ajoute une par objet.
    wb.remove(wb.active)
    bold = Font(bold=True)
    resume = {}

    for builder in _SHEET_BUILDERS:
        title, headers, rows = builder(user)
        ws = wb.create_sheet(title=title[:31] or 'Feuille')
        ws.append(list(headers))
        for cell in ws[1]:
            cell.font = bold
        for row in rows:
            ws.append(['' if v is None else v for v in row])
        for idx, _ in enumerate(headers, start=1):
            longest = len(str(headers[idx - 1]))
            for row in rows:
                if idx - 1 < len(row) and row[idx - 1] is not None:
                    longest = max(longest, len(str(row[idx - 1])))
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = \
                min(max(longest + 2, 10), 50)
        resume[title] = len(rows)

    return wb, resume


@api_view(['GET'])
@permission_classes([IsAdminRole])
def data_backup(request):
    """GET /reporting/backup/ → sauvegarde .xlsx complète de la société (admin).

    ?summary=1 renvoie seulement le résumé JSON {feuille: nb_lignes} sans
    produire le fichier (pour afficher un aperçu avant le téléchargement)."""
    from django.http import HttpResponse
    from django.utils import timezone

    user = request.user
    if not (user.company_id or user.is_superuser):
        return Response({'detail': 'Accès refusé.'}, status=403)

    wb, resume = build_backup_workbook(user)

    if request.query_params.get('summary'):
        return Response({'total': sum(resume.values()), 'feuilles': resume})

    stamp = timezone.now().strftime('%Y%m%d-%H%M')
    response = HttpResponse(
        content_type=(
            'application/vnd.openxmlformats-officedocument.'
            'spreadsheetml.sheet'
        ),
    )
    response['Content-Disposition'] = \
        f'attachment; filename="sauvegarde-taqinor-{stamp}.xlsx"'
    wb.save(response)
    return response
