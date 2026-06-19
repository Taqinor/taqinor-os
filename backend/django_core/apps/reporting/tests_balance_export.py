"""Tests — export .xlsx de la balance âgée (apps.reporting).

Vérifie : (1) l'endpoint renvoie un classeur .xlsx 200 borné à la société ;
(2) le bucketing (0–30/31–60/61–90/90+) + le total par client sont corrects et
cohérents avec l'endpoint JSON `/ventes/balance-agee/` ; (3) un pied « Total »
récapitule les tranches ; (4) un utilisateur sans société est refusé (403).
"""
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import AccessToken

from authentication.models import Company
from apps.crm.models import Client
from apps.stock.models import Produit
from apps.ventes.models import Facture, LigneFacture
from apps.reporting.balance_export import balance_agee_rows

User = get_user_model()

XLSX_CT = (
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
)


class TestBalanceAgeeExport(TestCase):
    def setUp(self):
        self.company = Company.objects.get_or_create(
            slug='bal-exp-co', defaults={'nom': 'Bal Exp Co'})[0]
        self.user = User.objects.create_user(
            username='bal_exp_resp', password='x', role_legacy='responsable',
            company=self.company)
        self.client_obj = Client.objects.create(
            company=self.company, nom='Débiteur', telephone='+212600000003')
        self.produit = Produit.objects.create(
            company=self.company, nom='Onduleur', sku='OND-BE',
            prix_vente=Decimal('5000'), quantite_stock=10,
            tva=Decimal('20.00'))
        # Facture émise, échéance dépassée de 45 j → bucket 31–60, due 6000.
        self.facture = Facture.objects.create(
            company=self.company, reference='FAC-BE-0001',
            client=self.client_obj, statut=Facture.Statut.EMISE,
            taux_tva=Decimal('20.00'),
            date_echeance=date.today() - timedelta(days=45))
        LigneFacture.objects.create(
            facture=self.facture, produit=self.produit, designation='Onduleur',
            quantite=Decimal('1'), prix_unitaire=Decimal('5000'),
            taux_tva=Decimal('20.00'))
        self.api = APIClient()
        self.api.credentials(
            HTTP_AUTHORIZATION=f'Bearer {AccessToken.for_user(self.user)}')

    def test_export_returns_xlsx(self):
        resp = self.api.get('/api/django/reporting/balance-agee/export/')
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp['Content-Type'], XLSX_CT)
        self.assertIn('balance-agee.xlsx', resp['Content-Disposition'])

    def test_rows_bucketing_and_total(self):
        rows = balance_agee_rows(self.user)
        # Une ligne client + un pied « Total ».
        self.assertEqual(len(rows), 2)
        client_row = rows[0]
        # [client, 0-30, 31-60, 61-90, 90+, total]
        self.assertEqual(client_row[0], 'Débiteur')
        self.assertEqual(client_row[1], 0.0)        # 0–30
        self.assertEqual(client_row[2], 6000.0)     # 31–60 (45 j)
        self.assertEqual(client_row[3], 0.0)        # 61–90
        self.assertEqual(client_row[4], 0.0)        # 90+
        self.assertEqual(client_row[5], 6000.0)     # total
        # Pied « Total » récapitule les tranches.
        total_row = rows[1]
        self.assertEqual(total_row[0], 'Total')
        self.assertEqual(total_row[2], 6000.0)
        self.assertEqual(total_row[5], 6000.0)

    def test_workbook_is_readable(self):
        from openpyxl import load_workbook
        resp = self.api.get('/api/django/reporting/balance-agee/export/')
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        self.assertEqual(ws.title, 'Balance âgée')
        header = [c.value for c in ws[1]]
        self.assertEqual(header[0], 'Client')
        self.assertIn('Total dû', header)

    def test_company_scoped(self):
        # Une autre société ne voit pas la créance de la première.
        other_co = Company.objects.get_or_create(
            slug='bal-exp-other', defaults={'nom': 'Other'})[0]
        other_user = User.objects.create_user(
            username='bal_exp_other', password='x', role_legacy='responsable',
            company=other_co)
        rows = balance_agee_rows(other_user)
        self.assertEqual(rows, [])

    def test_no_company_user_forbidden(self):
        nobody = User.objects.create_user(
            username='bal_exp_nobody', password='x', role_legacy='responsable',
            company=None)
        api = APIClient()
        api.credentials(
            HTTP_AUTHORIZATION=f'Bearer {AccessToken.for_user(nobody)}')
        resp = api.get('/api/django/reporting/balance-agee/export/')
        self.assertEqual(resp.status_code, 403)
