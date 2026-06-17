"""N97 — sauvegarde / export complet des données du tenant (admin, scopé société)."""
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import AccessToken

from apps.crm.models import Client, Lead
from authentication.models import Company

User = get_user_model()


class TestDataBackup(TestCase):
    def setUp(self):
        self.company = Company.objects.get_or_create(
            slug='backup-co', defaults={'nom': 'Backup Co'})[0]
        self.other = Company.objects.create(slug='backup-other', nom='Autre')
        self.admin = User.objects.create_user(
            username='backup_admin', password='x', role_legacy='admin',
            company=self.company)
        self.normal = User.objects.create_user(
            username='backup_normal', password='x', role_legacy='normal',
            company=self.company)
        # Données : un lead par société + un client.
        Lead.objects.create(company=self.company, nom='Alpha', prenom='Un')
        Lead.objects.create(company=self.other, nom='Beta', prenom='Deux')
        Client.objects.create(company=self.company, nom='ClientA')
        self.api = APIClient()

    def _auth(self, user):
        self.api.credentials(
            HTTP_AUTHORIZATION=f'Bearer {AccessToken.for_user(user)}')

    def test_summary_is_company_scoped(self):
        self._auth(self.admin)
        resp = self.api.get('/api/django/reporting/backup/?summary=1')
        self.assertEqual(resp.status_code, 200)
        # Une seule société : un seul lead et un seul client comptés.
        self.assertEqual(resp.data['feuilles']['Leads'], 1)
        self.assertEqual(resp.data['feuilles']['Clients'], 1)
        self.assertIn('Devis', resp.data['feuilles'])

    def test_download_returns_valid_xlsx(self):
        self._auth(self.admin)
        resp = self.api.get('/api/django/reporting/backup/')
        self.assertEqual(resp.status_code, 200)
        self.assertIn('spreadsheetml.sheet', resp['Content-Type'])
        self.assertIn('attachment', resp['Content-Disposition'])
        # Le fichier s'ouvre réellement et porte les feuilles attendues.
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(resp.getvalue()))
        self.assertIn('Leads', wb.sheetnames)
        self.assertIn('Tickets SAV', wb.sheetnames)
        # En-tête + 1 lead.
        self.assertEqual(wb['Leads'].max_row, 2)

    def test_non_admin_forbidden(self):
        self._auth(self.normal)
        resp = self.api.get('/api/django/reporting/backup/?summary=1')
        self.assertEqual(resp.status_code, 403)
