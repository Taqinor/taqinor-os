"""Tests de l'archive documentaire (N32) — reporting/archive.py.

Vérifie l'agrégation des documents d'un client et d'un chantier, l'isolation
multi-tenant (404 inter-société), et la présence des URLs de téléchargement.
"""
from decimal import Decimal

from django.test import TestCase
from django.contrib.auth import get_user_model
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import AccessToken

from authentication.models import Company
from apps.crm.models import Client
from apps.stock.models import Produit
from apps.ventes.models import Devis, LigneDevis
from apps.installations.models import Installation

User = get_user_model()


def _company(slug='arch-co', nom='Arch Co'):
    company, _ = Company.objects.get_or_create(slug=slug, defaults={'nom': nom})
    return company


def _user(company, username):
    return User.objects.create_user(
        username=username, password='pw',
        role_legacy='responsable', company=company,
    )


class ArchiveDocumentsTest(TestCase):
    def setUp(self):
        self.company = _company()
        self.user = _user(self.company, 'arch_user')
        self.client_obj = Client.objects.create(
            nom='Bennani', prenom='Karim', telephone='0600000000',
            company=self.company)
        self.devis = Devis.objects.create(
            reference='DEV-ARCH-0001', client=self.client_obj,
            statut='accepte', taux_tva=Decimal('20.00'),
            remise_globale=Decimal('0'), created_by=self.user,
            company=self.company)
        self.produit = Produit.objects.create(
            nom='Panneau 550W', sku='PV-ARCH',
            prix_vente=Decimal('1500.00'), prix_achat=Decimal('900.00'),
            quantite_stock=50, company=self.company)
        LigneDevis.objects.create(
            devis=self.devis, produit=self.produit, designation='Panneau',
            quantite=Decimal('10'), prix_unitaire=Decimal('1500.00'),
            remise=Decimal('0'))
        self.chantier = Installation.objects.create(
            company=self.company, reference='CH-ARCH-0001',
            client=self.client_obj, devis=self.devis)

        self.other_company = _company(slug='arch-other', nom='Other')
        self.other_user = _user(self.other_company, 'arch_other')

        self.api = APIClient()
        token = str(AccessToken.for_user(self.user))
        self.api.credentials(HTTP_AUTHORIZATION=f'Bearer {token}')

    def test_archive_client_aggregates_docs(self):
        r = self.api.get(
            f'/api/django/reporting/archive/client/{self.client_obj.id}/')
        self.assertEqual(r.status_code, 200)
        types = {d['type'] for d in r.data['documents']}
        self.assertIn('devis', types)
        self.assertIn('pv_reception', types)
        # Une URL de téléchargement existe pour le devis (proposal).
        devis_doc = next(d for d in r.data['documents'] if d['type'] == 'devis')
        self.assertIn('/proposal/', devis_doc['download_url'])

    def test_archive_chantier_aggregates_docs(self):
        r = self.api.get(
            f'/api/django/reporting/archive/chantier/{self.chantier.id}/')
        self.assertEqual(r.status_code, 200)
        types = {d['type'] for d in r.data['documents']}
        self.assertIn('devis', types)
        self.assertIn('attestation', types)

    def test_archive_client_foreign_404(self):
        token = str(AccessToken.for_user(self.other_user))
        self.api.credentials(HTTP_AUTHORIZATION=f'Bearer {token}')
        r = self.api.get(
            f'/api/django/reporting/archive/client/{self.client_obj.id}/')
        self.assertEqual(r.status_code, 404)

    def test_archive_chantier_foreign_404(self):
        token = str(AccessToken.for_user(self.other_user))
        self.api.credentials(HTTP_AUTHORIZATION=f'Bearer {token}')
        r = self.api.get(
            f'/api/django/reporting/archive/chantier/{self.chantier.id}/')
        self.assertEqual(r.status_code, 404)

    # ── L862 — filtrage par type + tri par date ──────────────────────
    def test_archive_filter_by_type(self):
        url = f'/api/django/reporting/archive/client/{self.client_obj.id}/'
        r = self.api.get(url, {'type': 'devis'})
        self.assertEqual(r.status_code, 200)
        types = {d['type'] for d in r.data['documents']}
        self.assertEqual(types, {'devis'})
        # Le compte reflète le filtre.
        self.assertEqual(r.data['count'], len(r.data['documents']))

    def test_archive_sort_asc(self):
        url = f'/api/django/reporting/archive/client/{self.client_obj.id}/'
        dates = [d['date'] for d in self.api.get(
            url, {'sort': 'asc'}).data['documents'] if d['date']]
        self.assertEqual(dates, sorted(dates))

    # ── L863 — bon de commande : pas de PDF explicite ────────────────
    def test_bon_commande_has_pdf_false(self):
        from apps.ventes.models import BonCommande
        BonCommande.objects.create(
            reference='BC-ARCH-0001', client=self.client_obj,
            devis=self.devis, company=self.company)
        url = f'/api/django/reporting/archive/client/{self.client_obj.id}/'
        r = self.api.get(url, {'type': 'bon_commande'})
        self.assertTrue(r.data['documents'])
        bc = r.data['documents'][0]
        self.assertFalse(bc['has_pdf'])
        self.assertIsNone(bc['download_url'])

    def test_devis_has_pdf_true(self):
        url = f'/api/django/reporting/archive/client/{self.client_obj.id}/'
        r = self.api.get(url, {'type': 'devis'})
        self.assertTrue(r.data['documents'][0]['has_pdf'])

    # ── L864 — export .xlsx de la liste ──────────────────────────────
    def test_archive_export_xlsx(self):
        url = f'/api/django/reporting/archive/client/{self.client_obj.id}/'
        r = self.api.get(url, {'export': 'xlsx'})
        self.assertEqual(r.status_code, 200)
        self.assertIn('spreadsheetml', r['Content-Type'])
        self.assertIn('.xlsx', r['Content-Disposition'])
        self.assertTrue(r.content.startswith(b'PK'))

    def test_archive_export_xlsx_respects_type_filter(self):
        from openpyxl import load_workbook
        import io
        url = f'/api/django/reporting/archive/client/{self.client_obj.id}/'
        r = self.api.get(url, {'export': 'xlsx', 'type': 'devis'})
        ws = load_workbook(io.BytesIO(r.content)).active
        # En-tête + lignes devis uniquement.
        self.assertEqual(
            [c.value for c in ws[1]], ['Type', 'Référence', 'Date'])
        body_types = {ws.cell(row=i, column=1).value
                      for i in range(2, ws.max_row + 1)}
        self.assertTrue(body_types <= {'Devis'})
