"""Tests des rapports analytiques (lecture seule).

Couvre : le scoping par société (multi-tenant), la réconciliation TVA 10/20 du
journal comptable, et la forme attendue des réponses de chaque endpoint.
"""
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import AccessToken
from openpyxl import load_workbook

from authentication.models import Company
from apps.crm import stages
from apps.crm.models import Client, Lead
from apps.ventes.models import Devis, LigneDevis, Facture, LigneFacture
from apps.stock.models import Categorie, Produit, MouvementStock
from apps.installations.models import Installation
from apps.sav.models import Equipement, Ticket

User = get_user_model()

# Étapes canoniques (STAGES.py) — référencées par nom via le module canonique
# chargé par apps.crm.stages, jamais codées en dur dans les tests.
NEW = stages.NEW
QUOTE_SENT = stages._stages.QUOTE_SENT
SIGNED = stages._stages.SIGNED


def auth(company):
    user = User.objects.create_user(
        username=f'u_{company.slug}', password='x',
        role_legacy='responsable', company=company)
    api = APIClient()
    token = str(AccessToken.for_user(user))
    api.credentials(HTTP_AUTHORIZATION=f'Bearer {token}')
    return api, user


class BaseData(TestCase):
    def setUp(self):
        self.company = Company.objects.get_or_create(
            slug='co-a', defaults={'nom': 'Co A'})[0]
        self.other = Company.objects.get_or_create(
            slug='co-b', defaults={'nom': 'Co B'})[0]
        self.api, self.user = auth(self.company)
        self.client_a = Client.objects.create(
            company=self.company, nom='Client A', email='a@a.ma')
        self.client_b = Client.objects.create(
            company=self.other, nom='Client B', email='b@b.ma')
        self.cat = Categorie.objects.create(company=self.company, nom='Panneaux')
        self.prod = Produit.objects.create(
            company=self.company, nom='Panneau 550W', sku='PV550',
            prix_achat=Decimal('800'), prix_vente=Decimal('1200'),
            quantite_stock=10, seuil_alerte=3, categorie=self.cat,
            marque='JA Solar')
        self.prod_other = Produit.objects.create(
            company=self.other, nom='Autre', sku='OTH',
            prix_achat=Decimal('1'), prix_vente=Decimal('2'),
            quantite_stock=100)


class TestPipelineValue(BaseData):
    def test_shape_and_stage_keys(self):
        Lead.objects.create(company=self.company, nom='L1',
                            stage=QUOTE_SENT)
        Lead.objects.create(company=self.company, nom='L2',
                            stage=SIGNED)
        Lead.objects.create(company=self.company, nom='Perdu', perdu=True,
                            motif_perte='Prix trop élevé')
        resp = self.api.get('/api/django/reporting/pipeline-value/')
        self.assertEqual(resp.status_code, 200)
        data = resp.data
        self.assertIn('total_pipeline', data)
        self.assertIn('forecast_pondere', data)
        self.assertIn('par_etape', data)
        # Étapes exactement et dans l'ordre canonique STAGES.py.
        keys = [r['stage'] for r in data['par_etape']]
        self.assertEqual(keys, stages.STAGES)
        self.assertIn('devis_par_statut', data)
        self.assertIn('win_loss', data)
        self.assertEqual(data['win_loss']['perdus'], 1)
        self.assertEqual(data['win_loss']['par_motif'][0]['motif'],
                         'Prix trop élevé')

    def test_company_scoping(self):
        Lead.objects.create(company=self.company, nom='Mine',
                            stage=NEW)
        Lead.objects.create(company=self.other, nom='Theirs',
                            stage=NEW)
        resp = self.api.get('/api/django/reporting/pipeline-value/')
        new_bucket = next(r for r in resp.data['par_etape']
                          if r['stage'] == NEW)
        # Seul le lead de ma société est compté.
        self.assertEqual(new_bucket['nb'], 1)

    def test_weighted_forecast_uses_devis_value(self):
        lead = Lead.objects.create(company=self.company, nom='L',
                                   stage=QUOTE_SENT)
        devis = Devis.objects.create(
            company=self.company, reference='D1', client=self.client_a,
            lead=lead, taux_tva=Decimal('20'))
        LigneDevis.objects.create(
            devis=devis, produit=self.prod, designation='x',
            quantite=Decimal('1'), prix_unitaire=Decimal('1000'))
        resp = self.api.get('/api/django/reporting/pipeline-value/')
        b = next(r for r in resp.data['par_etape']
                 if r['stage'] == QUOTE_SENT)
        # TTC = 1200 ; poids QUOTE_SENT = 0.5 → pondéré = 600.
        self.assertEqual(b['valeur'], 1200.0)
        self.assertEqual(b['valeur_ponderee'], 600.0)


class TestSalesReport(BaseData):
    def test_shape(self):
        Lead.objects.create(company=self.company, nom='L', stage=NEW)
        resp = self.api.get('/api/django/reporting/sales/')
        self.assertEqual(resp.status_code, 200)
        for k in ('funnel', 'devis_par_statut', 'ca_par_responsable',
                  'ca_par_canal', 'ca_par_mois', 'win_loss'):
            self.assertIn(k, resp.data)
        self.assertEqual([r['stage'] for r in resp.data['funnel']],
                         stages.STAGES)

    def test_xlsx_export(self):
        resp = self.api.get('/api/django/reporting/sales/?export=xlsx')
        self.assertEqual(resp.status_code, 200)
        self.assertIn('spreadsheetml', resp['Content-Type'])
        wb = load_workbook(BytesIO(resp.content))
        self.assertIn('Entonnoir', wb.sheetnames)
        self.assertIn('Gagné-Perdu', wb.sheetnames)


class TestStockReport(BaseData):
    def test_shape_and_valuation(self):
        resp = self.api.get('/api/django/reporting/stock/')
        self.assertEqual(resp.status_code, 200)
        vt = resp.data['valorisation_totale']
        # Seul mon produit (10 × 1200 = 12000 vente, 10 × 800 = 8000 achat).
        self.assertEqual(vt['valeur_vente'], 12000.0)
        self.assertEqual(vt['valeur_achat'], 8000.0)
        self.assertEqual(vt['marge_potentielle'], 4000.0)

    def test_company_scoping(self):
        resp = self.api.get('/api/django/reporting/stock/')
        noms = [r['nom'] for r in resp.data['valorisation']]
        self.assertIn('Panneau 550W', noms)
        self.assertNotIn('Autre', noms)

    def test_movements_and_low_stock(self):
        MouvementStock.objects.create(
            company=self.company, produit=self.prod,
            type_mouvement=MouvementStock.TypeMouvement.ENTREE,
            quantite=5, quantite_avant=10, quantite_apres=15)
        low = Produit.objects.create(
            company=self.company, nom='Cable', sku='CBL',
            prix_vente=Decimal('10'), quantite_stock=1, seuil_alerte=5)
        resp = self.api.get('/api/django/reporting/stock/')
        self.assertEqual(len(resp.data['mouvements']), 1)
        sous_seuil_noms = [r['nom'] for r in resp.data['sous_seuil']]
        self.assertIn(low.nom, sous_seuil_noms)

    def test_xlsx_export(self):
        resp = self.api.get('/api/django/reporting/stock/?export=xlsx')
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(BytesIO(resp.content))
        self.assertIn('Valorisation', wb.sheetnames)


class TestServiceReport(BaseData):
    def setUp(self):
        super().setUp()
        self.inst = Installation.objects.create(
            company=self.company, reference='CH-1', client=self.client_a,
            statut=Installation.Statut.CLOTURE,
            date_pose_reelle=date.today())

    def test_shape(self):
        resp = self.api.get('/api/django/reporting/service/')
        self.assertEqual(resp.status_code, 200)
        for k in ('chantiers_par_statut', 'completion',
                  'activite_techniciens', 'sav', 'garanties_expirant'):
            self.assertIn(k, resp.data)

    def test_sav_open_vs_resolved(self):
        Ticket.objects.create(
            company=self.company, reference='T1', client=self.client_a,
            installation=self.inst, statut=Ticket.Statut.NOUVEAU)
        Ticket.objects.create(
            company=self.company, reference='T2', client=self.client_a,
            installation=self.inst, statut=Ticket.Statut.RESOLU,
            date_ouverture=date.today() - timedelta(days=4),
            date_resolution=date.today())
        resp = self.api.get('/api/django/reporting/service/')
        self.assertEqual(resp.data['sav']['ouverts'], 1)
        self.assertEqual(resp.data['sav']['resolus'], 1)
        self.assertEqual(resp.data['sav']['delai_resolution_moyen_jours'], 4.0)

    def test_warranty_expiring(self):
        Equipement.objects.create(
            company=self.company, produit=self.prod, installation=self.inst,
            numero_serie='SN1', date_pose=date.today(),
            date_fin_garantie=date.today() + timedelta(days=30))
        resp = self.api.get('/api/django/reporting/service/')
        self.assertEqual(len(resp.data['garanties_expirant']), 1)
        self.assertEqual(resp.data['garanties_expirant'][0]['numero_serie'],
                         'SN1')

    def test_company_scoping(self):
        # Un chantier d'une autre société ne doit pas apparaître.
        Installation.objects.create(
            company=self.other, reference='CH-OTHER', client=self.client_b,
            statut=Installation.Statut.A_PLANIFIER)
        resp = self.api.get('/api/django/reporting/service/')
        total = sum(r['nb'] for r in resp.data['chantiers_par_statut'])
        self.assertEqual(total, 1)  # uniquement CH-1

    def test_xlsx_export(self):
        resp = self.api.get('/api/django/reporting/service/?export=xlsx')
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(BytesIO(resp.content))
        self.assertIn('SAV', wb.sheetnames)


class TestJournalVentesTVA(BaseData):
    def _facture_mixte(self):
        """Facture à deux taux : 10 % (panneaux) et 20 % (reste)."""
        f = Facture.objects.create(
            company=self.company, reference='F1', client=self.client_a,
            statut=Facture.Statut.EMISE, taux_tva=Decimal('20'))
        # On force date_emission (auto_now_add) après création.
        Facture.objects.filter(pk=f.pk).update(date_emission=date.today())
        LigneFacture.objects.create(
            facture=f, produit=self.prod, designation='Panneaux',
            quantite=Decimal('10'), prix_unitaire=Decimal('1000'),
            taux_tva=Decimal('10'))
        LigneFacture.objects.create(
            facture=f, produit=self.prod, designation='Onduleur',
            quantite=Decimal('1'), prix_unitaire=Decimal('5000'),
            taux_tva=Decimal('20'))
        return f

    def test_tva_split_10_20_reconciled(self):
        self._facture_mixte()
        resp = self.api.get('/api/django/reporting/journal-ventes/')
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(BytesIO(resp.content))
        ws = wb['Récapitulatif TVA']
        rows = list(ws.iter_rows(values_only=True))
        # En-tête + une ligne par taux.
        data_rows = [r for r in rows[1:] if r[0] and '%' in str(r[0])]
        recap = {r[0]: (Decimal(str(r[1])), Decimal(str(r[2])))
                 for r in data_rows}
        # Base 10 % = 10 000 → TVA = 1 000.00 ; base 20 % = 5 000 → 1 000.00.
        self.assertEqual(recap['10 %'][0], Decimal('10000.00'))
        self.assertEqual(recap['10 %'][1], Decimal('1000.00'))
        self.assertEqual(recap['20 %'][0], Decimal('5000.00'))
        self.assertEqual(recap['20 %'][1], Decimal('1000.00'))
        # Réconciliation au centime : somme des TVA = total TVA déclaré.
        somme_tva = recap['10 %'][1] + recap['20 %'][1]
        totaux = {r[0]: r[1] for r in rows if r and r[0] in (
            'Total HT', 'Total TVA', 'Total TTC')}
        self.assertEqual(Decimal(str(totaux['Total TVA'])), somme_tva)
        self.assertEqual(Decimal(str(totaux['Total HT'])),
                         Decimal('15000.00'))
        self.assertEqual(Decimal(str(totaux['Total TTC'])),
                         Decimal('17000.00'))

    def test_excludes_brouillon_and_other_company(self):
        # Brouillon exclu.
        Facture.objects.create(
            company=self.company, reference='FB', client=self.client_a,
            statut=Facture.Statut.BROUILLON, taux_tva=Decimal('20'))
        # Facture émise d'une autre société exclue (scoping).
        f_other = Facture.objects.create(
            company=self.other, reference='FX', client=self.client_b,
            statut=Facture.Statut.EMISE, taux_tva=Decimal('20'))
        LigneFacture.objects.create(
            facture=f_other, produit=self.prod_other, designation='x',
            quantite=Decimal('1'), prix_unitaire=Decimal('9999'))
        self._facture_mixte()
        resp = self.api.get('/api/django/reporting/journal-ventes/')
        wb = load_workbook(BytesIO(resp.content))
        ws = wb['Journal des ventes']
        refs = {r[1] for r in ws.iter_rows(min_row=2, values_only=True)
                if r[1]}
        self.assertIn('F1', refs)
        self.assertNotIn('FB', refs)
        self.assertNotIn('FX', refs)


class TestAccessControl(BaseData):
    def test_requires_auth(self):
        anon = APIClient()
        for url in ('/api/django/reporting/pipeline-value/',
                    '/api/django/reporting/sales/',
                    '/api/django/reporting/stock/',
                    '/api/django/reporting/service/',
                    '/api/django/reporting/journal-ventes/'):
            self.assertIn(anon.get(url).status_code, (401, 403), url)
