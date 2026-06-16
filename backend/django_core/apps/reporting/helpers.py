"""Helpers partagés des rapports (lecture seule).

Tout est filtré par société (multi-tenant) au niveau des vues ; ces fonctions
reçoivent des querysets déjà scopés ou des kwargs de société. Aucune écriture.
"""
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO

from django.http import HttpResponse


def company_filter(user):
    """Retourne les kwargs de filtre société, ou None si accès refusé.

    - utilisateur rattaché à une société → {'company': user.company}
    - superuser sans société → {} (voit tout)
    - sinon → None (accès refusé)
    """
    if user.company_id:
        return {'company': user.company}
    if user.is_superuser:
        return {}
    return None


def parse_periode(request):
    """Lit les paramètres de période de la requête.

    Accepte :
      - ?date_debut=YYYY-MM-DD&date_fin=YYYY-MM-DD (prioritaire)
      - ?annee=YYYY&mois=MM        → un mois civil
      - ?annee=YYYY&trimestre=1..4 → un trimestre civil
      - ?annee=YYYY                → l'année civile entière

    Retourne (date_debut|None, date_fin|None, label). Bornes incluses.
    """
    def _d(value):
        try:
            return date.fromisoformat(value)
        except (TypeError, ValueError):
            return None

    debut = _d(request.GET.get('date_debut'))
    fin = _d(request.GET.get('date_fin'))
    if debut or fin:
        label = f"{debut or '…'} → {fin or '…'}"
        return debut, fin, label

    annee = request.GET.get('annee')
    if not annee:
        return None, None, 'Toutes périodes'
    try:
        annee = int(annee)
    except (TypeError, ValueError):
        return None, None, 'Toutes périodes'

    mois = request.GET.get('mois')
    trimestre = request.GET.get('trimestre')
    if mois:
        try:
            m = int(mois)
        except (TypeError, ValueError):
            m = 1
        m = max(1, min(12, m))
        debut = date(annee, m, 1)
        fin = date(annee + 1, 1, 1) if m == 12 else date(annee, m + 1, 1)
        from datetime import timedelta
        fin = fin - timedelta(days=1)
        return debut, fin, f"{m:02d}/{annee}"
    if trimestre:
        try:
            t = int(trimestre)
        except (TypeError, ValueError):
            t = 1
        t = max(1, min(4, t))
        first_month = (t - 1) * 3 + 1
        debut = date(annee, first_month, 1)
        last_month = first_month + 2
        from datetime import timedelta
        fin_excl = (date(annee + 1, 1, 1) if last_month == 12
                    else date(annee, last_month + 1, 1))
        fin = fin_excl - timedelta(days=1)
        return debut, fin, f"T{t} {annee}"

    return date(annee, 1, 1), date(annee, 12, 31), str(annee)


def apply_date_range(qs, field, debut, fin):
    """Filtre un queryset sur un champ date entre debut et fin (bornes incluses)."""
    if debut:
        qs = qs.filter(**{f'{field}__gte': debut})
    if fin:
        qs = qs.filter(**{f'{field}__lte': fin})
    return qs


def q2(value):
    """Arrondit un Decimal au centime (ROUND_HALF_UP)."""
    return Decimal(value).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)


def f2(value):
    """Float arrondi au centime — pour les réponses JSON."""
    return float(q2(value))


# ── Export .xlsx (openpyxl) ──────────────────────────────────────────────────

def make_workbook():
    from openpyxl import Workbook
    wb = Workbook()
    # Retire la feuille par défaut ; on crée des feuilles nommées explicitement.
    wb.remove(wb.active)
    return wb


def style_header(ws, row=1):
    """Met en gras la ligne d'en-tête d'une feuille."""
    from openpyxl.styles import Font, PatternFill
    fill = PatternFill('solid', fgColor='1E293B')
    font = Font(bold=True, color='FFFFFF')
    for cell in ws[row]:
        cell.font = font
        cell.fill = fill


def autosize(ws):
    """Ajuste (approximativement) la largeur des colonnes."""
    from openpyxl.utils import get_column_letter
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            col = cell.column
            widths[col] = max(widths.get(col, 10), len(str(cell.value)) + 2)
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(width, 60)


def xlsx_response(wb, filename):
    """Sérialise un Workbook openpyxl en réponse HTTP .xlsx téléchargeable."""
    for ws in wb.worksheets:
        autosize(ws)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type=(
            'application/vnd.openxmlformats-officedocument.'
            'spreadsheetml.sheet'
        ),
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


# ── Calculs ventes (factures) ────────────────────────────────────────────────

def facture_totaux(facture):
    """Retourne (ht, tva, ttc) d'une facture — utilise les propriétés modèle
    (montants figés des tranches, sinon somme des lignes)."""
    return (
        Decimal(facture.total_ht),
        Decimal(facture.total_tva),
        Decimal(facture.total_ttc),
    )
