"""NTMFG36 — Export CSV/XLSX des Ordres de Fabrication + historique des
opérations.

Critère : l'export contient les mêmes données que l'écran NTMFG9 sur la
période choisie, deux onglets cohérents entre eux (mêmes OF), isolation
tenant."""
from decimal import Decimal

from django.test import TestCase
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import AccessToken

from apps.mrp.models import Gamme, OperationGamme, OrdreFabrication, PosteDeCharge
from apps.mrp.selectors import export_ordres_fabrication
from apps.mrp.services import confirmer_of
from apps.stock.models import Produit

from ._fixtures import make_company, make_user


def auth(user):
    api = APIClient()
    api.credentials(HTTP_AUTHORIZATION=f'Bearer {AccessToken.for_user(user)}')
    return api


def make_produit(company, nom='Produit'):
    return Produit.objects.create(company=company, nom=nom, prix_vente=0, tva=20)


class ExportOrdresFabricationSelectorTests(TestCase):
    def setUp(self):
        self.company = make_company('mrp-ntmfg36-1', 'MRP NTMFG36 1')
        self.produit = make_produit(self.company)
        self.poste = PosteDeCharge.objects.create(
            company=self.company, code='P-36', nom='Poste 36')
        self.gamme = Gamme.objects.create(
            company=self.company, nom='Gamme 36', produit=self.produit)
        OperationGamme.objects.create(
            gamme=self.gamme, ordre=1, poste_charge=self.poste, libelle='Découpe',
            temps_prepa_min=Decimal('5'), temps_unitaire_min=Decimal('1'))

    def test_onglets_coherents_memes_of(self):
        of = OrdreFabrication.objects.create(
            company=self.company, produit=self.produit, quantite=10,
            gamme=self.gamme)
        confirmer_of(of)

        lignes_of, lignes_operations = export_ordres_fabrication(self.company)
        self.assertEqual(len(lignes_of), 1)
        self.assertEqual(lignes_of[0]['of_id'], of.id)
        self.assertEqual(len(lignes_operations), 1)
        self.assertEqual(lignes_operations[0]['of_id'], of.id)

    def test_filtre_statut(self):
        OrdreFabrication.objects.create(
            company=self.company, produit=self.produit, quantite=1,
            statut=OrdreFabrication.Statut.BROUILLON)
        of_planifie = OrdreFabrication.objects.create(
            company=self.company, produit=self.produit, quantite=1,
            statut=OrdreFabrication.Statut.PLANIFIE)
        lignes_of, _ = export_ordres_fabrication(
            self.company, statut=OrdreFabrication.Statut.PLANIFIE)
        self.assertEqual(
            [ligne['of_id'] for ligne in lignes_of], [of_planifie.id])

    def test_isolation_tenant(self):
        autre_company = make_company('mrp-ntmfg36-2', 'MRP NTMFG36 2')
        autre_produit = make_produit(autre_company, 'Autre produit')
        OrdreFabrication.objects.create(
            company=autre_company, produit=autre_produit, quantite=1)
        lignes_of, _ = export_ordres_fabrication(self.company)
        self.assertEqual(lignes_of, [])


class ExportOrdresFabricationApiTests(TestCase):
    def setUp(self):
        self.company = make_company('mrp-ntmfg36-api-1', 'MRP NTMFG36 API 1')
        self.responsable = make_user(
            self.company, 'mrp-ntmfg36-resp', role='responsable')
        self.technicien = make_user(
            self.company, 'mrp-ntmfg36-tech', role='normal')
        self.produit = make_produit(self.company)
        OrdreFabrication.objects.create(
            company=self.company, produit=self.produit, quantite=3)

    def test_export_xlsx_deux_feuilles(self):
        resp = auth(self.responsable).get(
            '/api/django/mrp/ordres-fabrication/export/?format=xlsx')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        import io

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(resp.content))
        self.assertEqual(
            wb.sheetnames, ['Ordres de fabrication', 'Opérations'])

    def test_export_csv(self):
        resp = auth(self.responsable).get(
            '/api/django/mrp/ordres-fabrication/export/?format=csv')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'text/csv')
        self.assertIn(b'Produit', resp.content)

    def test_format_invalide_400(self):
        resp = auth(self.responsable).get(
            '/api/django/mrp/ordres-fabrication/export/?format=pdf')
        self.assertEqual(resp.status_code, 400)

    def test_technicien_403(self):
        resp = auth(self.technicien).get(
            '/api/django/mrp/ordres-fabrication/export/?format=xlsx')
        self.assertEqual(resp.status_code, 403)
