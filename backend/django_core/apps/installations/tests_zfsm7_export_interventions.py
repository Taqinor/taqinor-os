"""ZFSM7 — Export xlsx de la liste des interventions (list export Odoo).

`InterventionViewSet` n'avait aucun export ; le pattern `?export=xlsx` via
openpyxl existe partout ailleurs. Couvre :

  * `?export=xlsx` produit un xlsx filtré correct (respecte `statut`) ;
  * aucune colonne coût/marge (`prix_achat`) n'apparaît ;
  * isolation tenant ;
  * rôle responsable/admin requis (technicien → 403).

Run :
    python manage.py test apps.installations.tests_zfsm7_export_interventions -v2
"""
import io
import itertools

from django.contrib.auth import get_user_model
from django.test import TestCase
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import AccessToken

from apps.crm.models import Client
from apps.installations.models import Installation, Intervention

User = get_user_model()
_seq = itertools.count(1)
BASE = '/api/django/installations'


def make_company():
    from authentication.models import Company
    n = next(_seq)
    company, _ = Company.objects.get_or_create(
        slug=f'zfsm7-co-{n}', defaults={'nom': f'ZFSM7 Co {n}'})
    return company


def make_user(company, role='responsable'):
    return User.objects.create_user(
        username=f'zfsm7-{next(_seq)}', password='x',
        role_legacy=role, company=company)


def auth(user):
    api = APIClient()
    api.credentials(HTTP_AUTHORIZATION=f'Bearer {AccessToken.for_user(user)}')
    return api


def make_installation(company, ville='Casablanca'):
    n = next(_seq)
    client = Client.objects.create(
        company=company, nom='Client', prenom='ZFSM7',
        email=f'zfsm7-{company.id}-{n}@example.invalid')
    return Installation.objects.create(
        company=company, reference=f'CHT-ZFSM7-{n}', client=client,
        statut=Installation.Statut.PLANIFIE, site_ville=ville)


class TestExportInterventionsXlsx(TestCase):
    def setUp(self):
        self.company = make_company()
        self.user = make_user(self.company)
        self.api = auth(self.user)
        self.inst = make_installation(self.company)
        self.interv1 = Intervention.objects.create(
            company=self.company, installation=self.inst,
            type_intervention=Intervention.Type.POSE,
            statut=Intervention.Statut.A_PREPARER)
        self.interv2 = Intervention.objects.create(
            company=self.company, installation=self.inst,
            type_intervention=Intervention.Type.DEPANNAGE,
            statut=Intervention.Statut.TERMINEE)

    def test_export_produit_un_xlsx(self):
        r = self.api.get(f'{BASE}/interventions/?export=xlsx')
        self.assertEqual(r.status_code, 200)
        self.assertIn('spreadsheetml', r['Content-Type'])

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        # En-tête + 2 interventions.
        self.assertEqual(len(rows), 3)
        header = rows[0]
        self.assertIn('Chantier', header)
        self.assertIn('Ville', header)

    def test_export_respecte_filtre_statut(self):
        r = self.api.get(
            f'{BASE}/interventions/?export=xlsx&statut=terminee')
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(len(rows), 2)  # en-tête + 1 seule intervention.

    def test_aucune_colonne_cout_ou_marge(self):
        r = self.api.get(f'{BASE}/interventions/?export=xlsx')
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content), read_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                cell_str = str(cell).lower()
                self.assertNotIn('prix_achat', cell_str)
                self.assertNotIn('marge', cell_str)
                self.assertNotIn('cout', cell_str)

    def test_isolation_tenant(self):
        other_company = make_company()
        other_user = make_user(other_company)
        other_api = auth(other_user)
        r = other_api.get(f'{BASE}/interventions/?export=xlsx')
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(len(rows), 1)  # en-tête seulement, aucune fuite.

    def test_role_technicien_refuse(self):
        tech = make_user(self.company, role='technicien')
        tech_api = auth(tech)
        r = tech_api.get(f'{BASE}/interventions/?export=xlsx')
        self.assertEqual(r.status_code, 403)
