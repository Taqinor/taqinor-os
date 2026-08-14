"""Services du module Marketing (``apps.marketing``).

ODX10 — ré-export TRANSITOIRE des fonctions de service marketing qui vivent
encore physiquement dans ``apps.compta.services`` (elles y étaient interleavées
avec la logique comptable ; les extraire du fichier de 500 Ko en un seul move
serait un risque de régression non vérifiable hors suite complète). Ce module
donne au reste du code (receivers marketing, urls, appelants cross-app) un point
d'accès ``apps.marketing.services`` stable ; ODX22 re-logera le corps des
fonctions ici et retirera ce shim.

``marketing`` ne lit crm/ventes QUE via leurs selectors/services ou par
référence opaque — jamais leurs ``models`` (invariant CLAUDE.md déjà tenu par
les fonctions ré-exportées, qui référencent lead_id/devis_id opaques).
"""

from decimal import Decimal

from apps.compta.services import (  # noqa: F401
    annuler_campagne,
    appliquer_mouvement_fidelite,
    approuver_envoi_campagne,
    campagnes_par_statut,
    clics_par_lien,
    compteurs_par_etape,
    cout_total_campagne,
    creer_enquete,
    decider_gagnant_ab,
    demander_ou_envoyer_campagne,
    dupliquer_campagne,
    enregistrer_ouverture_partage,
    enregistrer_relance_devis_abandonne,
    envelopper_liens_campagne,
    envoyer_campagne,
    envoyer_campagnes_planifiees,
    envoyer_enquete_nps,
    envoyer_test_campagne,
    executer_etapes_dues,
    inscrire_lead_sequence,
    leads_source_roi,
    nb_participants_actifs,
    participants_sequence,
    planifier_campagne,
    planifier_etapes_sequence,
    pousser_avis_google,
    precheck_sante_campagne,
    questions_visibles,
    recalculer_compteurs_campagne,
    rejeter_envoi_campagne,
    rendre_pour_lead,
    renvoyer_echecs_campagne,
    reporting_campagnes,
    repondre_enquete_nps,
    roi_campagne,
    score_nps,
    sortir_inscription,
    sortir_inscriptions_pour_lead,
    suggestions_upsell,
    traces_sequence,
    valider_questions_enquete,
    variante_pour_langue,
    webhook_brevo_evenement,
)


# ── WIR64 / FG206 — Capture de lead publique depuis un FormulaireIntake ──────
# Écrit dans le domaine crm UNIQUEMENT via ``apps.crm.services`` (jamais un
# import des modèles crm — invariant CLAUDE.md/M3). La société vient TOUJOURS
# du formulaire résolu côté serveur, jamais du corps de la requête publique.

# ── NTMKT22 — Centre de préférences self-service (public, tokenisé) ────────
# XMKT3 ne sait que DÉSINSCRIRE totalement. Ici le contact choisit PAR CANAL
# et PAR LISTE ce qu'il veut recevoir. Le registre de consentement est celui
# de la plateforme (``core.ConsentRecord``) — jamais un second registre — et
# les listes sont les ``AbonnementListe`` existantes (XMKT5).

_PREFERENCES_SALT = 'marketing.ntmkt22.preferences'

#: canaux exposés au contact, et leur finalité ``ConsentRecord`` (même
#: correspondance que XMKT4, pour que le prochain envoi respecte le choix).
CANAUX_PREFERENCES = (
    ('email', 'email'),
    ('sms', 'sms'),
    ('whatsapp', 'whatsapp'),
)


def generer_token_preferences(company_id, destinataire):
    """Jeton signé du centre de préférences (NTMKT22) — même modèle de
    confiance que le lien de désinscription XMKT3 (signature ``SECRET_KEY``,
    société + destinataire portés par le jeton, jamais par l'URL en clair)."""
    from django.core import signing
    return signing.dumps(
        {'company_id': company_id, 'destinataire': destinataire},
        salt=_PREFERENCES_SALT)


def lien_preferences(company, destinataire):
    """Chemin public du centre de préférences, à placer à côté du lien de
    désinscription dans le pied de campagne (``{lien_preferences}``)."""
    return f'/api/django/marketing/preferences/{generer_token_preferences(company.id, destinataire)}/'


#: NTMKT33 — durée de vie du jeton de préférences (90 jours), au-delà duquel
#: un lien partagé/oublié n'est plus utilisable — même modèle de confiance
#: que XMKT3 mais AVEC expiration (``django.core.signing`` porte déjà
#: l'horodatage de signature ; aucun enregistrement DB n'est nécessaire).
TOKEN_PREFERENCES_MAX_AGE_SECONDS = 90 * 24 * 60 * 60


def lire_token_preferences(token, *, max_age=TOKEN_PREFERENCES_MAX_AGE_SECONDS):
    """Résout un jeton de préférences en ``(company, destinataire)``.

    Jeton invalide/corrompu/société supprimée/expiré (>90j, NTMKT33) →
    ``(None, None)`` : l'appelant répond proprement, jamais une 500 ni une
    fuite d'existence.
    """
    from django.core import signing
    try:
        payload = signing.loads(token, salt=_PREFERENCES_SALT, max_age=max_age)
    except signing.BadSignature:
        return None, None
    from authentication.models import Company
    company = Company.objects.filter(id=payload.get('company_id')).first()
    destinataire = (payload.get('destinataire') or '').strip()
    if company is None or not destinataire:
        return None, None
    return company, destinataire


def preferences_actuelles(company, destinataire):
    """État courant des préférences d'un contact (NTMKT22).

    Un canal sans aucune entrée de consentement est considéré ACCORDÉ —
    comportement historique XMKT4 strictement préservé.
    """
    from core.models import ConsentRecord
    from .models import AbonnementListe, ListeDiffusion
    canaux = {}
    for canal, purpose in CANAUX_PREFERENCES:
        dernier = (ConsentRecord.objects
                   .filter(company=company, subject_identifier=destinataire,
                           purpose=purpose)
                   .order_by('-id').first())
        canaux[canal] = True if dernier is None else bool(dernier.granted)
    abonnements = {
        a.liste_id: a.statut
        for a in AbonnementListe.objects.filter(
            company=company, destinataire=destinataire)
    }
    listes = [
        {
            'id': liste.id,
            'nom': liste.nom,
            'abonne': abonnements.get(liste.id) == AbonnementListe.Statut.INSCRIT,
        }
        for liste in ListeDiffusion.objects.filter(company=company)
    ]
    return {'destinataire': destinataire, 'canaux': canaux, 'listes': listes}


def enregistrer_preferences(company, destinataire, data=None,
                            *, source='centre_preferences'):
    """Enregistre les choix du contact (NTMKT22).

    ``data`` = {'canaux': {'email': False, …}, 'listes': {'<id>': True, …}}.
    Chaque canal cité crée une entrée ``core.ConsentRecord`` (traçabilité
    loi 09-08 : le registre est un JOURNAL, jamais réécrit sur place) ; chaque
    liste citée bascule son ``AbonnementListe``. Une clé absente n'est pas
    touchée — couper l'email ne désinscrit jamais de WhatsApp.
    """
    from django.utils import timezone
    from core.models import ConsentRecord
    from .models import AbonnementListe, ListeDiffusion
    data = data or {}
    purposes = dict(CANAUX_PREFERENCES)
    choix_canaux = data.get('canaux') or {}
    for canal, valeur in choix_canaux.items():
        purpose = purposes.get(str(canal).strip().lower())
        if purpose is None:
            continue
        ConsentRecord.objects.create(
            company=company,
            subject_identifier=destinataire,
            purpose=purpose,
            granted=bool(valeur),
            source=source,
            occurred_at=timezone.now(),
        )
    choix_listes = data.get('listes') or {}
    for liste_id, valeur in choix_listes.items():
        try:
            liste_id = int(liste_id)
        except (TypeError, ValueError):
            continue
        liste = ListeDiffusion.objects.filter(
            company=company, id=liste_id).first()
        if liste is None:
            continue
        statut = (AbonnementListe.Statut.INSCRIT if valeur
                  else AbonnementListe.Statut.DESINSCRIT)
        AbonnementListe.objects.update_or_create(
            liste=liste, destinataire=destinataire,
            defaults={'company': company, 'statut': statut},
        )
    return preferences_actuelles(company, destinataire)


def derniere_version_publiee(formulaire):
    """NTMKT16 — dernière version PUBLIÉE d'une landing page, ou ``None``.

    La page publique n'affiche jamais un brouillon : tant qu'aucune version
    n'est publiée, le rendu reste EXACTEMENT celui d'avant NTMKT16.
    """
    return (formulaire.versions.filter(publie=True)
            .order_by('-version', '-id')
            .first())


def creer_version_formulaire(formulaire, data=None):
    """NTMKT16 — crée une nouvelle version BROUILLON du contenu de page.

    Le numéro de version vient d'un compteur PERSISTANT côté serveur
    (``FormulaireIntake.dernier_numero_version`` — plus haut numéro
    ATTRIBUÉ + 1, jamais un count()/Max() sur les versions restantes, qui
    régresserait après la suppression de la version la plus récente) et
    n'est jamais accepté du corps de la requête ; la société vient du
    formulaire.
    """
    from django.db import transaction
    from django.db.models import Max

    from .models import FormulaireIntake, VersionFormulaireIntake
    data = data or {}
    with transaction.atomic():
        # select_for_update : deux créations concurrentes sur le même
        # formulaire ne doivent jamais recevoir le même numéro.
        verrouille = (FormulaireIntake.objects
                      .select_for_update()
                      .get(pk=formulaire.pk))
        # Backfill défensif : une ligne créée avant NTMKT16 (ou dont le
        # compteur n'a jamais rattrapé des versions déjà existantes) part du
        # plus haut numéro déjà présent, jamais en dessous.
        plus_haute_existante = (
            verrouille.versions.aggregate(m=Max('version'))['m'] or 0)
        depart = max(verrouille.dernier_numero_version, plus_haute_existante)
        nouveau_numero = depart + 1
        FormulaireIntake.objects.filter(pk=formulaire.pk).update(
            dernier_numero_version=nouveau_numero)
        formulaire.dernier_numero_version = nouveau_numero
        version = VersionFormulaireIntake.objects.create(
            company=formulaire.company,
            formulaire=formulaire,
            version=nouveau_numero,
            titre=data.get('titre') or '',
            pitch=data.get('pitch') or '',
            image_key=data.get('image_key') or '',
            publie=False,
        )
    return version


def publier_version_formulaire(version):
    """NTMKT16 — publie une version : la page publique bascule dessus.

    Les versions antérieures restent consultables (historique complet) ;
    seule la PLUS RÉCENTE publiée est rendue.
    """
    from django.utils import timezone
    if not version.publie:
        version.publie = True
        version.date_publication = timezone.now()
        version.save(update_fields=['publie', 'date_publication'])
    return version


def formulaire_intake_actif_par_slug(slug):
    """WIR64 — résout un FormulaireIntake ACTIF par son slug public (lookup
    par slug, cf. NTMKT16). Renvoie le formulaire ou ``None``. La société est
    portée par le formulaire trouvé — jamais lue d'un paramètre public."""
    from .models import FormulaireIntake
    return (FormulaireIntake.objects
            .filter(slug=slug, actif=True)
            .order_by('id')
            .first())


def creer_lead_depuis_intake(formulaire, data):
    """WIR64/FG206 — crée un lead depuis la soumission publique d'un
    ``FormulaireIntake``, via ``crm.services`` (jamais d'import des modèles
    crm). Le ``type_installation`` par défaut du formulaire pré-remplit le
    lead ; le ``tag_prefill`` est posé comme tag après création. ``nom`` est
    obligatoire (``ValueError`` sinon, remonté en 400 par la vue). La société
    est forcée depuis ``formulaire.company``."""
    from apps.crm.services import create_lead_from_public_api, poser_tag_lead
    data = data or {}
    fields = {
        'nom': data.get('nom'),
        'prenom': data.get('prenom'),
        'societe': data.get('societe'),
        'email': data.get('email'),
        'telephone': data.get('telephone'),
        'ville': data.get('ville'),
    }
    if formulaire.type_installation:
        fields['type_installation'] = formulaire.type_installation
    lead = create_lead_from_public_api(company=formulaire.company, fields=fields)
    if formulaire.tag_prefill:
        poser_tag_lead(lead, None, formulaire.tag_prefill)
    return lead


# ── NTMKT17 — Progressive profiling (champs déjà connus masqués) ───────────

def _champ_code(entree):
    if isinstance(entree, dict):
        return entree.get('code') or entree.get('nom')
    return entree


def champs_publics_a_afficher(formulaire, identifiant):
    """NTMKT17 — filtre ``formulaire.champs`` pour un visiteur RECONNU.

    ``identifiant`` (email OU téléphone) vient du navigateur du visiteur
    (cookie/stockage déjà géré côté client, aucune nouvelle dépendance
    backend). Un visiteur INCONNU (``identifiant`` vide, ou aucun lead
    correspondant — dédup QJ8) voit le formulaire COMPLET, comportement
    actuel inchangé. Un visiteur RECONNU ne revoit QUE les champs non encore
    renseignés sur son lead le plus récent — HubSpot-style."""
    champs = formulaire.champs or []
    identifiant = (identifiant or '').strip()
    if not identifiant:
        return champs

    from apps.crm.selectors import lead_known_field_codes

    if '@' in identifiant:
        connus = lead_known_field_codes(formulaire.company, email=identifiant)
    else:
        connus = lead_known_field_codes(formulaire.company, phone=identifiant)
    if not connus:
        return champs
    return [c for c in champs if _champ_code(c) not in connus]


# ── NTMKT12 — Parcours en GRAPHE d'une séquence de relance ──────────────────
# Extension strictement ADDITIVE du moteur XMKT1 : une séquence SANS nœud n'est
# jamais vue par ce code (``sequence_a_graphe`` faux) et continue d'être
# déroulée à l'octet par ``compta.services.executer_etapes_dues``. Une séquence
# AVEC nœuds n'a pas d'``EtapeSequence`` : ses inscriptions portent
# ``etape_courante=None`` et sont donc naturellement ignorées par le moteur
# linéaire (qui filtre sur ``etape_courante__isnull=False``) — aucune double
# exécution possible.

# Garde-fou anti-boucle : un cycle dans le graphe ne doit jamais faire tourner
# le tick indéfiniment.
MAX_NOEUDS_PAR_TICK = 50


def sequence_a_graphe(sequence):
    """La séquence porte-t-elle un graphe de journey (NTMKT12) ?"""
    from .models import NoeudJourney
    return NoeudJourney.objects.filter(sequence=sequence).exists()


def noeud_initial(sequence):
    """Nœud d'entrée du graphe : le premier ``declencheur``, à défaut le nœud
    de plus petit id. ``None`` si la séquence n'a pas de graphe."""
    from .models import NoeudJourney
    noeuds = NoeudJourney.objects.filter(sequence=sequence).order_by('id')
    return (noeuds.filter(type_noeud=NoeudJourney.Type.DECLENCHEUR).first()
            or noeuds.first())


def _lead_du_parcours(inscription):
    """Lead de l'inscription, lu via le selector crm (jamais ses models)."""
    from apps.crm.selectors import get_company_lead
    try:
        return get_company_lead(inscription.company, inscription.lead_id)
    except Exception:
        return None


def _condition_arc_vraie(inscription, arc):
    """NTMKT12 — évalue la condition d'un arc sur les traces EXISTANTES
    (``EnvoiCampagne`` XMKT2 pour l'ouverture/le clic) et sur le lead lu via
    ``crm.selectors`` (score/tag). Toute condition inconnue est fausse."""
    from .models import ArcJourney, EnvoiCampagne
    condition = arc.condition or ArcJourney.Condition.TOUJOURS
    if condition == ArcJourney.Condition.TOUJOURS:
        return True
    if condition in (ArcJourney.Condition.A_OUVERT,
                     ArcJourney.Condition.A_CLIQUE):
        champ = ('ouvert_le' if condition == ArcJourney.Condition.A_OUVERT
                 else 'clique_le')
        filtre = {
            f'{champ}__isnull': False,
            f'{champ}__gte': inscription.declenchee_le,
        }
        return EnvoiCampagne.objects.filter(
            company=inscription.company,
            contact_ref=f'lead:{inscription.lead_id}',
            **filtre).exists()
    lead = _lead_du_parcours(inscription)
    if lead is None:
        return False
    if condition == ArcJourney.Condition.SCORE_SEUIL:
        try:
            seuil = float(arc.valeur)
        except (TypeError, ValueError):
            return False
        return float(getattr(lead, 'score', 0) or 0) >= seuil
    if condition == ArcJourney.Condition.TAG_PRESENT:
        tag = (arc.valeur or '').strip().lower()
        if not tag:
            return False
        tags = (getattr(lead, 'tags', '') or '').lower()
        return tag in [t.strip() for t in tags.split(',') if t.strip()]
    return False


def arc_suivant(inscription, noeud):
    """Premier arc sortant (par ``ordre``) dont la condition est vraie."""
    for arc in noeud.arcs_sortants.all().order_by('ordre', 'id'):
        if _condition_arc_vraie(inscription, arc):
            return arc
    return None


def prochaine_echeance_ouvree(depuis, company, *, heure=9, jour_semaine=None,
                              max_jours=60):
    """NTMKT14 — prochain créneau OUVRÉ à ``heure`` strictement après
    ``depuis``.

    Les jours fériés / non ouvrés et la nuit viennent de
    ``apps.notifications.selectors.est_hors_fenetre_silence`` (jamais un
    import des modèles ``notifications``). ``jour_semaine`` (0 = lundi) permet
    « le prochain lundi 9h » ; sans lui, c'est le prochain jour ouvré.
    """
    from django.utils import timezone
    from apps.notifications.selectors import est_hors_fenetre_silence
    heure = max(0, min(23, int(heure or 0)))
    candidat = timezone.localtime(depuis).replace(
        hour=heure, minute=0, second=0, microsecond=0)
    if candidat <= timezone.localtime(depuis):
        candidat += timezone.timedelta(days=1)
    for _ in range(max_jours):
        jour_ok = (jour_semaine is None
                   or candidat.weekday() == int(jour_semaine))
        if jour_ok and not est_hors_fenetre_silence(candidat, company):
            return candidat
        candidat += timezone.timedelta(days=1)
    return candidat


def echeance_noeud(inscription, noeud):
    """Instant à partir duquel le nœud d'attente est franchissable.

    ``attente``         = J+``config.delai_jours`` depuis l'entrée sur le nœud.
    ``attente_jusqu_a`` = NTMKT14, selon ``config.mode`` :
        * ``date``      -> date/heure absolue (``config.date``, ISO) ;
        * sinon         -> prochain créneau ouvré (``config.heure``, défaut 9h,
                           et ``config.jour_semaine`` optionnel, 0 = lundi).
    Tout autre type est franchissable immédiatement.
    """
    from django.utils import timezone
    from django.utils.dateparse import parse_datetime
    from .models import NoeudJourney
    depuis = inscription.noeud_depuis or inscription.declenchee_le
    config = noeud.config or {}
    if noeud.type_noeud == NoeudJourney.Type.ATTENTE:
        try:
            jours = int(config.get('delai_jours') or 0)
        except (TypeError, ValueError):
            jours = 0
        return depuis + timezone.timedelta(days=max(jours, 0))
    if noeud.type_noeud == NoeudJourney.Type.ATTENTE_JUSQU_A:
        if (config.get('mode') or '') == 'date':
            brut = parse_datetime(str(config.get('date') or '')) or None
            if brut is None:
                return depuis
            if timezone.is_naive(brut):
                brut = timezone.make_aware(
                    brut, timezone.get_current_timezone())
            return brut
        jour_semaine = config.get('jour_semaine')
        try:
            jour_semaine = (None if jour_semaine in (None, '')
                            else int(jour_semaine))
        except (TypeError, ValueError):
            jour_semaine = None
        return prochaine_echeance_ouvree(
            depuis, inscription.company,
            heure=config.get('heure', 9), jour_semaine=jour_semaine)
    return depuis


def _tracer_noeud(inscription, noeud, *, resultat='planifie', canal='',
                  erreur=''):
    """Trace l'exécution d'un nœud dans le journal EXISTANT
    (``ExecutionEtapeSequence``) — jamais un second registre de traces."""
    from .models import ExecutionEtapeSequence
    return ExecutionEtapeSequence.objects.create(
        company=inscription.company,
        inscription=inscription,
        etape=None,
        noeud=noeud,
        canal=canal or '',
        resultat=resultat,
        erreur=erreur or '',
    )


def _positionner(inscription, noeud, *, maintenant):
    """Place l'inscription sur ``noeud`` (ou la termine si ``None``)."""
    from .models import InscriptionSequence
    if noeud is None:
        inscription.noeud_courant = None
        inscription.noeud_depuis = None
        inscription.statut = InscriptionSequence.Statut.TERMINE
        inscription.save(update_fields=[
            'noeud_courant', 'noeud_depuis', 'statut'])
        return
    inscription.noeud_courant = noeud
    inscription.noeud_depuis = maintenant
    inscription.save(update_fields=['noeud_courant', 'noeud_depuis'])


def avancer_journey(inscription, *, maintenant=None):
    """Fait avancer UNE inscription dans le graphe de sa séquence (NTMKT12).

    Renvoie la liste des traces créées. S'arrête sur un nœud d'attente non
    échu, sur un nœud de sortie, ou quand aucun arc sortant n'est empruntable
    (fin de parcours → inscription terminée).
    """
    from django.utils import timezone
    from .models import InscriptionSequence, NoeudJourney
    maintenant = maintenant or timezone.now()
    traces = []
    if inscription.statut != InscriptionSequence.Statut.ACTIF:
        return traces
    if inscription.noeud_courant is None:
        depart = noeud_initial(inscription.sequence)
        if depart is None:
            return traces
        _positionner(inscription, depart, maintenant=maintenant)
    for _ in range(MAX_NOEUDS_PAR_TICK):
        noeud = inscription.noeud_courant
        if noeud is None:
            break
        if noeud.type_noeud == NoeudJourney.Type.SORTIE:
            inscription.noeud_courant = None
            inscription.noeud_depuis = None
            inscription.statut = InscriptionSequence.Statut.TERMINE
            inscription.save(update_fields=[
                'noeud_courant', 'noeud_depuis', 'statut'])
            break
        if noeud.type_noeud in (NoeudJourney.Type.ATTENTE,
                                NoeudJourney.Type.ATTENTE_JUSQU_A):
            if maintenant < echeance_noeud(inscription, noeud):
                break
        elif noeud.type_noeud == NoeudJourney.Type.ACTION:
            config = noeud.config or {}
            traces.append(_tracer_noeud(
                inscription, noeud, canal=str(config.get('canal') or '')))
        arc = arc_suivant(inscription, noeud)
        _positionner(inscription, arc.cible if arc else None,
                     maintenant=maintenant)
        if arc is None:
            break
    return traces


def instancier_modele_journey(company, modele, *, nom=None,
                              stage_declencheur=''):
    """NTMKT15 — instancie un ``ModeleJourney`` en une NOUVELLE séquence
    éditable (nœuds + arcs NTMKT12), désactivée par défaut.

    Le graphe du modèle référence ses nœuds par une ``cle`` textuelle ; la
    correspondance clé → nœud créé est locale à cette instanciation (aucune
    fuite entre modèles, aucune modification du modèle source).
    """
    from .models import ArcJourney, NoeudJourney, SequenceRelance
    graphe = modele.graphe or {}
    sequence = SequenceRelance.objects.create(
        company=company,
        nom=nom or modele.nom,
        stage_declencheur=stage_declencheur or '',
        actif=False,
    )
    par_cle = {}
    for brut in graphe.get('noeuds') or []:
        noeud = NoeudJourney.objects.create(
            company=company,
            sequence=sequence,
            type_noeud=brut.get('type_noeud') or NoeudJourney.Type.ACTION,
            libelle=brut.get('libelle') or '',
            position_x=int(brut.get('position_x') or 0),
            position_y=int(brut.get('position_y') or 0),
            config=brut.get('config') or {},
        )
        if brut.get('cle'):
            par_cle[str(brut['cle'])] = noeud
    for brut in graphe.get('arcs') or []:
        source = par_cle.get(str(brut.get('source')))
        cible = par_cle.get(str(brut.get('cible')))
        if source is None or cible is None:
            continue
        ArcJourney.objects.create(
            company=company,
            source=source,
            cible=cible,
            condition=(brut.get('condition')
                       or ArcJourney.Condition.TOUJOURS),
            valeur=brut.get('valeur') or '',
            ordre=int(brut.get('ordre') or 1),
        )
    return sequence


def executer_journeys_dus(company, *, maintenant=None):
    """Tick des séquences EN GRAPHE d'une société (NTMKT12).

    Pendant graphe de ``compta.services.executer_etapes_dues`` (linéaire), qui
    reste la seule voie pour les séquences sans nœud.
    """
    from django.utils import timezone
    from .models import InscriptionSequence, NoeudJourney
    maintenant = maintenant or timezone.now()
    sequences_graphe = set(
        NoeudJourney.objects.filter(company=company)
        .values_list('sequence_id', flat=True))
    if not sequences_graphe:
        return []
    traces = []
    inscriptions = InscriptionSequence.objects.filter(
        company=company,
        statut=InscriptionSequence.Statut.ACTIF,
        sequence_id__in=sequences_graphe,
    ).select_related('sequence', 'noeud_courant')
    for inscription in inscriptions:
        traces.extend(avancer_journey(inscription, maintenant=maintenant))
    return traces


# ── NTMKT26 — Import de coûts publicitaires externes (Meta/Google Ads) ─────
# Aucun appel API externe (pas de jeton requis) : un fichier CSV exporté à la
# main depuis Meta Ads Manager / Google Ads est importé et réconcilié par nom
# de campagne — réutilise le parseur d'en-têtes de ``apps.dataimport``
# (import function-local, lecture seule, jamais un target du registre
# ``dataimport`` puisqu'on ne crée/maj aucun modèle dataimport ici).

#: Alias de colonnes CSV tolérés (Meta Ads Manager / Google Ads exports natifs).
_COLONNES_NOM_CAMPAGNE = ('nom_campagne', 'campaign name', 'campaign', 'nom')
_COLONNES_COUT = ('cout', 'cout_mad', 'amount spent', 'amount spent (mad)',
                  'cost', 'montant')


def _colonne_normalisee(headers, alias):
    for h in headers:
        if (h or '').strip().lower() in alias:
            return h
    return None


def importer_couts_publicitaires(company, file_bytes, filename):
    """NTMKT26 — importe un CSV de coûts publicitaires et met à jour
    ``Campagne.cout_reel_mad`` par correspondance de NOM (insensible à la
    casse/aux espaces).

    Renvoie un rapport ``{'matched': [...], 'unmatched': [...]}`` — jamais
    d'exception sur une ligne malformée (elle finit simplement en
    ``unmatched``).
    """
    from apps.dataimport.parsing import iter_rows

    from .models import Campagne

    headers, rows = iter_rows(file_bytes, filename)
    col_nom = _colonne_normalisee(headers, _COLONNES_NOM_CAMPAGNE)
    col_cout = _colonne_normalisee(headers, _COLONNES_COUT)
    matched, unmatched = [], []
    if col_nom is None or col_cout is None:
        return {'matched': matched, 'unmatched': unmatched,
                'erreur': 'colonnes nom/coût introuvables dans le CSV'}

    campagnes_par_nom = {
        c.nom.strip().lower(): c
        for c in Campagne.objects.filter(company=company)
    }
    for row in rows:
        nom_brut = (row.get(col_nom) or '').strip()
        cout_brut = (row.get(col_cout) or '').strip()
        campagne = campagnes_par_nom.get(nom_brut.lower())
        if campagne is None:
            unmatched.append({'nom_campagne': nom_brut, 'cout': cout_brut,
                              'raison': 'aucune campagne de ce nom'})
            continue
        try:
            montant = Decimal(cout_brut.replace(',', '.').replace(' ', ''))
        except Exception:
            unmatched.append({'nom_campagne': nom_brut, 'cout': cout_brut,
                              'raison': 'coût illisible'})
            continue
        campagne.cout_reel_mad = montant
        campagne.save(update_fields=['cout_reel_mad'])
        matched.append({'campagne_id': campagne.id, 'nom_campagne': campagne.nom,
                        'cout_reel_mad': str(montant)})
    return {'matched': matched, 'unmatched': unmatched}


# ── NTMKT27 — Rapport imprimable « Bilan de campagne » (PDF interne) ───────

def rapport_campagne_donnees(campagne):
    """Données du bilan PDF (NTMKT27) : entonnoir, top liens, coût/ROI.

    JAMAIS ``Produit.prix_achat`` (hors sujet marketing, de toute façon) — le
    PDF reste strictement un bilan de campagne, aucune marge produit.
    """
    liens = sorted(
        clics_par_lien(campagne), key=lambda lien: -lien['nb_clics'])[:10]
    roi = roi_campagne(campagne)
    return {
        'campagne': campagne,
        'entonnoir': {
            'envoyes': campagne.nb_envois,
            'ouverts': campagne.nb_ouvertures,
            'cliques': campagne.nb_clics,
            'convertis': roi.get('nb_signes', 0) if isinstance(roi, dict) else 0,
        },
        'top_liens': liens,
        'roi': roi,
    }


def rapport_campagne_pdf(campagne):
    """Rend le bilan de campagne (NTMKT27) en PDF via ``core.pdf.render_pdf``
    (même moteur WeasyPrint que ``reporting``/``compta`` — jamais
    ``quote_engine``, règle #4)."""
    from html import escape

    from core.pdf import render_pdf

    donnees = rapport_campagne_donnees(campagne)
    ent = donnees['entonnoir']
    roi = donnees['roi'] or {}
    lignes_liens = ''.join(
        f"<tr><td>{escape(lien['url_cible'][:70])}</td>"
        f"<td style='text-align:right'>{lien['nb_clics']}</td></tr>"
        for lien in donnees['top_liens']
    ) or "<tr><td colspan='2'>Aucun lien tracké</td></tr>"
    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><style>
  body {{ font-family: sans-serif; margin: 40px; color: #222; }}
  h1 {{ font-size: 22px; }} h2 {{ font-size: 15px; margin-top: 28px; }}
  table {{ width: 100%; border-collapse: collapse; margin-top: 8px; }}
  td, th {{ border-bottom: 1px solid #ddd; padding: 4px 6px; font-size: 12px; }}
  .watermark {{ position: fixed; top: 40%; left: 15%; font-size: 48px;
               color: #eee; transform: rotate(-25deg); z-index: -1; }}
</style></head><body>
  <div class="watermark">USAGE INTERNE</div>
  <h1>Bilan de campagne — {escape(campagne.nom)}</h1>
  <h2>Entonnoir</h2>
  <table>
    <tr><td>Envoyés</td><td style="text-align:right">{ent['envoyes']}</td></tr>
    <tr><td>Ouverts</td><td style="text-align:right">{ent['ouverts']}</td></tr>
    <tr><td>Cliqués</td><td style="text-align:right">{ent['cliques']}</td></tr>
    <tr><td>Convertis</td><td style="text-align:right">{ent['convertis']}</td></tr>
  </table>
  <h2>Top 10 liens trackés</h2>
  <table><tr><th>URL</th><th>Clics</th></tr>{lignes_liens}</table>
  <h2>Coût réel vs revenu attribué</h2>
  <table>
    <tr><td>Coût réel (MAD)</td><td style="text-align:right">
    {escape(str(roi.get('cout_mad', '0')))}</td></tr>
    <tr><td>Revenu attribué (MAD)</td><td style="text-align:right">
    {escape(str(roi.get('revenu_ttc_mad', '0')))}</td></tr>
    <tr><td>ROI (%)</td><td style="text-align:right">
    {escape(str(roi.get('roi_pct', '0')))}</td></tr>
  </table>
</body></html>"""
    return render_pdf(html=html)


# ── NTMKT28 — Rapport imprimable « Registre de consentement » (export CNDP) ─
# Lecture seule sur ``core.ConsentRecord``/``SuppressionMarketing`` — jamais
# un second registre créé.

def registre_consentement_export(company, *, date_debut=None, date_fin=None,
                                 contact=None):
    """NTMKT28 — entrées du registre de consentement de la société sur une
    période (recevable pour un contrôle CNDP), filtrable par contact.

    Lecture seule, bornée société — aucune fuite inter-sociétés possible
    (filtre ``company=company`` systématique).
    """
    from core.models import ConsentRecord

    from .models import SuppressionMarketing

    qs = ConsentRecord.objects.filter(company=company).order_by('-id')
    if date_debut:
        qs = qs.filter(created_at__date__gte=date_debut)
    if date_fin:
        qs = qs.filter(created_at__date__lte=date_fin)
    if contact:
        qs = qs.filter(subject_identifier__icontains=contact)
    entrees = [
        {
            'subject_identifier': r.subject_identifier,
            'purpose': r.purpose,
            'granted': r.granted,
            'source': r.source,
            'version_texte': r.version_texte,
            'date_collecte': r.occurred_at or r.created_at,
        }
        for r in qs
    ]
    suppressions_qs = SuppressionMarketing.objects.filter(company=company)
    if date_debut:
        suppressions_qs = suppressions_qs.filter(date_creation__date__gte=date_debut)
    if date_fin:
        suppressions_qs = suppressions_qs.filter(date_creation__date__lte=date_fin)
    if contact:
        suppressions_qs = suppressions_qs.filter(destinataire__icontains=contact)
    suppressions = [
        {
            'destinataire': s.destinataire,
            'motif': s.motif,
            'source': s.source,
            'date_retrait': s.date_creation,
        }
        for s in suppressions_qs.order_by('-date_creation')
    ]
    return {'consentements': entrees, 'suppressions': suppressions}


def registre_consentement_pdf(company, *, date_debut=None, date_fin=None,
                              contact=None):
    """NTMKT28 — export PDF du registre de consentement (contrôle CNDP)."""
    from html import escape

    from core.pdf import render_pdf

    donnees = registre_consentement_export(
        company, date_debut=date_debut, date_fin=date_fin, contact=contact)
    lignes = ''.join(
        f"<tr><td>{escape(e['subject_identifier'])}</td>"
        f"<td>{escape(e['purpose'])}</td>"
        f"<td>{'Accordé' if e['granted'] else 'Retiré'}</td>"
        f"<td>{escape(e['source'])}</td>"
        f"<td>{escape(str(e['date_collecte']))}</td></tr>"
        for e in donnees['consentements']
    ) or "<tr><td colspan='5'>Aucune entrée sur la période</td></tr>"
    lignes_suppr = ''.join(
        f"<tr><td>{escape(s['destinataire'])}</td>"
        f"<td>{escape(s['motif'])}</td>"
        f"<td>{escape(str(s['date_retrait']))}</td></tr>"
        for s in donnees['suppressions']
    ) or "<tr><td colspan='3'>Aucune entrée sur la période</td></tr>"
    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><style>
  body {{ font-family: sans-serif; margin: 40px; color: #222; }}
  h1 {{ font-size: 20px; }} h2 {{ font-size: 14px; margin-top: 24px; }}
  table {{ width: 100%; border-collapse: collapse; margin-top: 8px; }}
  td, th {{ border: 1px solid #ddd; padding: 4px 6px; font-size: 11px; }}
</style></head><body>
  <h1>Registre de consentement — export CNDP</h1>
  <p>Traitement : marketing direct (email/SMS/WhatsApp) — base légale :
  consentement (loi 09-08).</p>
  <h2>Consentements ({len(donnees['consentements'])})</h2>
  <table><tr><th>Contact</th><th>Finalité</th><th>État</th><th>Source</th>
  <th>Date</th></tr>{lignes}</table>
  <h2>Désinscriptions / suppressions ({len(donnees['suppressions'])})</h2>
  <table><tr><th>Contact</th><th>Motif</th><th>Date</th></tr>
  {lignes_suppr}</table>
</body></html>"""
    return render_pdf(html=html)


# ── NTMKT31 — Réglages tenant « Marketing » ─────────────────────────────────

def parametres_marketing_pour(company):
    """Renvoie (en le créant si besoin) l'unique ``ParametresMarketing`` de la
    société — get_or_create côté service, jamais géré par l'écran (NTMKT31)."""
    from .models import ParametresMarketing
    obj, _ = ParametresMarketing.objects.get_or_create(company=company)
    return obj


def plafond_envois_atteint(company, *, aujourdhui=None):
    """NTMKT31 — True si le plafond d'envois/jour configuré est atteint ou
    dépassé pour AUJOURD'HUI. Plafond NULL/0 = désactivé (comportement actuel,
    jamais bloquant) — no-op par défaut tant que rien n'est configuré."""
    from django.utils import timezone

    from .models import EnvoiCampagne

    parametres = parametres_marketing_pour(company)
    if not parametres.plafond_envois_jour:
        return False
    aujourdhui = aujourdhui or timezone.localdate()
    nb = EnvoiCampagne.objects.filter(
        company=company, envoye_le__date=aujourdhui).count()
    return nb >= parametres.plafond_envois_jour


# ── NTMKT18 — Score de maturité marketing multi-signal (additif à QJ6) ─────
# ``lead_id`` reste OPAQUE : jamais d'import de ``crm.Lead`` ici. Le lien vers
# le lead se fait uniquement via la convention ``contact_ref = f'lead:{id}'``
# déjà utilisée par ``EnvoiCampagne`` (XMKT6/XMKT16) et via
# ``apps.crm.selectors.lead_devis_ids_by_id`` (lecture cross-app sanctionnée).

def _signaux_maturite(company, lead_id):
    """Compte les signaux marketing bruts d'un lead (NTMKT18) : ouvertures,
    clics (sur les envois de campagne le référençant) et visites de sa page
    proposition (``OuverturePartage`` sur ses devis, via crm.selectors)."""
    from .models import EnvoiCampagne, OuverturePartage

    contact_ref = f'lead:{lead_id}'
    envois = EnvoiCampagne.objects.filter(company=company, contact_ref=contact_ref)
    nb_ouvertures = envois.filter(ouvert_le__isnull=False).count()
    nb_clics = envois.filter(clique_le__isnull=False).count()

    from apps.crm.selectors import lead_devis_ids_by_id
    devis_ids = lead_devis_ids_by_id(company, lead_id)
    nb_visites = 0
    if devis_ids:
        nb_visites = OuverturePartage.objects.filter(
            company=company, cible=OuverturePartage.Cible.DEVIS,
            cible_reference__in=devis_ids).count()
    return nb_ouvertures, nb_clics, nb_visites


def recalculer_score_maturite(company, lead_id, *, dernier_contact=None,
                              now=None):
    """NTMKT18 — recalcule (jamais un delta incrémental non rejouable) le
    score de maturité (0-100) d'un lead à partir de ses signaux marketing
    bruts, pondérés par les réglages société (``ParametresMarketing``).

    No-op complet (renvoie ``None``, AUCUNE ligne créée) tant que la société
    n'a pas activé ``score_maturite_actif`` — comportement actuel inchangé
    par défaut. Applique la pénalité d'inactivité 30j (NTMKT34) si
    ``dernier_contact`` est fourni. Journalise une ``VariationScoreMaturite``
    UNIQUEMENT quand la valeur change effectivement (jamais un historique qui
    grossit sans raison)."""
    from django.utils import timezone as _tz

    from .models import ScoreMaturite, VariationScoreMaturite

    parametres = parametres_marketing_pour(company)
    if not parametres.score_maturite_actif:
        return None

    now = now or _tz.now()
    nb_ouvertures, nb_clics, nb_visites = _signaux_maturite(company, lead_id)
    valeur = (
        nb_ouvertures * parametres.ponderation_maturite_ouverture
        + nb_clics * parametres.ponderation_maturite_clic
        + nb_visites * parametres.ponderation_maturite_visite_proposition
    )
    motif_parts = [
        f'{nb_ouvertures} ouverture(s)', f'{nb_clics} clic(s)',
        f'{nb_visites} visite(s) proposition',
    ]
    if dernier_contact is not None:
        jours_inactif = (now.date() - dernier_contact.date()).days
        if jours_inactif >= 30:
            valeur -= parametres.penalite_maturite_inactivite
            motif_parts.append(f'inactif {jours_inactif}j')
    valeur = max(0, min(100, valeur))

    score, cree = ScoreMaturite.objects.get_or_create(
        company=company, lead_id=lead_id, defaults={'valeur': valeur})
    if not cree and score.valeur != valeur:
        delta = valeur - score.valeur
        score.valeur = valeur
        score.save(update_fields=['valeur', 'updated_at'])
        VariationScoreMaturite.objects.create(
            company=company, lead_id=lead_id, delta=delta,
            valeur_apres=valeur, motif=', '.join(motif_parts))
    elif cree and valeur:
        VariationScoreMaturite.objects.create(
            company=company, lead_id=lead_id, delta=valeur,
            valeur_apres=valeur, motif=', '.join(motif_parts))
    return score


def score_maturite_pour(company, lead_id):
    """NTMKT18 — recalcule puis renvoie le ``ScoreMaturite`` courant d'un
    lead (jamais une valeur périmée). ``None`` si le module est désactivé
    pour la société (comportement par défaut)."""
    return recalculer_score_maturite(company, lead_id)


def historique_maturite(company, lead_id, limite=30):
    """NTMKT18/NTMKT19 — historique horodaté (le plus récent en premier) des
    variations du score de maturité d'un lead."""
    from .models import VariationScoreMaturite
    return list(VariationScoreMaturite.objects.filter(
        company=company, lead_id=lead_id).order_by('-created_at')[:limite])


# ── NTMKT34 — Recalcul quotidien du score de maturité (pénalité inactivité) ─
# Le calcul ÉVÉNEMENTIEL de NTMKT18 (``recalculer_score_maturite``) ne couvre
# QUE les événements ENTRANTS (ouverture/clic/visite) : il ne réagit jamais au
# SILENCE d'un lead. Cette tâche quotidienne applique la pénalité
# d'inactivité 30j sur les leads qui ont DÉJÀ un ``ScoreMaturite`` (créé par
# un premier événement NTMKT18) — jamais un balayage de tous les leads de la
# société (no-op complet si NTMKT18 n'a jamais rien créé, cf.
# ``score_maturite_actif``).

def recalculer_scores_maturite_inactivite(company, *, now=None):
    """NTMKT34 — recalcule le score de maturité de chaque lead PORTANT DÉJÀ
    un ``ScoreMaturite`` pour ``company``, en tenant compte du dernier point
    de contact (``apps.crm.selectors.dernier_contact_lead``) pour appliquer
    la pénalité d'inactivité 30j. Émet ``core.events.lead_maturite_changee``
    pour CHAQUE changement effectif (jamais à un tick no-op). Renvoie la
    liste des ``lead_id`` dont le score a changé."""
    from django.utils import timezone as _tz

    from core import events

    from apps.crm.selectors import dernier_contact_lead

    from .models import ScoreMaturite

    now = now or _tz.now()
    changes = []
    for score in ScoreMaturite.objects.filter(company=company):
        avant = score.valeur
        dernier_contact = dernier_contact_lead(company, score.lead_id)
        recalcule = recalculer_score_maturite(
            company, score.lead_id, dernier_contact=dernier_contact, now=now)
        apres = recalcule.valeur if recalcule is not None else avant
        if apres != avant:
            changes.append(score.lead_id)
            events.lead_maturite_changee.send(
                sender='marketing.recalculer_scores_maturite_inactivite',
                lead_id=score.lead_id, company=company,
                ancienne_valeur=avant, nouvelle_valeur=apres)
    return changes


# ── NTMKT20 — Modèles d'attribution configurables (étend FG204/XMKT17) ─────

def modele_attribution_pour(company):
    """NTMKT20 — modèle d'attribution configuré pour la société (défaut
    ``dernier_touche`` = comportement XMKT17 actuel, inchangé)."""
    return parametres_marketing_pour(company).modele_attribution


def _cellule_export(valeur):
    """NTMKT39/40 — rendu d'une valeur pour une cellule XLSX (même esprit que
    ``apps.dataimport.exporters._cell``, sans en dépendre — ce module reste
    autonome)."""
    import datetime
    from decimal import Decimal

    if valeur is None:
        return ''
    if isinstance(valeur, (datetime.datetime, datetime.date)):
        return valeur.isoformat()
    if isinstance(valeur, Decimal):
        return str(valeur)
    return valeur


# ── NTMKT39 — Export CSV/XLSX des campagnes et de leur trace d'envoi ───────

_EXPORT_CAMPAGNES_COLONNES = [
    ('nom', 'Nom'), ('canal', 'Canal'), ('statut', 'Statut'),
    ('nb_destinataires', 'Destinataires'), ('nb_envois', 'Envoyés'),
    ('nb_ouvertures', 'Ouvertures'), ('nb_clics', 'Clics'),
    ('cout_reel_mad', 'Coût réel (MAD)'), ('date_creation', 'Créée le'),
]


def export_campagnes_xlsx(company, *, statut=None, canal=None):
    """NTMKT39 — export XLSX des campagnes de la société, MÊMES colonnes que
    la liste (``CampagnesList.jsx``), filtrable par statut/canal comme
    l'écran. ``openpyxl`` est déjà une dépendance du repo (cf.
    ``apps.dataimport.parsing``) — aucune nouvelle dépendance ajoutée."""
    import io

    from openpyxl import Workbook

    from .models import Campagne

    qs = Campagne.objects.filter(company=company).order_by('-date_creation')
    if statut:
        qs = qs.filter(statut=statut)
    if canal:
        qs = qs.filter(canal=canal)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Campagnes'
    ws.append([libelle for _, libelle in _EXPORT_CAMPAGNES_COLONNES])
    for campagne in qs:
        ws.append([
            _cellule_export(getattr(campagne, champ))
            for champ, _ in _EXPORT_CAMPAGNES_COLONNES
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_envois_campagne_csv(campagne):
    """NTMKT39 — export CSV de la trace ``EnvoiCampagne`` d'UNE campagne
    (destinataire/statut/date_envoi/date_ouverture), correctement échappé
    (virgules/accents) via le module ``csv`` standard — jamais une nouvelle
    lib."""
    import csv
    import io

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(
        ['Destinataire', 'Statut', 'Envoyé le', 'Ouvert le', 'Cliqué le'])
    for envoi in campagne.envois.all().order_by('-date_creation'):
        writer.writerow([
            envoi.destinataire, envoi.statut,
            envoi.envoye_le.isoformat() if envoi.envoye_le else '',
            envoi.ouvert_le.isoformat() if envoi.ouvert_le else '',
            envoi.clique_le.isoformat() if envoi.clique_le else '',
        ])
    # BOM UTF-8 : Excel ouvre alors correctement les accents FR (même
    # convention que ``apps.dataimport.exporters.export_csv``).
    return ('﻿' + buf.getvalue()).encode('utf-8')


def export_membres_segment_xlsx(segment):
    """NTMKT40 — export XLSX SNAPSHOT (horodaté) des membres résolus d'un
    segment marketing AU MOMENT DE L'EXPORT (jamais une vue live) — utile
    pour justifier une base d'envoi lors d'un contrôle RGPD/CNDP.

    Réutilise ``apps.compta.services.evaluer_segment`` — LA MÊME fonction que
    ``previsualiser_segment`` (XMKT6) — pour que le nombre de lignes exportées
    corresponde TOUJOURS exactement au compte affiché à l'écran."""
    import io

    from django.utils import timezone
    from openpyxl import Workbook

    from apps.compta.services import evaluer_segment
    from apps.crm.selectors import leads_export_rows

    lead_ids = evaluer_segment(segment)
    rows = leads_export_rows(segment.company, lead_ids)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Membres du segment'
    ws.append(['Segment', segment.nom])
    ws.append(['Snapshot le', timezone.now().isoformat()])
    ws.append([])
    ws.append(['Nom', 'Prénom', 'Email', 'Téléphone', 'Ville'])
    for r in rows:
        ws.append([
            r['nom'], r['prenom'] or '', r['email'] or '',
            r['telephone'] or '', r['ville'] or '',
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def importer_inscriptions_evenement(evenement, file_bytes, filename):
    """NTMKT41 — crée des ``InscriptionEvenement`` en MASSE depuis un
    CSV/XLSX (ex. liste de participants d'un salon partenaire), SANS passer
    par le formulaire public.

    Réutilise le PARSEUR générique partagé
    ``apps.dataimport.parsing.iter_rows`` (CSV+XLSX, détection encodage/
    séparateur — conçu explicitement pour tout call-site hors ``dataimport``,
    voir sa docstring) : lecture seule d'une fonction utilitaire, jamais un
    import des modèles/services ``dataimport``, jamais une nouvelle lib.

    Colonnes reconnues (insensibles à la casse) : ``nom``, ``email``,
    ``telephone``/``tel``. ``nom`` obligatoire par ligne (sinon signalée en
    invalide). Doublon = email DÉJÀ inscrit à CET événement (ignoré avec
    rapport, jamais silencieux). Renvoie
    ``{'crees', 'doublons', 'lignes_invalides', 'total'}``."""
    import uuid

    from apps.dataimport.parsing import iter_rows

    from .models import InscriptionEvenement

    headers, rows = iter_rows(file_bytes, filename)
    index_by_norm = {(h or '').strip().lower(): h for h in headers}

    def _valeur(row, *alias):
        for a in alias:
            cle = index_by_norm.get(a)
            if cle is not None and row.get(cle):
                return str(row[cle]).strip()
        return ''

    emails_existants = {
        e.lower() for e in InscriptionEvenement.objects.filter(
            evenement=evenement, email__gt='')
        .values_list('email', flat=True) if e
    }

    crees = 0
    doublons = 0
    invalides = 0
    for row in rows:
        nom = _valeur(row, 'nom', 'name')
        if not nom:
            invalides += 1
            continue
        email = _valeur(row, 'email', 'e-mail', 'mail')
        telephone = _valeur(row, 'telephone', 'tel', 'téléphone')
        if email and email.lower() in emails_existants:
            doublons += 1
            continue
        InscriptionEvenement.objects.create(
            company=evenement.company, evenement=evenement, nom=nom,
            email=email, telephone=telephone, qr_token=uuid.uuid4().hex,
        )
        if email:
            emails_existants.add(email.lower())
        crees += 1
    return {
        'crees': crees, 'doublons': doublons, 'lignes_invalides': invalides,
        'total': len(rows),
    }


def notifier_si_nps_detracteur(enquete):
    """NTMKT44 — notifie le commercial du lead d'un client DÉTRACTEUR
    (score <= 6) à une enquête NPS, lien vers la fiche lead.

    Indépendant du suivi YSERV11 (``apps.compta.services._declencher_suivi_nps``
    — gated ``CompanyProfile.referral_enabled``, une ACTIVITÉ de rappel
    distincte) : ce chemin passe TOUJOURS par ``notifications.Notification``
    (jamais un second système de notification). Jamais de doublon :
    ``repondre_enquete_nps`` ne rejoue jamais une enquête déjà répondue, donc
    ce déclencheur ne s'exécute qu'une fois par enquête."""
    if enquete.score is None or enquete.score > 6:
        return None
    from apps.crm.selectors import get_latest_lead_for_client

    lead = get_latest_lead_for_client(enquete.company, enquete.client_id)
    if lead is None or not getattr(lead, 'owner_id', None):
        return None
    from apps.notifications.models import EventType
    from apps.notifications.services import notify

    notify(
        lead.owner, EventType.DIGEST, f'Détracteur NPS : {lead.nom}',
        body=(f'Note {enquete.score}/10 — {enquete.commentaire}'
              if enquete.commentaire else f'Note {enquete.score}/10.'),
        company=enquete.company, link=f'/crm/leads/{lead.id}',
    )
    return lead.owner_id


def notifier_inscription_evenement(inscription):
    """NTMKT44 — notifie le commercial du lead qui vient de s'inscrire à un
    événement marketing (résolu/dédupliqué par
    ``apps.compta.services.inscrire_evenement``, XMKT28), lien vers la fiche
    lead. No-op silencieux si l'inscription n'a résolu aucun lead ou que le
    lead n'a pas d'owner assigné."""
    if not inscription.lead_id:
        return None
    from apps.crm.selectors import get_company_lead

    lead = get_company_lead(inscription.company, inscription.lead_id)
    if lead is None or not getattr(lead, 'owner_id', None):
        return None
    from apps.notifications.models import EventType
    from apps.notifications.services import notify

    notify(
        lead.owner, EventType.DIGEST, f'Inscription événement : {lead.nom}',
        body=f"{lead.nom} s'est inscrit(e) à « {inscription.evenement.nom} ».",
        company=inscription.company, link=f'/crm/leads/{lead.id}',
    )
    return lead.owner_id


def inscrire_evenement_et_notifier(evenement, *, nom, email='', telephone='',
                                   billet=None, reponses_questions=None):
    """NTMKT44 — enveloppe ``apps.compta.services.inscrire_evenement``
    (XMKT28) SANS le modifier : notifie en plus le commercial du lead résolu
    une fois l'inscription créée. Propage toute ``ValueError`` inchangée
    (billet hors fenêtre/quota, question obligatoire manquante)."""
    from apps.compta.services import inscrire_evenement

    inscription = inscrire_evenement(
        evenement, nom=nom, email=email, telephone=telephone, billet=billet,
        reponses_questions=reponses_questions)
    notifier_inscription_evenement(inscription)
    return inscription


def attribution_comparaison(company, devis_id):
    """NTMKT20 — comparaison des 4 modèles d'attribution pour UN devis signé
    (aide à la décision, jamais un recalcul persistant). ``None`` si le devis
    n'existe pas, n'appartient pas à la société ou n'est pas accepté.

    Lit ``ventes`` UNIQUEMENT via son selector ``get_devis_by_pk`` (jamais un
    import de ``apps.ventes.models``) puis délègue le calcul à
    ``apps.crm.selectors`` (jamais un import de ``apps.crm.models``)."""
    from apps.crm.selectors import attribution_comparaison_devis
    from apps.ventes.selectors import get_devis_by_pk

    devis = get_devis_by_pk(devis_id)
    if devis is None or devis.company_id != company.id:
        return None
    resultat = attribution_comparaison_devis(devis)
    if resultat is not None:
        resultat['modele_actuel'] = modele_attribution_pour(company)
    return resultat


# ── NTMKT33 — Purge des tokens expirés (désinscription/préférences) ────────

def purger_tokens_expires(company=None):
    """NTMKT33 — balaie les artefacts liés aux jetons publics expirés.

    Les jetons de désinscription (XMKT3, ``apps.compta.services``) et de
    préférences (NTMKT22, ci-dessus) sont des jetons SIGNÉS
    (``django.core.signing``), jamais stockés en base — il n'existe donc
    AUCUN enregistrement à supprimer/anonymiser pour eux : leur expiration
    est déjà imposée à la LECTURE par ``max_age`` (voir
    ``lire_token_preferences``, NTMKT33 ci-dessus), ce qui les rend inutilisables
    passé 90 jours sans qu'aucune donnée ne soit conservée. Cette tâche reste
    la sentinelle du beat (QX11) et le point d'extension si un futur jeton
    marketing devient un jour persisté — elle ne touche JAMAIS
    ``core.ConsentRecord`` (rétention légale distincte).
    """
    return {'jetons_purges': 0}


# ── NTMKT35 — Rappel d'approbation d'envoi en attente ───────────────────────

def rappeler_approbations_envoi_en_attente(company, *, maintenant=None,
                                           delai_heures=24):
    """NTMKT35 — notifie les approbateurs (rôles admin/responsable de la
    société) pour toute ``ApprobationEnvoiCampagne`` EN ATTENTE depuis plus de
    ``delai_heures`` — une seule relance par demande (``rappel_envoye_le``
    posé au premier envoi, jamais réinitialisé). Ne notifie jamais HORS des
    heures ouvrées de la société (``notifications.selectors`` — la tâche beat
    tourne toutes les 4h mais reste no-op la nuit/jour non ouvré)."""
    from django.utils import timezone

    from apps.notifications.selectors import est_hors_fenetre_silence

    from .models import ApprobationEnvoiCampagne

    maintenant = maintenant or timezone.now()
    if est_hors_fenetre_silence(maintenant, company):
        return []
    seuil = maintenant - timezone.timedelta(hours=delai_heures)
    demandes = ApprobationEnvoiCampagne.objects.filter(
        company=company,
        statut=ApprobationEnvoiCampagne.Statut.EN_ATTENTE,
        date_creation__lte=seuil,
        rappel_envoye_le__isnull=True,
    ).select_related('campagne')
    if not demandes.exists():
        return []
    from authentication.models import CustomUser
    approbateurs = list(CustomUser.objects.filter(
        company=company, is_active=True,
        role_legacy__in=[CustomUser.ROLE_ADMIN, CustomUser.ROLE_RESPONSABLE]))
    if not approbateurs:
        return []
    from apps.notifications.models import EventType
    from apps.notifications.services import notify_many

    notifiees = []
    for demande in demandes:
        notify_many(
            approbateurs, EventType.DIGEST,
            f'Approbation d\'envoi en attente : {demande.campagne.nom}',
            body=(f'{demande.nb_destinataires_demandes} destinataires — '
                  f'en attente depuis plus de {delai_heures}h.'),
            company=company,
        )
        demande.rappel_envoye_le = maintenant
        demande.save(update_fields=['rappel_envoye_le'])
        notifiees.append(demande.id)
    return notifiees
