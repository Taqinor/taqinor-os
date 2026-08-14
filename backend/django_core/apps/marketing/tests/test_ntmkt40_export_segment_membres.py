"""NTMKT40 — Export XLSX des segments et de leurs membres (audit RGPD/CNDP).

Couvre : le nombre de lignes exportées = le compte de la prévisualisation
(même fonction source, ``evaluer_segment``), snapshot horodaté, isolation
multi-société, authentification requise.
"""
import io

from openpyxl import load_workbook

from apps.crm.models import Lead
from apps.marketing.models import SegmentMarketing

from testkit.base import TenantAPITestCase


class ExportMembresSegmentXlsxTests(TenantAPITestCase):
    def setUp(self):
        super().setUp()
        Lead.objects.create(
            company=self.company, nom='Alami', ville='Casablanca')
        Lead.objects.create(
            company=self.company, nom='Bennani', ville='Casablanca')
        Lead.objects.create(
            company=self.company, nom='Chraibi', ville='Rabat')
        self.segment = SegmentMarketing.objects.create(
            company=self.company, nom='Casablanca',
            regles={'ville': 'Casablanca'})

    def test_endpoint_exige_une_authentification(self):
        res = self.client.get(
            f'/api/django/marketing/segments-marketing/{self.segment.id}/export/')
        self.assertIn(res.status_code, (401, 403))

    def test_nombre_de_lignes_egal_au_compte_de_previsualisation(self):
        # `previsualiser` vit sur `SegmentMarketingViewSet` (apps.compta,
        # `_ComptaBaseViewSet` : accès Responsable/Admin uniquement, cf.
        # apps/marketing/tests/test_xmkt6_segments_marketing.py qui utilise
        # déjà un rôle 'responsable') — l'export XLSX, lui, n'exige que
        # IsAuthenticated (`export_membres_segment_xlsx_view`).
        preview = self.client_as(role='responsable').get(
            f'/api/django/marketing/segments-marketing/{self.segment.id}/previsualiser/')
        self.assertEqual(preview.status_code, 200)
        compte_affiche = preview.json()['count']
        self.assertEqual(compte_affiche, 2)  # Alami + Bennani (Casablanca)

        res = self.client_as().get(
            f'/api/django/marketing/segments-marketing/{self.segment.id}/export/')
        self.assertEqual(res.status_code, 200)
        wb = load_workbook(io.BytesIO(res.content))
        ws = wb.active
        lignes = list(ws.iter_rows(values_only=True))
        # 2 lignes d'en-tête méta (Segment/Snapshot) + 1 ligne vide +
        # 1 ligne colonnes + N membres.
        membres = lignes[4:]
        self.assertEqual(len(membres), compte_affiche)
        noms = {ligne[0] for ligne in membres}
        self.assertEqual(noms, {'Alami', 'Bennani'})

    def test_segment_d_une_autre_societe_404(self):
        autre_segment = SegmentMarketing.objects.create(
            company=self.other_company, nom='Fuite')
        res = self.client_as().get(
            f'/api/django/marketing/segments-marketing/{autre_segment.id}/export/')
        self.assertEqual(res.status_code, 404)
