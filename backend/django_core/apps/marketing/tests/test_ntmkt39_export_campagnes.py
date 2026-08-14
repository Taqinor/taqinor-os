"""NTMKT39 — Export CSV/XLSX des campagnes et de leur trace d'envoi.

Couvre : l'export XLSX contient exactement les lignes filtrées, l'export CSV
d'une campagne est correctement échappé (virgules/accents), authentification
requise, isolation multi-société.
"""
import io

from django.utils import timezone
from openpyxl import load_workbook

from apps.marketing.models import Campagne, EnvoiCampagne

from testkit.base import TenantAPITestCase


class ExportCampagnesXlsxTests(TenantAPITestCase):
    def setUp(self):
        super().setUp()
        Campagne.objects.create(
            company=self.company, nom='Relance été', canal='email',
            statut='brouillon')
        Campagne.objects.create(
            company=self.company, nom='Promo agricole', canal='sms',
            statut='envoyee')
        # Jamais visible : autre société.
        Campagne.objects.create(
            company=self.other_company, nom='Fuite', canal='email',
            statut='envoyee')

    def test_endpoint_exige_une_authentification(self):
        res = self.client.get('/api/django/marketing/campagnes/export/')
        self.assertIn(res.status_code, (401, 403))

    def test_export_contient_exactement_les_lignes_filtrees(self):
        res = self.client_as().get(
            '/api/django/marketing/campagnes/export/', {'statut': 'envoyee'})
        self.assertEqual(res.status_code, 200)
        wb = load_workbook(io.BytesIO(res.content))
        ws = wb.active
        lignes = list(ws.iter_rows(values_only=True))
        # en-tête + 1 seule campagne (statut=envoyee, hors autre société).
        self.assertEqual(len(lignes), 2)
        self.assertEqual(lignes[1][0], 'Promo agricole')

    def test_export_sans_filtre_couvre_toute_la_societe_jamais_les_autres(self):
        res = self.client_as().get('/api/django/marketing/campagnes/export/')
        wb = load_workbook(io.BytesIO(res.content))
        ws = wb.active
        noms = [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]
        self.assertEqual(set(noms), {'Relance été', 'Promo agricole'})


class ExportEnvoisCampagneCsvTests(TenantAPITestCase):
    def setUp(self):
        super().setUp()
        self.campagne = Campagne.objects.create(
            company=self.company, nom='Campagne, avec virgule', canal='email')
        EnvoiCampagne.objects.create(
            company=self.company, campagne=self.campagne,
            destinataire='accentué@ex.ma', statut='ouvert',
            envoye_le=timezone.now(), ouvert_le=timezone.now())
        EnvoiCampagne.objects.create(
            company=self.company, campagne=self.campagne,
            destinataire='b@ex.ma', statut='queued')

    def test_export_csv_contient_toutes_les_lignes_correctement_echappees(self):
        res = self.client_as().get(
            f'/api/django/marketing/campagnes/{self.campagne.id}/envois/export/')
        self.assertEqual(res.status_code, 200)
        texte = res.content.decode('utf-8-sig')
        self.assertIn('accentué@ex.ma', texte)
        self.assertIn('b@ex.ma', texte)
        # 1 en-tête + 2 lignes de données.
        self.assertEqual(len(texte.strip().splitlines()), 3)

    def test_campagne_d_une_autre_societe_404(self):
        autre = Campagne.objects.create(
            company=self.other_company, nom='Autre', canal='email')
        res = self.client_as().get(
            f'/api/django/marketing/campagnes/{autre.id}/envois/export/')
        self.assertEqual(res.status_code, 404)
