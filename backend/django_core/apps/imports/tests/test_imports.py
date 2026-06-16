"""Tests T9 — import réutilisable (dry-run) + export .xlsx sur les listes.

Couvre :
  - dry-run : mapping correct + colonnes non mappées, sans persistance ;
  - import complet : fiches créées, scopées société, origin-taggées ;
  - aucun écrasement silencieux (doublon ignoré) ;
  - exports : scopés société, honorent un filtre, excluent le prix d'achat.
"""
import io
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import TestCase
from openpyxl import Workbook, load_workbook
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import AccessToken

from authentication.models import Company
from apps.crm.models import Client, Lead
from apps.imports.models import ImportBatch
from apps.stock.models import Produit

User = get_user_model()


def make_company(slug='imp-co', nom='Imp Co'):
    company, _ = Company.objects.get_or_create(slug=slug, defaults={'nom': nom})
    return company


def csv_bytes(text):
    return io.BytesIO(text.encode('utf-8'))


def xlsx_bytes(rows):
    wb = Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = 'data.xlsx'
    return buf


class _Base(TestCase):
    def setUp(self):
        from apps.roles.models import Role, ALL_PERMISSIONS
        self.company = make_company()
        self.other = make_company(slug='imp-other', nom='Imp Other')
        admin_role = Role.objects.create(
            company=self.company, nom='Administrateur',
            permissions=ALL_PERMISSIONS, est_systeme=True)
        self.admin = User.objects.create_user(
            username='imp_admin', password='x', role=admin_role,
            role_legacy='admin', company=self.company)
        self.api = self._api(self.admin)

    def _api(self, u):
        api = APIClient()
        api.credentials(HTTP_AUTHORIZATION=f'Bearer {AccessToken.for_user(u)}')
        return api


class TestImportPreview(_Base):
    def test_dryrun_mapping_and_unmapped_without_persist(self):
        f = csv_bytes(
            'Nom,Prénom,Email,Couleur préférée\n'
            'Dupont,Ali,ali@example.com,bleu\n'
            'Bennani,Sara,sara@example.com,vert\n')
        f.name = 'leads.csv'
        r = self.api.post(
            '/api/django/imports/preview/',
            {'target': 'lead', 'file': f}, format='multipart')
        self.assertEqual(r.status_code, 200, r.data)
        mapped = {m['field'] for m in r.data['mapping']}
        self.assertEqual(mapped, {'nom', 'prenom', 'email'})
        self.assertIn('Couleur préférée', r.data['unmapped_columns'])
        self.assertEqual(r.data['total_rows'], 2)
        self.assertEqual(r.data['will_create'], 2)
        # Rien n'est persisté par le dry-run.
        self.assertEqual(Lead.objects.count(), 0)
        self.assertEqual(ImportBatch.objects.count(), 0)

    def test_required_field_missing_flagged(self):
        f = csv_bytes('Email\nbob@example.com\n')
        f.name = 'leads.csv'
        r = self.api.post(
            '/api/django/imports/preview/',
            {'target': 'lead', 'file': f}, format='multipart')
        self.assertEqual(r.status_code, 200, r.data)
        self.assertEqual(r.data['will_create'], 0)
        self.assertEqual(r.data['preview'][0]['status'], 'error')

    def test_duplicate_flagged_in_preview(self):
        Client.objects.create(company=self.company, nom='X',
                              email='dup@example.com')
        f = csv_bytes('Nom,Email\nY,dup@example.com\n')
        f.name = 'clients.csv'
        r = self.api.post(
            '/api/django/imports/preview/',
            {'target': 'client', 'file': f}, format='multipart')
        self.assertEqual(r.status_code, 200, r.data)
        self.assertEqual(r.data['preview'][0]['status'], 'duplicate')
        self.assertEqual(r.data['will_skip'], 1)


class TestImportConfirm(_Base):
    def test_confirm_creates_company_scoped_origin_tagged(self):
        f = csv_bytes(
            'Nom,Prénom,Email\nDupont,Ali,ali@example.com\n'
            'Bennani,Sara,sara@example.com\n')
        f.name = 'leads.csv'
        r = self.api.post(
            '/api/django/imports/confirm/',
            {'target': 'lead', 'file': f}, format='multipart')
        self.assertEqual(r.status_code, 201, r.data)
        self.assertEqual(r.data['created'], 2)
        leads = Lead.objects.all()
        self.assertEqual(leads.count(), 2)
        batch = ImportBatch.objects.get()
        for le in leads:
            self.assertEqual(le.company_id, self.company.id)
            self.assertEqual(le.import_batch_id, batch.id)
        self.assertEqual(batch.target, 'lead')
        self.assertEqual(batch.company_id, self.company.id)

    def test_confirm_no_silent_overwrite_of_duplicate(self):
        existing = Client.objects.create(
            company=self.company, nom='Original', email='dup@example.com')
        f = csv_bytes('Nom,Email,Téléphone\nCHANGED,dup@example.com,0600\n')
        f.name = 'clients.csv'
        r = self.api.post(
            '/api/django/imports/confirm/',
            {'target': 'client', 'file': f}, format='multipart')
        self.assertEqual(r.status_code, 201, r.data)
        self.assertEqual(r.data['created'], 0)
        self.assertEqual(r.data['skipped'], 1)
        existing.refresh_from_db()
        # La fiche existante n'est PAS modifiée.
        self.assertEqual(existing.nom, 'Original')
        self.assertIsNone(existing.telephone)

    def test_confirm_xlsx_products(self):
        buf = xlsx_bytes([
            ['Nom', 'SKU', 'Prix vente HT', 'Quantité'],
            ['Panneau 550W', 'PAN-550', '1200,50', '10'],
        ])
        r = self.api.post(
            '/api/django/imports/confirm/',
            {'target': 'produit', 'file': buf}, format='multipart')
        self.assertEqual(r.status_code, 201, r.data)
        p = Produit.objects.get()
        self.assertEqual(p.sku, 'PAN-550')
        self.assertEqual(p.prix_vente, Decimal('1200.50'))
        self.assertEqual(p.quantite_stock, 10)
        self.assertIsNotNone(p.import_batch_id)

    def test_import_never_trusts_company_from_body(self):
        f = csv_bytes('Nom,Email\nA,a@example.com\n')
        f.name = 'leads.csv'
        r = self.api.post(
            '/api/django/imports/confirm/',
            {'target': 'lead', 'file': f, 'company': self.other.id},
            format='multipart')
        self.assertEqual(r.status_code, 201, r.data)
        self.assertEqual(Lead.objects.get().company_id, self.company.id)


class TestExports(_Base):
    def _read_xlsx(self, response):
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        return [list(row) for row in ws.iter_rows(values_only=True)]

    def test_export_company_scoped(self):
        Client.objects.create(company=self.company, nom='Mine',
                              email='mine@example.com')
        Client.objects.create(company=self.other, nom='Theirs',
                              email='theirs@example.com')
        r = self.api.get('/api/django/crm/clients/export/')
        self.assertEqual(r.status_code, 200)
        rows = self._read_xlsx(r)
        names = [row[1] for row in rows[1:]]
        self.assertIn('Mine', names)
        self.assertNotIn('Theirs', names)

    def test_export_honors_filter(self):
        Produit.objects.create(company=self.company, nom='Inverter Huawei',
                               sku='INV-1', prix_vente=Decimal('100'))
        Produit.objects.create(company=self.company, nom='Cable',
                               sku='CAB-1', prix_vente=Decimal('5'))
        r = self.api.get('/api/django/stock/produits/export/?search=Huawei')
        self.assertEqual(r.status_code, 200)
        rows = self._read_xlsx(r)
        names = [row[1] for row in rows[1:]]
        self.assertEqual(names, ['Inverter Huawei'])

    def test_export_excludes_buy_price(self):
        Produit.objects.create(
            company=self.company, nom='Secret margin', sku='S-1',
            prix_vente=Decimal('999'), prix_achat=Decimal('500'))
        r = self.api.get('/api/django/stock/produits/export/')
        self.assertEqual(r.status_code, 200)
        text = r.content.decode('latin-1', errors='ignore')
        header_lower = self._read_xlsx(r)[0]
        joined = ' '.join(str(h or '') for h in header_lower).lower()
        self.assertNotIn('achat', joined)
        self.assertNotIn('marge', joined)
        # La valeur du prix d'achat ne doit apparaître dans aucune cellule.
        for row in self._read_xlsx(r)[1:]:
            self.assertNotIn(500.0, row)
            self.assertNotIn('500', [str(c) for c in row])
        self.assertNotIn('prix_achat', text)

    def test_devis_facture_installations_sav_export_endpoints(self):
        # Endpoints répondent 200 et sont scopés (listes vides ici).
        for url in (
            '/api/django/ventes/devis/export/',
            '/api/django/ventes/factures/export/',
            '/api/django/installations/chantiers/export/',
            '/api/django/sav/equipements/export/',
            '/api/django/sav/tickets/export/',
            '/api/django/crm/leads/export/',
        ):
            r = self.api.get(url)
            self.assertEqual(r.status_code, 200, f'{url} -> {r.status_code}')
            self.assertIn('spreadsheet', r['Content-Type'])
