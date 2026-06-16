"""Service d'import réutilisable : parse → aperçu (dry-run) → confirmation.

Garanties :
  - RIEN n'est persisté tant que l'utilisateur n'a pas approuvé l'aperçu ;
  - RIEN n'est écrasé silencieusement : l'import est CRÉATION SEULE. Une ligne
    en doublon (clé de dédup déjà présente dans la société) est IGNORÉE et
    visible dans l'aperçu (« doublon ») ;
  - la société est TOUJOURS imposée côté serveur (jamais lue du fichier).
"""
import csv
import io

from django.db import transaction

from .models import ImportBatch
from .specs import get_spec

PREVIEW_ROWS = 10
MAX_ROWS = 5000


class ImportError_(Exception):
    """Erreur d'import présentable à l'utilisateur (message FR)."""


def _read_table(filename, content):
    """Lit un CSV ou XLSX en (headers, rows) — rows = liste de listes de cellules.

    `content` : octets bruts du fichier téléversé.
    """
    name = (filename or '').lower()
    if name.endswith('.xlsx') or name.endswith('.xlsm'):
        return _read_xlsx(content)
    if name.endswith('.csv') or name.endswith('.txt'):
        return _read_csv(content)
    # Heuristique : tente XLSX (signature ZIP) sinon CSV.
    if content[:2] == b'PK':
        return _read_xlsx(content)
    return _read_csv(content)


def _read_csv(content):
    text = None
    for enc in ('utf-8-sig', 'utf-8', 'latin-1'):
        try:
            text = content.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ImportError_('Fichier illisible (encodage non supporté).')
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=',;\t')
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ';' if sample.count(';') > sample.count(',') else ','
    reader = csv.reader(io.StringIO(text), dialect)
    rows = [list(r) for r in reader]
    if not rows:
        raise ImportError_('Fichier vide.')
    headers = [str(h).strip() for h in rows[0]]
    return headers, rows[1:]


def _read_xlsx(content):
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise ImportError_('Fichier Excel illisible.')
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ImportError_('Fichier vide.')
    headers = ['' if h is None else str(h).strip() for h in header_row]
    rows = []
    for r in rows_iter:
        rows.append(['' if c is None else c for c in r])
    return headers, rows


def _infer_mapping(spec, headers):
    """Retourne (mapping, unmapped).

    mapping : liste de {column, field, label} pour les en-têtes reconnus.
    unmapped : liste des en-têtes non reconnus (libellés).
    Une même colonne ne peut alimenter qu'un champ ; le premier en-tête
    correspondant gagne pour un champ donné.
    """
    mapping = []
    unmapped = []
    used_fields = set()
    for idx, header in enumerate(headers):
        if not str(header).strip():
            continue
        key = spec.match_header(header)
        if key and key not in used_fields:
            used_fields.add(key)
            field = spec.field(key)
            mapping.append({
                'column': header,
                'index': idx,
                'field': key,
                'label': field.label,
            })
        else:
            unmapped.append(header)
    return mapping, unmapped


def _row_to_values(spec, mapping, row):
    """Convertit une ligne brute en (values, problems) selon le mapping.

    values : {field_key: coerced_value}. problems : liste de messages FR.
    """
    values = {}
    problems = []
    for m in mapping:
        idx = m['index']
        raw = row[idx] if idx < len(row) else None
        field = spec.field(m['field'])
        try:
            values[m['field']] = field.coerce(raw)
        except ValueError as exc:
            problems.append(f"{field.label} : {exc}")
    # Champs requis manquants.
    for key in spec.required_keys():
        if values.get(key) in (None, ''):
            field = spec.field(key)
            problems.append(f"{field.label} requis")
    return values, problems


def _is_blank_row(row):
    return all((c is None or str(c).strip() == '') for c in row)


def _dedup_lookup(spec, company, values):
    """Retourne True si une ligne en doublon existe déjà (création seule).

    On teste les clés de dédup dans l'ordre : la première clé renseignée sert
    de critère. Strictement scopé société.
    """
    model = spec.model
    for key in spec.dedup_keys:
        val = values.get(key)
        if val in (None, ''):
            continue
        if model.objects.filter(company=company, **{key: val}).exists():
            return True, key
        return False, key  # clé renseignée mais pas de doublon
    return False, None


def build_preview(target, filename, content, company):
    """Aperçu dry-run : mapping inféré + 10 lignes + colonnes/lignes en
    problème. NE PERSISTE RIEN."""
    spec = get_spec(target)
    if spec is None:
        raise ImportError_('Type d\'import inconnu.')
    headers, rows = _read_table(filename, content)
    mapping, unmapped = _infer_mapping(spec, headers)
    if not mapping:
        raise ImportError_(
            'Aucune colonne reconnue. Vérifiez la ligne d\'en-tête.')

    data_rows = [r for r in rows if not _is_blank_row(r)]
    total = len(data_rows)
    if total > MAX_ROWS:
        raise ImportError_(
            f'Trop de lignes ({total}). Maximum {MAX_ROWS} par import.')

    preview = []
    will_create = 0
    will_skip = 0
    for row in data_rows:
        values, problems = _row_to_values(spec, mapping, row)
        status = 'create'
        if problems:
            status = 'error'
        else:
            is_dup, _ = _dedup_lookup(spec, company, values)
            if is_dup:
                status = 'duplicate'
                problems = ['Doublon — déjà présent (ignoré, jamais écrasé)']
        if status == 'create':
            will_create += 1
        else:
            will_skip += 1
        if len(preview) < PREVIEW_ROWS:
            preview.append({
                'values': {k: (str(v) if v is not None else '')
                           for k, v in values.items()},
                'status': status,
                'problems': problems,
            })

    return {
        'target': target,
        'label': spec.label,
        'filename': filename,
        'mapping': [{'column': m['column'], 'field': m['field'],
                     'label': m['label']} for m in mapping],
        'unmapped_columns': unmapped,
        'total_rows': total,
        'will_create': will_create,
        'will_skip': will_skip,
        'preview': preview,
        'preview_rows': PREVIEW_ROWS,
    }


@transaction.atomic
def run_import(target, filename, content, company, user):
    """Exécute l'import complet (création seule, doublons ignorés). Crée un
    ImportBatch et marque chaque fiche créée avec son `import_batch`."""
    spec = get_spec(target)
    if spec is None:
        raise ImportError_('Type d\'import inconnu.')
    headers, rows = _read_table(filename, content)
    mapping, _ = _infer_mapping(spec, headers)
    if not mapping:
        raise ImportError_('Aucune colonne reconnue.')

    data_rows = [r for r in rows if not _is_blank_row(r)]
    if len(data_rows) > MAX_ROWS:
        raise ImportError_(
            f'Trop de lignes ({len(data_rows)}). Maximum {MAX_ROWS}.')

    batch = ImportBatch.objects.create(
        company=company, target=target,
        filename=(filename or '')[:255], created_by=user)

    model = spec.model
    created = 0
    skipped = 0
    for row in data_rows:
        values, problems = _row_to_values(spec, mapping, row)
        if problems:
            skipped += 1
            continue
        is_dup, _ = _dedup_lookup(spec, company, values)
        if is_dup:
            skipped += 1
            continue
        obj = model(company=company, import_batch=batch, **values)
        obj.save()
        created += 1

    batch.created_count = created
    batch.skipped_count = skipped
    batch.save(update_fields=['created_count', 'skipped_count'])
    return {
        'batch_id': batch.id,
        'target': target,
        'created': created,
        'skipped': skipped,
        'total_rows': len(data_rows),
    }
