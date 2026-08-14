"""NTSCM27 — Rapport S&OP exportable (compte-rendu de réunion), archivé
automatiquement en GED à la clôture d'un cycle.

Critère d'acceptation : clôturer un cycle produit un fichier téléchargeable
avec les 3 feuilles correctement remplies et une copie visible dans
``/ged``.

``apps.ged.models.Document`` consulté directement UNIQUEMENT pour vérifier
la copie GED dans le test (frontière cross-app, CLAUDE.md — la PRODUCTION
passe par ``apps.ged.services.deposit_document``, jamais un import de modèle
côté ``apps.scm``)."""
import io

from django.test import TestCase
from openpyxl import load_workbook

from apps.ged.models import Document
from apps.scm import services as scm_services
from apps.scm.models import CyclePlanificationSOP

from .helpers import auth, make_company, make_user


class ComptesRendusSopTests(TestCase):
    def setUp(self):
        self.company = make_company('scm-compte-rendu', 'Supply Compte-rendu')
        self.admin = make_user(self.company, 'scm-compte-rendu-admin', 'admin')
        self.cycle = CyclePlanificationSOP.objects.create(
            company=self.company, periode='2026-09')

    def _cloturer(self):
        cycle = self.cycle
        for _ in range(6):  # brouillon -> ... -> clos (6 transitions)
            cycle = scm_services.avancer_statut_cycle(cycle, self.admin)
        return cycle

    def test_classeur_a_3_feuilles(self):
        wb = scm_services._construire_classeur_sop(self.cycle)
        self.assertEqual(
            wb.sheetnames,
            ['Demande consensuelle', 'Offre et écarts', 'Impact financier'])

    def test_cloture_archive_en_ged(self):
        self.assertEqual(Document.objects.filter(company=self.company).count(), 0)
        cycle_clos = self._cloturer()
        self.assertEqual(cycle_clos.statut, CyclePlanificationSOP.Statut.CLOS)

        documents = Document.objects.filter(company=self.company)
        self.assertEqual(documents.count(), 1)
        self.assertIn('S&OP', documents.first().nom)

    def test_cloture_idempotente_pas_de_doublon_ged(self):
        self._cloturer()
        scm_services.generer_compte_rendu_sop(self.cycle, user=self.admin)
        self.assertEqual(Document.objects.filter(company=self.company).count(), 1)

    def test_endpoint_telechargement_xlsx(self):
        resp = auth(self.admin).get(
            f'/api/django/scm/cycles-sop/{self.cycle.id}/compte-rendu/')
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(io.BytesIO(resp.content))
        self.assertEqual(len(wb.sheetnames), 3)
