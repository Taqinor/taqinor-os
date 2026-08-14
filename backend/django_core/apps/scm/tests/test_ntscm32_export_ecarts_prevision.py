"""NTSCM32 — Rapport « Écarts de prévision » exportable (.xlsx, réutilise
NTSCM24).

Critère d'acceptation : le fichier exporté contient une ligne par produit
avec les colonnes attendues et un total en pied de tableau."""
import io

from django.test import TestCase
from django.utils import timezone

from apps.scm.models import PrevisionDemande
from apps.stock.models import MouvementStock, Produit

from .helpers import auth, make_company, make_user


class ExportEcartsPrevisionTests(TestCase):
    def setUp(self):
        self.company = make_company('scm-export-ecarts', 'Supply Export Écarts')
        self.admin = make_user(self.company, 'scm-export-ecarts-admin', 'admin')
        self.produit = Produit.objects.create(
            company=self.company, nom='Onduleur 5kW', prix_vente=8000,
            quantite_stock=50)

    def _periode(self, offset_mois):
        today = timezone.localdate()
        idx = today.year * 12 + (today.month - 1) - offset_mois
        y, m0 = divmod(idx, 12)
        return y, m0 + 1, f'{y:04d}-{m0 + 1:02d}'

    def _seed(self, offset, prevu, reel):
        y, m, periode = self._periode(offset)
        PrevisionDemande.objects.create(
            company=self.company, produit=self.produit, segment='',
            periode=periode, quantite_prevue=prevu)
        mvt = MouvementStock.objects.create(
            company=self.company, produit=self.produit,
            type_mouvement=MouvementStock.TypeMouvement.SORTIE,
            quantite=reel, quantite_avant=1000, quantite_apres=1000 - reel)
        mvt.date = timezone.make_aware(timezone.datetime(y, m, 10))
        mvt.save(update_fields=['date'])

    def test_export_contient_une_ligne_par_produit_et_un_total(self):
        self._seed(1, 40, 60)

        resp = auth(self.admin).get('/api/django/scm/precision-previsions/export/')
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        self.assertIn('spreadsheetml', resp['Content-Type'])

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(resp.content), read_only=True)
        ws = wb.active
        lignes = list(ws.iter_rows(values_only=True))

        self.assertEqual(
            lignes[0], ('Produit', 'Prévision totale', 'Réel total',
                        'Écart absolu', 'Écart %'))
        # Une ligne pour le produit + une ligne TOTAL en pied de tableau.
        self.assertEqual(len(lignes), 3)
        ligne_produit = lignes[1]
        self.assertEqual(ligne_produit[0], 'Onduleur 5kW')
        self.assertEqual(str(ligne_produit[1]), '40')
        self.assertEqual(str(ligne_produit[2]), '60')
        self.assertEqual(str(ligne_produit[3]), '20')

        ligne_total = lignes[2]
        self.assertEqual(ligne_total[0], 'TOTAL')
        self.assertEqual(str(ligne_total[1]), '40')
        self.assertEqual(str(ligne_total[2]), '60')

    def test_export_refuse_role_non_responsable(self):
        normal = make_user(self.company, 'scm-export-ecarts-normal', 'normal')
        resp = auth(normal).get('/api/django/scm/precision-previsions/export/')
        self.assertEqual(resp.status_code, 403)
