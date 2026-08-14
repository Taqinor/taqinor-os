"""NTSCM41 — Export CSV/XLSX des suggestions d'achat groupées et des
suggestions de transfert.

Critère d'acceptation : les deux exports produisent un fichier avec une ligne
par suggestion active au moment de l'export."""
import io

from django.test import TestCase

from apps.scm.models import PolitiqueStock
from apps.stock.models import EmplacementStock, Fournisseur, MouvementStock, PrixFournisseur, Produit, StockEmplacement

from .helpers import auth, make_company, make_user


class ExportSuggestionsAchatGroupeTests(TestCase):
    def setUp(self):
        self.company = make_company('scm-export-achat', 'Supply Export Achat')
        self.admin = make_user(self.company, 'scm-export-achat-admin', 'admin')
        self.fournisseur = Fournisseur.objects.create(
            company=self.company, nom='Fournisseur Export')
        self.produit = Produit.objects.create(
            company=self.company, nom='Batterie 5kWh', prix_vente=15000,
            quantite_stock=0, fournisseur=self.fournisseur)
        PrixFournisseur.objects.create(
            company=self.company, produit=self.produit,
            fournisseur=self.fournisseur, prix_achat=9000)
        # Politique de stock avec point de commande > stock actuel (0) pour
        # apparaître dans le tableau de bord réappro -> suggestion d'achat.
        PolitiqueStock.objects.create(
            company=self.company, produit=self.produit, classe_abc='A',
            service_level_pct=95, point_commande=10,
            stock_securite_calcule=5)
        for _i in range(3):
            MouvementStock.objects.create(
                company=self.company, produit=self.produit,
                type_mouvement=MouvementStock.TypeMouvement.SORTIE,
                quantite=10, quantite_avant=100, quantite_apres=90)

    def test_export_xlsx_une_ligne_par_suggestion(self):
        resp = auth(self.admin).get(
            '/api/django/scm/suggestions-achat-groupe/export/')
        self.assertEqual(resp.status_code, 200, resp.content[:300])
        self.assertIn('spreadsheetml', resp['Content-Type'])

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(resp.content), read_only=True)
        ws = wb.active
        lignes = list(ws.iter_rows(values_only=True))
        self.assertEqual(
            lignes[0],
            ('Fournisseur', 'Produit', 'Besoin net', 'Décision', 'Quantité',
             'Prix unitaire', 'Coût total', 'Justification'))
        self.assertGreaterEqual(len(lignes), 2)
        self.assertEqual(lignes[1][0], 'Fournisseur Export')
        self.assertEqual(lignes[1][1], 'Batterie 5kWh')


class ExportSuggestionsTransfertTests(TestCase):
    def setUp(self):
        self.company = make_company('scm-export-transfert', 'Supply Export Transfert')
        self.admin = make_user(self.company, 'scm-export-transfert-admin', 'admin')
        self.produit = Produit.objects.create(
            company=self.company, nom='Onduleur 3kW', prix_vente=6000)
        self.source = EmplacementStock.objects.create(
            company=self.company, nom='Dépôt Casablanca', is_principal=False)
        self.destination = EmplacementStock.objects.create(
            company=self.company, nom='Dépôt Rabat', is_principal=False)
        StockEmplacement.objects.create(
            company=self.company, produit=self.produit,
            emplacement=self.source, quantite=100, seuil_max=20)
        StockEmplacement.objects.create(
            company=self.company, produit=self.produit,
            emplacement=self.destination, quantite=0, seuil_min=10)

        from apps.scm.models import PrevisionDemande
        PrevisionDemande.objects.create(
            company=self.company, produit=self.produit,
            segment='Dépôt Casablanca', periode='2099-01', quantite_prevue=0)
        PrevisionDemande.objects.create(
            company=self.company, produit=self.produit,
            segment='Dépôt Rabat', periode='2099-01', quantite_prevue=50)

    def test_export_xlsx_une_ligne_par_suggestion(self):
        resp = auth(self.admin).get('/api/django/scm/suggestions-transfert/export/')
        self.assertEqual(resp.status_code, 200, resp.content[:300])

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(resp.content), read_only=True)
        ws = wb.active
        lignes = list(ws.iter_rows(values_only=True))
        self.assertEqual(
            lignes[0],
            ('Produit', 'Dépôt source', 'Dépôt destination',
             'Quantité suggérée', 'Justification'))
        self.assertGreaterEqual(len(lignes), 2)
