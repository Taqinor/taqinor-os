"""Services (écriture/orchestration) de planification supply chain (Groupe
NTSCM). Toute lecture cross-app (``apps.stock``, ``apps.ventes``…) passe par
le ``selectors.py``/``services.py`` de l'app cible, jamais un import de
modèle — voir chaque fonction pour le détail de sa frontière.
"""
import math
from calendar import monthrange
from datetime import date
from decimal import Decimal

from django.db.models import Q
from django.utils import timezone

from core.demand_forecast import forecast_demand
from core.safety_stock import compute_safety_stock

from .models import (
    ClassificationABC, CyclePlanificationSOP, EvenementDemande, LigneDemandeSOP,
    LigneOffreSOP, PolitiqueStock, PrevisionDemande,
)

# Convention du module : les historiques exposés par ``_historique_sorties_
# mensuelles`` sont MENSUELS ; ``core.safety_stock.compute_safety_stock``
# attend une consommation JOURNALIÈRE (mêmes unités que son ``lead_time_days``).
# Conversion mensuel -> journalier au mois normalisé de 30 jours (même
# convention que ``core.stock_reorder``, qui travaille déjà en jours).
JOURS_PAR_MOIS = 30.0

# NTSCM6 — niveau de service PAR DÉFAUT selon la classe ABC (NTSCM4), à la
# création d'une PolitiqueStock UNIQUEMENT (un recalcul n'écrase jamais un
# niveau déjà personnalisé par l'acheteur).
SERVICE_LEVEL_PAR_CLASSE = {'A': Decimal('95'), 'B': Decimal('90'), 'C': Decimal('85')}

# Délai fournisseur (jours) utilisé quand aucun historique de livraison
# n'est exploitable (produit sans fournisseur, ou fournisseur jamais livré).
DEFAULT_LEAD_TIME_DAYS = 15.0


def _historique_sorties_mensuelles(company, produit_id, fenetre_mois):
    """Sorties ``stock.MouvementStock`` mensuelles d'un produit, sur
    ``fenetre_mois`` mois glissants jusqu'à aujourd'hui.

    ``apps.stock.selectors.mouvements_agreges`` n'agrège que sur UNE
    dimension à la fois (produit OU mois, jamais les deux ensemble — voir
    ``apps/stock/selectors.py``) : il ne peut donc pas fournir un historique
    PAR PRODUIT ET PAR MOIS. Faute d'un sélecteur adapté côté ``apps.stock``
    (que cette lane n'a pas le droit d'écrire — frontière cross-app,
    CLAUDE.md), le modèle est résolu DYNAMIQUEMENT via
    ``django.apps.apps.get_model`` pour une agrégation en LECTURE SEULE —
    exactement le même patron que ``installations.selectors`` (FG294/FG295)
    pour ses agrégats cross-app : jamais un ``from apps.stock.models import
    MouvementStock`` statique, jamais une écriture dans ``apps.stock``.
    Renvoie ``[(periode:'YYYY-MM', quantite:float), ...]`` trié."""
    from django.apps import apps as django_apps
    from django.db.models import Sum
    from django.db.models.functions import TruncMonth

    MouvementStock = django_apps.get_model('stock', 'MouvementStock')

    today = timezone.localdate()
    idx = today.year * 12 + (today.month - 1) - max(0, int(fenetre_mois))
    y0, m0 = divmod(idx, 12)
    debut = date(y0, m0 + 1, 1)

    qs = (
        MouvementStock.objects
        .filter(
            company=company, produit_id=produit_id,
            type_mouvement=MouvementStock.TypeMouvement.SORTIE,
            date__date__gte=debut)
        .annotate(mois=TruncMonth('date'))
        .values('mois')
        .annotate(total=Sum('quantite'))
        .order_by('mois')
    )
    return [
        (f'{row["mois"].year:04d}-{row["mois"].month:02d}', float(row['total'] or 0))
        for row in qs
    ]


def _evenements_du_mois(company, produit, periode):
    """NTSCM3 — événements chevauchant le mois ``periode`` ('YYYY-MM') pour
    ce produit : match explicite sur le produit, sur sa catégorie (événement
    à ``produit`` vide), ou événement GLOBAL (``produit`` et ``categorie``
    vides tous les deux)."""
    y, m = int(periode[:4]), int(periode[5:7])
    debut_mois = date(y, m, 1)
    fin_mois = date(y, m, monthrange(y, m)[1])

    qs = EvenementDemande.objects.filter(
        company=company, date_debut__lte=fin_mois, date_fin__gte=debut_mois,
    ).filter(
        Q(produit=produit)
        | Q(produit__isnull=True, categorie_id=produit.categorie_id)
        | Q(produit__isnull=True, categorie__isnull=True)
    )
    return qs


def _appliquer_evenements(company, produit, periode, quantite_base):
    """NTSCM3 — applique l'impact cumulé (somme des ``impact_pct``, jamais en
    dessous de 0) des événements chevauchant ``periode`` à la quantité de
    base issue de la prévision (NTSCM2)."""
    facteur = Decimal('1')
    for ev in _evenements_du_mois(company, produit, periode):
        facteur += (ev.impact_pct or Decimal('0')) / Decimal('100')
    if facteur < 0:
        facteur = Decimal('0')
    return (Decimal(str(quantite_base)) * facteur).quantize(Decimal('0.01'))


def generer_previsions(
    produit, horizon_mois, company, *, segment='', fenetre_mois=24, user=None,
):
    """NTSCM2/3 — génère/rafraîchit les ``PrevisionDemande`` d'un produit sur
    ``horizon_mois`` mois, à partir de son historique de sorties (fenêtre
    ``fenetre_mois`` mois, réutilise :func:`core.demand_forecast.forecast_demand`),
    puis applique l'impact cumulé des ``EvenementDemande`` (NTSCM3)
    chevauchant chaque mois cible.

    ``produit`` est une instance ``stock.Produit`` déjà résolue par
    l'APPELANT (via ``apps.stock.selectors`` — jamais un import de modèle
    ici). Idempotent par ``(company, produit, segment, periode)``
    (``update_or_create``). Renvoie la liste des ``PrevisionDemande`` créées
    ou mises à jour."""
    historique = _historique_sorties_mensuelles(company, produit.id, fenetre_mois)
    resultat = forecast_demand(historique, horizon_mois=horizon_mois)

    methode = (
        PrevisionDemande.Methode.MOYENNE_MOBILE if resultat.used_fallback
        else PrevisionDemande.Methode.SAISONNIER
    )

    previsions = []
    for periode, quantite in resultat.previsions:
        quantite_ajustee = _appliquer_evenements(company, produit, periode, quantite)
        obj, _ = PrevisionDemande.objects.update_or_create(
            company=company, produit=produit, segment=segment or '', periode=periode,
            defaults={
                'quantite_prevue': quantite_ajustee,
                'methode': methode,
                'genere_le': timezone.now(),
                'genere_par': user,
            },
        )
        previsions.append(obj)
    return previsions


def appliquer_politique_stock(
    produit, service_level_pct, company, *, lead_time_days=0, fenetre_mois=12,
):
    """NTSCM5 — calcule le stock de sécurité au niveau de service pour un
    produit, à partir de son historique de sorties mensuelles (réutilise
    ``_historique_sorties_mensuelles``, comme NTSCM2) et
    :func:`core.safety_stock.compute_safety_stock`.

    NE PERSISTE RIEN : ``PolitiqueStock`` n'existe pas encore à ce stade de
    la lane (NTSCM6, tâche SUIVANTE, définit le modèle et appelle CETTE
    fonction pour écrire ``PolitiqueStock.stock_securite_calcule``) —
    fonction pure de calcul, réutilisable.

    Renvoie un dict ``{stock_securite, avg_daily_consumption, std_dev_daily,
    nb_mois_historique}``."""
    historique = _historique_sorties_mensuelles(company, produit.id, fenetre_mois)
    valeurs_mensuelles = [q for _, q in historique]
    n = len(valeurs_mensuelles)
    moyenne_mensuelle = (sum(valeurs_mensuelles) / n) if n else 0.0
    if n > 1:
        variance = sum((v - moyenne_mensuelle) ** 2 for v in valeurs_mensuelles) / n
        ecart_type_mensuel = math.sqrt(variance)
    else:
        ecart_type_mensuel = 0.0

    avg_daily = moyenne_mensuelle / JOURS_PAR_MOIS
    std_daily = ecart_type_mensuel / JOURS_PAR_MOIS

    stock_securite = compute_safety_stock(
        avg_daily, std_daily, lead_time_days, service_level_pct)

    return {
        'stock_securite': round(stock_securite, 2),
        'avg_daily_consumption': round(avg_daily, 4),
        'std_dev_daily': round(std_daily, 4),
        'nb_mois_historique': n,
    }


def lead_time_moyen_fournisseur(company, produit):
    """Délai fournisseur moyen (jours) — réutilise le scorecard FG59 déjà
    bâti (``apps.stock.services.supplier_performance``, LECTURE SEULE,
    import fonction-local) plutôt que de dupliquer un calcul de délai côté
    ``apps.scm``. Repli :data:`DEFAULT_LEAD_TIME_DAYS` si le produit n'a pas
    de fournisseur ou qu'aucune livraison n'a encore été confirmée."""
    if not produit.fournisseur_id:
        return DEFAULT_LEAD_TIME_DAYS
    from apps.stock.services import supplier_performance

    perf = supplier_performance(company, produit.fournisseur)
    avg = perf.get('avg_lead_time_days')
    return float(avg) if avg is not None else DEFAULT_LEAD_TIME_DAYS


def recalculer_politiques_stock(company):
    """NTSCM6 — recalcule (tâche périodique) les ``PolitiqueStock`` de tous
    les produits actifs d'une société :

      * ``classe_abc`` — snapshot depuis ``ClassificationABC`` (NTSCM4,
        défaut 'C' si le produit n'a encore jamais été classé) ;
      * ``service_level_pct`` — initialisé selon la classe
        (:data:`SERVICE_LEVEL_PAR_CLASSE`) À LA CRÉATION SEULEMENT, jamais
        écrasé si déjà personnalisé par l'acheteur ;
      * ``stock_securite_calcule`` — NTSCM5 (:func:`appliquer_politique_stock`) ;
      * ``point_commande`` (ROP) = ``conso_moy_journalière × délai
        fournisseur moyen + stock_securite`` — où ``stock_securite`` est
        ``stock_securite_manuel`` si renseigné (override acheteur), sinon
        ``stock_securite_calcule``.

    Idempotent (``get_or_create``/``save`` par produit). Renvoie la liste des
    ``PolitiqueStock`` recalculées."""
    from django.apps import apps as django_apps

    Produit = django_apps.get_model('stock', 'Produit')
    produits = Produit.objects.filter(company=company, is_archived=False)
    classes = dict(
        ClassificationABC.objects.filter(company=company)
        .values_list('produit_id', 'classe'))

    resultats = []
    for produit in produits:
        classe = classes.get(produit.id, 'C')
        niveau_defaut = SERVICE_LEVEL_PAR_CLASSE.get(classe, Decimal('95'))

        politique, created = PolitiqueStock.objects.get_or_create(
            company=company, produit=produit,
            defaults={'service_level_pct': niveau_defaut, 'classe_abc': classe},
        )
        service_level = niveau_defaut if created else politique.service_level_pct

        lead_time_days = lead_time_moyen_fournisseur(company, produit)
        calc = appliquer_politique_stock(
            produit, service_level, company, lead_time_days=lead_time_days)

        stock_securite_effectif = (
            politique.stock_securite_manuel
            if politique.stock_securite_manuel is not None
            else Decimal(str(calc['stock_securite']))
        )
        point_commande = (
            Decimal(str(calc['avg_daily_consumption'])) * Decimal(str(lead_time_days))
            + stock_securite_effectif
        )

        politique.classe_abc = classe
        politique.service_level_pct = service_level
        politique.stock_securite_calcule = Decimal(str(calc['stock_securite']))
        politique.point_commande = point_commande.quantize(Decimal('0.01'))
        politique.revise_le = timezone.now()
        politique.save(update_fields=[
            'classe_abc', 'service_level_pct', 'stock_securite_calcule',
            'point_commande', 'revise_le', 'updated_at',
        ])
        resultats.append(politique)
    return resultats


def avancer_statut_cycle(cycle, user, *, statut_cible=None):
    """NTSCM12 — avance un ``CyclePlanificationSOP`` à l'étape SUIVANTE de sa
    machine à états (jamais de saut).

    ``statut_cible`` optionnel : si fourni, DOIT être exactement l'étape
    suivante (``cycle.prochain_statut()``), sinon ``ValueError`` (mappé 400
    côté vue) — c'est ce qui refuse explicitement un saut d'étape (ex.
    brouillon -> approuve directement). Sans argument, avance simplement à
    l'étape suivante. Journalise la transition via
    ``apps.records.services.log_field_change`` (primitive plateforme
    réutilisée — jamais un nouveau modèle ``*Activity``, foundation app
    exemptée de la frontière cross-app)."""
    from apps.records.services import log_field_change

    prochain = cycle.prochain_statut()
    if prochain is None:
        raise ValueError(
            f'Le cycle est déjà à son étape finale ({cycle.get_statut_display()}).')
    if statut_cible is not None and statut_cible != prochain:
        raise ValueError(
            f'Transition invalide : depuis « {cycle.get_statut_display()} », '
            f'seule l\'étape suivante est autorisée — jamais un saut d\'étape.')

    ancien = cycle.statut
    cycle.statut = prochain
    cycle.save(update_fields=['statut'])
    log_field_change(
        cycle, 'statut', ancien, prochain, user=user,
        field_label='Statut du cycle S&OP', company=cycle.company)

    # NTSCM13 — le gel de la demande consensuelle se déclenche EXACTEMENT au
    # passage brouillon -> revue_demande (jamais rejoué à une étape
    # ultérieure : c'est ce qui rend les lignes gelées immuables une fois le
    # cycle en revue).
    if prochain == CyclePlanificationSOP.Statut.REVUE_DEMANDE:
        geler_previsions_cycle(cycle)

    # NTSCM27 — le compte-rendu S&OP (3 feuilles xlsx) est généré et archivé
    # en GED EXACTEMENT à la clôture (jamais rejoué à une étape antérieure).
    # Best-effort : un accroc GED/stockage objet ne doit JAMAIS bloquer la
    # transition d'état elle-même (même contrat que `_notify_rapport_envoye`,
    # apps/monitoring/report.py) — le compte-rendu reste régénérable à la
    # demande via l'action `compte-rendu` du viewset.
    if prochain == CyclePlanificationSOP.Statut.CLOS:
        try:
            generer_compte_rendu_sop(cycle, user=user)
        except Exception:  # noqa: BLE001 — best-effort, jamais bloquant
            import logging
            logging.getLogger(__name__).warning(
                'avancer_statut_cycle: génération du compte-rendu S&OP '
                'échouée (cycle %s)', cycle.id, exc_info=True)

    return cycle


def geler_previsions_cycle(cycle):
    """NTSCM13 — copie l'état COURANT de ``PrevisionDemande`` du mois cible
    (``cycle.periode``) dans les ``LigneDemandeSOP`` du cycle (agrégées par
    produit, tous segments confondus — le cycle S&OP raisonne au niveau
    produit).

    Appelé UNIQUEMENT au passage brouillon -> revue_demande
    (``avancer_statut_cycle``) : le snapshot devient alors IMMUABLE — modifier
    ``PrevisionDemande`` par la suite n'affecte plus les lignes déjà gelées,
    puisque cette fonction n'est plus rappelée aux étapes suivantes.
    Idempotent (``update_or_create`` par produit) pour un appel manuel de
    rattrapage. Renvoie la liste des ``LigneDemandeSOP``."""
    previsions = PrevisionDemande.objects.filter(
        company=cycle.company, periode=cycle.periode)

    par_produit = {}
    for prevision in previsions:
        par_produit[prevision.produit_id] = (
            par_produit.get(prevision.produit_id, Decimal('0'))
            + prevision.quantite_prevue
        )

    lignes = []
    for produit_id, quantite in par_produit.items():
        ligne, _ = LigneDemandeSOP.objects.update_or_create(
            cycle=cycle, produit_id=produit_id,
            defaults={
                'company': cycle.company,
                'quantite_prevision_systeme': quantite,
            },
        )
        lignes.append(ligne)
    return lignes


def calculer_offre_cycle(cycle):
    """NTSCM14 — peuple les ``LigneOffreSOP`` d'un cycle depuis
    ``apps.stock.selectors`` (jamais un import de modèle, jamais une
    duplication de logique) : pour chaque produit ayant une
    ``LigneDemandeSOP`` gelée (NTSCM13), snapshotte le stock disponible
    (``Produit.quantite_stock``, lu via ``apps.get_model('stock','Produit')``
    en LECTURE SEULE — cf. le patron déjà établi par
    ``_historique_sorties_mensuelles``) et la quantité déjà en commande
    fournisseur (``apps.stock.selectors.quantite_en_commande_produit``,
    YPROC9, déjà bâti). ``ecart_offre_demande`` = (stock disponible +
    capacité appro) − ``LigneDemandeSOP.quantite_finale``.

    Idempotent (``update_or_create`` par produit). Renvoie la liste des
    ``LigneOffreSOP``, TRIÉE par écart croissant (pénurie la plus sévère —
    la plus négative — en tête), comme ``Meta.ordering``."""
    from django.apps import apps as django_apps

    from apps.stock.selectors import quantite_en_commande_produit

    Produit = django_apps.get_model('stock', 'Produit')

    lignes_demande = LigneDemandeSOP.objects.filter(cycle=cycle).select_related('produit')

    lignes = []
    for ligne_demande in lignes_demande:
        produit = ligne_demande.produit
        stock_disponible = Decimal(str(Produit.objects.get(pk=produit.pk).quantite_stock))
        capacite_appro = Decimal(str(
            quantite_en_commande_produit(cycle.company, produit.id)))
        ecart = (stock_disponible + capacite_appro) - ligne_demande.quantite_finale

        ligne, _ = LigneOffreSOP.objects.update_or_create(
            cycle=cycle, produit=produit,
            defaults={
                'company': cycle.company,
                'stock_disponible_snapshot': stock_disponible,
                'capacite_appro_fournisseur_estimee': capacite_appro,
                'ecart_offre_demande': ecart,
            },
        )
        lignes.append(ligne)

    return sorted(lignes, key=lambda ligne_offre: ligne_offre.ecart_offre_demande)


def reouvrir_cycle(cycle, user, *, motif=''):
    """NTSCM12 — réouverture ADMIN EXPLICITE d'un cycle (retour à
    ``brouillon``), journalisée. C'est le SEUL chemin de retour en arrière de
    la machine à états (jamais via ``avancer_statut_cycle``)."""
    from apps.records.services import log_field_change

    ancien = cycle.statut
    cycle.statut = CyclePlanificationSOP.Statut.BROUILLON
    cycle.save(update_fields=['statut'])
    log_field_change(
        cycle, 'statut', ancien, CyclePlanificationSOP.Statut.BROUILLON,
        user=user, company=cycle.company,
        field_label=f'Réouverture admin ({motif or "sans motif précisé"})')
    return cycle


# ── NTSCM16 — suggestion d'achat groupée multi-fournisseurs (MOQ/paliers) ────

def _prix_pour_quantite(prix_fournisseur, quantite):
    """Prix unitaire applicable pour ``quantite`` : le palier
    (``stock.PalierPrixFournisseur``, XPUR14, déjà modélisé) dont ``qte_min``
    est le plus élevé sans dépasser ``quantite``, sinon le prix de base
    (``prix_achat``). Un ``PrixFournisseur`` sans palier garde le
    comportement historique (jamais de logique dupliquée avec XPUR14)."""
    applicable = prix_fournisseur.prix_achat
    for palier in prix_fournisseur.paliers.all():
        if palier.qte_min <= quantite:
            applicable = palier.prix
    return applicable


def _decider_quantite_achat(besoin_net, prix_fournisseur):
    """NTSCM16 — décide la quantité à commander pour couvrir ``besoin_net``
    (JAMAIS en dessous) au coût total le plus bas, MOQ et paliers de prix
    respectés.

    ADAPTATION DE PÉRIMÈTRE : le champ ``PrixFournisseur.
    quantite_minimale_commande`` (NTSCM17) n'existe pas encore — hors
    périmètre de cette lane (``apps.achats``, pas ``apps.scm``). Lu via
    ``getattr(..., None)`` : ``None`` aujourd'hui (comportement inchangé,
    comme un produit sans MOQ), la vraie valeur dès que NTSCM17 atterrit —
    SANS modification de cette fonction.

    Renvoie soit ``{'decision': 'commander', 'quantite', 'prix_unitaire',
    'cout_total'}`` (besoin >= MOQ, ou pas de MOQ connu), soit
    ``{'decision': 'sous_moq', 'moq', 'besoin_net', 'options': [...]}`` — dans
    ce dernier cas AUCUNE quantité < MOQ n'est jamais proposée : les deux
    SEULES options sont « attendre » ou « commander le MOQ » (+ alerte
    surstock)."""
    moq = getattr(prix_fournisseur, 'quantite_minimale_commande', None)

    if moq is not None and besoin_net < moq:
        return {
            'decision': 'sous_moq',
            'moq': moq,
            'besoin_net': str(besoin_net),
            'options': [
                {
                    'action': 'attendre',
                    'motif': (
                        f'Besoin net {besoin_net} sous le MOQ {moq} — '
                        'grouper avec le prochain cycle.'),
                },
                {
                    'action': 'commander_moq',
                    'quantite': moq,
                    'surstock': str(Decimal(str(moq)) - besoin_net),
                    'alerte_surstock': True,
                    'prix_unitaire': str(_prix_pour_quantite(prix_fournisseur, moq)),
                },
            ],
        }

    candidats = sorted({besoin_net} | {
        Decimal(str(p.qte_min)) for p in prix_fournisseur.paliers.all()
        if p.qte_min >= besoin_net
    })
    meilleure_quantite = min(
        candidats,
        key=lambda q: _prix_pour_quantite(prix_fournisseur, q) * q)
    prix_unitaire = _prix_pour_quantite(prix_fournisseur, meilleure_quantite)
    return {
        'decision': 'commander',
        'quantite': str(meilleure_quantite),
        'prix_unitaire': str(prix_unitaire),
        'cout_total': str(prix_unitaire * meilleure_quantite),
    }


def suggerer_achats_groupes(company):
    """NTSCM16 — regroupe PAR FOURNISSEUR le moins cher éligible les produits
    en statut « à_commander »/« rupture_imminente » (NTSCM7,
    ``selectors.tableau_bord_reappro``), en choisissant pour chacun la
    quantité qui minimise le coût total tout en couvrant le besoin net
    (jamais en dessous), MOQ et paliers de prix respectés (voir
    ``_decider_quantite_achat``).

    LECTURE SEULE — la création du BCF brouillon reste l'action séparée
    ``tableau-bord-reappro/creer-bcf/`` (NTSCM7, réutilisée telle quelle).

    Renvoie ``[{'fournisseur_id', 'fournisseur_nom', 'lignes': [...]}, ...]``."""
    from django.apps import apps as django_apps

    from apps.stock.services import cheapest_prix_fournisseur

    from . import selectors

    Produit = django_apps.get_model('stock', 'Produit')

    lignes_tableau = [
        ligne for ligne in selectors.tableau_bord_reappro(company)
        if ligne['statut'] != 'ok' and ligne['quantite_suggeree']
    ]

    groupes = {}
    for ligne in lignes_tableau:
        produit = Produit.objects.filter(
            company=company, pk=ligne['produit_id']).first()
        if produit is None:
            continue
        prix_fournisseur = cheapest_prix_fournisseur(produit)
        if prix_fournisseur is None:
            continue

        besoin_net = Decimal(str(ligne['quantite_suggeree']))
        decision = _decider_quantite_achat(besoin_net, prix_fournisseur)

        fid = prix_fournisseur.fournisseur_id
        groupe = groupes.setdefault(fid, {
            'fournisseur_id': fid,
            'fournisseur_nom': prix_fournisseur.fournisseur.nom,
            'lignes': [],
        })
        groupe['lignes'].append({
            'produit_id': produit.id,
            'produit_nom': produit.nom,
            'besoin_net': str(besoin_net),
            **decision,
        })

    return list(groupes.values())


# ── NTSCM18 — simulation « et si… » de rupture (lecture seule, en mémoire) ───

def simuler_rupture(produit, scenario, company, *, today=None):
    """NTSCM18 — simulation « et si… » EN MÉMOIRE (AUCUNE écriture DB) de
    l'impact d'un scénario hypothétique sur le réappro d'un produit.
    Réutilise ``core.stock_reorder.predict_reorder`` (FG364) — aucune
    logique de projection dupliquée.

    ``scenario`` (dict, toutes clés optionnelles) :
      - ``delai_fournisseur_jours_supplementaires`` (peut être négatif) —
        ajouté au délai fournisseur moyen mesuré ;
      - ``demande_pct`` (ex. ``20`` = +20% de conso journalière) ;
      - ``commande_annulee_quantite`` — retranchée du stock actuel AVANT
        simulation (une commande en cours qu'on suppose annulée).

    ADAPTATION : ``predict_reorder`` calcule la date de rupture PHYSIQUE
    (``stock actuel / conso``), qui NE DÉPEND PAS du délai fournisseur (un
    stock se vide à la même vitesse quel que soit le délai de la PROCHAINE
    livraison) — un ``lead_time`` allongé ne peut donc jamais, à conso et
    stock constants, avancer la date de rupture PHYSIQUE. La métrique qui
    EN DÉPEND directement, à parts égales, est la ``date_limite_commande``
    (dernier jour où passer commande pour la recevoir avant la rupture =
    ``rupture_date − lead_time_days``) : c'est CETTE date qu'un délai
    fournisseur plus long avance d'exactement le même nombre de jours — la
    métrique renvoyée pour le scénario « délai +N jours ».

    Renvoie ``{'produit_id', 'scenario', 'base': {...}, 'simule': {...},
    'delta_jours_rupture', 'delta_jours_date_limite_commande'}``."""
    from core.stock_reorder import predict_reorder

    today = today or timezone.localdate()
    scenario = scenario or {}

    lead_time_base = lead_time_moyen_fournisseur(company, produit)
    calc_base = appliquer_politique_stock(
        produit, Decimal('95'), company, lead_time_days=lead_time_base)

    politique = PolitiqueStock.objects.filter(
        company=company, produit=produit).first()
    if politique is not None:
        stock_securite = float(
            politique.stock_securite_manuel
            if politique.stock_securite_manuel is not None
            else politique.stock_securite_calcule)
    else:
        stock_securite = calc_base['stock_securite']

    stock_actuel = float(produit.quantite_stock or 0)

    resultat_base = predict_reorder(
        current_stock=stock_actuel, today=today,
        avg_daily_consumption=calc_base['avg_daily_consumption'],
        lead_time_days=lead_time_base, safety_stock=stock_securite)

    lead_time_simule = lead_time_base + float(
        scenario.get('delai_fournisseur_jours_supplementaires') or 0)
    if lead_time_simule < 0:
        lead_time_simule = 0.0
    demande_pct = float(scenario.get('demande_pct') or 0)
    conso_simulee = calc_base['avg_daily_consumption'] * (1 + demande_pct / 100)
    if conso_simulee < 0:
        conso_simulee = 0.0
    stock_simule = stock_actuel - float(
        scenario.get('commande_annulee_quantite') or 0)

    resultat_simule = predict_reorder(
        current_stock=stock_simule, today=today,
        avg_daily_consumption=conso_simulee,
        lead_time_days=lead_time_simule, safety_stock=stock_securite)

    def _serialize(resultat, lead_time):
        date_limite = None
        if resultat.rupture_date is not None:
            from datetime import timedelta
            date_limite = resultat.rupture_date - timedelta(days=lead_time)
        return {
            'reorder_now': resultat.reorder_now,
            'days_until_rupture': resultat.days_until_rupture,
            'rupture_date': (
                resultat.rupture_date.isoformat() if resultat.rupture_date else None),
            'date_limite_commande': (
                date_limite.isoformat() if date_limite else None),
            'suggested_quantity': resultat.suggested_quantity,
        }

    base = _serialize(resultat_base, lead_time_base)
    simule = _serialize(resultat_simule, lead_time_simule)

    def _delta_jours(iso_simule, iso_base):
        if not iso_simule or not iso_base:
            return None
        from datetime import date as _date
        return (_date.fromisoformat(iso_simule) - _date.fromisoformat(iso_base)).days

    return {
        'produit_id': produit.id,
        'scenario': scenario,
        'base': base,
        'simule': simule,
        'delta_jours_rupture': _delta_jours(
            simule['rupture_date'], base['rupture_date']),
        'delta_jours_date_limite_commande': _delta_jours(
            simule['date_limite_commande'], base['date_limite_commande']),
    }


# ── NTSCM19 — allocation en pénurie multi-clients (proposition, jamais une
# réservation automatique) ───────────────────────────────────────────────────

def proposer_allocation_penurie(produit, company, *, mode='fifo'):
    """NTSCM19 — propose une répartition PROPOSÉE (jamais une réservation
    automatique — l'acheteur/commercial confirme manuellement via l'action
    existante de réservation stock) du stock disponible d'un produit entre
    les ``ventes.Devis`` OUVERTS (statut ``envoye``/``accepte`` — ni
    brouillon, ni mort) qui en dépendent.

    Lu en cross-app via ``django.apps.apps.get_model`` (LECTURE SEULE, même
    patron que ``_historique_sorties_mensuelles``/``classifier_abc`` — jamais
    un ``from apps.ventes.models import ...`` statique). ``ventes.
    BonCommande`` n'est PAS interrogé séparément : il est en OneToOne avec
    son ``Devis`` source (``BonCommande.devis``) et partage donc EXACTEMENT
    sa quantité — compter le Devis suffit, jamais de double comptage.

    ``mode`` :
      - ``'fifo'`` (défaut) — par ``date_creation`` croissante (premier
        arrivé, premier servi) ;
      - ``'priorite'`` — par ``crm.Lead.priorite`` du devis (via
        ``Devis.lead``, string-FK déjà posée — haute > normale > basse), puis
        ``date_creation`` en cas d'égalité/absence de lead.

    Renvoie ``{'produit_id', 'stock_disponible', 'mode', 'propositions': [
    {'devis_id', 'reference', 'client_nom', 'quantite_demandee',
    'quantite_allouee', 'quantite_non_couverte'}, ...]}`` — jamais une
    allocation totale supérieure au disponible."""
    from django.apps import apps as django_apps

    LigneDevis = django_apps.get_model('ventes', 'LigneDevis')

    ORDRE_PRIORITE = {'haute': 0, 'normale': 1, 'basse': 2}

    lignes = (
        LigneDevis.objects
        .filter(
            devis__company=company, produit_id=produit.id,
            devis__statut__in=['envoye', 'accepte'])
        .select_related('devis', 'devis__client', 'devis__lead'))

    par_devis = {}
    for ligne in lignes:
        devis = ligne.devis
        entree = par_devis.setdefault(devis.id, {
            'devis': devis, 'quantite_demandee': Decimal('0'),
        })
        entree['quantite_demandee'] += (ligne.quantite or Decimal('0'))

    def _cle_tri(entree):
        devis = entree['devis']
        if mode == 'priorite':
            lead = getattr(devis, 'lead', None)
            priorite = getattr(lead, 'priorite', 'normale') if lead else 'normale'
            return (ORDRE_PRIORITE.get(priorite, 1), devis.date_creation)
        return (devis.date_creation,)

    entrees = sorted(par_devis.values(), key=_cle_tri)

    disponible = Decimal(str(produit.quantite_stock or 0))
    restant = disponible
    propositions = []
    for entree in entrees:
        devis = entree['devis']
        demande = entree['quantite_demandee']
        alloue = max(Decimal('0'), min(demande, restant))
        restant -= alloue
        propositions.append({
            'devis_id': devis.id,
            'reference': devis.reference,
            'client_nom': (
                devis.client.nom if devis.client_id and devis.client else ''),
            'quantite_demandee': str(demande),
            'quantite_allouee': str(alloue),
            'quantite_non_couverte': str(demande - alloue),
        })

    return {
        'produit_id': produit.id,
        'stock_disponible': str(disponible),
        'mode': mode,
        'propositions': propositions,
    }


# ── NTSCM22 — réglages opt-in du cycle S&OP automatique ─────────────────────

def parametres_scm(company):
    """NTSCM22 — réglages SCM de la société (lazy get_or_create, défaut
    ``sop_actif=False`` — n'affecte AUCUNE société qui n'a encore rien
    configuré). Voir ``models.ParametresSCM`` pour l'adaptation de
    périmètre."""
    from .models import ParametresSCM

    obj, _created = ParametresSCM.objects.get_or_create(company=company)
    return obj


# ── NTSCM25 — détection d'anomalie de demande (pic/creux inattendu) ─────────

def detecter_anomalies_demande(company, *, fenetre_mois=13):
    """NTSCM25 — réutilise ``core.anomaly`` (FG360, fondation pure — AUCUNE
    logique de détection dupliquée) sur la série de consommation mensuelle
    RÉELLE par produit (``_historique_sorties_mensuelles``, même source que
    NTSCM2/5/6). Flague le DERNIER mois écoulé quand il s'écarte de plus de
    2 écarts-types de la série (``z_threshold=2.0``).

    ADAPTATION DE PÉRIMÈTRE : le plan visait un ``AnomalyFlag`` taggé
    ``type='demande'`` — absent des choix FERMÉS de
    ``core.models.AnomalyFlag.CATEGORY_CHOICES`` (stock/paiement/fraude/
    autre, contrat de fondation ``core``, hors périmètre d'extension de
    cette lane). Catégorie ``'stock'`` (la plus proche) + ``subject_type``/
    ``metric`` portant le tag ``'scm.demande'`` à la place — filtrable de
    façon identique. Idempotent (dédupe intégrée à
    ``core.anomaly.record_anomaly`` : un flag OUVERT du même (société, mois,
    produit) n'est jamais dupliqué).

    Renvoie la liste des ``AnomalyFlag`` créés/réutilisés."""
    from django.apps import apps as django_apps

    from core.anomaly import record_outliers, scan_for_outliers

    Produit = django_apps.get_model('stock', 'Produit')

    flags = []
    for produit in Produit.objects.filter(company=company, is_archived=False):
        historique = _historique_sorties_mensuelles(company, produit.id, fenetre_mois)
        if len(historique) < 4:
            continue
        points = [
            {'id': periode, 'value': quantite, 'label': produit.nom}
            for periode, quantite in historique
        ]
        candidats = scan_for_outliers(points, z_threshold=2.0)
        dernier_periode = historique[-1][0]
        candidats_dernier_mois = [
            c for c in candidats if c.subject_id == dernier_periode]
        if not candidats_dernier_mois:
            continue
        flags.extend(record_outliers(
            candidats_dernier_mois, company=company, category='stock',
            subject_type='scm.demande',
            metric=f'consommation_mensuelle:{produit.id}'))
    return flags


# ── NTSCM27 — rapport S&OP exportable, archivé en GED à la clôture ──────────

def _construire_classeur_sop(cycle):
    """NTSCM27 — classeur .xlsx à 3 feuilles (Demande consensuelle, Offre et
    écarts, Impact financier) d'un cycle S&OP. Réutilise le style du builder
    partagé (``apps.records.xlsx.coerce_cell`` — coercition fr-MA identique
    aux autres exports) ; construit ICI en 3 feuilles car
    ``apps.records.xlsx.build_workbook`` n'en produit qu'UNE."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    from apps.records.xlsx import coerce_cell

    from . import selectors

    wb = Workbook()
    bold = Font(bold=True)

    def _en_tete(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = bold

    ws1 = wb.active
    ws1.title = 'Demande consensuelle'
    _en_tete(ws1, [
        'Produit', 'Prévision système', 'Ajusté commercial', 'Quantité finale',
        "Motif de l'ajustement"])
    for ligne in cycle.lignes_demande.select_related('produit'):
        ws1.append([coerce_cell(v) for v in [
            ligne.produit.nom, ligne.quantite_prevision_systeme,
            ligne.quantite_ajustee_commercial, ligne.quantite_finale,
            ligne.motif_ajustement]])

    ws2 = wb.create_sheet('Offre et écarts')
    _en_tete(ws2, [
        'Produit', 'Stock disponible', 'Capacité appro fournisseur',
        'Écart offre − demande'])
    for ligne in cycle.lignes_offre.select_related('produit'):
        ws2.append([coerce_cell(v) for v in [
            ligne.produit.nom, ligne.stock_disponible_snapshot,
            ligne.capacite_appro_fournisseur_estimee, ligne.ecart_offre_demande]])

    impact = selectors.impact_financier_cycle(cycle)
    ws3 = wb.create_sheet('Impact financier')
    _en_tete(ws3, ['Produit', 'Quantité finale', 'Prix de vente', 'Valeur HT'])
    for ligne in impact['lignes']:
        ws3.append([coerce_cell(v) for v in [
            ligne['produit_nom'], ligne['quantite_finale'],
            ligne['prix_vente'], ligne['valeur_ht']]])
    ws3.append([])
    ws3.append(['CA prévisionnel HT', coerce_cell(impact['ca_previsionnel_ht'])])
    ws3.append(['CA forecast HT', coerce_cell(impact['ca_forecast_ht'])])
    ws3.append(['Écart %', coerce_cell(impact['ecart_pct'])])

    return wb


def generer_compte_rendu_sop(cycle, *, user=None):
    """NTSCM27 — assemble le compte-rendu .xlsx d'un cycle S&OP
    (``_construire_classeur_sop``, 3 feuilles) et l'archive dans GED
    (``apps.ged.services.deposit_document`` — cross-app en SERVICE, jamais un
    import de modèle GED) sous un dossier « S&OP » auto-créé PAR SOCIÉTÉ.

    IDEMPOTENT par cycle (``source_type='scm.cycleplanificationsop'`` +
    ``source_id=str(cycle.id)``, mécanisme natif de ``deposit_document``) :
    un second appel pour le MÊME cycle retrouve le document déjà déposé
    plutôt que d'en créer un doublon.

    Renvoie ``(document_ged, xlsx_bytes)``."""
    import io

    from apps.ged.services import deposit_document
    from apps.records.xlsx import XLSX_CONTENT_TYPE

    wb = _construire_classeur_sop(cycle)
    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    document, _created = deposit_document(
        company=cycle.company,
        nom=f'Compte-rendu S&OP {cycle.periode}',
        source_type='scm.cycleplanificationsop', source_id=str(cycle.id),
        contenu_bytes=xlsx_bytes, mime=XLSX_CONTENT_TYPE,
        description=f'Compte-rendu du cycle de planification S&OP {cycle.periode}.',
        cabinet_nom='S&OP', folder_nom='S&OP', created_by=user)
    return document, xlsx_bytes


# ── NTSCM29 — fiche PDF interne « Politique de stock » ──────────────────────

_FICHE_POLITIQUE_STOCK_CSS = """
@page { size: A4; margin: 18mm 16mm; }
* { font-family: 'Helvetica Neue', Arial, sans-serif; color: #1a1a1a; }
h1 { font-size: 18px; margin: 0 0 4px; }
.bandeau { background: #fde68a; color: #78350f; font-weight: 700;
  font-size: 11px; padding: 6px 10px; border-radius: 6px; margin-bottom: 12px; }
table { width: 100%; border-collapse: collapse; font-size: 12px; margin-top: 10px; }
td, th { padding: 5px 6px; border-bottom: 1px solid #eee; text-align: left; }
.section-h { font-weight: 700; font-size: 13px; margin: 16px 0 6px;
  border-bottom: 1px solid #ddd; padding-bottom: 3px; }
"""


def generer_fiche_politique_stock(politique):
    """NTSCM29 — PDF INTERNE (WeasyPrint via ``core.pdf.render_pdf``, ARC12 —
    JAMAIS ``quote_engine``/``/proposal``, ce document n'est PAS un document
    client, règle #4 hors périmètre) récapitulant une ``PolitiqueStock``
    (NTSCM6) : classe ABC, niveau de service, ROP, stock min/max, stock de
    sécurité calculé vs manuel, historique des révisions (chatter générique
    ``records``, s'il en existe — NTSCM44, tâche future, alimentera ce fil).

    AUCUN ``prix_achat``/marge n'apparaît — uniquement des quantités et
    pourcentages. Bandeau « document interne » toujours visible."""
    from html import escape

    from core.pdf import render_pdf

    produit_nom = politique.produit.nom
    stock_securite_effectif = (
        politique.stock_securite_manuel
        if politique.stock_securite_manuel is not None
        else politique.stock_securite_calcule)

    historique_html = ''
    try:
        from apps.records.services import chatter_qs
        entries = list(chatter_qs(politique, politique.company)[:12])
        if entries:
            lignes_html = ''.join(
                '<tr>'
                f'<td>{escape(str(e.created_at.date()) if e.created_at else "—")}</td>'
                f'<td>{escape(e.field_label or e.body or "—")}</td>'
                f'<td>{escape(str(e.old_value) if e.old_value else "—")}</td>'
                f'<td>{escape(str(e.new_value) if e.new_value else "—")}</td>'
                '</tr>'
                for e in entries)
            historique_html = (
                '<div class="section-h">Historique des révisions (12 dernières)</div>'
                '<table><tr><th>Date</th><th>Champ</th><th>Avant</th>'
                f'<th>Après</th></tr>{lignes_html}</table>')
    except Exception:  # noqa: BLE001 - chatter optionnel, jamais bloquant
        historique_html = ''

    body = (
        f'<h1>Politique de stock — {escape(produit_nom)}</h1>'
        '<div class="bandeau">Document interne — usage acheteur, '
        'ne jamais transmettre au client.</div>'
        '<table>'
        f'<tr><td>Classe ABC</td><td>{escape(politique.classe_abc or "—")}</td></tr>'
        f'<tr><td>Niveau de service</td><td>{politique.service_level_pct}%</td></tr>'
        f'<tr><td>Point de commande (ROP)</td><td>{politique.point_commande}</td></tr>'
        f'<tr><td>Stock min</td><td>{politique.stock_min}</td></tr>'
        f'<tr><td>Stock max</td><td>{politique.stock_max}</td></tr>'
        f'<tr><td>Stock de sécurité calculé</td><td>{politique.stock_securite_calcule}</td></tr>'
        '<tr><td>Stock de sécurité manuel (override)</td>'
        f'<td>{politique.stock_securite_manuel if politique.stock_securite_manuel is not None else "—"}</td></tr>'
        f'<tr><td>Stock de sécurité effectif</td><td>{stock_securite_effectif}</td></tr>'
        '</table>'
        + historique_html
    )
    html = (
        '<!DOCTYPE html><html lang="fr"><head><meta charset="UTF-8">'
        f'<title>Politique de stock — {escape(produit_nom)}</title>'
        f'<style>{_FICHE_POLITIQUE_STOCK_CSS}</style></head><body>{body}</body></html>'
    )
    return render_pdf(html=html)


# ── NTSCM30 — assistant guidé « Créer une politique de stock » en lot ───────

def creer_politiques_en_lot(produits, service_level_pct, company):
    """NTSCM30 — applique NTSCM6 (``get_or_create`` + calcul ROP/stock de
    sécurité, NTSCM5) EN LOT à une liste de produits, pour l'assistant guidé
    ``/scm/politiques-stock/nouveau``.

    ``service_level_pct`` s'applique UNIQUEMENT à la CRÉATION d'une
    politique — jamais sur une politique déjà existante (override acheteur
    respecté, même contrat que ``recalculer_politiques_stock``).

    Renvoie la liste des ``PolitiqueStock`` créées/mises à jour, dans le même
    ordre que ``produits``."""
    classes = dict(
        ClassificationABC.objects.filter(company=company, produit__in=produits)
        .values_list('produit_id', 'classe'))

    resultats = []
    for produit in produits:
        classe = classes.get(produit.id, 'C')
        politique, created = PolitiqueStock.objects.get_or_create(
            company=company, produit=produit,
            defaults={'service_level_pct': service_level_pct, 'classe_abc': classe},
        )
        niveau = service_level_pct if created else politique.service_level_pct

        lead_time_days = lead_time_moyen_fournisseur(company, produit)
        calc = appliquer_politique_stock(
            produit, niveau, company, lead_time_days=lead_time_days)

        stock_securite_effectif = (
            politique.stock_securite_manuel
            if politique.stock_securite_manuel is not None
            else Decimal(str(calc['stock_securite'])))
        point_commande = (
            Decimal(str(calc['avg_daily_consumption'])) * Decimal(str(lead_time_days))
            + stock_securite_effectif
        )

        politique.classe_abc = classe
        politique.service_level_pct = niveau
        politique.stock_securite_calcule = Decimal(str(calc['stock_securite']))
        politique.point_commande = point_commande.quantize(Decimal('0.01'))
        politique.revise_le = timezone.now()
        politique.save(update_fields=[
            'classe_abc', 'service_level_pct', 'stock_securite_calcule',
            'point_commande', 'revise_le', 'updated_at',
        ])
        resultats.append(politique)
    return resultats
