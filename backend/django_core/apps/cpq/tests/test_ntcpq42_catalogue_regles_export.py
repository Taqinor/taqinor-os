"""NTCPQ42 — Export CSV/XLSX du catalogue de règles de compatibilité."""
from django.test import TestCase
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import AccessToken

from apps.cpq import reports
from apps.cpq.models import ContrainteCompatibilite, RegleProduitCPQ
from authentication.models import CustomUser
from testkit.factories import CompanyFactory, ProduitFactory, UserFactory

CATALOGUE_URL = '/api/django/cpq/rapports/catalogue-regles/'


def auth(user):
    api = APIClient()
    api.credentials(HTTP_AUTHORIZATION=f'Bearer {AccessToken.for_user(user)}')
    return api


class TestCatalogueReglesCompatibilite(TestCase):
    def setUp(self):
        self.company = CompanyFactory()
        self.staff = UserFactory(
            company=self.company, role_legacy=CustomUser.ROLE_RESPONSABLE)
        self.normal = UserFactory(
            company=self.company, role_legacy=CustomUser.ROLE_NORMAL)
        self.pa = ProduitFactory(company=self.company, nom='Onduleur')
        self.pb = ProduitFactory(company=self.company, nom='Batterie')
        ContrainteCompatibilite.objects.create(
            company=self.company, produit_a=self.pa, produit_b=self.pb,
            type=ContrainteCompatibilite.TypeContrainte.INCOMPATIBLE,
            message_utilisateur='Ces deux produits ne sont pas compatibles.')
        RegleProduitCPQ.objects.create(
            company=self.company, nom='Triphasé requis',
            condition_group={'field': 'kwc', 'operator': 'gte', 'value': 9},
            actions=[{'produit_id': self.pa.id, 'quantite': 2}])

    def test_catalogue_json_libelles_fr(self):
        data = reports.catalogue_regles_compatibilite(self.company)
        self.assertEqual(len(data['contraintes']), 1)
        self.assertEqual(data['contraintes'][0]['type'], 'Incompatible')
        self.assertEqual(data['contraintes'][0]['produit_a'], 'Onduleur')
        self.assertEqual(len(data['regles_produit']), 1)
        self.assertIn('kwc', data['regles_produit'][0]['condition'])
        self.assertIn('ajoute produit', data['regles_produit'][0]['actions'])

    def test_endpoint_json_reserve_staff(self):
        resp = auth(self.normal).get(CATALOGUE_URL)
        self.assertEqual(resp.status_code, 403)
        resp2 = auth(self.staff).get(CATALOGUE_URL)
        self.assertEqual(resp2.status_code, 200, resp2.data)
        self.assertEqual(len(resp2.data['contraintes']), 1)

    def test_export_xlsx_une_feuille_par_type(self):
        resp = auth(self.staff).get(CATALOGUE_URL, {'export': 'xlsx'})
        self.assertEqual(resp.status_code, 200)
        self.assertIn('spreadsheetml', resp['Content-Type'])
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(resp.content))
        self.assertEqual(
            wb.sheetnames,
            ['Contraintes de compatibilité', 'Règles produit'])
        ws1 = wb['Contraintes de compatibilité']
        self.assertEqual(ws1.cell(row=2, column=1).value, 'Onduleur')

    def test_isolation_multi_tenant(self):
        autre = CompanyFactory()
        data = reports.catalogue_regles_compatibilite(autre)
        self.assertEqual(data['contraintes'], [])
        self.assertEqual(data['regles_produit'], [])
