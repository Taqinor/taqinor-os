"""NTUX1/2 — Tests de l'app `apps.uxviews` (fondation SavedView).

Couvre : société+owner posés côté serveur (jamais du corps), isolation
multi-société, visibilité PERSONNELLE (owner uniquement) vs EQUIPE (toute la
société), garde-fou `definir-par-defaut-role` (Directeur/Admin uniquement, un
seul défaut actif par rôle+écran), garde-fous de suppression.
"""
import json

from django.contrib.auth import get_user_model
from django.contrib.contenttypes.models import ContentType
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import AccessToken

from authentication.models import Company
from apps.roles.models import Role

from .models import FavoriUtilisateur, SavedView, UxParametres

User = get_user_model()


def make_company(slug, nom):
    company, _ = Company.objects.get_or_create(slug=slug, defaults={'nom': nom})
    return company


def make_user(company, username, role_legacy='normal'):
    return User.objects.create_user(
        username=username, password='x', company=company, role_legacy=role_legacy)


def make_role(company, nom='Commercial'):
    return Role.objects.create(company=company, nom=nom, permissions=[])


def auth(user):
    api = APIClient()
    api.credentials(HTTP_AUTHORIZATION=f'Bearer {AccessToken.for_user(user)}')
    return api


def rows(resp):
    data = resp.data
    return data['results'] if isinstance(data, dict) and 'results' in data else data


class SavedViewApiTests(TestCase):
    BASE = '/api/django/uxviews/saved-views/'

    def setUp(self):
        self.co_a = make_company('uxv-a', 'A')
        self.co_b = make_company('uxv-b', 'B')
        self.directeur = make_user(self.co_a, 'uxv-directeur', role_legacy='responsable')
        self.commercial1 = make_user(self.co_a, 'uxv-com1', role_legacy='normal')
        self.commercial2 = make_user(self.co_a, 'uxv-com2', role_legacy='normal')
        self.other_co_user = make_user(self.co_b, 'uxv-b-user', role_legacy='normal')
        self.role_commercial = make_role(self.co_a, 'Commercial')

    def _payload(self, **kw):
        base = {'ecran': 'crm.leads', 'nom': 'Mes leads chauds', 'configuration': {'filtres': {}}}
        base.update(kw)
        return base

    # ── Création : société + propriétaire côté serveur ─────────────────────
    def test_create_forces_company_and_owner_server_side(self):
        resp = auth(self.commercial1).post(self.BASE, self._payload(), format='json')
        self.assertEqual(resp.status_code, 201, resp.data)
        obj = SavedView.objects.get(id=resp.data['id'])
        self.assertEqual(obj.company, self.co_a)
        self.assertEqual(obj.owner, self.commercial1)
        self.assertFalse(obj.est_defaut_role)

    def test_create_ignores_owner_and_company_in_body(self):
        payload = self._payload(owner=self.commercial2.id, company=self.co_b.id)
        resp = auth(self.commercial1).post(self.BASE, payload, format='json')
        self.assertEqual(resp.status_code, 201, resp.data)
        obj = SavedView.objects.get(id=resp.data['id'])
        self.assertEqual(obj.owner, self.commercial1)
        self.assertEqual(obj.company, self.co_a)

    # ── Visibilité : personnelle vs équipe, isolation multi-société ────────
    def test_personal_view_hidden_from_other_users(self):
        SavedView.objects.create(
            company=self.co_a, owner=self.commercial1, ecran='crm.leads', nom='Perso',
            visibilite=SavedView.Visibilite.PERSONNELLE,
        )
        resp = auth(self.commercial2).get(self.BASE, {'ecran': 'crm.leads'})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(len(rows(resp)), 0)

    def test_team_view_visible_to_company_but_not_other_company(self):
        SavedView.objects.create(
            company=self.co_a, owner=self.directeur, ecran='crm.leads', nom='Équipe',
            visibilite=SavedView.Visibilite.EQUIPE,
        )
        resp_same_co = auth(self.commercial1).get(self.BASE, {'ecran': 'crm.leads'})
        self.assertEqual(len(rows(resp_same_co)), 1)
        resp_other_co = auth(self.other_co_user).get(self.BASE, {'ecran': 'crm.leads'})
        self.assertEqual(len(rows(resp_other_co)), 0)

    def test_ecran_filter(self):
        SavedView.objects.create(
            company=self.co_a, owner=self.commercial1, ecran='crm.leads', nom='A')
        SavedView.objects.create(
            company=self.co_a, owner=self.commercial1, ecran='ventes.devis', nom='B')
        resp = auth(self.commercial1).get(self.BASE, {'ecran': 'ventes.devis'})
        self.assertEqual(len(rows(resp)), 1)
        self.assertEqual(rows(resp)[0]['nom'], 'B')

    # ── NTUX2 — définir-par-défaut-rôle : Directeur/Admin uniquement ───────
    def test_definir_par_defaut_role_forbidden_for_commercial(self):
        view = SavedView.objects.create(
            company=self.co_a, owner=self.directeur, ecran='crm.leads', nom='Équipe',
            visibilite=SavedView.Visibilite.EQUIPE, role=self.role_commercial,
        )
        resp = auth(self.commercial1).post(f'{self.BASE}{view.id}/definir-par-defaut-role/')
        self.assertEqual(resp.status_code, 403)

    def test_definir_par_defaut_role_single_default_per_role_ecran(self):
        v1 = SavedView.objects.create(
            company=self.co_a, owner=self.directeur, ecran='crm.leads', nom='V1',
            role=self.role_commercial, est_defaut_role=True,
        )
        v2 = SavedView.objects.create(
            company=self.co_a, owner=self.directeur, ecran='crm.leads', nom='V2',
            role=self.role_commercial,
        )
        resp = auth(self.directeur).post(
            f'{self.BASE}{v2.id}/definir-par-defaut-role/',
            {'role': self.role_commercial.id}, format='json',
        )
        self.assertEqual(resp.status_code, 200, resp.data)
        v1.refresh_from_db()
        v2.refresh_from_db()
        self.assertFalse(v1.est_defaut_role)
        self.assertTrue(v2.est_defaut_role)
        self.assertEqual(v2.visibilite, SavedView.Visibilite.EQUIPE)

    def test_definir_par_defaut_role_requires_role(self):
        view = SavedView.objects.create(
            company=self.co_a, owner=self.directeur, ecran='crm.leads', nom='V',
        )
        resp = auth(self.directeur).post(f'{self.BASE}{view.id}/definir-par-defaut-role/')
        self.assertEqual(resp.status_code, 400)

    # ── Suppression : garde-fous ────────────────────────────────────────────
    def test_delete_own_personal_view(self):
        view = SavedView.objects.create(
            company=self.co_a, owner=self.commercial1, ecran='crm.leads', nom='V')
        resp = auth(self.commercial1).delete(f'{self.BASE}{view.id}/')
        self.assertEqual(resp.status_code, 204)

    def test_cannot_delete_default_role_view_without_permission(self):
        view = SavedView.objects.create(
            company=self.co_a, owner=self.directeur, ecran='crm.leads', nom='V',
            role=self.role_commercial, est_defaut_role=True,
            visibilite=SavedView.Visibilite.EQUIPE,
        )
        resp = auth(self.commercial1).delete(f'{self.BASE}{view.id}/')
        self.assertEqual(resp.status_code, 403)
        self.assertTrue(SavedView.objects.filter(id=view.id).exists())

    def test_directeur_can_delete_default_role_view(self):
        view = SavedView.objects.create(
            company=self.co_a, owner=self.directeur, ecran='crm.leads', nom='V',
            role=self.role_commercial, est_defaut_role=True,
            visibilite=SavedView.Visibilite.EQUIPE,
        )
        resp = auth(self.directeur).delete(f'{self.BASE}{view.id}/')
        self.assertEqual(resp.status_code, 204)

    # ── NTUX23 — rapport « configuration des vues actives » (gouvernance) ──
    def test_toutes_company_forbidden_for_commercial(self):
        resp = auth(self.commercial1).get(f'{self.BASE}toutes-company/')
        self.assertEqual(resp.status_code, 403)

    def test_toutes_company_lists_every_view_of_the_company_beyond_perso_equipe_filter(self):
        # Vue PERSONNELLE d'un AUTRE utilisateur — invisible via list() normal,
        # mais visible ici (rapport de gouvernance Directeur/Admin).
        SavedView.objects.create(
            company=self.co_a, owner=self.commercial1, ecran='crm.leads', nom='Perso com1',
            visibilite=SavedView.Visibilite.PERSONNELLE,
        )
        SavedView.objects.create(
            company=self.co_a, owner=self.commercial2, ecran='ventes.devis', nom='Perso com2',
            visibilite=SavedView.Visibilite.PERSONNELLE,
        )
        # Vue d'une AUTRE company — ne doit jamais apparaître.
        SavedView.objects.create(
            company=self.co_b, owner=self.other_co_user, ecran='crm.leads', nom='Autre société',
        )
        resp = auth(self.directeur).get(f'{self.BASE}toutes-company/')
        self.assertEqual(resp.status_code, 200)
        noms = {r['nom'] for r in rows(resp)}
        self.assertEqual(noms, {'Perso com1', 'Perso com2'})

    def test_export_xlsx_forbidden_for_commercial(self):
        resp = auth(self.commercial1).get(f'{self.BASE}export-xlsx/')
        self.assertEqual(resp.status_code, 403)

    def test_export_xlsx_returns_a_workbook_with_one_row_per_view_and_default_role(self):
        SavedView.objects.create(
            company=self.co_a, owner=self.directeur, ecran='crm.leads', nom='Équipe',
            visibilite=SavedView.Visibilite.EQUIPE, role=self.role_commercial, est_defaut_role=True,
        )
        SavedView.objects.create(
            company=self.co_a, owner=self.commercial1, ecran='ventes.devis', nom='Perso',
        )
        resp = auth(self.directeur).get(f'{self.BASE}export-xlsx/')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        # En-tête + 2 lignes de données.
        self.assertEqual(ws.max_row, 3)
        header = [c.value for c in ws[1]]
        self.assertEqual(header, ['Écran', 'Nom', 'Propriétaire', 'Visibilité', 'Rôle par défaut', 'Dernière modification'])
        noms = {ws.cell(row=r, column=2).value for r in (2, 3)}
        self.assertEqual(noms, {'Équipe', 'Perso'})
        # Le rôle par défaut n'apparaît que pour la vue qui le porte.
        roles_col = {ws.cell(row=r, column=2).value: ws.cell(row=r, column=5).value for r in (2, 3)}
        self.assertEqual(roles_col['Équipe'], 'Commercial')
        # openpyxl relit une cellule chaîne-vide comme None (l'export écrit bien '').
        self.assertIn(roles_col['Perso'], (None, ''))

    # ── NTUX34 — import CSV de vues sauvegardées entre environnements ──────
    def _csv(self, rows):
        lines = ['ecran,nom,configuration']
        for ecran, nom, configuration in rows:
            lines.append(f'{ecran},{nom},"{configuration}"')
        return SimpleUploadedFile(
            'vues.csv', ('\n'.join(lines)).encode('utf-8'), content_type='text/csv')

    def test_importer_forbidden_for_commercial(self):
        fichier = self._csv([('crm.leads', 'V1', json.dumps({}))])
        resp = auth(self.commercial1).post(f'{self.BASE}importer/', {'fichier': fichier}, format='multipart')
        self.assertEqual(resp.status_code, 403)

    def test_importer_creates_valid_rows_as_personal_views_owned_by_caller(self):
        config = json.dumps({'colonnes_visibles': ['nom'], 'filtres': {'op': 'AND', 'conditions': []}})
        fichier = self._csv([('crm.leads', 'Mes leads chauds', config.replace('"', '""'))])
        resp = auth(self.directeur).post(f'{self.BASE}importer/', {'fichier': fichier}, format='multipart')
        self.assertEqual(resp.status_code, 200, resp.data)
        self.assertEqual(len(resp.data['created']), 1)
        self.assertEqual(resp.data['erreurs'], [])
        view = SavedView.objects.get(id=resp.data['created'][0]['id'])
        self.assertEqual(view.owner, self.directeur)
        self.assertEqual(view.company, self.co_a)
        self.assertEqual(view.visibilite, SavedView.Visibilite.PERSONNELLE)
        self.assertEqual(view.configuration['colonnes_visibles'], ['nom'])

    def test_importer_reports_invalid_json_line_with_its_number_and_still_imports_valid_rows(self):
        good = json.dumps({'colonnes_visibles': ['nom']}).replace('"', '""')
        lines = [
            'ecran,nom,configuration',
            f'crm.leads,Vue valide,"{good}"',
            'crm.leads,Vue cassée,"{pas du json valide"',
        ]
        fichier = SimpleUploadedFile(
            'vues.csv', ('\n'.join(lines)).encode('utf-8'), content_type='text/csv')
        resp = auth(self.directeur).post(f'{self.BASE}importer/', {'fichier': fichier}, format='multipart')
        self.assertEqual(resp.status_code, 200, resp.data)
        self.assertEqual(len(resp.data['created']), 1)
        self.assertEqual(len(resp.data['erreurs']), 1)
        self.assertEqual(resp.data['erreurs'][0]['ligne'], 2)
        self.assertIn('JSON', resp.data['erreurs'][0]['message'])

    def test_importer_never_silently_overwrites_renames_with_import_suffix(self):
        SavedView.objects.create(
            company=self.co_a, owner=self.directeur, ecran='crm.leads', nom='Mes leads',
            configuration={'ancien': True},
        )
        config = json.dumps({'nouveau': True}).replace('"', '""')
        fichier = self._csv([('crm.leads', 'Mes leads', config)])
        resp = auth(self.directeur).post(f'{self.BASE}importer/', {'fichier': fichier}, format='multipart')
        self.assertEqual(resp.status_code, 200, resp.data)
        self.assertEqual(SavedView.objects.filter(company=self.co_a, owner=self.directeur, ecran='crm.leads').count(), 2)
        imported = SavedView.objects.get(
            company=self.co_a, owner=self.directeur, ecran='crm.leads', nom='Mes leads (import)')
        self.assertEqual(imported.configuration, {'nouveau': True})
        original = SavedView.objects.get(
            company=self.co_a, owner=self.directeur, ecran='crm.leads', nom='Mes leads')
        self.assertEqual(original.configuration, {'ancien': True})

    def test_importer_rejects_malformed_configuration_structure(self):
        bad = json.dumps({'colonnes_visibles': 'pas-une-liste'}).replace('"', '""')
        fichier = self._csv([('crm.leads', 'Vue mal formée', bad)])
        resp = auth(self.directeur).post(f'{self.BASE}importer/', {'fichier': fichier}, format='multipart')
        self.assertEqual(resp.status_code, 200, resp.data)
        self.assertEqual(resp.data['created'], [])
        self.assertEqual(len(resp.data['erreurs']), 1)


class FavoriUtilisateurApiTests(TestCase):
    """NTUX12 — favoris épinglés, STRICTEMENT personnels."""

    BASE = '/api/django/uxviews/favoris/'

    def setUp(self):
        self.co_a = make_company('uxfav-a', 'Fav A')
        self.co_b = make_company('uxfav-b', 'Fav B')
        self.com1 = make_user(self.co_a, 'uxfav-com1')
        self.com2 = make_user(self.co_a, 'uxfav-com2')
        self.other_co_user = make_user(self.co_b, 'uxfav-b-user')
        # Cible générique : n'importe quel enregistrement fait l'affaire —
        # `uxviews` n'importe aucune app métier (contenttypes suffit). On épingle
        # des `SavedView`, ce qui garde le test dans le périmètre de l'app.
        self.cible1 = SavedView.objects.create(
            company=self.co_a, owner=self.com1, ecran='crm.leads', nom='Cible 1')
        self.cible2 = SavedView.objects.create(
            company=self.co_a, owner=self.com1, ecran='ventes.devis', nom='Cible 2')

    def _epingler(self, api, cible, **extra):
        payload = {'modele': 'uxviews.savedview', 'object_id': cible.pk}
        payload.update(extra)
        return api.post(self.BASE, payload, format='json')

    def test_epingler_pose_company_et_owner_cote_serveur(self):
        resp = self._epingler(auth(self.com1), self.cible1)
        self.assertEqual(resp.status_code, 201, resp.data)
        favori = FavoriUtilisateur.objects.get()
        self.assertEqual(favori.company, self.co_a)
        self.assertEqual(favori.owner, self.com1)
        self.assertEqual(favori.cle_modele, 'uxviews.savedview')
        self.assertEqual(favori.object_id, self.cible1.pk)

    def test_company_et_owner_du_corps_sont_ignores(self):
        resp = self._epingler(
            auth(self.com1), self.cible1,
            company=self.co_b.pk, owner=self.com2.pk)
        self.assertEqual(resp.status_code, 201, resp.data)
        favori = FavoriUtilisateur.objects.get()
        self.assertEqual(favori.company, self.co_a)
        self.assertEqual(favori.owner, self.com1)

    def test_modele_inconnu_est_rejete(self):
        api = auth(self.com1)
        self.assertEqual(
            api.post(self.BASE, {'modele': 'nimporte.quoi', 'object_id': 1},
                     format='json').status_code, 400)
        self.assertEqual(
            api.post(self.BASE, {'modele': 'pasdepoint', 'object_id': 1},
                     format='json').status_code, 400)

    def test_epingler_deux_fois_est_un_no_op(self):
        api = auth(self.com1)
        self._epingler(api, self.cible1)
        self._epingler(api, self.cible1)
        self.assertEqual(FavoriUtilisateur.objects.count(), 1)

    def test_favoris_strictement_personnels(self):
        self._epingler(auth(self.com1), self.cible1)
        favori = FavoriUtilisateur.objects.get()
        # Un collègue de la MÊME société ne voit rien.
        self.assertEqual(len(rows(auth(self.com2).get(self.BASE))), 0)
        self.assertEqual(
            auth(self.com2).get(f'{self.BASE}{favori.pk}/').status_code, 404)
        self.assertEqual(
            auth(self.com2).delete(f'{self.BASE}{favori.pk}/').status_code, 404)
        # Une autre société non plus.
        self.assertEqual(len(rows(auth(self.other_co_user).get(self.BASE))), 0)

    def test_libelle_resolu_depuis_la_cible(self):
        self._epingler(auth(self.com1), self.cible1)
        ligne = rows(auth(self.com1).get(self.BASE))[0]
        self.assertEqual(ligne['libelle'], str(self.cible1))
        self.assertEqual(ligne['modele'], 'uxviews.savedview')

    def test_libelle_none_si_la_cible_a_disparu(self):
        self._epingler(auth(self.com1), self.cible1)
        self.cible1.delete()
        ligne = rows(auth(self.com1).get(self.BASE))[0]
        self.assertIsNone(ligne['libelle'])

    def test_ordre_par_defaut_ajoute_en_fin_de_liste(self):
        api = auth(self.com1)
        self._epingler(api, self.cible1)
        self._epingler(api, self.cible2)
        ordres = list(FavoriUtilisateur.objects.order_by('ordre')
                      .values_list('object_id', 'ordre'))
        self.assertEqual(ordres, [(self.cible1.pk, 0), (self.cible2.pk, 1)])

    def test_reordonner_remonte_le_favori_en_tete(self):
        api = auth(self.com1)
        self._epingler(api, self.cible1)
        self._epingler(api, self.cible2)
        dernier = FavoriUtilisateur.objects.get(object_id=self.cible2.pk)
        resp = api.post(f'{self.BASE}{dernier.pk}/reordonner/',
                        {'ordre': 0}, format='json')
        self.assertEqual(resp.status_code, 200, resp.data)
        self.assertEqual([ligne['object_id'] for ligne in resp.data],
                         [self.cible2.pk, self.cible1.pk])
        # Persisté : le rechargement conserve le nouvel ordre (NTUX21).
        self.assertEqual(
            [ligne['object_id'] for ligne in rows(api.get(self.BASE))],
            [self.cible2.pk, self.cible1.pk])

    def test_reordonner_refuse_une_position_invalide(self):
        api = auth(self.com1)
        self._epingler(api, self.cible1)
        favori = FavoriUtilisateur.objects.get()
        self.assertEqual(
            api.post(f'{self.BASE}{favori.pk}/reordonner/', {},
                     format='json').status_code, 400)
        self.assertEqual(
            api.post(f'{self.BASE}{favori.pk}/reordonner/', {'ordre': -1},
                     format='json').status_code, 400)

    def test_reordonner_refuse_le_favori_dun_autre(self):
        self._epingler(auth(self.com1), self.cible1)
        favori = FavoriUtilisateur.objects.get()
        resp = auth(self.com2).post(f'{self.BASE}{favori.pk}/reordonner/',
                                    {'ordre': 0}, format='json')
        self.assertEqual(resp.status_code, 404)

    def test_desepingler(self):
        self._epingler(auth(self.com1), self.cible1)
        favori = FavoriUtilisateur.objects.get()
        self.assertEqual(
            auth(self.com1).delete(f'{self.BASE}{favori.pk}/').status_code, 204)
        self.assertEqual(FavoriUtilisateur.objects.count(), 0)


class UxParametresApiTests(TestCase):
    """NTUX27 — réglages UX par société (singleton) et ce qu'ils GOUVERNENT."""

    BASE = '/api/django/uxviews/parametres/'
    VUES = '/api/django/uxviews/saved-views/'

    def setUp(self):
        self.co_a = make_company('uxprm-a', 'Prm A')
        self.co_b = make_company('uxprm-b', 'Prm B')
        self.directeur = make_user(self.co_a, 'uxprm-directeur', role_legacy='responsable')
        self.commercial = make_user(self.co_a, 'uxprm-com', role_legacy='normal')
        self.directeur_b = make_user(self.co_b, 'uxprm-b-directeur', role_legacy='responsable')

    def _desactiver_partage(self):
        parametres = UxParametres.get_or_default(self.co_a)
        parametres.permettre_vues_partagees_equipe = False
        parametres.save(update_fields=['permettre_vues_partagees_equipe'])
        return parametres

    def test_lecture_cree_les_reglages_au_defaut(self):
        self.assertEqual(UxParametres.objects.count(), 0)
        resp = auth(self.commercial).get(self.BASE)
        self.assertEqual(resp.status_code, 200, resp.data)
        self.assertEqual(resp.data['duree_hover_peek_ms'], 400)
        self.assertEqual(resp.data['duree_undo_toast_s'], 10)
        self.assertTrue(resp.data['permettre_vues_partagees_equipe'])
        self.assertEqual(resp.data['roles_autorises_definir_defaut'], [])
        self.assertEqual(UxParametres.objects.count(), 1)

    def test_ecriture_reservee_directeur_admin(self):
        self.assertEqual(
            auth(self.commercial).patch(
                self.BASE, {'duree_undo_toast_s': 30}, format='json').status_code, 403)
        resp = auth(self.directeur).patch(
            self.BASE, {'duree_undo_toast_s': 30}, format='json')
        self.assertEqual(resp.status_code, 200, resp.data)
        self.assertEqual(
            UxParametres.get_or_default(self.co_a).duree_undo_toast_s, 30)

    def test_reglages_isoles_par_societe(self):
        auth(self.directeur).patch(self.BASE, {'duree_hover_peek_ms': 900}, format='json')
        resp = auth(self.directeur_b).get(self.BASE)
        self.assertEqual(resp.data['duree_hover_peek_ms'], 400)
        self.assertEqual(UxParametres.objects.count(), 2)

    def test_role_dune_autre_societe_refuse(self):
        role_b = make_role(self.co_b, 'Directeur B')
        resp = auth(self.directeur).patch(
            self.BASE, {'roles_autorises_definir_defaut': [role_b.pk]}, format='json')
        self.assertEqual(resp.status_code, 400, resp.data)

    # ── Ce que le réglage GOUVERNE (NTUX1/NTUX2) ────────────────────────────
    def test_partage_desactive_refuse_une_nouvelle_vue_equipe(self):
        self._desactiver_partage()
        resp = auth(self.commercial).post(self.VUES, {
            'ecran': 'crm.leads', 'nom': 'Partagée', 'configuration': {},
            'visibilite': 'EQUIPE',
        }, format='json')
        self.assertEqual(resp.status_code, 400, resp.data)
        # Une vue PERSONNELLE reste évidemment créable.
        self.assertEqual(auth(self.commercial).post(self.VUES, {
            'ecran': 'crm.leads', 'nom': 'Perso', 'configuration': {},
        }, format='json').status_code, 201)

    def test_partage_desactive_ne_supprime_ni_ne_departage_lexistant(self):
        vue = SavedView.objects.create(
            company=self.co_a, owner=self.commercial, ecran='crm.leads',
            nom='Déjà partagée', visibilite=SavedView.Visibilite.EQUIPE)
        self._desactiver_partage()
        # Toujours en base, toujours partagée, toujours visible de l'équipe.
        vue.refresh_from_db()
        self.assertEqual(vue.visibilite, SavedView.Visibilite.EQUIPE)
        noms = [ligne['nom'] for ligne in rows(auth(self.directeur).get(self.VUES))]
        self.assertIn('Déjà partagée', noms)
        # Une édition SANS bascule de visibilité passe (la vue n'est pas gelée
        # pour son propriétaire, elle est juste non re-partageable).
        resp = auth(self.commercial).patch(
            f'{self.VUES}{vue.pk}/', {'nom': 'Renommée'}, format='json')
        self.assertEqual(resp.status_code, 200, resp.data)
        vue.refresh_from_db()
        self.assertEqual(vue.visibilite, SavedView.Visibilite.EQUIPE)

    def test_partage_desactive_refuse_une_vue_par_defaut_de_role(self):
        role = make_role(self.co_a, 'Commercial')
        vue = SavedView.objects.create(
            company=self.co_a, owner=self.directeur, ecran='crm.leads', nom='Défaut')
        self._desactiver_partage()
        resp = auth(self.directeur).post(
            f'{self.VUES}{vue.pk}/definir-par-defaut-role/',
            {'role': role.pk}, format='json')
        self.assertEqual(resp.status_code, 400, resp.data)

    def test_roles_autorises_restreint_qui_pose_le_defaut(self):
        from apps.roles.models import Role

        role_autorise = Role.objects.create(
            company=self.co_a, nom='Directeur commercial', permissions=['crm_gerer'])
        role_autre = Role.objects.create(
            company=self.co_a, nom='Chef de projet', permissions=['crm_gerer'])
        parametres = UxParametres.get_or_default(self.co_a)
        parametres.roles_autorises_definir_defaut.set([role_autorise])

        refuse = make_user(self.co_a, 'uxprm-refuse')
        refuse.role = role_autre
        refuse.save(update_fields=['role'])
        permis = make_user(self.co_a, 'uxprm-permis')
        permis.role = role_autorise
        permis.save(update_fields=['role'])

        # Partagée à l'équipe : sinon la vue d'un AUTRE propriétaire est hors
        # `get_queryset` et l'appel renverrait 404 avant d'atteindre le garde
        # de rôle que ce test vise.
        vue = SavedView.objects.create(
            company=self.co_a, owner=self.directeur, ecran='crm.leads',
            nom='Défaut', visibilite=SavedView.Visibilite.EQUIPE)
        url = f'{self.VUES}{vue.pk}/definir-par-defaut-role/'
        self.assertEqual(
            auth(refuse).post(url, {'role': role_autorise.pk}, format='json'
                              ).status_code, 403)
        self.assertEqual(
            auth(permis).post(url, {'role': role_autorise.pk}, format='json'
                              ).status_code, 200)

    def test_liste_vide_de_roles_autorises_ne_restreint_rien(self):
        role = make_role(self.co_a, 'Commercial')
        vue = SavedView.objects.create(
            company=self.co_a, owner=self.directeur, ecran='crm.leads', nom='Défaut')
        resp = auth(self.directeur).post(
            f'{self.VUES}{vue.pk}/definir-par-defaut-role/',
            {'role': role.pk}, format='json')
        self.assertEqual(resp.status_code, 200, resp.data)


class NTUX30DigestFavorisObsoletesTests(TestCase):
    """NTUX30 — digest hebdomadaire des favoris pointant vers une cible
    supprimée : jamais de suppression automatique, une notification par
    propriétaire (jamais une par favori)."""

    def setUp(self):
        self.co = make_company('uxdig-a', 'Dig A')
        self.user = make_user(self.co, 'uxdig-user')
        self.cible_vivante = SavedView.objects.create(
            company=self.co, owner=self.user, ecran='crm.leads', nom='Vivante')
        self.cible_a_supprimer = SavedView.objects.create(
            company=self.co, owner=self.user, ecran='crm.leads', nom='À supprimer')

    def _epingler(self, cible):
        return FavoriUtilisateur.objects.create(
            company=self.co, owner=self.user,
            content_type=ContentType.objects.get_for_model(SavedView),
            object_id=cible.pk,
        )

    def test_favori_vivant_ne_declenche_rien(self):
        from .tasks import digest_favoris_obsoletes_hebdo

        self._epingler(self.cible_vivante)
        resultat = digest_favoris_obsoletes_hebdo()
        self.assertEqual(resultat['proprietaires_notifies'], 0)

    def test_favori_mort_notifie_le_proprietaire_une_seule_fois(self):
        from apps.notifications.models import Notification

        from .tasks import digest_favoris_obsoletes_hebdo

        self._epingler(self.cible_vivante)
        favori_mort_1 = self._epingler(self.cible_a_supprimer)
        cible2 = SavedView.objects.create(
            company=self.co, owner=self.user, ecran='ventes.devis', nom='À supprimer aussi')
        favori_mort_2 = self._epingler(cible2)
        self.cible_a_supprimer.delete()
        cible2.delete()

        resultat = digest_favoris_obsoletes_hebdo()
        self.assertEqual(resultat['proprietaires_notifies'], 1)
        # Deux favoris morts, UNE seule notification (pas de spam par favori).
        notifs = Notification.objects.filter(
            recipient=self.user, event_type='uxviews_favoris_obsoletes')
        self.assertEqual(notifs.count(), 1)
        self.assertIn('2', notifs.first().body)
        # Aucune suppression automatique des favoris morts.
        self.assertTrue(FavoriUtilisateur.objects.filter(pk=favori_mort_1.pk).exists())
        self.assertTrue(FavoriUtilisateur.objects.filter(pk=favori_mort_2.pk).exists())

    def test_tache_beat_est_planifiee(self):
        from erp_agentique.celery import app as celery_app

        taches = {v['task'] for v in celery_app.conf.beat_schedule.values()}
        self.assertIn('uxviews.digest_favoris_obsoletes_hebdo', taches)
