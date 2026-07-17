"""Tests de l'export .xlsx des idées (NTIDE12).

Couvre : en-têtes/contenu du classeur, filtres statut appliqués, palier
d'accès admin/responsable.
"""
import io

from django.contrib.auth import get_user_model
from django.test import TestCase
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import AccessToken

from authentication.models import Company

from apps.innovation.models import Idee

User = get_user_model()


def make_company(slug, nom):
    company, _ = Company.objects.get_or_create(slug=slug, defaults={'nom': nom})
    return company


def make_user(company, username, role='responsable'):
    return User.objects.create_user(
        username=username, password='x', company=company, role_legacy=role)


def auth(user):
    api = APIClient()
    api.credentials(HTTP_AUTHORIZATION=f'Bearer {AccessToken.for_user(user)}')
    return api


def _xlsx_rows(content):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    return list(ws.iter_rows(values_only=True))


class ExportXlsxTests(TestCase):
    BASE = '/api/django/innovation/idees/'

    def setUp(self):
        self.co_a = make_company('innov-exp-a', 'A')
        self.admin_a = make_user(self.co_a, 'innov-exp-admin', role='admin')
        self.normal_a = make_user(self.co_a, 'innov-exp-normal', role='normal')

    def test_export_headers_and_row(self):
        Idee.objects.create(
            company=self.co_a, titre='Idée exportée', contexte='SAV',
            votes_count=3)
        resp = auth(self.admin_a).get(f'{self.BASE}export-xlsx/')
        self.assertEqual(resp.status_code, 200)
        rows = _xlsx_rows(resp.content)
        self.assertEqual(
            rows[0],
            ('Titre', 'Auteur', 'Contexte', 'Statut', 'Votes', 'Créée le', 'Notes'))
        self.assertEqual(rows[1][0], 'Idée exportée')
        self.assertEqual(rows[1][2], 'SAV')
        self.assertEqual(rows[1][4], 3)

    def test_export_respects_statut_filter(self):
        Idee.objects.create(
            company=self.co_a, titre='Ouverte', statut=Idee.Statut.OUVERT)
        Idee.objects.create(
            company=self.co_a, titre='Retenue', statut=Idee.Statut.RETENUE)
        resp = auth(self.admin_a).get(
            f'{self.BASE}export-xlsx/', {'statut': 'retenue'})
        rows = _xlsx_rows(resp.content)
        self.assertEqual(len(rows), 2)  # en-tête + 1 ligne
        self.assertEqual(rows[1][0], 'Retenue')

    def test_export_permission_admin_only(self):
        resp = auth(self.normal_a).get(f'{self.BASE}export-xlsx/')
        self.assertEqual(resp.status_code, 403)
