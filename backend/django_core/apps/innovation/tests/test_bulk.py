"""Tests des actions en masse sur idées (NTIDE13).

Couvre : palier d'accès admin/responsable, bulk set-statut (avec ignorées si
transition illégale), bulk add/remove tag (réutilise ``records.Tag``/
``TaggedItem``, FG9 — aucune nouvelle table), bulk export sur la sélection
(NTIDE12 réutilisé, court-circuite le JSON).
"""
import io

from django.contrib.auth import get_user_model
from django.test import TestCase
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import AccessToken

from authentication.models import Company

from apps.innovation.models import Idee

User = get_user_model()


def make_company(slug, nom):
    company, _ = Company.objects.get_or_create(slug=slug, defaults={'nom': nom})
    return company


def make_user(company, username, role='responsable'):
    return User.objects.create_user(
        username=username, password='x', company=company, role_legacy=role)


def auth(user):
    api = APIClient()
    api.credentials(HTTP_AUTHORIZATION=f'Bearer {AccessToken.for_user(user)}')
    return api


def _xlsx_rows(content):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    return list(ws.iter_rows(values_only=True))


class BulkActionsTests(TestCase):
    BASE = '/api/django/innovation/idees/bulk/'

    def setUp(self):
        self.co_a = make_company('innov-bulk-a', 'A')
        self.admin_a = make_user(self.co_a, 'innov-bulk-admin', role='admin')
        self.normal_a = make_user(self.co_a, 'innov-bulk-normal', role='normal')
        self.idee1 = Idee.objects.create(company=self.co_a, titre='Une')
        self.idee2 = Idee.objects.create(company=self.co_a, titre='Deux')

    def test_bulk_requires_ids(self):
        resp = auth(self.admin_a).post(
            self.BASE, {'action': 'set_statut', 'ids': []}, format='json')
        self.assertEqual(resp.status_code, 400)

    def test_bulk_permission_admin_only(self):
        resp = auth(self.normal_a).post(
            self.BASE,
            {'action': 'set_statut', 'ids': [self.idee1.id], 'statut': 'examinee'},
            format='json')
        self.assertEqual(resp.status_code, 403)

    def test_bulk_set_statut_applies_and_ignores_invalid(self):
        self.idee2.statut = Idee.Statut.RETENUE
        self.idee2.save(update_fields=['statut'])
        resp = auth(self.admin_a).post(
            self.BASE,
            {
                'action': 'set_statut',
                'ids': [self.idee1.id, self.idee2.id],
                'statut': 'examinee',
            },
            format='json')
        self.assertEqual(resp.status_code, 200, resp.data)
        self.idee1.refresh_from_db()
        self.idee2.refresh_from_db()
        # idee1 : ouvert → examinée (légal). idee2 : retenue → examinée
        # (illégal, ignorée — reste retenue).
        self.assertEqual(self.idee1.statut, Idee.Statut.EXAMINEE)
        self.assertEqual(self.idee2.statut, Idee.Statut.RETENUE)
        self.assertEqual(resp.data['appliquees'], [self.idee1.id])
        self.assertEqual(resp.data['ignorees'], [self.idee2.id])

    def test_bulk_add_tag(self):
        from apps.records.models import Tag, TaggedItem

        resp = auth(self.admin_a).post(
            self.BASE,
            {
                'action': 'add_tag',
                'ids': [self.idee1.id, self.idee2.id],
                'tag': 'Prioritaire',
            },
            format='json')
        self.assertEqual(resp.status_code, 200, resp.data)
        tag = Tag.objects.get(company=self.co_a, nom='Prioritaire')
        from django.contrib.contenttypes.models import ContentType
        ct = ContentType.objects.get_for_model(Idee)
        self.assertEqual(
            TaggedItem.objects.filter(tag=tag, content_type=ct).count(), 2)

    def test_bulk_remove_tag(self):
        from apps.records.models import Tag, TaggedItem
        from django.contrib.contenttypes.models import ContentType

        tag = Tag.objects.create(company=self.co_a, nom='Prioritaire')
        ct = ContentType.objects.get_for_model(Idee)
        TaggedItem.objects.create(tag=tag, content_type=ct, object_id=self.idee1.id)
        TaggedItem.objects.create(tag=tag, content_type=ct, object_id=self.idee2.id)

        resp = auth(self.admin_a).post(
            self.BASE,
            {
                'action': 'remove_tag',
                'ids': [self.idee1.id],
                'tag': 'Prioritaire',
            },
            format='json')
        self.assertEqual(resp.status_code, 200, resp.data)
        self.assertEqual(TaggedItem.objects.filter(tag=tag, content_type=ct).count(), 1)

    def test_bulk_add_tag_requires_tag_name(self):
        resp = auth(self.admin_a).post(
            self.BASE,
            {'action': 'add_tag', 'ids': [self.idee1.id], 'tag': ''},
            format='json')
        self.assertEqual(resp.status_code, 400)

    def test_bulk_export_returns_xlsx_for_selection(self):
        resp = auth(self.admin_a).post(
            self.BASE,
            {'action': 'export', 'ids': [self.idee1.id]},
            format='json')
        self.assertEqual(resp.status_code, 200)
        rows = _xlsx_rows(resp.content)
        self.assertEqual(len(rows), 2)  # en-tête + 1 ligne (idee1 seule)
        self.assertEqual(rows[1][0], 'Une')

    def test_bulk_unknown_action_rejected(self):
        resp = auth(self.admin_a).post(
            self.BASE,
            {'action': 'delete_everything', 'ids': [self.idee1.id]},
            format='json')
        self.assertEqual(resp.status_code, 400)
