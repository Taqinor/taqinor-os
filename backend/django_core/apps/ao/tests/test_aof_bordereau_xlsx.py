"""AOF127 — le classeur du bordereau : des NOMBRES, pas des chaînes.

    python -m unittest apps.ao.tests.test_aof_bordereau_xlsx -v
"""
import io
import re
import unittest
from decimal import Decimal

from openpyxl import load_workbook

from apps.ao.fabrique.contexte import construire_contexte
from apps.ao.fabrique.rendus import bordereau_xlsx as xlsx
from apps.ao.tests.aof_fixtures import (ARRETE_TTC, SOUS_TOTAUX_ATTENDUS,
                                        TOTAL_HT, TOTAL_TTC, TVA,
                                        bordereau_depose, contexte_dossier)

COL_TOTAL = 'G'


def classeur():
    lignes = bordereau_depose()
    octets = xlsx.rendre(lignes, construire_contexte(contexte_dossier(lignes)))
    return load_workbook(io.BytesIO(octets)), lignes


def cellules(feuille):
    return [cellule for rangee in feuille.iter_rows() for cellule in rangee
            if cellule.value is not None]


class TestClasseurOuvrable(unittest.TestCase):

    def setUp(self):
        self.classeur, self.lignes = classeur()
        self.feuille = self.classeur.active

    def test_le_classeur_s_ouvre(self):
        self.assertEqual(self.feuille.title, 'Bordereau des prix')

    def test_les_colonnes_attendues(self):
        entetes = [c.value for c in cellules(self.feuille)
                   if c.value in [titre for titre, _ in xlsx.COLONNES]]
        for titre, _ in xlsx.COLONNES:
            self.assertIn(titre, entetes, titre)

    def test_toutes_les_lignes_sont_presentes(self):
        designations = {c.value for c in cellules(self.feuille)}
        for ligne in self.lignes:
            self.assertIn(ligne['designation'], designations,
                          ligne['designation'])

    def test_les_sections_sont_ecrites(self):
        valeurs = {c.value for c in cellules(self.feuille)}
        for section in SOUS_TOTAUX_ATTENDUS:
            self.assertIn(section, valeurs, section)

    def test_mise_en_page_a4_paysage(self):
        self.assertEqual(self.feuille.page_setup.orientation, 'landscape')
        self.assertEqual(str(self.feuille.page_setup.paperSize),
                         str(self.feuille.PAPERSIZE_A4))

    def test_la_feuille_est_protegee(self):
        self.assertTrue(self.feuille.protection.sheet)


class TestAucunMontantEnTexte(unittest.TestCase):
    """Le test d'introspection : un montant-chaîne casse les sommes Excel."""

    def setUp(self):
        self.classeur, self.lignes = classeur()
        self.feuille = self.classeur.active

    def test_les_cellules_de_montant_ne_sont_jamais_des_chaines(self):
        fautes = []
        for rangee in self.feuille.iter_rows():
            for cellule in rangee:
                if cellule.value is None:
                    continue
                if cellule.number_format != xlsx.FORMAT_MONTANT:
                    continue
                if cellule.data_type not in ('n', 'f'):
                    fautes.append('%s = %r (type %s)'
                                  % (cellule.coordinate, cellule.value,
                                     cellule.data_type))
        self.assertEqual(fautes, [], '\n'.join(fautes))

    def test_aucune_cellule_ne_contient_un_montant_formate_en_texte(self):
        """Ni « 4 999 920,00 DH » ni « 2 950,00 DH » en valeur de cellule."""
        motif = re.compile(r'^\s*[\d  .,]+\s*(DH|MAD|dirhams)\s*$',
                           re.IGNORECASE)
        fautes = [c.coordinate for c in cellules(self.feuille)
                  if isinstance(c.value, str) and motif.match(c.value)]
        self.assertEqual(fautes, [])

    def test_les_prix_unitaires_sont_des_nombres(self):
        prix = [c.value for c in cellules(self.feuille)
                if c.column_letter == 'E' and c.data_type == 'n']
        self.assertIn(2950.0, prix)
        self.assertIn(78000.0, prix)

    def test_les_quantites_sont_des_nombres(self):
        quantites = [c.value for c in cellules(self.feuille)
                     if c.column_letter == 'D' and c.data_type == 'n']
        self.assertIn(152.0, quantites)
        self.assertIn(288.0, quantites)

    def test_seul_l_arrete_en_lettres_est_une_chaine_de_montant(self):
        chaines = [c.value for c in cellules(self.feuille)
                   if isinstance(c.value, str) and 'DIRHAMS' in c.value]
        self.assertTrue(any(ARRETE_TTC in valeur for valeur in chaines))


class TestSommesExcelJustes(unittest.TestCase):
    """Les totaux sont des FORMULES — et elles somment les bonnes cellules."""

    def setUp(self):
        self.classeur, self.lignes = classeur()
        self.feuille = self.classeur.active
        self.valeurs = xlsx.valeurs_de_controle(self.lignes)

    def valeur_numerique(self, coordonnee):
        """Évalue une cellule : nombre, `=SUM(a:b)`, `=A*B` ou `=a+b`."""
        cellule = self.feuille[coordonnee]
        valeur = cellule.value
        if valeur is None:
            return None
        if not isinstance(valeur, str):
            return Decimal(str(valeur))
        formule = valeur.lstrip('=')
        somme = re.fullmatch(r'SUM\((\w+?)(\d+):(\w+?)(\d+)\)', formule)
        if somme:
            colonne, debut, fin = somme.group(1), int(somme.group(2)), \
                int(somme.group(4))
            total = Decimal('0')
            for rang in range(debut, fin + 1):
                partiel = self.valeur_numerique('%s%d' % (colonne, rang))
                if partiel is not None:
                    total += partiel
            return total
        produit = re.fullmatch(r'(\w+\d+)\*(\w+\d+)', formule)
        if produit:
            return (self.valeur_numerique(produit.group(1))
                    * self.valeur_numerique(produit.group(2)))
        facteur = re.fullmatch(r'(\w+\d+)\*([\d.]+)', formule)
        if facteur:
            return (self.valeur_numerique(facteur.group(1))
                    * Decimal(facteur.group(2)))
        addition = re.fullmatch(r'(\w+\d+)\+(\w+\d+)', formule)
        if addition:
            return (self.valeur_numerique(addition.group(1))
                    + self.valeur_numerique(addition.group(2)))
        chaine = re.fullmatch(r'(\w+\d+(?:\+\w+\d+)+)', formule)
        if chaine:
            return sum((self.valeur_numerique(ref)
                        for ref in formule.split('+')), Decimal('0'))
        raise AssertionError('formule non évaluée : %r' % valeur)

    def cellule_de(self, libelle):
        """La cellule de MONTANT de la ligne portant ce libellé.

        La recherche est restreinte à la colonne des libellés : « Total HT »
        est aussi un en-tête de colonne, et le chercher partout renverrait
        l'en-tête au lieu de la ligne de total.
        """
        for rangee in self.feuille.iter_rows(min_col=2, max_col=2):
            for cellule in rangee:
                if cellule.value == libelle:
                    return '%s%d' % (COL_TOTAL, cellule.row)
        raise AssertionError('libellé introuvable : %s' % libelle)

    def test_les_totaux_de_ligne_sont_des_formules(self):
        formules = [c.value for c in cellules(self.feuille)
                    if c.column_letter == COL_TOTAL
                    and isinstance(c.value, str) and c.value.startswith('=')]
        self.assertTrue(any(f.startswith('=D') for f in formules))

    def test_les_sous_totaux_de_section_sont_justes(self):
        for section, attendu in SOUS_TOTAUX_ATTENDUS.items():
            coordonnee = self.cellule_de('Sous-total %s' % section)
            self.assertEqual(self.valeur_numerique(coordonnee), attendu,
                             section)

    def test_le_total_ht_est_juste(self):
        self.assertEqual(self.valeur_numerique(self.cellule_de('Total HT')),
                         TOTAL_HT)

    def test_la_tva_est_juste(self):
        coordonnee = self.cellule_de('TVA (20 %)')
        self.assertEqual(self.valeur_numerique(coordonnee), TVA)

    def test_le_total_ttc_est_juste(self):
        self.assertEqual(self.valeur_numerique(self.cellule_de('Total TTC')),
                         TOTAL_TTC)

    def test_les_valeurs_de_controle_python_concordent(self):
        self.assertEqual(self.valeurs['total_ht'], TOTAL_HT)
        self.assertEqual(self.valeurs['total_ttc'], TOTAL_TTC)


class TestClauseEtLettres(unittest.TestCase):

    def setUp(self):
        self.classeur, _ = classeur()
        self.feuille = self.classeur.active

    def test_la_clause_est_en_pied(self):
        from apps.ao.fabrique.clauses import CLAUSE_RESERVE_QUANTITES
        valeurs = [c.value for c in cellules(self.feuille)
                   if isinstance(c.value, str)]
        self.assertIn(CLAUSE_RESERVE_QUANTITES, valeurs)

    def test_la_colonne_des_pu_en_lettres_est_remplie(self):
        lettres = [c.value for c in cellules(self.feuille)
                   if c.column_letter == 'F' and isinstance(c.value, str)]
        self.assertIn('DEUX MILLE NEUF CENT CINQUANTE DIRHAMS', lettres)


class TestSeuilDeJob(unittest.TestCase):

    def test_seuil(self):
        self.assertFalse(xlsx.doit_passer_en_job(xlsx.SEUIL_JOB))
        self.assertTrue(xlsx.doit_passer_en_job(xlsx.SEUIL_JOB + 1))
        self.assertFalse(xlsx.doit_passer_en_job(0))
        self.assertFalse(xlsx.doit_passer_en_job(None))


class TestAucuneFuiteDeCout(unittest.TestCase):

    def test_le_classeur_ne_porte_aucun_champ_de_cout(self):
        cls, _ = classeur()
        textes = ' '.join(str(c.value) for c in cellules(cls.active)).lower()
        for interdit in ('prix_achat', "prix d'achat", 'coût de revient',
                         'marge', 'bénéfice'):
            self.assertNotIn(interdit, textes, interdit)


if __name__ == '__main__':
    unittest.main()
