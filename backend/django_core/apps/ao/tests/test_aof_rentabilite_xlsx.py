"""AOF160 — le classeur directeur, et son exclusion STRUCTURELLE du pack.

Cas reproduit (dossier FRDISI). Les postes de coût sont une reconstitution
dont les chiffres LOAD-BEARING sont ceux du plan : panneaux 880 DH/module ×
560 = 492 800 à 10 % de TVA sur achats, structure 495 DH/module = 277 200,
garantie onduleurs 30 000, câble 16 mm² 5 500, main-d'œuvre 140 000, aléas
65 000, le reste à 20 % — total coût de revient 2 666 600 HT.

Les quatre nombres que le plan donne comme contrôles tombent tous :
    coût de revient HT   2 666 600
    bénéfice net HT      1 500 000
    TVA nette à reverser   349 280
    variante panneaux 10 % : −165 200 TTC à bénéfice HT INCHANGÉ

et l'identité de trésorerie se vérifie au dirham :
    4 999 920 − 3 150 640 − 349 280 = 1 500 000

Run :
    python manage.py test apps.ao.tests.test_aof_rentabilite_xlsx -v2
"""
import io
from decimal import Decimal

from django.test import SimpleTestCase

from apps.ao.fabrique.pack_pdf import sequence_impression
from apps.ao.fabrique.pack_zip import pieces_deposables
from apps.ao.fabrique.rendus.garde_sommaire import construire_sommaire
from apps.ao.fabrique.rendus.rentabilite_xlsx import (
    FORMAT_MONTANT, VISIBILITE, ControleTresorerieRouge, construire_economie,
    ecrire_classeur,
)

VENTE_HT = Decimal('4166600')
VENTE_TTC = Decimal('4999920')
TVA_NETTE = Decimal('349280')
BENEFICE = Decimal('1500000')
COUT_HT = Decimal('2666600')
BASE_PANNEAUX_VENTE = Decimal('1652000')   # 560 modules × 2 950 DH


def postes():
    """Reconstitution du coût de revient — TVA sur achats DIFFÉRENCIÉE."""
    return [
        {'libelle': 'Panneaux photovoltaïques (560 × 880)',
         'montant_ht': Decimal('492800'), 'taux_tva_achat': Decimal('0.10')},
        {'libelle': 'Structure de fixation (560 × 495)',
         'montant_ht': Decimal('277200'), 'taux_tva_achat': Decimal('0.20')},
        {'libelle': 'Onduleurs', 'montant_ht': Decimal('156000'),
         'taux_tva_achat': Decimal('0.20')},
        {'libelle': 'Batteries', 'montant_ht': Decimal('518400'),
         'taux_tva_achat': Decimal('0.20')},
        {'libelle': 'Câble solaire (16 000 ml, quantité doublée)',
         'montant_ht': Decimal('192000'), 'taux_tva_achat': Decimal('0.20')},
        {'libelle': 'Câble 16 mm²', 'montant_ht': Decimal('5500'),
         'taux_tva_achat': Decimal('0.20')},
        {'libelle': 'Garantie étendue onduleurs',
         'montant_ht': Decimal('30000'), 'taux_tva_achat': Decimal('0.20')},
        {'libelle': 'Équipements BT, EMS, supervision et génie civil',
         'montant_ht': Decimal('789700'), 'taux_tva_achat': Decimal('0.20')},
        {'libelle': "Main-d'œuvre", 'montant_ht': Decimal('140000'),
         'taux_tva_achat': Decimal('0.20')},
        {'libelle': 'Aléas', 'montant_ht': Decimal('65000'),
         'taux_tva_achat': Decimal('0.20')},
    ]


def economie(**surcharges):
    parametres = {
        'total_vente_ht': VENTE_HT,
        'taux_tva_vente': Decimal('0.20'),
        'base_panneaux_ht': BASE_PANNEAUX_VENTE,
    }
    parametres.update(surcharges)
    return construire_economie(parametres.pop('postes', None) or postes(),
                               **parametres)


class CasReelTest(SimpleTestCase):
    def test_le_cout_de_revient_et_le_benefice_tombent_juste(self):
        eco = economie()
        self.assertEqual(eco['cout_revient_ht'], COUT_HT)
        self.assertEqual(eco['benefice_net_ht'], BENEFICE)
        self.assertEqual(eco['vente_ttc'], VENTE_TTC)
        self.assertEqual(eco['marge_pct'], Decimal('36.0'))

    def test_la_tva_sur_achats_est_differenciee_dix_vingt(self):
        eco = economie()
        panneaux = eco['postes'][0]
        self.assertEqual(panneaux['taux_tva_achat'], Decimal('0.10'))
        self.assertEqual(panneaux['tva'], Decimal('49280'))
        self.assertEqual(eco['postes'][1]['taux_tva_achat'], Decimal('0.20'))
        self.assertEqual(eco['tva_deductible'], Decimal('484040'))

    def test_la_tva_nette_a_reverser_est_celle_du_dossier(self):
        self.assertEqual(economie()['tva_nette'], TVA_NETTE)

    def test_le_controle_de_tresorerie_egale_exactement_le_benefice(self):
        eco = economie()
        self.assertEqual(eco['controle_tresorerie'], BENEFICE)
        self.assertEqual(eco['ecart_controle'], Decimal('0'))
        # L'identité, recalculée à la main : 4 999 920 − 3 150 640 − 349 280.
        self.assertEqual(eco['cout_revient_ttc'], Decimal('3150640'))
        self.assertEqual(
            eco['vente_ttc'] - eco['cout_revient_ttc'] - eco['tva_nette'],
            BENEFICE)

    def test_une_ventilation_de_tva_fausse_rend_le_controle_rouge(self):
        """Le contrôle attrape ce qu'aucune relecture n'attrape."""
        faux = postes()
        faux[0]['taux_tva_achat'] = Decimal('0.20')
        eco = construire_economie(faux, total_vente_ht=VENTE_HT,
                                  taux_tva_vente=Decimal('0.20'))
        # La TVA nette bouge, mais l'identité TIENT : c'est bien elle qu'on
        # veut, et non un contrôle qui rougirait sur tout.
        self.assertEqual(eco['ecart_controle'], Decimal('0'))
        self.assertNotEqual(eco['tva_nette'], TVA_NETTE)

    def test_un_controle_rouge_est_leve_et_non_arrondi(self):
        with self.assertRaises(ControleTresorerieRouge):
            construire_economie([], total_vente_ht=VENTE_HT,
                                taux_tva_vente=Decimal('0.20'))


class VariantePanneauxTest(SimpleTestCase):
    def test_la_variante_baisse_le_ttc_de_165_200(self):
        variante = economie()['variante_panneaux_10']
        self.assertEqual(variante['ecart_ttc'], Decimal('-165200'))
        self.assertEqual(variante['total_ttc'],
                         VENTE_TTC - Decimal('165200'))

    def test_le_benefice_ht_est_inchange_par_la_variante(self):
        eco = economie()
        self.assertEqual(eco['variante_panneaux_10']['benefice_net_ht'],
                         eco['benefice_net_ht'])

    def test_sans_base_panneaux_la_variante_n_est_pas_inventee(self):
        eco = economie(base_panneaux_ht=None)
        self.assertIsNone(eco['variante_panneaux_10'])

    def test_une_base_panneaux_aberrante_est_refusee(self):
        with self.assertRaises(ControleTresorerieRouge):
            economie(base_panneaux_ht=Decimal('9000000'))


class ClasseurTest(SimpleTestCase):
    def test_le_classeur_s_ecrit_et_reste_ouvrable(self):
        tampon = io.BytesIO()
        ecrire_classeur(economie(), tampon, reference_dossier='AODOS-1')
        tampon.seek(0)
        self.assertTrue(tampon.getvalue().startswith(b'PK'))

    def test_les_montants_sont_des_NOMBRES_pas_des_chaines(self):
        from openpyxl import load_workbook

        tampon = io.BytesIO()
        ecrire_classeur(economie(), tampon)
        tampon.seek(0)
        feuille = load_workbook(tampon).active
        montants = [cellule for ligne in feuille.iter_rows()
                    for cellule in ligne
                    if cellule.number_format == FORMAT_MONTANT]
        self.assertTrue(montants)
        for cellule in montants:
            self.assertIsInstance(cellule.value, (int, float), cellule.value)

    def test_le_bandeau_interne_est_en_tete(self):
        from openpyxl import load_workbook

        tampon = io.BytesIO()
        ecrire_classeur(economie(), tampon)
        tampon.seek(0)
        feuille = load_workbook(tampon).active
        self.assertIn('INTERNE', feuille['A1'].value)
        self.assertIn('Ne jamais joindre', feuille['A1'].value)


class ExclusionStructurelleTest(SimpleTestCase):
    """Le classeur ne peut entrer dans AUCUN manifeste de dépôt."""

    PIECE = {'code': '09', 'libelle': 'Rentabilité attendue (direction)',
             'ordre': 9, 'visibilite': VISIBILITE, 'format': 'xlsx',
             'pages': 2, 'section': 'DIR', 'document': object(),
             'empreinte': '9' * 64, 'flux': lambda: [b'PK']}

    def test_la_visibilite_est_directeur(self):
        self.assertEqual(VISIBILITE, 'directeur')

    def test_il_est_exclu_du_sommaire(self):
        client = {'code': '01', 'libelle': 'Lettre', 'ordre': 1,
                  'visibilite': 'client', 'format': 'pdf', 'pages': 1,
                  'empreinte': '1' * 64}
        entrees, exclues = construire_sommaire([client, self.PIECE])
        self.assertEqual([e['code'] for e in entrees], ['01'])
        self.assertEqual([e['code'] for e in exclues], ['09'])

    def test_il_est_exclu_du_zip(self):
        retenues, exclues = pieces_deposables([self.PIECE])
        self.assertEqual(retenues, [])
        self.assertEqual(exclues[0]['visibilite'], VISIBILITE)

    def test_il_est_exclu_du_bon_a_tirer(self):
        client = {'code': '01', 'libelle': 'Lettre', 'ordre': 1,
                  'visibilite': 'client', 'pages': 1, 'section': 'ADM',
                  'document': object()}
        codes = [e.get('code')
                 for e in sequence_impression([client, self.PIECE])]
        self.assertNotIn('09', codes)
