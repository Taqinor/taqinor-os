"""AOF161 — le TÉLÉCHARGEMENT du classeur directeur : job, artefact, garde.

Le front appelait ``/ao/<id>/rentabilite/telecharger/`` — une route que
personne n'avait enregistrée, tandis que le rendu (AOF160) et la tâche
``ao.produire_rentabilite_xlsx`` existaient déjà des deux côtés sans être
reliés. Pire, la tâche passait l'INSTANCE de modèle à un rendu qui attend un
dictionnaire : le classeur ne pouvait pas être produit, et l'écrivait dans un
tampon mémoire qu'elle jetait ensuite.

Ce qui est prouvé ici :

  1. l'action ``telecharger`` est routée sur la ressource RÉELLE
     (``/api/django/ao/economie/<id>/telecharger/``) ;
  2. elle est fermée à tout non-directeur (Responsable, Commercial), en GET
     comme en POST — la permission est ``ao_rentabilite_voir``, ÉLEVÉE ;
  3. le POST passe par ``core.jobs.submit`` (``BackgroundJob``) : AUCUN rendu
     synchrone dans la requête ;
  4. la tâche produit un VRAI classeur et DÉPOSE l'artefact (clé posée sur le
     job) au lieu de le jeter ;
  5. le classeur porte bien les nombres du cas réel — le traducteur
     modèle→rendu ne fabrique pas une économie différente ;
  6. le job d'une AUTRE société est INTROUVABLE (404), jamais « interdit » ;
  7. **étanchéité** : l'artefact n'est pas déposé en ``records.Attachment`` —
     la liste générique des pièces jointes est ouverte à tout rôle et servirait
     le coût de revient à un non-directeur.

Run :
    python manage.py test apps.ao.tests.test_aof_rentabilite_telechargement -v2
"""
import io
from decimal import Decimal
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.test import TestCase
from openpyxl import load_workbook
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import AccessToken

from apps.ao import services_directeur
from apps.ao.models import AppelOffre, EconomieAO, LigneCoutRevient
from apps.ao.tasks import produire_rentabilite_xlsx_task
from apps.ao.views_directeur import KIND_RENTABILITE_XLSX
from apps.roles.models import (
    COMMERCIAL_PERMISSIONS, DIRECTEUR_PERMISSIONS, RESPONSABLE_PERMISSIONS,
    Role,
)
from authentication.models import Company
from core.models import BackgroundJob

User = get_user_model()

#: Les postes RÉELS du dossier (mêmes chiffres que ``test_economie_directeur``
#: et ``test_aof_rentabilite_xlsx`` : coût de revient 2 666 600 HT, bénéfice
#: visé 1 500 000, donc 4 166 600 HT).
POSTES = (
    ('panneaux', 'Modules photovoltaïques', '560', '880.00', 'reduit'),
    ('structure', 'Structure de pose', '560', '495.00', 'standard'),
    ('garantie_onduleurs', 'Extension de garantie onduleurs', '1', '30000.00',
     'standard'),
    ('cable_solaire', 'Câble solaire (métré DOUBLÉ)', '16000', '11.00',
     'standard'),
    ('cable_ac', 'Câble 16 mm²', '1', '5500.00', 'standard'),
    ('main_oeuvre', "Main d'œuvre", '1', '140000.00', 'standard'),
    ('aleas', 'Aléas', '1', '65000.00', 'standard'),
    ('onduleurs', 'Onduleurs, stockage, coffrets et équipements', '1',
     '1480100.00', 'standard'),
)


def url_telecharger(economie):
    return f'/api/django/ao/economie/{economie.pk}/telecharger/'


class TacheImmediate:
    """Substitut de la tâche Celery : ``.delay()`` exécute tout de suite.

    Aucun broker n'est joint. Le CHEMIN testé reste le vrai — la vue appelle
    ``core.jobs.submit``, qui crée le ``BackgroundJob`` puis ``.delay`` ; seul
    le transport est court-circuité (même patron que ``test_api_calepinage``).
    """

    def __init__(self, resultat=None):
        self.appels = []
        self.resultat = resultat

    def delay(self, **kwargs):
        self.appels.append(kwargs)
        if self.resultat is not None:
            return self.resultat
        return produire_rentabilite_xlsx_task(**kwargs)


class BaseTelechargement(TestCase):
    def setUp(self):
        self.company = Company.objects.create(nom='AOF161 Co',
                                              slug='aof161-co')
        self.ao = AppelOffre.objects.create(
            company=self.company, reference='AO-161/2026',
            objet='Classeur directeur')
        self.economie = services_directeur.creer_economie(
            self.ao, benefice_net_cible_ht=Decimal('1500000.00'),
            motif='Cible initiale')
        for ordre, (poste, designation, quantite, pu, regime) in enumerate(
                POSTES):
            LigneCoutRevient.objects.create(
                company=self.company, economie=self.economie, poste=poste,
                designation=designation, quantite=Decimal(quantite),
                prix_unitaire_ht=Decimal(pu), regime_tva=regime, ordre=ordre)
        self.directeur = self._api(DIRECTEUR_PERMISSIONS, 'aof161_dir')

    def _api(self, permissions, username):
        role = Role.objects.create(
            company=self.company, nom=f'R-{username}',
            permissions=list(permissions))
        user = User.objects.create_user(
            username=username, password='x', company=self.company, role=role)
        api = APIClient()
        api.credentials(
            HTTP_AUTHORIZATION=f'Bearer {AccessToken.for_user(user)}')
        api.utilisateur = user
        return api


class TestGarde(BaseTelechargement):
    """La permission est ÉLEVÉE : personne d'autre que la direction."""

    def test_le_responsable_ne_peut_ni_lancer_ni_lire(self):
        api = self._api(RESPONSABLE_PERMISSIONS, 'aof161_resp')
        self.assertEqual(api.post(url_telecharger(self.economie)).status_code,
                         403)
        self.assertEqual(api.get(url_telecharger(self.economie)).status_code,
                         403)

    def test_le_commercial_ne_peut_ni_lancer_ni_lire(self):
        api = self._api(COMMERCIAL_PERMISSIONS, 'aof161_com')
        self.assertEqual(api.post(url_telecharger(self.economie)).status_code,
                         403)
        self.assertEqual(api.get(url_telecharger(self.economie)).status_code,
                         403)

    def test_l_economie_d_une_autre_societe_est_introuvable(self):
        autre = Company.objects.create(nom='Autre AOF161', slug='autre-161')
        ao = AppelOffre.objects.create(company=autre, reference='AO-X',
                                       objet='Ailleurs')
        economie = EconomieAO.objects.create(company=autre, appel_offre=ao)
        self.assertEqual(
            self.directeur.post(url_telecharger(economie)).status_code, 404)


class TestProductionParJob(BaseTelechargement):
    def test_le_post_lance_un_job_et_ne_rend_rien_synchroniquement(self):
        tache = TacheImmediate(resultat={'produit': True})
        with patch('apps.ao.tasks.produire_rentabilite_xlsx_task', tache):
            reponse = self.directeur.post(url_telecharger(self.economie))
        self.assertEqual(reponse.status_code, 202, reponse.data)
        job = BackgroundJob.objects.get(pk=reponse.data['job'])
        self.assertEqual(job.kind, KIND_RENTABILITE_XLSX)
        self.assertEqual(job.company_id, self.company.pk)
        self.assertFalse(reponse.data['pret'])
        # La tâche a bien reçu le PROJET, jamais un id d'économie deviné.
        self.assertEqual(tache.appels[0]['projet_id'], self.ao.pk)

    def test_un_broker_injoignable_donne_503_et_non_500(self):
        """Une file morte est une indisponibilité, pas une erreur du serveur."""
        class BrokerMort:
            def delay(self, **_kwargs):
                raise OSError('broker injoignable')

        with patch('apps.ao.tasks.produire_rentabilite_xlsx_task',
                   BrokerMort()):
            reponse = self.directeur.post(url_telecharger(self.economie))
        self.assertEqual(reponse.status_code, 503)

    def test_le_get_sans_job_est_un_404_explicite(self):
        reponse = self.directeur.get(url_telecharger(self.economie))
        self.assertEqual(reponse.status_code, 404)

    def test_le_job_d_une_autre_societe_est_introuvable(self):
        autre = Company.objects.create(nom='Tierce AOF161', slug='tierce-161')
        role = Role.objects.create(company=autre, nom='Dir',
                                   permissions=list(DIRECTEUR_PERMISSIONS))
        etranger = User.objects.create_user(
            username='aof161_ailleurs', password='x', company=autre,
            role=role)
        job = BackgroundJob.objects.create(
            company=autre, user=etranger, kind=KIND_RENTABILITE_XLSX)
        reponse = self.directeur.get(url_telecharger(self.economie),
                                     {'job': job.pk})
        self.assertEqual(reponse.status_code, 404)


class TestArtefactDepose(BaseTelechargement):
    """La tâche PRODUIT et DÉPOSE — elle ne jetait rien de moins qu'un fichier."""

    def _produire(self):
        """Lance la production en interceptant le seul appel au stockage."""
        depose = {}

        def faux_depot(data, *, company_id, job_id, ext='xlsx',
                       content_type=None):
            octets = data if isinstance(data, bytes) else data.read()
            cle = f'exports/{company_id}/{job_id}.{ext}'
            depose[cle] = octets
            return cle

        tache = TacheImmediate()
        with patch('apps.records.storage.store_export_result', faux_depot), \
                patch('apps.ao.tasks.produire_rentabilite_xlsx_task', tache):
            reponse = self.directeur.post(url_telecharger(self.economie))
        job = BackgroundJob.objects.get(pk=reponse.data['job'])
        return job, depose

    def test_le_job_se_termine_avec_la_cle_de_l_artefact(self):
        job, depose = self._produire()
        self.assertEqual(job.statut, BackgroundJob.STATUT_DONE)
        self.assertTrue(job.result_file_key, "aucun artefact déposé")
        self.assertIn(job.result_file_key, depose)
        # Clé PRÉFIXÉE PAR SOCIÉTÉ (isolation du stockage objet).
        self.assertTrue(
            job.result_file_key.startswith(f'exports/{self.company.pk}/'))

    def test_le_classeur_porte_les_nombres_du_cas_reel(self):
        """Le traducteur modèle→rendu ne fabrique pas une autre économie."""
        _job, depose = self._produire()
        octets = next(iter(depose.values()))
        feuille = load_workbook(io.BytesIO(octets)).active
        valeurs = {cellule.value for rangee in feuille.iter_rows()
                   for cellule in rangee if cellule.value is not None}
        for attendu in (2666600.0, 4166600.0, 4999920.0, 349280.0, 1500000.0):
            self.assertIn(attendu, valeurs, attendu)
        # L'écart de contrôle de trésorerie est NUL, ou le classeur est rouge.
        self.assertIn('Écart de contrôle (doit être nul)', valeurs)
        self.assertIn(0.0, valeurs)

    def test_le_directeur_retire_les_octets_par_l_endpoint_garde(self):
        job, depose = self._produire()
        octets = depose[job.result_file_key]
        with patch('apps.records.storage.fetch_attachment',
                   lambda cle: (depose.get(cle), None)):
            reponse = self.directeur.get(
                url_telecharger(self.economie),
                {'job': job.pk, 'fichier': 1})
        self.assertEqual(reponse.status_code, 200)
        self.assertEqual(reponse.content, octets)
        self.assertIn('spreadsheetml', reponse['Content-Type'])
        # Pièce INTERNE : téléchargement forcé, jamais un aperçu en ligne.
        self.assertIn('attachment;', reponse['Content-Disposition'])
        self.assertIn('AO-161-2026', reponse['Content-Disposition'])

    def test_le_classeur_d_un_AUTRE_dossier_n_est_pas_servi_sous_ce_nom(self):
        """La clé porte l'AO produit : un job d'un autre dossier est refusé.

        Ce n'est pas une fuite de permission (l'économie est company-wide pour
        un directeur) mais un fichier ÉTIQUETÉ FAUX — invisible à l'usage,
        donc pire.
        """
        job, depose = self._produire()
        autre_ao = AppelOffre.objects.create(
            company=self.company, reference='AO-161/AUTRE', objet='Autre')
        autre = EconomieAO.objects.create(company=self.company,
                                          appel_offre=autre_ao)
        with patch('apps.records.storage.fetch_attachment',
                   lambda cle: (depose.get(cle), None)):
            reponse = self.directeur.get(url_telecharger(autre),
                                         {'job': job.pk, 'fichier': 1})
        self.assertEqual(reponse.status_code, 404)

    def test_le_suivi_sans_fichier_reste_du_json(self):
        job, _depose = self._produire()
        reponse = self.directeur.get(url_telecharger(self.economie),
                                     {'job': job.pk})
        self.assertEqual(reponse.status_code, 200)
        self.assertTrue(reponse.data['pret'])
        self.assertEqual(reponse.data['statut'], BackgroundJob.STATUT_DONE)

    def test_aucune_piece_jointe_generique_n_est_creee(self):
        """ÉTANCHÉITÉ : ``/records/attachments/` est ouvert à TOUT rôle.

        Déposer le coût de revient en ``records.Attachment`` le servirait à un
        non-directeur (liste ``IsAnyRole`` + action ``download``). L'artefact
        passe donc par le livrable de job, jamais par la pièce jointe
        générique.
        """
        from apps.records.models import Attachment

        avant = Attachment.objects.count()
        self._produire()
        self.assertEqual(Attachment.objects.count(), avant)


class TestEconomieAbsente(BaseTelechargement):
    def test_un_ao_sans_economie_echoue_avec_un_motif_francais(self):
        ao = AppelOffre.objects.create(
            company=self.company, reference='AO-161/VIDE', objet='Sans coût')
        job = BackgroundJob.objects.create(
            company=self.company, user=self.directeur.utilisateur,
            kind=KIND_RENTABILITE_XLSX)
        resultat = produire_rentabilite_xlsx_task(
            job_id=job.pk, company_id=self.company.pk, projet_id=ao.pk)
        job.refresh_from_db()
        self.assertFalse(resultat['produit'])
        self.assertEqual(job.statut, BackgroundJob.STATUT_FAILED)
        self.assertIn('économie', job.message_erreur)

    def test_une_economie_sans_poste_echoue_sans_produire_de_classeur_faux(self):
        ao = AppelOffre.objects.create(
            company=self.company, reference='AO-161/NUL', objet='Aucun poste')
        EconomieAO.objects.create(company=self.company, appel_offre=ao)
        job = BackgroundJob.objects.create(
            company=self.company, user=self.directeur.utilisateur,
            kind=KIND_RENTABILITE_XLSX)
        resultat = produire_rentabilite_xlsx_task(
            job_id=job.pk, company_id=self.company.pk, projet_id=ao.pk)
        job.refresh_from_db()
        self.assertFalse(resultat['produit'])
        self.assertEqual(job.statut, BackgroundJob.STATUT_FAILED)
