"""AOF135 — simulation de rentabilité 25 ans : PIÈCE CLIENT, sans AUCUN coût.

Ce qui est prouvé ici :

* **aucun champ de coût de revient ni de marge** sur le modèle ni dans les
  sorties (ratchet local ; AOF129 l'étend à tous les artefacts) ;
* les valeurs sont DÉRIVÉES du bordereau (CAPEX) et du calepinage (puissance),
  jamais saisies deux fois ;
* le cas réel est reproduit : CAPEX hors stockage 3 449 000, montant remis
  4 999 920 TTC → retour simple 5,40 ans, retour sur le TTC 7,83 ans ;
* ``source_hash`` est porté et bouge dès qu'une entrée bouge ;
* le classeur XLSX porte des FORMULES vivantes, pas des valeurs mortes.

Run :
    python manage.py test apps.ao.tests.test_aof_simulation -v2
"""
from decimal import Decimal

from django.test import SimpleTestCase, TestCase

from apps.ao.fabrique.rendus import simulation as rendu_simulation
from apps.ao.models import (
    AppelOffre, BatimentAO, BordereauPrix, LigneBordereau,
    SimulationRentabilite, ToitureAO, VarianteCalepinage,
)
from authentication.models import Company

CLAUSE = 'Marché à prix unitaires — quantités prévisionnelles.'


class TestAucunCout(SimpleTestCase):
    def test_aucun_champ_de_cout_sur_la_simulation(self):
        noms = {f.name for f in SimulationRentabilite._meta.get_fields()}
        for interdit in ('cout_revient', 'cout_de_revient', 'marge',
                         'marge_pct', 'benefice', 'benefice_net',
                         'prix_achat'):
            self.assertNotIn(interdit, noms)

    def test_le_modele_est_distinct_de_l_economie_directeur(self):
        """AOF157 vit dans SES tables ; les deux ne se confondent jamais."""
        self.assertEqual(SimulationRentabilite._meta.db_table,
                         'ao_simulation_rentabilite')


class BaseSimulation(TestCase):
    """Le cas réel, reconstitué à partir de ses seules ENTRÉES."""

    def setUp(self):
        self.company = Company.objects.create(nom='AOF135 Co',
                                              slug='aof135-co')
        self.ao = AppelOffre.objects.create(
            company=self.company, reference='AO-135-1',
            objet='Centrale photovoltaïque', soumissionnaire='PARTENAIRE SA')
        batiment = BatimentAO.objects.create(
            company=self.company, appel_offre=self.ao, code='C')
        toiture = ToitureAO.objects.create(
            company=self.company, batiment=batiment, code_document='05H')
        VarianteCalepinage.objects.create(
            company=self.company, toiture=toiture, appel_offre=self.ao,
            nom='Retenue', est_retenue=True,
            resultat={'total_modules': 560, 'kwc': '350.000'})
        self.bordereau = BordereauPrix.objects.create(
            company=self.company, appel_offre=self.ao, clause_reserve=CLAUSE,
            taux_tva_defaut=Decimal('20.00'))
        # Bordereau réel : total TTC 4 999 920, dont stockage 1 550 920 TTC
        # → CAPEX hors stockage 3 449 000 TTC.
        LigneBordereau.objects.create(
            company=self.company, bordereau=self.bordereau, numero=1,
            designation='Fourniture et pose', quantite=Decimal('1'),
            prix_unitaire=Decimal('2874166.6667'))
        LigneBordereau.objects.create(
            company=self.company, bordereau=self.bordereau, numero=2,
            designation='Stockage', quantite=Decimal('1'),
            prix_unitaire=Decimal('1292433.3333'), est_stockage=True)
        self.simulation = SimulationRentabilite.objects.create(
            company=self.company, appel_offre=self.ao,
            bordereau=self.bordereau, duree_annees=25,
            productible_kwh_par_kwc_an=Decimal('1600.00'),
            productible_source='Étude PVGIS du site',
            tarif_kwh=Decimal('1.1405'),
            part_autoconsommee_pct=Decimal('100.00'),
            degradation_annuelle_pct=Decimal('0.00'),
            inflation_tarif_pct=Decimal('0.00'),
            taux_actualisation_pct=Decimal('0.00'))


class TestGrandeursDerivees(BaseSimulation):
    def test_la_puissance_vient_du_calepinage(self):
        self.assertEqual(self.simulation.puissance_kwc, Decimal('350.000'))

    def test_le_capex_vient_du_bordereau(self):
        self.assertEqual(self.bordereau.total_ttc, Decimal('4999920.00'))
        self.assertEqual(self.simulation.capex_total, Decimal('4999920.00'))

    def test_le_capex_hors_stockage_exclut_les_lignes_de_stockage(self):
        self.assertEqual(self.simulation.capex_hors_stockage,
                         Decimal('3449000.00'))

    def test_le_productible_est_derive_de_la_puissance(self):
        self.assertEqual(self.simulation.productible_kwh_an,
                         Decimal('560000.000'))

    def test_l_economie_annuelle_est_derivee(self):
        self.assertEqual(self.simulation.economie_annuelle_initiale,
                         Decimal('638680.00'))


class TestCasReel(BaseSimulation):
    """Les deux retours du dossier réel, au centième d'année."""

    def test_retour_simple_sur_le_capex_hors_stockage(self):
        self.assertEqual(self.simulation.payback_simple_ans,
                         Decimal('5.40'))

    def test_retour_sur_le_montant_ttc_remis(self):
        self.assertEqual(self.simulation.roi_sur_ttc_ans, Decimal('7.83'))

    def test_le_tableau_couvre_la_duree_declaree(self):
        tableau = self.simulation.tableau_annuel
        self.assertEqual(len(tableau), 25)
        self.assertEqual(tableau[0]['annee'], 1)
        self.assertEqual(tableau[-1]['annee'], 25)

    def test_economies_cumulees(self):
        self.assertEqual(self.simulation.economies_cumulees,
                         Decimal('638680.00') * 25)

    def test_la_degradation_fait_baisser_le_productible(self):
        self.simulation.degradation_annuelle_pct = Decimal('0.50')
        tableau = self.simulation.tableau_annuel
        self.assertGreater(tableau[0]['productible_kwh'],
                           tableau[-1]['productible_kwh'])

    def test_l_actualisation_allonge_le_retour(self):
        simple = self.simulation.payback_simple_ans
        self.simulation.taux_actualisation_pct = Decimal('5.00')
        self.assertGreater(self.simulation.payback_actualise_ans, simple)


class TestEmpreinte(BaseSimulation):
    def test_l_empreinte_est_posee_et_stable(self):
        empreinte, a_change = rendu_simulation.rafraichir_empreinte(
            self.simulation)
        self.assertEqual(len(empreinte), 64)
        self.assertTrue(a_change)
        _, a_change_2 = rendu_simulation.rafraichir_empreinte(self.simulation)
        self.assertFalse(a_change_2)

    def test_un_changement_de_tarif_change_l_empreinte(self):
        avant = rendu_simulation.empreinte_simulation(self.simulation)
        self.simulation.tarif_kwh = Decimal('1.2000')
        self.assertNotEqual(
            rendu_simulation.empreinte_simulation(self.simulation), avant)

    def test_un_changement_de_bordereau_change_l_empreinte(self):
        avant = rendu_simulation.empreinte_simulation(self.simulation)
        LigneBordereau.objects.create(
            company=self.company, bordereau=self.bordereau, numero=3,
            designation='Ajout', quantite=Decimal('1'),
            prix_unitaire=Decimal('1000.00'))
        self.assertNotEqual(
            rendu_simulation.empreinte_simulation(self.simulation), avant)


class TestSorties(BaseSimulation):
    def test_le_contexte_ne_porte_aucun_cout(self):
        contexte = rendu_simulation.contexte_simulation(self.simulation)
        serialise = str(contexte).lower()
        for mot in rendu_simulation.MOTS_INTERDITS_CLIENT:
            self.assertNotIn(mot, serialise, mot)

    def test_l_html_client_est_propre(self):
        html = rendu_simulation.html_simulation(self.simulation)
        self.assertEqual(
            rendu_simulation.controler_absence_de_cout(html), [])
        self.assertIn('4999920.00', html)
        self.assertIn('Étude PVGIS du site', html)

    def test_le_classeur_porte_des_formules_vivantes(self):
        from io import BytesIO

        from openpyxl import load_workbook

        octets = rendu_simulation.classeur_xlsx(self.simulation)
        classeur = load_workbook(BytesIO(octets))
        feuille = classeur['Rentabilité']
        # Ligne 13 = première année du tableau (en-tête ligne 12).
        self.assertTrue(str(feuille['B13'].value).startswith('='))
        self.assertTrue(str(feuille['D13'].value).startswith('='))
        self.assertIn('$B$5', str(feuille['C13'].value))
        self.assertEqual(feuille['A13'].value, 1)

    def test_le_classeur_couvre_toutes_les_annees(self):
        from io import BytesIO

        from openpyxl import load_workbook

        classeur = load_workbook(
            BytesIO(rendu_simulation.classeur_xlsx(self.simulation)))
        feuille = classeur['Rentabilité']
        self.assertEqual(feuille[f'A{12 + 25}'].value, 25)

    def test_le_ratchet_attrape_un_mot_de_cout(self):
        self.assertEqual(
            rendu_simulation.controler_absence_de_cout(
                'Notre marge est de 36 %.'),
            ['marge'])
