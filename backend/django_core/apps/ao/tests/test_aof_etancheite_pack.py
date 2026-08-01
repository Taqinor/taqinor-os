"""AOF129 — RATCHET d'étanchéité du pack client. Écrit AVANT les pièces.

**Pourquoi ici et pas plus tard.** Ce test est placé délibérément avant les
écrans et avant l'économie du dossier : un ratchet écrit après la dernière
pièce ne protège rien. Une simple inclusion de serializer ou un
`select_related` bavard suffit à faire sortir une marge dans un dossier remis
au maître d'ouvrage — et c'est un défaut IRRATTRAPABLE une fois le pli déposé.

**Comment il se maintient.** `ARTEFACTS_COUVERTS` est la liste, publiée ici
même, de tout ce que le ratchet ouvre. **Chaque tâche qui livre un nouvel
artefact client l'ajoute à cette liste DANS LE MÊME COMMIT** (critère de Done
repris par AOF131-135, 138-139, 151-152, 160). Un artefact absent de la liste
n'est pas protégé, et c'est la liste qui rend l'oubli visible en revue.

Calqué sur le test d'exclusion du moteur de devis
(`apps/ventes/tests/test_quote_engine.py`) et sur le ratchet
`core/tests/test_sensitive_field_leak.py`.

    python -m unittest apps.ao.tests.test_aof_etancheite_pack -v
"""
import html as _html
import io
import re
import unittest
from decimal import Decimal

from openpyxl import load_workbook

from apps.ao.fabrique import approvisionnement, bibliotheque_prix
from apps.ao.fabrique.contexte import construire_contexte
from apps.ao.fabrique.rendus import (acte_engagement, bordereau_pdf,
                                     bordereau_xlsx, lettre)
from apps.ao.tests.aof_fixtures import (bordereau_depose, contexte_dossier,
                                        rendre_gabarit)

# ---------------------------------------------------------------------------
# Ce que le ratchet interdit. Chaque motif vise le sens ÉCONOMIQUE du mot : le
# calepinage publie légitimement des « marges de robustesse » en centimètres,
# elles n'ont rien à voir avec une marge commerciale et ne doivent pas faire
# rougir le test (un ratchet qui crie à tort finit désactivé).
# ---------------------------------------------------------------------------
MOTIFS_INTERDITS = (
    ('prix_achat', re.compile(r"prix[_ ]d?['’ ]?achat", re.IGNORECASE)),
    ('cout_revient', re.compile(r"co[uû]ts?[_ ]de[_ ]revient|cout_revient",
                                re.IGNORECASE)),
    ('taux_marge', re.compile(
        r"taux[_ ]de[_ ]marge|taux_marge|marge[_ ](?:brute|nette|"
        r"commerciale|pct|pourcent)|marge_|%\s*de\s*marge", re.IGNORECASE)),
    ('benefice', re.compile(r"b[ée]n[ée]fice", re.IGNORECASE)),
    ('coefficient_commercial', re.compile(
        r"coefficient[_ ]?(?:de[_ ])?(?:vente|marge|majoration|prix)",
        re.IGNORECASE)),
    ('maximum_agrege_site', re.compile(
        r"maximum[_ ](?:agr[ée]g[ée]|du[_ ]site|site)|"
        r"potentiel[_ ]maximal[_ ]du[_ ]site|"
        r"capacit[ée][_ ]maximale[_ ]du[_ ]site", re.IGNORECASE)),
)

#: Noms du bureau d'études en marque blanche : ils ne doivent JAMAIS apparaître
#: sur une pièce que NOUS remettons. La liste est volontairement vide tant que
#: le fondateur ne l'a pas arrêtée (voir NOTES_LANE_AO_FABRIQUE.md) — le
#: mécanisme, lui, est testé ci-dessous avec un nom injecté, de sorte qu'il
#: suffira de remplir cette constante pour qu'il morde.
NOMS_MARQUE_BLANCHE = ()


# ---------------------------------------------------------------------- outils
def texte_visible(html):
    """Le texte LU par la commission : sans style, sans script, sans balise."""
    texte = re.sub(r'<style\b.*?</style>|<script\b.*?</script>', ' ', html,
                   flags=re.DOTALL | re.IGNORECASE)
    texte = re.sub(r'\{#.*?#\}|<!--.*?-->', ' ', texte, flags=re.DOTALL)
    texte = re.sub(r'<[^>]*>', ' ', texte)
    return _html.unescape(re.sub(r'\s+', ' ', texte))


def texte_du_classeur(octets):
    """Toutes les valeurs de cellule d'un classeur, concaténées."""
    classeur = load_workbook(io.BytesIO(octets))
    morceaux = []
    for feuille in classeur.worksheets:
        morceaux.append(feuille.title)
        for rangee in feuille.iter_rows():
            for cellule in rangee:
                if cellule.value is not None:
                    morceaux.append(str(cellule.value))
    return ' '.join(morceaux)


def texte_de_charge(objet, _profondeur=0):
    """Texte d'une charge utile d'API — CLÉS COMPRISES.

    Une fuite de serializer se voit d'abord dans le NOM du champ : scanner les
    seules valeurs laisserait passer `{"prix_achat": 2100}` si la valeur est
    nulle.
    """
    if _profondeur > 12:
        return ''
    if objet is None or isinstance(objet, (bool, int, float, Decimal)):
        return str(objet)
    if isinstance(objet, str):
        return objet
    if hasattr(objet, 'vers_dict'):
        return texte_de_charge(objet.vers_dict(), _profondeur + 1)
    if hasattr(objet, 'items'):
        return ' '.join(
            '%s %s' % (cle, texte_de_charge(val, _profondeur + 1))
            for cle, val in objet.items())
    if isinstance(objet, (list, tuple, set, frozenset)):
        return ' '.join(texte_de_charge(item, _profondeur + 1)
                        for item in objet)
    return str(objet)


def fuites(texte, *, noms_marque_blanche=NOMS_MARQUE_BLANCHE):
    """Les motifs interdits trouvés dans un texte. Vide = étanche."""
    trouvees = []
    for code, motif in MOTIFS_INTERDITS:
        occurrence = motif.search(texte)
        if occurrence:
            trouvees.append('%s : « %s »' % (code, occurrence.group(0)))
    for nom in noms_marque_blanche:
        if re.search(re.escape(nom), texte, re.IGNORECASE):
            trouvees.append('marque_blanche : « %s »' % nom)
    return tuple(trouvees)


def scanner(artefact, *, noms_marque_blanche=NOMS_MARQUE_BLANCHE):
    """Point d'entrée réutilisable : accepte HTML, octets XLSX ou charge API."""
    if isinstance(artefact, bytes):
        texte = texte_du_classeur(artefact)
    elif isinstance(artefact, str):
        texte = texte_visible(artefact) if '<' in artefact else artefact
    else:
        texte = texte_de_charge(artefact)
    return fuites(texte, noms_marque_blanche=noms_marque_blanche)


# ------------------------------------------------------- artefacts couverts
def _lignes_et_contexte():
    lignes = bordereau_depose()
    return lignes, construire_contexte(contexte_dossier(lignes))


def artefact_bordereau_pdf():
    lignes, contexte = _lignes_et_contexte()
    donnees = bordereau_pdf.contexte_gabarit(lignes, contexte)
    return rendre_gabarit(bordereau_pdf.NOM_GABARIT, donnees)


def artefact_bordereau_xlsx():
    lignes, contexte = _lignes_et_contexte()
    return bordereau_xlsx.rendre(lignes, contexte)


def artefact_lettre_soumission():
    lignes, contexte = _lignes_et_contexte()
    return rendre_gabarit(lettre.NOM_GABARIT,
                          lettre.contexte_gabarit(lignes, contexte))


def artefact_acte_engagement():
    lignes, contexte = _lignes_et_contexte()
    return rendre_gabarit(acte_engagement.NOM_GABARIT,
                          acte_engagement.contexte_gabarit(lignes, contexte))


def artefact_acte_fiche_de_report():
    lignes, contexte = _lignes_et_contexte()
    vue = acte_engagement.contexte_gabarit(
        lignes, contexte,
        modele_acheteur={'reference': 'DCE-03',
                         'libelle': "Acte d'engagement (E3)"})
    return rendre_gabarit(acte_engagement.NOM_GABARIT, vue)


def artefact_contexte_dossier():
    _, contexte = _lignes_et_contexte()
    return dict(contexte)


def artefact_rapport_approvisionnement():
    catalogue = {'MOD-625': {'existe': True, 'archive': False,
                             'prix_renseigne': True, 'disponible': 600,
                             'deja_approvisionne': True,
                             'designation': 'Module 625 Wc'}}
    equipements = [{'role': 'module', 'reference': 'MOD-625',
                    'quantite': 560, 'designation': 'Module 625 Wc'}]
    return approvisionnement.controler(equipements, catalogue).vers_dict()


def artefact_proposition_de_prix():
    historique = [{'reference': 'MOD-625', 'famille': 'modules',
                   'prix_unitaire': '2950', 'date': '2026-07-27',
                   'dossier': 'AO-202607-0002', 'unite': 'U'}]
    return {
        'proposition': bibliotheque_prix.proposer(
            historique, reference='MOD-625').vers_dict(),
        'fourchettes': {famille: bande.vers_dict() for famille, bande
                        in bibliotheque_prix.fourchettes(historique).items()},
    }


#: LA LISTE. Tout artefact remis au maître d'ouvrage — ou lisible par un
#: profil non-directeur — figure ici. Une tâche qui livre une pièce SANS
#: l'ajouter laisse un trou ; c'est cette liste qu'on relit en revue.
ARTEFACTS_COUVERTS = (
    ('bordereau des prix — PDF (HTML rendu)', artefact_bordereau_pdf),
    ('bordereau des prix — classeur XLSX', artefact_bordereau_xlsx),
    ('lettre de soumission — PDF (HTML rendu)', artefact_lettre_soumission),
    ("acte d'engagement — PDF autonome", artefact_acte_engagement),
    ("acte d'engagement — fiche de report des valeurs",
     artefact_acte_fiche_de_report),
    ('contexte de dossier — charge utile API non-directeur',
     artefact_contexte_dossier),
    ("contrôle d'approvisionnement — charge utile API",
     artefact_rapport_approvisionnement),
    ('bibliothèque de prix — charge utile API',
     artefact_proposition_de_prix),
)


class TestRatchetVert(unittest.TestCase):
    """Aucun artefact client existant ne laisse fuir une donnée d'économie."""

    def test_tous_les_artefacts_couverts_sont_etanches(self):
        fautes = []
        for libelle, produire in ARTEFACTS_COUVERTS:
            for fuite in scanner(produire()):
                fautes.append('%s → %s' % (libelle, fuite))
        self.assertEqual(fautes, [], '\n'.join(fautes))

    def test_la_liste_des_artefacts_est_publiee_et_non_vide(self):
        self.assertGreaterEqual(len(ARTEFACTS_COUVERTS), 5)
        for libelle, produire in ARTEFACTS_COUVERTS:
            self.assertTrue(libelle.strip())
            self.assertTrue(produire(), libelle)

    def test_le_pdf_porte_bien_les_montants_clients(self):
        """Contre-épreuve : le ratchet n'est pas vert parce que la pièce est vide."""
        texte = texte_visible(artefact_bordereau_pdf())
        self.assertIn('Total TTC', texte)
        self.assertIn('QUATRE MILLIONS', texte)


class TestRatchetRouge(unittest.TestCase):
    """Le ratchet doit MORDRE : chaque motif est prouvé sur un cas injecté."""

    def test_un_champ_de_cout_injecte_dans_un_serializer_est_detecte(self):
        charge = artefact_rapport_approvisionnement()
        charge['controles'][0]['prix_achat'] = 2100
        self.assertTrue(scanner(charge))

    def test_un_cout_de_revient_dans_une_charge_api(self):
        self.assertTrue(scanner({'total': 1, 'cout_de_revient': 2666600}))
        self.assertTrue(scanner({'coût de revient': '2 666 600'}))

    def test_un_taux_de_marge(self):
        self.assertTrue(scanner({'taux_de_marge': '36.0'}))
        self.assertTrue(scanner({'marge_pct': 36}))
        self.assertTrue(scanner('Le taux de marge est de 36 %'))

    def test_un_benefice(self):
        self.assertTrue(scanner({'benefice_net_ht': 1500000}))
        self.assertTrue(scanner('Bénéfice visé : 1 500 000'))

    def test_un_coefficient_commercial(self):
        self.assertTrue(scanner({'coefficient_de_vente': '1.56'}))

    def test_un_maximum_agrege_de_site(self):
        self.assertTrue(scanner({'maximum_du_site': 640}))
        self.assertTrue(scanner('Potentiel maximal du site : 640 modules'))

    def test_une_fuite_dans_un_pdf_rendu(self):
        html = artefact_bordereau_pdf().replace(
            'Génie civil', 'Génie civil (marge brute 32 %)')
        self.assertTrue(scanner(html))

    def test_une_fuite_dans_un_classeur(self):
        lignes = bordereau_depose()
        lignes[0] = dict(lignes[0],
                         designation='Modules (prix d\'achat 2 100)')
        _, contexte = _lignes_et_contexte()
        self.assertTrue(scanner(bordereau_xlsx.rendre(lignes, contexte)))

    def test_le_mecanisme_marque_blanche_mord(self):
        """La liste est vide par décision ; le mécanisme, lui, fonctionne."""
        self.assertEqual(
            scanner('Étude réalisée par le bureau NOM-DU-BUREAU'), ())
        self.assertTrue(scanner('Étude réalisée par le bureau NOM-DU-BUREAU',
                                noms_marque_blanche=('NOM-DU-BUREAU',)))

    def test_les_cles_sont_scannees_meme_a_valeur_nulle(self):
        """`{"prix_achat": null}` est une fuite de SCHÉMA, pas de valeur."""
        self.assertTrue(scanner({'prix_achat': None}))


class TestPasDeFauxPositif(unittest.TestCase):
    """Un ratchet qui crie à tort finit désactivé — donc il ne crie pas à tort."""

    def test_les_marges_de_robustesse_du_calepinage_ne_sont_pas_une_fuite(self):
        self.assertEqual(scanner({
            'marges': {'troncon_min_cm': 4.15, 'bande_min_cm': 4.9,
                       'rangee_critique': 'R3'}}), ())

    def test_les_marges_typographiques_de_la_feuille_de_style(self):
        from apps.ao.fabrique.styles import css
        self.assertEqual(scanner(css()), ())

    def test_le_vocabulaire_courant_d_un_bordereau(self):
        self.assertEqual(scanner(
            'Prix unitaire, quantité, sous-total, TVA, total TTC, coffret'), ())

    def test_un_prix_de_vente_n_est_pas_un_prix_d_achat(self):
        self.assertEqual(scanner({'prix_unitaire': 2950,
                                  'prix_vente': 2950}), ())


if __name__ == '__main__':
    unittest.main()
