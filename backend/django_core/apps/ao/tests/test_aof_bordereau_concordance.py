"""AOF128 — le PDF du bordereau et son classeur disent la MÊME chose.

    python -m unittest apps.ao.tests.test_aof_bordereau_concordance -v
"""
import html
import io
import pathlib
import re
import unittest
from decimal import Decimal

from openpyxl import load_workbook

from apps.ao.fabrique.contexte import construire_contexte, litteraux_chiffres
from apps.ao.fabrique.rendus import bordereau_pdf as pdf
from apps.ao.fabrique.rendus import bordereau_xlsx as xlsx
from apps.ao.tests.aof_fixtures import (ARRETE_TTC, SOUS_TOTAUX_ATTENDUS,
                                        TOTAL_HT, TOTAL_TTC, TVA,
                                        bordereau_depose, contexte_dossier,
                                        rendre_gabarit)

GABARITS = pathlib.Path(__file__).resolve().parents[3] / 'templates'


def rendu_html(donnees):
    """Rend le gabarit sans base ni WeasyPrint (moteur Django autonome).

    Le HTML est déséchappé pour les assertions de contenu : Django écrit
    « Études d&#x27;exécution », le lecteur voit « Études d'exécution ».
    """
    return html.unescape(rendre_gabarit(pdf.NOM_GABARIT, donnees))


def artefacts():
    lignes = bordereau_depose()
    contexte = construire_contexte(contexte_dossier(lignes))
    donnees = pdf.contexte_gabarit(lignes, contexte)
    classeur = load_workbook(io.BytesIO(xlsx.rendre(lignes, contexte)))
    return lignes, contexte, donnees, classeur


def montants_du_classeur(classeur, lignes):
    """Relit les montants du classeur en évaluant ses formules simples."""
    feuille = classeur.active

    def valeur(coordonnee):
        brute = feuille[coordonnee].value
        if brute is None:
            return None
        if not isinstance(brute, str):
            return Decimal(str(brute))
        formule = brute.lstrip('=')
        produit = re.fullmatch(r'(\w+\d+)\*(\w+\d+)', formule)
        if produit:
            return valeur(produit.group(1)) * valeur(produit.group(2))
        somme = re.fullmatch(r'SUM\((\w+?)(\d+):(\w+?)(\d+)\)', formule)
        if somme:
            colonne = somme.group(1)
            total = Decimal('0')
            for rang in range(int(somme.group(2)), int(somme.group(4)) + 1):
                partiel = valeur('%s%d' % (colonne, rang))
                if partiel is not None:
                    total += partiel
            return total
        facteur = re.fullmatch(r'(\w+\d+)\*([\d.]+)', formule)
        if facteur:
            return valeur(facteur.group(1)) * Decimal(facteur.group(2))
        addition = re.fullmatch(r'(\w+\d+)\+(\w+\d+)', formule)
        if addition:
            return valeur(addition.group(1)) + valeur(addition.group(2))
        if re.fullmatch(r'\w+\d+(\+\w+\d+)+', formule):
            return sum((valeur(ref) for ref in formule.split('+')),
                       Decimal('0'))
        return None

    # Indexation par NUMÉRO de ligne (colonne A) : la désignation « Modules
    # photovoltaïques 625 Wc » apparaît dans les trois sections, indexer
    # dessus ferait lire trois fois la même rangée.
    par_numero = {}
    for rangee in feuille.iter_rows(min_col=1, max_col=1):
        for cellule in rangee:
            if isinstance(cellule.value, str) and cellule.value.isdigit():
                par_numero[cellule.value] = cellule.row

    par_libelle = {}
    for rangee in feuille.iter_rows(min_col=2, max_col=2):
        for cellule in rangee:
            if isinstance(cellule.value, str):
                par_libelle.setdefault(cellule.value, cellule.row)

    montants_lignes = {}
    for ligne in lignes:
        rang = par_numero.get(str(ligne.get('numero') or ''))
        if rang:
            montant = valeur('G%d' % rang)
            montants_lignes[str(ligne['cle'])] = (
                None if montant is None else montant.quantize(Decimal('0.01')))

    def total(libelle):
        rang = par_libelle.get(libelle)
        montant = valeur('G%d' % rang) if rang else None
        return None if montant is None else montant.quantize(Decimal('0.01'))

    return {'total_ht': total('Total HT'), 'tva': total('TVA (20 %)'),
            'total_ttc': total('Total TTC'), 'lignes': montants_lignes,
            'sous_totaux': {section: total('Sous-total %s' % section)
                            for section in SOUS_TOTAUX_ATTENDUS}}


class TestConcordance(unittest.TestCase):
    """Le cœur d'AOF128 : comparaison LIGNE À LIGNE et AU CENTIME."""

    def setUp(self):
        self.lignes, self.contexte, self.donnees, self.classeur = artefacts()
        self.pdf = pdf.valeurs_de_controle(self.donnees)
        self.xlsx = montants_du_classeur(self.classeur, self.lignes)

    def test_les_deux_artefacts_concordent(self):
        self.assertEqual(pdf.comparer(self.pdf, self.xlsx), ())
        self.assertTrue(pdf.exiger_concordance(self.pdf, self.xlsx))

    def test_ligne_a_ligne(self):
        self.assertEqual(sorted(self.pdf['lignes']),
                         sorted(self.xlsx['lignes']))
        for cle, montant in self.pdf['lignes'].items():
            self.assertEqual(montant, self.xlsx['lignes'][cle], cle)

    def test_sous_totaux_par_section(self):
        for section, attendu in SOUS_TOTAUX_ATTENDUS.items():
            self.assertEqual(self.pdf['sous_totaux'][section], attendu,
                             section)
            self.assertEqual(self.xlsx['sous_totaux'][section], attendu,
                             section)

    def test_recalcul_independant_du_cas_reel(self):
        """A 1 034 100 + B 744 200 + C 1 511 300 + communes 877 000."""
        somme = sum(SOUS_TOTAUX_ATTENDUS.values(), Decimal('0'))
        self.assertEqual(somme, TOTAL_HT)
        self.assertEqual((TOTAL_HT * Decimal('20') / Decimal('100')), TVA)
        self.assertEqual(TOTAL_HT + TVA, TOTAL_TTC)
        self.assertEqual(self.pdf['total_ht'], TOTAL_HT)
        self.assertEqual(self.pdf['tva'], TVA)
        self.assertEqual(self.pdf['total_ttc'], TOTAL_TTC)

    def test_une_divergence_injectee_est_detectee(self):
        fausse = dict(self.xlsx, total_ttc=Decimal('5219280'))
        divergences = pdf.comparer(self.pdf, fausse)
        self.assertEqual([d.repere for d in divergences], ['total_ttc'])
        with self.assertRaises(pdf.ConcordanceRompue):
            pdf.exiger_concordance(self.pdf, fausse)

    def test_une_divergence_de_ligne_est_detectee(self):
        lignes_faussees = dict(self.xlsx['lignes'])
        lignes_faussees['mod-a'] = Decimal('448500.00')
        divergences = pdf.comparer(self.pdf,
                                   dict(self.xlsx, lignes=lignes_faussees))
        self.assertEqual([d.repere for d in divergences], ['ligne mod-a'])
        self.assertIn('448500', divergences[0].motif)


class TestRenduPdf(unittest.TestCase):

    def setUp(self):
        self.lignes, self.contexte, self.donnees, _ = artefacts()
        self.html = rendu_html(self.donnees)

    def test_le_gabarit_se_rend(self):
        self.assertIn('Bordereau des prix', self.html)
        self.assertIn('TAQINOR SARL', self.html)

    def test_toutes_les_lignes_sont_imprimees(self):
        for ligne in self.lignes:
            self.assertIn(ligne['designation'], self.html,
                          ligne['designation'])

    def test_les_sections_et_leurs_sous_totaux(self):
        for section in SOUS_TOTAUX_ATTENDUS:
            self.assertIn(section, self.html, section)

    def test_l_arrete_en_lettres_est_imprime(self):
        self.assertIn(ARRETE_TTC, self.html)

    def test_la_clause_de_reserve_est_en_pied(self):
        from apps.ao.fabrique.clauses import CLAUSE_RESERVE_QUANTITES
        self.assertIn(CLAUSE_RESERVE_QUANTITES[:60], self.html)

    def test_les_pu_en_lettres_par_ligne(self):
        self.assertIn('DEUX MILLE NEUF CENT CINQUANTE DIRHAMS', self.html)

    def test_l_empreinte_du_contexte_est_imprimee(self):
        self.assertIn(self.contexte['empreinte'], self.html)

    def test_le_gabarit_ne_contient_aucun_chiffre_litteral(self):
        source = (GABARITS / 'ao' / 'bordereau.html').read_text(
            encoding='utf-8')
        self.assertEqual(litteraux_chiffres(source), ())

    def test_la_chaine_des_totaux_est_visible(self):
        for libelle in ('Sous-total HT', 'Total HT', 'TVA', 'Total TTC'):
            self.assertIn(libelle, self.html, libelle)

    def test_la_date_est_formatee_par_la_fabrique_pas_par_le_gabarit(self):
        """Sans cela, une pièce marocaine s'imprime « Aug. 1, 2026 »."""
        self.assertIn('01/08/2026', self.html)
        self.assertNotIn('Aug', self.html)

    def test_aucune_fuite_de_cout(self):
        texte = re.sub(r'<style\b.*?</style>|\{#.*?#\}', ' ', self.html,
                       flags=re.DOTALL | re.IGNORECASE)
        texte = re.sub(r'<[^>]+>', ' ', texte).lower()
        for interdit in ('prix_achat', "prix d'achat", 'coût de revient',
                         'marge', 'bénéfice'):
            self.assertNotIn(interdit, texte, interdit)


class TestContexteGabarit(unittest.TestCase):

    def test_la_remise_n_apparait_que_si_elle_existe(self):
        lignes = bordereau_depose()
        self.assertFalse(pdf.contexte_gabarit(lignes)['avec_remise'])

    def test_le_gabarit_recoit_la_feuille_de_style(self):
        self.assertIn('css_fabrique', pdf.contexte_gabarit(bordereau_depose()))

    def test_une_ligne_sans_pu_n_invente_rien(self):
        lignes = bordereau_depose() + [
            {'cle': 'a-chiffrer', 'section': 'Prestations communes',
             'designation': 'Poste à chiffrer', 'unite': 'ENS',
             'quantite': Decimal('1'), 'prix_unitaire': None}]
        donnees = pdf.contexte_gabarit(lignes)
        derniere = donnees['sections'][-1]['lignes'][-1]
        self.assertEqual(derniere['prix_unitaire_texte'], '')
        self.assertEqual(derniere['total_texte'], '')
        self.assertEqual(donnees['totaux'].total_ttc, TOTAL_TTC)


if __name__ == '__main__':
    unittest.main()
