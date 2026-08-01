"""AOF135 — simulation de rentabilité 25 ans : PIÈCE CLIENT, sans AUCUN coût.

Cette pièce part chez le maître d'ouvrage. Elle ne porte donc NI coût de
revient, NI marge, NI bénéfice — l'économie de l'AO vit dans une table séparée
derrière ``ao_rentabilite_voir`` (AOF157). **Les fusionner « parce que ça parle
de rentabilité » est le chemin le plus court vers la fuite de marge.**

Trois sorties, une seule source :

* :func:`contexte_simulation` — le dictionnaire dérivé (CAPEX du bordereau,
  puissance du calepinage, tableau annuel) ;
* :func:`classeur_xlsx` — un classeur à **formules VIVANTES** (le lecteur peut
  changer le tarif et voir bouger le retour, ce qu'un tableau de valeurs mortes
  ne permet pas) ;
* :func:`rendre_simulation_pdf` — le PDF via ``core.pdf.render_pdf`` (jamais un
  import direct de WeasyPrint).

L'empreinte :func:`empreinte_simulation` fige les entrées : une simulation dont
l'empreinte ne correspond plus au dossier est PÉRIMÉE.
"""
from __future__ import annotations

import hashlib
import json

__all__ = [
    'classeur_xlsx',
    'contexte_simulation',
    'controler_absence_de_cout',
    'empreinte_simulation',
    'html_simulation',
    'rafraichir_empreinte',
    'rendre_simulation_pdf',
]

#: Mots qui ne doivent JAMAIS apparaître dans cette pièce (garde local ; le
#: ratchet AOF129 l'étend à tous les artefacts client).
MOTS_INTERDITS_CLIENT = (
    "prix d'achat", 'coût de revient', 'marge', 'bénéfice',
    'maximum posable',
)


def _entrees_canoniques(simulation):
    """Le dictionnaire CANONIQUE des entrées — trié, stable, débogable."""
    return {
        'appel_offre': simulation.appel_offre.reference,
        'bordereau_total_ttc': str(simulation.capex_total),
        'bordereau_hors_stockage_ttc': str(simulation.capex_hors_stockage),
        'puissance_kwc': str(simulation.puissance_kwc),
        'duree_annees': int(simulation.duree_annees or 0),
        'productible_kwh_par_kwc_an': str(
            simulation.productible_kwh_par_kwc_an),
        'tarif_kwh': str(simulation.tarif_kwh),
        'inflation_tarif_pct': str(simulation.inflation_tarif_pct),
        'degradation_annuelle_pct': str(simulation.degradation_annuelle_pct),
        'taux_actualisation_pct': str(simulation.taux_actualisation_pct),
        'part_autoconsommee_pct': str(simulation.part_autoconsommee_pct),
    }


def empreinte_simulation(simulation):
    """SHA-256 des entrées de la simulation (même patron qu'AOF29)."""
    charge = json.dumps(_entrees_canoniques(simulation), sort_keys=True,
                        ensure_ascii=False, separators=(',', ':'))
    return hashlib.sha256(charge.encode('utf-8')).hexdigest()


def rafraichir_empreinte(simulation, *, save=True):
    """Recalcule et pose ``source_hash``. Renvoie ``(empreinte, a_change)``."""
    empreinte = empreinte_simulation(simulation)
    a_change = empreinte != (simulation.source_hash or '')
    if a_change and save:
        simulation.source_hash = empreinte
        simulation.save(update_fields=['source_hash', 'updated_at'])
    else:
        simulation.source_hash = empreinte
    return empreinte, a_change


def _texte(valeur):
    return '' if valeur is None else str(valeur)


def contexte_simulation(simulation):
    """Contexte DÉRIVÉ de la simulation — aucune valeur saisie deux fois."""
    return {
        'appel_offre': {
            'reference': simulation.appel_offre.reference,
            'objet': simulation.appel_offre.objet,
            'maitre_ouvrage': simulation.appel_offre.maitre_ouvrage
            or simulation.appel_offre.acheteur,
        },
        'soumissionnaire': simulation.appel_offre.soumissionnaire
        or simulation.appel_offre.company.nom,
        'duree_annees': int(simulation.duree_annees or 0),
        'puissance_kwc': _texte(simulation.puissance_kwc),
        'productible_kwh_an': _texte(simulation.productible_kwh_an),
        'productible_source': simulation.productible_source,
        'tarif_kwh': _texte(simulation.tarif_kwh),
        'inflation_tarif_pct': _texte(simulation.inflation_tarif_pct),
        'degradation_annuelle_pct': _texte(
            simulation.degradation_annuelle_pct),
        'taux_actualisation_pct': _texte(simulation.taux_actualisation_pct),
        'capex_total': _texte(simulation.capex_total),
        'capex_hors_stockage': _texte(simulation.capex_hors_stockage),
        'economie_annuelle_initiale': _texte(
            simulation.economie_annuelle_initiale),
        'economies_cumulees': _texte(simulation.economies_cumulees),
        'payback_simple_ans': _texte(simulation.payback_simple_ans),
        'payback_actualise_ans': _texte(simulation.payback_actualise_ans),
        'roi_sur_ttc_ans': _texte(simulation.roi_sur_ttc_ans),
        'tableau': simulation.tableau_annuel,
        'source_hash': empreinte_simulation(simulation),
    }


def classeur_xlsx(simulation):
    """Classeur XLSX à FORMULES VIVANTES (bytes).

    Les paramètres sont écrits en cellules NOMMÉES et le tableau annuel est
    construit en FORMULES qui les référencent : changer le tarif dans le
    classeur recalcule tout. Un tableau de valeurs mortes ne serait qu'une
    capture d'écran chiffrée.
    """
    from io import BytesIO

    from openpyxl import Workbook

    classeur = Workbook()
    feuille = classeur.active
    feuille.title = 'Rentabilité'

    feuille['A1'] = 'Paramètres'
    parametres = (
        ('Puissance installée (kWc)', simulation.puissance_kwc),
        ('Productible spécifique (kWh/kWc/an)',
         simulation.productible_kwh_par_kwc_an),
        ('Part autoconsommée (%)', simulation.part_autoconsommee_pct),
        ('Tarif du kWh évité (MAD)', simulation.tarif_kwh),
        ('Inflation annuelle du tarif (%)', simulation.inflation_tarif_pct),
        ('Dégradation annuelle (%)', simulation.degradation_annuelle_pct),
        ("Taux d'actualisation (%)", simulation.taux_actualisation_pct),
        ('CAPEX hors stockage (MAD TTC)', simulation.capex_hors_stockage),
        ('CAPEX total remis (MAD TTC)', simulation.capex_total),
    )
    for index, (libelle, valeur) in enumerate(parametres, start=2):
        feuille.cell(row=index, column=1, value=libelle)
        feuille.cell(row=index, column=2, value=float(valeur or 0))

    # Références des paramètres, utilisées par les formules du tableau.
    ref_puissance, ref_productible = '$B$2', '$B$3'
    ref_part, ref_tarif = '$B$4', '$B$5'
    ref_inflation, ref_degradation = '$B$6', '$B$7'
    ref_actualisation = '$B$8'
    #: Ligne 9 = « CAPEX hors stockage » (ligne 10 = CAPEX total remis).
    ref_capex_hs = '$B$9'

    entete = 12
    for colonne, titre in enumerate(
            ('Année', 'Productible (kWh)', 'Tarif (MAD/kWh)',
             'Économie (MAD)', 'Économie cumulée (MAD)',
             'Économie actualisée (MAD)', 'Reste à couvrir (MAD)'), start=1):
        feuille.cell(row=entete, column=colonne, value=titre)

    duree = int(simulation.duree_annees or 0)
    for rang in range(1, duree + 1):
        ligne = entete + rang
        feuille.cell(row=ligne, column=1, value=rang)
        feuille.cell(row=ligne, column=2, value=(
            f'={ref_puissance}*{ref_productible}'
            f'*(1-{ref_degradation}/100)^(A{ligne}-1)'))
        feuille.cell(row=ligne, column=3, value=(
            f'={ref_tarif}*(1+{ref_inflation}/100)^(A{ligne}-1)'))
        feuille.cell(row=ligne, column=4, value=(
            f'=B{ligne}*C{ligne}*{ref_part}/100'))
        precedent = f'E{ligne - 1}' if rang > 1 else '0'
        feuille.cell(row=ligne, column=5, value=f'={precedent}+D{ligne}')
        feuille.cell(row=ligne, column=6, value=(
            f'=D{ligne}/(1+{ref_actualisation}/100)^A{ligne}'))
        feuille.cell(row=ligne, column=7, value=(
            f'=MAX(0,{ref_capex_hs}-E{ligne})'))

    flux = BytesIO()
    classeur.save(flux)
    return flux.getvalue()


def html_simulation(simulation, *, contexte=None):
    """HTML de la pièce client — construit depuis le contexte, jamais saisi."""
    from django.utils.html import escape

    contexte = contexte or contexte_simulation(simulation)
    lignes = ''.join(
        '<tr>'
        f'<td>{ligne["annee"]}</td>'
        f'<td>{ligne["productible_kwh"]}</td>'
        f'<td>{ligne["economie"]}</td>'
        f'<td>{ligne["economie_cumulee"]}</td>'
        '</tr>'
        for ligne in contexte['tableau'])
    return (
        '<article class="ao-simulation">'
        f'<h1>Simulation de rentabilité sur {contexte["duree_annees"]} ans</h1>'
        f'<p>{escape(contexte["appel_offre"]["objet"])}</p>'
        '<dl>'
        f'<dt>Puissance installée (kWc)</dt><dd>{contexte["puissance_kwc"]}</dd>'
        f'<dt>Productible (kWh/an)</dt>'
        f'<dd>{contexte["productible_kwh_an"]} '
        f'({escape(contexte["productible_source"])})</dd>'
        f'<dt>CAPEX hors stockage (MAD TTC)</dt>'
        f'<dd>{contexte["capex_hors_stockage"]}</dd>'
        f'<dt>Montant remis (MAD TTC)</dt><dd>{contexte["capex_total"]}</dd>'
        f'<dt>Économie de la première année (MAD)</dt>'
        f'<dd>{contexte["economie_annuelle_initiale"]}</dd>'
        f'<dt>Retour simple (ans)</dt><dd>{contexte["payback_simple_ans"]}</dd>'
        f'<dt>Retour actualisé (ans)</dt>'
        f'<dd>{contexte["payback_actualise_ans"]}</dd>'
        f'<dt>Retour sur le montant remis (ans)</dt>'
        f'<dd>{contexte["roi_sur_ttc_ans"]}</dd>'
        f'<dt>Économies cumulées (MAD)</dt>'
        f'<dd>{contexte["economies_cumulees"]}</dd>'
        '</dl>'
        '<table><thead><tr><th>Année</th><th>Productible (kWh)</th>'
        '<th>Économie (MAD)</th><th>Cumul (MAD)</th></tr></thead>'
        f'<tbody>{lignes}</tbody></table>'
        f'<p class="ao-simulation__empreinte">Empreinte : '
        f'{contexte["source_hash"][:8]}</p>'
        '</article>'
    )


def rendre_simulation_pdf(simulation, *, contexte=None):
    """PDF de la pièce client via ``core.pdf.render_pdf`` (ARC11)."""
    from core.pdf import render_pdf

    return render_pdf(
        html=html_simulation(simulation, contexte=contexte),
        company=simulation.company, header=False, footer=False)


def controler_absence_de_cout(texte):
    """Ratchet local : refuse un rendu client contenant un mot de coût.

    Renvoie la liste des mots trouvés (vide = propre). AOF129 étend ce contrôle
    à TOUS les artefacts ; ici on garde la pièce la plus exposée.
    """
    minuscule = (texte or '').lower()
    return [mot for mot in MOTS_INTERDITS_CLIENT if mot in minuscule]
