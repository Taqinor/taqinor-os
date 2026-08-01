"""AOF160 — classeur DIRECTEUR de rentabilité attendue. Jamais dans un pack.

Objet distinct, pas une variante
================================
La « simulation de rentabilité 25 ans » (AOF135) est une pièce CLIENT sans
aucun coût. Ce classeur-ci est son opposé : il ne contient QUE des coûts. Les
fusionner « parce que ça parle de rentabilité » est le chemin le plus court
vers la fuite de marge — c'est écrit noir sur blanc dans le plan, et ce module
matérialise la séparation.

L'exclusion est STRUCTURELLE : le classeur porte ``VISIBILITE = 'directeur'``,
et les trois assembleurs de pack livrés par cette lane (sommaire AOF139, ZIP
AOF151, bon à tirer AOF152) filtrent sur la visibilité. Il ne peut donc pas
entrer dans un manifeste de dépôt, même si on le demande explicitement.

L'identité qui rend le classeur vérifiable
------------------------------------------
    trésorerie = encaissements TTC − décaissements achats TTC − TVA nette
               = bénéfice net HT

Le cas réel (dossier FRDISI) la vérifie au dirham :
    4 999 920 − 3 150 640 − 349 280 = 1 500 000

Si la cellule de contrôle de trésorerie ne tombe pas exactement sur le
bénéfice, le classeur est ROUGE : c'est ce contrôle, et non une relecture, qui
attrape une erreur de ventilation de TVA.

Un montant est un NOMBRE
------------------------
Comme pour le bordereau XLSX (AOF127) : chaque montant est écrit en valeur
numérique avec un format de cellule ``# ##0,00 "DH"``. Un montant-chaîne casse
les sommes Excel, se comporte différemment selon la locale du poste, et rend
le classeur illisible par tout contrôle automatique.
"""
from __future__ import annotations

from decimal import ROUND_HALF_UP, Decimal

__all__ = [
    'VISIBILITE',
    'FORMAT_MONTANT',
    'FORMAT_TAUX',
    'ControleTresorerieRouge',
    'construire_economie',
    'ecrire_classeur',
]

VISIBILITE = 'directeur'
FORMAT_MONTANT = '# ##0,00 "DH"'
FORMAT_TAUX = '0,0 %'
CENTIME = Decimal('0.01')

#: Taux de TVA de vente applicable aux panneaux dans la VARIANTE étudiée
#: (point ouvert signalé au comptable), contre le taux standard par défaut.
TAUX_VARIANTE_PANNEAUX = Decimal('0.10')


class ControleTresorerieRouge(Exception):
    """Levée quand la trésorerie ne tombe pas exactement sur le bénéfice."""


def _q(valeur):
    return Decimal(str(valeur)).quantize(CENTIME, rounding=ROUND_HALF_UP)


def construire_economie(postes, *, total_vente_ht, taux_tva_vente,
                        base_panneaux_ht=None):
    """Calcule toute l'économie du dossier à partir des POSTES de coût.

    Args:
        postes: ``[{'libelle', 'montant_ht', 'taux_tva_achat'}]`` — la TVA sur
            achats est DIFFÉRENCIÉE poste par poste (10 % sur les panneaux,
            20 % sur le reste) : un taux global unique fausserait la TVA nette
            et donc le contrôle de trésorerie.
        total_vente_ht: total HT du bordereau (sortie de la cascade AOF158).
        taux_tva_vente: taux standard de TVA de vente (0,20).
        base_panneaux_ht: part panneaux du prix de VENTE, pour la variante
            « panneaux facturés à 10 % ». Sans elle, la variante n'est pas
            calculée (plutôt que calculée faux).

    Returns:
        dict complet, avec ``controle_tresorerie`` et ``ecart_controle``.
    """
    postes = list(postes or ())
    if not postes:
        raise ControleTresorerieRouge(
            "Aucun poste de coût : il n'y a rien à contrôler.")

    lignes = []
    cout_revient_ht = Decimal('0')
    tva_deductible = Decimal('0')
    for poste in postes:
        montant = _q(poste['montant_ht'])
        taux = Decimal(str(poste.get('taux_tva_achat', '0.20')))
        tva = _q(montant * taux)
        lignes.append({
            'libelle': poste.get('libelle', ''),
            'montant_ht': montant,
            'taux_tva_achat': taux,
            'tva': tva,
            'montant_ttc': _q(montant + tva),
        })
        cout_revient_ht += montant
        tva_deductible += tva
    cout_revient_ht = _q(cout_revient_ht)
    tva_deductible = _q(tva_deductible)
    cout_revient_ttc = _q(cout_revient_ht + tva_deductible)

    vente_ht = _q(total_vente_ht)
    taux_vente = Decimal(str(taux_tva_vente))
    tva_collectee = _q(vente_ht * taux_vente)
    vente_ttc = _q(vente_ht + tva_collectee)
    tva_nette = _q(tva_collectee - tva_deductible)
    benefice_net_ht = _q(vente_ht - cout_revient_ht)

    # L'identité qui rend le classeur vérifiable — et non « relu ».
    tresorerie = _q(vente_ttc - cout_revient_ttc - tva_nette)
    ecart = _q(tresorerie - benefice_net_ht)
    if ecart != 0:
        raise ControleTresorerieRouge(
            "Contrôle de trésorerie ROUGE : {} ≠ bénéfice net {} (écart {}). "
            "La ventilation de TVA est fausse quelque part.".format(
                tresorerie, benefice_net_ht, ecart))

    variante = None
    if base_panneaux_ht is not None:
        base = _q(base_panneaux_ht)
        if base > vente_ht:
            raise ControleTresorerieRouge(
                "Base panneaux {} supérieure au total de vente {}.".format(
                    base, vente_ht))
        tva_variante = _q(base * TAUX_VARIANTE_PANNEAUX
                          + (vente_ht - base) * taux_vente)
        ttc_variante = _q(vente_ht + tva_variante)
        variante = {
            'base_panneaux_ht': base,
            'taux_panneaux': TAUX_VARIANTE_PANNEAUX,
            'tva_collectee': tva_variante,
            'total_ttc': ttc_variante,
            'ecart_ttc': _q(ttc_variante - vente_ttc),
            'benefice_net_ht': benefice_net_ht,  # INCHANGÉ, c'est le point
            'tva_nette': _q(tva_variante - tva_deductible),
        }

    return {
        'visibilite': VISIBILITE,
        'postes': lignes,
        'cout_revient_ht': cout_revient_ht,
        'tva_deductible': tva_deductible,
        'cout_revient_ttc': cout_revient_ttc,
        'vente_ht': vente_ht,
        'taux_tva_vente': taux_vente,
        'tva_collectee': tva_collectee,
        'vente_ttc': vente_ttc,
        'tva_nette': tva_nette,
        'benefice_net_ht': benefice_net_ht,
        'marge_pct': (benefice_net_ht / vente_ht * Decimal('100')).quantize(
            Decimal('0.1')) if vente_ht else Decimal('0'),
        'controle_tresorerie': tresorerie,
        'ecart_controle': ecart,
        'variante_panneaux_10': variante,
    }


def _montant(feuille, ligne, colonne, valeur):
    cellule = feuille.cell(row=ligne, column=colonne, value=float(valeur))
    cellule.number_format = FORMAT_MONTANT
    return cellule


def ecrire_classeur(economie, destination, *, reference_dossier=''):
    """Écrit le classeur .xlsx dans ``destination`` (fichier binaire).

    openpyxl est déjà en production (``openpyxl==3.1.5``) — import
    fonction-local, comme partout dans ce dépôt pour les libs lourdes.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    classeur = Workbook()
    feuille = classeur.active
    feuille.title = 'Rentabilité'
    gras = Font(bold=True)

    feuille['A1'] = 'INTERNE — DIRECTION. Ne jamais joindre au dossier.'
    feuille['A1'].font = Font(bold=True, color='AA0000')
    feuille['A2'] = 'Dossier'
    feuille['B2'] = reference_dossier

    ligne = 4
    feuille.cell(row=ligne, column=1, value='Coût de revient par poste').font \
        = gras
    ligne += 1
    for index, entete in enumerate(
            ('Poste', 'Montant HT', 'Taux TVA achat', 'TVA', 'Montant TTC'),
            start=1):
        feuille.cell(row=ligne, column=index, value=entete).font = gras
    ligne += 1
    for poste in economie['postes']:
        feuille.cell(row=ligne, column=1, value=poste['libelle'])
        _montant(feuille, ligne, 2, poste['montant_ht'])
        taux = feuille.cell(row=ligne, column=3,
                            value=float(poste['taux_tva_achat']))
        taux.number_format = FORMAT_TAUX
        _montant(feuille, ligne, 4, poste['tva'])
        _montant(feuille, ligne, 5, poste['montant_ttc'])
        ligne += 1

    feuille.cell(row=ligne, column=1, value='Total coût de revient').font = gras
    _montant(feuille, ligne, 2, economie['cout_revient_ht'])
    _montant(feuille, ligne, 4, economie['tva_deductible'])
    _montant(feuille, ligne, 5, economie['cout_revient_ttc'])
    ligne += 2

    feuille.cell(row=ligne, column=1, value='Vente').font = gras
    ligne += 1
    for libelle, valeur in (('Total HT (bordereau)', economie['vente_ht']),
                            ('TVA collectée', economie['tva_collectee']),
                            ('Total TTC', economie['vente_ttc'])):
        feuille.cell(row=ligne, column=1, value=libelle)
        _montant(feuille, ligne, 2, valeur)
        ligne += 1
    ligne += 1

    feuille.cell(row=ligne, column=1, value='Résultat').font = gras
    ligne += 1
    for libelle, valeur in (
            ('TVA nette à reverser', economie['tva_nette']),
            ('Bénéfice net HT', economie['benefice_net_ht']),
            ('CONTRÔLE DE TRÉSORERIE', economie['controle_tresorerie']),
            ('Écart de contrôle (doit être nul)',
             economie['ecart_controle'])):
        cellule = feuille.cell(row=ligne, column=1, value=libelle)
        if libelle.startswith('CONTRÔLE'):
            cellule.font = gras
        _montant(feuille, ligne, 2, valeur)
        ligne += 1
    feuille.cell(row=ligne, column=1, value='Marge')
    pourcentage = feuille.cell(row=ligne, column=2,
                               value=float(economie['marge_pct']) / 100)
    pourcentage.number_format = FORMAT_TAUX
    ligne += 2

    variante = economie.get('variante_panneaux_10')
    if variante:
        feuille.cell(row=ligne, column=1,
                     value='Variante — panneaux facturés à 10 %').font = gras
        ligne += 1
        for libelle, valeur in (
                ('Base panneaux HT', variante['base_panneaux_ht']),
                ('TVA collectée (variante)', variante['tva_collectee']),
                ('Total TTC (variante)', variante['total_ttc']),
                ('Écart TTC vs offre', variante['ecart_ttc']),
                ('Bénéfice net HT (inchangé)',
                 variante['benefice_net_ht'])):
            feuille.cell(row=ligne, column=1, value=libelle)
            _montant(feuille, ligne, 2, valeur)
            ligne += 1

    classeur.save(destination)
    return destination
