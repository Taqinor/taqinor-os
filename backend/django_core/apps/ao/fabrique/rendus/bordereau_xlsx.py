"""AOF127 — export XLSX du bordereau : un montant est un NOMBRE.

**Le point dur.** Un montant écrit « 4 999 920,00 DH » dans une cellule est une
CHAÎNE. Conséquences, toutes constatées sur des bordereaux réels :

* les sommes Excel de l'acheteur renvoient 0 — il refait donc les additions à
  la main, et c'est là que naissent les écarts ;
* le contrôleur de cohérence ne peut plus relire le classeur ;
* le rendu change selon la locale du poste qui l'ouvre.

Ici, la valeur de cellule est un nombre et le séparateur de milliers comme la
virgule décimale sont portés par le FORMAT de cellule. Le format stocké est
`[$-40C]#,##0.00\\ "DH"` : dans un fichier XLSX, les séparateurs du format
s'écrivent en notation neutre (`,` = milliers, `.` = décimale) et le préfixe
`[$-40C]` force le rendu français — Excel affiche donc « 4 999 920,00 DH »
quelle que soit la locale du lecteur. C'est la traduction exacte du
`# ##0,00 "DH"` demandé, écrite dans la syntaxe que le format de fichier
impose.

**Les sommes sont des FORMULES.** Sous-totaux, total HT, TVA et TTC sont écrits
en `=SUM(...)` / `=…*…` et non en valeurs figées : l'acheteur qui touche une
cellule voit le classeur se recalculer, au lieu de lire un total devenu faux.

**Seule chaîne autorisée : l'arrêté en lettres.** Par nature.

Le module ne fait AUCUNE E/S : il construit un classeur openpyxl et sait le
sérialiser en octets. Le téléversement (`records.Attachment`) et le passage en
tâche de fond (`core.jobs.submit`) appartiennent à la couche Django.
"""
from __future__ import annotations

import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..clauses import CLAUSE_RESERVE_QUANTITES
from ..montants import arrete, en_lettres
from ..ordonnancement import montant_ligne, sections_et_lignes, totaux

#: Format de cellule des montants (voir l'en-tête du module).
FORMAT_MONTANT = '[$-40C]#,##0.00\\ "DH"'
FORMAT_QUANTITE = '[$-40C]#,##0.###'

#: Au-delà de ce nombre de lignes, le rendu part en tâche de fond (patron
#: SCA41) : un classeur de plusieurs centaines de lignes ne se construit pas
#: dans le temps d'une requête HTTP.
SEUIL_JOB = 200

COLONNES = (
    ('N°', 6), ('Désignation', 52), ('Unité', 9), ('Quantité', 12),
    ('Prix unitaire HT', 16), ('Prix unitaire en toutes lettres', 46),
    ('Total HT', 16))

COL_NUMERO, COL_DESIGNATION, COL_UNITE = 1, 2, 3
COL_QUANTITE, COL_PU, COL_LETTRES, COL_TOTAL = 4, 5, 6, 7

_TRAIT = Side(style='thin', color='FF9A9A9A')
_BORDURE = Border(left=_TRAIT, right=_TRAIT, top=_TRAIT, bottom=_TRAIT)
_FOND_ENTETE = PatternFill('solid', fgColor='FFEEEEEE')


def doit_passer_en_job(nb_lignes):
    """Le rendu doit-il partir en tâche de fond ?"""
    return int(nb_lignes or 0) > SEUIL_JOB


def _d(valeur):
    if valeur is None or valeur == '':
        return None
    return valeur if isinstance(valeur, Decimal) else Decimal(str(valeur))


def _nombre(valeur):
    """`Decimal` → `float` pour openpyxl, sans passer par une chaîne."""
    return None if valeur is None else float(valeur)


def construire_classeur(lignes, contexte=None, *, texte_clause=None,
                        taux_tva=Decimal('20')):
    """Construit le classeur du bordereau. Retourne un `Workbook` openpyxl.

    :param lignes: lignes du bordereau (déjà renumérotées, AOF123).
    :param contexte: contexte de dossier (AOF111) — identité, objet, empreinte.
    :param texte_clause: clause de réserve (AOF126) ; défaut = texte de
        référence.
    """
    contexte = contexte or {}
    classeur = Workbook()
    feuille = classeur.active
    feuille.title = 'Bordereau des prix'

    rang = _ecrire_entete(feuille, contexte)
    rang = _ecrire_colonnes(feuille, rang)
    premiere_ligne = rang
    rangs_sous_totaux = []

    for section, lignes_section in sections_et_lignes(lignes):
        rang = _ecrire_section(feuille, rang, section, lignes_section,
                               rangs_sous_totaux)

    rang = _ecrire_totaux(feuille, rang, rangs_sous_totaux, lignes, taux_tva)
    rang = _ecrire_arrete(feuille, rang, lignes, taux_tva)
    _ecrire_clause(feuille, rang, texte_clause)
    _mise_en_page(feuille, premiere_ligne)
    return classeur


def _ecrire_entete(feuille, contexte):
    identite = contexte.get('identite') or {}
    marche = contexte.get('marche') or {}
    rang = 1
    for texte, gras, taille in (
            (identite.get('raison_sociale', ''), True, 12),
            (marche.get('objet', ''), False, 11),
            (marche.get('reference_acheteur', ''), False, 10)):
        if not texte:
            continue
        cellule = feuille.cell(row=rang, column=COL_NUMERO, value=texte)
        cellule.font = Font(bold=gras, size=taille)
        feuille.merge_cells(start_row=rang, start_column=COL_NUMERO,
                            end_row=rang, end_column=COL_TOTAL)
        rang += 1
    return rang + 1


def _ecrire_colonnes(feuille, rang):
    for index, (titre, largeur) in enumerate(COLONNES, start=1):
        cellule = feuille.cell(row=rang, column=index, value=titre)
        cellule.font = Font(bold=True)
        cellule.fill = _FOND_ENTETE
        cellule.border = _BORDURE
        cellule.alignment = Alignment(horizontal='center', vertical='center',
                                      wrap_text=True)
        feuille.column_dimensions[get_column_letter(index)].width = largeur
    return rang + 1


def _ecrire_section(feuille, rang, section, lignes_section, rangs_sous_totaux):
    if section:
        cellule = feuille.cell(row=rang, column=COL_NUMERO, value=section)
        cellule.font = Font(bold=True)
        cellule.fill = _FOND_ENTETE
        feuille.merge_cells(start_row=rang, start_column=COL_NUMERO,
                            end_row=rang, end_column=COL_TOTAL)
        rang += 1

    premiere = rang
    for ligne in lignes_section:
        _ecrire_ligne(feuille, rang, ligne)
        rang += 1
    derniere = rang - 1

    libelle = 'Sous-total %s' % section if section else 'Sous-total'
    cellule = feuille.cell(row=rang, column=COL_DESIGNATION, value=libelle)
    cellule.font = Font(bold=True)
    total = feuille.cell(row=rang, column=COL_TOTAL)
    if derniere >= premiere:
        colonne = get_column_letter(COL_TOTAL)
        total.value = '=SUM(%s%d:%s%d)' % (colonne, premiere, colonne,
                                           derniere)
    else:
        total.value = 0
    total.number_format = FORMAT_MONTANT
    total.font = Font(bold=True)
    total.border = _BORDURE
    rangs_sous_totaux.append(rang)
    return rang + 1


def _ecrire_ligne(feuille, rang, ligne):
    quantite = _d(ligne.get('quantite'))
    prix = _d(ligne.get('prix_unitaire'))

    feuille.cell(row=rang, column=COL_NUMERO,
                 value=str(ligne.get('numero') or ''))
    feuille.cell(row=rang, column=COL_DESIGNATION,
                 value=str(ligne.get('designation') or ''))
    feuille.cell(row=rang, column=COL_UNITE,
                 value=str(ligne.get('unite') or ''))

    cellule_quantite = feuille.cell(row=rang, column=COL_QUANTITE,
                                    value=_nombre(quantite))
    cellule_quantite.number_format = FORMAT_QUANTITE

    cellule_prix = feuille.cell(row=rang, column=COL_PU, value=_nombre(prix))
    cellule_prix.number_format = FORMAT_MONTANT

    feuille.cell(row=rang, column=COL_LETTRES,
                 value=en_lettres(prix) or None).alignment = Alignment(
        wrap_text=True, vertical='top')

    cellule_total = feuille.cell(row=rang, column=COL_TOTAL)
    if quantite is not None and prix is not None:
        cellule_total.value = '=%s%d*%s%d' % (
            get_column_letter(COL_QUANTITE), rang,
            get_column_letter(COL_PU), rang)
    else:
        cellule_total.value = None
    cellule_total.number_format = FORMAT_MONTANT

    for colonne in range(1, len(COLONNES) + 1):
        cellule = feuille.cell(row=rang, column=colonne)
        cellule.border = _BORDURE
        # Les cellules de PRIX sont verrouillées ; le reste ne l'est pas, pour
        # que l'acheteur puisse annoter sans casser le chiffrage.
        cellule.protection = _protection(colonne in (COL_PU, COL_TOTAL))
    feuille.cell(row=rang, column=COL_DESIGNATION).alignment = Alignment(
        wrap_text=True, vertical='top')


def _protection(verrouillee):
    from openpyxl.styles import Protection
    return Protection(locked=bool(verrouillee))


def _ecrire_totaux(feuille, rang, rangs_sous_totaux, lignes, taux_tva):
    colonne = get_column_letter(COL_TOTAL)
    rang += 1
    somme = '+'.join('%s%d' % (colonne, r) for r in rangs_sous_totaux) or '0'

    rang_ht = rang
    _ecrire_total(feuille, rang, 'Total HT', '=%s' % somme)
    rang += 1
    _ecrire_total(feuille, rang, 'TVA (%s %%)' % _taux(taux_tva),
                  '=%s%d*%s' % (colonne, rang_ht,
                                _taux(taux_tva / Decimal('100'))))
    rang_tva = rang
    rang += 1
    _ecrire_total(feuille, rang, 'Total TTC',
                  '=%s%d+%s%d' % (colonne, rang_ht, colonne, rang_tva),
                  double=True)
    return rang + 1


def _taux(valeur):
    """Un taux s'écrit avec un point dans une formule Excel, jamais autrement."""
    texte = format(Decimal(str(valeur)).normalize(), 'f')
    return texte


def _ecrire_total(feuille, rang, libelle, formule, double=False):
    cellule = feuille.cell(row=rang, column=COL_DESIGNATION, value=libelle)
    cellule.font = Font(bold=True)
    total = feuille.cell(row=rang, column=COL_TOTAL, value=formule)
    total.number_format = FORMAT_MONTANT
    total.font = Font(bold=True, size=12 if double else 11)
    total.border = _BORDURE
    total.protection = _protection(True)


def _ecrire_arrete(feuille, rang, lignes, taux_tva):
    rang += 1
    calcules = totaux(lignes, taux_defaut=taux_tva)
    cellule = feuille.cell(row=rang, column=COL_NUMERO,
                           value=arrete(calcules.total_ttc))
    cellule.font = Font(bold=True)
    cellule.alignment = Alignment(wrap_text=True, vertical='center')
    feuille.merge_cells(start_row=rang, start_column=COL_NUMERO,
                        end_row=rang, end_column=COL_TOTAL)
    feuille.row_dimensions[rang].height = 34
    return rang + 2


def _ecrire_clause(feuille, rang, texte_clause):
    cellule = feuille.cell(row=rang, column=COL_NUMERO,
                           value=texte_clause or CLAUSE_RESERVE_QUANTITES)
    cellule.alignment = Alignment(wrap_text=True, vertical='top')
    cellule.font = Font(italic=True, size=9)
    feuille.merge_cells(start_row=rang, start_column=COL_NUMERO,
                        end_row=rang, end_column=COL_TOTAL)
    feuille.row_dimensions[rang].height = 58


def _mise_en_page(feuille, premiere_ligne):
    feuille.page_setup.orientation = 'landscape'
    feuille.page_setup.paperSize = feuille.PAPERSIZE_A4
    feuille.page_setup.fitToWidth = 1
    feuille.page_setup.fitToHeight = 0
    feuille.sheet_properties.pageSetUpPr.fitToPage = True
    feuille.print_title_rows = '%d:%d' % (premiere_ligne - 1,
                                          premiere_ligne - 1)
    feuille.freeze_panes = feuille.cell(row=premiere_ligne, column=1)
    feuille.protection.sheet = True
    feuille.protection.enable()


def vers_octets(classeur):
    """Sérialise le classeur — la seule sortie du module."""
    tampon = io.BytesIO()
    classeur.save(tampon)
    return tampon.getvalue()


def rendre(lignes, contexte=None, **options):
    """Raccourci : lignes → octets XLSX."""
    return vers_octets(construire_classeur(lignes, contexte, **options))


def valeurs_de_controle(lignes, *, taux_tva=Decimal('20')):
    """Les montants que le classeur DOIT afficher — recalculés côté Python.

    Sert au test de concordance (AOF128) et au contrôleur de cohérence : le
    classeur porte des formules, cette fonction porte la vérité arithmétique.
    """
    calcules = totaux(lignes, taux_defaut=taux_tva)
    return {'total_ht': calcules.total_ht, 'tva': calcules.tva,
            'total_ttc': calcules.total_ttc,
            'lignes': {str(ligne.get('cle') or ligne.get('numero') or ''):
                       montant_ligne(ligne) for ligne in lignes or ()}}
