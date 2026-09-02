"""CRX16 — la portée de rôle s'applique aussi à l'export et à la recherche.

Deux trous constatés par l'audit L3 du CRM :

* ``LeadViewSet.export_xlsx`` construisait son queryset à la main
  (``Lead.objects.filter(company=…)``) au lieu de passer par
  ``self.get_queryset()`` — un rôle restreint exportait donc en .xlsx des
  leads que sa propre liste lui masque (parité rompue avec
  ``ClientViewSet.export_xlsx``, qui passe bien par ``get_queryset()``).
  Au passage, ``exports.lead_row`` faisait ``lead.devis.order_by(…).first()``
  par ligne : un N+1 qu'un prefetch NU ne referme pas (le ``order_by`` sur le
  manager reconstruit un queryset et ignore le cache).
* ``ClientViewSet.search`` (autocomplete QC1) interrogeait directement
  ``Client.objects.filter(company=…)`` — nom, adresse, téléphone et email
  d'enregistrements hors portée passaient par là.
"""
import io

from django.contrib.auth import get_user_model
from django.test import TestCase
from openpyxl import load_workbook
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import AccessToken

from authentication.models import Company
from apps.roles.models import Role
from core.test_utils import AssertQueryBudgetMixin

from .models import Client, Lead

User = get_user_model()

_PERMS = ['crm_voir', 'crm_creer', 'crm_modifier']


def _lignes_xlsx(response):
    """Valeurs de la 1re colonne (Nom) des lignes de données d'un export."""
    classeur = load_workbook(io.BytesIO(response.content))
    feuille = classeur.active
    return [row[0] for row in feuille.iter_rows(min_row=2, values_only=True)]


class Crx16Base(TestCase):
    def setUp(self):
        self.company = Company.objects.create(
            nom='CRX16 Co', slug='crx16-co')
        # Rôle RESTREINT à l'équipe : sans superviseur, sa portée se limite à
        # lui-même (``core.scoping.peer_user_ids``).
        self.role_restreint = Role.objects.create(
            company=self.company, nom='Commercial',
            permissions=_PERMS + ['records_scope_equipe'], est_systeme=False)
        # Rôle SANS marqueur de portée → voit tout (comportement historique).
        self.role_large = Role.objects.create(
            company=self.company, nom='Responsable',
            permissions=_PERMS, est_systeme=False)

        self.restreint = User.objects.create_user(
            username='crx16_restreint', password='x', company=self.company,
            role=self.role_restreint, role_legacy='responsable')
        self.autre = User.objects.create_user(
            username='crx16_autre', password='x', company=self.company,
            role=self.role_large, role_legacy='responsable')

        self.api = APIClient()

    def auth(self, user):
        self.api.credentials(
            HTTP_AUTHORIZATION=f'Bearer {AccessToken.for_user(user)}')


class ExportLeadsPorteeTests(Crx16Base):
    def setUp(self):
        super().setUp()
        self.lead_mien = Lead.objects.create(
            company=self.company, nom='Lead Mien', owner=self.restreint)
        self.lead_autre = Lead.objects.create(
            company=self.company, nom='Lead Voisin', owner=self.autre)

    def test_role_restreint_n_exporte_que_ses_leads(self):
        self.auth(self.restreint)
        resp = self.api.post(
            '/api/django/crm/leads/export-xlsx/',
            {'ids': [self.lead_mien.pk, self.lead_autre.pk]}, format='json')
        self.assertEqual(resp.status_code, 200, getattr(resp, 'data', resp))
        noms = _lignes_xlsx(resp)
        self.assertIn('Lead Mien', noms)
        self.assertNotIn(
            'Lead Voisin', noms,
            "L'export .xlsx contourne la portée de visibilité du rôle : un "
            'lead hors portée est sorti en clair dans le fichier.')

    def test_role_sans_portee_exporte_tout(self):
        """Non-régression : une portée « all » exporte comme avant."""
        self.auth(self.autre)
        resp = self.api.post(
            '/api/django/crm/leads/export-xlsx/',
            {'ids': [self.lead_mien.pk, self.lead_autre.pk]}, format='json')
        self.assertEqual(resp.status_code, 200, getattr(resp, 'data', resp))
        noms = _lignes_xlsx(resp)
        self.assertIn('Lead Mien', noms)
        self.assertIn('Lead Voisin', noms)

    def test_lead_d_une_autre_societe_reste_hors_export(self):
        autre_societe = Company.objects.create(
            nom='CRX16 Voisine', slug='crx16-voisine')
        etranger = Lead.objects.create(
            company=autre_societe, nom='Lead Étranger')
        self.auth(self.autre)
        resp = self.api.post(
            '/api/django/crm/leads/export-xlsx/',
            {'ids': [self.lead_mien.pk, etranger.pk]}, format='json')
        self.assertEqual(resp.status_code, 200, getattr(resp, 'data', resp))
        self.assertNotIn('Lead Étranger', _lignes_xlsx(resp))


class ExportLeadsBudgetRequetesTests(AssertQueryBudgetMixin, Crx16Base):
    """Le coût de l'export ne grandit PLUS avec le nombre de lignes."""

    def _creer_leads(self, nombre, prefixe):
        from decimal import Decimal
        from apps.ventes.models import Devis, LigneDevis
        ids = []
        for i in range(nombre):
            lead = Lead.objects.create(
                company=self.company, nom=f'{prefixe}{i}', owner=self.autre)
            client = Client.objects.create(
                company=self.company, nom=f'Client {prefixe}{i}')
            # Deux devis par lead : c'est le PLUS RÉCENT qui part à l'export.
            for j in range(2):
                devis = Devis.objects.create(
                    company=self.company, reference=f'{prefixe}{i}-{j}',
                    lead=lead, client=client)
                LigneDevis.objects.create(
                    devis=devis, designation='Panneau',
                    quantite=Decimal('2'), prix_unitaire=Decimal('1000'))
            ids.append(lead.pk)
        return ids

    def test_le_nombre_de_requetes_ne_depend_pas_du_nombre_de_leads(self):
        self.auth(self.autre)
        petits = self._creer_leads(3, 'P')
        grands = self._creer_leads(9, 'G')
        url = '/api/django/crm/leads/export-xlsx/'

        # Chauffe (contenttypes, permissions… mis en cache) pour que la
        # comparaison ne porte que sur le coût des lignes exportées.
        self.assertEqual(
            self.api.post(url, {'ids': petits}, format='json').status_code,
            200)

        with self.assertMaxQueries(40) as petit_ctx:
            resp = self.api.post(url, {'ids': petits}, format='json')
            self.assertEqual(resp.status_code, 200)
        with self.assertMaxQueries(40) as grand_ctx:
            resp = self.api.post(url, {'ids': grands}, format='json')
            self.assertEqual(resp.status_code, 200)
            self.assertEqual(len(_lignes_xlsx(resp)), 9)

        self.assertEqual(
            len(petit_ctx.captured_queries), len(grand_ctx.captured_queries),
            'Le coût de l\'export grandit avec le nombre de leads (N+1) : '
            'le prefetch ORDONNÉ du dernier devis a été perdu.')


class RechercheEntreprisePorteeTests(Crx16Base):
    """QC1 — l'autocomplete respecte la portée de rôle (CRX16)."""

    def setUp(self):
        super().setUp()
        # Un client est « visible » d'un rôle restreint via un document ou un
        # lead visible (cf. ``core.scoping.scope_client_queryset``).
        self.client_visible = Client.objects.create(
            company=self.company, nom='Zenith Visible')
        self.client_cache = Client.objects.create(
            company=self.company, nom='Zenith Cache')
        Lead.objects.create(
            company=self.company, nom='Rattachement visible',
            owner=self.restreint, client=self.client_visible)
        Lead.objects.create(
            company=self.company, nom='Rattachement voisin',
            owner=self.autre, client=self.client_cache)

    def _noms(self, user):
        self.auth(user)
        resp = self.api.get('/api/django/crm/clients/search/', {'q': 'Zenith'})
        self.assertEqual(resp.status_code, 200, getattr(resp, 'data', resp))
        return [h['nom'] for h in resp.data['results']]

    def test_role_restreint_ne_voit_que_ses_clients(self):
        noms = self._noms(self.restreint)
        self.assertIn('Zenith Visible', noms)
        self.assertNotIn(
            'Zenith Cache', noms,
            "L'autocomplete entreprise sert des coordonnées hors portée du "
            'rôle (nom/adresse/téléphone/email).')

    def test_role_sans_portee_voit_tout(self):
        noms = self._noms(self.autre)
        self.assertIn('Zenith Visible', noms)
        self.assertIn('Zenith Cache', noms)

    def test_provider_externe_garde_sa_signature(self):
        """Le seam QC2 n'est PAS cassé : un provider ``(company, q, *, limit)``
        est toujours appelé sans ``user``."""
        from .company_search import search_companies
        appels = []

        def faux_provider(company, q, *, limit):
            appels.append((company, q, limit))
            return [{'source': 'registre', 'nom': q}]

        out = search_companies(None, 'acme', provider=faux_provider,
                               user=self.restreint)
        self.assertEqual(len(appels), 1)
        self.assertEqual(out[0]['nom'], 'acme')
