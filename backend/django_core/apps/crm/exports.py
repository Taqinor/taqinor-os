"""Export Excel (.xlsx) — leads CRM.

`openpyxl` est une dépendance pré-approuvée (voir docs/PLAN.md). L'import est
fait À LA DEMANDE dans la fonction pour que l'app démarre même si la lib n'est
pas présente dans un contexte qui n'exporte jamais.

Le helper `build_xlsx_response` est volontairement générique (en-têtes + lignes)
pour être réutilisé par les autres exports à venir (clients, produits…).
"""
from django.http import HttpResponse

from .stages import STAGE_LABELS

# Libellés FR des canaux/priorités — alignés sur le modèle Lead. On lit les
# choices du modèle quand c'est possible pour ne jamais diverger.


def build_xlsx_response(filename, headers, rows, sheet_title='Export'):
    """Construit une réponse HTTP .xlsx à partir d'en-têtes + lignes.

    headers : liste de chaînes (1re ligne, en gras).
    rows    : liste de listes (valeurs déjà mises en forme, str/num/None).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] or 'Export'

    ws.append(list(headers))
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold

    for row in rows:
        ws.append(['' if v is None else v for v in row])

    # Largeurs lisibles : on borne à la valeur la plus longue par colonne.
    for idx, _ in enumerate(headers, start=1):
        longest = len(str(headers[idx - 1]))
        for row in rows:
            if idx - 1 < len(row) and row[idx - 1] is not None:
                longest = max(longest, len(str(row[idx - 1])))
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = \
            min(max(longest + 2, 10), 50)

    response = HttpResponse(
        content_type=(
            'application/vnd.openxmlformats-officedocument.'
            'spreadsheetml.sheet'
        ),
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


LEAD_EXPORT_HEADERS = [
    'Nom', 'Prénom', 'Société', 'Email', 'Téléphone', 'WhatsApp',
    'Ville', 'Étape', 'Canal', 'Priorité', 'Responsable', 'Tags',
    'Perdu', 'Motif de perte', 'Relance', 'Archivé', 'Créé le',
    'Modifié le', 'Dernier devis (TTC)', 'Statut devis',
]

_PRIORITE_LABELS = {'basse': 'Basse', 'normale': 'Normale', 'haute': 'Haute'}

# Libellés FR des statuts de devis (miroir d'apps.ventes — gardé local pour
# que l'export ne dépende pas du module ventes au chargement).
_DEVIS_STATUT_LABELS = {
    'brouillon': 'Brouillon', 'envoye': 'Envoyé', 'accepte': 'Accepté',
    'refuse': 'Refusé', 'expire': 'Expiré',
}


def lead_row(lead):
    """Une ligne d'export pour un lead (valeurs FR lisibles)."""
    canal_label = dict(
        lead._meta.get_field('canal').choices or []
    ).get(lead.canal, lead.canal or '')
    # Dernier devis du lead (le plus récent) — montant TTC + statut. Aucun prix
    # d'achat ni marge : uniquement le total client-facing.
    dernier = lead.devis.order_by('-date_creation').first()
    devis_total = str(dernier.total_ttc) if dernier else ''
    devis_statut = (
        _DEVIS_STATUT_LABELS.get(dernier.statut, dernier.statut)
        if dernier else '')
    return [
        lead.nom or '',
        lead.prenom or '',
        lead.societe or '',
        lead.email or '',
        lead.telephone or '',
        lead.whatsapp or '',
        lead.ville or '',
        STAGE_LABELS.get(lead.stage, lead.stage),
        canal_label,
        _PRIORITE_LABELS.get(lead.priorite, lead.priorite or ''),
        getattr(lead.owner, 'username', '') if lead.owner_id else '',
        lead.tags or '',
        'Oui' if lead.perdu else 'Non',
        lead.motif_perte or '',
        lead.relance_date.isoformat() if lead.relance_date else '',
        'Oui' if lead.is_archived else 'Non',
        lead.date_creation.strftime('%Y-%m-%d %H:%M') if lead.date_creation else '',
        lead.date_modification.strftime('%Y-%m-%d %H:%M')
        if lead.date_modification else '',
        devis_total,
        devis_statut,
    ]


def export_leads_xlsx(leads):
    """Réponse .xlsx pour une sélection de leads."""
    rows = [lead_row(lead) for lead in leads]
    return build_xlsx_response(
        'leads.xlsx', LEAD_EXPORT_HEADERS, rows, sheet_title='Leads')


_TYPE_CLIENT_LABELS = {'particulier': 'Particulier', 'entreprise': 'Entreprise'}

CLIENT_EXPORT_HEADERS = [
    'Nom', 'Prénom', 'Type', 'Email', 'Téléphone', 'Adresse',
    'ICE', 'IF', 'RC', 'CIN', 'RIB', 'Créé le',
]


def export_clients_xlsx(clients):
    """Export .xlsx clients avec les identifiants légaux marocains (ICE/IF/RC/
    CIN/RIB). Aucun prix d'achat ni marge — données d'identité uniquement.
    RIB lu via getattr (champ optionnel/à venir) → vide s'il n'existe pas."""
    rows = [[
        c.nom or '', c.prenom or '',
        _TYPE_CLIENT_LABELS.get(
            getattr(c, 'type_client', ''), getattr(c, 'type_client', '') or ''),
        c.email or '', c.telephone or '', c.adresse or '',
        getattr(c, 'ice', '') or '', getattr(c, 'if_fiscal', '') or '',
        getattr(c, 'rc', '') or '', getattr(c, 'cin', '') or '',
        getattr(c, 'rib', '') or '',
        c.date_creation.strftime('%Y-%m-%d') if getattr(c, 'date_creation', None) else '',
    ] for c in clients]
    return build_xlsx_response(
        'clients.xlsx', CLIENT_EXPORT_HEADERS, rows, sheet_title='Clients')
