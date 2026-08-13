"""NTCRM28 — Export CSV/XLSX du classement de défi.

Critère d'acceptation : le fichier exporté contient le même classement que
l'endpoint JSON, colonnes rang/nom/score.
"""
import datetime
import io

import openpyxl
from django.contrib.auth import get_user_model
from django.test import TestCase
from django.utils import timezone
from rest_framework.test import APIClient

from authentication.models import Company
from apps.crm.exports import export_defi_classement_xlsx
from apps.crm.models import Defi, Lead
from apps.crm.selectors import classement_defi
from apps.roles.models import Role

User = get_user_model()


class ExportDefiClassementTests(TestCase):
    def setUp(self):
        self.company = Company.objects.create(nom='Taqinor NTCRM28', slug='taqinor-ntcrm28')
        self.role = Role.objects.create(
            company=self.company, nom='Commercial', permissions=['crm_creer'])
        self.com1 = User.objects.create_user(
            username='com1_ntcrm28', password='x', company=self.company, role=self.role)
        self.com2 = User.objects.create_user(
            username='com2_ntcrm28', password='x', company=self.company, role=self.role)
        today = timezone.now().date()
        self.defi = Defi.objects.create(
            company=self.company, nom='Défi export',
            periode_debut=today - datetime.timedelta(days=5),
            periode_fin=today + datetime.timedelta(days=5),
            metrique='nb_leads')
        for _ in range(3):
            Lead.objects.create(company=self.company, owner=self.com1, nom='L')
        Lead.objects.create(company=self.company, owner=self.com2, nom='L')

    def test_export_contient_le_meme_classement_que_le_json(self):
        classement = classement_defi(self.defi)
        resp = export_defi_classement_xlsx(self.defi, classement)
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), len(classement))
        self.assertEqual(rows[0], (1, self.com1.get_username(), '3'))
        self.assertEqual(rows[1], (2, self.com2.get_username(), '1'))

    def test_endpoint_export_xlsx(self):
        api = APIClient()
        api.force_authenticate(self.com1)
        resp = api.get(f'/api/django/crm/defis/{self.defi.pk}/export-xlsx/')
        self.assertEqual(resp.status_code, 200)
        self.assertIn('spreadsheetml', resp['Content-Type'])
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 2)
