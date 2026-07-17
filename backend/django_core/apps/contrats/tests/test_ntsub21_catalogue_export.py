"""Tests NTSUB21 — Export .xlsx du catalogue d'offres (plans/add-ons/paliers)."""
import io
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import TestCase
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import AccessToken

from authentication.models import Company

from apps.contrats.models import (
    AddOnAbonnement,
    PalierUsage,
    PlanAbonnement,
    PlanRecurrent,
)

User = get_user_model()

EXPORT = "/api/django/contrats/plans-abonnement/export/"


def make_company(slug, nom):
    company, _ = Company.objects.get_or_create(slug=slug, defaults={"nom": nom})
    return company


def make_user(company, username):
    return User.objects.create_user(
        username=username, password="x", company=company, role_legacy="admin")


def auth(user):
    api = APIClient()
    api.credentials(HTTP_AUTHORIZATION=f"Bearer {AccessToken.for_user(user)}")
    return api


class CatalogueExportTests(TestCase):
    def setUp(self):
        self.co = make_company("ntsub21", "Ntsub21")
        self.admin = make_user(self.co, "ntsub21-admin")
        pr = PlanRecurrent.objects.create(
            company=self.co, nom="Mensuel",
            unite=PlanRecurrent.Unite.MENSUEL, intervalle=1)
        self.plan = PlanAbonnement.objects.create(
            company=self.co, code="PRO", nom="Offre Pro",
            plan_recurrent=pr, prix_base=Decimal("800"))
        self.addon = AddOnAbonnement.objects.create(
            company=self.co, code="SUP", nom="Supervision",
            prix_unitaire=Decimal("150"), plan_abonnement=self.plan)
        PalierUsage.objects.create(
            company=self.co, addon=self.addon, seuil_min=Decimal("0"),
            seuil_max=Decimal("100"), prix_unitaire=Decimal("2"))

    def test_export_ok_trois_feuilles(self):
        api = auth(self.admin)
        res = api.get(EXPORT)
        self.assertEqual(res.status_code, 200, res.content)
        self.assertIn('spreadsheetml', res['Content-Type'])

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(res.content))
        self.assertEqual(
            set(wb.sheetnames), {'Plans', 'Add-ons', 'Paliers'})
        # Le plan PRO figure sur la feuille Plans.
        plans_vals = [c.value for row in wb['Plans'].iter_rows()
                      for c in row]
        self.assertIn('PRO', plans_vals)

    def test_export_scope_societe(self):
        autre = make_company("ntsub21-a", "Ntsub21A")
        pr = PlanRecurrent.objects.create(
            company=autre, nom="M", unite=PlanRecurrent.Unite.MENSUEL,
            intervalle=1)
        PlanAbonnement.objects.create(
            company=autre, code="AUTRE", nom="Autre",
            plan_recurrent=pr, prix_base=Decimal("1"))
        api = auth(self.admin)
        res = api.get(EXPORT)
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(res.content))
        plans_vals = [c.value for row in wb['Plans'].iter_rows()
                      for c in row]
        self.assertNotIn('AUTRE', plans_vals)
