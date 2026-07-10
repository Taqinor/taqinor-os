"""T12 — export comptable : journal des ventes + résumé TVA (.xlsx).

Deux feuilles : (1) Journal des ventes = une ligne par ligne de facture émise
sur la période, avec sa TVA par ligne ; (2) Résumé TVA = HT/TVA/TTC répartis
par taux (10 % / 20 %…), réconciliés au centime, + totaux. Lecture seule,
borné à la société. openpyxl (pré-approuvé). Groundwork DGI (per-ligne + ICE).
"""
from datetime import date
from decimal import Decimal, ROUND_HALF_UP

from django.http import HttpResponse

# Statuts « émis » (comptablement sortis) — jamais brouillon ni annulée.
ISSUED_STATUTS = ('emise', 'payee', 'en_retard')


def _q2(d):
    return Decimal(d).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)


def period_bounds(params):
    """Calcule (debut, fin) depuis ?month=YYYY-MM, ?quarter=YYYY-Q ou
    ?start=&end=. Défaut : mois courant."""
    month = params.get('month')
    quarter = params.get('quarter')
    start = params.get('start')
    end = params.get('end')
    if start and end:
        return date.fromisoformat(start), date.fromisoformat(end)
    if month:
        y, m = (int(x) for x in month.split('-'))
        debut = date(y, m, 1)
        fin = date(y + (m == 12), (m % 12) + 1, 1)
        return debut, fin
    if quarter:
        y, q = (int(x) for x in quarter.upper().replace('Q', '').split('-') if x)
        m0 = (q - 1) * 3 + 1
        debut = date(y, m0, 1)
        fin = date(y + (m0 + 3 > 12), ((m0 + 2) % 12) + 1, 1)
        return debut, fin
    today = date.today()
    return date(today.year, today.month, 1), \
        date(today.year + (today.month == 12), (today.month % 12) + 1, 1)


def export_journal_ventes(company, debut, fin):
    """Construit le classeur .xlsx (journal + résumé TVA) pour [debut, fin[."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from apps.ventes.models import Avoir, Facture

    factures = (Facture.objects
                .filter(company=company, statut__in=ISSUED_STATUTS,
                        date_emission__gte=debut, date_emission__lt=fin)
                .select_related('client').prefetch_related('lignes')
                .order_by('date_emission', 'reference'))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Journal des ventes'
    bold = Font(bold=True)
    headers = ['Facture', 'Date', 'Type', 'Client', 'ICE client',
               'Désignation', 'Qté', 'P.U. HT', 'Total HT', 'TVA %',
               'Montant TVA', 'Total TTC']
    ws.append(headers)
    for c in ws[1]:
        c.font = bold

    par_taux = {}
    tot_ht = tot_tva = tot_ttc = Decimal('0')
    for f in factures:
        type_libelle = f.get_type_facture_display()
        for ligne in f.lignes.all():
            ht = _q2(ligne.total_ht)
            taux = Decimal(ligne.taux_tva_effectif or 0)
            tva = _q2(ht * taux / Decimal('100'))
            ttc = _q2(ht + tva)
            ws.append([
                f.reference,
                f.date_emission.isoformat() if f.date_emission else '',
                type_libelle,
                getattr(f.client, 'nom', '') or '',
                getattr(f.client, 'ice', '') or '',
                ligne.designation, str(ligne.quantite), str(ligne.prix_unitaire),
                float(ht), float(taux), float(tva), float(ttc),
            ])
            bucket = par_taux.setdefault(taux, {'ht': Decimal('0'), 'tva': Decimal('0')})
            bucket['ht'] += ht
            bucket['tva'] += tva
            tot_ht += ht
            tot_tva += tva
            tot_ttc += ttc

    # Avoirs (notes de crédit) émis sur la période : lignes NÉGATIVES pour
    # réconcilier le CA. Ventilés par taux (10/20) comme les factures, et
    # déduits du résumé TVA. Les avoirs annulés sont exclus.
    avoirs = (Avoir.objects
              .filter(company=company, statut=Avoir.Statut.EMISE,
                      date_emission__gte=debut, date_emission__lt=fin)
              .select_related('client', 'facture').prefetch_related('lignes')
              .order_by('date_emission', 'reference'))
    for a in avoirs:
        date_a = a.date_emission.isoformat() if a.date_emission else ''
        nom = getattr(a.client, 'nom', '') or ''
        ice = getattr(a.client, 'ice', '') or ''
        desig = f'Avoir sur {a.facture.reference}' if a.facture_id else 'Avoir'
        for b in a.tva_par_taux:
            taux = Decimal(b['taux'] or 0)
            ht = -_q2(b['base_ht'])
            tva = -_q2(b['montant'])
            ttc = _q2(ht + tva)
            ws.append([
                a.reference, date_a, 'Avoir', nom, ice,
                desig, '', '',
                float(ht), float(taux), float(tva), float(ttc),
            ])
            bucket = par_taux.setdefault(taux, {'ht': Decimal('0'), 'tva': Decimal('0')})
            bucket['ht'] += ht
            bucket['tva'] += tva
            tot_ht += ht
            tot_tva += tva
            tot_ttc += ttc

    # Feuille résumé TVA.
    ws2 = wb.create_sheet('Résumé TVA')
    ws2.append(['Taux TVA', 'Base HT', 'Montant TVA', 'Total TTC'])
    for c in ws2[1]:
        c.font = bold
    for taux in sorted(par_taux):
        b = par_taux[taux]
        ws2.append([f'{taux} %', float(_q2(b['ht'])), float(_q2(b['tva'])),
                    float(_q2(b['ht'] + b['tva']))])
    ws2.append([])
    total_row = ['TOTAL', float(_q2(tot_ht)), float(_q2(tot_tva)), float(_q2(tot_ttc))]
    ws2.append(total_row)
    for c in ws2[ws2.max_row]:
        c.font = bold

    resp = HttpResponse(content_type=(
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))
    resp['Content-Disposition'] = (
        f'attachment; filename="journal-ventes-{debut.isoformat()}.xlsx"')
    wb.save(resp)
    return resp


# ── Export comptable DGI (groundwork) : factures VALIDÉES d'une plage from→to ──
# Ventilation TVA par ligne, ICE client, totaux HT/TVA/TTC. Excel ET CSV.
# Lecture seule, borné société, AUCUNE transmission. (Distinct du journal :
# focalisé sur les factures validées, sans avoirs.)

_COMPTA_HEADERS = [
    'Facture', 'Date', 'Type', 'Client', 'ICE client', 'Désignation',
    'Qté', 'P.U. HT', 'Total HT', 'TVA %', 'Montant TVA', 'Total TTC',
]


def _compta_rows(company, debut, fin):
    """Lignes du export comptable + totaux, pour [debut, fin[.

    Renvoie (rows, totals) où rows = liste de listes (ordre _COMPTA_HEADERS)
    et totals = (tot_ht, tot_tva, tot_ttc) en Decimal.
    """
    from apps.ventes.models import Facture
    factures = (Facture.objects
                .filter(company=company, statut__in=ISSUED_STATUTS,
                        date_emission__gte=debut, date_emission__lt=fin)
                .select_related('client').prefetch_related('lignes')
                .order_by('date_emission', 'reference'))
    rows = []
    tot_ht = tot_tva = tot_ttc = Decimal('0')
    for f in factures:
        type_libelle = f.get_type_facture_display()
        nom = getattr(f.client, 'nom', '') or ''
        ice = getattr(f.client, 'ice', '') or ''
        date_f = f.date_emission.isoformat() if f.date_emission else ''
        lignes = list(f.lignes.all())
        if lignes:
            for ligne in lignes:
                ht = _q2(ligne.total_ht)
                taux = Decimal(ligne.taux_tva_effectif or 0)
                tva = _q2(ht * taux / Decimal('100'))
                ttc = _q2(ht + tva)
                rows.append([
                    f.reference, date_f, type_libelle, nom, ice,
                    ligne.designation, str(ligne.quantite),
                    str(ligne.prix_unitaire),
                    float(ht), float(taux), float(tva), float(ttc),
                ])
                tot_ht += ht
                tot_tva += tva
                tot_ttc += ttc
        else:
            # Facture de tranche sans lignes : montants figés (un seul taux).
            ht = _q2(f.total_ht)
            taux = Decimal(f.taux_tva or 0)
            tva = _q2(f.total_tva)
            ttc = _q2(f.total_ttc)
            rows.append([
                f.reference, date_f, type_libelle, nom, ice,
                type_libelle or 'Facture', '', '',
                float(ht), float(taux), float(tva), float(ttc),
            ])
            tot_ht += ht
            tot_tva += tva
            tot_ttc += ttc
    return rows, (tot_ht, tot_tva, tot_ttc)


def export_comptable_xlsx(company, debut, fin):
    """Classeur .xlsx : une ligne par ligne de facture validée + ligne TOTAL."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    rows, (tot_ht, tot_tva, tot_ttc) = _compta_rows(company, debut, fin)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Export comptable'
    bold = Font(bold=True)
    ws.append(_COMPTA_HEADERS)
    for c in ws[1]:
        c.font = bold
    for r in rows:
        ws.append(r)
    ws.append([])
    total_row = ['TOTAL', '', '', '', '', '', '', '',
                 float(_q2(tot_ht)), '', float(_q2(tot_tva)), float(_q2(tot_ttc))]
    ws.append(total_row)
    for c in ws[ws.max_row]:
        c.font = bold

    resp = HttpResponse(content_type=(
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))
    resp['Content-Disposition'] = (
        f'attachment; filename="export-comptable-{debut.isoformat()}'
        f'_{fin.isoformat()}.xlsx"')
    wb.save(resp)
    return resp


def export_comptable_csv(company, debut, fin):
    """Fichier .csv : mêmes colonnes que l'export .xlsx + ligne TOTAL."""
    import csv
    import io

    rows, (tot_ht, tot_tva, tot_ttc) = _compta_rows(company, debut, fin)
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=';')
    writer.writerow(_COMPTA_HEADERS)
    for r in rows:
        writer.writerow(r)
    writer.writerow([])
    writer.writerow(['TOTAL', '', '', '', '', '', '', '',
                     float(_q2(tot_ht)), '', float(_q2(tot_tva)),
                     float(_q2(tot_ttc))])
    # BOM pour qu'Excel lise correctement les accents (UTF-8).
    resp = HttpResponse('﻿' + buf.getvalue(),
                        content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = (
        f'attachment; filename="export-comptable-{debut.isoformat()}'
        f'_{fin.isoformat()}.csv"')
    return resp


# ── FG49 — Export comptable CODÉ PAR COMPTE (grand-livre CGNC / fiduciaire) ──
# Troisième format : un grand-livre des ventes mappé sur des numéros de compte
# CGNC marocains, prêt pour un import direct chez le fiduciaire (mise en page
# type PCG/Sage). Chaque facture/avoir devient un jeu d'écritures ÉQUILIBRÉES :
#   • DÉBIT  3421 Clients .......... TTC
#   • CRÉDIT 7111 Ventes ........... HT  (une ligne par taux de TVA)
#   • CRÉDIT 4455 TVA collectée .... TVA (une ligne par taux non nul)
# Les avoirs inversent débit/crédit. Lecture seule, borné société. C'est un
# export AR/ventes : il ne contient QUE des comptes de produit/TVA/créance —
# jamais d'achat ni de marge.
#
# DECISION (codes par défaut) : 7111 = ventes de marchandises (produits),
# 4455 = État - TVA facturée (collectée), 3421 = clients. Défauts CGNC sains,
# alignés sur le plan semé par apps/compta. CONFIGURABLES : un override global
# par ``settings.VENTES_COMPTA_ACCOUNT_CODES`` (dict) et un override par société
# via ``settings.VENTES_COMPTA_ACCOUNT_CODES_BY_COMPANY`` (clé = slug société).

# Défauts CGNC — surchargeables (voir account_codes_for).
_DEFAULT_ACCOUNT_CODES = {
    'ventes': '7111',         # Ventes de marchandises (produit)
    'tva_collectee': '4455',  # État - TVA facturée (collectée)
    'clients': '3421',        # Clients (créance / tiers)
}

# Intitulés CGNC usuels des comptes par défaut (mise en page fiduciaire).
_ACCOUNT_LABELS = {
    '7111': 'Ventes de marchandises',
    '7121': 'Ventes de biens et services produits',
    '4455': 'État - TVA facturée',
    '3421': 'Clients',
}

_GRAND_LIVRE_HEADERS = [
    'Compte', 'Intitulé', 'Date', 'Journal', 'Pièce', 'Libellé',
    'Tiers', 'ICE tiers', 'TVA %', 'Débit', 'Crédit',
]


def account_codes_for(company):
    """Renvoie le mapping seau → numéro de compte CGNC pour ``company``.

    Part des défauts CGNC sains (7111/4455/3421), puis applique, dans l'ordre :
    un override global ``settings.VENTES_COMPTA_ACCOUNT_CODES`` puis un override
    par société ``settings.VENTES_COMPTA_ACCOUNT_CODES_BY_COMPANY[slug]``. Seules
    les clés connues (ventes/tva_collectee/clients) sont prises en compte.
    """
    from django.conf import settings
    codes = dict(_DEFAULT_ACCOUNT_CODES)
    glob = getattr(settings, 'VENTES_COMPTA_ACCOUNT_CODES', None) or {}
    for k in codes:
        if glob.get(k):
            codes[k] = str(glob[k])
    by_company = (
        getattr(settings, 'VENTES_COMPTA_ACCOUNT_CODES_BY_COMPANY', None) or {})
    slug = getattr(company, 'slug', None)
    per = by_company.get(slug) if slug else None
    if isinstance(per, dict):
        for k in codes:
            if per.get(k):
                codes[k] = str(per[k])
    return codes


def _account_label(code):
    return _ACCOUNT_LABELS.get(str(code), '')


def _grand_livre_rows(company, debut, fin):
    """Construit les écritures du grand-livre des ventes pour [debut, fin[.

    Renvoie (rows, totals) où rows = listes (ordre _GRAND_LIVRE_HEADERS) et
    totals = (tot_debit, tot_credit) en Decimal. Factures (signe +) et avoirs
    (signe -) émis sur la période ; ventilation HT par taux, TVA par taux.
    """
    from apps.ventes.models import Avoir, Facture
    codes = account_codes_for(company)
    c_clients = codes['clients']
    c_ventes = codes['ventes']
    c_tva = codes['tva_collectee']
    journal = 'VTE'
    rows = []
    tot_debit = tot_credit = Decimal('0')

    def _emit(debit, credit, compte, taux, date_s, piece, libelle,
              tiers, ice):
        nonlocal tot_debit, tot_credit
        debit = _q2(debit)
        credit = _q2(credit)
        rows.append([
            compte, _account_label(compte), date_s, journal, piece, libelle,
            tiers, ice, (float(taux) if taux is not None else ''),
            float(debit), float(credit),
        ])
        tot_debit += debit
        tot_credit += credit

    factures = (Facture.objects
                .filter(company=company, statut__in=ISSUED_STATUTS,
                        date_emission__gte=debut, date_emission__lt=fin)
                .select_related('client').prefetch_related('lignes')
                .order_by('date_emission', 'reference'))
    for f in factures:
        date_s = f.date_emission.isoformat() if f.date_emission else ''
        nom = getattr(f.client, 'nom', '') or ''
        ice = getattr(f.client, 'ice', '') or ''
        piece = f.reference
        # Ventilation HT/TVA par taux sur les lignes (ou montants figés).
        par_taux = {}
        ttc_total = Decimal('0')
        lignes = list(f.lignes.all())
        if lignes:
            for ligne in lignes:
                ht = _q2(ligne.total_ht)
                taux = Decimal(ligne.taux_tva_effectif or 0)
                tva = _q2(ht * taux / Decimal('100'))
                b = par_taux.setdefault(
                    taux, {'ht': Decimal('0'), 'tva': Decimal('0')})
                b['ht'] += ht
                b['tva'] += tva
                ttc_total += ht + tva
        else:
            ht = _q2(f.total_ht)
            taux = Decimal(f.taux_tva or 0)
            tva = _q2(f.total_tva)
            par_taux[taux] = {'ht': ht, 'tva': tva}
            ttc_total = _q2(f.total_ttc)
        # DÉBIT 3421 Clients (TTC) — une ligne par pièce.
        _emit(ttc_total, Decimal('0'), c_clients, None, date_s, piece,
              f'Facture {piece}', nom, ice)
        # CRÉDIT 7111 Ventes (HT) + 4455 TVA (par taux).
        for taux in sorted(par_taux):
            b = par_taux[taux]
            if b['ht']:
                _emit(Decimal('0'), b['ht'], c_ventes, taux, date_s, piece,
                      f'Vente {piece}', nom, ice)
            if b['tva']:
                _emit(Decimal('0'), b['tva'], c_tva, taux, date_s, piece,
                      f'TVA {piece}', nom, ice)

    avoirs = (Avoir.objects
              .filter(company=company, statut=Avoir.Statut.EMISE,
                      date_emission__gte=debut, date_emission__lt=fin)
              .select_related('client', 'facture').prefetch_related('lignes')
              .order_by('date_emission', 'reference'))
    for a in avoirs:
        date_s = a.date_emission.isoformat() if a.date_emission else ''
        nom = getattr(a.client, 'nom', '') or ''
        ice = getattr(a.client, 'ice', '') or ''
        piece = a.reference
        ttc_total = Decimal('0')
        # Avoir = contre-passation : CRÉDIT 3421 Clients (TTC), DÉBIT
        # 7111 Ventes (HT) + 4455 TVA (par taux).
        for b in a.tva_par_taux:
            taux = Decimal(b['taux'] or 0)
            ht = _q2(b['base_ht'])
            tva = _q2(b['montant'])
            ttc_total += ht + tva
            if ht:
                _emit(ht, Decimal('0'), c_ventes, taux, date_s, piece,
                      f'Avoir {piece} (ventes)', nom, ice)
            if tva:
                _emit(tva, Decimal('0'), c_tva, taux, date_s, piece,
                      f'Avoir {piece} (TVA)', nom, ice)
        _emit(Decimal('0'), ttc_total, c_clients, None, date_s, piece,
              f'Avoir {piece}', nom, ice)

    return rows, (tot_debit, tot_credit)


def export_grand_livre_xlsx(company, debut, fin):
    """Classeur .xlsx : grand-livre codé par compte + ligne TOTAL (débit/crédit)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    rows, (tot_debit, tot_credit) = _grand_livre_rows(company, debut, fin)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Grand-livre ventes'
    bold = Font(bold=True)
    ws.append(_GRAND_LIVRE_HEADERS)
    for c in ws[1]:
        c.font = bold
    for r in rows:
        ws.append(r)
    ws.append([])
    total_row = ['TOTAL', '', '', '', '', '', '', '', '',
                 float(_q2(tot_debit)), float(_q2(tot_credit))]
    ws.append(total_row)
    for c in ws[ws.max_row]:
        c.font = bold

    resp = HttpResponse(content_type=(
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))
    resp['Content-Disposition'] = (
        f'attachment; filename="grand-livre-ventes-{debut.isoformat()}'
        f'_{fin.isoformat()}.xlsx"')
    wb.save(resp)
    return resp


def export_grand_livre_csv(company, debut, fin):
    """Fichier .csv : grand-livre codé par compte (mise en page fiduciaire)."""
    import csv
    import io

    rows, (tot_debit, tot_credit) = _grand_livre_rows(company, debut, fin)
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=';')
    writer.writerow(_GRAND_LIVRE_HEADERS)
    for r in rows:
        writer.writerow(r)
    writer.writerow([])
    writer.writerow(['TOTAL', '', '', '', '', '', '', '', '',
                     float(_q2(tot_debit)), float(_q2(tot_credit))])
    resp = HttpResponse('﻿' + buf.getvalue(),
                        content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = (
        f'attachment; filename="grand-livre-ventes-{debut.isoformat()}'
        f'_{fin.isoformat()}.csv"')
    return resp


# ─────────────────────────────────────────────────────────────────────────────
# SCA41 — exports xlsx asynchrones au-delà d'un seuil (pilote de NTPLT29/30).
#
# En dessous du seuil (défaut 2 000 lignes, surchargeable par env), le chemin
# synchrone actuel reste STRICTEMENT inchangé (même octets, UI inchangée).
# Au-delà, l'endpoint bascule sur une tâche Celery (queue `interactive`) qui
# construit le MÊME classeur, l'upload dans MinIO sous une clé préfixée société
# (motif ERR75), et répond 202 + endpoint de statut/téléchargement pré-signé.
#
# Conçu comme le PILOTE ventes que NTPLT29/30 (BackgroundJob générique + heavy
# async exports, nommés) généraliseront — même signature de sortie (bytes +
# filename), pour qu'un futur refactor pointe simplement le task générique sur
# ces builders sans changer la forme de sortie.
# ─────────────────────────────────────────────────────────────────────────────

# Layouts d'export xlsx supportés par la voie asynchrone. `journal` et les deux
# variantes comptables partagent le même driver de lignes (les factures émises
# de la période), donc un seul compteur suffit.
XLSX_LAYOUTS = ('journal', 'comptable', 'grand-livre')

# Mapping layout → (fonction builder synchrone, préfixe de nom de fichier).
_XLSX_BUILDERS = {
    'journal': (export_journal_ventes, 'journal-ventes'),
    'comptable': (export_comptable_xlsx, 'export-comptable'),
    'grand-livre': (export_grand_livre_xlsx, 'grand-livre-ventes'),
}


def export_async_row_threshold():
    """Seuil (nombre de factures de la période) au-delà duquel un export xlsx
    part en tâche Celery plutôt qu'en synchrone. Surcharge via l'env
    ``VENTES_EXPORT_ASYNC_ROW_THRESHOLD`` ; défaut 2 000."""
    from django.conf import settings
    try:
        return int(getattr(
            settings, 'VENTES_EXPORT_ASYNC_ROW_THRESHOLD', 2000) or 2000)
    except (TypeError, ValueError):
        return 2000


def count_export_rows(company, debut, fin):
    """Nombre de factures émises de la période — le driver commun aux trois
    exports xlsx (journal / comptable / grand-livre). Requête COUNT bornée
    société, sans construire le classeur."""
    from apps.ventes.models import Facture
    return (Facture.objects
            .filter(company=company, statut__in=ISSUED_STATUTS,
                    date_emission__gte=debut, date_emission__lt=fin)
            .count())


def build_export_xlsx_bytes(company_id, layout, debut, fin):
    """Construit le classeur xlsx via le MÊME builder synchrone et renvoie
    ``(bytes, filename)``. Réutiliser le builder existant garantit des octets
    STRICTEMENT identiques entre la voie synchrone et la voie asynchrone.

    ``debut``/``fin`` sont des ``datetime.date``. ``layout`` ∈ XLSX_LAYOUTS.
    Résolution société INTERNE (jamais depuis le corps de requête)."""
    from authentication.models import Company
    if layout not in _XLSX_BUILDERS:
        raise ValueError(f'layout xlsx inconnu : {layout!r}')
    company = Company.objects.get(pk=company_id)
    builder, prefix = _XLSX_BUILDERS[layout]
    resp = builder(company, debut, fin)
    # Le nom de fichier reproduit exactement celui posé par le builder
    # synchrone (Content-Disposition), pour un téléchargement identique.
    if layout == 'journal':
        filename = f'{prefix}-{debut.isoformat()}.xlsx'
    else:
        filename = f'{prefix}-{debut.isoformat()}_{fin.isoformat()}.xlsx'
    return bytes(resp.content), filename


def export_object_key(company_id, layout, debut, fin, token):
    """Clé MinIO préfixée société (motif ERR75) pour un export asynchrone. Le
    ``token`` (id de tâche) rend la clé unique par job et empêche toute
    collision inter-tenant même à période/layout identiques."""
    return (f'exports/{company_id}/{layout}/'
            f'{debut.isoformat()}_{fin.isoformat()}_{token}.xlsx')
