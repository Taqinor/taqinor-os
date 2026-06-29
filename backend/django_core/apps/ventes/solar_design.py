"""FG246 / FG247 / FG249 / … / FG257 — Calculs d'ingénierie solaire (conception électrique).

Module PUR (aucune écriture base, aucun effet de bord) regroupant trois
calculateurs réutilisés par l'écran de devis et, à terme, le pont toiture 3D :

* ``string_design`` (FG246) — répartit N panneaux sur les entrées MPPT d'un
  onduleur, vérifie Vmp/Voc des chaînes à FROID contre la fenêtre de tension
  onduleur, et rapporte le ratio DC/AC.
* ``match_inverter`` (FG247) — propose l'onduleur compatible du catalogue pour
  une configuration panneaux donnée, en gardant les mots-clés de classification
  ALIGNÉS sur ``quote_engine/builder.py`` (réseau/injection, hybride, batterie,
  panneau).
* ``optimize_orientation`` (FG249) — balaie inclinaison/azimut autour du site
  via l'intégration PVGIS EXISTANTE (``apps.parametres.pvgis``) pour retourner
  l'orientation optimale ; repli gracieux hors-ligne (la fonction PVGIS retombe
  déjà sur l'hypothèse manuelle, jamais d'exception réseau).

CONTRAINTES :
* Les onduleurs/panneaux du catalogue (``stock.Produit``) ne portent PAS de
  fiche électrique complète (Voc/Vmp/Vmppt/Vmax). On utilise donc des
  paramètres électriques par défaut SENSÉS (silicium cristallin) avec repli sûr,
  en extrayant du nom produit ce qui est extractible (puissance W, kW onduleur).
* Aucun prix d'achat / marge n'apparaît jamais dans une sortie — ce module ne
  manipule que des grandeurs électriques publiques.
* Aucune dépendance pip nouvelle ; PVGIS via le client stdlib existant.
"""
from __future__ import annotations

import math
import re

# ── Paramètres électriques par défaut (module silicium cristallin) ────────────
# Valeurs marché conservatrices pour un panneau PV mono/poly courant. Tout est
# surchargeable par l'appelant via le dict ``module``.
#
# * ``vmp`` / ``voc`` aux conditions STC (25 °C) — référence d'un panneau ~60–72
#   cellules. À défaut de fiche produit, on s'appuie sur ces valeurs et le
#   coefficient de température pour borner la fenêtre.
# * ``temp_coeff_voc`` : coefficient de température du Voc (%/°C), négatif :
#   le Voc MONTE quand il fait FROID (cas dimensionnant pour la borne haute).
DEFAULT_MODULE = {
    "vmp": 34.0,           # tension au point de puissance max (V), STC
    "voc": 41.0,           # tension circuit ouvert (V), STC
    "temp_coeff_voc": -0.27,   # %/°C (négatif)
    "temp_coeff_vmp": -0.35,   # %/°C (négatif) — Vmp chute plus vite
    "puissance_w": 450,    # puissance crête (W) — repli si non lisible du nom
}

# Fenêtre onduleur par défaut (onduleur string résidentiel/commercial typique).
# Surchargeable via ``inverter``.
DEFAULT_INVERTER_WINDOW = {
    "v_min": 90.0,       # tension de démarrage / minimale DC (V)
    "v_max": 600.0,      # tension DC max absolue (V) — ne JAMAIS dépasser
    "v_mppt_min": 120.0,  # bas de la plage MPPT (V)
    "v_mppt_max": 500.0,  # haut de la plage MPPT (V)
    "n_mppt": 2,         # nombre d'entrées MPPT
    "ac_kw": None,       # puissance AC nominale (kW) — pour le ratio DC/AC
}

# Conditions de température de référence pour le calcul à froid / à chaud.
STC_TEMP_C = 25.0          # conditions standard (Voc/Vmp donnés à 25 °C)
DEFAULT_COLD_TEMP_C = -5.0   # température cellule mini de dimensionnement (hiver Maroc montagne)
DEFAULT_HOT_TEMP_C = 70.0    # température cellule maxi (été, module chaud)

# Ratio DC/AC maximal toléré pour considérer un onduleur « assez gros ».
MAX_DC_AC = 1.35

_WATT_RE = re.compile(r"(\d{3,4})\s*(?:wc|w)\b", re.IGNORECASE)
_KW_RE = re.compile(r"(\d+(?:[.,]\d+)?)\s*(?:kw|kva)\b", re.IGNORECASE)
_KWH_RE = re.compile(r"(\d+(?:[.,]\d+)?)\s*kwh\b", re.IGNORECASE)


def parse_watt(text: str):
    """Puissance panneau (W) lue dans un nom/désignation, sinon None."""
    m = _WATT_RE.search(text or "")
    return int(m.group(1)) if m else None


def parse_kw(text: str):
    """Puissance (kW/kVA) lue dans un nom — exclut d'abord les « kWh »."""
    cleaned = _KWH_RE.sub(" ", text or "")
    m = _KW_RE.search(cleaned)
    return float(m.group(1).replace(",", ".")) if m else None


# ── Classification ALIGNÉE sur quote_engine/builder.py (mots-clés identiques) ──
def is_panel(designation: str, produit_nom: str = "") -> bool:
    blob = f"{designation} {produit_nom}".lower()
    return "panneau" in blob or "panneaux" in blob


def is_battery(designation: str) -> bool:
    return "batterie" in (designation or "").lower()


def is_hybrid_inverter(designation: str) -> bool:
    d = (designation or "").lower()
    return "onduleur" in d and "hybride" in d


def is_reseau_inverter(designation: str) -> bool:
    d = (designation or "").lower()
    return "onduleur" in d and (
        "réseau" in d or "reseau" in d or "injection" in d)


def is_any_inverter(designation: str) -> bool:
    return is_hybrid_inverter(designation) or is_reseau_inverter(designation)


def _as_kw(value):
    try:
        v = float(value)
        return v if v > 0 else None
    except (TypeError, ValueError):
        return None


# ═════════════════════════════════════════════════════════════════════════════
# FG246 — Calcul de chaînes (string design) & vérification du ratio DC/AC
# ═════════════════════════════════════════════════════════════════════════════
def _voltage_at_temp(v_stc: float, temp_coeff_pct_per_c: float,
                     cell_temp_c: float) -> float:
    """Tension à ``cell_temp_c`` à partir d'une tension STC (25 °C).

    ``temp_coeff_pct_per_c`` est en %/°C (négatif). À FROID (temp < 25 °C) la
    tension MONTE (coeff négatif × écart négatif = positif).
    """
    delta = cell_temp_c - STC_TEMP_C
    return v_stc * (1.0 + (temp_coeff_pct_per_c / 100.0) * delta)


def string_design(n_panels, module=None, inverter=None,
                  cold_temp_c=DEFAULT_COLD_TEMP_C,
                  hot_temp_c=DEFAULT_HOT_TEMP_C):
    """FG246 — répartit ``n_panels`` panneaux sur les MPPT et vérifie la fenêtre.

    Distribue les panneaux en chaînes série équilibrées sur les ``n_mppt``
    entrées de l'onduleur, choisit une longueur de chaîne qui respecte les
    bornes de tension, puis VÉRIFIE :

    * Voc de la chaîne à FROID (``cold_temp_c``) ≤ ``v_max`` (sécurité absolue).
    * Vmp de la chaîne à FROID dans la plage MPPT haute (``v_mppt_max``).
    * Vmp de la chaîne à CHAUD (``hot_temp_c``) ≥ ``v_mppt_min`` (démarrage MPPT).
    * Vmp à chaud ≥ ``v_min`` (démarrage onduleur).

    Et rapporte le ratio DC/AC (puissance crête DC ÷ puissance AC onduleur).

    Paramètres
    ----------
    n_panels : nombre total de panneaux.
    module : dict de paramètres électriques module (défaut ``DEFAULT_MODULE``).
    inverter : dict de fenêtre onduleur (défaut ``DEFAULT_INVERTER_WINDOW``).
    cold_temp_c / hot_temp_c : températures cellule de dimensionnement (°C).

    Retourne un dict JSON-sérialisable ``{n_panels, n_mppt, strings,
    panels_per_string, dc_kw, ac_kw, dc_ac_ratio, voltages{...}, checks{...},
    ok, warnings[]}``. Ne lève jamais sur des entrées dégradées : sur 0 panneau,
    retourne une structure vide cohérente.
    """
    mod = {**DEFAULT_MODULE, **(module or {})}
    inv = {**DEFAULT_INVERTER_WINDOW, **(inverter or {})}

    try:
        n = int(n_panels)
    except (TypeError, ValueError):
        n = 0
    n_mppt = max(1, int(inv.get("n_mppt") or 1))

    warnings = []

    vmp_stc = float(mod["vmp"])
    voc_stc = float(mod["voc"])
    panel_w = float(mod.get("puissance_w") or DEFAULT_MODULE["puissance_w"])

    v_max = float(inv["v_max"])
    v_min = float(inv["v_min"])
    v_mppt_min = float(inv["v_mppt_min"])
    v_mppt_max = float(inv["v_mppt_max"])

    # Tensions unitaires aux températures de dimensionnement.
    voc_cold = _voltage_at_temp(voc_stc, mod["temp_coeff_voc"], cold_temp_c)
    vmp_cold = _voltage_at_temp(vmp_stc, mod["temp_coeff_vmp"], cold_temp_c)
    vmp_hot = _voltage_at_temp(vmp_stc, mod["temp_coeff_vmp"], hot_temp_c)

    if n <= 0:
        return {
            "n_panels": 0, "n_mppt": n_mppt, "strings": 0,
            "panels_per_string": 0, "string_layout": [],
            "dc_kw": 0.0, "ac_kw": _as_kw(inv.get("ac_kw")),
            "dc_ac_ratio": None,
            "voltages": {}, "checks": {}, "ok": False,
            "warnings": ["aucun panneau à répartir"],
        }

    # ── Longueur de chaîne admissible (bornée par le Voc à froid) ──
    # Max modules avant de dépasser v_max au Voc froid (sécurité absolue).
    max_by_voc = int(math.floor(v_max / voc_cold)) if voc_cold > 0 else n
    # Max modules avant de dépasser le haut de la plage MPPT au Vmp froid.
    max_by_mppt = int(math.floor(v_mppt_max / vmp_cold)) if vmp_cold > 0 else n
    # Min modules pour démarrer le MPPT au Vmp chaud (le pire cas bas).
    min_by_mppt = int(math.ceil(v_mppt_min / vmp_hot)) if vmp_hot > 0 else 1
    min_by_start = int(math.ceil(v_min / vmp_hot)) if vmp_hot > 0 else 1
    min_len = max(1, min_by_mppt, min_by_start)
    max_len = max(1, min(max_by_voc, max_by_mppt))

    window_too_narrow = False
    if max_len < min_len:
        window_too_narrow = True
        warnings.append(
            "fenêtre de tension trop étroite pour ce module : aucune longueur "
            "de chaîne ne respecte à la fois la borne haute (froid) et le "
            "démarrage MPPT (chaud) — vérifier le couple module/onduleur")
        # On garde quand même une répartition « best effort » bornée par v_max.
        max_len = max(1, max_by_voc)
        min_len = 1

    # ── Répartition équilibrée sur les MPPT ──
    # On vise une longueur de chaîne qui partitionne n en chaînes ÉGALES sur les
    # entrées MPPT, dans [min_len, max_len], la plus longue possible.
    panels_per_string, strings = _choose_string_layout(
        n, n_mppt, min_len, max_len)

    uneven = False
    if panels_per_string == 0:
        # Aucun découpage propre — repli : chaînes de longueur bornée.
        uneven = True
        panels_per_string = min(n, max_len)
        strings = int(math.ceil(n / panels_per_string))
        warnings.append(
            "répartition non homogène : les chaînes ne sont pas toutes de "
            "longueur égale (vérifier la conception)")

    # Disposition réelle des chaînes par MPPT (équilibrée).
    string_layout = _distribute_strings(strings, n_mppt)

    # ── Tensions au niveau CHAÎNE ──
    string_voc_cold = round(voc_cold * panels_per_string, 1)
    string_vmp_cold = round(vmp_cold * panels_per_string, 1)
    string_vmp_hot = round(vmp_hot * panels_per_string, 1)
    string_vmp_stc = round(vmp_stc * panels_per_string, 1)

    # ── Vérifications ──
    checks = {
        # Sécurité absolue : Voc froid sous le V_max DC.
        "voc_cold_under_vmax": string_voc_cold <= v_max,
        # Vmp froid sous le haut de plage MPPT.
        "vmp_cold_under_mppt_max": string_vmp_cold <= v_mppt_max,
        # Vmp chaud au-dessus du bas de plage MPPT.
        "vmp_hot_over_mppt_min": string_vmp_hot >= v_mppt_min,
        # Vmp chaud au-dessus du démarrage onduleur.
        "vmp_hot_over_vmin": string_vmp_hot >= v_min,
    }
    if not checks["voc_cold_under_vmax"]:
        warnings.append(
            f"Voc à froid {string_voc_cold} V > V_max onduleur {v_max} V — "
            "chaîne trop longue, RISQUE matériel (réduire le nombre de modules "
            "par chaîne)")
    if not checks["vmp_cold_under_mppt_max"]:
        warnings.append(
            f"Vmp à froid {string_vmp_cold} V > haut de plage MPPT {v_mppt_max} "
            "V — l'onduleur écrête, perte de production")
    if not checks["vmp_hot_over_mppt_min"]:
        warnings.append(
            f"Vmp à chaud {string_vmp_hot} V < bas de plage MPPT {v_mppt_min} V "
            "— chaîne trop courte, MPPT hors plage en été")
    if not checks["vmp_hot_over_vmin"]:
        warnings.append(
            f"Vmp à chaud {string_vmp_hot} V < démarrage onduleur {v_min} V")

    # ── Ratio DC/AC ──
    dc_kw = round(n * panel_w / 1000.0, 3)
    ac_kw = _as_kw(inv.get("ac_kw"))
    dc_ac_ratio = round(dc_kw / ac_kw, 3) if ac_kw and ac_kw > 0 else None
    if dc_ac_ratio is not None:
        if dc_ac_ratio > 1.5:
            warnings.append(
                f"ratio DC/AC {dc_ac_ratio} élevé (> 1.5) — surdimensionnement "
                "DC important, écrêtage probable")
        elif dc_ac_ratio < 1.0:
            warnings.append(
                f"ratio DC/AC {dc_ac_ratio} faible (< 1.0) — onduleur "
                "surdimensionné par rapport au champ PV")

    ok = all(checks.values()) and not uneven and not window_too_narrow

    return {
        "n_panels": n,
        "n_mppt": n_mppt,
        "strings": strings,
        "panels_per_string": panels_per_string,
        "string_layout": string_layout,
        "dc_kw": dc_kw,
        "ac_kw": ac_kw,
        "dc_ac_ratio": dc_ac_ratio,
        "voltages": {
            "voc_cold": string_voc_cold,
            "vmp_cold": string_vmp_cold,
            "vmp_hot": string_vmp_hot,
            "vmp_stc": string_vmp_stc,
            "cold_temp_c": cold_temp_c,
            "hot_temp_c": hot_temp_c,
        },
        "checks": checks,
        "ok": ok,
        "warnings": warnings,
    }


def _choose_string_layout(n, n_mppt, min_len, max_len):
    """Choisit (panels_per_string, strings) : chaînes ÉGALES, longueur valide.

    Cherche une partition de ``n`` en chaînes ÉGALES telle que la longueur de
    chaîne ∈ [min_len, max_len], en privilégiant un usage équilibré des entrées
    MPPT (nombre de chaînes multiple de n_mppt) puis les chaînes les plus
    longues (moins de câblage). Renvoie ``(0, 0)`` si aucune partition égale
    n'existe.
    """
    best = (0, 0)
    best_score = None
    # On essaie tous les nombres de chaînes de 1 à n.
    for strings in range(1, n + 1):
        if n % strings != 0:
            continue
        length = n // strings
        if length < min_len or length > max_len:
            continue
        # Score : préférer un nombre de chaînes multiple de n_mppt (réparti
        # également), puis les chaînes les plus longues (moins de câblage).
        balanced = 0 if strings % n_mppt == 0 else 1
        score = (balanced, -length)
        if best_score is None or score < best_score:
            best_score = score
            best = (length, strings)
    return best


def _distribute_strings(strings, n_mppt):
    """Répartit ``strings`` chaînes sur ``n_mppt`` entrées, le plus égal possible.

    Renvoie une liste de longueur ``n_mppt`` : nombre de chaînes par MPPT.
    """
    base = strings // n_mppt
    extra = strings % n_mppt
    return [base + (1 if i < extra else 0) for i in range(n_mppt)]


# ═════════════════════════════════════════════════════════════════════════════
# FG247 — Appariement module–onduleur depuis le catalogue
# ═════════════════════════════════════════════════════════════════════════════
def match_inverter(produits, *, n_panels, panel_w=None, hybrid=False,
                   module=None, inverter_window=None,
                   cold_temp_c=DEFAULT_COLD_TEMP_C,
                   hot_temp_c=DEFAULT_HOT_TEMP_C):
    """FG247 — propose l'onduleur catalogue compatible pour une config panneaux.

    Parcourt ``produits`` (itérable de ``stock.Produit``), retient les onduleurs
    de la bonne FAMILLE — hybride si ``hybrid`` sinon réseau/injection, mots-clés
    ALIGNÉS sur ``builder.py`` — puis choisit le plus petit onduleur (kW) dont :

    * la puissance AC supporte un ratio DC/AC raisonnable (≤ ``MAX_DC_AC``) ;
    * la fenêtre de tension accepte au moins une longueur de chaîne valide pour
      le module (réutilise ``string_design`` pour la vérif Vmp/Voc à froid).

    Le nom du produit fournit la puissance kW (``parse_kw``) ; faute de fiche
    électrique au catalogue, la fenêtre onduleur reste celle par défaut
    (surchargeable via ``inverter_window``). Aucun prix d'achat n'est lu.

    Retourne un dict ``{inverter, ac_kw, dc_kw, dc_ac_ratio, string_design,
    compatible, candidates_considered, reason}`` ; ``inverter`` est le
    ``Produit`` choisi (ou None si aucun ne convient).
    """
    mod = {**DEFAULT_MODULE, **(module or {})}
    if panel_w:
        try:
            mod["puissance_w"] = float(panel_w)
        except (TypeError, ValueError):
            pass

    try:
        n = int(n_panels)
    except (TypeError, ValueError):
        n = 0

    dc_kw = round(n * float(mod["puissance_w"]) / 1000.0, 3)

    family_pred = is_hybrid_inverter if hybrid else is_reseau_inverter
    # Candidats : bonne famille, puissance lisible, prix de vente réel (jamais
    # un produit « prix à renseigner » — même garde que l'auto-fill).
    candidates = []
    for p in produits:
        nom = getattr(p, "nom", "") or ""
        if not family_pred(nom):
            continue
        kw = parse_kw(nom)
        if kw is None or kw <= 0:
            continue
        prix = getattr(p, "prix_vente", None)
        try:
            priced = prix is not None and float(prix) > 0
        except (TypeError, ValueError):
            priced = False
        if not priced:
            continue
        candidates.append((p, kw))

    # Tri : plus petite puissance d'abord (puis id stable).
    candidates.sort(key=lambda x: (x[1], getattr(x[0], "id", 0) or 0))

    considered = len(candidates)
    chosen = None
    chosen_kw = None
    chosen_design = None
    chosen_ratio = None
    reason = ""

    for p, kw in candidates:
        ratio = (dc_kw / kw) if kw > 0 else None
        # Onduleur trop petit pour le champ (ratio DC/AC excessif) → suivant.
        if ratio is not None and ratio > MAX_DC_AC:
            continue
        window = {**DEFAULT_INVERTER_WINDOW, **(inverter_window or {}),
                  "ac_kw": kw}
        design = string_design(
            n, module=mod, inverter=window,
            cold_temp_c=cold_temp_c, hot_temp_c=hot_temp_c)
        if design["ok"]:
            chosen, chosen_kw = p, kw
            chosen_design, chosen_ratio = design, design["dc_ac_ratio"]
            reason = ("onduleur le plus petit respectant ratio DC/AC et "
                      "fenêtre de tension")
            break

    if chosen is None and candidates:
        # Aucun candidat parfait : on retient le plus gros (meilleur ratio) en
        # le signalant, plutôt que de ne rien proposer.
        p, kw = candidates[-1]
        window = {**DEFAULT_INVERTER_WINDOW, **(inverter_window or {}),
                  "ac_kw": kw}
        chosen_design = string_design(
            n, module=mod, inverter=window,
            cold_temp_c=cold_temp_c, hot_temp_c=hot_temp_c)
        chosen, chosen_kw = p, kw
        chosen_ratio = chosen_design["dc_ac_ratio"]
        reason = ("aucun onduleur catalogue parfaitement compatible — plus "
                  "grosse puissance retenue, vérifier la conception")

    if chosen is None:
        reason = ("aucun onduleur "
                  + ("hybride" if hybrid else "réseau/injection")
                  + " chiffrable au catalogue pour cette configuration")

    return {
        "inverter": chosen,
        "ac_kw": chosen_kw,
        "dc_kw": dc_kw,
        "dc_ac_ratio": chosen_ratio,
        "string_design": chosen_design,
        "compatible": bool(chosen_design and chosen_design.get("ok")),
        "candidates_considered": considered,
        "reason": reason,
    }


# ═════════════════════════════════════════════════════════════════════════════
# FG249 — Optimisation inclinaison / azimut (balayage via PVGIS existant)
# ═════════════════════════════════════════════════════════════════════════════
# Convention d'azimut founder (= PVGIS aspect) : Sud 0 / Est −90 / Ouest +90.
DEFAULT_TILT_RANGE = (0, 60, 5)        # (min, max, pas) en degrés
DEFAULT_AZIMUTH_RANGE = (-90, 90, 15)  # (min, max, pas), Sud=0


def _frange(start, stop, step):
    """Plage inclusive d'entiers de ``start`` à ``stop`` par pas ``step``."""
    if step <= 0:
        return [int(start)]
    vals = []
    v = int(start)
    while v <= int(stop):
        vals.append(v)
        v += int(step)
    if vals and vals[-1] != int(stop):
        vals.append(int(stop))
    return vals


def optimize_orientation(settings, lat, lon, *, peakpower_kwc=1.0,
                         tilt_range=DEFAULT_TILT_RANGE,
                         azimuth_range=DEFAULT_AZIMUTH_RANGE,
                         fetch=None):
    """FG249 — balaie (inclinaison, azimut) → orientation au productible maximal.

    Réutilise l'intégration PVGIS EXISTANTE
    (``apps.parametres.pvgis.fetch_productible``, injectable via ``fetch`` pour
    les tests) : pour chaque couple (tilt, azimuth) de la grille, on demande le
    productible (kWh/kWc/an) et on retient le maximum. La fonction PVGIS retombe
    déjà sur l'hypothèse manuelle hors-ligne (jamais d'exception réseau), donc ce
    balayage fonctionne aussi sans réseau (toutes les cases renvoient alors la
    même valeur manuelle → l'orientation par défaut société est retenue).

    Paramètres
    ----------
    settings : ``TariffSettings`` (défauts inclinaison/azimut + repli manuel).
    lat, lon : coordonnées GPS du site.
    peakpower_kwc : puissance crête pour la requête (1 → kWh/kWc/an direct).
    tilt_range / azimuth_range : ``(min, max, pas)`` en degrés.
    fetch : surcharge de ``fetch_productible`` (signature identique) — pour les
        tests, on injecte un stub qui ne touche pas le réseau.

    Retourne ``{best{tilt, azimuth, productible_kwh_kwc, source}, grid[...],
    evaluated, source, default_orientation{tilt, azimuth}, gain_vs_default_pct}``.
    Ne lève jamais sur PVGIS indisponible.
    """
    if fetch is None:
        from apps.parametres.pvgis import fetch_productible as fetch

    tilts = _frange(*tilt_range)
    azimuths = _frange(*azimuth_range)

    default_tilt = int(getattr(settings, "inclinaison_defaut_deg", 30) or 30)
    default_azimuth = int(getattr(settings, "azimut_defaut_deg", 0) or 0)

    grid = []
    best = None
    default_prod = None
    any_pvgis = False

    for tilt in tilts:
        for az in azimuths:
            res = fetch(settings, lat, lon, peakpower_kwc=peakpower_kwc,
                        tilt=tilt, azimuth=az)
            prod = res.get("productible_kwh_kwc")
            src = res.get("source")
            if src == "pvgis":
                any_pvgis = True
            cell = {
                "tilt": tilt, "azimuth": az,
                "productible_kwh_kwc": prod, "source": src,
            }
            grid.append(cell)
            if prod is not None and (
                    best is None or prod > best["productible_kwh_kwc"]):
                best = cell
            if tilt == default_tilt and az == default_azimuth and prod is not None:
                default_prod = prod

    # Productible à l'orientation par défaut société : on l'interroge à part
    # s'il n'est pas tombé sur une case de la grille (gain de référence).
    if default_prod is None:
        res = fetch(settings, lat, lon, peakpower_kwc=peakpower_kwc,
                    tilt=default_tilt, azimuth=default_azimuth)
        default_prod = res.get("productible_kwh_kwc")
        if res.get("source") == "pvgis":
            any_pvgis = True

    gain_pct = None
    if best and default_prod and default_prod > 0:
        gain_pct = round(
            (best["productible_kwh_kwc"] - default_prod) / default_prod * 100, 1)

    return {
        "best": best,
        "grid": grid,
        "evaluated": len(grid),
        "source": "pvgis" if any_pvgis else "manual",
        "default_orientation": {
            "tilt": default_tilt, "azimuth": default_azimuth,
            "productible_kwh_kwc": default_prod,
        },
        "gain_vs_default_pct": gain_pct,
    }


# ── FG250 — Analyse d'ombrage & profil d'horizon ─────────────────────────────
# Transforme l'ombrage QUALITATIF (obstacles + profil d'horizon) en une PERTE
# d'ombrage CHIFFRÉE, mensuelle. Module PUR : aucune base, aucun réseau.
#
# Modèle simplifié et transparent (pas de tracé de rayon par minute) :
#   * Le profil d'horizon est une liste d'élévations (degrés au-dessus de
#     l'horizontale) par secteur d'azimut. Plus l'horizon est haut au Sud, plus
#     la perte est forte ; l'Est/Ouest pèse moins, le Nord (hémisphère N) ~0.
#   * Des obstacles ponctuels (arbre, mur, cheminée) ajoutent une perte locale
#     pondérée par leur azimut et par la saison (soleil bas en hiver → pertes
#     accrues).
# Le résultat donne une perte mensuelle (%) et une perte annuelle moyenne (%),
# un facteur de production (1 - perte) prêt à multiplier un productible PVGIS.

# Poids saisonnier de l'élévation solaire : en hiver le soleil culmine bas, donc
# un horizon haut masque davantage. Index 0 = janvier … 11 = décembre.
_SHADING_SEASON_WEIGHT = [
    1.35, 1.25, 1.10, 0.95, 0.85, 0.80,
    0.82, 0.90, 1.05, 1.20, 1.32, 1.40,
]

_MONTHS_FR = [
    "janvier", "février", "mars", "avril", "mai", "juin",
    "juillet", "août", "septembre", "octobre", "novembre", "décembre",
]


def _azimuth_solar_weight(azimuth, hemisphere_north=True):
    """Poids de pénalité d'un obstacle selon son azimut (0=N, 90=E, 180=S, 270=O).

    Le soleil utile est au Sud (hémisphère Nord) : un masque plein Sud coûte le
    plus (poids ~1), l'Est/Ouest ~0.5, le Nord ~0 (le soleil n'y passe jamais).
    """
    try:
        az = float(azimuth) % 360.0
    except (TypeError, ValueError):
        return 0.5
    # Azimut solaire de référence : 180° (Sud) au Nord, 0° (Nord) au Sud.
    ref = 180.0 if hemisphere_north else 0.0
    delta = abs((az - ref + 180.0) % 360.0 - 180.0)  # écart angulaire 0..180
    # cos décroît du Sud (0°) vers le Nord (180°) ; borné à [0, 1].
    return max(0.0, math.cos(math.radians(min(delta, 90.0))))


def shading_analysis(horizon_profile=None, obstacles=None, *,
                     hemisphere_north=True):
    """FG250 — perte d'ombrage mensuelle depuis l'horizon + les obstacles.

    Paramètres
    ----------
    horizon_profile : liste de points ``{azimuth, elevation}`` (degrés). Une
        élévation d'horizon ``e`` au secteur d'azimut ``a`` masque le soleil
        bas ; sa pénalité est ``e/90`` pondérée par le poids solaire de ``a``.
    obstacles : liste de masques ponctuels ``{azimuth, elevation, type?}`` —
        même pénalité, additive et bornée.
    hemisphere_north : True (Maroc) → soleil au Sud.

    Retourne un dict JSON-sérialisable ::

        {monthly_loss_pct: [12], annual_loss_pct, production_factor,
         monthly_production_factor: [12], horizon_severity, n_obstacles,
         warnings: []}

    Ne lève jamais : entrées vides → perte 0, facteur 1.0 (aucun ombrage).
    """
    horizon_profile = horizon_profile or []
    obstacles = obstacles or []
    warnings = []

    def _pt_penalty(pt):
        """Pénalité de base (0..1) d'un point d'horizon/obstacle, hors saison."""
        try:
            elev = float(pt.get("elevation") or 0.0)
        except (TypeError, ValueError, AttributeError):
            return 0.0
        if elev <= 0:
            return 0.0
        elev = min(elev, 90.0)
        az_w = _azimuth_solar_weight(pt.get("azimuth"), hemisphere_north)
        # Fraction du ciel utile masquée par cette élévation à cet azimut.
        return (elev / 90.0) * az_w

    # Sévérité de base = moyenne des pénalités d'horizon + somme bornée des
    # obstacles (un obstacle plein Sud à 30° pèse lourd, plusieurs s'ajoutent).
    horizon_pen = 0.0
    if horizon_profile:
        horizon_pen = sum(_pt_penalty(p) for p in horizon_profile) \
            / len(horizon_profile)
    obstacle_pen = sum(_pt_penalty(p) for p in obstacles)

    # Pénalité de référence (annuelle, avant pondération saisonnière), bornée.
    base_pen = min(0.6, horizon_pen + obstacle_pen)

    if base_pen <= 0:
        return {
            "monthly_loss_pct": [0.0] * 12,
            "annual_loss_pct": 0.0,
            "production_factor": 1.0,
            "monthly_production_factor": [1.0] * 12,
            "horizon_severity": 0.0,
            "n_obstacles": len(obstacles),
            "warnings": ["aucun ombrage significatif détecté"]
            if not (horizon_profile or obstacles) else [],
        }

    monthly_loss_pct = []
    monthly_factor = []
    for w in _SHADING_SEASON_WEIGHT:
        # Perte du mois = pénalité de base × poids saisonnier, en %, bornée à 90.
        loss = min(90.0, round(base_pen * w * 100.0, 1))
        monthly_loss_pct.append(loss)
        monthly_factor.append(round(1.0 - loss / 100.0, 4))

    annual_loss = round(sum(monthly_loss_pct) / 12.0, 1)
    production_factor = round(1.0 - annual_loss / 100.0, 4)

    if annual_loss >= 20.0:
        warnings.append(
            "ombrage important (perte annuelle ≥ 20 %) — envisager un "
            "repositionnement des panneaux ou des optimiseurs/micro-onduleurs")
    elif annual_loss >= 8.0:
        warnings.append(
            "ombrage modéré — vérifier la disposition des chaînes pour limiter "
            "l'impact d'un panneau masqué sur sa chaîne")

    return {
        "monthly_loss_pct": monthly_loss_pct,
        "monthly_labels": list(_MONTHS_FR),
        "annual_loss_pct": annual_loss,
        "production_factor": production_factor,
        "monthly_production_factor": monthly_factor,
        "horizon_severity": round(base_pen, 4),
        "n_obstacles": len(obstacles),
        "warnings": warnings,
    }


# ── FG251 — Générateur de nomenclature électrique (BOQ) ───────────────────────
# Déduit du design (nb panneaux, kWc, conception de chaînes, type d'installation)
# une nomenclature électrique : câbles DC/AC, disjoncteurs, parafoudres,
# coffrets, mise à la terre, structure. Module PUR ; aucun prix (le BOQ liste des
# QUANTITÉS et des spécifications, jamais de prix d'achat/marge).

# Section de câble AC (mm²) par tranche de courant — barème prudent (cuivre).
_AC_CABLE_BY_AMP = [
    (16, 2.5), (25, 4.0), (32, 6.0), (40, 10.0),
    (63, 16.0), (80, 25.0), (100, 35.0), (125, 50.0),
]


def _ac_cable_section(amps):
    for amp_max, section in _AC_CABLE_BY_AMP:
        if amps <= amp_max:
            return section
    return 70.0


def generate_boq(*, n_panels=0, kwc=None, string_result=None,
                 installation_type="reseau", phases=1, has_battery=False,
                 ac_cable_length_m=15.0, dc_cable_length_m=None):
    """FG251 — nomenclature électrique (BOQ) déduite du design.

    Paramètres
    ----------
    n_panels : nombre de panneaux.
    kwc : puissance crête DC (kWc) ; déduite de ``string_result.dc_kw`` sinon.
    string_result : sortie de ``string_design`` (strings/panels_per_string/
        ac_kw…) — pilote la longueur DC et le nombre de protections de chaîne.
    installation_type : 'reseau' | 'hybride' | 'autonome' (pompage exclu — pas
        de BOQ PV classique).
    phases : 1 (mono) ou 3 (triphasé) → calibre disjoncteur AC + section câble.
    has_battery : ajoute le câblage/protections batterie (DC).
    ac_cable_length_m / dc_cable_length_m : longueurs estimées (m) ; le DC est
        déduit du nombre de chaînes si non fourni.

    Retourne ``{items: [{categorie, designation, quantite, unite, spec}],
    summary: {...}, warnings: []}``. JSON-sérialisable, jamais de prix.
    """
    warnings = []
    sr = string_result or {}
    if kwc is None:
        kwc = sr.get("dc_kw")
    try:
        kwc = float(kwc or 0.0)
    except (TypeError, ValueError):
        kwc = 0.0
    try:
        n_panels = int(n_panels or sr.get("n_panels") or 0)
    except (TypeError, ValueError):
        n_panels = 0
    strings = int(sr.get("strings") or (1 if n_panels else 0))
    phases = 3 if int(phases or 1) == 3 else 1

    items = []

    def add(categorie, designation, quantite, unite, spec=""):
        items.append({
            "categorie": categorie,
            "designation": designation,
            "quantite": quantite,
            "unite": unite,
            "spec": spec,
        })

    if n_panels <= 0:
        return {
            "items": [],
            "summary": {"kwc": kwc, "n_panels": 0, "strings": 0,
                        "phases": phases},
            "warnings": ["aucun panneau — pas de nomenclature à générer"],
        }

    # ── Courant & calibre AC ──
    # Puissance AC de référence = ac_kw onduleur si connu, sinon kWc / 1.2.
    ac_kw = sr.get("ac_kw") or (kwc / 1.2 if kwc else 0.0)
    voltage = 400.0 if phases == 3 else 230.0
    sqrt3 = math.sqrt(3) if phases == 3 else 1.0
    ac_amps = (ac_kw * 1000.0) / (voltage * sqrt3) if ac_kw else 0.0
    # Calibre disjoncteur AC : 1.25× le courant nominal, arrondi au calibre std.
    breaker_amp = _round_breaker(ac_amps * 1.25)

    # ── Câble DC (chaînes) ──
    if dc_cable_length_m is None:
        # ~2 conducteurs (+/-) par chaîne, longueur estimée par chaîne.
        dc_cable_length_m = max(10.0, strings * 20.0)
    add("Câblage DC", "Câble solaire DC 6 mm² (PV1-F)",
        round(dc_cable_length_m, 1), "m",
        "1000 V DC, double isolation, résistant UV")

    # ── Câble AC ──
    ac_section = _ac_cable_section(max(ac_amps, 1.0))
    n_cond_ac = 5 if phases == 3 else 3  # 3P+N+T ou P+N+T
    add("Câblage AC",
        f"Câble AC {ac_section:g} mm² ({'triphasé' if phases == 3 else 'monophasé'})",
        round(ac_cable_length_m * n_cond_ac, 1), "m",
        f"{n_cond_ac} conducteurs, U-1000 R2V")

    # ── Protections DC ──
    add("Protection DC", "Parafoudre DC Type 2", 1, "u",
        "1000 V DC, pour string box / entrée onduleur")
    add("Protection DC", "Sectionneur-fusible DC par chaîne", max(strings, 1),
        "u", "porte-fusible + fusible gPV 1000 V DC")
    add("Coffret", "Coffret de chaîne DC (string box)",
        1 if strings <= 2 else 2, "u",
        "IP65, presse-étoupes, embase parafoudre")

    # ── Protections AC ──
    add("Protection AC", f"Disjoncteur AC {breaker_amp} A "
        f"{'tétrapolaire' if phases == 3 else 'bipolaire'}", 1, "u",
        f"courbe C, {voltage:g} V")
    add("Protection AC", "Parafoudre AC Type 2", 1, "u",
        f"{'triphasé' if phases == 3 else 'monophasé'}, In 20 kA")
    add("Coffret", "Coffret de protection AC", 1, "u",
        "IP65, prêt à raccorder au tableau")

    # ── Mise à la terre ──
    add("Mise à la terre", "Piquet de terre + barrette de coupure", 1, "ens",
        "≤ 100 Ω, conforme NF C 15-100")
    add("Mise à la terre", "Câble de terre cuivre nu 25 mm²",
        round(dc_cable_length_m * 0.6 + ac_cable_length_m, 1), "m",
        "liaison équipotentielle structure + masses")

    # ── Structure ──
    add("Structure", "Rail de fixation aluminium", n_panels * 2, "u",
        "rail anodisé, longueur ajustée au module")
    add("Structure", "Pince de fixation (milieu + extrémité)",
        n_panels * 2 + 4, "u", "inox A2, milieu et extrémité")
    add("Structure", "Crochet / patte de fixation toiture",
        max(4, int(math.ceil(n_panels * 0.6))), "u",
        "selon couverture (tuile/bac acier)")

    # ── Batterie (hybride/autonome) ──
    if has_battery or installation_type in ("hybride", "autonome"):
        add("Batterie", "Câble batterie DC 25 mm²", 6.0, "m",
            "section forte courant, cosses serties")
        add("Protection batterie", "Fusible / disjoncteur DC batterie", 1, "u",
            "calibre selon courant batterie")

    if string_result and not string_result.get("ok", True):
        warnings.append(
            "la conception de chaînes signale des avertissements — vérifier la "
            "compatibilité tension avant de figer la nomenclature")

    summary = {
        "kwc": round(kwc, 3),
        "n_panels": n_panels,
        "strings": strings,
        "phases": phases,
        "ac_breaker_amp": breaker_amp,
        "ac_cable_section_mm2": ac_section,
        "n_lignes": len(items),
    }
    return {"items": items, "summary": summary, "warnings": warnings}


# Calibres de disjoncteur AC normalisés (A).
_STD_BREAKERS = [6, 10, 16, 20, 25, 32, 40, 50, 63, 80, 100, 125, 160]


def _round_breaker(amps):
    """Arrondit au calibre normalisé immédiatement supérieur."""
    try:
        amps = float(amps)
    except (TypeError, ValueError):
        amps = 0.0
    for b in _STD_BREAKERS:
        if amps <= b:
            return b
    return _STD_BREAKERS[-1]


# ── FG255 — Dimensionnement borne de recharge VE couplée au PV ───────────────
# Dimensionne une borne de recharge de véhicule électrique (puissance kW, mono
# ou triphasé, nombre de sessions/jour) et chiffre son IMPACT sur
# l'autoconsommation du champ PV : combien de l'énergie VE peut être couverte
# par le surplus solaire journalier, et de combien la borne augmente le taux
# d'autoconsommation global de l'installation. Module PUR : aucune base, aucun
# réseau, aucun prix. Les entrées numériques ne sont JAMAIS rejetées (la liberté
# de saisie du founder est préservée) — seules les valeurs absurdes (≤ 0) sont
# bornées à un défaut sensé pour éviter une division par zéro.

# Bornes monophasées usuelles au Maroc (230 V) et triphasées (400 V).
_EV_VOLTAGE_MONO = 230.0
_EV_VOLTAGE_TRI = 400.0
# Rendement de charge AC→batterie (pertes chargeur embarqué + câble).
_EV_CHARGE_EFFICIENCY = 0.90
# Calibres de borne courants (kW) pour l'aide au choix.
_EV_STD_POWER_KW = [3.7, 7.4, 11.0, 22.0]


def ev_charger_sizing(*, borne_kw=7.4, phases=1, sessions_per_day=1,
                      energy_per_session_kwh=None, kwh_per_100km=18.0,
                      km_per_session=40.0, pv_kwc=None,
                      pv_daily_production_kwh=None,
                      pv_self_consumption_kwh=None,
                      pv_surplus_kwh=None, charge_window_h=None,
                      productible_kwh_kwc_year=1700.0):
    """FG255 — dimensionne une borne VE et chiffre son impact autoconsommation.

    Côté BORNE, à partir de la puissance ``borne_kw``, du nombre de phases
    (``phases`` = 1 mono / 3 tri) et des ``sessions_per_day`` :

    * courant de ligne (A) et calibre disjoncteur dédié (1.25× le nominal,
      arrondi au calibre normalisé) ;
    * énergie VE journalière requise — soit ``energy_per_session_kwh`` fournie,
      soit déduite de la conso véhicule (``kwh_per_100km`` × ``km_per_session``
      / 100) — multipliée par les sessions et divisée par le rendement de
      charge ;
    * durée de charge d'une session à ``borne_kw`` (h) et vérification qu'elle
      tient dans la fenêtre de charge ``charge_window_h`` si fournie.

    Côté PV (IMPACT AUTOCONSOMMATION), si un contexte PV est donné :

    * le SURPLUS solaire journalier (``pv_surplus_kwh`` direct, sinon
      production − autoconsommation existante, sinon déduit de ``pv_kwc`` via
      ``productible_kwh_kwc_year``) est la réserve disponible pour le VE ;
    * la part de l'énergie VE couvrable par ce surplus (``solar_covered_kwh``)
      et donc importée du réseau (``grid_kwh``) ;
    * le NOUVEAU taux d'autoconsommation : (autoconsommation existante + énergie
      VE couverte par le solaire) ÷ production — la borne RECYCLE le surplus
      qui partait au réseau, donc le taux MONTE.

    Toutes les sorties sont JSON-sérialisables et sûres sur entrées dégradées
    (jamais d'exception, division par zéro bornée). Retourne ::

        {borne: {kw, phases, line_current_a, breaker_a, voltage_v,
                 session_charge_h, fits_window, recommended_kw},
         energy: {per_session_kwh, daily_demand_kwh, sessions_per_day},
         pv_impact: {available_surplus_kwh, solar_covered_kwh, grid_kwh,
                     solar_coverage_pct, base_self_consumption_pct,
                     new_self_consumption_pct, self_consumption_gain_pts,
                     pv_daily_production_kwh},
         warnings: []}
    """
    warnings = []

    # ── Normalisation des entrées (bornage minimal, jamais de rejet) ──
    def _pos(value, default):
        try:
            v = float(value)
        except (TypeError, ValueError):
            return float(default)
        return v if v > 0 else float(default)

    def _nonneg(value):
        try:
            v = float(value)
        except (TypeError, ValueError):
            return None
        return v if v >= 0 else None

    kw = _pos(borne_kw, 7.4)
    ph = 3 if int(phases or 1) == 3 else 1
    sessions = _pos(sessions_per_day, 1)

    # ── Énergie d'une session ──
    if energy_per_session_kwh is not None:
        per_session = _pos(energy_per_session_kwh, 0.0)
        if per_session <= 0:
            per_session = 0.0
    else:
        # Déduite de la conso véhicule × km par session.
        kwh_100 = _pos(kwh_per_100km, 18.0)
        km = _pos(km_per_session, 40.0)
        per_session = kwh_100 * km / 100.0

    # Énergie à PRÉLEVER au tableau (pertes de charge incluses).
    daily_demand = round(
        per_session * sessions / _EV_CHARGE_EFFICIENCY, 2)

    # ── Électricité borne : courant de ligne & calibre ──
    voltage = _EV_VOLTAGE_TRI if ph == 3 else _EV_VOLTAGE_MONO
    sqrt3 = math.sqrt(3) if ph == 3 else 1.0
    line_current = (kw * 1000.0) / (voltage * sqrt3)
    breaker_a = _round_breaker(line_current * 1.25)

    # Durée d'une session à la puissance borne (pertes incluses).
    session_charge_h = round(
        (per_session / _EV_CHARGE_EFFICIENCY) / kw, 2) if kw > 0 else None

    fits_window = None
    win = _nonneg(charge_window_h)
    if win is not None and session_charge_h is not None:
        fits_window = session_charge_h <= win
        if not fits_window:
            warnings.append(
                f"charge d'une session {session_charge_h} h > fenêtre "
                f"disponible {win} h — augmenter la puissance de borne ou "
                "réduire l'énergie par session")

    # Borne recommandée : la plus petite puissance standard ≥ celle saisie.
    recommended_kw = next((p for p in _EV_STD_POWER_KW if p >= kw - 1e-6), kw)

    # ── Impact PV / autoconsommation ──
    prod = _nonneg(pv_daily_production_kwh)
    if prod is None:
        kwc = _nonneg(pv_kwc)
        if kwc is not None and kwc > 0:
            prod = round(
                kwc * _pos(productible_kwh_kwc_year, 1700.0) / 365.0, 2)

    surplus = _nonneg(pv_surplus_kwh)
    base_self = _nonneg(pv_self_consumption_kwh)
    if surplus is None and prod is not None and base_self is not None:
        surplus = max(0.0, round(prod - base_self, 2))

    pv_impact = {
        "pv_daily_production_kwh": prod,
        "available_surplus_kwh": surplus,
        "solar_covered_kwh": None,
        "grid_kwh": None,
        "solar_coverage_pct": None,
        "base_self_consumption_pct": None,
        "new_self_consumption_pct": None,
        "self_consumption_gain_pts": None,
    }

    if surplus is not None and daily_demand > 0:
        # Le VE consomme d'abord le surplus solaire, le reste vient du réseau.
        solar_covered = round(min(surplus, daily_demand), 2)
        grid_kwh = round(max(0.0, daily_demand - solar_covered), 2)
        pv_impact["solar_covered_kwh"] = solar_covered
        pv_impact["grid_kwh"] = grid_kwh
        pv_impact["solar_coverage_pct"] = round(
            solar_covered / daily_demand * 100.0, 1)
        if solar_covered < daily_demand * 0.5:
            warnings.append(
                "le surplus solaire couvre moins de la moitié des besoins VE — "
                "envisager d'agrandir le champ PV ou de programmer la charge en "
                "milieu de journée")

        # Nouveau taux d'autoconsommation : la borne recycle le surplus.
        if prod is not None and prod > 0:
            base_sc = base_self if base_self is not None else max(
                0.0, prod - surplus)
            pv_impact["base_self_consumption_pct"] = round(
                base_sc / prod * 100.0, 1)
            new_sc = base_sc + solar_covered
            pv_impact["new_self_consumption_pct"] = round(
                min(new_sc, prod) / prod * 100.0, 1)
            pv_impact["self_consumption_gain_pts"] = round(
                pv_impact["new_self_consumption_pct"]
                - pv_impact["base_self_consumption_pct"], 1)

    return {
        "borne": {
            "kw": round(kw, 2),
            "phases": ph,
            "voltage_v": voltage,
            "line_current_a": round(line_current, 1),
            "breaker_a": breaker_a,
            "session_charge_h": session_charge_h,
            "fits_window": fits_window,
            "recommended_kw": recommended_kw,
        },
        "energy": {
            "per_session_kwh": round(per_session, 2),
            "daily_demand_kwh": daily_demand,
            "sessions_per_day": round(sessions, 2),
            "charge_efficiency": _EV_CHARGE_EFFICIENCY,
        },
        "pv_impact": pv_impact,
        "warnings": warnings,
    }


# ── FG256 — Étude de stockage & dispatch batterie (backup) ───────────────────
# Dimensionne le parc batterie d'une installation PV pour DEUX objectifs
# distincts, puis désigne la contrainte DIMENSIONNANTE (binding) :
#
#   (a) AUTOCONSOMMATION MAX — stocker le SURPLUS solaire de la journée pour le
#       restituer le soir/la nuit. La capacité utile vise à absorber le surplus
#       journalier (production − autoconsommation directe), elle-même bornée par
#       l'énergie réellement déchargée la nuit (besoin du soir). La puissance
#       utile suit le pic de décharge nocturne.
#   (b) BACKUP N heures critiques — tenir une charge critique (kW) pendant un
#       nombre d'heures de coupure. La capacité utile = charge critique × heures,
#       la puissance utile = charge critique (avec une marge de pointe).
#
# On part toujours de la capacité UTILE (kWh utilisables) puis on remonte à la
# capacité NOMINALE installée en divisant par la profondeur de décharge (DoD) et
# le rendement aller-retour (round-trip) — deux pertes physiques bien réelles.
# Module PUR : aucune base, aucun réseau, aucun prix. Les entrées numériques ne
# sont JAMAIS rejetées (liberté de saisie du founder) — seules les valeurs
# absurdes (≤ 0) sont bornées à un défaut sensé pour éviter une division par
# zéro.

# Profondeur de décharge utilisable par défaut (lithium LFP courant : 90 %).
_BATTERY_DEFAULT_DOD = 0.90
# Rendement aller-retour (charge → décharge) d'un parc lithium + onduleur.
_BATTERY_DEFAULT_ROUND_TRIP = 0.90
# Facteur de pointe pour la puissance backup (démarrages moteurs, appels).
_BATTERY_BACKUP_PEAK_FACTOR = 1.25
# Fraction du surplus journalier réellement restituée le soir si le besoin
# nocturne n'est pas précisé (le reste serait réinjecté/perdu).
_BATTERY_DEFAULT_NIGHT_FRACTION = 0.80


def battery_storage_sizing(*, mode="autoconso",
                           pv_daily_production_kwh=None,
                           pv_self_consumption_kwh=None,
                           daily_surplus_kwh=None,
                           night_load_kwh=None,
                           pv_kwc=None,
                           productible_kwh_kwc_year=1700.0,
                           critical_load_kw=None,
                           backup_hours=None,
                           evening_peak_kw=None,
                           depth_of_discharge=_BATTERY_DEFAULT_DOD,
                           round_trip_efficiency=_BATTERY_DEFAULT_ROUND_TRIP,
                           system_voltage_v=48.0):
    """FG256 — capacité (kWh) et puissance (kW) batterie utiles + nominales.

    Calcule le dimensionnement pour le ou les objectifs demandés et désigne la
    contrainte DIMENSIONNANTE (``binding_objective``).

    Modes (``mode``) :

    * ``"autoconso"`` — autoconsommation max : stocker le surplus journalier.
    * ``"backup"`` — autonomie : tenir la charge critique N heures.
    * ``"both"`` — calcule les deux ; la capacité retenue est la PLUS GRANDE et
      ``binding_objective`` indique laquelle dimensionne le parc.

    AUTOCONSOMMATION : le surplus journalier (``daily_surplus_kwh`` direct,
    sinon production − autoconsommation directe, sinon déduit de ``pv_kwc`` via
    ``productible_kwh_kwc_year``) est l'énergie EXCÉDENTAIRE de la journée. La
    capacité UTILE visée est ``min(surplus, besoin nocturne)`` : inutile de
    stocker plus que ce qui sera redéchargé le soir (``night_load_kwh`` ; à
    défaut une fraction du surplus). La puissance UTILE suit le pic de décharge
    nocturne (``evening_peak_kw`` si fourni, sinon estimée du besoin nocturne).

    BACKUP : capacité UTILE = ``critical_load_kw`` × ``backup_hours`` ; puissance
    UTILE = ``critical_load_kw`` × marge de pointe (``_BATTERY_BACKUP_PEAK_FACTOR``).

    De l'UTILE au NOMINAL : la capacité nominale installée =
    capacité utile ÷ (DoD × √rendement_round_trip) — la profondeur de décharge
    limite la fraction exploitable et le rendement aller-retour ajoute des pertes
    de stockage. Le courant batterie indicatif = puissance utile ÷ tension
    système.

    Toutes les sorties sont JSON-sérialisables et sûres sur entrées dégradées
    (jamais d'exception, division par zéro bornée). Retourne ::

        {mode, autoconso: {usable_kwh, usable_kw, nominal_kwh, ...} | None,
         backup: {usable_kwh, usable_kw, nominal_kwh, ...} | None,
         recommended: {usable_kwh, usable_kw, nominal_kwh, current_a},
         binding_objective, depth_of_discharge, round_trip_efficiency,
         warnings: []}
    """
    warnings = []

    def _pos(value, default):
        try:
            v = float(value)
        except (TypeError, ValueError):
            return float(default)
        return v if v > 0 else float(default)

    def _nonneg(value):
        try:
            v = float(value)
        except (TypeError, ValueError):
            return None
        return v if v >= 0 else None

    # DoD et rendement bornés à ]0, 1] (jamais une division par zéro).
    dod = _pos(depth_of_discharge, _BATTERY_DEFAULT_DOD)
    if dod > 1.0:
        dod = 1.0
        warnings.append("profondeur de décharge plafonnée à 100 %")
    rte = _pos(round_trip_efficiency, _BATTERY_DEFAULT_ROUND_TRIP)
    if rte > 1.0:
        rte = 1.0
        warnings.append("rendement aller-retour plafonné à 100 %")
    voltage = _pos(system_voltage_v, 48.0)

    mode = (mode or "autoconso").lower()
    want_autoconso = mode in ("autoconso", "both")
    want_backup = mode in ("backup", "both")
    if mode not in ("autoconso", "backup", "both"):
        want_autoconso = True
        warnings.append(
            f"mode « {mode} » inconnu — autoconsommation par défaut")

    def _usable_to_nominal(usable_kwh):
        # Capacité nominale = utile / (DoD × √rendement aller-retour). Le √
        # répartit la perte round-trip entre charge et décharge (modèle simple).
        denom = dod * math.sqrt(rte)
        return round(usable_kwh / denom, 2) if denom > 0 else None

    # ── (a) AUTOCONSOMMATION MAX ──────────────────────────────────────────────
    autoconso = None
    if want_autoconso:
        prod = _nonneg(pv_daily_production_kwh)
        if prod is None:
            kwc = _nonneg(pv_kwc)
            if kwc is not None and kwc > 0:
                prod = round(
                    kwc * _pos(productible_kwh_kwc_year, 1700.0) / 365.0, 2)

        surplus = _nonneg(daily_surplus_kwh)
        base_self = _nonneg(pv_self_consumption_kwh)
        if surplus is None and prod is not None and base_self is not None:
            surplus = max(0.0, round(prod - base_self, 2))

        if surplus is None:
            autoconso = {
                "usable_kwh": None, "usable_kw": None, "nominal_kwh": None,
                "daily_surplus_kwh": None, "night_load_kwh": _nonneg(night_load_kwh),
                "stored_kwh": None, "spilled_surplus_kwh": None,
                "pv_daily_production_kwh": prod,
            }
            warnings.append(
                "autoconsommation : surplus solaire inconnu — fournir production "
                "+ autoconsommation, surplus direct, ou kWc")
        else:
            # Énergie redéchargée le soir : besoin nocturne fourni, sinon une
            # fraction du surplus (le reste serait réinjecté/perdu).
            night = _nonneg(night_load_kwh)
            if night is None:
                night = round(surplus * _BATTERY_DEFAULT_NIGHT_FRACTION, 2)
            # On ne stocke pas plus que ce qui sera redéchargé (ni que le surplus).
            usable_kwh = round(min(surplus, night), 2)
            spilled = round(max(0.0, surplus - usable_kwh), 2)
            # Pic de décharge nocturne : fourni, sinon ~le besoin réparti sur une
            # soirée de pointe de 4 h (modèle simple).
            peak = _nonneg(evening_peak_kw)
            if peak is None or peak <= 0:
                peak = round(usable_kwh / 4.0, 2) if usable_kwh > 0 else 0.0
            autoconso = {
                "usable_kwh": usable_kwh,
                "usable_kw": round(peak, 2),
                "nominal_kwh": _usable_to_nominal(usable_kwh),
                "daily_surplus_kwh": round(surplus, 2),
                "night_load_kwh": round(night, 2),
                "stored_kwh": usable_kwh,
                "spilled_surplus_kwh": spilled,
                "pv_daily_production_kwh": prod,
            }
            if spilled > 0:
                warnings.append(
                    "autoconsommation : une partie du surplus dépasse le besoin "
                    "nocturne et ne sera pas stockée (réinjection/écrêtage)")

    # ── (b) BACKUP N heures critiques ─────────────────────────────────────────
    backup = None
    if want_backup:
        crit_kw = _nonneg(critical_load_kw)
        hours = _nonneg(backup_hours)
        if crit_kw is None or hours is None:
            backup = {
                "usable_kwh": None, "usable_kw": None, "nominal_kwh": None,
                "critical_load_kw": crit_kw, "backup_hours": hours,
            }
            warnings.append(
                "backup : charge critique (kW) et heures d'autonomie requises")
        else:
            usable_kwh = round(crit_kw * hours, 2)
            # Puissance utile = charge critique avec marge de pointe (appels
            # moteurs, démarrages) — l'onduleur batterie doit la soutenir.
            usable_kw = round(crit_kw * _BATTERY_BACKUP_PEAK_FACTOR, 2)
            backup = {
                "usable_kwh": usable_kwh,
                "usable_kw": usable_kw,
                "nominal_kwh": _usable_to_nominal(usable_kwh),
                "critical_load_kw": round(crit_kw, 2),
                "backup_hours": round(hours, 2),
                "peak_factor": _BATTERY_BACKUP_PEAK_FACTOR,
            }
            if usable_kwh <= 0:
                warnings.append(
                    "backup : énergie d'autonomie nulle (charge ou heures à 0)")

    # ── Contrainte dimensionnante (binding) ───────────────────────────────────
    candidates = []
    if autoconso and autoconso.get("usable_kwh") is not None:
        candidates.append(("autoconso", autoconso))
    if backup and backup.get("usable_kwh") is not None:
        candidates.append(("backup", backup))

    binding = None
    recommended = {
        "usable_kwh": None, "usable_kw": None,
        "nominal_kwh": None, "current_a": None,
    }
    if candidates:
        # La contrainte dimensionnante est celle qui exige la plus grande
        # capacité utile ; la puissance retenue est le max des deux pics.
        binding, chosen = max(
            candidates, key=lambda c: c[1].get("usable_kwh") or 0.0)
        usable_kwh = chosen.get("usable_kwh") or 0.0
        usable_kw = max(
            (c[1].get("usable_kw") or 0.0) for c in candidates)
        nominal_kwh = _usable_to_nominal(usable_kwh)
        current_a = round(usable_kw * 1000.0 / voltage, 1) \
            if voltage > 0 and usable_kw else 0.0
        recommended = {
            "usable_kwh": round(usable_kwh, 2),
            "usable_kw": round(usable_kw, 2),
            "nominal_kwh": nominal_kwh,
            "current_a": current_a,
        }

    return {
        "mode": mode,
        "autoconso": autoconso,
        "backup": backup,
        "recommended": recommended,
        "binding_objective": binding,
        "depth_of_discharge": round(dod, 4),
        "round_trip_efficiency": round(rte, 4),
        "system_voltage_v": round(voltage, 1),
        "warnings": warnings,
    }


# ── FG257 — Simulation bankable P50/P90 avec modèle de pertes ─────────────────
# Transforme une production de BASE (le productible « brut », p. ex. PVGIS au
# point de fonctionnement idéal × kWc) en une production FINANCIÈREMENT EXPLOITABLE
# (« bankable ») en deux temps :
#
#   1) MODÈLE DE PERTES → RATIO DE PERFORMANCE (PR). Chaque source de perte
#      physique (température, salissure/poussière, câblage DC+AC, rendement
#      onduleur, et des pertes diverses : mismatch, ombrage déjà chiffré ailleurs,
#      indisponibilité) ronge une fraction de l'énergie. Le PR est le produit des
#      rendements de chaque poste : PR = Π (1 − perte_i). La production P50 (valeur
#      médiane attendue, « 1 année sur 2 on fait au moins ça ») = base × PR.
#
#   2) VARIABILITÉ INTERANNUELLE → P90 / P75. L'ensoleillement varie d'une année
#      à l'autre (météo) ; on modélise la production annuelle comme une gaussienne
#      de moyenne P50 et d'écart-type relatif σ (typiquement 5–7 %). Les banques
#      financent sur le P90 : la production DÉPASSÉE 9 années sur 10. Avec le
#      quantile gaussien z₉₀ = 1.282, P90 = P50 × (1 − z₉₀·σ) ; de même
#      P75 = P50 × (1 − z₇₅·σ) avec z₇₅ = 0.674.
#
# Module PUR : aucune base, aucun réseau, aucun prix. Les entrées numériques ne
# sont JAMAIS rejetées (liberté de saisie du founder) — seuls les facteurs de
# perte sont bornés à [0, 1] (une perte hors de cet intervalle n'a pas de sens
# physique) et σ est borné ≥ 0 pour éviter un P90 > P50.

# Postes de perte par défaut (fractions), valeurs marché conservatrices pour une
# centrale PV au Maroc bien conçue. Tout est surchargeable par l'appelant.
DEFAULT_LOSS_FACTORS = {
    "temperature": 0.08,   # échauffement cellule au-dessus du STC (climat chaud)
    "soiling": 0.03,       # salissure / poussière (Maroc : sable, à nettoyer)
    "wiring": 0.02,        # pertes ohmiques câblage DC + AC
    "inverter": 0.025,     # rendement de conversion onduleur (≈ 97.5 %)
    "mismatch": 0.02,      # dispersion modules + connectique + LID/dégradation 1re année
    "availability": 0.01,  # indisponibilité réseau / maintenance
}

# Quantiles de la loi normale centrée réduite (borne basse) :
#   P90 = production dépassée 90 % du temps → z tel que Φ(−z) = 10 % → z = 1.282.
#   P75 = production dépassée 75 % du temps → Φ(−z) = 25 % → z = 0.674.
Z_P90 = 1.282
Z_P75 = 0.674

# Écart-type relatif interannuel par défaut (variabilité météo), typiquement 5–7 %.
DEFAULT_ANNUAL_VARIABILITY = 0.06


def _clamp01(value, default):
    """Borne un facteur de perte à [0, 1]. Entrée illisible → ``default``."""
    try:
        v = float(value)
    except (TypeError, ValueError):
        return float(default)
    if v < 0.0:
        return 0.0
    if v > 1.0:
        return 1.0
    return v


def simulate_bankable_yield(base_production_kwh, *, loss_factors=None,
                            annual_variability=DEFAULT_ANNUAL_VARIABILITY,
                            kwc=None, include_p75=True):
    """FG257 — simulation bankable P50/P90 (+P75) avec modèle de pertes & PR.

    Calcule un RATIO DE PERFORMANCE (PR) à partir des postes de perte, l'applique
    à une production de base pour obtenir le P50 (médiane), puis dérive le P90
    (et optionnellement le P75) via le quantile gaussien d'une variabilité
    interannuelle ``annual_variability`` (σ relatif).

    Paramètres
    ----------
    base_production_kwh : production de base AVANT pertes (kWh/an) — typiquement
        le productible idéal (PVGIS au point de fonctionnement) × kWc. Une valeur
        ≤ 0 ou illisible → 0 (toutes les sorties à 0, jamais d'exception).
    loss_factors : dict ``{poste: fraction}`` surchargeant ``DEFAULT_LOSS_FACTORS``
        (température, soiling/salissure, wiring/câblage, inverter/onduleur, …).
        Chaque facteur est borné à [0, 1] ; les postes inconnus sont acceptés et
        comptés (extensibilité). Le PR = produit des (1 − perte_i).
    annual_variability : écart-type relatif σ de la production annuelle (météo),
        typiquement 0.05–0.07. Borné ≥ 0 (un σ négatif est ramené à 0 → P90=P50).
    kwc : puissance crête (kWc) facultative, pour rapporter un ``specific_yield``
        (kWh/kWc/an) au P50 — purement indicatif, jamais de prix.
    include_p75 : ajoute le P75 (médiane des banques moins conservatrices).

    Retourne un dict JSON-sérialisable ::

        {base_production_kwh, performance_ratio, total_loss_pct,
         loss_breakdown: {poste: {fraction, pct}}, applied_losses,
         p50_kwh, p90_kwh, p75_kwh|None, annual_variability,
         z_p90, z_p75, specific_yield_kwh_kwc|None, warnings: []}

    Ne lève jamais : entrées dégradées → structure cohérente à 0.
    """
    warnings = []

    try:
        base = float(base_production_kwh)
    except (TypeError, ValueError):
        base = 0.0
    if base < 0.0:
        base = 0.0
        warnings.append("production de base négative ramenée à 0")

    factors = {**DEFAULT_LOSS_FACTORS, **(loss_factors or {})}

    # ── PR = produit des rendements (1 − perte) de chaque poste ──
    loss_breakdown = {}
    pr = 1.0
    for poste, raw in factors.items():
        frac = _clamp01(raw, DEFAULT_LOSS_FACTORS.get(poste, 0.0))
        loss_breakdown[poste] = {
            "fraction": round(frac, 4),
            "pct": round(frac * 100.0, 2),
        }
        pr *= (1.0 - frac)

    performance_ratio = round(pr, 4)
    total_loss_pct = round((1.0 - pr) * 100.0, 2)
    if performance_ratio < 0.70 and factors:
        warnings.append(
            "ratio de performance < 0.70 — pertes cumulées élevées, vérifier les "
            "postes (température/salissure/câblage/onduleur)")

    # ── P50 = base × PR (production médiane attendue) ──
    p50 = round(base * pr, 1)

    # ── Variabilité interannuelle → P90 / P75 (quantile gaussien borne basse) ──
    try:
        sigma = float(annual_variability)
    except (TypeError, ValueError):
        sigma = DEFAULT_ANNUAL_VARIABILITY
    if sigma < 0.0:
        sigma = 0.0
        warnings.append("variabilité interannuelle négative ramenée à 0")
    if sigma > 0.30:
        warnings.append(
            "variabilité interannuelle > 30 % — valeur inhabituelle, vérifier σ")

    # P90/P75 = P50 × (1 − z·σ), borné ≥ 0 (un σ énorme ne donne pas un négatif).
    p90 = round(p50 * max(0.0, 1.0 - Z_P90 * sigma), 1)
    p75 = round(p50 * max(0.0, 1.0 - Z_P75 * sigma), 1) if include_p75 else None

    specific_yield = None
    try:
        k = float(kwc) if kwc is not None else None
    except (TypeError, ValueError):
        k = None
    if k is not None and k > 0:
        specific_yield = round(p50 / k, 1)

    return {
        "base_production_kwh": round(base, 1),
        "performance_ratio": performance_ratio,
        "total_loss_pct": total_loss_pct,
        "loss_breakdown": loss_breakdown,
        "applied_losses": sorted(loss_breakdown.keys()),
        "p50_kwh": p50,
        "p90_kwh": p90,
        "p75_kwh": p75,
        "annual_variability": round(sigma, 4),
        "z_p90": Z_P90,
        "z_p75": Z_P75,
        "specific_yield_kwh_kwc": specific_yield,
        "warnings": warnings,
    }


# ── FG258 — Profil d'autoconsommation horaire depuis la courbe de charge ──────
# Croise une COURBE DE CHARGE horaire (consommation, kWh par heure) avec un
# PROFIL DE PRODUCTION horaire (PV, kWh par heure) pour calculer le taux
# d'autoconsommation RÉEL — heure par heure, l'autoconsommé instantané vaut
# min(charge, production) (on ne peut pas autoconsommer plus que ce que l'on
# produit NI plus que ce que l'on consomme à cet instant). Le surplus
# (production − autoconsommé) part au réseau ; le complément (charge −
# autoconsommé) est importé.
#
#   taux d'autoconsommation = Σ autoconsommé / Σ production
#       → quelle part de MA production je consomme moi-même (le reste injecté).
#   taux de couverture (autoproduction) = Σ autoconsommé / Σ charge
#       → quelle part de MA consommation est couverte par le solaire.
#
# Le calcul est PUR (aucune base, aucun réseau, aucun prix) et tolère des
# courbes de longueurs différentes (on aligne sur la plus courte, en signalant).
# Un profil annuel 8760 h fonctionne exactement comme un profil type 24 h : la
# routine ne suppose AUCUNE longueur particulière. La liberté de saisie est
# préservée : aucune valeur n'est rejetée — les valeurs illisibles ou négatives
# sont ramenées à 0 (une charge/production négative n'a pas de sens physique),
# jamais d'exception, division par zéro bornée (Σproduction = 0 → taux = 0).
#
# Le PARSING d'un classeur .xlsx (openpyxl, déjà une dépendance du projet) est
# tenu SÉPARÉ du calcul : ``load_curve_from_xlsx`` lit une colonne en liste de
# floats, puis on passe cette liste à ``hourly_self_consumption``. Aucune
# dépendance pip nouvelle n'est introduite.

# ── Profils horaires types (24 valeurs, part de la grandeur journalière) ──────
# Profil de CHARGE résidentiel marocain type : creux la nuit, pics matin & soir
# (cuisine, éclairage, clim/TV en soirée). Somme = 1.0 (fractions de la conso
# journalière). Index 0 = 00 h … 23 = 23 h.
TYPICAL_LOAD_PROFILE_RESIDENTIAL = [
    0.020, 0.018, 0.016, 0.016, 0.018, 0.025,  # 00–05 h
    0.040, 0.055, 0.050, 0.040, 0.035, 0.035,  # 06–11 h
    0.038, 0.035, 0.030, 0.030, 0.035, 0.050,  # 12–17 h
    0.075, 0.090, 0.085, 0.065, 0.045, 0.029,  # 18–23 h
]

# Profil de CHARGE tertiaire / commerce (journée ouvrée) : plat la nuit, plateau
# diurne aligné sur le soleil — beaucoup plus favorable à l'autoconsommation.
TYPICAL_LOAD_PROFILE_COMMERCIAL = [
    0.010, 0.010, 0.010, 0.010, 0.010, 0.015,  # 00–05 h
    0.030, 0.055, 0.075, 0.085, 0.090, 0.085,  # 06–11 h
    0.075, 0.085, 0.090, 0.085, 0.070, 0.045,  # 12–17 h
    0.025, 0.015, 0.010, 0.010, 0.010, 0.010,  # 18–23 h
]

# Profil de PRODUCTION PV type (jour clair) : cloche centrée sur midi solaire,
# nulle la nuit. Somme = 1.0 (fractions de la production journalière).
TYPICAL_PV_PROFILE = [
    0.0, 0.0, 0.0, 0.0, 0.0, 0.005,            # 00–05 h
    0.020, 0.045, 0.075, 0.100, 0.120, 0.130,  # 06–11 h
    0.130, 0.120, 0.100, 0.075, 0.045, 0.020,  # 12–17 h
    0.010, 0.005, 0.0, 0.0, 0.0, 0.0,          # 18–23 h
]

_TYPICAL_LOAD_PROFILES = {
    "residential": TYPICAL_LOAD_PROFILE_RESIDENTIAL,
    "residentiel": TYPICAL_LOAD_PROFILE_RESIDENTIAL,
    "commercial": TYPICAL_LOAD_PROFILE_COMMERCIAL,
    "tertiaire": TYPICAL_LOAD_PROFILE_COMMERCIAL,
}


def _coerce_series(values):
    """Convertit un itérable en liste de floats ≥ 0 (illisible/<0 → 0.0).

    Préserve la longueur : chaque case impossible à lire ou négative devient
    0.0 (jamais de rejet, jamais d'exception). ``None`` → liste vide.
    """
    if values is None:
        return []
    out = []
    for v in values:
        try:
            f = float(v)
        except (TypeError, ValueError):
            f = 0.0
        if f < 0.0 or f != f:  # négatif ou NaN → 0
            f = 0.0
        out.append(f)
    return out


def _scaled_typical_load(total_kwh, profile_key="residential"):
    """Distribue ``total_kwh`` sur 24 h selon un profil type (somme→total)."""
    try:
        total = float(total_kwh)
    except (TypeError, ValueError):
        total = 0.0
    if total < 0.0:
        total = 0.0
    profile = _TYPICAL_LOAD_PROFILES.get(
        (profile_key or "residential").lower(),
        TYPICAL_LOAD_PROFILE_RESIDENTIAL)
    s = sum(profile)
    if s <= 0:
        return [0.0] * len(profile)
    return [total * (p / s) for p in profile]


def _scaled_typical_pv(total_kwh):
    """Distribue ``total_kwh`` de production sur 24 h selon ``TYPICAL_PV_PROFILE``."""
    try:
        total = float(total_kwh)
    except (TypeError, ValueError):
        total = 0.0
    if total < 0.0:
        total = 0.0
    s = sum(TYPICAL_PV_PROFILE)
    if s <= 0:
        return [0.0] * len(TYPICAL_PV_PROFILE)
    return [total * (p / s) for p in TYPICAL_PV_PROFILE]


def hourly_self_consumption(load_curve=None, production_curve=None, *,
                            daily_load_kwh=None, daily_production_kwh=None,
                            load_profile="residential"):
    """FG258 — taux d'autoconsommation RÉEL depuis courbes horaires.

    Aligne heure par heure une COURBE DE CHARGE (consommation, kWh/h) et un
    PROFIL DE PRODUCTION (PV, kWh/h). Pour chaque heure :

        autoconsommé[h] = min(charge[h], production[h])
        surplus_injecté[h] = production[h] − autoconsommé[h]
        importé_réseau[h] = charge[h] − autoconsommé[h]

    puis agrège ::

        taux_autoconsommation = Σ autoconsommé / Σ production
        taux_couverture       = Σ autoconsommé / Σ charge

    Paramètres
    ----------
    load_curve : itérable de consommations horaires (kWh/h) — 24 h, 8760 h ou
        toute longueur. Les valeurs illisibles/négatives sont ramenées à 0
        (jamais de rejet). Si absent, on synthétise un profil type
        (``load_profile``) calé sur ``daily_load_kwh``.
    production_curve : itérable de productions PV horaires (kWh/h). Si absent,
        on synthétise ``TYPICAL_PV_PROFILE`` calé sur ``daily_production_kwh``.
    daily_load_kwh / daily_production_kwh : énergies journalières servant à
        générer les profils type quand une courbe n'est pas fournie.
    load_profile : clé de profil type de charge (``"residential"`` |
        ``"commercial"`` / ``"tertiaire"``) utilisée comme repli.

    Retourne un dict JSON-sérialisable ::

        {hours, total_load_kwh, total_production_kwh, self_consumed_kwh,
         surplus_kwh, grid_import_kwh,
         self_consumption_rate, self_consumption_pct,
         coverage_rate, coverage_pct,
         load_source, production_source, warnings: []}

    Ne lève jamais : Σproduction = 0 → taux d'autoconso 0 ; Σcharge = 0 →
    couverture 0. Des courbes de longueurs différentes sont alignées sur la plus
    courte (avec un avertissement).
    """
    warnings = []

    load = _coerce_series(load_curve)
    load_source = "courbe fournie"
    if not load:
        load = _scaled_typical_load(daily_load_kwh, load_profile)
        load_source = f"profil type ({load_profile})"

    prod = _coerce_series(production_curve)
    production_source = "courbe fournie"
    if not prod:
        prod = _scaled_typical_pv(daily_production_kwh)
        production_source = "profil type PV"

    # ── Alignement des longueurs (on borne sur la plus courte) ──
    n = min(len(load), len(prod))
    if len(load) != len(prod) and load and prod:
        warnings.append(
            f"courbes de longueurs différentes (charge={len(load)} h, "
            f"production={len(prod)} h) — alignées sur {n} h")

    total_load = 0.0
    total_prod = 0.0
    self_consumed = 0.0
    for i in range(n):
        c = load[i]
        p = prod[i]
        total_load += c
        total_prod += p
        self_consumed += c if c < p else p   # min(charge, production)

    total_load = round(total_load, 3)
    total_prod = round(total_prod, 3)
    self_consumed = round(self_consumed, 3)
    surplus = round(max(0.0, total_prod - self_consumed), 3)
    grid_import = round(max(0.0, total_load - self_consumed), 3)

    sc_rate = round(self_consumed / total_prod, 4) if total_prod > 0 else 0.0
    cov_rate = round(self_consumed / total_load, 4) if total_load > 0 else 0.0

    if total_prod <= 0:
        warnings.append("production horaire nulle — taux d'autoconsommation 0")
    elif sc_rate >= 0.95:
        warnings.append(
            "autoconsommation quasi totale (≥ 95 %) — peu/pas de surplus "
            "injecté ; un champ plus grand resterait autoconsommé")
    elif sc_rate <= 0.30 and total_prod > 0:
        warnings.append(
            "autoconsommation faible (≤ 30 %) — fort surplus injecté au "
            "réseau ; envisager stockage, décalage des usages ou champ réduit")

    return {
        "hours": n,
        "total_load_kwh": total_load,
        "total_production_kwh": total_prod,
        "self_consumed_kwh": self_consumed,
        "surplus_kwh": surplus,
        "grid_import_kwh": grid_import,
        "self_consumption_rate": sc_rate,
        "self_consumption_pct": round(sc_rate * 100.0, 1),
        "coverage_rate": cov_rate,
        "coverage_pct": round(cov_rate * 100.0, 1),
        "load_source": load_source,
        "production_source": production_source,
        "warnings": warnings,
    }


def load_curve_from_xlsx(file_or_path, *, sheet=None, column=1,
                         skip_header=True, max_rows=8760):
    """FG258 (I/O) — lit une courbe de charge depuis un classeur .xlsx.

    PARSING SÉPARÉ du calcul : extrait UNE colonne d'un classeur openpyxl en
    liste de floats, prête à passer à :func:`hourly_self_consumption`. openpyxl
    est déjà une dépendance du projet (exports .xlsx) — aucune nouvelle
    dépendance n'est ajoutée.

    Paramètres
    ----------
    file_or_path : chemin, objet fichier ou flux ouvert par ``load_workbook``.
    sheet : nom de la feuille (défaut : feuille active).
    column : index de colonne 1-based contenant les valeurs (défaut 1 = A).
    skip_header : ignore la première ligne (en-tête) si True.
    max_rows : nombre maximal de lignes de DONNÉES lues (8760 = année horaire).

    Retourne la liste de floats (cellule vide/illisible → 0.0 pour préserver
    l'alignement horaire). Ne lève pas sur une cellule illisible.
    """
    import openpyxl  # déjà une dépendance projet (cf. requirements.txt)

    wb = openpyxl.load_workbook(file_or_path, read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet else wb.active
        col_idx = max(1, int(column or 1))
        limit = int(max_rows) + (1 if skip_header else 0)
        values = []
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx,
                                values_only=True):
            cell = row[0] if row else None
            try:
                values.append(float(cell))
            except (TypeError, ValueError):
                values.append(0.0)
            if len(values) >= limit:
                break
    finally:
        wb.close()

    if skip_header and values:
        values = values[1:]
    return values[:max_rows]


# ── FG259 — Économie net-metering / injection surplus (loi 13-09, MT/HT) ──────
# Valorise le SURPLUS solaire INJECTÉ au réseau (issu de FG258 :
# production − autoconsommation, heure par heure) selon la règle de
# compensation marocaine (loi 13-09, moyenne/haute tension) et un tarif
# TIME-OF-USE par tranche horaire (pointe / pleine / creuse).
#
# RÈGLE DE COMPENSATION (net-metering 13-09, modèle transparent) :
#   * L'énergie INJECTÉE pendant une tranche n'est COMPENSÉE qu'à hauteur de
#     l'énergie SOUTIRÉE (importée du réseau) dans la même tranche — c'est le
#     principe du « net » : on solde les flux d'une même période tarifaire.
#     Au-delà, le surplus est « excédentaire » (spill) : par défaut NON
#     valorisé (la loi 13-09 ne rémunère pas l'excédent au MT), valorisable à
#     un tarif résiduel facultatif ``spill_tariff`` si l'appelant le fournit.
#   * Le surplus compensé est valorisé au TARIF de sa propre tranche (il efface
#     un kWh qui aurait été facturé à ce tarif-là).
#   * Un PLAFOND annuel global facultatif (``annual_cap_kwh``) borne l'énergie
#     totale compensée (certains contrats limitent la compensation à un % de la
#     conso annuelle) ; il s'applique sur l'énergie déjà éligible, tranche par
#     tranche, des plus chères aux moins chères (maximise l'économie client).
#   * Le réglage existant ``surplus_injecte_compense`` (toggle) : OFF →
#     l'injection n'est PAS compensée du tout → économie 0 (et tout le surplus
#     bascule en « non compensé »).
#
# Module PUR : aucune base, aucun réseau, aucun prix d'achat/marge. Les entrées
# numériques ne sont JAMAIS rejetées (liberté de saisie du founder) — les
# valeurs illisibles/négatives sont ramenées à 0, jamais d'exception, division
# par zéro bornée.

# ── Tranches time-of-use ONEE par défaut (MT général, ordre des heures) ───────
# Affectation heure-de-journée → tranche (24 valeurs, index 0 = 00 h … 23 h).
# Modèle marché courant : creuse la nuit, pleine en journée, pointe le soir
# (18 h–22 h, le pic de demande nationale). Surchargeable via ``hour_tranches``.
DEFAULT_HOUR_TRANCHES = [
    "creuse", "creuse", "creuse", "creuse", "creuse", "creuse",  # 00–05 h
    "creuse", "pleine", "pleine", "pleine", "pleine", "pleine",  # 06–11 h
    "pleine", "pleine", "pleine", "pleine", "pleine", "pleine",  # 12–17 h
    "pointe", "pointe", "pointe", "pointe", "pleine", "creuse",  # 18–23 h
]

# Tarifs TTC indicatifs (MAD/kWh) par tranche — valeurs marché conservatrices,
# à CONFIRMER par le founder selon le contrat ONEE réel. Surchargeable via
# ``tranche_tariffs``. La pointe est la plus chère, la creuse la moins chère.
DEFAULT_TRANCHE_TARIFFS = {
    "pointe": 1.45,
    "pleine": 1.15,
    "creuse": 0.85,
}

# Ordre canonique des tranches (du plus cher au moins cher) pour l'allocation
# du plafond annuel : on compense d'abord les kWh les plus chers.
_TRANCHE_ORDER = ["pointe", "pleine", "creuse"]


def net_metering_savings(injected_curve=None, import_curve=None, *,
                         hour_tranches=None, tranche_tariffs=None,
                         surplus_injecte_compense=True,
                         days_per_year=365, spill_tariff=None,
                         annual_cap_kwh=None,
                         compensation_ratio=1.0):
    """FG259 — économie annuelle du surplus injecté, valorisé par tranche TOU.

    Croise le SURPLUS INJECTÉ horaire (kWh/h, typiquement
    ``production[h] − autoconsommé[h]`` issu de :func:`hourly_self_consumption`)
    et le SOUTIRAGE RÉSEAU horaire (import, kWh/h) avec un découpage TIME-OF-USE
    (pointe / pleine / creuse) pour appliquer la règle de compensation
    net-metering (loi 13-09, MT/HT) :

        compensé[tranche] = min(injecté[tranche],
                                import[tranche] × compensation_ratio)
        économie[tranche] = compensé[tranche] × tarif[tranche]

    Le reste (``injecté − compensé``) est l'EXCÉDENT non absorbé par le
    soutirage de la même tranche : non valorisé par défaut, ou au tarif
    résiduel ``spill_tariff`` si fourni.

    Paramètres
    ----------
    injected_curve : itérable du surplus injecté horaire (kWh/h) — n'importe
        quelle longueur multiple de 24 (journée type, semaine, année 8760 h).
        Les heures sont mappées sur ``hour_tranches`` modulo 24. Valeurs
        illisibles/négatives → 0 (jamais de rejet).
    import_curve : itérable du soutirage réseau horaire (kWh/h), même mapping.
        Absent → soutirage nul → rien à compenser (économie 0).
    hour_tranches : liste de 24 libellés de tranche par heure (défaut
        ``DEFAULT_HOUR_TRANCHES``). Un libellé inconnu retombe sur « pleine ».
    tranche_tariffs : dict ``{tranche: MAD/kWh}`` (défaut
        ``DEFAULT_TRANCHE_TARIFFS``). Tarif manquant/illisible → 0.
    surplus_injecte_compense : réglage EXISTANT (toggle). False → l'injection
        n'est PAS compensée → économie 0, tout le surplus en « non compensé ».
    days_per_year : facteur d'annualisation si la courbe est une journée type
        (24 h → ×365). Si la courbe couvre déjà l'année (≥ 8760 h), passer 1.
    spill_tariff : tarif résiduel (MAD/kWh) du surplus EXCÉDENTAIRE non
        compensé (rachat éventuel) ; None → non valorisé (cas 13-09 MT).
    annual_cap_kwh : plafond annuel d'énergie compensée (kWh/an) ; None → pas
        de plafond. Appliqué des tranches les plus chères aux moins chères.
    compensation_ratio : fraction du soutirage compensable par l'injection
        (1.0 = compensation intégrale) ; borné à [0, 1].

    Retourne un dict JSON-sérialisable ::

        {compense, hours, periods, days_per_year,
         tranches: {nom: {injected_kwh, import_kwh, compensated_kwh,
                          spilled_kwh, tariff, savings_mad}},
         injected_kwh, import_kwh, compensated_kwh, spilled_kwh,
         annual_compensated_kwh, annual_injected_kwh,
         savings_mad_per_period, annual_savings_mad,
         annual_spill_value_mad, annual_cap_kwh, compensation_ratio,
         warnings: []}

    Ne lève jamais : toggle OFF / courbes vides / tarifs nuls → économie 0,
    division par zéro bornée.
    """
    warnings = []

    injected = _coerce_series(injected_curve)
    imported = _coerce_series(import_curve)

    tranches_by_hour = list(hour_tranches) if hour_tranches \
        else list(DEFAULT_HOUR_TRANCHES)
    if not tranches_by_hour:
        tranches_by_hour = list(DEFAULT_HOUR_TRANCHES)

    tariffs = {**DEFAULT_TRANCHE_TARIFFS, **(tranche_tariffs or {})}

    def _tariff(name):
        try:
            t = float(tariffs.get(name, 0.0))
        except (TypeError, ValueError):
            return 0.0
        return t if t >= 0 else 0.0

    try:
        ratio = float(compensation_ratio)
    except (TypeError, ValueError):
        ratio = 1.0
    if ratio < 0.0:
        ratio = 0.0
    if ratio > 1.0:
        ratio = 1.0

    try:
        days = float(days_per_year)
    except (TypeError, ValueError):
        days = 365.0
    if days <= 0:
        days = 1.0

    # ── Agrégation des flux PAR TRANCHE (modèle sur la période fournie) ──
    names = []
    for label in tranches_by_hour:
        key = (label or "pleine").lower()
        if key not in names:
            names.append(key)
    # Toujours exposer les tranches tarifées même si la courbe ne les touche pas.
    for key in tariffs:
        k = (key or "").lower()
        if k and k not in names:
            names.append(k)

    agg = {n: {"injected": 0.0, "import": 0.0} for n in names}

    n_hours = max(len(injected), len(imported))
    for h in range(n_hours):
        inj = injected[h] if h < len(injected) else 0.0
        imp = imported[h] if h < len(imported) else 0.0
        label = tranches_by_hour[h % len(tranches_by_hour)]
        key = (label or "pleine").lower()
        if key not in agg:
            key = "pleine"
            if "pleine" not in agg:
                agg["pleine"] = {"injected": 0.0, "import": 0.0}
                names.append("pleine")
        agg[key]["injected"] += inj
        agg[key]["import"] += imp

    compense = bool(surplus_injecte_compense)
    if not compense:
        warnings.append(
            "compensation de l'injection désactivée "
            "(surplus_injecte_compense = False) — économie nulle, surplus non "
            "valorisé")

    # ── Compensation par tranche (cap = soutirage de la tranche × ratio) ──
    tranche_out = {}
    total_injected = 0.0
    total_import = 0.0
    for name in names:
        inj = round(agg[name]["injected"], 6)
        imp = round(agg[name]["import"], 6)
        total_injected += inj
        total_import += imp
        if compense:
            compensable = min(inj, imp * ratio)
            if compensable < 0.0:
                compensable = 0.0
        else:
            compensable = 0.0
        tranche_out[name] = {
            "injected_kwh": round(inj, 3),
            "import_kwh": round(imp, 3),
            "eligible_kwh": round(compensable, 6),  # avant plafond annuel
            "compensated_kwh": 0.0,
            "spilled_kwh": 0.0,
            "tariff": round(_tariff(name), 4),
            "savings_mad": 0.0,
        }

    # ── Plafond annuel : allouer l'énergie compensée des tranches chères → pas ──
    # Le plafond est exprimé en kWh/AN ; on le convertit en kWh sur la PÉRIODE
    # fournie (÷ days) pour l'appliquer sur l'énergie de la période.
    cap_period = None
    try:
        if annual_cap_kwh is not None:
            cap_annual = float(annual_cap_kwh)
            if cap_annual >= 0:
                cap_period = cap_annual / days if days > 0 else cap_annual
    except (TypeError, ValueError):
        cap_period = None

    # Ordre d'allocation : tranches connues du plus cher au moins cher, puis le
    # reste par tarif décroissant (toute tranche custom).
    def _alloc_key(name):
        try:
            rank = _TRANCHE_ORDER.index(name)
        except ValueError:
            rank = len(_TRANCHE_ORDER)
        return (-_tariff(name), rank, name)

    remaining_cap = cap_period
    for name in sorted(names, key=_alloc_key):
        out = tranche_out[name]
        eligible = out["eligible_kwh"]
        if remaining_cap is not None:
            comp = min(eligible, max(0.0, remaining_cap))
            remaining_cap = max(0.0, remaining_cap - comp)
        else:
            comp = eligible
        spilled = max(0.0, out["injected_kwh"] - comp)
        out["compensated_kwh"] = round(comp, 3)
        out["spilled_kwh"] = round(spilled, 3)
        out["savings_mad"] = round(comp * _tariff(name), 2)

    if cap_period is not None and total_injected > 0:
        total_eligible = sum(t["eligible_kwh"] for t in tranche_out.values())
        if total_eligible > cap_period:
            warnings.append(
                "plafond annuel de compensation atteint — une partie du "
                "surplus éligible n'est pas valorisée")

    # ── Totaux sur la période + annualisation ──
    period_compensated = round(
        sum(t["compensated_kwh"] for t in tranche_out.values()), 3)
    period_spilled = round(
        sum(t["spilled_kwh"] for t in tranche_out.values()), 3)
    savings_per_period = round(
        sum(t["savings_mad"] for t in tranche_out.values()), 2)

    # Valeur du surplus excédentaire (spill) au tarif résiduel facultatif.
    spill_value_period = 0.0
    spill_rate = None
    if spill_tariff is not None:
        try:
            spill_rate = float(spill_tariff)
        except (TypeError, ValueError):
            spill_rate = None
        if spill_rate is not None and spill_rate < 0:
            spill_rate = 0.0
    if spill_rate:
        spill_value_period = round(period_spilled * spill_rate, 2)

    annual_savings = round(savings_per_period * days, 2)
    annual_spill_value = round(spill_value_period * days, 2)
    annual_compensated = round(period_compensated * days, 3)
    annual_injected = round(total_injected * days, 3)

    if compense and total_injected > 0 and period_compensated <= 0:
        warnings.append(
            "surplus injecté mais aucun soutirage simultané par tranche à "
            "compenser — vérifier la courbe d'import ou envisager du stockage")

    return {
        "compense": compense,
        "hours": n_hours,
        "periods": days,
        "days_per_year": round(days, 2),
        "tranches": tranche_out,
        "injected_kwh": round(total_injected, 3),
        "import_kwh": round(total_import, 3),
        "compensated_kwh": period_compensated,
        "spilled_kwh": period_spilled,
        "annual_injected_kwh": annual_injected,
        "annual_compensated_kwh": annual_compensated,
        "savings_mad_per_period": savings_per_period,
        "annual_savings_mad": annual_savings,
        "annual_spill_value_mad": annual_spill_value,
        "spill_tariff": spill_rate,
        "annual_cap_kwh": annual_cap_kwh,
        "compensation_ratio": round(ratio, 4),
        "warnings": warnings,
    }


# ── FG260 — Escalade tarifaire ONEE sur 20–25 ans + VAN/TRI ──────────────────
# Projette, année par année, la facture d'électricité ÉVITÉE (économie) sur un
# horizon long (20–25 ans) en tenant compte de DEUX dérives bien réelles :
#
#   * l'ESCALADE TARIFAIRE annuelle (le kWh ONEE renchérit chaque année — taux
#     éditable, ~6 %/an par défaut côté marché marocain) qui POUSSE l'économie
#     vers le HAUT (chaque kWh autoconsommé évite un kWh de plus en plus cher) ;
#   * la DÉGRADATION des modules (~0,5 %/an) qui ÉRODE la production donc
#     l'énergie évitée, TIRANT l'économie vers le bas.
#
# L'économie de l'année ``y`` (base 1) = économie_année1
#     × (1 + escalade)^(y-1)        ← le tarif monte
#     × (1 - dégradation)^(y-1)     ← la production baisse
#
# À partir du flux de trésorerie (économies annuelles − coût initial en année 0)
# on calcule la VAN (NPV, valeur actualisée nette au taux d'actualisation) et le
# TRI (IRR, taux qui annule la VAN), ce dernier par BISSECTION puis affinage
# Newton — STDLIB pure, jamais de boucle infinie (itérations plafonnées), renvoie
# None si le flux n'admet pas de TRI (pas de changement de signe / non-convergence).
#
# Module PUR : aucune base, aucun réseau, aucun prix d'achat/marge. Les entrées
# numériques ne sont JAMAIS rejetées (liberté de saisie du founder) — seules les
# valeurs absurdes sont bornées pour éviter une division par zéro.

# Taux d'escalade tarifaire ONEE annuel par défaut (éditable par l'appelant).
DEFAULT_TARIFF_ESCALATION = 0.06        # 6 %/an
# Dégradation annuelle des modules par défaut (lithium/silicium courant).
DEFAULT_MODULE_DEGRADATION = 0.005      # 0,5 %/an
# Taux d'actualisation par défaut pour la VAN.
DEFAULT_DISCOUNT_RATE = 0.05            # 5 %/an
# Horizon de projection par défaut (années).
DEFAULT_HORIZON_YEARS = 25
# Bornes de l'horizon admissible (le métier vise 20–25 ans, on tolère 1..40).
_MIN_HORIZON_YEARS = 1
_MAX_HORIZON_YEARS = 40


def _npv(rate, cashflows):
    """Valeur actualisée nette d'une suite de flux (flux[0] = année 0).

    ``cashflows[t]`` est actualisé par ``(1 + rate)**t``. ``rate`` = -1 est
    interdit (division par zéro) — renvoie ``float('inf')`` pour rester monotone
    côté solveur (sans lever).
    """
    if rate <= -1.0:
        return float("inf")
    total = 0.0
    for t, cf in enumerate(cashflows):
        total += cf / ((1.0 + rate) ** t)
    return total


def _irr(cashflows, *, low=-0.9999, high=10.0, tol=1e-7, max_iter=200):
    """TRI (taux annulant la VAN) par bissection + affinage Newton, ou None.

    Cherche la racine de ``_npv(rate)`` sur ``[low, high]`` quand la VAN y change
    de signe ; sinon renvoie None (flux sans TRI). Plafonné à ``max_iter`` — pas
    de boucle infinie. Un dernier pas de Newton affine la racine de bissection.
    """
    # Il faut au moins un flux positif ET un flux négatif pour qu'un TRI existe.
    if not any(cf > 0 for cf in cashflows) or not any(cf < 0 for cf in cashflows):
        return None

    f_low = _npv(low, cashflows)
    f_high = _npv(high, cashflows)
    if f_low == 0.0:
        return round(low, 6)
    if f_high == 0.0:
        return round(high, 6)
    # Pas de changement de signe sur l'intervalle → pas de racine bracketée ici.
    if (f_low > 0) == (f_high > 0):
        return None

    a, b = low, high
    fa = f_low
    rate = a
    for _ in range(max_iter):
        rate = (a + b) / 2.0
        fr = _npv(rate, cashflows)
        if abs(fr) < tol or (b - a) / 2.0 < tol:
            break
        if (fr > 0) == (fa > 0):
            a, fa = rate, fr
        else:
            b = rate

    # Affinage Newton (dérivée numérique) sans jamais sortir des bornes.
    for _ in range(20):
        fr = _npv(rate, cashflows)
        if abs(fr) < tol:
            break
        h = 1e-6
        deriv = (_npv(rate + h, cashflows) - fr) / h
        if deriv == 0.0:
            break
        step = fr / deriv
        new_rate = rate - step
        if not (low < new_rate < high):
            break
        rate = new_rate

    return round(rate, 6)


def tariff_escalation_projection(*, annual_savings_year1,
                                 upfront_cost=0.0,
                                 escalation_rate=DEFAULT_TARIFF_ESCALATION,
                                 degradation_rate=DEFAULT_MODULE_DEGRADATION,
                                 horizon_years=DEFAULT_HORIZON_YEARS,
                                 discount_rate=DEFAULT_DISCOUNT_RATE,
                                 baseline_bill_year1=None):
    """FG260 — projette facture/économies sur 20–25 ans + VAN (NPV) & TRI (IRR).

    Construit le tableau année par année du flux de trésorerie d'une installation
    PV : l'économie de l'année 1 escalade chaque année au rythme tarifaire ONEE
    (``escalation_rate``) tout en s'érodant de la dégradation modules
    (``degradation_rate``), face à un coût initial unique (``upfront_cost``) en
    année 0.

    Paramètres
    ----------
    annual_savings_year1 : économie (facture évitée) de la 1re année (MAD/an).
    upfront_cost : investissement initial TTC (MAD), placé en flux d'année 0.
    escalation_rate : taux d'escalade tarifaire annuel (éditable, défaut ~6 %).
    degradation_rate : dégradation annuelle de la production (défaut ~0,5 %).
    horizon_years : durée de projection (20–25 visés ; borné 1..40).
    discount_rate : taux d'actualisation pour la VAN (défaut 5 %).
    baseline_bill_year1 : facture ONEE de base year-1 (MAD/an), pour projeter la
        facture brute escaladée (optionnel ; sinon non renseignée).

    Retourne un dict JSON-sérialisable ::

        {schedule: [{year, escalated_tariff_factor, degradation_factor,
                     annual_savings, projected_bill, cumulative_savings,
                     net_cumulative, discounted_savings}],
         summary: {horizon_years, escalation_rate, degradation_rate,
                   discount_rate, upfront_cost, total_savings,
                   total_discounted_savings, npv, irr, payback_year,
                   discounted_payback_year, savings_year1, savings_last_year},
         warnings: []}

    Ne lève JAMAIS sur entrées dégradées : valeurs illisibles → 0, horizon hors
    bornes ramené dans [1, 40], division par zéro gardée, TRI = None si le flux
    n'en admet pas (pas de boucle infinie).
    """
    warnings = []

    def _num(value, default=0.0):
        try:
            return float(value)
        except (TypeError, ValueError):
            return float(default)

    savings1 = _num(annual_savings_year1)
    cost0 = _num(upfront_cost)
    esc = _num(escalation_rate, DEFAULT_TARIFF_ESCALATION)
    deg = _num(degradation_rate, DEFAULT_MODULE_DEGRADATION)
    disc = _num(discount_rate, DEFAULT_DISCOUNT_RATE)

    # Horizon : entier borné dans [1, 40] (le métier vise 20–25).
    try:
        horizon = int(round(_num(horizon_years, DEFAULT_HORIZON_YEARS)))
    except (TypeError, ValueError):
        horizon = DEFAULT_HORIZON_YEARS
    if horizon < _MIN_HORIZON_YEARS:
        horizon = _MIN_HORIZON_YEARS
        warnings.append("horizon ramené à 1 an (minimum)")
    elif horizon > _MAX_HORIZON_YEARS:
        horizon = _MAX_HORIZON_YEARS
        warnings.append("horizon plafonné à 40 ans (maximum)")

    base_bill1 = None
    if baseline_bill_year1 is not None:
        base_bill1 = _num(baseline_bill_year1)

    # Garde-fous métier (avertissements, jamais de rejet).
    if esc < 0:
        warnings.append("taux d'escalade négatif — le tarif baisse, inhabituel")
    if not (0.0 <= deg < 1.0):
        warnings.append("taux de dégradation hors [0, 1[ — vérifier la saisie")
    if disc <= -1.0:
        warnings.append(
            "taux d'actualisation ≤ -100 % — VAN non calculable, ramené à 0")
        disc = 0.0

    schedule = []
    cumulative = 0.0
    discounted_total = 0.0
    payback_year = None
    discounted_payback_year = None
    discounted_cumulative_net = -cost0
    # Flux de trésorerie : année 0 = -coût initial, puis économies escaladées.
    cashflows = [-cost0]

    for y in range(1, horizon + 1):
        esc_factor = (1.0 + esc) ** (y - 1)
        # Dégradation bornée : (1 - deg) ne doit pas devenir négatif.
        deg_step = 1.0 - deg if deg < 1.0 else 0.0
        deg_factor = deg_step ** (y - 1)
        annual_savings = savings1 * esc_factor * deg_factor

        projected_bill = None
        if base_bill1 is not None:
            # La facture ONEE brute escalade au même rythme tarifaire.
            projected_bill = round(base_bill1 * esc_factor, 2)

        cumulative += annual_savings
        net_cumulative = cumulative - cost0

        # Actualisation : le flux de l'année y est divisé par (1+disc)^y.
        if disc > -1.0:
            discounted = annual_savings / ((1.0 + disc) ** y)
        else:
            discounted = 0.0
        discounted_total += discounted
        discounted_cumulative_net += discounted

        cashflows.append(annual_savings)

        if payback_year is None and net_cumulative >= 0:
            payback_year = y
        if discounted_payback_year is None and discounted_cumulative_net >= 0:
            discounted_payback_year = y

        schedule.append({
            "year": y,
            "escalated_tariff_factor": round(esc_factor, 6),
            "degradation_factor": round(deg_factor, 6),
            "annual_savings": round(annual_savings, 2),
            "projected_bill": projected_bill,
            "cumulative_savings": round(cumulative, 2),
            "net_cumulative": round(net_cumulative, 2),
            "discounted_savings": round(discounted, 2),
        })

    npv = round(_npv(disc, cashflows), 2)
    irr = _irr(cashflows)

    if payback_year is None and cost0 > 0:
        warnings.append(
            "retour sur investissement non atteint sur l'horizon — l'économie "
            "cumulée ne couvre pas le coût initial")
    if irr is None and cost0 > 0:
        warnings.append(
            "TRI non calculable (flux sans changement de signe ou non "
            "convergent) — vérifier coût initial et économies")

    summary = {
        "horizon_years": horizon,
        "escalation_rate": round(esc, 6),
        "degradation_rate": round(deg, 6),
        "discount_rate": round(disc, 6),
        "upfront_cost": round(cost0, 2),
        "savings_year1": round(savings1, 2),
        "savings_last_year": (
            round(schedule[-1]["annual_savings"], 2) if schedule else 0.0),
        "total_savings": round(cumulative, 2),
        "total_discounted_savings": round(discounted_total, 2),
        "net_total": round(cumulative - cost0, 2),
        "npv": npv,
        "irr": irr,
        "payback_year": payback_year,
        "discounted_payback_year": discounted_payback_year,
    }

    return {
        "schedule": schedule,
        "summary": summary,
        "warnings": warnings,
    }


# ── FG261 — Optimisation de la puissance souscrite (C&I) après PV ─────────────
# Pour un client COMMERCIAL / INDUSTRIEL facturé sur une PUISSANCE SOUSCRITE
# (kVA ou kW, « prime fixe » / redevance de puissance MT-BT) : le PV écrête la
# pointe de soutirage RÉSEAU pendant les heures ensoleillées. En croisant la
# courbe de charge horaire avec la production PV horaire, on calcule la demande
# RÉSEAU nette heure par heure (charge − PV, plancher 0), on en tire la NOUVELLE
# pointe de soutirage, et on recommande une puissance souscrite réduite. La
# baisse de redevance de puissance (MAD/kVA/an) est alors chiffrée.
#
# RÈGLE :
#   * demande_réseau[h] = max(0, charge[h] − production[h])  (en kW)
#   * pointe_post_pv = max(demande_réseau)  ;  pointe_pre_pv = max(charge)
#   * puissance recommandée = ceil(pointe_post_pv × marge_sécurité)
#       — jamais SUPÉRIEURE à la souscription actuelle (on n'augmente pas) ;
#       — bornée au plus à la souscription actuelle même si la pointe ne baisse
#         pas (recommandation = pas de réduction → économie 0).
#   * facteur de puissance / conversion kW→kVA : si ``power_factor`` est fourni
#     (cos φ), la demande en kVA = demande_kW / cos φ ; sinon on raisonne dans
#     l'unité de la souscription telle quelle (la souscription est déjà en kVA
#     ou en kW selon le contrat — on ne convertit que si on a le cos φ).
#   * économie annuelle = (souscription_actuelle − recommandée) × tarif_capacité.
#     Le tarif est en MAD par unité de puissance et par AN (``capacity_tariff``,
#     ``tariff_period="year"``) ou par MOIS (``tariff_period="month"`` → ×12).
#
# Module PUR : aucune base, aucun réseau, aucun prix d'achat/marge. Les entrées
# numériques ne sont JAMAIS rejetées (liberté de saisie du founder) — les
# valeurs illisibles/négatives sont ramenées à 0, jamais d'exception, division
# par zéro bornée.

# Marge de sécurité par défaut sur la pointe post-PV pour dimensionner la
# souscription recommandée (aléas météo, croissance de charge, démarrages).
DEFAULT_SUBSCRIBED_SAFETY_MARGIN = 1.10


def optimize_subscribed_power(load_curve=None, production_curve=None, *,
                              current_subscribed_kva=None,
                              capacity_tariff=0.0,
                              tariff_period="year",
                              safety_margin=DEFAULT_SUBSCRIBED_SAFETY_MARGIN,
                              power_factor=None,
                              daily_load_kwh=None,
                              daily_production_kwh=None,
                              load_profile="commercial"):
    """FG261 — recommande une puissance souscrite réduite après PV (C&I).

    Croise la COURBE DE CHARGE horaire (kW/h ≈ kWh par heure) et la PRODUCTION
    PV horaire (kW/h) pour calculer la demande RÉSEAU nette heure par heure
    (``max(0, charge − PV)``), en déduit la pointe de soutirage POST-PV, puis
    recommande une puissance souscrite réduite et chiffre l'économie sur la
    redevance de puissance (prime fixe).

    Paramètres
    ----------
    load_curve : itérable de demandes horaires (kW, ≈ kWh/h) — 24 h, 8760 h ou
        toute longueur. Valeurs illisibles/négatives → 0 (jamais de rejet). Si
        absent, on synthétise un profil type (``load_profile``) calé sur
        ``daily_load_kwh``.
    production_curve : production PV horaire (kW/h). Si absent, on synthétise
        ``TYPICAL_PV_PROFILE`` calé sur ``daily_production_kwh``.
    current_subscribed_kva : puissance souscrite ACTUELLE (kVA ou kW selon le
        contrat). Si absente/illisible → la pointe pré-PV sert de référence.
    capacity_tariff : redevance de puissance (MAD par unité de puissance et par
        période ``tariff_period``).
    tariff_period : ``"year"`` (défaut) ou ``"month"`` (×12 pour l'annuel).
    safety_margin : marge sur la pointe post-PV pour dimensionner la
        souscription recommandée (défaut 1.10). Bornée ≥ 1.0.
    power_factor : cos φ facultatif — convertit la demande kW en kVA
        (kVA = kW / cos φ) quand la souscription est en kVA. Si absent, on
        raisonne dans l'unité de la souscription telle quelle.
    daily_load_kwh / daily_production_kwh : énergies journalières pour générer
        les profils type quand une courbe n'est pas fournie.
    load_profile : clé de profil type de charge (repli) — ``"commercial"`` par
        défaut (la cible C&I).

    Retourne un dict JSON-sérialisable ::

        {hours, peak_pre_pv_kw, peak_post_pv_kw, peak_reduction_kw,
         peak_reduction_pct, demand_unit, power_factor,
         peak_post_pv_demand, current_subscribed, recommended_subscribed,
         subscribed_reduction, annual_capacity_tariff,
         annual_saving, monthly_saving, load_source, production_source,
         warnings: []}

    Ne lève jamais : courbes vides → pointe 0 ; pas de réduction possible →
    recommandation = souscription actuelle et économie 0 ; division par zéro
    bornée (cos φ ≤ 0 ignoré).
    """
    warnings = []

    load = _coerce_series(load_curve)
    load_source = "courbe fournie"
    if not load:
        load = _scaled_typical_load(daily_load_kwh, load_profile)
        load_source = f"profil type ({load_profile})"

    prod = _coerce_series(production_curve)
    production_source = "courbe fournie"
    if not prod:
        prod = _scaled_typical_pv(daily_production_kwh)
        production_source = "profil type PV"

    # Alignement des longueurs (on borne sur la plus courte).
    n = min(len(load), len(prod)) if prod else len(load)
    if load and prod and len(load) != len(prod):
        warnings.append(
            f"courbes de longueurs différentes (charge={len(load)} h, "
            f"production={len(prod)} h) — alignées sur {n} h")

    # ── Pointe de soutirage réseau pré-PV et post-PV (en kW) ──
    peak_pre = 0.0
    peak_post = 0.0
    if prod:
        for i in range(n):
            c = load[i]
            net = c - prod[i]
            if net < 0.0:
                net = 0.0
            if c > peak_pre:
                peak_pre = c
            if net > peak_post:
                peak_post = net
    else:
        # Aucune production fournie ni synthétisable : pointe post-PV = pré-PV.
        for c in load:
            if c > peak_pre:
                peak_pre = c
        peak_post = peak_pre

    peak_pre = round(peak_pre, 3)
    peak_post = round(peak_post, 3)
    peak_reduction_kw = round(max(0.0, peak_pre - peak_post), 3)
    peak_reduction_pct = (
        round(peak_reduction_kw / peak_pre * 100.0, 1) if peak_pre > 0 else 0.0)

    # ── Conversion kW → kVA si cos φ fourni (souscription en kVA) ──
    pf = None
    try:
        pf_val = float(power_factor)
        if 0.0 < pf_val <= 1.0:
            pf = pf_val
    except (TypeError, ValueError):
        pf = None
    if pf is not None:
        demand_unit = "kVA"
        peak_post_demand = round(peak_post / pf, 3)
    else:
        demand_unit = "kW/kVA"
        peak_post_demand = peak_post

    # ── Souscription actuelle (référence) ──
    try:
        current = float(current_subscribed_kva)
    except (TypeError, ValueError):
        current = None
    if current is None or current <= 0:
        # Pas de souscription fournie : on prend la pointe pré-PV convertie
        # comme référence (l'économie est alors purement indicative).
        ref = peak_pre / pf if pf else peak_pre
        current = round(ref, 3)
        if current > 0:
            warnings.append(
                "puissance souscrite actuelle non fournie — pointe pré-PV "
                "utilisée comme référence (économie indicative)")

    # ── Marge de sécurité ──
    try:
        margin = float(safety_margin)
    except (TypeError, ValueError):
        margin = DEFAULT_SUBSCRIBED_SAFETY_MARGIN
    if margin < 1.0:
        margin = 1.0

    # Souscription recommandée = ceil(pointe post-PV × marge), bornée à la
    # souscription actuelle (on ne RECOMMANDE jamais d'augmenter).
    recommended_raw = peak_post_demand * margin
    recommended = int(math.ceil(recommended_raw)) if recommended_raw > 0 else 0
    if current > 0 and recommended > current:
        # La pointe post-PV (× marge) dépasse déjà la souscription : aucune
        # réduction possible — on garde la souscription actuelle.
        recommended = round(current, 3)
        warnings.append(
            "la pointe réseau après PV (avec marge) atteint la puissance "
            "souscrite actuelle — aucune réduction recommandée")

    subscribed_reduction = round(max(0.0, current - recommended), 3)

    # ── Tarif annuel de capacité ──
    try:
        tariff = float(capacity_tariff)
    except (TypeError, ValueError):
        tariff = 0.0
    if tariff < 0.0:
        tariff = 0.0
    annual_tariff = tariff * 12.0 if (tariff_period or "year") == "month" \
        else tariff

    annual_saving = round(subscribed_reduction * annual_tariff, 2)
    monthly_saving = round(annual_saving / 12.0, 2)

    if peak_reduction_kw <= 0 and prod:
        warnings.append(
            "le PV n'écrête pas la pointe de soutirage (pointe hors heures "
            "solaires) — pas de gain sur la puissance souscrite")
    if subscribed_reduction > 0 and tariff <= 0:
        warnings.append(
            "réduction de puissance possible mais tarif de capacité non "
            "fourni — économie non chiffrée (renseigner capacity_tariff)")

    return {
        "hours": n,
        "peak_pre_pv_kw": peak_pre,
        "peak_post_pv_kw": peak_post,
        "peak_reduction_kw": peak_reduction_kw,
        "peak_reduction_pct": peak_reduction_pct,
        "demand_unit": demand_unit,
        "power_factor": pf,
        "peak_post_pv_demand": peak_post_demand,
        "current_subscribed": round(current, 3),
        "recommended_subscribed": recommended,
        "subscribed_reduction": subscribed_reduction,
        "safety_margin": round(margin, 4),
        "annual_capacity_tariff": round(annual_tariff, 4),
        "annual_saving": annual_saving,
        "monthly_saving": monthly_saving,
        "load_source": load_source,
        "production_source": production_source,
        "warnings": warnings,
    }
