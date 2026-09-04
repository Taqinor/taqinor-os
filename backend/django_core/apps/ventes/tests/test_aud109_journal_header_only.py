"""AUD109 — le journal des ventes n'OMET plus les factures header-only.

``export_journal_ventes`` bouclait ``for ligne in f.lignes.all():`` sans aucun
``else`` : une facture ÉMISE sans ``LigneFacture`` n'écrivait ZÉRO ligne dans
la feuille « Journal des ventes » et n'ajoutait ZÉRO montant au résumé TVA —
alors qu'elle est bien captée par le filtre ``ISSUED_STATUTS``, donc présente
dans le queryset. Ses deux jumeaux du même fichier (``_compta_rows``,
``_grand_livre_rows``) avaient déjà le filet.

Ces factures existent réellement en production : ``contrats.services`` crée des
factures ÉMISES à montants figés et SANS ligne (abonnements/maintenance
récurrents), tout comme les tranches d'échéancier — tout ce CA disparaissait du
journal du mois, et le cabinet recevait un CA sous-déclaré.
"""
from datetime import date
from decimal import Decimal

from django.test import TestCase

from apps.crm.models import Client
from apps.stock.models import Produit
from apps.ventes.exports import export_journal_ventes
from apps.ventes.models import Facture, LigneFacture
from authentication.models import Company

_CTR = [0]


def _nxt():
    _CTR[0] += 1
    return _CTR[0]


def _feuilles(reponse):
    """Le classeur .xlsx renvoyé, relu — on lit la SORTIE, pas le calcul."""
    import io

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(reponse.content))
    return wb


class TestJournalVentesHeaderOnly(TestCase):
    DEBUT = date(2026, 3, 1)
    FIN = date(2026, 4, 1)

    def setUp(self):
        self.company = Company.objects.create(
            nom='AUD109 Co', slug=f'aud109-{_nxt()}')
        self.client_obj = Client.objects.create(
            company=self.company, nom='AUD109', prenom='Client',
            telephone='+212600000110')
        self.produit = Produit.objects.create(
            company=self.company, nom='Onduleur', sku=f'AUD109-{_nxt()}',
            prix_vente=Decimal('5000'), quantite_stock=10)

    def _facture(self, *, montants=None, avec_ligne=False):
        facture = Facture.objects.create(
            company=self.company, reference=f'FAC-AUD109-{_nxt()}',
            client=self.client_obj, statut=Facture.Statut.EMISE,
            taux_tva=Decimal('20.00'),
            libelle='Abonnement maintenance' if montants else '',
            montant_ht=montants[0] if montants else None,
            montant_tva=montants[1] if montants else None,
            montant_ttc=montants[2] if montants else None)
        Facture.objects.filter(pk=facture.pk).update(
            date_emission=date(2026, 3, 15))
        facture.refresh_from_db()
        if avec_ligne:
            LigneFacture.objects.create(
                facture=facture, produit=self.produit,
                designation='Onduleur', quantite=Decimal('1'),
                prix_unitaire=Decimal('5000'), taux_tva=Decimal('20.00'))
        return facture

    def _journal(self):
        return _feuilles(
            export_journal_ventes(self.company, self.DEBUT, self.FIN))

    def _resume_tva(self, wb):
        """{taux: (base_ht, tva, ttc)} + la ligne TOTAL de « Résumé TVA »."""
        ws = wb['Résumé TVA']
        lignes = [tuple(c.value for c in row) for row in ws.iter_rows(min_row=2)]
        par_taux = {}
        total = None
        for ligne in lignes:
            if not ligne or ligne[0] is None:
                continue
            if ligne[0] == 'TOTAL':
                total = ligne[1:4]
            else:
                par_taux[ligne[0]] = ligne[1:4]
        return par_taux, total

    def test_facture_header_only_apparait_dans_le_journal(self):
        self._facture(montants=(Decimal('10000'), Decimal('2000'),
                                Decimal('12000')))
        wb = self._journal()
        ws = wb['Journal des ventes']
        lignes = [tuple(c.value for c in row)
                  for row in ws.iter_rows(min_row=2)]
        self.assertEqual(len(lignes), 1, lignes)
        # Colonnes : ... Total HT (index 8), TVA % (9), Montant TVA (10), TTC (11)
        self.assertEqual(lignes[0][8], 10000.0)
        self.assertEqual(lignes[0][10], 2000.0)
        self.assertEqual(lignes[0][11], 12000.0)

    def test_le_resume_tva_ventile_la_facture_header_only(self):
        self._facture(montants=(Decimal('10000'), Decimal('2000'),
                                Decimal('12000')))
        par_taux, total = self._resume_tva(self._journal())
        self.assertEqual(total, (10000.0, 2000.0, 12000.0))
        self.assertEqual(par_taux.get('20.00 %'), (10000.0, 2000.0, 12000.0))

    def test_non_regression_une_facture_a_lignes_produit_les_memes_lignes(self):
        self._facture(avec_ligne=True)
        wb = self._journal()
        ws = wb['Journal des ventes']
        lignes = [tuple(c.value for c in row)
                  for row in ws.iter_rows(min_row=2)]
        self.assertEqual(len(lignes), 1)
        self.assertEqual(lignes[0][5], 'Onduleur')
        self.assertEqual(lignes[0][6], '1.00')
        self.assertEqual(lignes[0][8], 5000.0)
        self.assertEqual(lignes[0][10], 1000.0)
        self.assertEqual(lignes[0][11], 6000.0)

    def test_les_deux_types_de_facture_coexistent_dans_le_total(self):
        self._facture(avec_ligne=True)
        self._facture(montants=(Decimal('10000'), Decimal('2000'),
                                Decimal('12000')))
        _par_taux, total = self._resume_tva(self._journal())
        self.assertEqual(total, (15000.0, 3000.0, 18000.0))
