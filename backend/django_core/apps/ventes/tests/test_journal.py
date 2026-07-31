"""T12 — export comptable journal des ventes + résumé TVA."""
from datetime import date
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import TestCase
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import AccessToken

from apps.crm.models import Client
from apps.stock.models import Produit
from apps.ventes.models import Facture, LigneFacture
from apps.ventes.exports import period_bounds
from authentication.models import Company

User = get_user_model()


class TestPeriodBounds(TestCase):
    def test_month(self):
        d, f = period_bounds({'month': '2026-06'})
        self.assertEqual((d, f), (date(2026, 6, 1), date(2026, 7, 1)))

    def test_quarter(self):
        d, f = period_bounds({'quarter': '2026-Q2'})
        self.assertEqual((d, f), (date(2026, 4, 1), date(2026, 7, 1)))


class TestJournalExport(TestCase):
    def setUp(self):
        self.company = Company.objects.get_or_create(
            slug='jr-co', defaults={'nom': 'JR Co'})[0]
        self.user = User.objects.create_user(
            username='jr_u', password='x', role_legacy='responsable',
            company=self.company)
        self.api = APIClient()
        self.api.credentials(
            HTTP_AUTHORIZATION=f'Bearer {AccessToken.for_user(self.user)}')
        c = Client.objects.create(company=self.company, nom='C', ice='000111222')
        p = Produit.objects.create(company=self.company, nom='Panneau', sku='J-1',
                                   prix_vente=Decimal('1000'), quantite_stock=5)
        fac = Facture.objects.create(
            company=self.company, reference='FAC-JR-1', client=c, statut='emise',
            taux_tva=Decimal('20'))
        LigneFacture.objects.create(
            facture=fac, produit=p, designation='Panneaux', quantite=Decimal('10'),
            prix_unitaire=Decimal('1000'), remise=Decimal('0'), taux_tva=Decimal('10'))
        LigneFacture.objects.create(
            facture=fac, produit=p, designation='Pose', quantite=Decimal('1'),
            prix_unitaire=Decimal('2000'), remise=Decimal('0'), taux_tva=Decimal('20'))

    def test_export_returns_xlsx(self):
        resp = self.api.get('/api/django/ventes/journal-ventes/?month=%s'
                            % date.today().strftime('%Y-%m'))
        self.assertEqual(resp.status_code, 200)
        body = b''.join(resp.streaming_content) if resp.streaming else resp.content
        self.assertTrue(body.startswith(b'PK'))
        self.assertGreater(len(body), 2000)

    def _load_journal_rows(self):
        from datetime import date as _date
        from apps.ventes.exports import export_journal_ventes, period_bounds
        from openpyxl import load_workbook
        from io import BytesIO
        debut, fin = period_bounds({'month': _date.today().strftime('%Y-%m')})
        resp = export_journal_ventes(self.company, debut, fin)
        body = b''.join(resp.streaming_content) if resp.streaming else resp.content
        wb = load_workbook(BytesIO(body))
        ws = wb['Journal des ventes']
        return [list(r) for r in ws.iter_rows(values_only=True)]

    def test_type_column_present(self):
        rows = self._load_journal_rows()
        self.assertEqual(rows[0][2], 'Type')
        # Les factures classiques (chaîne BC) sont de type « Facture complète ».
        self.assertEqual(rows[1][2], 'Facture complète')

    def test_avoirs_appear_as_negative_lines(self):
        from apps.ventes.models import Avoir, LigneAvoir
        from apps.crm.models import Client
        fac = Facture.objects.get(reference='FAC-JR-1')
        avoir = Avoir.objects.create(
            company=self.company, reference='AVO-JR-1', facture=fac,
            client=Client.objects.get(nom='C'), statut='emise',
            taux_tva=Decimal('20'))
        LigneAvoir.objects.create(
            avoir=avoir, designation='Retour', quantite=Decimal('1'),
            prix_unitaire=Decimal('2000'), remise=Decimal('0'),
            taux_tva=Decimal('20'))
        rows = self._load_journal_rows()
        avoir_rows = [r for r in rows if r[2] == 'Avoir']
        self.assertEqual(len(avoir_rows), 1)
        # Total HT (col index 8) et Total TTC (col index 11) sont négatifs.
        self.assertLess(avoir_rows[0][8], 0)
        self.assertLess(avoir_rows[0][11], 0)

    def test_export_comptable_xlsx(self):
        from datetime import timedelta
        d = date.today()
        start = d.strftime('%Y-%m-01')
        end = (d + timedelta(days=1)).strftime('%Y-%m-%d')
        resp = self.api.get(
            f'/api/django/ventes/export-comptable/?start={start}&end={end}'
            f'&fmt=xlsx')
        self.assertEqual(resp.status_code, 200)
        body = (b''.join(resp.streaming_content)
                if resp.streaming else resp.content)
        self.assertTrue(body.startswith(b'PK'))
        # Le contenu contient l'ICE client + une ligne TOTAL.
        from openpyxl import load_workbook
        from io import BytesIO
        ws = load_workbook(BytesIO(body)).active
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        self.assertEqual(rows[0][4], 'ICE client')
        self.assertIn('000111222', [r[4] for r in rows[1:] if r[4]])
        self.assertTrue(any(r[0] == 'TOTAL' for r in rows))

    def test_export_comptable_csv(self):
        from datetime import timedelta
        d = date.today()
        start = d.strftime('%Y-%m-01')
        end = (d + timedelta(days=1)).strftime('%Y-%m-%d')
        resp = self.api.get(
            f'/api/django/ventes/export-comptable/?start={start}&end={end}'
            f'&fmt=csv')
        self.assertEqual(resp.status_code, 200)
        self.assertIn('text/csv', resp['Content-Type'])
        text = resp.content.decode('utf-8')
        self.assertIn('ICE client', text)
        self.assertIn('000111222', text)
        self.assertIn('TOTAL', text)

    def test_export_comptable_excludes_brouillon(self):
        # Une facture brouillon ne doit jamais apparaître dans l'export.
        from apps.crm.models import Client
        Facture.objects.create(
            company=self.company, reference='FAC-JR-DRAFT',
            client=Client.objects.get(nom='C'), statut='brouillon',
            taux_tva=Decimal('20'))
        from datetime import timedelta
        d = date.today()
        start = d.strftime('%Y-%m-01')
        end = (d + timedelta(days=1)).strftime('%Y-%m-%d')
        resp = self.api.get(
            f'/api/django/ventes/export-comptable/?start={start}&end={end}'
            f'&fmt=csv')
        self.assertNotIn('FAC-JR-DRAFT', resp.content.decode('utf-8'))


class TestGrandLivreExport(TestCase):
    """FG49 — grand-livre codé par compte CGNC (3421 / 7111 / 4455)."""

    def setUp(self):
        self.company = Company.objects.get_or_create(
            slug='gl-co', defaults={'nom': 'GL Co'})[0]
        self.user = User.objects.create_user(
            username='gl_u', password='x', role_legacy='responsable',
            company=self.company)
        self.api = APIClient()
        self.api.credentials(
            HTTP_AUTHORIZATION=f'Bearer {AccessToken.for_user(self.user)}')
        self.client_obj = Client.objects.create(
            company=self.company, nom='GLClient', ice='009988776')
        p = Produit.objects.create(
            company=self.company, nom='Panneau', sku='GL-1',
            prix_vente=Decimal('1000'), quantite_stock=5)
        fac = Facture.objects.create(
            company=self.company, reference='FAC-GL-1', client=self.client_obj,
            statut='emise', taux_tva=Decimal('20'))
        # 10000 HT @ 10 % → 1000 TVA ; 2000 HT @ 20 % → 400 TVA.
        LigneFacture.objects.create(
            facture=fac, produit=p, designation='Panneaux',
            quantite=Decimal('10'), prix_unitaire=Decimal('1000'),
            remise=Decimal('0'), taux_tva=Decimal('10'))
        LigneFacture.objects.create(
            facture=fac, produit=p, designation='Pose', quantite=Decimal('1'),
            prix_unitaire=Decimal('2000'), remise=Decimal('0'),
            taux_tva=Decimal('20'))

    def _params(self):
        from datetime import timedelta
        d = date.today()
        return (d.strftime('%Y-%m-01'),
                (d + timedelta(days=1)).strftime('%Y-%m-%d'))

    def _gl_rows(self):
        from datetime import date as _date
        from apps.ventes.exports import (
            export_grand_livre_xlsx, period_bounds)
        from openpyxl import load_workbook
        from io import BytesIO
        debut, fin = period_bounds(
            {'month': _date.today().strftime('%Y-%m')})
        resp = export_grand_livre_xlsx(self.company, debut, fin)
        body = (b''.join(resp.streaming_content)
                if resp.streaming else resp.content)
        ws = load_workbook(BytesIO(body)).active
        return [list(r) for r in ws.iter_rows(values_only=True)]

    def test_default_account_codes(self):
        from apps.ventes.exports import account_codes_for
        codes = account_codes_for(self.company)
        self.assertEqual(codes['ventes'], '7111')
        self.assertEqual(codes['tva_collectee'], '4455')
        self.assertEqual(codes['clients'], '3421')

    def test_endpoint_returns_xlsx(self):
        start, end = self._params()
        resp = self.api.get(
            f'/api/django/ventes/export-comptable/?start={start}&end={end}'
            f'&layout=grand-livre&fmt=xlsx')
        self.assertEqual(resp.status_code, 200)
        body = (b''.join(resp.streaming_content)
                if resp.streaming else resp.content)
        self.assertTrue(body.startswith(b'PK'))

    def test_account_coded_lines_present(self):
        rows = self._gl_rows()
        header = rows[0]
        self.assertEqual(header[0], 'Compte')
        codes = [str(r[0]) for r in rows[1:] if r[0]]
        # Les trois comptes CGNC apparaissent.
        self.assertIn('3421', codes)
        self.assertIn('7111', codes)
        self.assertIn('4455', codes)

    def test_entries_balance(self):
        rows = self._gl_rows()
        # Colonnes Débit (9) / Crédit (10) ; somme débit == somme crédit.
        tot_debit = sum(r[9] for r in rows[1:]
                        if r[0] not in (None, '', 'TOTAL') and r[9])
        tot_credit = sum(r[10] for r in rows[1:]
                         if r[0] not in (None, '', 'TOTAL') and r[10])
        self.assertEqual(round(tot_debit, 2), round(tot_credit, 2))
        # TTC attendu : 10000 + 2000 + 1000 + 400 = 13400.
        self.assertEqual(round(tot_debit, 2), 13400.0)

    def test_client_debit_is_ttc(self):
        rows = self._gl_rows()
        clients_lines = [r for r in rows[1:] if str(r[0]) == '3421']
        self.assertEqual(len(clients_lines), 1)
        # 3421 Clients est en DÉBIT (col 9) pour le TTC, crédit (col 10) nul.
        self.assertEqual(round(clients_lines[0][9], 2), 13400.0)
        self.assertFalse(clients_lines[0][10])
        # ICE du tiers reporté.
        self.assertEqual(clients_lines[0][7], '009988776')

    def test_configurable_account_codes(self):
        from django.test import override_settings
        from apps.ventes.exports import account_codes_for
        with override_settings(
                VENTES_COMPTA_ACCOUNT_CODES={'ventes': '7121'}):
            codes = account_codes_for(self.company)
            self.assertEqual(codes['ventes'], '7121')
            # Les autres restent aux défauts.
            self.assertEqual(codes['clients'], '3421')

    def test_per_company_account_codes(self):
        from django.test import override_settings
        from apps.ventes.exports import account_codes_for
        with override_settings(
                VENTES_COMPTA_ACCOUNT_CODES_BY_COMPANY={
                    'gl-co': {'clients': '3424'}}):
            codes = account_codes_for(self.company)
            self.assertEqual(codes['clients'], '3424')

    def test_csv_layout_grand_livre(self):
        start, end = self._params()
        resp = self.api.get(
            f'/api/django/ventes/export-comptable/?start={start}&end={end}'
            f'&layout=grand-livre&fmt=csv')
        self.assertEqual(resp.status_code, 200)
        self.assertIn('text/csv', resp['Content-Type'])
        text = resp.content.decode('utf-8')
        self.assertIn('3421', text)
        self.assertIn('7111', text)
        self.assertIn('4455', text)

    def test_avoir_reverses_signs(self):
        from apps.ventes.models import Avoir, LigneAvoir
        fac = Facture.objects.get(reference='FAC-GL-1')
        avoir = Avoir.objects.create(
            company=self.company, reference='AVO-GL-1', facture=fac,
            client=self.client_obj, statut='emise', taux_tva=Decimal('20'))
        LigneAvoir.objects.create(
            avoir=avoir, designation='Retour', quantite=Decimal('1'),
            prix_unitaire=Decimal('2000'), remise=Decimal('0'),
            taux_tva=Decimal('20'))
        rows = self._gl_rows()
        # L'avoir met 3421 Clients en CRÉDIT (col 10) — contre-passation.
        clients_lines = [r for r in rows[1:] if str(r[0]) == '3421']
        credit_clients = [r for r in clients_lines if r[10]]
        self.assertTrue(credit_clients)
        # Le grand-livre reste équilibré avec l'avoir.
        tot_debit = sum(r[9] for r in rows[1:]
                        if r[0] not in (None, '', 'TOTAL') and r[9])
        tot_credit = sum(r[10] for r in rows[1:]
                         if r[0] not in (None, '', 'TOTAL') and r[10])
        self.assertEqual(round(tot_debit, 2), round(tot_credit, 2))


class TestQuickBooksIifExport(TestCase):
    """NTAPI37 — pont comptable QuickBooks (.iif, à côté de FG49/FG377) :
    MÊMES écritures que le grand-livre (FG49), sérialisées en IIF General
    Journal. Réutilise le fixture de TestGrandLivreExport (2 lignes, TVA
    10 %/20 %)."""

    def setUp(self):
        self.company = Company.objects.get_or_create(
            slug='qb-co', defaults={'nom': 'QB Co'})[0]
        self.user = User.objects.create_user(
            username='qb_u', password='x', role_legacy='responsable',
            company=self.company)
        self.api = APIClient()
        self.api.credentials(
            HTTP_AUTHORIZATION=f'Bearer {AccessToken.for_user(self.user)}')
        client_obj = Client.objects.create(
            company=self.company, nom='QBClient', ice='001122334')
        p = Produit.objects.create(
            company=self.company, nom='Panneau', sku='QB-1',
            prix_vente=Decimal('1000'), quantite_stock=5)
        self.fac = Facture.objects.create(
            company=self.company, reference='FAC-QB-1', client=client_obj,
            statut='emise', taux_tva=Decimal('20'))
        LigneFacture.objects.create(
            facture=self.fac, produit=p, designation='Panneaux',
            quantite=Decimal('10'), prix_unitaire=Decimal('1000'),
            remise=Decimal('0'), taux_tva=Decimal('10'))
        LigneFacture.objects.create(
            facture=self.fac, produit=p, designation='Pose',
            quantite=Decimal('1'), prix_unitaire=Decimal('2000'),
            remise=Decimal('0'), taux_tva=Decimal('20'))

    def _params(self):
        from datetime import timedelta
        d = date.today()
        return (d.strftime('%Y-%m-01'),
                (d + timedelta(days=1)).strftime('%Y-%m-%d'))

    def test_endpoint_returns_iif_file(self):
        start, end = self._params()
        resp = self.api.get(
            f'/api/django/ventes/export-comptable/?start={start}&end={end}'
            f'&layout=quickbooks')
        self.assertEqual(resp.status_code, 200)
        self.assertIn('application/octet-stream', resp['Content-Type'])
        self.assertIn('.iif', resp['Content-Disposition'])
        text = resp.content.decode('utf-8')
        self.assertTrue(text.startswith('!TRNS\t'))
        self.assertIn('!SPL\t', text)
        self.assertIn('!ENDTRNS', text)
        self.assertIn('GENERAL JOURNAL', text)

    def test_account_codes_match_grand_livre(self):
        # Mêmes codes CGNC configurables que FG49 (3421/7111/4455 par défaut).
        from apps.ventes.exports import account_codes_for
        codes = account_codes_for(self.company)
        start, end = self._params()
        resp = self.api.get(
            f'/api/django/ventes/export-comptable/?start={start}&end={end}'
            f'&layout=quickbooks')
        text = resp.content.decode('utf-8')
        self.assertIn(codes['clients'], text)
        self.assertIn(codes['ventes'], text)
        self.assertIn(codes['tva_collectee'], text)

    def test_transaction_balanced(self):
        from datetime import date as _date
        from apps.ventes.exports import (
            _grand_livre_rows_as_entries, period_bounds)
        debut, fin = period_bounds(
            {'month': _date.today().strftime('%Y-%m')})
        entries, _totals = _grand_livre_rows_as_entries(
            self.company, debut, fin)
        tot_debit = sum(float(e.get('debit') or 0) for e in entries)
        tot_credit = sum(float(e.get('credit') or 0) for e in entries)
        self.assertEqual(round(tot_debit, 2), round(tot_credit, 2))
        self.assertGreater(tot_debit, 0)

    def test_avoir_reverses_signs_stays_balanced(self):
        from apps.ventes.models import Avoir, LigneAvoir
        avoir = Avoir.objects.create(
            company=self.company, reference='AVO-QB-1', facture=self.fac,
            client=self.fac.client, statut='emise', taux_tva=Decimal('20'))
        LigneAvoir.objects.create(
            avoir=avoir, designation='Retour', quantite=Decimal('1'),
            prix_unitaire=Decimal('2000'), remise=Decimal('0'),
            taux_tva=Decimal('20'))
        start, end = self._params()
        resp = self.api.get(
            f'/api/django/ventes/export-comptable/?start={start}&end={end}'
            f'&layout=quickbooks')
        self.assertEqual(resp.status_code, 200)
        text = resp.content.decode('utf-8')
        self.assertIn('AVO-QB-1', text)
