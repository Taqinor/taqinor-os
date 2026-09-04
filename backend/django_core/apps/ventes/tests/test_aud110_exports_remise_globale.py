"""AUD110 — les QUATRE exports comptables honorent enfin ``remise_globale``.

Les trois boucles d'``exports.py`` (journal, ``_compta_rows`` qui alimente
l'export comptable xlsx/csv, et ``_grand_livre_rows`` réutilisé par le pont
QuickBooks) sommaient directement ``ligne.total_ht``. Or
``LigneFacture.total_ht`` n'applique QUE la remise de LIGNE, jamais la remise
GLOBALE du document, alors que ``Facture.total_ht``/``total_tva``/``total_ttc``
l'appliquent via la chaîne canonique QX1 — la source de vérité utilisée partout
ailleurs (PDF facture legacy, UBL DGI, repli note de débit,
``compta.services.ecriture_pour_facture``). Les quatre exports ne l'appelaient
JAMAIS.

Scénario chiffré : facture remisée à 15 % sur 20 000 HT, TVA 20 % → chacun des
quatre documents doit porter 17 000 HT / 3 400 TVA / 20 400 TTC. Rouge avant
AUD110 sur les quatre (20 000 / 4 000 / 24 000).
"""
import io
from datetime import date
from decimal import Decimal

from django.test import TestCase

from apps.crm.models import Client
from apps.stock.models import Produit
from apps.ventes.exports import (
    _compta_rows, _grand_livre_rows, export_journal_ventes,
    export_quickbooks_iif,
)
from apps.ventes.models import Facture, LigneFacture
from authentication.models import Company

_CTR = [0]


def _nxt():
    _CTR[0] += 1
    return _CTR[0]


class TestExportsRemiseGlobale(TestCase):
    DEBUT = date(2026, 3, 1)
    FIN = date(2026, 4, 1)

    HT_NET = Decimal('17000.00')
    TVA = Decimal('3400.00')
    TTC = Decimal('20400.00')

    def setUp(self):
        self.company = Company.objects.create(
            nom='AUD110 Co', slug=f'aud110-{_nxt()}')
        self.client_obj = Client.objects.create(
            company=self.company, nom='AUD110', prenom='Client',
            telephone='+212600000111')
        self.produit = Produit.objects.create(
            company=self.company, nom='Kit PV', sku=f'AUD110-{_nxt()}',
            prix_vente=Decimal('20000'), quantite_stock=10)
        self.facture = Facture.objects.create(
            company=self.company, reference=f'FAC-AUD110-{_nxt()}',
            client=self.client_obj, statut=Facture.Statut.EMISE,
            taux_tva=Decimal('20.00'), remise_globale=Decimal('15.00'))
        LigneFacture.objects.create(
            facture=self.facture, produit=self.produit, designation='Kit PV',
            quantite=Decimal('1'), prix_unitaire=Decimal('20000'),
            taux_tva=Decimal('20.00'))
        Facture.objects.filter(pk=self.facture.pk).update(
            date_emission=date(2026, 3, 12))
        self.facture.refresh_from_db()

    # ── 1/4 — Journal des ventes (.xlsx) ──────────────────────────────────

    def test_journal_des_ventes(self):
        from openpyxl import load_workbook

        resp = export_journal_ventes(self.company, self.DEBUT, self.FIN)
        wb = load_workbook(io.BytesIO(resp.content))
        lignes = [tuple(c.value for c in row)
                  for row in wb['Journal des ventes'].iter_rows(min_row=2)]
        self.assertEqual(len(lignes), 1, lignes)
        self.assertEqual(lignes[0][8], float(self.HT_NET))
        self.assertEqual(lignes[0][10], float(self.TVA))
        self.assertEqual(lignes[0][11], float(self.TTC))
        total = [tuple(c.value for c in row)
                 for row in wb['Résumé TVA'].iter_rows(min_row=2)
                 if row[0].value == 'TOTAL'][0]
        self.assertEqual(total[1:4],
                         (float(self.HT_NET), float(self.TVA),
                          float(self.TTC)))

    # ── 2/4 — Export comptable (xlsx/csv, mêmes lignes) ───────────────────

    def test_export_comptable(self):
        rows, (tot_ht, tot_tva, tot_ttc) = _compta_rows(
            self.company, self.DEBUT, self.FIN)
        self.assertEqual(len(rows), 1, rows)
        self.assertEqual(rows[0][8], float(self.HT_NET))
        self.assertEqual(rows[0][10], float(self.TVA))
        self.assertEqual(rows[0][11], float(self.TTC))
        self.assertEqual((tot_ht, tot_tva, tot_ttc),
                         (self.HT_NET, self.TVA, self.TTC))

    # ── 3/4 — Grand-livre ventes ──────────────────────────────────────────

    def test_grand_livre(self):
        rows, (tot_debit, tot_credit) = _grand_livre_rows(
            self.company, self.DEBUT, self.FIN)
        par_compte = {}
        for compte, _lib, _d, _j, _p, _l, _t, _i, _taux, debit, credit in rows:
            entree = par_compte.setdefault(compte, [0.0, 0.0])
            entree[0] += debit
            entree[1] += credit
        codes = {'clients': '3421', 'ventes': '7111', 'tva': '4455'}
        self.assertEqual(par_compte[codes['clients']][0], float(self.TTC))
        self.assertEqual(par_compte[codes['ventes']][1], float(self.HT_NET))
        self.assertEqual(par_compte[codes['tva']][1], float(self.TVA))
        # Une écriture équilibrée reste équilibrée.
        self.assertEqual(tot_debit, tot_credit)
        self.assertEqual(tot_debit, self.TTC)

    # ── 4/4 — Pont QuickBooks (.iif), dérivé du grand-livre ───────────────

    def test_quickbooks_iif(self):
        contenu = export_quickbooks_iif(
            self.company, self.DEBUT, self.FIN).content.decode('utf-8')
        self.assertIn('20400.00', contenu)
        self.assertIn('17000.00', contenu)
        self.assertIn('3400.00', contenu)
        self.assertNotIn('24000.00', contenu)
        self.assertNotIn('20000.00', contenu)

    # ── Réconciliation avec le document lui-même ──────────────────────────

    def test_les_totaux_du_document_reconcilient_avec_la_facture(self):
        _rows, (tot_ht, tot_tva, tot_ttc) = _compta_rows(
            self.company, self.DEBUT, self.FIN)
        self.assertEqual(tot_ht, self.facture.total_ht)
        self.assertEqual(tot_tva, self.facture.total_tva)
        self.assertEqual(tot_ttc, self.facture.total_ttc)

    def test_taux_mixtes_repartis_au_centime(self):
        """Deux paniers 10/20 : la somme des lignes égale le document."""
        panneaux = Produit.objects.create(
            company=self.company, nom='Panneaux', sku=f'AUD110P-{_nxt()}',
            prix_vente=Decimal('1000'), quantite_stock=100)
        LigneFacture.objects.create(
            facture=self.facture, produit=panneaux, designation='Panneaux',
            quantite=Decimal('10'), prix_unitaire=Decimal('1000'),
            taux_tva=Decimal('10.00'))
        self.facture.refresh_from_db()
        rows, (tot_ht, tot_tva, tot_ttc) = _compta_rows(
            self.company, self.DEBUT, self.FIN)
        self.assertEqual(len(rows), 2)
        self.assertEqual(tot_ht, self.facture.total_ht)
        self.assertEqual(tot_tva, self.facture.total_tva)
        self.assertEqual(tot_ttc, self.facture.total_ttc)
        # 30 000 brut − 15 % = 25 500 HT net ; TVA 3 400 + 850 = 4 250.
        self.assertEqual(tot_ht, Decimal('25500.00'))
        self.assertEqual(tot_tva, Decimal('4250.00'))

    def test_sans_remise_les_montants_sont_inchanges(self):
        facture = Facture.objects.create(
            company=self.company, reference=f'FAC-AUD110-N{_nxt()}',
            client=self.client_obj, statut=Facture.Statut.EMISE,
            taux_tva=Decimal('20.00'))
        LigneFacture.objects.create(
            facture=facture, produit=self.produit, designation='Kit PV',
            quantite=Decimal('1'), prix_unitaire=Decimal('5000'),
            taux_tva=Decimal('20.00'))
        Facture.objects.filter(pk=facture.pk).update(
            date_emission=date(2026, 3, 20))
        rows, _totaux = _compta_rows(self.company, self.DEBUT, self.FIN)
        ligne_nue = [r for r in rows if r[0] == facture.reference][0]
        self.assertEqual(ligne_nue[8], 5000.0)
        self.assertEqual(ligne_nue[10], 1000.0)
        self.assertEqual(ligne_nue[11], 6000.0)
