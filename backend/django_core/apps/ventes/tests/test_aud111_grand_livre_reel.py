"""AUD111 — le grand-livre ventes cesse de contourner ``apps.compta``.

``ventes/exports.py`` n'importait JAMAIS ``apps.compta`` : ``_grand_livre_rows``
et ``export_quickbooks_iif`` RECALCULAIENT depuis Facture/Avoir un jeu
d'écritures Débit/Crédit sur des comptes CGNC, sans lire ni écrire une seule
``LigneEcriture``. Or le seul écrivain de GL réel est
``compta.services.ecriture_pour_facture`` et le FEC officiel ne lit QUE
``LigneEcriture`` : le cabinet recevait deux documents qui ne se recoupaient sur
aucune ligne. Même toggle actif, les deux calculs divergeaient
structurellement — le reconstitué ignorait la remise globale (AUD110) et
créditait 7111/4455 pour TOUTE facture, faisant apparaître un ACOMPTE en
produit.

Deux régimes désormais :
  * ``COMPTA_AUTO_ECRITURES`` actif (le cas nominal visé) → le document est un
    extrait du grand livre RÉEL (journal VTE), donc aligné ligne pour ligne sur
    ce que lit le FEC ;
  * inactif → reconstitution assumée, DITE en clair sur le document
    (``BANDEAU_RECONSTITUTION``) et alignée a minima sur ``facture.total_ht``
    et le traitement CGNC de l'acompte (4421, jamais 71xx/4455).
"""
import io
from datetime import date
from decimal import Decimal

from django.test import TestCase, override_settings

from apps.crm.models import Client
from apps.stock.models import Produit
from apps.ventes.exports import (
    BANDEAU_RECONSTITUTION, _grand_livre_rows, export_grand_livre_csv,
    export_grand_livre_xlsx, grand_livre_est_reel, period_bounds,
)
from apps.ventes.models import Facture, LigneFacture
from authentication.models import Company

_CTR = [0]


def _nxt():
    _CTR[0] += 1
    return _CTR[0]


class _BaseGrandLivre(TestCase):
    def setUp(self):
        self.company = Company.objects.create(
            nom='AUD111 Co', slug=f'aud111-{_nxt()}')
        self.client_obj = Client.objects.create(
            company=self.company, nom='AUD111', prenom='Client',
            ice='001122334', telephone='+212600000112')
        self.produit = Produit.objects.create(
            company=self.company, nom='Kit PV', sku=f'AUD111-{_nxt()}',
            prix_vente=Decimal('20000'), quantite_stock=50)
        self.debut, self.fin = period_bounds(
            {'month': date.today().strftime('%Y-%m')})

    def _facture(self, *, remise=Decimal('0'), lignes=True,
                 type_facture=Facture.TypeFacture.COMPLETE, montants=None):
        facture = Facture.objects.create(
            company=self.company, reference=f'FAC-AUD111-{_nxt()}',
            client=self.client_obj, statut=Facture.Statut.BROUILLON,
            taux_tva=Decimal('20.00'), remise_globale=remise,
            type_facture=type_facture,
            libelle='Tranche' if montants else '',
            montant_ht=montants[0] if montants else None,
            montant_tva=montants[1] if montants else None,
            montant_ttc=montants[2] if montants else None)
        if lignes:
            LigneFacture.objects.create(
                facture=facture, produit=self.produit, designation='Kit PV',
                quantite=Decimal('1'), prix_unitaire=Decimal('20000'),
                taux_tva=Decimal('20.00'))
        return facture

    def _emettre(self, facture):
        from apps.ventes.domain.facturation_ops import emettre_facture
        emettre_facture(facture, source='test_aud111')
        return facture


@override_settings(COMPTA_AUTO_ECRITURES=True)
class TestGrandLivreLitLeGrandLivreReel(_BaseGrandLivre):
    """Toggle ACTIF — le document est un extrait du journal VTE réel."""

    def _lignes_du_journal_reel(self):
        from datetime import timedelta

        from apps.compta.selectors import journal_items
        return journal_items(
            self.company, journal='VTE', date_debut=self.debut,
            date_fin=self.fin - timedelta(days=1))

    def _comparer_au_journal(self):
        """Le document et le journal VTE portent EXACTEMENT les mêmes lignes."""
        rows, (tot_debit, tot_credit) = _grand_livre_rows(
            self.company, self.debut, self.fin)
        attendu = [
            (item['compte_numero'], float(item['debit'] or 0),
             float(item['credit'] or 0))
            for item in self._lignes_du_journal_reel()
        ]
        obtenu = [(r[0], r[9], r[10]) for r in rows]
        self.assertEqual(obtenu, attendu)
        self.assertEqual(tot_debit, tot_credit)
        self.assertTrue(attendu, 'aucune écriture postée — test sans valeur')
        return rows

    def test_source_reelle_active(self):
        self.assertTrue(grand_livre_est_reel(self.company))

    def test_facture_a_lignes(self):
        self._emettre(self._facture())
        self._comparer_au_journal()

    def test_facture_header_only(self):
        self._emettre(self._facture(
            lignes=False,
            montants=(Decimal('10000'), Decimal('2000'), Decimal('12000'))))
        self._comparer_au_journal()

    def test_facture_remisee(self):
        self._emettre(self._facture(remise=Decimal('15.00')))
        rows = self._comparer_au_journal()
        # Le HT NET (17 000) est au grand livre, jamais le brut (20 000).
        credits = [r[10] for r in rows]
        self.assertIn(17000.0, credits)
        self.assertNotIn(20000.0, credits)

    def test_facture_dacompte(self):
        self._emettre(self._facture(
            type_facture=Facture.TypeFacture.ACOMPTE))
        rows = self._comparer_au_journal()
        comptes = {str(r[0]) for r in rows}
        # CGNC : l'acompte crédite 4421, JAMAIS un compte de produit 71xx.
        self.assertIn('4421', comptes)
        self.assertFalse({c for c in comptes if c.startswith('71')}, comptes)

    def test_aucun_bandeau_quand_la_source_est_reelle(self):
        self._emettre(self._facture())
        resp = export_grand_livre_csv(self.company, self.debut, self.fin)
        self.assertNotIn(
            BANDEAU_RECONSTITUTION, resp.content.decode('utf-8'))


class TestReconstitutionAssumeeEtDite(_BaseGrandLivre):
    """Toggle INACTIF (le défaut) — approximation assumée et DITE."""

    def test_source_reelle_inactive_par_defaut(self):
        self.assertFalse(grand_livre_est_reel(self.company))

    def test_bandeau_present_sur_le_csv(self):
        self._emettre(self._facture())
        resp = export_grand_livre_csv(self.company, self.debut, self.fin)
        self.assertIn(BANDEAU_RECONSTITUTION, resp.content.decode('utf-8'))

    def test_bandeau_present_sur_le_xlsx(self):
        from openpyxl import load_workbook

        self._emettre(self._facture())
        resp = export_grand_livre_xlsx(self.company, self.debut, self.fin)
        ws = load_workbook(io.BytesIO(resp.content)).active
        textes = [c.value for row in ws.iter_rows() for c in row
                  if isinstance(c.value, str)]
        self.assertIn(BANDEAU_RECONSTITUTION, textes)

    def test_lentete_et_la_zone_de_donnees_restent_en_place(self):
        """Le bandeau est POSÉ APRÈS le TOTAL : rien ne bouge au-dessus."""
        from openpyxl import load_workbook

        self._emettre(self._facture())
        resp = export_grand_livre_xlsx(self.company, self.debut, self.fin)
        ws = load_workbook(io.BytesIO(resp.content)).active
        lignes = [tuple(c.value for c in row) for row in ws.iter_rows()]
        self.assertEqual(lignes[0][0], 'Compte')
        self.assertEqual(lignes[1][0], '3421')

    def test_acompte_credite_4421_jamais_un_compte_de_produit(self):
        self._emettre(self._facture(
            type_facture=Facture.TypeFacture.ACOMPTE))
        rows, (tot_debit, tot_credit) = _grand_livre_rows(
            self.company, self.debut, self.fin)
        comptes = {str(r[0]) for r in rows}
        self.assertIn('4421', comptes)
        self.assertNotIn('7111', comptes)
        self.assertNotIn('4455', comptes)
        # L'écriture reste équilibrée (TTC au débit clients, TTC en avances).
        self.assertEqual(tot_debit, tot_credit)
        self.assertEqual(tot_debit, Decimal('24000.00'))

    def test_facture_complete_reste_inchangee(self):
        self._emettre(self._facture())
        rows, (tot_debit, tot_credit) = _grand_livre_rows(
            self.company, self.debut, self.fin)
        comptes = {str(r[0]) for r in rows}
        self.assertEqual(comptes, {'3421', '7111', '4455'})
        self.assertEqual(tot_debit, tot_credit)
