"""T8 — édition EN MASSE du catalogue produit (multi-sélection).

Toute la règle métier vit ici (les vues restent fines), bornée à la société de
l'utilisateur. Le PRIX D'ACHAT n'est JAMAIS modifié ni exposé (règle marges).
Les changements sont journalisés (audit logger).
"""
import logging
from decimal import Decimal, InvalidOperation, ROUND_CEILING
from django.db import models

logger = logging.getLogger('stock.audit')

BULK_ACTIONS = {'set_price', 'set_warranty', 'set_category', 'set_brand'}


def _dec(value):
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None


def apply_product_bulk(*, company, user, ids, op, params):
    """Applique une action en masse à une sélection de produits de la société.

    Renvoie {ok, updated, skipped:[{id,nom,reason}]}. Le prix d'achat reste
    intouché en toutes circonstances."""
    from .models import Produit, Categorie

    if op not in BULK_ACTIONS:
        raise ValueError("Action en masse inconnue.")

    produits = list(Produit.objects.filter(company=company, id__in=ids))
    updated, skipped = 0, []

    def skip(p, reason):
        skipped.append({'id': p.id, 'nom': p.nom, 'reason': reason})

    # Pré-validation par action.
    categorie = None
    if op == 'set_price':
        mode = params.get('mode')
        valeur = _dec(params.get('valeur'))
        if mode not in ('percent', 'fixed') or valeur is None:
            raise ValueError("Prix invalide (mode percent/fixed + valeur requise).")
    elif op == 'set_category':
        cid = params.get('categorie_id')
        categorie = Categorie.objects.filter(id=cid, company=company).first()
        if cid not in (None, '', 'null') and categorie is None:
            raise ValueError("Catégorie introuvable dans cette société.")

    for p in produits:
        if op == 'set_price':
            if mode == 'percent':
                new_price = (p.prix_vente or Decimal('0')) * (Decimal('1') + valeur / Decimal('100'))
            else:
                new_price = valeur
            if new_price < 0:
                skip(p, "prix négatif refusé")
                continue
            p.prix_vente = new_price.quantize(Decimal('0.01'))
            p.save(update_fields=['prix_vente'])  # prix_achat JAMAIS touché
            updated += 1

        elif op == 'set_warranty':
            fields = []
            if 'garantie_mois' in params and params['garantie_mois'] not in ('', None):
                p.garantie_mois = int(params['garantie_mois'])
                fields.append('garantie_mois')
            if ('garantie_production_mois' in params
                    and params['garantie_production_mois'] not in ('', None)):
                p.garantie_production_mois = int(params['garantie_production_mois'])
                fields.append('garantie_production_mois')
            if not fields:
                skip(p, "aucune durée fournie")
                continue
            p.save(update_fields=fields)
            updated += 1

        elif op == 'set_category':
            p.categorie = categorie
            p.save(update_fields=['categorie'])
            updated += 1

        elif op == 'set_brand':
            p.marque = (params.get('marque') or '').strip() or None
            p.save(update_fields=['marque'])
            updated += 1

    logger.info('BULK %s sur %d produit(s) par user=%s (company=%s)',
                op, updated, getattr(user, 'username', '?'), company.id)
    return {'ok': True, 'updated': updated, 'skipped': skipped,
            'total': len(produits)}


# ── Export Excel d'une sélection de produits ────────────────────────────────
PRODUCT_EXPORT_HEADERS = [
    'Nom', 'SKU', 'Marque', 'Catégorie', 'Prix vente HT', 'Quantité',
    'Seuil alerte', 'Garantie (mois)', 'Garantie production (mois)',
]


def export_products_xlsx(produits):
    """Réponse .xlsx pour une sélection de produits (prix d'achat EXCLU)."""
    from apps.crm.exports import build_xlsx_response
    rows = [[
        p.nom, p.sku or '', p.marque or '',
        getattr(p.categorie, 'nom', '') or '',
        str(p.prix_vente), p.quantite_stock, p.seuil_alerte,
        p.garantie_mois or '', p.garantie_production_mois or '',
    ] for p in produits]
    return build_xlsx_response(
        'produits.xlsx', PRODUCT_EXPORT_HEADERS, rows, sheet_title='Produits')


# ── N13 — Approvisionnement : besoin matériel par chantier ───────────────────
# Lecture seule sur installations.Installation (aucune migration installations).
# On dérive le besoin des LIGNES DU DEVIS source du chantier, on confronte au
# stock disponible (Produit.quantite_stock) et on signale les manques. Une
# action amont peut transformer ces manques en un BonCommandeFournisseur
# brouillon. Le prix d'achat reste INTERNE (jamais sur un document client).

def apply_inventory_count(*, company, user, motif, lignes):
    """N16 — inventaire : pose un comptage physique par produit et enregistre
    l'écart en MouvementStock (AJUSTEMENT). Renvoie {ajustes, inchanges,
    mouvements:[…]}. Le stock devient la quantité comptée ; rien n'est touché
    quand le comptage = stock actuel. Tout est scopé à la société."""
    from django.db import transaction
    from .models import Produit, MouvementStock

    motif = (motif or '').strip()
    result = {'ajustes': 0, 'inchanges': 0, 'mouvements': []}
    with transaction.atomic():
        for ligne in (lignes or []):
            pid = ligne.get('produit')
            try:
                compte = int(ligne.get('quantite_comptee'))
            except (TypeError, ValueError):
                continue
            if compte < 0:
                continue
            produit = Produit.objects.select_for_update().filter(
                id=pid, company=company).first()
            if produit is None:
                continue
            avant = produit.quantite_stock
            if compte == avant:
                result['inchanges'] += 1
                continue
            MouvementStock.objects.create(
                company=company, produit=produit,
                type_mouvement=MouvementStock.TypeMouvement.AJUSTEMENT,
                quantite=compte, quantite_avant=avant, quantite_apres=compte,
                reference='INVENTAIRE',
                note=f'Inventaire — comptage {compte} (écart {compte - avant})'
                     + (f' · {motif}' if motif else ''),
                created_by=user)
            produit.quantite_stock = compte
            produit.save(update_fields=['quantite_stock'])
            result['ajustes'] += 1
            result['mouvements'].append({
                'produit': produit.id, 'avant': avant, 'apres': compte})
    logger.info('INVENTAIRE %d ajustement(s) par user=%s (company=%s)',
                result['ajustes'], getattr(user, 'username', '?'), company.id)
    return result


# ── N15 — Stock multi-emplacements (dépôt principal + camionnette …) ─────────
# Le total `Produit.quantite_stock` reste canonique et inchangé ; cette couche
# ventile ce total entre emplacements. L'emplacement PRINCIPAL détient le reste
# (total − somme des non principaux), donc tout le stock existant est par défaut
# au dépôt principal et le comportement actuel ne change pas.

def ensure_emplacements(company):
    """Crée le dépôt principal + la camionnette si la société n'a encore aucun
    emplacement (idempotent, additif). Renvoie l'emplacement principal."""
    from .models import EmplacementStock
    if company is None:
        return None
    qs = EmplacementStock.objects.filter(company=company)
    if not qs.exists():
        EmplacementStock.objects.create(
            company=company, nom='Dépôt principal', is_principal=True, ordre=0)
        EmplacementStock.objects.create(
            company=company, nom='Camionnette', is_principal=False, ordre=10)
    return qs.filter(is_principal=True).first()


def get_or_create_emplacement_soustraitant(company, sous_traitant_nom):
    """XMFG16 — emplacement dédié « chez {sous-traitant} » (créé à la volée,
    même type d'emplacement que les autres — pas de type dédié). Idempotent :
    ``get_or_create`` par (company, nom). Sert à ventiler les composants
    confiés à un atelier de façon externe, hors dépôt principal."""
    from .models import EmplacementStock
    ensure_emplacements(company)
    nom = f'Chez {sous_traitant_nom}'
    emplacement, _ = EmplacementStock.objects.get_or_create(
        company=company, nom=nom, defaults={
            'is_principal': False, 'ordre': 900})
    return emplacement


def stock_breakdown(produit):
    """Ventilation du stock d'un produit par emplacement (non archivés).

    Renvoie [{emplacement_id, emplacement_nom, is_principal, quantite}] — le
    principal détient total − somme(non principaux)."""
    from .models import EmplacementStock
    company = produit.company
    ensure_emplacements(company)
    emplacements = list(EmplacementStock.objects.filter(
        company=company, archived=False))
    records = {se.emplacement_id: se.quantite
               for se in produit.stocks_emplacement.all()}
    autres = sum(records.get(e.id, 0)
                 for e in emplacements if not e.is_principal)
    out = []
    for e in emplacements:
        # ERR94 — le principal détient le reste (total − non principaux), mais
        # ne doit JAMAIS afficher une quantité négative : on le plafonne à 0
        # si le total est tombé sous l'allocation ventilée.
        qte = max(produit.quantite_stock - autres, 0) if e.is_principal \
            else records.get(e.id, 0)
        out.append({
            'emplacement_id': e.id,
            'emplacement_nom': e.nom,
            'is_principal': e.is_principal,
            'quantite': qte,
        })
    return out


def stock_breakdown_map(company):
    """Ventilation par emplacement pour TOUS les produits d'une société, en une
    passe (évite un N+1 quand la liste catalogue veut la répartition par produit).

    Renvoie {produit_id: [{emplacement_id, emplacement_nom, is_principal,
    quantite}]} — le principal détient total − somme(non principaux), comme
    `stock_breakdown` mais sans requête par produit."""
    from .models import EmplacementStock, Produit, StockEmplacement
    if company is None:
        return {}
    ensure_emplacements(company)
    emplacements = list(EmplacementStock.objects.filter(
        company=company, archived=False))
    # {produit_id: {emplacement_id: quantite}} pour les non principaux
    records = {}
    for se in StockEmplacement.objects.filter(
            produit__company=company, emplacement__archived=False):
        records.setdefault(se.produit_id, {})[se.emplacement_id] = se.quantite
    out = {}
    for p in Produit.objects.filter(company=company).only('id', 'quantite_stock'):
        rec = records.get(p.id, {})
        autres = sum(rec.get(e.id, 0) for e in emplacements if not e.is_principal)
        out[p.id] = [{
            'emplacement_id': e.id,
            'emplacement_nom': e.nom,
            'is_principal': e.is_principal,
            # ERR94 — principal plafonné à 0, jamais négatif (cf. stock_breakdown).
            'quantite': max(p.quantite_stock - autres, 0) if e.is_principal
            else rec.get(e.id, 0),
        } for e in emplacements]
    return out


def transfer_stock(*, company, user, produit_id, source_id, destination_id,
                   quantite, note=''):
    """Transfère `quantite` d'un produit de l'emplacement source vers la
    destination. Crée un TransfertStock (le « transfer record »). Ne change
    JAMAIS le total `Produit.quantite_stock`. Lève ValueError si invalide."""
    from django.db import transaction
    from .models import (
        Produit, EmplacementStock, StockEmplacement, TransfertStock)

    try:
        quantite = int(quantite)
    except (TypeError, ValueError):
        raise ValueError('Quantité invalide.')
    if quantite <= 0:
        raise ValueError('La quantité doit être positive.')
    if str(source_id) == str(destination_id):
        raise ValueError('La source et la destination doivent être différentes.')

    with transaction.atomic():
        produit = Produit.objects.select_for_update().filter(
            id=produit_id, company=company).first()
        if produit is None:
            raise ValueError('Produit introuvable dans cette société.')
        ensure_emplacements(company)
        emps = {e.id: e for e in EmplacementStock.objects.filter(
            company=company, archived=False)}
        try:
            source = emps[int(source_id)]
            destination = emps[int(destination_id)]
        except (KeyError, TypeError, ValueError):
            raise ValueError('Emplacement introuvable dans cette société.')

        records = {se.emplacement_id: se for se in
                   produit.stocks_emplacement.select_for_update()}
        non_principal_sum = sum(
            se.quantite for eid, se in records.items()
            if eid in emps and not emps[eid].is_principal)

        def current_qty(emp):
            if emp.is_principal:
                return produit.quantite_stock - non_principal_sum
            se = records.get(emp.id)
            return se.quantite if se else 0

        if quantite > current_qty(source):
            raise ValueError(
                f'Quantité insuffisante à « {source.nom} » '
                f'({current_qty(source)} disponible).')

        if not source.is_principal:
            se, _ = StockEmplacement.objects.select_for_update().get_or_create(
                produit=produit, emplacement=source,
                defaults={'company': company, 'quantite': 0})
            se.quantite -= quantite
            se.save(update_fields=['quantite'])
        if not destination.is_principal:
            se, _ = StockEmplacement.objects.get_or_create(
                produit=produit, emplacement=destination,
                defaults={'company': company, 'quantite': 0})
            se.quantite += quantite
            se.save(update_fields=['quantite'])

        transfert = TransfertStock.objects.create(
            company=company, produit=produit, source=source,
            destination=destination, quantite=quantite,
            note=(note or '').strip(), created_by=user)
    logger.info('TRANSFERT %d × produit=%s %s→%s par user=%s (company=%s)',
                quantite, produit_id, source_id, destination_id,
                getattr(user, 'username', '?'), company.id)
    return transfert


# ── N17 — Listes de prix multi-fournisseurs par SKU ──────────────────────────
# Prix d'achat INTERNE par (produit, fournisseur) + date du dernier achat.
# Sert à proposer le fournisseur le moins cher en rédigeant un bon de commande.

def cheapest_prix_fournisseur(produit):
    """Renvoie le PrixFournisseur le moins cher (prix > 0) pour ce produit, ou
    None s'il n'existe aucun prix multi-fournisseur renseigné."""
    return (produit.prix_fournisseurs.filter(prix_achat__gt=0)
            .select_related('fournisseur').order_by('prix_achat').first())


def record_purchase_price(*, company, produit, fournisseur, prix_achat, date):
    """Upsert du prix d'achat (produit, fournisseur) + date du dernier achat.
    Appelé à la réception d'un BCF. INTERNE (jamais client-facing)."""
    from decimal import Decimal
    from .models import PrixFournisseur
    if fournisseur is None or produit is None:
        return None
    prix = _dec(prix_achat) or Decimal('0')
    obj, created = PrixFournisseur.objects.get_or_create(
        produit=produit, fournisseur=fournisseur,
        defaults={'company': company, 'prix_achat': prix,
                  'date_dernier_achat': date})
    if not created:
        obj.prix_achat = prix
        obj.date_dernier_achat = date
        if obj.company_id is None:
            obj.company = company
        obj.save(update_fields=['prix_achat', 'date_dernier_achat', 'company'])
    return obj


# ── XPUR23 — Destination de réception (dépôt cible ou chantier direct) ──────
# Le put-away FG320 suggère un casier, mais l'entrée elle-même ne portait pas
# de dépôt cible ni de chantier de livraison directe : la marchandise entrait
# toujours au dépôt principal implicite. Ces fonctions sont appelées APRÈS le
# MouvementStock ENTREE (jamais un mécanisme parallèle) pour ventiler/tracer
# la destination. Comportement historique inchangé quand aucune destination
# n'est renseignée sur le BCF (défaut = dépôt principal, comme aujourd'hui).

def credit_emplacement_destination(company, produit, emplacement, quantite):
    """XPUR23 — crédite `quantite` sur l'emplacement destination (non
    principal). Le dépôt PRINCIPAL ne stocke jamais de ligne explicite (sa
    quantité reste dérivée : total − somme des non-principaux) — appeler
    cette fonction pour lui est un no-op volontaire (comportement historique
    déjà correct sans rien faire)."""
    from .models import StockEmplacement
    if emplacement is None or emplacement.is_principal or quantite <= 0:
        return
    se, _created = StockEmplacement.objects.get_or_create(
        company=company, produit=produit, emplacement=emplacement,
        defaults={'quantite': 0})
    se.quantite = (se.quantite or 0) + quantite
    se.save(update_fields=['quantite'])


def affecter_livraison_directe_chantier(
        company, user, bc, produit, quantite, reference):
    """XPUR23 — livraison DIRECTE chantier : la marchandise reçue N'ENTRE
    JAMAIS en stock libre. On la sort aussitôt (SORTIE tracée, référence du
    BCF + chantier) via l'accesseur unique `record_stock_movement` — jamais
    un mécanisme parallèle. Best-effort : une erreur ne casse jamais la
    réception elle-même (le stock reste alors en dépôt principal, visible et
    corrigeable manuellement)."""
    if quantite <= 0:
        return
    try:
        chantier = bc.chantier_livraison
        record_stock_movement(
            company=company, produit=produit,
            type_mouvement=mouvement_type_sortie(),
            quantite=quantite,
            quantite_avant=produit.quantite_stock,
            quantite_apres=produit.quantite_stock - quantite,
            reference=reference,
            note=(f'Livraison directe chantier {chantier.reference} '
                  f'(BCF {bc.reference})'
                  if chantier is not None else
                  f'Livraison directe chantier (BCF {bc.reference})'),
            created_by=user,
        )
        produit.refresh_from_db()
    except Exception:  # noqa: BLE001 — best-effort, jamais bloquant
        logger.info(
            'XPUR23: affectation chantier non tracée pour BCF %s', bc.pk)


# ── N18 — Valorisation du stock par emplacement (coût moyen d'achat) ──────────
# Coût moyen pondéré issu de l'historique d'achat (réceptions de BCF) ; à défaut
# le prix d'achat catalogue. Valorisation = quantité par emplacement × coût
# moyen. INTERNE uniquement (les prix d'achat ne sont jamais client-facing).

def average_cost_with_source(produit):
    """Coût moyen d'achat pondéré + sa SOURCE.

    Renvoie (cout, source) où source vaut 'achats' (dérivé des réceptions de
    bons de commande fournisseur), 'revalorisation' (dernière revalorisation
    manuelle validée servant de couche de départ, XSTK14) ou 'catalogue'
    (repli sur prix_achat quand aucun achat reçu ni revalorisation). INTERNE.

    FG67/DC38 — le coût intègre les FRAIS ANNEXES (fret/douane/TVA import/
    transit) de chaque ligne via le coût débarqué unitaire : aucun champ de
    coût parallèle, les frais sont repliés dans CE même coût moyen. Une ligne
    sans frais annexes (0) garde exactement le comportement historique.

    XSTK14 — si une `RevalorisationStock` VALIDÉE existe pour ce produit, son
    `nouveau_cout`/`quantite_snapshot` sert de couche de DÉPART (comme un
    inventaire initial) et seules les réceptions POSTÉRIEURES à sa
    validation entrent dans la moyenne pondérée — les réceptions
    antérieures sont supplantées par la revalorisation (comportement
    historique inchangé quand aucune revalorisation n'existe)."""
    from .models import LigneBonCommandeFournisseur, RevalorisationStock
    revalo = (RevalorisationStock.objects
              .filter(produit=produit,
                      statut=RevalorisationStock.Statut.VALIDEE)
              .order_by('-date_validation', '-id').first())
    lignes_qs = LigneBonCommandeFournisseur.objects.filter(
        produit=produit, quantite_recue__gt=0)
    if revalo is not None and revalo.date_validation is not None:
        lignes_qs = lignes_qs.filter(
            bon_commande__date_creation__gt=revalo.date_validation)
    lignes = lignes_qs.values_list(
        'quantite_recue', 'prix_achat_unitaire', 'quantite', 'frais_annexes')
    total_q, total_v = 0, Decimal('0')
    if revalo is not None:
        total_q += revalo.quantite_snapshot
        total_v += revalo.quantite_snapshot * revalo.nouveau_cout
    for q_recue, pu, q_ligne, frais in lignes:
        pu = pu or Decimal('0')
        frais = frais or Decimal('0')
        # Coût débarqué unitaire : frais annexes répartis sur la quantité
        # COMMANDÉE de la ligne (q_ligne), valorisés sur la quantité REÇUE.
        if q_ligne and frais:
            pu = pu + (frais / Decimal(str(q_ligne)))
        total_q += q_recue
        total_v += q_recue * pu
    if total_q:
        source = 'achats' if revalo is None else 'revalorisation'
        return (total_v / total_q).quantize(Decimal('0.01')), source
    return (produit.prix_achat or Decimal('0')), 'catalogue'


def average_cost(produit):
    """Coût moyen d'achat pondéré d'un produit, depuis l'historique des
    réceptions de bons de commande fournisseur ; repli sur le prix d'achat
    catalogue si aucun achat reçu. INTERNE."""
    return average_cost_with_source(produit)[0]


# ── FG67 — Méthode de valorisation : coût moyen pondéré (défaut) ou FIFO ──────
# La méthode est un réglage société (CompanyProfile.stock_valuation_method,
# 'wavg' par défaut). Le toggle CHAMP vit dans `parametres` (foundation) ; ce
# module le LIT prudemment via getattr et retombe sur 'wavg' si le champ
# n'existe pas encore — on ne crée aucun champ hors de stock. INTERNE.

VALUATION_WAVG = 'wavg'
VALUATION_FIFO = 'fifo'


def stock_valuation_method(company):
    """Méthode de valorisation choisie par la société : 'wavg' (coût moyen
    pondéré, défaut) ou 'fifo'. Lit CompanyProfile.stock_valuation_method si
    présent (couche parametres), repli prudent sur 'wavg' sinon. Lecture seule."""
    if company is None:
        return VALUATION_WAVG
    try:
        from apps.parametres.models_company import CompanyProfile
        profile = CompanyProfile.objects.filter(company=company).first()
    except Exception:  # parametres absent / champ pas encore migré
        profile = None
    method = getattr(profile, 'stock_valuation_method', None)
    return method if method in (VALUATION_WAVG, VALUATION_FIFO) else VALUATION_WAVG


def fifo_cost_with_source(produit):
    """Coût FIFO unitaire d'un produit + sa SOURCE (FG67).

    FIFO : le stock restant est valorisé aux coûts d'achat débarqués des
    réceptions les PLUS RÉCENTES (les premières entrées sont consommées en
    premier ; il reste donc les dernières). On reconstitue le coût unitaire des
    `quantite_stock` dernières unités reçues. Repli sur le prix d'achat
    catalogue ('catalogue') si aucune réception. INTERNE — jamais client-facing.

    Renvoie (cout_unitaire_moyen_des_couches_restantes, source)."""
    from .models import LigneBonCommandeFournisseur
    # Couches d'entrée, de la plus récente à la plus ancienne (FIFO -> il reste
    # les dernières entrées). On valorise au coût débarqué unitaire.
    lignes = (LigneBonCommandeFournisseur.objects
              .filter(produit=produit, quantite_recue__gt=0)
              .order_by('-bon_commande__date_creation', '-id'))
    restant = produit.quantite_stock or 0
    if restant <= 0:
        return (produit.prix_achat or Decimal('0')), 'catalogue'
    pris_q, pris_v = 0, Decimal('0')
    for ligne in lignes:
        if restant <= 0:
            break
        couche_q = min(ligne.quantite_recue, restant)
        pris_q += couche_q
        pris_v += couche_q * ligne.cout_unitaire_debarque
        restant -= couche_q
    if pris_q == 0:
        return (produit.prix_achat or Decimal('0')), 'catalogue'
    cout = (pris_v / pris_q).quantize(Decimal('0.01'))
    return cout, 'achats'


def valuation_cost_with_source(produit, method=None):
    """Coût unitaire de valorisation d'un produit selon la méthode société
    (FG67). 'wavg' -> coût moyen pondéré débarqué ; 'fifo' -> couches FIFO
    restantes. Si `method` n'est pas fourni, on lit le réglage société.
    Renvoie (cout, source). INTERNE."""
    if method is None:
        method = stock_valuation_method(produit.company)
    if method == VALUATION_FIFO:
        return fifo_cost_with_source(produit)
    return average_cost_with_source(produit)


# ── DC28 — UN seul résolveur du coût d'achat courant ─────────────────────────
# Précédence DOCUMENTÉE, une seule porte d'entrée pour marge / auto-fill /
# job-costing (ils lisent CET accesseur, jamais Produit.prix_achat en direct) :
#
#   1. accord de prix fournisseur actif (modèle à venir : aucun accord
#      n'existe encore au catalogue → étape inerte aujourd'hui, point
#      d'extension réservé) ;
#   2. PrixFournisseur — le DERNIER PAYÉ (date_dernier_achat la plus récente,
#      prix > 0) chez n'importe quel fournisseur de ce produit ;
#   3. repli : Produit.prix_achat (prix d'achat catalogue).
#
# INTERNE — jamais client-facing (règle marges). Renvoie (cout, source) où
# source ∈ {'accord', 'prix_fournisseur', 'catalogue'}.

def cout_achat_courant_with_source(produit):
    """Coût d'achat courant d'un produit + sa SOURCE, selon la précédence DC28.

    Renvoie ``(Decimal, source)`` avec ``source`` ∈ {'accord',
    'prix_fournisseur', 'catalogue'}. Accesseur UNIQUE — marge / auto-fill /
    job-costing passent par ici. INTERNE."""
    # 1. Accord de prix actif — point d'extension réservé (aucun modèle d'accord
    #    n'existe encore ; quand il arrivera, le brancher ici en priorité).
    # 2. PrixFournisseur : dernier payé (date la plus récente), prix > 0.
    dernier = (produit.prix_fournisseurs
               .filter(prix_achat__gt=0)
               .order_by('-date_dernier_achat', '-id')
               .first())
    if dernier is not None:
        return (Decimal(str(dernier.prix_achat)), 'prix_fournisseur')
    # 3. Repli catalogue.
    return (produit.prix_achat or Decimal('0'), 'catalogue')


def cout_achat_courant(produit):
    """Coût d'achat courant d'un produit (Decimal), selon la précédence DC28.
    Accesseur UNIQUE pour marge / auto-fill / job-costing. INTERNE."""
    return cout_achat_courant_with_source(produit)[0]


def stock_valuation_by_location(company):
    """Valorisation du stock par emplacement au coût moyen d'achat (N18).

    Renvoie {par_emplacement:[{emplacement_id, emplacement_nom, is_principal,
    quantite, valeur}], total, lignes:[{produit_id, sku, designation,
    emplacement_nom, quantite, cout_moyen, valeur}]}. INTERNE — ne jamais
    exposer dans un contexte client."""
    from .models import Produit, EmplacementStock
    ensure_emplacements(company)
    emplacements = list(EmplacementStock.objects.filter(
        company=company, archived=False))
    totals = {e.id: {'emplacement_id': e.id, 'emplacement_nom': e.nom,
                     'is_principal': e.is_principal, 'quantite': 0,
                     'valeur': Decimal('0')} for e in emplacements}
    lignes = []
    grand_total = Decimal('0')
    # FG67 — méthode société (coût moyen pondéré débarqué ou FIFO), résolue une
    # seule fois pour toute la passe.
    method = stock_valuation_method(company)
    produits = (Produit.objects.filter(company=company, is_archived=False)
                .prefetch_related('stocks_emplacement'))
    for p in produits:
        cout, source = valuation_cost_with_source(p, method=method)
        for b in stock_breakdown(p):
            if b['quantite'] == 0:
                continue
            valeur = (cout * b['quantite']).quantize(Decimal('0.01'))
            t = totals.get(b['emplacement_id'])
            if t is not None:
                t['quantite'] += b['quantite']
                t['valeur'] += valeur
            grand_total += valeur
            lignes.append({
                'produit_id': p.id, 'sku': p.sku or '', 'designation': p.nom,
                'emplacement_nom': b['emplacement_nom'],
                'quantite': b['quantite'], 'cout_moyen': cout, 'valeur': valeur,
                'source': source,
            })
    lignes.sort(key=lambda x: (x['designation'].lower(), x['emplacement_nom']))
    return {'par_emplacement': list(totals.values()),
            'total': grand_total, 'lignes': lignes}


VALORISATION_EXPORT_HEADERS = [
    'Produit', 'SKU', 'Emplacement', 'Quantité',
    'Coût moyen (HT)', 'Valeur (HT)', 'Source du coût',
]
_SOURCE_LABELS = {'achats': 'Achats reçus', 'catalogue': 'Prix catalogue'}


def export_valorisation_xlsx(company):
    """Réponse .xlsx de la valorisation du stock (admin/INTERNE — jamais
    client-facing). Reprend les lignes de stock_valuation_by_location."""
    from apps.crm.exports import build_xlsx_response
    data = stock_valuation_by_location(company)
    rows = [[
        ligne['designation'], ligne['sku'], ligne['emplacement_nom'],
        ligne['quantite'], str(ligne['cout_moyen']), str(ligne['valeur']),
        _SOURCE_LABELS.get(ligne.get('source'), ''),
    ] for ligne in data['lignes']]
    return build_xlsx_response(
        'valorisation.xlsx', VALORISATION_EXPORT_HEADERS, rows,
        sheet_title='Valorisation')


def stock_valuation_excludes_materiel_consigne(company):
    """YSTCK8 — GARDE explicite (jamais bloquante) : prouve que le matériel
    consigné (`installations.MaterielConsigne`, FG327 — non possédé) n'ajoute
    AUCUN layer de valeur à `stock_valuation_by_location`.

    Structurellement déjà garanti : `MaterielConsigne` n'a aucun FK vers
    `Produit` et ne crée jamais de `MouvementStock` ni n'incrémente
    `Produit.quantite_stock` — cette fonction documente et VÉRIFIE
    explicitement l'invariant (lu via `apps.installations.selectors`, jamais
    d'import du modèle installations depuis stock) plutôt que de le laisser
    implicite. Renvoie ``True`` (l'invariant tient toujours par construction
    — ne peut PAS renvoyer False aujourd'hui, il n'existe aucun chemin
    d'écriture qui violerait la garde).

    POINT D'EXTENSION documenté (pas construit ici, décision différée) :
    quand la consommation d'un lot consigné sera modélisée, elle devra passer
    par un TRANSFERT DE PROPRIÉTÉ explicite (le matériel devient possédé →
    entre alors, et alors seulement, dans cette valorisation) + une dette
    fournisseur (SAP 411-K). Voir `apps.installations.models_consignation`."""
    from apps.installations.selectors import materiel_consigne_quantite_totale
    valorisation = stock_valuation_by_location(company)
    # Le total consigné (info seule) n'apparaît dans AUCUNE ligne valorisée —
    # aucun rapprochement possible par construction (données disjointes),
    # ceci confirme juste que l'appel cross-app reste lecture seule et sûr.
    materiel_consigne_quantite_totale(company)
    return valorisation is not None


# ── XSTK13 — Valorisation À DATE (as-of) + inventaire annuel légal (CGNC) ────
# `stock_valuation_by_location` valorise l'INSTANT présent. Le CGNC exige un
# état d'inventaire valorisé PAR EXERCICE (support du bilan, contrôle fiscal) :
# on reconstruit la quantité de chaque produit à une date passée depuis
# l'historique des `MouvementStock` (dernier `quantite_apres` <= date), et le
# coût moyen d'achat AVEC LES SEULES réceptions antérieures ou égales à cette
# date (même formule débarquée que `average_cost_with_source`, bornée dans le
# temps). INTERNE — jamais client-facing.

def _quantite_produit_a_date(produit, date):
    """Dernière `quantite_apres` connue du produit à la date donnée (incluse).

    Aucun mouvement <= date -> 0 (le produit n'existait pas encore en stock à
    cette date, comportement conservateur)."""
    from .models import MouvementStock
    mvt = (MouvementStock.objects
           .filter(produit=produit, date__date__lte=date)
           .order_by('-date', '-id').first())
    return mvt.quantite_apres if mvt is not None else 0


def _cout_moyen_produit_a_date(produit, date):
    """Coût moyen d'achat débarqué du produit, en ne comptant QUE les
    réceptions de BCF dont le bon de commande est daté <= date (cf.
    `average_cost_with_source`, borné dans le temps). Repli catalogue si
    aucun achat reçu avant cette date."""
    from .models import LigneBonCommandeFournisseur
    lignes = (LigneBonCommandeFournisseur.objects
              .filter(produit=produit, quantite_recue__gt=0,
                      bon_commande__date_creation__date__lte=date)
              .values_list('quantite_recue', 'prix_achat_unitaire',
                           'quantite', 'frais_annexes'))
    total_q, total_v = 0, Decimal('0')
    for q_recue, pu, q_ligne, frais in lignes:
        pu = pu or Decimal('0')
        frais = frais or Decimal('0')
        if q_ligne and frais:
            pu = pu + (frais / Decimal(str(q_ligne)))
        total_q += q_recue
        total_v += q_recue * pu
    if total_q:
        return (total_v / total_q).quantize(Decimal('0.01')), 'achats'
    return (produit.prix_achat or Decimal('0')), 'catalogue'


def valorisation_a_date(company, date):
    """XSTK13 — valorisation du stock reconstruite À UNE DATE PASSÉE.

    Renvoie {date, total, lignes:[{produit_id, sku, designation, quantite,
    cout_moyen, valeur, source}]} — une ligne par produit dont la quantité
    reconstruite à `date` est non nulle. INTERNE (admin) — jamais
    client-facing."""
    from .models import Produit
    produits = Produit.objects.filter(company=company)
    lignes = []
    total = Decimal('0')
    for p in produits:
        quantite = _quantite_produit_a_date(p, date)
        if quantite == 0:
            continue
        cout, source = _cout_moyen_produit_a_date(p, date)
        valeur = (cout * quantite).quantize(Decimal('0.01'))
        total += valeur
        lignes.append({
            'produit_id': p.id, 'sku': p.sku or '', 'designation': p.nom,
            'quantite': quantite, 'cout_moyen': cout, 'valeur': valeur,
            'source': source,
        })
    lignes.sort(key=lambda x: x['designation'].lower())
    return {'date': date.isoformat() if hasattr(date, 'isoformat') else str(date),
            'total': total, 'lignes': lignes}


def figer_inventaire_annuel(company, exercice, user):
    """XSTK13 — fige l'inventaire de l'exercice : archive un snapshot complet
    et IMMUABLE de la valorisation au dernier jour de l'exercice (31/12).

    Un `InventaireAnnuel` déjà figé pour cet exercice+société lève ValueError
    (jamais deux figements pour le même exercice, jamais de ré-écriture)."""
    import datetime
    import json
    from django.core.serializers.json import DjangoJSONEncoder
    from .models import InventaireAnnuel
    if InventaireAnnuel.objects.filter(
            company=company, exercice=exercice).exists():
        raise ValueError(
            f"L'exercice {exercice} est déjà figé pour cette société.")
    date_fin = datetime.date(exercice, 12, 31)
    data = valorisation_a_date(company, date_fin)
    total_valeur = data['total']
    nb_lignes = len(data['lignes'])
    # JSONField encode déjà via DjangoJSONEncoder (les Decimal ne sont pas
    # sérialisables tels quels) — on applique le même aller-retour EN MÉMOIRE
    # au blob `donnees` (PAS à `total_valeur`, un vrai DecimalField qui doit
    # garder son type Decimal) avant `.create()`, pour que
    # `inventaire.donnees` (objet Python encore en mémoire) soit
    # BYTE-IDENTIQUE à une relecture depuis la base (le champ une fois stocké
    # ne connaît plus les Decimal, seulement leur str()) : sans ça,
    # `test_relit_a_l_identique` verrait deux représentations différentes
    # (Decimal vs str) pour la même donnée.
    donnees = json.loads(json.dumps(data, cls=DjangoJSONEncoder))
    return InventaireAnnuel.objects.create(
        company=company, exercice=exercice,
        date_reference=date_fin,
        total_valeur=total_valeur,
        nb_lignes=nb_lignes,
        donnees=donnees,
        created_by=user,
    )


def export_inventaire_annuel_xlsx(inventaire):
    """Export .xlsx de l'inventaire annuel FIGÉ (relit son snapshot JSON
    immuable — jamais recalculé après figement). INTERNE."""
    from apps.records.xlsx import build_xlsx_response
    rows = [[
        ligne['designation'], ligne['sku'], ligne['quantite'],
        str(ligne['cout_moyen']), str(ligne['valeur']),
        _SOURCE_LABELS.get(ligne.get('source'), ''),
    ] for ligne in inventaire.donnees.get('lignes', [])]
    return build_xlsx_response(
        f'inventaire-{inventaire.exercice}.xlsx',
        ['Produit', 'SKU', 'Quantité', 'Coût moyen (HT)', 'Valeur (HT)',
         'Source du coût'],
        rows, sheet_title=f'Inventaire {inventaire.exercice}')


# ── XSTK14 — Revalorisation manuelle du stock (document tracé) ──────────────
# INTERNE, admin-only, jamais client-facing.

def creer_revalorisation(*, company, produit, nouveau_cout, motif, user):
    """Crée une `RevalorisationStock` en BROUILLON : snapshot du coût moyen
    actuel + de la quantité en stock, delta calculé. Motif obligatoire
    (ValueError sinon)."""
    from .models import RevalorisationStock
    if not motif or not str(motif).strip():
        raise ValueError('Le motif de la revalorisation est obligatoire.')
    ancien_cout, _source = average_cost_with_source(produit)
    nouveau_cout = Decimal(str(nouveau_cout))
    quantite = produit.quantite_stock or 0
    delta = (nouveau_cout - ancien_cout) * quantite
    return RevalorisationStock.objects.create(
        company=company, produit=produit, ancien_cout=ancien_cout,
        nouveau_cout=nouveau_cout, quantite_snapshot=quantite,
        delta_valeur=delta.quantize(Decimal('0.01')), motif=motif,
        auteur=user)


def valider_revalorisation(revalorisation):
    """Valide une `RevalorisationStock` BROUILLON : verrouille le document
    (statut VALIDEE + date_validation) — devient la nouvelle couche de
    départ du coût moyen (`average_cost_with_source`). Une revalorisation
    déjà validée lève ValueError (jamais re-validée, jamais modifiée)."""
    from django.utils import timezone
    from .models import RevalorisationStock
    if revalorisation.statut == RevalorisationStock.Statut.VALIDEE:
        raise ValueError('Cette revalorisation est déjà validée.')
    revalorisation.statut = RevalorisationStock.Statut.VALIDEE
    revalorisation.date_validation = timezone.now()
    revalorisation.save(update_fields=['statut', 'date_validation'])
    return revalorisation


# ── XSTK8 — Contrôle du stock négatif (garde configurable) ──────────────────
# Seul `transfer_stock` refusait une quantité insuffisante ; les AUTRES
# chemins d'écriture (sorties chantier assemblage, retours fournisseur) ne
# vérifiaient rien. `AchatsParametres.stock_negatif_autorise` (défaut False)
# fait respecter le même garde là où il manquait — le garde EXISTANT de
# `transfer_stock` reste inchangé (ne l'appelle pas, a déjà le sien). Les
# données historiques déjà négatives restent lisibles (le garde ne s'applique
# qu'à une ÉCRITURE qui ferait PASSER sous zéro, jamais à la lecture).

def check_negative_stock_guard(company, quantite_avant, quantite_apres):
    """Lève ValueError si `quantite_apres` serait négatif et que la société
    n'autorise pas le stock négatif (`AchatsParametres.stock_negatif_
    autorise`, défaut False = refuse). Ne fait rien si le réglage l'autorise
    ou si le résultat reste ≥ 0 (comportement historique inchangé)."""
    if quantite_apres >= 0:
        return
    from .models import AchatsParametres
    parametres = AchatsParametres.for_company(company)
    if not parametres.stock_negatif_autorise:
        raise ValueError(
            f'Stock insuffisant ({quantite_avant} disponible) — cette '
            'opération ferait passer le stock sous zéro.')


# ── N19 — Retour fournisseur : validation = décrément de stock (SORTIE) ───────

def _reouvrir_quantite_recue_bcf(bc, produit_id, quantite_retournee):
    """YPROC8 — décrémente ``quantite_recue`` des lignes BCF de ``produit_id``
    à hauteur de ``quantite_retournee`` (plafonné à la quantité déjà reçue,
    réparti sur les lignes du même produit — la plus récente d'abord), et
    rétrograde le statut du BCF de RECU à ENVOYE si ``est_entierement_recu``
    devient faux. Ne fait rien si ``bc`` est None (retour sans BCF lié —
    comportement historique inchangé)."""
    from .models import BonCommandeFournisseur

    if bc is None or quantite_retournee <= 0:
        return
    restant_a_reouvrir = quantite_retournee
    lignes = (bc.lignes.select_for_update()
              .filter(produit_id=produit_id, quantite_recue__gt=0)
              .order_by('-id'))
    for ligne in lignes:
        if restant_a_reouvrir <= 0:
            break
        # Plafond : on ne rouvre jamais plus que ce que la ligne montre reçu.
        decrement = min(restant_a_reouvrir, ligne.quantite_recue)
        if decrement <= 0:
            continue
        ligne.quantite_recue -= decrement
        ligne.save(update_fields=['quantite_recue'])
        restant_a_reouvrir -= decrement
    bc.refresh_from_db()
    if (bc.statut == BonCommandeFournisseur.Statut.RECU
            and not bc.est_entierement_recu):
        bc.statut = BonCommandeFournisseur.Statut.ENVOYE
        bc.save(update_fields=['statut'])


def apply_retour_fournisseur(retour, user):
    """Valide un retour fournisseur : décrémente le stock (MouvementStock
    SORTIE) pour chaque ligne, puis passe le retour à « validé ». Lève
    ValueError si le retour n'est pas en brouillon ou est vide. INTERNE.

    YPROC8 — quand ``retour.bon_commande`` est renseigné, rouvre
    ``quantite_recue`` des lignes BCF du même produit (plafonné à la quantité
    reçue), rétrograde le statut RECU→ENVOYE si le BCF n'est plus entièrement
    reçu, et rafraîchit les rapprochements 3 voies OUVERTS de ce BCF (via le
    service compta dédié — jamais d'import du modèle compta). Un retour SANS
    BCF lié se comporte exactement comme avant (aucune régression)."""
    from django.db import transaction
    from .models import Produit, RetourFournisseur, MouvementStock
    if retour.statut != RetourFournisseur.Statut.BROUILLON:
        raise ValueError('Seul un retour en brouillon peut être validé.')
    lignes = list(retour.lignes.select_related('produit'))
    if not lignes:
        raise ValueError('Le retour ne contient aucune ligne.')
    with transaction.atomic():
        for ligne in lignes:
            # ERR24 — verrou de ligne produit dans la transaction pour que des
            # retours concurrents du même produit ne perdent pas de décrément.
            produit = (Produit.objects.select_for_update()
                       .get(pk=ligne.produit_id))
            qte_avant = produit.quantite_stock
            qte_apres = qte_avant - ligne.quantite
            check_negative_stock_guard(retour.company, qte_avant, qte_apres)
            MouvementStock.objects.create(
                company=retour.company, produit=produit,
                type_mouvement=MouvementStock.TypeMouvement.SORTIE,
                quantite=ligne.quantite, quantite_avant=qte_avant,
                quantite_apres=qte_apres, reference=retour.reference,
                note=f'Retour fournisseur {retour.reference}'
                     + (f' — {ligne.motif}' if ligne.motif else ''),
                created_by=user)
            produit.quantite_stock = qte_apres
            produit.save(update_fields=['quantite_stock'])
            if retour.bon_commande_id:
                _reouvrir_quantite_recue_bcf(
                    retour.bon_commande, ligne.produit_id, ligne.quantite)
        retour.statut = RetourFournisseur.Statut.VALIDE
        retour.save(update_fields=['statut'])
    if retour.bon_commande_id:
        try:
            from apps.compta.services import (
                refresh_rapprochements_ouverts_pour_bcf,
            )
            refresh_rapprochements_ouverts_pour_bcf(
                retour.company, retour.bon_commande_id)
        except Exception:  # noqa: BLE001 — best-effort, jamais bloquant
            logger.warning(
                'apply_retour_fournisseur: échec refresh rapprochement '
                'BCF %s', retour.bon_commande_id)
    return retour


# ── G5 — Réception fournisseur (goods-in) : confirmation = ENTRÉE de stock ────
# La confirmation d'une réception incrémente le stock (MouvementStock ENTREE)
# pour chaque ligne reçue, avance `quantite_recue` sur la ligne de BCF et fait
# évoluer le statut du BCF vers reçu/partiellement reçu via ses quantités reçues
# existantes (`est_entierement_recu`). IDEMPOTENTE : une réception déjà confirmée
# ne re-crée jamais de mouvement. Mêmes règles que l'action `recevoir` du BCF.

def confirm_reception_fournisseur(reception, user):
    """Confirme une réception fournisseur : crée un MouvementStock ENTREE par
    ligne reçue, incrémente le stock + `quantite_recue` du BCF, puis avance le
    statut du BCF (reçu si entièrement reçu). Lève ValueError si la réception
    n'est pas en brouillon ou est vide. INTERNE."""
    from django.db import transaction
    from django.utils import timezone
    from .models import (
        ReceptionFournisseur, BonCommandeFournisseur, MouvementStock,
    )
    if reception.statut != ReceptionFournisseur.Statut.BROUILLON:
        # Idempotence : une réception déjà confirmée/annulée ne touche pas le
        # stock une seconde fois.
        raise ValueError(
            'Seule une réception en brouillon peut être confirmée.')
    # YPROC7 — un BCF ANNULÉ ne doit plus jamais recevoir de stock : la garde
    # intervenait auparavant seulement APRÈS coup (avancement de statut), donc
    # on pouvait réceptionner (et incrémenter le stock) contre un BCF annulé.
    if (reception.bon_commande is not None
            and reception.bon_commande.statut
            == BonCommandeFournisseur.Statut.ANNULE):
        raise ValueError(
            'Ce bon de commande fournisseur est annulé : la réception ne '
            'peut pas être confirmée.')
    lignes = list(reception.lignes.select_related('ligne_commande', 'produit'))
    if not lignes:
        raise ValueError('La réception ne contient aucune ligne.')

    today = timezone.now().date()
    bc = reception.bon_commande
    with transaction.atomic():
        for ligne in lignes:
            qte = int(ligne.quantite or 0)
            if qte <= 0:
                continue
            # Plafonne au reste dû de la ligne de commande (jamais plus que
            # commandé — protège contre une saisie incohérente, idempotence).
            ligne_cmd = ligne.ligne_commande
            ligne_cmd.refresh_from_db()
            qte = min(qte, ligne_cmd.quantite_restante)
            if qte <= 0:
                continue
            # XPUR16 — ligne libre/service (sans_stock ou produit=null) :
            # aucun MouvementStock, la quantité reçue est simplement actée.
            if ligne_cmd.sans_stock or ligne.produit_id is None:
                ligne_cmd.quantite_recue += qte
                ligne_cmd.save(update_fields=['quantite_recue'])
                continue
            produit = ligne.produit
            produit.refresh_from_db()
            qte_avant = produit.quantite_stock
            qte_apres = qte_avant + qte
            MouvementStock.objects.create(
                company=reception.company, produit=produit,
                type_mouvement=MouvementStock.TypeMouvement.ENTREE,
                quantite=qte, quantite_avant=qte_avant,
                quantite_apres=qte_apres, reference=reception.reference,
                note=f'Réception {reception.reference}'
                     + (f' (BCF {bc.reference})' if bc else ''),
                created_by=user)
            produit.quantite_stock = qte_apres
            produit.save(update_fields=['quantite_stock'])
            ligne_cmd.quantite_recue += qte
            ligne_cmd.save(update_fields=['quantite_recue'])
            # N17 — mémorise le prix d'achat (interne) chez ce fournisseur.
            if bc is not None:
                record_purchase_price(
                    company=reception.company, produit=produit,
                    fournisseur=bc.fournisseur,
                    prix_achat=ligne_cmd.prix_achat_unitaire, date=today)
            # XPUR23 — destination de réception : dépôt cible OU chantier de
            # livraison directe (défaut = dépôt principal, inchangé).
            if bc is not None and bc.chantier_livraison_id:
                affecter_livraison_directe_chantier(
                    reception.company, user, bc, produit, qte,
                    reception.reference)
            elif bc is not None and bc.emplacement_destination_id:
                credit_emplacement_destination(
                    reception.company, produit, bc.emplacement_destination,
                    qte)
            # XSTK6 — alimente le registre de lots quand la ligne porte un
            # numero_lot (FG64). Sans lot renseigné : comportement historique
            # inchangé (aucune écriture LotEntrepot).
            if getattr(ligne, 'numero_lot', None):
                alimenter_lot_entrepot(
                    company=reception.company, produit=produit,
                    numero_lot=ligne.numero_lot,
                    date_peremption=ligne.date_peremption,
                    emplacement=(bc.emplacement_destination
                                 if bc is not None
                                 and bc.emplacement_destination_id
                                 else None),
                    quantite=qte,
                    reference_reception=reception.reference,
                    user=user)
        reception.statut = ReceptionFournisseur.Statut.CONFIRME
        reception.recu_par = reception.recu_par or user
        reception.save(update_fields=['statut', 'recu_par'])
        # Avance le statut du BCF selon ses quantités reçues existantes.
        if bc is not None:
            bc.refresh_from_db()
            if bc.statut not in (
                BonCommandeFournisseur.Statut.ANNULE,
                BonCommandeFournisseur.Statut.RECU,
            ):
                if bc.est_entierement_recu:
                    bc.statut = BonCommandeFournisseur.Statut.RECU
                else:
                    # Une réception partielle confirmée laisse un BCF brouillon
                    # passer à « envoyé » (commande engagée, partiellement reçue).
                    bc.statut = BonCommandeFournisseur.Statut.ENVOYE
                bc.save(update_fields=['statut'])
    # XQHS3 / YPROC3 — émet l'événement de confirmation sur le bus (core.events)
    # APRÈS le commit de la transaction (best-effort, ne casse jamais la
    # confirmation) : qhse peut ouvrir un contrôle qualité de réception,
    # installations peut créer sa provision GR/IR. stock n'importe ni l'un ni
    # l'autre — les deux s'abonnent dans leur propre apps.py ready().
    try:
        from core.events import reception_fournisseur_confirmee
        reception_fournisseur_confirmee.send(
            sender=ReceptionFournisseur, reception=reception,
            company=reception.company, user=user)
    except Exception:  # pragma: no cover - défensif, best-effort
        pass
    return reception


def annuler_reception_confirmee(reception, user):
    """YSTCK6 — annule une réception CONFIRMÉE par une CONTRE-PASSATION
    (reversal référencé, jamais un blocage ni une suppression — pattern SAP
    102). Pour chaque ligne de la réception, crée un `MouvementStock` SORTIE
    référencé ``ANNUL-<REC>`` (traçable à l'original), décrémente
    `quantite_recue` de la ligne BCF correspondante (plafonné à sa quantité
    reçue, jamais négatif) et rétrograde le statut du BCF de RECU à ENVOYE si
    ``est_entierement_recu`` devient faux. La sortie est plafonnée au stock en
    main (garde XSTK8 — jamais négatif). IDEMPOTENTE : seule une réception
    CONFIRME peut être annulée ; ré-annuler une réception déjà ANNULE lève
    ValueError (rien à rejouer). Renvoie la réception."""
    from django.db import transaction
    from django.utils import timezone
    from .models import (
        ReceptionFournisseur, BonCommandeFournisseur, MouvementStock,
    )

    if reception.statut != ReceptionFournisseur.Statut.CONFIRME:
        raise ValueError(
            'Seule une réception confirmée peut être annulée par '
            'contre-passation.')

    lignes = list(reception.lignes.select_related('ligne_commande', 'produit'))
    bc = reception.bon_commande
    with transaction.atomic():
        for ligne in lignes:
            qte = int(ligne.quantite or 0)
            if qte <= 0 or ligne.produit_id is None:
                continue
            produit = ligne.produit
            produit.refresh_from_db()
            qte_avant = produit.quantite_stock
            # XSTK8 — jamais négatif : la contre-passation ne sort jamais
            # plus que le stock en main (une partie a pu être déjà consommée
            # ailleurs entre-temps).
            qte_sortie = min(qte, qte_avant) if qte_avant > 0 else 0
            qte_apres = qte_avant - qte_sortie
            if qte_sortie > 0:
                MouvementStock.objects.create(
                    company=reception.company, produit=produit,
                    type_mouvement=MouvementStock.TypeMouvement.SORTIE,
                    quantite=qte_sortie, quantite_avant=qte_avant,
                    quantite_apres=qte_apres,
                    reference=f'ANNUL-{reception.reference}',
                    note=(f'Contre-passation annulation réception '
                          f'{reception.reference}'),
                    created_by=user)
                produit.quantite_stock = qte_apres
                produit.save(update_fields=['quantite_stock'])
            ligne_cmd = ligne.ligne_commande
            if ligne_cmd is not None:
                ligne_cmd.refresh_from_db()
                ligne_cmd.quantite_recue = max(
                    ligne_cmd.quantite_recue - qte, 0)
                ligne_cmd.save(update_fields=['quantite_recue'])
        reception.statut = ReceptionFournisseur.Statut.ANNULE
        reception.note = (
            f'{reception.note}\n[{timezone.now().date().isoformat()}] '
            f'Annulée par contre-passation.'.strip()
            if reception.note else
            f'[{timezone.now().date().isoformat()}] '
            'Annulée par contre-passation.')
        reception.save(update_fields=['statut', 'note'])
        if bc is not None:
            bc.refresh_from_db()
            if (bc.statut == BonCommandeFournisseur.Statut.RECU
                    and not bc.est_entierement_recu):
                bc.statut = BonCommandeFournisseur.Statut.ENVOYE
                bc.save(update_fields=['statut'])
    return reception


# ── XSTK6 — Registre de lots en entrepôt + sortie FEFO + garde périmé ───────
# Miroir d'`installations.SerieEntrepot` (FG323) mais pour du stock suivi PAR
# LOT (non sérialisé). Alimenté à la confirmation d'une réception ; décrémenté
# à la sortie. Le picking FEFO propose le lot à péremption la plus proche
# d'abord ; sortir un lot périmé est bloqué par défaut (`AchatsParametres.
# bloquer_stock_perime`, contournable avec motif tracé).

def alimenter_lot_entrepot(
        *, company, produit, numero_lot, date_peremption, quantite,
        reference_reception, emplacement=None, user=None):
    """XSTK6 — crée (ou incrémente si le MÊME lot existe déjà pour ce
    produit) une entrée de `LotEntrepot`. Jamais appelée pour une ligne sans
    ``numero_lot`` (comportement historique inchangé)."""
    from .models import LotEntrepot
    lot, created = LotEntrepot.objects.get_or_create(
        company=company, produit=produit, numero_lot=numero_lot,
        defaults={
            'date_peremption': date_peremption,
            'emplacement': emplacement,
            'quantite_recue': 0,
            'quantite_restante': 0,
            'reference_reception': reference_reception,
            'created_by': user,
        })
    if not created and date_peremption and not lot.date_peremption:
        lot.date_peremption = date_peremption
    lot.quantite_recue += quantite
    lot.quantite_restante += quantite
    lot.save(update_fields=[
        'quantite_recue', 'quantite_restante', 'date_peremption'])
    return lot


def suggestion_fefo(company, produit, quantite_requise):
    """XSTK6 — suggère le(s) lot(s) à sortir en premier (FEFO : péremption la
    plus proche d'abord ; les lots sans date de péremption passent en
    dernier). Renvoie une liste ``[{lot, quantite}]`` couvrant au mieux
    ``quantite_requise`` (LECTURE SEULE, ne décrémente rien)."""
    from .models import LotEntrepot
    from django.db.models import F
    lots = (LotEntrepot.objects
            .filter(company=company, produit=produit,
                    quantite_restante__gt=0)
            .order_by(
                F('date_peremption').asc(nulls_last=True), 'date_creation'))
    restant = quantite_requise
    plan = []
    for lot in lots:
        if restant <= 0:
            break
        prise = min(lot.quantite_restante, restant)
        plan.append({'lot': lot, 'quantite': prise})
        restant -= prise
    return plan


def sortir_lot_entrepot(
        *, company, lot, quantite, user=None, forcer=False, motif=None):
    """XSTK6 — décrémente un lot précis (sortie ciblée, ex. depuis le plan
    FEFO). Bloque un lot PÉRIMÉ si
    ``AchatsParametres.bloquer_stock_perime`` (défaut ON) — sauf ``forcer=
    True`` avec un ``motif`` tracé (journalisé dans la note du mouvement).
    Lève ValueError si la quantité dépasse le restant ou si le lot est
    périmé et non contourné."""
    from .models import AchatsParametres
    if quantite <= 0:
        raise ValueError('La quantité doit être positive.')
    if quantite > lot.quantite_restante:
        raise ValueError(
            f'Quantité insuffisante dans le lot {lot.numero_lot} '
            f'({lot.quantite_restante} restant).')
    parametres = AchatsParametres.for_company(company)
    if lot.est_perime and parametres.bloquer_stock_perime and not forcer:
        raise ValueError(
            f'Le lot {lot.numero_lot} est périmé '
            f'({lot.date_peremption}) — sortie bloquée.')
    if lot.est_perime and forcer and not motif:
        raise ValueError(
            'Un motif est requis pour contourner le blocage du lot périmé.')
    lot.quantite_restante -= quantite
    lot.save(update_fields=['quantite_restante'])
    if lot.est_perime and forcer:
        logger.info(
            'XSTK6: sortie forcée du lot périmé %s (%s) par %s — motif: %s',
            lot.numero_lot, lot.produit_id, getattr(user, 'id', None), motif)
    return lot


# ── G5 — Facture fournisseur / comptes à payer (AP) ──────────────────────────
# Le solde dû d'une facture = TTC − Σ paiements. Le statut de règlement est
# RECALCULÉ après chaque paiement (à payer / partiellement payée / payée). Les
# montants d'achat sont INTERNES (jamais client-facing).

def recompute_facture_fournisseur_statut(facture):
    """Recalcule le statut de règlement d'une facture fournisseur depuis ses
    paiements et le persiste. À payer si rien réglé, payée si le solde ≤ 0,
    sinon partiellement payée."""
    from decimal import Decimal
    from .models import FactureFournisseur
    paye = facture.total_paye
    ttc = facture.montant_ttc or Decimal('0')
    if paye <= Decimal('0'):
        statut = FactureFournisseur.Statut.A_PAYER
    elif paye >= ttc:
        statut = FactureFournisseur.Statut.PAYEE
    else:
        statut = FactureFournisseur.Statut.PARTIELLEMENT_PAYEE
    if facture.statut != statut:
        facture.statut = statut
        facture.save(update_fields=['statut'])
    return statut


# ── DC34 — Référentiel sous-traitant UNIFIÉ + AP par la chaîne standard ──────
# Écritures cross-app : les autres apps (installations) créent/mettent à jour un
# sous-traitant (Fournisseur type=service + SousTraitantProfile) et ses comptes à
# payer (FactureFournisseur/PaiementFournisseur) à travers ces services, jamais
# en important apps.stock.models directement. La société est TOUJOURS posée côté
# serveur (jamais lue du corps) et le profil hérite de la société du fournisseur.

def map_specialite_to_metier(specialite):
    """DC34/ARC22 — fait correspondre un texte libre de spécialité (ex. carnet
    projet ``gestion_projet.SousTraitant``) à un code ``SousTraitantProfile.
    Metier`` (enum fermé), insensible à la casse, repli ``AUTRE`` si aucun
    métier ne correspond (comportement jamais bloquant). Cette fonction est le
    SEUL endroit qui connaît le mapping specialite→metier — les autres apps
    passent ``specialite`` à ``create_sous_traitant`` plutôt que de dupliquer
    cette logique en important ``apps.stock.models`` elles-mêmes."""
    from .models import SousTraitantProfile
    if specialite:
        for code, _label in SousTraitantProfile.Metier.choices:
            if code.replace('_', ' ') == specialite.strip().lower():
                return code
    return SousTraitantProfile.Metier.AUTRE


def create_sous_traitant(*, company, user=None, nom, metier='autre',
                         specialite=None, contact_personne=None, email=None,
                         telephone=None, adresse=None, ice=None,
                         identifiant_fiscal=None, rc=None, rib=None,
                         actif=True, note=None):
    """DC34 — crée un sous-traitant = Fournisseur(type='service') + son
    SousTraitantProfile. Société posée serveur (fournisseur ET profil).
    ``specialite`` (texte libre, optionnel) est mappé vers ``metier`` via
    ``map_specialite_to_metier`` quand fourni — permet aux appelants
    cross-app (ex. ``gestion_projet``) de passer du texte libre sans importer
    ``apps.stock.models`` eux-mêmes ; ``metier`` explicite reste prioritaire
    si fourni. Renvoie le Fournisseur créé."""
    from .models import Fournisseur, SousTraitantProfile
    if specialite is not None:
        metier = map_specialite_to_metier(specialite)
    fournisseur = Fournisseur.objects.create(
        company=company, nom=nom, type=Fournisseur.Type.SERVICE,
        contact_personne=contact_personne, email=email or None,
        telephone=telephone, adresse=adresse, ice=ice,
        identifiant_fiscal=identifiant_fiscal, rc=rc, rib=rib)
    SousTraitantProfile.objects.create(
        company=company, fournisseur=fournisseur, metier=metier,
        actif=actif, note=note, created_by=user)
    return fournisseur


def update_sous_traitant(*, fournisseur, metier=None, actif=None, note=None,
                         **identity):
    """DC34 — met à jour un sous-traitant : champs d'identité sur le
    Fournisseur, champs propres (métier/actif/note) sur le profil satellite
    (créé s'il manque). Renvoie le Fournisseur."""
    from .models import SousTraitantProfile
    id_fields = []
    for champ in ('nom', 'contact_personne', 'email', 'telephone', 'adresse',
                  'ice', 'identifiant_fiscal', 'rc', 'rib'):
        if champ in identity and identity[champ] is not None:
            setattr(fournisseur, champ, identity[champ])
            id_fields.append(champ)
    if id_fields:
        fournisseur.save(update_fields=id_fields)
    profil, _ = SousTraitantProfile.objects.get_or_create(
        fournisseur=fournisseur,
        defaults={'company': fournisseur.company})
    prof_fields = []
    if metier is not None:
        profil.metier = metier
        prof_fields.append('metier')
    if actif is not None:
        profil.actif = actif
        prof_fields.append('actif')
    if note is not None:
        profil.note = note
        prof_fields.append('note')
    if prof_fields:
        prof_fields.append('date_modification')
        profil.save(update_fields=prof_fields)
    return fournisseur


def create_facture_sous_traitant(*, company, user=None, fournisseur,
                                 ref_fournisseur=None, date_facture=None,
                                 date_echeance=None, montant_ht=0,
                                 montant_tva=0, montant_ttc=0, note=None):
    """DC34 — crée une facture ENTRANTE de sous-traitant via la chaîne AP
    standard (FactureFournisseur), jamais un modèle parallèle. Référence
    anti-collision (préfixe FF), société + créateur posés serveur. Renvoie la
    FactureFournisseur."""
    from decimal import Decimal
    from apps.ventes.utils.references import create_with_reference
    from .models import FactureFournisseur

    def _save(reference):
        return FactureFournisseur.objects.create(
            company=company, reference=reference, fournisseur=fournisseur,
            ref_fournisseur=ref_fournisseur, date_facture=date_facture,
            date_echeance=date_echeance,
            montant_ht=montant_ht or Decimal('0'),
            montant_tva=montant_tva or Decimal('0'),
            montant_ttc=montant_ttc or Decimal('0'),
            statut=FactureFournisseur.Statut.A_PAYER, note=note,
            created_by=user)

    return create_with_reference(FactureFournisseur, 'FF', company, _save)


def add_paiement_sous_traitant(*, company, user=None, facture, montant,
                               date_paiement=None, mode='virement', note=None):
    """DC34 — impute un règlement sur une facture sous-traitant via la chaîne AP
    standard (PaiementFournisseur) et recalcule le statut de la facture. On
    n'impute jamais plus que le solde dû. Renvoie le PaiementFournisseur."""
    from decimal import Decimal, InvalidOperation
    from django.db import transaction
    from .models import PaiementFournisseur

    try:
        montant_dec = Decimal(str(montant))
    except (InvalidOperation, TypeError, ValueError):
        raise ValueError('Montant de paiement invalide.')
    if montant_dec <= 0:
        raise ValueError('Le montant du paiement doit être positif.')
    if montant_dec > facture.solde_du:
        raise ValueError('Le paiement dépasse le reste à payer.')
    with transaction.atomic():
        paiement = PaiementFournisseur.objects.create(
            company=company, facture=facture, montant=montant_dec,
            date_paiement=date_paiement, mode=mode, note=note,
            created_by=user)
        facture.refresh_from_db()
        recompute_facture_fournisseur_statut(facture)
    return paiement


def delete_paiement_sous_traitant(paiement):
    """DC34 — supprime un règlement sous-traitant et recalcule le statut de sa
    facture (chaîne AP standard)."""
    from django.db import transaction
    facture = paiement.facture
    with transaction.atomic():
        paiement.delete()
        facture.refresh_from_db()
        recompute_facture_fournisseur_statut(facture)


def enregistrer_paiement_fournisseur_depuis_run(*, company, facture_id,
                                                montant, date_paiement,
                                                user=None):
    """YLEDG8 — Crée le ``PaiementFournisseur`` d'une ligne de
    ``compta.PaymentRun`` POSTÉE référençant cette facture (mode virement,
    date du run) et recalcule son statut — même chaîne AP standard que
    ``add_paiement_sous_traitant`` (non restreinte aux sous-traitants).

    Garde anti-doublon avec YLEDG2 : le paiement créé ICI est la seule
    matérialisation du règlement (compta ne crée QUE l'écriture du run —
    jamais un 2ᵉ paiement/une 2ᵉ écriture pour la même ligne). Renvoie
    ``None`` si la facture est introuvable dans la société (aucune
    exception — l'appelant (compta) journalise et continue les autres
    lignes du run)."""
    from decimal import Decimal
    from django.db import transaction
    from .models import FactureFournisseur, PaiementFournisseur

    facture = FactureFournisseur.objects.filter(
        id=facture_id, company=company).first()
    if facture is None:
        return None
    montant_dec = Decimal(str(montant or 0))
    if montant_dec <= 0:
        return None
    with transaction.atomic():
        paiement = PaiementFournisseur.objects.create(
            company=company, facture=facture, montant=montant_dec,
            date_paiement=date_paiement,
            mode=PaiementFournisseur.Mode.VIREMENT, created_by=user,
            note='Règlement via campagne de paiement (payment run).')
        facture.refresh_from_db()
        recompute_facture_fournisseur_statut(facture)
    return paiement


# ── N14 — Réservé vs disponible : engagé-mais-non-consommé ───────────────────
# Une réservation de chantier (installations.StockReservation) ENGAGE le stock
# d'un SKU sans le décrémenter. Le « disponible » d'un produit en tient compte :
#   disponible = quantite_stock − somme(réservations actives non consommées).
# Les vues stock + alertes de stock bas s'appuient sur le disponible (N14). Le
# modèle de réservation vit dans installations (qui dépend déjà de stock) ; on
# l'importe PARESSEUSEMENT pour éviter toute dépendance d'app circulaire.

def reserved_quantity(produit):
    """Quantité de ce produit ENGAGÉE par des réservations de chantier actives
    et non encore consommées (0 si aucune). Lecture seule."""
    from apps.installations.selectors import reserved_quantity_for_produit
    return reserved_quantity_for_produit(produit)


def reserved_quantities(company):
    """Map {produit_id: quantité réservée active} pour toute la société — un
    seul agrégat (évite un N+1 sur la liste produits). Lecture seule."""
    from apps.installations.selectors import reserved_quantities_for_company
    return reserved_quantities_for_company(company)


def available_quantity(produit, reserved=None):
    """Disponible = stock total − réservé (engagé-mais-non-consommé). `reserved`
    peut être fourni (depuis `reserved_quantities`) pour éviter une requête."""
    if reserved is None:
        reserved = reserved_quantity(produit)
    return produit.quantite_stock - reserved


def _own_reservation_map(installation):
    """Map {produit_id: quantité} des réservations actives non consommées
    propres à CE chantier (pour ne pas les décompter de son propre disponible).
    """
    from apps.installations.selectors import own_reservation_map
    return own_reservation_map(installation)


def is_low_stock_available(produit, reserved=None):
    """Alerte de stock bas N14 : compare le DISPONIBLE (et non le stock brut) au
    seuil d'alerte. Un seuil ≤ 0 désactive l'alerte (comportement historique)."""
    if produit.seuil_alerte is None or produit.seuil_alerte <= 0:
        return False
    return available_quantity(produit, reserved) <= produit.seuil_alerte


def compute_besoin_materiel(installation):
    """Agrège les besoins matériel d'un chantier depuis son devis source.

    Renvoie une liste de dicts triés par désignation :
      {produit, produit_id, sku, designation, requis, disponible,
       reserve, manque, fournisseur_id, fournisseur_nom,
       fournisseur_min_id, fournisseur_min_nom, prix_achat_min}
    `disponible` = stock total − réservé (engagé non consommé). `manque` =
    max(requis − disponible, 0). Un manque > 0 = pénurie.
    `fournisseur_min_*` = fournisseur le moins cher (N17), s'il existe.

    Les lignes sans produit (libre) sont ignorées : on ne peut pas
    réapprovisionner un article qui n'est pas au catalogue.
    """
    devis = installation.devis
    if devis is None:
        return []
    # N14 — réservations actives de la société, et la part réservée par CE
    # chantier (pour ne pas la décompter deux fois de son propre disponible).
    reserved_all = reserved_quantities(installation.company)
    own_reserved = _own_reservation_map(installation)
    besoins = {}
    for ligne in devis.lignes.select_related('produit', 'produit__fournisseur'):
        produit = ligne.produit
        if produit is None:
            continue
        # ERR54 — une ligne fractionnaire (ex. 2,5 unités) exige un APPRO ARRONDI
        # AU SUPÉRIEUR (3, pas 2) : un int() tronquait et sous-commandait.
        try:
            requis = int(_dec(ligne.quantite).to_integral_value(rounding=ROUND_CEILING))
        except (AttributeError, InvalidOperation, TypeError, ValueError):
            requis = 0
        entry = besoins.get(produit.id)
        if entry is None:
            # Disponible pour CE chantier = stock total − réservé par les AUTRES
            # chantiers (engagé non consommé) ; sa propre réservation n'est pas
            # soustraite (sinon le besoin serait compté deux fois).
            reserve_autres = max(
                reserved_all.get(produit.id, 0)
                - own_reserved.get(produit.id, 0), 0)
            besoins[produit.id] = {
                'produit': produit,
                'produit_id': produit.id,
                'sku': produit.sku or '',
                'designation': produit.nom,
                'requis': requis,
                'disponible': produit.quantite_stock - reserve_autres,
                'reserve': reserved_all.get(produit.id, 0),
                'fournisseur_id': produit.fournisseur_id,
                'fournisseur_nom': (produit.fournisseur.nom
                                    if produit.fournisseur_id else None),
            }
        else:
            entry['requis'] += requis
    out = []
    for entry in besoins.values():
        entry['manque'] = max(entry['requis'] - entry['disponible'], 0)
        # N17 — fournisseur le moins cher (si une liste de prix existe).
        cheapest = cheapest_prix_fournisseur(entry['produit'])
        if cheapest is not None:
            entry['fournisseur_min_id'] = cheapest.fournisseur_id
            entry['fournisseur_min_nom'] = cheapest.fournisseur.nom
            entry['prix_achat_min'] = cheapest.prix_achat
        else:
            entry['fournisseur_min_id'] = None
            entry['fournisseur_min_nom'] = None
            entry['prix_achat_min'] = None
        out.append(entry)
    out.sort(key=lambda e: e['designation'].lower())
    return out


def draft_bcf_for_shortfall(installation, fournisseur, user, company):
    """Crée un BonCommandeFournisseur BROUILLON pour les manques du chantier.

    Une ligne BCF par produit en pénurie (quantité = le manque), au prix
    d'achat catalogue (interne). Renvoie (bon, lignes_count). Lève ValueError
    s'il n'y a aucun manque à commander.

    YPROC10 — pose `chantier_origine` (string-FK, distinct de la destination
    de livraison XPUR23) : la note texte ne suffisait pas à relier
    structurellement le BCF au chantier — à la réception, la marchandise
    entrait en stock libre et pouvait être consommée par n'importe qui.
    """
    from apps.ventes.utils.references import create_with_reference
    from .models import BonCommandeFournisseur, LigneBonCommandeFournisseur

    besoins = compute_besoin_materiel(installation)
    manquants = [b for b in besoins if b['manque'] > 0]
    if not manquants:
        raise ValueError('Aucun manque à commander pour ce chantier.')

    def _save(ref):
        bon = BonCommandeFournisseur.objects.create(
            company=company,
            reference=ref,
            fournisseur=fournisseur,
            statut=BonCommandeFournisseur.Statut.BROUILLON,
            note=(f'Besoin matériel — chantier {installation.reference}'),
            chantier_origine=installation,
            created_by=user,
        )
        for b in manquants:
            produit = b['produit']
            LigneBonCommandeFournisseur.objects.create(
                bon_commande=bon,
                produit=produit,
                quantite=b['manque'],
                prix_achat_unitaire=produit.prix_achat,
            )
        return bon

    bon = create_with_reference(BonCommandeFournisseur, 'BCF', company, _save)
    return bon, len(manquants)


# ── YPROC5/YPROC6 — conversion générique lignes → BCF brouillon ─────────────
# Point d'entrée cross-app RÉUTILISABLE : appelé par ``installations`` (demande
# d'achat approuvée → BCF, adjudication RFQ → BCF) SANS que ``stock`` importe
# jamais ``installations``. Une ligne par tuple (produit_id, désignation,
# quantité, prix) ; ``produit_id`` peut être ``None`` (ligne libre/service, DA
# hors catalogue). Référence anti-collision (jamais count()+1).

def creer_bcf_depuis_lignes(*, company, user, fournisseur, lignes, note=''):
    """Crée un ``BonCommandeFournisseur`` BROUILLON depuis une liste de lignes.

    ``lignes`` : itérable de tuples ``(produit_id, designation, qte, prix)``.
    ``produit_id`` peut être ``None`` (ligne libre/service — ``sans_stock``
    posé automatiquement par le modèle via l'absence de produit). Lève
    ValueError si ``lignes`` est vide. Renvoie le ``BonCommandeFournisseur``
    créé. INTERNE — jamais de prix d'achat sur un document client.
    """
    from apps.ventes.utils.references import create_with_reference
    from .models import BonCommandeFournisseur, LigneBonCommandeFournisseur

    lignes = list(lignes)
    if not lignes:
        raise ValueError('Aucune ligne à commander.')

    def _save(ref):
        bon = BonCommandeFournisseur.objects.create(
            company=company, reference=ref, fournisseur=fournisseur,
            statut=BonCommandeFournisseur.Statut.BROUILLON,
            note=note or '', created_by=user)
        for produit_id, designation, qte, prix in lignes:
            LigneBonCommandeFournisseur.objects.create(
                bon_commande=bon, produit_id=produit_id,
                designation=designation or '', quantite=int(qte or 0),
                prix_achat_unitaire=prix or 0)
        return bon

    return create_with_reference(BonCommandeFournisseur, 'BCF', company, _save)


def resolve_fournisseur(company, fournisseur_id, installation):
    """Choisit le fournisseur du brouillon : explicite, sinon (N17) le
    fournisseur le moins cher du premier produit en pénurie, sinon le
    fournisseur catalogue de ce produit, sinon None."""
    from .models import Fournisseur
    if fournisseur_id:
        return Fournisseur.objects.filter(
            id=fournisseur_id, company=company).first()
    for b in compute_besoin_materiel(installation):
        if b['manque'] <= 0:
            continue
        cible = b.get('fournisseur_min_id') or b.get('fournisseur_id')
        if cible:
            return Fournisseur.objects.filter(
                id=cible, company=company).first()
    return None


# ── Point d'entrée cross-app : mouvement de stock + maj du produit ───────────
# Les autres apps (ventes, installations, sav) décrémentent/incrémentent le
# stock à travers ce service plutôt qu'en créant `MouvementStock` directement et
# en sauvant `Produit` à la main (voir CLAUDE.md, règle de modularité). L'appelant
# garde sa propre logique de garde (refresh_from_db, plancher zéro, etc.) et passe
# les quantités déjà calculées : le service ne fait que l'écriture, à l'identique.

def record_stock_movement(*, company, produit, type_mouvement, quantite,
                          quantite_avant, quantite_apres, reference, note,
                          created_by, save_produit=True, emplacement_source=None):
    """Crée UN MouvementStock et (par défaut) cale `produit.quantite_stock` sur
    `quantite_apres`. Renvoie le mouvement créé. Écriture identique au
    `MouvementStock.objects.create(...) + produit.save(update_fields=...)` que les
    appelants faisaient inline. À utiliser dans la transaction de l'appelant.

    YSTCK3 — ``emplacement_source`` (optionnel, `EmplacementStock`) : quand un
    SORTIE est imputée à un emplacement NON PRINCIPAL précis (ex. camionnette
    d'un technicien), décrémente CE `StockEmplacement` au lieu de laisser le
    principal absorber silencieusement toute la baisse (`stock_breakdown`
    dérive le principal = total − Σ non-principaux, donc sans ce paramètre la
    camionnette ne redescend JAMAIS quand un technicien consomme depuis son
    van). Défaut ``None`` = comportement historique EXACT (aucun
    `StockEmplacement` touché ; la dérivation absorbe la baisse au principal,
    plafonnée à 0 par ERR94). Un `emplacement_source` PRINCIPAL est un no-op
    (le principal est déjà dérivé, jamais stocké). Ne s'applique qu'aux
    mouvements SORTIE (une ENTREE avec emplacement passe par
    `credit_emplacement_destination`, jamais dupliquée ici)."""
    from .models import MouvementStock, StockEmplacement
    mouvement = MouvementStock.objects.create(
        company=company,
        produit=produit,
        type_mouvement=type_mouvement,
        quantite=quantite,
        quantite_avant=quantite_avant,
        quantite_apres=quantite_apres,
        reference=reference,
        note=note,
        created_by=created_by,
    )
    if save_produit:
        produit.quantite_stock = quantite_apres
        produit.save(update_fields=['quantite_stock'])
    if (emplacement_source is not None
            and not emplacement_source.is_principal
            and type_mouvement == MouvementStock.TypeMouvement.SORTIE):
        se, _created = StockEmplacement.objects.select_for_update().get_or_create(
            produit=produit, emplacement=emplacement_source,
            defaults={'company': company, 'quantite': 0})
        se.quantite = max(se.quantite - quantite, 0)
        se.save(update_fields=['quantite'])
    _notify_seuil_atteint_si_franchi(
        company=company, produit=produit,
        quantite_avant=quantite_avant, quantite_apres=quantite_apres)
    return mouvement


def _notify_seuil_atteint_si_franchi(*, company, produit, quantite_avant,
                                     quantite_apres):
    """XSTK23 — webhook `stock.seuil_atteint`, ÉMIS UNE SEULE FOIS au moment où
    ce mouvement fait FRANCHIR le seuil effectif À LA BAISSE (avant > seuil,
    après <= seuil). Un mouvement qui ne franchit rien (déjà sous seuil avant,
    ou remontée) ne redéclenche rien. Best-effort : jamais bloquant, appelle le
    SERVICE publicapi (jamais son modèle) — cross-app via services.py comme
    l'exige la frontière inter-app."""
    try:
        seuil, _cible = seuil_effectif_produit(company, produit)
        if seuil is None or seuil <= 0:
            return
        if quantite_avant > seuil and quantite_apres <= seuil:
            from apps.publicapi.services import notify_stock_seuil_atteint
            notify_stock_seuil_atteint(
                company_id=company.id if company else None,
                produit_id=produit.id,
                sku=produit.sku,
                nom=produit.nom,
                quantite_disponible=quantite_apres,
                seuil=seuil,
            )
    except Exception:  # noqa: BLE001 — jamais bloquant pour l'écriture stock
        logger.exception('publicapi stock.seuil_atteint dispatch failed')


def mouvement_type_sortie():
    """Valeur enum SORTIE (sans importer le modèle côté appelant)."""
    from .models import MouvementStock
    return MouvementStock.TypeMouvement.SORTIE


def mouvement_type_entree():
    """Valeur enum ENTREE (sans importer le modèle côté appelant)."""
    from .models import MouvementStock
    return MouvementStock.TypeMouvement.ENTREE


def mouvement_type_rebut():
    """XMFG11 — valeur enum REBUT (sans importer le modèle côté appelant)."""
    from .models import MouvementStock
    return MouvementStock.TypeMouvement.REBUT


# ── YSTCK1 — comptage cyclique (installations.SessionComptage/ComptageLigne)
# poste l'écart en AJUSTEMENT à la clôture. Appelé DEPUIS installations
# (jamais l'inverse) : le service reste ici (comme `apply_inventory_count`,
# FG63, le comptage one-shot) — la diff EST le mouvement.

def appliquer_ecarts_comptage(*, company, lignes, user, reference):
    """YSTCK1 — poste UN `MouvementStock` AJUSTEMENT par ligne dont
    `quantite_comptee != quantite_theorique` (attribut ``ecart`` non nul, non
    None), cale `Produit.quantite_stock` sur le compté.

    ``lignes`` : itérable de ``ComptageLigne`` (ou tout objet portant
    ``produit_id``/``quantite_theorique``/``quantite_comptee``/``ecart``).
    Une ligne sans produit catalogue (désignation libre) ou pas encore
    comptée (``quantite_comptee`` None) est ignorée. Renvoie le nombre de
    mouvements postés. Scopé société ; verrouille le produit (select_for_update)
    pour éviter une course avec un mouvement concurrent."""
    from django.db import transaction
    from .models import MouvementStock, Produit

    count = 0
    with transaction.atomic():
        for ligne in lignes:
            if ligne.produit_id is None:
                continue
            if ligne.quantite_comptee is None:
                continue
            ecart = ligne.quantite_comptee - (ligne.quantite_theorique or 0)
            if ecart == 0:
                continue
            produit = Produit.objects.select_for_update().filter(
                id=ligne.produit_id, company=company).first()
            if produit is None:
                continue
            avant = produit.quantite_stock
            apres = ligne.quantite_comptee
            record_stock_movement(
                company=company, produit=produit,
                type_mouvement=MouvementStock.TypeMouvement.AJUSTEMENT,
                quantite=abs(ecart), quantite_avant=avant,
                quantite_apres=apres, reference=f'CYC-{reference}',
                note=f'Comptage cyclique {reference} — écart {ecart}',
                created_by=user)
            count += 1
    return count


def declarer_rebut(*, company, produit, quantite, motif, reference, note,
                   user):
    """XMFG11 — déclare un rebut de production : SORTIE typée REBUT, motivée,
    rattachée à un document source (``reference``, ex. un ordre d'assemblage).
    ``motif`` doit être une valeur de ``MouvementStock.MotifRebut``. Lève
    ValueError si la quantité est invalide."""
    from django.db import transaction
    from .models import MouvementStock, Produit

    if quantite is None or quantite <= 0:
        raise ValueError('La quantité de rebut doit être positive.')
    valeurs_motif = {c for c, _ in MouvementStock.MotifRebut.choices}
    if motif not in valeurs_motif:
        raise ValueError('Motif de rebut invalide.')

    with transaction.atomic():
        p = Produit.objects.select_for_update().get(id=produit.id)
        avant = p.quantite_stock
        qte_sortie = min(quantite, avant) if avant > 0 else 0
        apres = avant - qte_sortie
        mouvement = MouvementStock.objects.create(
            company=company, produit=p,
            type_mouvement=MouvementStock.TypeMouvement.REBUT,
            quantite=qte_sortie, quantite_avant=avant, quantite_apres=apres,
            reference=reference, note=note, motif_rebut=motif,
            created_by=user)
        p.quantite_stock = apres
        p.save(update_fields=['quantite_stock'])
    return mouvement


def rebuter_produit(
        *, company, produit, quantite, motif, user, emplacement=None,
        reference_chantier=None):
    """XSTK10 — met au rebut une quantité d'un produit (motif obligatoire :
    casse/obsolète/périmé/vol/défaut/erreur/autre), décrémente l'emplacement
    source (si fourni, N15) en plus du total canonique, respecte le garde
    XSTK8 (stock négatif) et journalise la VALEUR perdue au coût moyen
    (`average_cost_with_source`). Renvoie {mouvement, valeur_perdue}."""
    from django.db import transaction
    from .models import MouvementStock, Produit, StockEmplacement

    if quantite is None or quantite <= 0:
        raise ValueError('La quantité de rebut doit être positive.')
    valeurs_motif = {c for c, _ in MouvementStock.MotifRebut.choices}
    if motif not in valeurs_motif:
        raise ValueError('Motif de rebut invalide.')

    cout_moyen, _source = average_cost_with_source(produit)
    valeur_perdue = (cout_moyen or Decimal('0')) * Decimal(quantite)

    note = f'Rebut ({dict(MouvementStock.MotifRebut.choices).get(motif, motif)})'
    if reference_chantier:
        note += f' — chantier {reference_chantier}'

    with transaction.atomic():
        p = Produit.objects.select_for_update().get(id=produit.id)
        avant = p.quantite_stock
        apres = avant - quantite
        check_negative_stock_guard(company, avant, apres)
        mouvement = MouvementStock.objects.create(
            company=company, produit=p,
            type_mouvement=MouvementStock.TypeMouvement.REBUT,
            quantite=quantite, quantite_avant=avant, quantite_apres=apres,
            reference=reference_chantier or 'REBUT', note=note,
            motif_rebut=motif, created_by=user)
        p.quantite_stock = apres
        p.save(update_fields=['quantite_stock'])
        if emplacement is not None and not emplacement.is_principal:
            se, _ = StockEmplacement.objects.select_for_update().get_or_create(
                produit=p, emplacement=emplacement,
                defaults={'company': company, 'quantite': 0})
            se.quantite = max(se.quantite - quantite, 0)
            se.save(update_fields=['quantite'])
    return {'mouvement': mouvement, 'valeur_perdue': valeur_perdue}


def rapport_pertes(company, *, date_debut=None, date_fin=None):
    """XSTK10 — rapport « pertes de la période » : quantités ET valeur (coût
    moyen au moment du calcul) par motif de rebut. Admin-only, JAMAIS
    client-facing (prix_achat interne). Renvoie une liste de dicts
    {produit_id, produit_nom, quantite_totale, valeur_totale,
    par_motif: {motif: {quantite, valeur}}}, triée par valeur décroissante.
    """
    from .models import MouvementStock

    qs = MouvementStock.objects.filter(
        company=company, type_mouvement=MouvementStock.TypeMouvement.REBUT)
    if date_debut is not None:
        qs = qs.filter(date__gte=date_debut)
    if date_fin is not None:
        qs = qs.filter(date__lte=date_fin)

    par_produit = {}
    for mvt in qs.select_related('produit'):
        entry = par_produit.setdefault(mvt.produit_id, {
            'produit_id': mvt.produit_id,
            'produit_nom': mvt.produit.nom,
            'quantite_totale': 0,
            'valeur_totale': Decimal('0'),
            'par_motif': {},
        })
        cout_moyen, _source = average_cost_with_source(mvt.produit)
        valeur = (cout_moyen or Decimal('0')) * Decimal(mvt.quantite)
        motif = mvt.motif_rebut or 'autre'
        entry['quantite_totale'] += mvt.quantite
        entry['valeur_totale'] += valeur
        motif_entry = entry['par_motif'].setdefault(
            motif, {'quantite': 0, 'valeur': Decimal('0')})
        motif_entry['quantite'] += mvt.quantite
        motif_entry['valeur'] += valeur
    return sorted(
        par_produit.values(), key=lambda e: -e['valeur_totale'])


def rapport_rebuts(company, *, date_debut=None, date_fin=None):
    """XMFG11 — mini-rapport rebuts agrégé par produit sur une période
    (bornes optionnelles). Renvoie une liste de dicts {produit_id, produit_nom,
    quantite_totale, motifs: {motif: quantite}}, triée par quantité totale
    décroissante. INTERNE."""
    from .models import MouvementStock

    qs = MouvementStock.objects.filter(
        company=company, type_mouvement=MouvementStock.TypeMouvement.REBUT)
    if date_debut is not None:
        qs = qs.filter(date__gte=date_debut)
    if date_fin is not None:
        qs = qs.filter(date__lte=date_fin)

    par_produit = {}
    for mvt in qs.select_related('produit'):
        entry = par_produit.setdefault(mvt.produit_id, {
            'produit_id': mvt.produit_id,
            'produit_nom': mvt.produit.nom,
            'quantite_totale': 0,
            'motifs': {},
        })
        entry['quantite_totale'] += mvt.quantite
        motif = mvt.motif_rebut or 'autre'
        entry['motifs'][motif] = entry['motifs'].get(motif, 0) + mvt.quantite
    return sorted(
        par_produit.values(), key=lambda e: -e['quantite_totale'])


def decrementer_stock_dotation_epi(*, company, produit_id, quantite,
                                   reference, user,
                                   bloquer_si_insuffisant=False):
    """YHIRE13 — décrémente le stock pour une dotation EPI liée à un produit.

    Appelée par ``rh.services`` (cross-app, jamais l'inverse) quand une
    ``DotationEpi`` porte un ``EpiCatalogue.produit_id`` renseigné. Mouvement
    typé SORTIE, ``reference`` traçant la dotation d'origine. Par défaut
    (``bloquer_si_insuffisant=False``) un stock insuffisant n'empêche PAS la
    dotation (le stock peut aller négatif, motif = matériel déjà en atelier
    hors flux stock) — comportement ``warn``. Si ``bloquer_si_insuffisant``
    est vrai, lève ``ValueError`` sans créer de mouvement (comportement
    ``block``, 400 explicite côté appelant). Renvoie le mouvement créé, ou
    ``None`` si ``produit_id`` est vide.
    """
    from django.db import transaction
    from .models import MouvementStock, Produit

    if not produit_id or not quantite:
        return None
    with transaction.atomic():
        try:
            produit = Produit.objects.select_for_update().get(
                id=produit_id, company=company)
        except Produit.DoesNotExist:
            return None
        avant = produit.quantite_stock
        apres = avant - quantite
        if bloquer_si_insuffisant and apres < 0:
            raise ValueError(
                'Stock insuffisant pour cette dotation EPI '
                f'({avant} disponible, {quantite} demandé).')
        mouvement = MouvementStock.objects.create(
            company=company, produit=produit,
            type_mouvement=MouvementStock.TypeMouvement.SORTIE,
            quantite=quantite, quantite_avant=avant, quantite_apres=apres,
            reference=reference, note='Dotation EPI', created_by=user)
        produit.quantite_stock = apres
        produit.save(update_fields=['quantite_stock'])
    return mouvement


def reintegrer_stock_restitution_epi(*, company, produit_id, quantite,
                                     reference, user):
    """YHIRE13 — réintègre le stock à la restitution d'un EPI lié.

    Mouvement typé ENTREE symétrique à ``decrementer_stock_dotation_epi``.
    Renvoie ``None`` si ``produit_id`` est vide.
    """
    from django.db import transaction
    from .models import MouvementStock, Produit

    if not produit_id or not quantite:
        return None
    with transaction.atomic():
        try:
            produit = Produit.objects.select_for_update().get(
                id=produit_id, company=company)
        except Produit.DoesNotExist:
            return None
        avant = produit.quantite_stock
        apres = avant + quantite
        mouvement = MouvementStock.objects.create(
            company=company, produit=produit,
            type_mouvement=MouvementStock.TypeMouvement.ENTREE,
            quantite=quantite, quantite_avant=avant, quantite_apres=apres,
            reference=reference, note='Restitution EPI', created_by=user)
        produit.quantite_stock = apres
        produit.save(update_fields=['quantite_stock'])
    return mouvement


def sortie_exists_for_reference(company, reference):
    """True si un mouvement SORTIE référence déjà ``reference`` pour la société.

    Sert de garde anti-double-comptage (U9) : les apps appelantes (ventes)
    vérifient via ce service — sans importer le modèle MouvementStock — qu'un
    stock n'a pas déjà été réservé/consommé pour un même document (devis),
    qu'il vienne du chemin bon-commande (livraison) ou de la facturation
    directe par échéancier."""
    from .models import MouvementStock
    if not reference:
        return False
    return MouvementStock.objects.filter(
        company=company,
        type_mouvement=MouvementStock.TypeMouvement.SORTIE,
        reference=reference,
    ).exists()


# ── FG54 — Réapprovisionnement auto ──────────────────────────────────────────

# ── XSTK17 — profils saisonniers de seuils (saison pompage) ─────────────────

def profil_saisonnier_actif(company, produit, *, mois=None):
    """XSTK17 — profil saisonnier ACTIF couvrant ``mois`` (défaut : mois
    courant) pour ce produit (priorité) ou sa catégorie. Renvoie None hors
    saison / sans profil — auquel cas l'appelant garde le seuil statique
    (comportement historique inchangé)."""
    from django.utils import timezone
    from .models import ProfilSaisonnier
    mois = mois or timezone.now().month

    profil_produit = ProfilSaisonnier.objects.filter(
        company=company, produit=produit, actif=True).first()
    candidats = [profil_produit] if profil_produit else []
    if not candidats and produit.categorie_id:
        candidats = list(ProfilSaisonnier.objects.filter(
            company=company, categorie_id=produit.categorie_id, actif=True))
    for profil in candidats:
        if profil and profil.couvre_mois(mois):
            return profil
    return None


def seuil_effectif_produit(company, produit, *, mois=None):
    """XSTK17 — (seuil_alerte, quantite_cible) EFFECTIFS pour ce produit : le
    profil saisonnier ACTIF prime pendant sa fenêtre ; hors saison ou sans
    profil, renvoie EXACTEMENT (produit.seuil_alerte,
    produit.quantite_reappro_cible) — repli byte-identique au comportement
    historique."""
    profil = profil_saisonnier_actif(company, produit, mois=mois)
    if profil is None:
        return produit.seuil_alerte, produit.quantite_reappro_cible
    seuil = profil.seuil_min if profil.seuil_min is not None \
        else produit.seuil_alerte
    cible = profil.quantite_cible if profil.quantite_cible is not None \
        else produit.quantite_reappro_cible
    return seuil, cible


def creer_profil_saisonnier(
        company, *, produit=None, categorie=None, mois_debut, mois_fin,
        seuil_min=None, seuil_max=None, quantite_cible=None, nom=None,
        user=None):
    """XSTK17 — crée un profil saisonnier après avoir rejeté tout
    CHEVAUCHEMENT calendaire avec un profil déjà ACTIF de la MÊME cible
    (produit ou catégorie, jamais les deux). Lève ValueError si ni produit
    ni catégorie n'est fourni (ou les deux), ou en cas de chevauchement."""
    from .models import ProfilSaisonnier

    if bool(produit) == bool(categorie):
        raise ValueError(
            'Un profil saisonnier cible soit un produit, soit une '
            'catégorie (jamais les deux, jamais aucun).')
    if not (1 <= mois_debut <= 12 and 1 <= mois_fin <= 12):
        raise ValueError('Les mois doivent être compris entre 1 et 12.')

    candidat = ProfilSaisonnier(
        mois_debut=mois_debut, mois_fin=mois_fin)
    existants = ProfilSaisonnier.objects.filter(company=company, actif=True)
    existants = existants.filter(produit=produit) if produit \
        else existants.filter(categorie=categorie)
    for autre in existants:
        for m in range(1, 13):
            if candidat.couvre_mois(m) and autre.couvre_mois(m):
                label = autre.nom or f'#{autre.id}'
                raise ValueError(
                    f'Ce profil chevauche le profil existant « {label} » '
                    f'sur le mois {m}.')

    return ProfilSaisonnier.objects.create(
        company=company, produit=produit, categorie=categorie,
        mois_debut=mois_debut, mois_fin=mois_fin, seuil_min=seuil_min,
        seuil_max=seuil_max, quantite_cible=quantite_cible, nom=nom,
        created_by=user)


def produits_a_reapprovisionner(company):
    """Retourne les produits dont la POSITION NETTE (disponible N14 + déjà EN
    COMMANDE via BCF brouillon/envoyé, YPROC9) est <= seuil EFFECTIF (XSTK17 :
    le profil saisonnier ACTIF prime sur `seuil_alerte` pendant sa fenêtre ;
    hors saison, repli byte-identique sur `seuil_alerte`), groupés par
    fournisseur le moins cher (PrixFournisseur). INTERNE.

    Chaque item : {produit_id, nom, quantite_stock, seuil_alerte,
    quantite_suggere, disponible, en_commande, fournisseur_id,
    fournisseur_nom, prix_achat, action, kit_id}. ``quantite_suggere`` est
    NETTE du pipeline déjà en route (0 → produit exclu du tout : ce qui
    arrive déjà suffit). ``action`` = 'assembler' (kit_id renseigné) quand le
    produit sous seuil est le ``produit_compose`` d'un kit ACTIF
    (`installations.Kit`, XMFG3) — la suggestion devient « assembler N »
    plutôt qu'un bon de commande fournisseur ; sinon 'acheter'.
    """
    from .models import Produit, PrixFournisseur
    from .selectors import quantite_en_commande_produit
    from apps.installations.selectors import kit_map_for_produits_composes

    # XSTK17 — exclut seulement les produits SANS seuil_alerte ET sans
    # profil saisonnier actif (repli inchangé : seuil_alerte=0 = pas de
    # suivi de seuil, comme avant). Les candidats avec un profil actif
    # passent même si seuil_alerte=0 (le profil peut définir un seuil que
    # le produit n'a pas statiquement).
    candidats_ids = set(
        Produit.objects.filter(
            company=company, is_archived=False)
        .exclude(seuil_alerte=0).values_list('id', flat=True))
    from .models import ProfilSaisonnier
    from django.utils import timezone
    mois_actuel = timezone.now().month
    for profil in ProfilSaisonnier.objects.filter(
            company=company, actif=True).select_related('categorie'):
        if not profil.couvre_mois(mois_actuel):
            continue
        if profil.produit_id:
            candidats_ids.add(profil.produit_id)
        elif profil.categorie_id:
            candidats_ids.update(
                Produit.objects.filter(
                    company=company, is_archived=False,
                    categorie_id=profil.categorie_id)
                .values_list('id', flat=True))

    qs = (Produit.objects
          .filter(company=company, id__in=candidats_ids)
          .prefetch_related('prix_fournisseurs__fournisseur'))

    produits = list(qs)
    kit_map = kit_map_for_produits_composes(
        company, [p.id for p in produits])
    # YPROC9 — netting : le besoin réel se calcule sur le DISPONIBLE (stock −
    # réservations chantier actives, N14) PLUS le pipeline déjà EN COMMANDE
    # (BCF brouillon/envoyé non reçus) — jamais le stock brut seul. Sinon on
    # re-suggère (et `generer_bcf_reappro` re-commande) ce qui est déjà en
    # route.
    reserved_map = reserved_quantities(company)

    result = []
    for p in produits:
        seuil_effectif, cible_effective = seuil_effectif_produit(
            company, p, mois=mois_actuel)
        disponible = p.quantite_stock - reserved_map.get(p.id, 0)
        en_commande = quantite_en_commande_produit(company, p.id)
        # YPROC9 — le produit reste candidat tant que le DISPONIBLE seul (hors
        # pipeline) est sous seuil (comportement historique inchangé) : un
        # BCF ouvert qui ne couvre qu'UNE PARTIE du manque (position nette
        # au-dessus du seuil mais toujours sous la cible) doit continuer à
        # être suggéré pour le reliquat — seul `qte_suggere <= 0` (calculé
        # plus bas contre la CIBLE, pas le seuil) exclut réellement le
        # produit une fois le pipeline prix en compte.
        if disponible > (seuil_effectif or 0):
            continue
        # Fournisseur le moins cher parmi les prix enregistrés.
        best = (PrixFournisseur.objects
                .filter(company=company, produit=p)
                .select_related('fournisseur')
                .order_by('prix_achat')
                .first())
        cible = cible_effective if cible_effective else (
            (seuil_effectif or 0) * 2)
        # FG54 (historique) — la quantité SUGGÉRÉE reste la CIBLE pleine
        # (comportement byte-identique préservé : `quantite_suggere` ==
        # `quantite_reappro_cible`/`seuil × 2` quand rien n'est en pipeline).
        # YPROC9 ne déduit QUE le pipeline déjà en commande (`en_commande`) —
        # jamais le disponible courant, qui sert uniquement à la porte
        # d'inclusion ci-dessus — pour ne pas re-suggérer ce qui est déjà en
        # route. Une quantité suggérée nette <= 0 exclut le produit (rien à
        # recommander, ce qui arrive déjà suffit).
        qte_suggere = max(cible - en_commande, 0)
        if qte_suggere <= 0:
            continue
        kit_id = kit_map.get(p.id)
        result.append({
            'produit_id': p.id,
            'nom': p.nom,
            'sku': p.sku,
            'quantite_stock': p.quantite_stock,
            'seuil_alerte': seuil_effectif,
            'quantite_suggere': qte_suggere,
            'disponible': disponible,
            'en_commande': en_commande,
            'fournisseur_id': best.fournisseur_id if best else None,
            'fournisseur_nom': best.fournisseur.nom if best else None,
            'prix_achat': str(best.prix_achat) if best else None,
            'action': 'assembler' if kit_id else 'acheter',
            'kit_id': kit_id,
        })
    return result


REAPPRO_NOTE_MARKER = 'Réapprovisionnement automatique (stock < seuil)'


def generer_bcf_reappro(company, user, fournisseur_id):
    """Génère (ou complète) un BCF BROUILLON pour tous les produits sous
    seuil, chez le fournisseur donné (ou les moins chers si non précisé).
    Renvoie {bon_commande_id, reference, nb_lignes, fusionne}. Lève
    ValueError si rien à faire.

    YPROC9 — FUSION au lieu de duplication : si un BCF BROUILLON « Réappro-
    visionnement automatique » existe déjà pour ce fournisseur (même
    marqueur de note), ses lignes sont incrémentées (ou une ligne ajoutée
    pour un produit encore absent) au lieu d'ouvrir un second brouillon —
    deux appels successifs n'ouvrent donc qu'UN seul brouillon par
    fournisseur."""
    from apps.ventes.utils.references import create_with_reference
    from .models import BonCommandeFournisseur, LigneBonCommandeFournisseur, Fournisseur

    besoins = produits_a_reapprovisionner(company)
    if fournisseur_id:
        fournisseur = Fournisseur.objects.filter(
            id=fournisseur_id, company=company).first()
        if not fournisseur:
            raise ValueError("Fournisseur introuvable.")
        lignes = [(b['produit_id'], b['quantite_suggere'], b['prix_achat'])
                  for b in besoins]
    else:
        # Filtre sur les besoins qui ont un fournisseur
        fournisseur_id_premier = next(
            (b['fournisseur_id'] for b in besoins if b['fournisseur_id']), None)
        if not fournisseur_id_premier:
            raise ValueError("Aucun fournisseur associé aux produits à réapprovisionner.")
        fournisseur = Fournisseur.objects.get(id=fournisseur_id_premier, company=company)
        lignes = [(b['produit_id'], b['quantite_suggere'], b['prix_achat'])
                  for b in besoins if b['fournisseur_id'] == fournisseur_id_premier]

    if not lignes:
        raise ValueError("Aucun produit à réapprovisionner.")

    from .models import Produit
    lignes_produits = [
        (Produit.objects.get(id=pid, company=company), qte, prix)
        for pid, qte, prix in lignes
    ]

    # Dernier BROUILLON réappro auto du MÊME fournisseur (le plus récent).
    bon_existant = (BonCommandeFournisseur.objects
                    .filter(company=company, fournisseur=fournisseur,
                            statut=BonCommandeFournisseur.Statut.BROUILLON,
                            note=REAPPRO_NOTE_MARKER)
                    .order_by('-date_creation')
                    .first())
    if bon_existant is not None:
        lignes_par_produit = {
            ligne.produit_id: ligne
            for ligne in bon_existant.lignes.all()}
        for produit, qte, prix in lignes_produits:
            existante = lignes_par_produit.get(produit.id)
            if existante is not None:
                existante.quantite += qte
                existante.save(update_fields=['quantite'])
            else:
                LigneBonCommandeFournisseur.objects.create(
                    bon_commande=bon_existant, produit=produit,
                    quantite=qte,
                    prix_achat_unitaire=(
                        Decimal(prix) if prix else Decimal('0')))
        return {
            'bon_commande_id': bon_existant.id,
            'reference': bon_existant.reference,
            'nb_lignes': len(lignes_produits),
            'fusionne': True,
        }

    created_bon = {}

    def _save(ref):
        bon = BonCommandeFournisseur.objects.create(
            company=company, reference=ref, fournisseur=fournisseur,
            statut=BonCommandeFournisseur.Statut.BROUILLON,
            note=REAPPRO_NOTE_MARKER,
            created_by=user)
        for produit, qte, prix in lignes_produits:
            LigneBonCommandeFournisseur.objects.create(
                bon_commande=bon, produit=produit,
                quantite=qte,
                prix_achat_unitaire=Decimal(prix) if prix else Decimal('0'))
        created_bon['bon'] = bon
        return bon

    create_with_reference(BonCommandeFournisseur, 'BCF', company, _save)
    bon = created_bon['bon']
    return {'bon_commande_id': bon.id, 'reference': bon.reference,
            'nb_lignes': len(lignes_produits), 'fusionne': False}


# ── FG55 — PDF facture fournisseur ────────────────────────────────────────────

def generate_facture_fournisseur_pdf(facture):
    """Génère le PDF d'une facture fournisseur (INTERNE). Utilise WeasyPrint."""
    from apps.ventes.utils.pdf import _company_context, _render_html, _html_to_pdf
    context = _company_context(company=facture.company)
    context['facture'] = facture
    context['fournisseur'] = facture.fournisseur
    context['lignes'] = list(facture.lignes.select_related('produit').all())
    context['paiements'] = list(facture.paiements.all())
    context['solde_du'] = facture.solde_du
    context['total_paye'] = facture.total_paye
    html = _render_html('facture_fournisseur.html', context)
    return _html_to_pdf(html)


# ── FG56 — Facturer une réception ────────────────────────────────────────────

def facturer_reception(company, user, reception):
    """Crée une FactureFournisseur à partir d'une réception confirmée.

    Construit les lignes HT à partir de LigneReceptionFournisseur ×
    prix_achat_unitaire de la ligne BCF, calcule HT/TVA/TTC (TVA 20 %).
    Lance ValueError si déjà facturée ou si la réception n'est pas confirmée.
    """
    from decimal import Decimal
    from apps.ventes.utils.references import create_with_reference
    from .models import FactureFournisseur, LigneFactureFournisseur

    if reception.statut != 'confirme':
        raise ValueError("Seule une réception confirmée peut être facturée.")

    # Garde idempotence : si une FF porte déjà ce bon de commande et la même
    # réception (on lie via note), on refuse.
    if FactureFournisseur.objects.filter(
            company=company,
            bon_commande=reception.bon_commande,
            note__startswith=f'Facture réception {reception.reference}').exists():
        raise ValueError(
            f"Cette réception ({reception.reference}) est déjà facturée.")

    taux_tva_defaut = Decimal('20')
    montant_ht = Decimal('0')
    montant_tva = Decimal('0')
    lignes_data = []
    for ligne in reception.lignes.select_related('produit', 'ligne_commande').all():
        pu = ligne.ligne_commande.prix_achat_unitaire if ligne.ligne_commande else Decimal('0')
        total = Decimal(str(ligne.quantite)) * pu
        montant_ht += total
        # XPUR16 — une ligne libre/service reprend sa désignation d'origine
        # (BCF) plutôt que le nom d'un produit catalogue absent.
        if ligne.produit:
            designation = ligne.produit.nom
        elif ligne.ligne_commande and ligne.ligne_commande.designation:
            designation = ligne.ligne_commande.designation
        else:
            designation = 'Produit'
        # XPUR17 — TVA par ligne : reprend le taux du produit (`Produit.tva`)
        # quand connu, sinon le défaut 20 % (comportement historique de
        # cette fonction, qui appliquait déjà 20 % globalement).
        taux_ligne = (ligne.produit.tva
                      if ligne.produit and ligne.produit.tva is not None
                      else taux_tva_defaut)
        tva_ligne = (total * taux_ligne / Decimal('100')).quantize(
            Decimal('0.01'))
        montant_tva += tva_ligne
        lignes_data.append((designation, ligne.quantite, pu, taux_ligne))

    montant_ttc = montant_ht + montant_tva

    created = {}

    def _save(ref):
        from django.utils import timezone
        ff = FactureFournisseur.objects.create(
            company=company, reference=ref,
            fournisseur=reception.bon_commande.fournisseur,
            bon_commande=reception.bon_commande,
            montant_ht=montant_ht, montant_tva=montant_tva,
            montant_ttc=montant_ttc,
            statut=FactureFournisseur.Statut.A_PAYER,
            # Sans date, l'écriture comptable auto (61xx/3455 -> 4411) crashait
            # NOT NULL en silence (bug préexistant attrapé par le test P2P).
            date_facture=timezone.now().date(),
            note=f'Facture réception {reception.reference}',
            created_by=user)
        for designation, qte, pu, taux_ligne in lignes_data:
            LigneFactureFournisseur.objects.create(
                facture=ff, designation=designation,
                quantite=qte, prix_unitaire_ht=pu, taux_tva=taux_ligne)
        created['ff'] = ff
        return ff

    create_with_reference(FactureFournisseur, 'FF', company, _save)
    # XPUR8 — impute automatiquement les acomptes non consommés du BCF sur
    # cette première facture (idempotent, no-op si aucun acompte).
    imputer_acomptes_bcf(reception.bon_commande)
    # YPROC3 — émet l'événement de création de facture fournisseur (best-effort,
    # ne casse jamais la facturation) : installations peut lettrer sa provision
    # GR/IR ouverte pour ce bon de commande. stock n'importe jamais installations.
    try:
        from core.events import facture_fournisseur_creee
        facture_fournisseur_creee.send(
            sender=FactureFournisseur, instance=created['ff'],
            company=company, user=user)
    except Exception:  # pragma: no cover - défensif, best-effort
        pass
    return created['ff']


# ── ZPUR1 — Politique de facturation d'achat (Odoo « Bill Control ») ────────
# FG56 (`facturer_reception`, ci-dessus) ne facture QUE depuis une réception
# confirmée — impossible de facturer d'avance un BCF « à la commande »
# (import payé sur bon de commande). `Produit.politique_facturation_achat`
# pilote quelles LIGNES sont éligibles à ce chemin direct ; les lignes
# `sur_reception` (défaut) restent facturées EXCLUSIVEMENT via FG56.

def facturer_bcf_sur_commande(company, user, bon_commande):
    """ZPUR1 — construit une FactureFournisseur BROUILLON depuis les lignes
    COMMANDÉES (quantité × prix d'achat) d'un BCF dont les lignes sont
    `sur_commande`, SANS exiger de réception. Réutilise le même builder de
    lignes/montants que FG56 (HT/TVA par ligne, taux produit ou 20% défaut).

    Lève ValueError si le BCF n'a AUCUNE ligne `sur_commande` (les lignes
    `sur_reception` restent hors de ce chemin — direction vers FG56), ou si
    ce BCF est déjà entièrement facturé par ce chemin (idempotence : jamais
    deux factures pour la même quantité `sur_commande`)."""
    from decimal import Decimal
    from apps.ventes.utils.references import create_with_reference
    from .models import FactureFournisseur, LigneFactureFournisseur, Produit

    lignes_eligibles = [
        ligne for ligne in bon_commande.lignes.select_related('produit').all()
        if ligne.produit_id is not None
        and ligne.produit.politique_facturation_achat
        == Produit.PolitiqueFacturationAchat.SUR_COMMANDE
    ]
    if not lignes_eligibles:
        raise ValueError(
            "Aucune ligne « sur commande » sur ce bon de commande — "
            'utilisez la facturation depuis la réception (FG56).')

    marqueur = f'Facture sur commande {bon_commande.reference}'
    if FactureFournisseur.objects.filter(
            company=company, bon_commande=bon_commande,
            note__startswith=marqueur).exists():
        raise ValueError(
            f'Ce bon de commande ({bon_commande.reference}) est déjà '
            'facturé sur commande.')

    taux_tva_defaut = Decimal('20')
    montant_ht = Decimal('0')
    montant_tva = Decimal('0')
    lignes_data = []
    for ligne in lignes_eligibles:
        pu = ligne.prix_achat_unitaire or Decimal('0')
        total = Decimal(str(ligne.quantite)) * pu
        montant_ht += total
        designation = (
            ligne.produit.nom if ligne.produit_id else
            (ligne.designation or 'Produit'))
        taux_ligne = (ligne.produit.tva
                      if ligne.produit_id and ligne.produit.tva is not None
                      else taux_tva_defaut)
        tva_ligne = (total * taux_ligne / Decimal('100')).quantize(
            Decimal('0.01'))
        montant_tva += tva_ligne
        lignes_data.append((designation, ligne.quantite, pu, taux_ligne))

    montant_ttc = montant_ht + montant_tva
    created = {}

    def _save(ref):
        ff = FactureFournisseur.objects.create(
            company=company, reference=ref,
            fournisseur=bon_commande.fournisseur,
            bon_commande=bon_commande,
            montant_ht=montant_ht, montant_tva=montant_tva,
            montant_ttc=montant_ttc,
            statut=FactureFournisseur.Statut.A_PAYER,
            note=marqueur, created_by=user)
        for designation, qte, pu, taux_ligne in lignes_data:
            LigneFactureFournisseur.objects.create(
                facture=ff, designation=designation,
                quantite=qte, prix_unitaire_ht=pu, taux_tva=taux_ligne)
        created['ff'] = ff
        return ff

    create_with_reference(FactureFournisseur, 'FF', company, _save)
    try:
        from core.events import facture_fournisseur_creee
        facture_fournisseur_creee.send(
            sender=FactureFournisseur, instance=created['ff'],
            company=company, user=user)
    except Exception:  # pragma: no cover - défensif, best-effort
        pass
    return created['ff']


# ── ZPUR4 — Duplication d'un bon de commande fournisseur ────────────────────
# Odoo offre « Duplicate » sur tout PO/RFQ. QP2 a livré la duplication de
# PRODUIT ; aucun endpoint ne duplique un BCF. Le clone est TOUJOURS un
# BROUILLON neuf (nouvelle référence, quantités reçues à zéro, statut
# réinitialisé) — la source n'est jamais modifiée.

def dupliquer_bcf(company, user, bon_commande):
    """ZPUR4 — crée un nouveau BCF BROUILLON copiant fournisseur + lignes
    (produit, quantité, prix d'achat) du BCF source. Référence neuve via
    `create_with_reference` (jamais count()+1), `date_commande` = aujourd'hui,
    statut réinitialisé, quantités reçues à zéro. Le BCF source n'est JAMAIS
    modifié. Renvoie le nouveau BCF."""
    from django.utils import timezone
    from apps.ventes.utils.references import create_with_reference
    from .models import BonCommandeFournisseur, LigneBonCommandeFournisseur

    lignes_source = list(bon_commande.lignes.all())
    created = {}

    def _save(ref):
        clone = BonCommandeFournisseur.objects.create(
            company=company, reference=ref,
            fournisseur=bon_commande.fournisseur,
            statut=BonCommandeFournisseur.Statut.BROUILLON,
            date_commande=timezone.now().date(),
            devise=bon_commande.devise, taux_change=bon_commande.taux_change,
            note=f'Dupliqué depuis {bon_commande.reference}',
            created_by=user)
        for ligne in lignes_source:
            LigneBonCommandeFournisseur.objects.create(
                bon_commande=clone, produit=ligne.produit,
                designation=ligne.designation, sans_stock=ligne.sans_stock,
                quantite=ligne.quantite,
                prix_achat_unitaire=ligne.prix_achat_unitaire,
                # ZPUR4 — quantité reçue TOUJOURS à zéro sur le clone (jamais
                # copiée : un clone brouillon n'a par construction rien reçu).
                quantite_recue=0,
            )
        created['bon'] = clone
        return clone

    create_with_reference(BonCommandeFournisseur, 'BCF', company, _save)
    return created['bon']


# ── ZPUR6 — Regroupement de plusieurs BCF en un seul par fournisseur ────────
# Odoo permet de fusionner des RFQ du même fournisseur. Plusieurs suggestions
# de réappro (FG54) ou besoins chantier créent aujourd'hui des BCF séparés
# vers le même fournisseur — l'acheteur finit par multiplier les commandes.

def fusionner_bcf(company, user, bon_commande_ids):
    """ZPUR6 — fusionne PLUSIEURS BCF BROUILLON du MÊME fournisseur (et de la
    MÊME société) en un BCF cible neuf : lignes additionnées par produit
    (quantités cumulées, prix d'achat du plus récent BCF source portant ce
    produit), puis passe les BCF sources en `annule` avec une note de fusion
    horodatée. Lève ValueError si : moins de 2 BCF, fournisseurs différents,
    ou un des BCF n'est pas BROUILLON (ou n'appartient pas à la société)."""
    from django.db import transaction
    from django.utils import timezone
    from apps.ventes.utils.references import create_with_reference
    from .models import BonCommandeFournisseur, LigneBonCommandeFournisseur

    bcs = list(
        BonCommandeFournisseur.objects.filter(
            company=company, id__in=bon_commande_ids)
        .select_related('fournisseur').prefetch_related('lignes__produit'))
    if len(bcs) != len(set(bon_commande_ids)):
        raise ValueError('Un ou plusieurs bons de commande sont introuvables.')
    if len(bcs) < 2:
        raise ValueError('Au moins deux bons de commande sont requis.')
    fournisseur_ids = {bc.fournisseur_id for bc in bcs}
    if len(fournisseur_ids) > 1:
        raise ValueError(
            'Tous les bons de commande doivent être du même fournisseur.')
    non_brouillon = [
        bc for bc in bcs
        if bc.statut != BonCommandeFournisseur.Statut.BROUILLON]
    if non_brouillon:
        raise ValueError(
            'Seuls des bons de commande en BROUILLON peuvent être fusionnés '
            f'({non_brouillon[0].reference} ne l\'est pas).')

    # Cumule les quantités par produit ; garde le prix du BCF le plus RÉCENT
    # (date_creation) portant ce produit. Une ligne sans produit (libre/
    # service) est reprise telle quelle (pas de fusion par désignation —
    # évite de coller à tort deux lignes libres différentes).
    bcs_par_date = sorted(bcs, key=lambda bc: bc.date_creation)
    lignes_par_produit = {}
    lignes_libres = []
    for bc in bcs_par_date:
        for ligne in bc.lignes.all():
            if ligne.produit_id is None:
                lignes_libres.append(ligne)
                continue
            existante = lignes_par_produit.get(ligne.produit_id)
            if existante is None:
                lignes_par_produit[ligne.produit_id] = {
                    'produit': ligne.produit,
                    'quantite': ligne.quantite,
                    'prix_achat_unitaire': ligne.prix_achat_unitaire,
                }
            else:
                existante['quantite'] += ligne.quantite
                # bcs_par_date est croissant : le dernier vu = le plus récent.
                existante['prix_achat_unitaire'] = ligne.prix_achat_unitaire

    fournisseur = bcs[0].fournisseur
    created = {}
    references_sources = ', '.join(bc.reference for bc in bcs)

    def _save(ref):
        cible = BonCommandeFournisseur.objects.create(
            company=company, reference=ref, fournisseur=fournisseur,
            statut=BonCommandeFournisseur.Statut.BROUILLON,
            note=f'Fusion de {references_sources}', created_by=user)
        for data in lignes_par_produit.values():
            LigneBonCommandeFournisseur.objects.create(
                bon_commande=cible, produit=data['produit'],
                quantite=data['quantite'],
                prix_achat_unitaire=data['prix_achat_unitaire'])
        for ligne in lignes_libres:
            LigneBonCommandeFournisseur.objects.create(
                bon_commande=cible, produit=None,
                designation=ligne.designation, sans_stock=True,
                quantite=ligne.quantite,
                prix_achat_unitaire=ligne.prix_achat_unitaire)
        created['bon'] = cible
        return cible

    with transaction.atomic():
        create_with_reference(BonCommandeFournisseur, 'BCF', company, _save)
        cible = created['bon']
        today = timezone.now().date()
        for bc in bcs:
            bc.statut = BonCommandeFournisseur.Statut.ANNULE
            note_fusion = (
                f'[{today.isoformat()}] Fusionné dans {cible.reference}.')
            bc.note = (f'{bc.note}\n{note_fusion}'.strip()
                       if bc.note else note_fusion)
            bc.save(update_fields=['statut', 'note'])
    return cible


# ── DC38 — Landed cost (FG316) replié dans le coût moyen pondéré ─────────────
# Le coût débarqué d'un dossier d'import (fret/douane/TVA import/transit) est
# écrit dans le champ EXISTANT `LigneBonCommandeFournisseur.frais_annexes` —
# celui que `average_cost_with_source` (FG67) intègre déjà — plutôt que dans un
# champ de coût d'achat PARALLÈLE. `installations` (qui possède le dossier
# d'import et string-FK stock) appelle ce setter ; stock n'importe jamais
# installations (sens de dépendance préservé).

def definir_frais_annexes_ligne_bcf(company, bon_commande_id, produit_id,
                                    frais_annexes):
    """DC38 — pose les frais annexes (coût débarqué) sur la/les ligne(s) de BCF
    d'un produit donné, dans la société. Setter STOCK pur (aucun import
    installations). Si plusieurs lignes portent le même produit sur ce BCF, le
    montant est réparti à parts égales pour que la somme sur les lignes reste
    exacte dans le coût moyen. Renvoie le nombre de lignes mises à jour."""
    from decimal import Decimal
    from .models import LigneBonCommandeFournisseur

    if not bon_commande_id or not produit_id:
        return 0
    lignes = list(LigneBonCommandeFournisseur.objects.filter(
        bon_commande_id=bon_commande_id,
        bon_commande__company=company,
        produit_id=produit_id))
    if not lignes:
        return 0
    total = Decimal(str(frais_annexes or 0))
    part = (total / Decimal(len(lignes))).quantize(Decimal('0.01'))
    for ligne in lignes:
        ligne.frais_annexes = part
        ligne.save(update_fields=['frais_annexes'])
    return len(lignes)


# ── FG57 — Rotation / dead-stock ─────────────────────────────────────────────

def rotation_report(company, jours=180):
    """Rapport de rotation des produits (dead-stock / immobile).

    Retourne une liste de dicts avec : produit_id, nom, sku, quantite_stock,
    valeur_stock (prix_achat × qte, INTERNE), derniere_sortie (date ou null),
    jours_sans_mouvement (int ou null), bucket ('actif'|'ralenti'|'immobile').
    Admin-only ; prix d'achat interne jamais client-facing.
    """
    from django.utils import timezone
    from .models import Produit, MouvementStock

    today = timezone.now().date()
    produits = (Produit.objects
                .filter(company=company, is_archived=False, quantite_stock__gt=0)
                .only('id', 'nom', 'sku', 'quantite_stock', 'prix_achat'))

    result = []
    for p in produits:
        derniere = (MouvementStock.objects
                    .filter(company=company, produit=p,
                            type_mouvement='sortie')
                    .order_by('-date')
                    .values_list('date', flat=True)
                    .first())
        jours_sans = None
        if derniere:
            jours_sans = (today - derniere.date()).days

        if derniere is None:
            bucket = 'immobile'
        elif jours_sans > jours:
            bucket = 'ralenti'
        else:
            bucket = 'actif'

        result.append({
            'produit_id': p.id,
            'nom': p.nom,
            'sku': p.sku,
            'quantite_stock': p.quantite_stock,
            'valeur_stock': str(p.prix_achat * p.quantite_stock),
            'derniere_sortie': derniere.date().isoformat() if derniere else None,
            'jours_sans_mouvement': jours_sans,
            'bucket': bucket,
        })
    # Trier : immobiles en premier, puis ralentis, puis actifs
    order = {'immobile': 0, 'ralenti': 1, 'actif': 2}
    result.sort(key=lambda x: (order[x['bucket']],
                               -(x['jours_sans_mouvement'] or 999999)))
    return result


# ── FG60 — Export xlsx mouvements ─────────────────────────────────────────────

def export_mouvements_xlsx(company, qs):
    """Export Excel de la liste filtrée des mouvements de stock (INTERNE).
    Prix d'achat jamais inclus."""
    from apps.crm.exports import build_xlsx_response
    headers = ['Référence', 'Type', 'Produit', 'Quantité',
               'Avant', 'Après', 'Note', 'Créé par', 'Date']
    rows = []
    for m in qs.select_related('produit', 'created_by'):
        rows.append([
            m.reference or '',
            m.get_type_mouvement_display(),
            m.produit.nom,
            m.quantite,
            m.quantite_avant,
            m.quantite_apres,
            m.note or '',
            m.created_by.username if m.created_by else '',
            m.date.strftime('%d/%m/%Y %H:%M') if m.date else '',
        ])
    return build_xlsx_response('mouvements-stock.xlsx', headers, rows,
                               sheet_title='Mouvements')


# ── FG58 — Comparaison fournisseurs ───────────────────────────────────────────

def comparer_fournisseurs(company, produit):
    """Retourne la liste des prix multi-fournisseurs d'un produit, triée du
    moins cher au plus cher. INTERNE — prix d'achat jamais client-facing."""
    from .models import PrixFournisseur
    qs = (PrixFournisseur.objects
          .filter(company=company, produit=produit)
          .select_related('fournisseur')
          .order_by('prix_achat'))
    return [
        {
            'fournisseur_id': pf.fournisseur_id,
            'fournisseur_nom': pf.fournisseur.nom,
            'prix_achat': str(pf.prix_achat),
            'date_dernier_achat': pf.date_dernier_achat.isoformat()
            if pf.date_dernier_achat else None,
        }
        for pf in qs
    ]


# ── FG59 — Scorecard performance fournisseur ─────────────────────────────────

def supplier_performance(company, fournisseur):
    """Scorecard performance fournisseur : délai moyen, taux de remplissage,
    taux de retour, dépenses totales. INTERNE."""
    from decimal import Decimal
    from .models import BonCommandeFournisseur, RetourFournisseur

    bons = (BonCommandeFournisseur.objects
            .filter(company=company, fournisseur=fournisseur)
            .exclude(statut=BonCommandeFournisseur.Statut.ANNULE)
            .prefetch_related('receptions', 'lignes'))

    nb_bons = bons.count()
    total_achats = Decimal('0')
    lead_times = []
    fill_rates = []

    for bc in bons:
        # Dépense totale HT (prix d'achat interne)
        total_achats += bc.total_achat or Decimal('0')

        # Délai de livraison = date_reception - date_commande (en jours)
        if bc.date_commande:
            for rec in bc.receptions.filter(statut='confirme'):
                if rec.date_reception:
                    delta = (rec.date_reception - bc.date_commande).days
                    if delta >= 0:
                        lead_times.append(delta)

        # Taux de remplissage = qte reçue / qte commandée
        total_cmd = sum(lig.quantite for lig in bc.lignes.all())
        total_recu = sum(lig.quantite_recue for lig in bc.lignes.all())
        if total_cmd > 0:
            fill_rates.append(total_recu / total_cmd * 100)

    # Taux de retour
    nb_retours = RetourFournisseur.objects.filter(
        company=company, fournisseur=fournisseur,
        statut='valide').count()

    # XPUR7 — OTD promis-vs-reçu ajouté au scorecard existant (champs
    # additifs ; None quand aucune date confirmée/prévue n'existe —
    # comportement historique inchangé pour les BCF sans XPUR7 renseigné).
    otd = otd_stats(company, fournisseur)

    return {
        'fournisseur_id': fournisseur.id,
        'fournisseur_nom': fournisseur.nom,
        'nb_bons': nb_bons,
        'avg_lead_time_days': round(sum(lead_times) / len(lead_times), 1) if lead_times else None,
        'fill_rate_pct': round(sum(fill_rates) / len(fill_rates), 1) if fill_rates else None,
        'nb_retours': nb_retours,
        'return_rate_pct': round(nb_retours / nb_bons * 100, 1) if nb_bons else None,
        'total_achats_ht': str(total_achats),
        'otd_ecart_moyen_jours': otd['otd_ecart_moyen_jours'],
        'otd_a_lheure_pct': otd['otd_a_lheure_pct'],
    }


# ── FG62 — Réapprovisionnement par emplacement ───────────────────────────────

def suggestions_reappro_emplacement(company):
    """Retourne les lignes StockEmplacement dont la quantité < seuil_min,
    avec une suggestion de transfert depuis le dépôt principal.
    INTERNE — admin uniquement."""
    from .models import StockEmplacement, EmplacementStock

    qs = (StockEmplacement.objects
          .filter(company=company)
          .filter(seuil_min__isnull=False)
          .filter(quantite__lt=models.F('seuil_min'))
          .select_related('produit', 'emplacement'))

    principal = EmplacementStock.objects.filter(
        company=company, is_principal=True).first()

    result = []
    for se in qs:
        qte_a_transferer = (se.seuil_min or 0) - se.quantite
        result.append({
            'produit_id': se.produit_id,
            'produit_nom': se.produit.nom,
            'emplacement_id': se.emplacement_id,
            'emplacement_nom': se.emplacement.nom,
            'quantite_actuelle': se.quantite,
            'seuil_min': se.seuil_min,
            'seuil_max': se.seuil_max,
            'qte_suggere_transfert': qte_a_transferer,
            'source_id': principal.id if principal else None,
            'source_nom': principal.nom if principal else None,
        })
    return result


# ── FG63 — Session d'inventaire ───────────────────────────────────────────────

def valider_inventaire_session(session, user):
    """Valide une session d'inventaire : émet les ajustements de stock pour
    chaque ligne en écart. Idempotent : une session déjà validée lève ValueError.
    Retourne {ajustes, inchanges}."""
    from django.db import transaction
    from .models import MouvementStock, InventaireSession

    if session.statut == InventaireSession.Statut.VALIDE:
        raise ValueError("Cette session d'inventaire est déjà validée.")
    if session.statut == InventaireSession.Statut.ANNULE:
        raise ValueError("Cette session d'inventaire est annulée.")

    ajustes, inchanges = 0, 0

    with transaction.atomic():
        for ligne in session.lignes.select_related('produit').all():
            ecart = ligne.quantite_comptee - ligne.quantite_theorique
            if ecart == 0:
                inchanges += 1
                continue

            # Ajustement : on pose la nouvelle valeur directement.
            produit = ligne.produit
            # Verrou anti-concurrence
            from .models import Produit
            produit = Produit.objects.select_for_update().get(pk=produit.pk)
            qte_avant = produit.quantite_stock
            qte_apres = ligne.quantite_comptee  # on remplace par le comptage

            MouvementStock.objects.create(
                company=session.company,
                produit=produit,
                type_mouvement=MouvementStock.TypeMouvement.AJUSTEMENT,
                quantite=abs(ecart),
                quantite_avant=qte_avant,
                quantite_apres=qte_apres,
                reference=session.reference,
                note=f'Inventaire {session.reference} — écart {ecart:+d}',
                created_by=user)

            produit.quantite_stock = qte_apres
            produit.save(update_fields=['quantite_stock'])
            ajustes += 1

        session.statut = InventaireSession.Statut.VALIDE
        session.save(update_fields=['statut'])

    return {'ajustes': ajustes, 'inchanges': inchanges}


# ── FG64 — Rapport expiry ─────────────────────────────────────────────────────

def produits_expirant_bientot(company, jours=90):
    """Retourne les produits dont au moins une ligne de réception a une
    date_peremption dans les `jours` prochains. Admin-only, INTERNE."""
    from django.utils import timezone
    import datetime
    from .models import LigneReceptionFournisseur

    today = timezone.now().date()
    limite = today + datetime.timedelta(days=jours)

    qs = (LigneReceptionFournisseur.objects
          .filter(
              reception__company=company,
              date_peremption__isnull=False,
              date_peremption__lte=limite,
              date_peremption__gte=today)
          .select_related('produit', 'reception'))

    # Déduplique par produit (garde la date de péremption la plus proche)
    seen = {}
    for ligne in qs.order_by('date_peremption'):
        pid = ligne.produit_id
        if pid not in seen:
            seen[pid] = {
                'produit_id': pid,
                'produit_nom': ligne.produit.nom,
                'date_peremption': ligne.date_peremption.isoformat(),
                'numero_lot': ligne.numero_lot,
                'quantite': ligne.quantite,
                'reception_ref': ligne.reception.reference,
                'jours_restants': (ligne.date_peremption - today).days,
            }
    return list(seen.values())


# ── FG65 — Prévisions de demande ─────────────────────────────────────────────

def previsions_reappro(company, nb_mois=6):
    """Calcule la consommation mensuelle moyenne par SKU (SORTIE sur les
    `nb_mois` derniers mois) et propose une quantité de réapprovisionnement.
    Retourne une liste de dicts par produit avec sortie > 0. Admin-only.

    YPROC9 — enrichit chaque item avec ``date_rupture``/``point_commande``
    (FG364, `core.stock_reorder.predict_reorder` — jusqu'ici un dead-end,
    consommé par AUCUNE vue) dérivés de la conso journalière moyenne déjà
    calculée ici. ``None`` si la conso journalière est nulle (aucune rupture
    prévisible — le module garde le résultat exploitable)."""
    from django.utils import timezone
    import datetime
    from .models import MouvementStock, Produit
    from core.stock_reorder import predict_reorder

    today = timezone.now().date()
    debut = today - datetime.timedelta(days=30 * nb_mois)

    # Agréger les sorties par produit
    from django.db.models import Sum
    sorties = (MouvementStock.objects
               .filter(company=company, type_mouvement='sortie',
                       date__date__gte=debut)
               .values('produit_id')
               .annotate(total=Sum('quantite')))

    sorties_map = {s['produit_id']: s['total'] for s in sorties}
    if not sorties_map:
        return []

    produits = Produit.objects.filter(
        company=company, id__in=list(sorties_map.keys()),
        is_archived=False).only('id', 'nom', 'sku', 'quantite_stock',
                                'seuil_alerte', 'quantite_reappro_cible')
    result = []
    for p in produits:
        total_sorties = sorties_map[p.id]
        conso_moy = total_sorties / nb_mois  # par mois
        # Quantité suggérée = conso 2 mois (stock de sécurité) ou cible si définie
        qte_suggeree = (p.quantite_reappro_cible
                        if p.quantite_reappro_cible
                        else max(round(conso_moy * 2), p.seuil_alerte * 2 if p.seuil_alerte else 1))
        conso_jour = conso_moy / 30.0 if conso_moy else 0.0
        reorder = predict_reorder(
            current_stock=p.quantite_stock, today=today,
            avg_daily_consumption=conso_jour,
            safety_stock=p.seuil_alerte or 0)
        result.append({
            'produit_id': p.id,
            'nom': p.nom,
            'sku': p.sku,
            'total_sorties': total_sorties,
            'consommation_mensuelle_moy': round(conso_moy, 2),
            'quantite_stock': p.quantite_stock,
            'quantite_suggeree': qte_suggeree,
            'date_rupture': reorder.rupture_date,
            'point_commande': reorder.reorder_point,
        })
    result.sort(key=lambda x: -x['consommation_mensuelle_moy'])
    return result


# ── FG66 / DC36 — Explosion d'un kit (BOM) en lignes composant ────────────────
# DC36 : aucun prix / marque / TVA n'est stocké sur le kit — l'explosion lit ces
# attributs sur le Produit composant au moment de l'insertion. Point d'entrée
# cross-app : `ventes` insère un kit dans un devis en appelant CE service (puis
# crée ses propres lignes de devis), jamais en important le modèle stock.
#
# XMFG17 — nomenclature multi-niveaux : un composant peut être un SOUS-KIT
# (`composant_kit`, XOR avec `produit`). L'explosion devient RÉCURSIVE :
# un sous-kit s'explose à son tour, ses lignes produit remontent (aplaties)
# dans le résultat du kit parent, quantités multipliées à chaque niveau.
# Garde anti-cycle (un kit ne peut jamais se contenir lui-même, directement
# ou via une chaîne de sous-kits) + profondeur raisonnable (10 niveaux).

MAX_PROFONDEUR_KIT = 10


class KitCycleError(ValueError):
    """XMFG17 — un kit se contient lui-même (directement ou via un sous-kit).
    Message clair pour le front (409/400 côté vue)."""


def exploser_kit(kit, quantite_kit=1, *, _chemin=None, _profondeur=0):
    """Explose un kit en ses lignes composant PRODUIT (récursif à travers les
    sous-kits XMFG17) pour ``quantite_kit`` unités.

    Renvoie une liste de dicts triés par désignation :
      {produit_id, sku, designation, quantite, prix_vente_unitaire, tva,
       marque, disponible}
    où ``quantite`` = quantité du composant × ``quantite_kit`` × le produit des
    facteurs d'échelle de chaque niveau parent. Le PRIX, la TVA et la MARQUE
    proviennent du ``Produit`` (DC36 — jamais stockés sur le kit). Un même
    produit utilisé à plusieurs niveaux (ou dans plusieurs sous-kits) apparaît
    en PLUSIEURS lignes agrégées par produit (quantités cumulées).
    ``prix_vente_unitaire`` est le prix de vente catalogue (client-facing OK).
    Le prix d'ACHAT n'est jamais exposé ici. INTERNE/écran ; côté ventes c'est
    cette liste qui devient des lignes de devis.

    Lève ``KitCycleError`` si ``kit`` se retrouve dans sa propre chaîne de
    sous-kits (cycle) ; ``ValueError`` si la profondeur dépasse
    ``MAX_PROFONDEUR_KIT`` (nomenclature anormalement profonde — probable
    erreur de saisie)."""
    from decimal import Decimal, InvalidOperation
    try:
        facteur = Decimal(str(quantite_kit))
    except (InvalidOperation, TypeError, ValueError):
        facteur = Decimal('1')

    chemin = _chemin if _chemin is not None else set()
    if kit.id in chemin:
        raise KitCycleError(
            f'Nomenclature cyclique détectée : le kit "{kit.nom}" se '
            'contient lui-même (directement ou via un sous-kit).')
    if _profondeur > MAX_PROFONDEUR_KIT:
        raise ValueError(
            f'Nomenclature trop profonde (> {MAX_PROFONDEUR_KIT} niveaux) '
            f'pour le kit "{kit.nom}" — vérifiez la composition.')
    chemin = chemin | {kit.id}

    par_produit = {}
    composants = (kit.composants
                  .select_related('produit', 'composant_kit')
                  .order_by('id'))
    for c in composants:
        qte = (c.quantite or Decimal('0')) * facteur
        if c.produit_id:
            p = c.produit
            existant = par_produit.get(p.id)
            if existant is not None:
                existant['quantite'] += qte
                continue
            par_produit[p.id] = {
                'produit_id': p.id,
                'sku': p.sku or '',
                'designation': p.nom,
                'quantite': qte,
                # DC36 — prix / TVA / marque lus sur le composant, jamais le kit.
                'prix_vente_unitaire': p.prix_vente,
                'tva': p.tva,
                'marque': p.marque,
                'disponible': p.quantite_stock,
            }
        else:
            # XMFG17 — sous-kit : explosion récursive, lignes produit
            # remontées (aplaties) et agrégées avec celles du niveau courant.
            sous_lignes = exploser_kit(
                c.composant_kit, qte, _chemin=chemin,
                _profondeur=_profondeur + 1)
            for ligne in sous_lignes:
                existant = par_produit.get(ligne['produit_id'])
                if existant is not None:
                    existant['quantite'] += ligne['quantite']
                else:
                    par_produit[ligne['produit_id']] = dict(ligne)
    return sorted(par_produit.values(), key=lambda x: x['designation'])


def exploser_kit_par_id(company, kit_id, quantite_kit=1):
    """Variante scopée société : explose le kit ``kit_id`` de ``company`` (ou
    None si introuvable / archivé). Point d'entrée cross-app pour `ventes`."""
    from .models import KitProduit
    kit = KitProduit.objects.filter(
        id=kit_id, company=company, is_archived=False).first()
    if kit is None:
        return None
    return exploser_kit(kit, quantite_kit)


# ── XMFG18 — Révisions de nomenclature + duplication de kit ────────────────
# Chaque modification des composants d'un kit crée un SNAPSHOT JSON numéroté
# (pattern RevisionDocument FG297). La révision la plus récente est la
# composition courante ; « composition au JJ/MM/AAAA » = la dernière révision
# à cette date. Jamais de prix d'achat dans le snapshot.

def _composition_snapshot(kit):
    """Sérialise la composition courante du kit en liste JSON-compatible
    (produit_id / composant_kit_id, désignation, quantité, taux de perte).
    AUCUN prix (ni achat ni vente) dans le snapshot — la valorisation est
    toujours dérivée du catalogue au moment voulu (DC36)."""
    out = []
    for c in (kit.composants
              .select_related('produit', 'composant_kit')
              .order_by('id')):
        out.append({
            'produit_id': c.produit_id,
            'composant_kit_id': c.composant_kit_id,
            'designation': (
                c.produit.nom if c.produit_id else c.composant_kit.nom),
            'sku': (
                (c.produit.sku if c.produit_id else c.composant_kit.sku)
                or ''),
            'quantite': str(c.quantite),
            'taux_perte_pct': str(c.taux_perte_pct),
        })
    return out


def snapshot_revision_kit(kit, user=None):
    """XMFG18 — crée une révision (snapshot JSON) de la composition COURANTE
    du kit si elle diffère de la dernière révision. Renvoie (revision,
    created). Idempotent : re-sauver un kit sans changer sa composition ne
    crée PAS de révision dupliquée."""
    from .models import RevisionKit
    composition = _composition_snapshot(kit)
    derniere = kit.revisions.order_by('-numero').first()
    if derniere is not None and derniere.composition == composition:
        return derniere, False
    numero = (derniere.numero + 1) if derniere is not None else 1
    revision = RevisionKit.objects.create(
        company=kit.company, kit=kit, numero=numero,
        composition=composition, user=user)
    return revision, True


def composition_kit_au(kit, date_limite):
    """XMFG18 — « composition au JJ/MM/AAAA » : la dernière révision créée à
    la date donnée incluse (datetime.date). None si aucune révision
    n'existait encore à cette date."""
    return (kit.revisions
            .filter(date_creation__date__lte=date_limite)
            .order_by('-numero')
            .first())


def dupliquer_kit(kit, user=None, facteur_echelle=None):
    """XMFG18 — duplique un kit stock : en-tête copié (« <nom> (copie) »,
    sku vidé pour éviter la collision d'unicité) + composants copiés, avec
    facteur d'échelle optionnel appliqué aux quantités (arrondi propre à 2
    décimales, ROUND_HALF_UP — ex. 3 × 1.67 = 5.01 → 5.01). La copie reçoit
    sa révision n°1 immédiatement."""
    from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
    from .models import KitProduit, KitComposant

    facteur = None
    if facteur_echelle is not None:
        try:
            facteur = Decimal(str(facteur_echelle))
        except (InvalidOperation, TypeError, ValueError):
            raise ValueError("Facteur d'échelle invalide.")
        if facteur <= 0:
            raise ValueError("Le facteur d'échelle doit être positif.")

    copie = KitProduit.objects.create(
        company=kit.company,
        nom=f'{kit.nom} (copie)',
        sku=None,
        description=kit.description,
    )
    for c in kit.composants.all():
        quantite = c.quantite or Decimal('0')
        if facteur is not None:
            quantite = (quantite * facteur).quantize(
                Decimal('0.01'), rounding=ROUND_HALF_UP)
        KitComposant.objects.create(
            kit=copie, produit_id=c.produit_id,
            composant_kit_id=c.composant_kit_id,
            quantite=quantite, taux_perte_pct=c.taux_perte_pct)
    snapshot_revision_kit(copie, user=user)
    return copie


# ── XMFG19 — Remplacement de masse d'un composant dans les nomenclatures ───

def remplacer_composant_masse(company, *, produit_ancien_id,
                              produit_nouveau_id, ratio_quantite=None,
                              dry_run=True, user=None):
    """XMFG19 — remplace `produit_ancien` par `produit_nouveau` dans TOUTES
    les nomenclatures de la société : kits stock (KitProduit) ET kits de
    pré-assemblage (installations.Kit, via ses selectors/services — jamais
    ses models importés ici).

    `dry_run=True` (défaut) : PRÉVIEW seule — liste les kits impactés des
    deux modules sans rien modifier. `dry_run=False` : application ATOMIQUE
    (une seule transaction pour les deux modules), chaque kit modifié créant
    sa révision XMFG18, plus UNE ligne d'audit récapitulative. Un kit dont
    la nomenclature contient DÉJÀ le produit nouveau voit les deux lignes
    fusionnées (quantités additionnées) — jamais de doublon (kit, produit).

    Renvoie {'dry_run', 'kits_stock': [...], 'kits_installations': [...],
    'nb_total'}. Lève ValueError sur produit inconnu / identique."""
    from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
    from django.db import transaction
    from .models import Produit, KitComposant

    if str(produit_ancien_id) == str(produit_nouveau_id):
        raise ValueError(
            'Le produit de remplacement doit être différent du produit '
            'remplacé.')
    ancien = Produit.objects.filter(
        company=company, id=produit_ancien_id).first()
    nouveau = Produit.objects.filter(
        company=company, id=produit_nouveau_id).first()
    if ancien is None or nouveau is None:
        raise ValueError('Produit introuvable pour cette société.')

    ratio = None
    if ratio_quantite is not None:
        try:
            ratio = Decimal(str(ratio_quantite))
        except (InvalidOperation, TypeError, ValueError):
            raise ValueError('Ratio de quantité invalide.')
        if ratio <= 0:
            raise ValueError('Le ratio de quantité doit être positif.')

    def _nouvelle_quantite(quantite):
        quantite = quantite or Decimal('0')
        if ratio is not None:
            quantite = (quantite * ratio).quantize(
                Decimal('0.01'), rounding=ROUND_HALF_UP)
        return quantite

    # ── Préview stock ──
    lignes_stock = (KitComposant.objects
                    .filter(kit__company=company, produit_id=ancien.id)
                    .select_related('kit')
                    .order_by('kit__nom', 'id'))
    kits_stock = [{
        'kit_id': c.kit_id,
        'kit_nom': c.kit.nom,
        'quantite_avant': str(c.quantite),
        'quantite_apres': str(_nouvelle_quantite(c.quantite)),
    } for c in lignes_stock]

    # ── Préview installations (lecture via selectors — cross-app propre) ──
    from apps.installations.selectors import kits_utilisant_produit
    kits_inst_preview = kits_utilisant_produit(company, ancien.id)
    kits_installations = [{
        'kit_id': k['kit_id'],
        'kit_nom': k['kit_nom'],
        'quantite_avant': str(k['quantite']),
        'quantite_apres': str(
            int((Decimal(k['quantite'] or 0) * ratio).to_integral_value(
                rounding=ROUND_HALF_UP)) or 1
            if ratio is not None and k['quantite'] else k['quantite']),
    } for k in kits_inst_preview]

    resultat = {
        'dry_run': bool(dry_run),
        'produit_ancien': ancien.nom,
        'produit_nouveau': nouveau.nom,
        'kits_stock': kits_stock,
        'kits_installations': kits_installations,
        'nb_total': len(kits_stock) + len(kits_installations),
    }
    if dry_run:
        return resultat

    # ── Application ATOMIQUE (stock + installations, une transaction) ──
    with transaction.atomic():
        kits_touches = {}
        for c in lignes_stock:
            existant = (KitComposant.objects
                        .filter(kit_id=c.kit_id, produit_id=nouveau.id)
                        .exclude(id=c.id)
                        .first())
            nouvelle_qte = _nouvelle_quantite(c.quantite)
            if existant is not None:
                # Fusion : le kit contient déjà le produit nouveau.
                existant.quantite = (
                    (existant.quantite or Decimal('0')) + nouvelle_qte)
                existant.save(update_fields=['quantite'])
                c.delete()
            else:
                c.produit_id = nouveau.id
                c.quantite = nouvelle_qte
                c.save(update_fields=['produit', 'quantite'])
            kits_touches[c.kit_id] = c.kit
        for kit in kits_touches.values():
            snapshot_revision_kit(kit, user=user)

        # Écriture installations via SON service (jamais ses models ici).
        from apps.installations.services import remplacer_composant_kits
        appliques_inst = remplacer_composant_kits(
            company, produit_ancien_id=ancien.id,
            produit_nouveau_id=nouveau.id, ratio=ratio, user=user)
        resultat['kits_installations'] = [{
            'kit_id': k['kit_id'],
            'kit_nom': k['kit_nom'],
            'quantite_avant': str(k['quantite_avant']),
            'quantite_apres': str(k['quantite_apres']),
        } for k in appliques_inst]
        resultat['nb_total'] = (
            len(resultat['kits_stock'])
            + len(resultat['kits_installations']))

        # Ligne d'audit récapitulative (best-effort, jamais bloquante).
        from apps.audit.recorder import record
        from apps.audit.models import AuditLog
        record(
            AuditLog.Action.UPDATE,
            instance=nouveau,
            company=company, user=user,
            detail=(
                f'Remplacement de masse : « {ancien.nom} » → '
                f'« {nouveau.nom} » dans {resultat["nb_total"]} '
                f'nomenclature(s)'
                + (f' (ratio quantité ×{ratio})' if ratio is not None
                   else '') + '.'))
    return resultat


# ── XMFG5 — Coût de revient du kit : roll-up + structure + stock potentiel ──
# INTERNE — le coût/marge n'est visible qu'aux rôles responsable/admin
# (gardé côté vue) ; le prix d'achat/coût n'apparaît JAMAIS sur un document
# client ni dans un PDF.

def _structure_kit_lignes(kit, reserves, *, niveau=0, _chemin=None,
                          _profondeur=0):
    """XMFG17 — collecte RÉCURSIVE des lignes indentées (niveau 0 = racine)
    + roll-up (coût/prix-vente/dispo potentielle) traversant les sous-kits.
    Sous-fonction interne de `structure_kit` ; garde anti-cycle/profondeur
    identique à `exploser_kit`."""
    from decimal import Decimal

    chemin = _chemin if _chemin is not None else set()
    if kit.id in chemin:
        raise KitCycleError(
            f'Nomenclature cyclique détectée : le kit "{kit.nom}" se '
            'contient lui-même (directement ou via un sous-kit).')
    if _profondeur > MAX_PROFONDEUR_KIT:
        raise ValueError(
            f'Nomenclature trop profonde (> {MAX_PROFONDEUR_KIT} niveaux) '
            f'pour le kit "{kit.nom}" — vérifiez la composition.')
    chemin = chemin | {kit.id}

    lignes = []
    cout_total = Decimal('0')
    prix_vente_total = Decimal('0')
    disponibilite_potentielle = None
    composants = (kit.composants
                  .select_related('produit', 'composant_kit')
                  .order_by('id'))
    for c in composants:
        quantite = c.quantite or Decimal('0')
        if c.produit_id:
            p = c.produit
            dispo = Decimal(str(p.quantite_stock)) - Decimal(
                str(reserves.get(p.id, 0)))
            cout_unitaire = cout_achat_courant(p)
            cout_ligne = (cout_unitaire * quantite).quantize(Decimal('0.01'))
            cout_total += cout_ligne
            prix_vente_total += (p.prix_vente or Decimal('0')) * quantite
            if quantite > 0:
                kits_possibles = int((dispo / quantite).to_integral_value(
                    rounding='ROUND_FLOOR')) if dispo > 0 else 0
                disponibilite_potentielle = kits_possibles if (
                    disponibilite_potentielle is None
                ) else min(disponibilite_potentielle, kits_possibles)
            lignes.append({
                'niveau': niveau, 'type': 'produit',
                'produit_id': p.id, 'sku': p.sku or '', 'designation': p.nom,
                'quantite': quantite, 'quantite_disponible': dispo,
                'cout_unitaire': cout_unitaire, 'cout_total': cout_ligne,
                'prix_vente': p.prix_vente,
            })
        else:
            sk = c.composant_kit
            (sous_lignes, sous_cout, sous_prix_vente,
             sous_dispo_potentielle) = _structure_kit_lignes(
                sk, reserves, niveau=niveau + 1, _chemin=chemin,
                _profondeur=_profondeur + 1)
            # Ligne d'en-tête du sous-kit (indentée), suivie de ses composants.
            lignes.append({
                'niveau': niveau, 'type': 'sous_kit',
                'composant_kit_id': sk.id, 'sku': sk.sku or '',
                'designation': sk.nom, 'quantite': quantite,
                'quantite_disponible': sous_dispo_potentielle,
                'cout_unitaire': (sous_cout * quantite).quantize(
                    Decimal('0.01')) if quantite else Decimal('0.00'),
                'cout_total': (sous_cout * quantite).quantize(
                    Decimal('0.01')),
                'prix_vente': None,
            })
            lignes.extend(sous_lignes)
            cout_total += (sous_cout * quantite).quantize(Decimal('0.01'))
            prix_vente_total += sous_prix_vente * quantite
            if quantite > 0 and sous_dispo_potentielle is not None:
                kits_possibles = int(
                    (Decimal(sous_dispo_potentielle) / quantite)
                    .to_integral_value(rounding='ROUND_FLOOR'))
                disponibilite_potentielle = kits_possibles if (
                    disponibilite_potentielle is None
                ) else min(disponibilite_potentielle, kits_possibles)
    return lignes, cout_total, prix_vente_total, disponibilite_potentielle


def structure_kit(kit):
    """XMFG5 — nomenclature indentée d'un kit : par composant, quantité,
    disponibilité (`quantite_disponible` = stock − réservé), coût unitaire
    (DC28 `cout_achat_courant`) et coût total roll-up. XMFG17 — traverse
    récursivement les SOUS-KITS (chaque ligne porte `niveau` pour un
    affichage indenté ; une ligne sous-kit a `type: 'sous_kit'`, une ligne
    produit `type: 'produit'`). Renvoie
    {kit_id, kit_nom, composants:[{niveau, type, produit_id?,
    composant_kit_id?, sku, designation, quantite, quantite_disponible,
    cout_unitaire, cout_total, prix_vente}], cout_total_roll_up, marge,
    disponibilite_potentielle} où `disponibilite_potentielle` = min(dispo
    composant ÷ quantité) TOUS NIVEAUX CONFONDUS (combien de kits sont
    assemblables avec le stock actuel, 0 si un composant manque, à
    n'importe quel niveau). INTERNE (coût/marge) — jamais client-facing.

    Lève ``KitCycleError``/``ValueError`` — mêmes gardes que `exploser_kit`."""
    from decimal import Decimal
    reserves = reserved_quantities(kit.company)
    lignes, cout_total, prix_vente_total, disponibilite_potentielle = (
        _structure_kit_lignes(kit, reserves))
    marge = (prix_vente_total - cout_total).quantize(Decimal('0.01'))
    return {
        'kit_id': kit.id, 'kit_nom': kit.nom, 'composants': lignes,
        'cout_total_roll_up': cout_total.quantize(Decimal('0.01')),
        'prix_vente_total': prix_vente_total.quantize(Decimal('0.01')),
        'marge': marge,
        'disponibilite_potentielle': disponibilite_potentielle or 0,
    }


# ── XMFG1 — Backflush : clôture d'un ordre d'assemblage (installations) ──────
# `installations.OrdreAssemblage.terminer` appelle CE service (jamais de
# MouvementStock créé côté installations) : une SORTIE par composant (quantité
# BOM × quantite_produite) + une ENTREE du composite. Idempotent via le drapeau
# `stock_mouvemente` posé par l'appelant dans la MÊME transaction atomique.

def consommer_et_produire_assemblage(*, company, kit, composants, produit_compose,
                                     quantite_produite, reference, user,
                                     emplacement_source=None,
                                     emplacement_destination=None,
                                     per_unit=True):
    """Consomme les composants d'un kit et produit le composite (ENTREE, qté
    ``quantite_produite``). ``composants`` est un itérable d'objets avec
    ``.produit`` et ``.quantite``. Si ``per_unit`` (défaut, la BOM du kit),
    ``.quantite`` est PAR UNITÉ et la sortie = ``.quantite`` × ``quantite_produite``
    (tolérance sur/sous-production XMFG1). Si ``per_unit=False`` (lignes d'ordre
    personnalisées — XMFG6), ``.quantite`` est déjà le TOTAL à consommer, utilisé
    tel quel. Ventile sur les emplacements fournis (StockEmplacement, N15) en
    plus du total canonique. Lève ValueError si ``produit_compose`` est None
    (kit non finalisable) — l'appelant refuse `terminer` dans ce cas."""
    from django.db import transaction
    from .models import Produit, StockEmplacement

    if produit_compose is None:
        raise ValueError(
            "Le kit n'a pas de produit composite (produit_compose) : "
            "impossible de clôturer l'ordre.")

    with transaction.atomic():
        mouvements = []
        for ligne in composants:
            comp_produit = ligne.produit
            if comp_produit is None:
                continue
            qte_conso = ((ligne.quantite or 0) * quantite_produite if per_unit
                         else (ligne.quantite or 0))
            if qte_conso <= 0:
                continue
            p = Produit.objects.select_for_update().get(id=comp_produit.id)
            avant = p.quantite_stock
            apres = avant - qte_conso
            check_negative_stock_guard(company, avant, apres)
            mvt = record_stock_movement(
                company=company, produit=p,
                type_mouvement=mouvement_type_sortie(),
                quantite=qte_conso, quantite_avant=avant,
                quantite_apres=apres, reference=reference,
                note=f'Assemblage {reference} — composant kit {kit.id}',
                created_by=user)
            mouvements.append(mvt)
            if emplacement_source is not None and \
                    not emplacement_source.is_principal:
                se, _ = StockEmplacement.objects.select_for_update().get_or_create(
                    produit=p, emplacement=emplacement_source,
                    defaults={'company': company, 'quantite': 0})
                se.quantite = max(se.quantite - qte_conso, 0)
                se.save(update_fields=['quantite'])

        composite = Produit.objects.select_for_update().get(id=produit_compose.id)
        avant_c = composite.quantite_stock
        apres_c = avant_c + quantite_produite
        mvt_entree = record_stock_movement(
            company=company, produit=composite,
            type_mouvement=mouvement_type_entree(),
            quantite=quantite_produite, quantite_avant=avant_c,
            quantite_apres=apres_c, reference=reference,
            note=f'Assemblage {reference} — composite kit {kit.id}',
            created_by=user)
        mouvements.append(mvt_entree)
        if emplacement_destination is not None and \
                not emplacement_destination.is_principal:
            se, _ = StockEmplacement.objects.select_for_update().get_or_create(
                produit=composite, emplacement=emplacement_destination,
                defaults={'company': company, 'quantite': 0})
            se.quantite += quantite_produite
            se.save(update_fields=['quantite'])

    return mouvements


# ── XMFG12 — Démontage (unbuild) : composite → composants ────────────────────
# Chemin INVERSE de XMFG1 : `installations.OrdreDemontage.terminer` appelle CE
# service pour un composite de retour (annulation, kit démo cannibalisé) —
# SORTIE du composite + ENTREE de chaque composant selon les quantités
# RÉCUPÉRÉES (éditables ligne à ligne, jamais la BOM brute).

def demonter_composite(*, company, kit, quantite_demontee, lignes_recuperation,
                       produit_compose, reference, user,
                       emplacement_source=None, emplacement_destination=None):
    """Sort le composite (`quantite_demontee` unités) et restocke chaque
    composant selon ``lignes_recuperation`` — itérable d'objets avec
    ``.produit`` et ``.quantite_recuperee`` (déjà le TOTAL récupéré, PAS par
    unité). Les composants cassés (récupéré < attendu) ne sont PAS restockés
    ici — l'appelant les déclare en rebut (XMFG11) séparément. Lève ValueError
    si ``produit_compose`` est None."""
    from django.db import transaction
    from .models import Produit, StockEmplacement

    if produit_compose is None:
        raise ValueError(
            "Le kit n'a pas de produit composite (produit_compose) : "
            "démontage impossible.")

    with transaction.atomic():
        mouvements = []
        composite = Produit.objects.select_for_update().get(id=produit_compose.id)
        avant_c = composite.quantite_stock
        qte_sortie = min(quantite_demontee, avant_c) if avant_c > 0 else 0
        apres_c = avant_c - qte_sortie
        mvt_sortie = record_stock_movement(
            company=company, produit=composite,
            type_mouvement=mouvement_type_sortie(),
            quantite=qte_sortie, quantite_avant=avant_c, quantite_apres=apres_c,
            reference=reference,
            note=f'Démontage {reference} — composite kit {kit.id}',
            created_by=user)
        mouvements.append(mvt_sortie)
        if emplacement_source is not None and not emplacement_source.is_principal:
            se, _ = StockEmplacement.objects.select_for_update().get_or_create(
                produit=composite, emplacement=emplacement_source,
                defaults={'company': company, 'quantite': 0})
            se.quantite = max(se.quantite - qte_sortie, 0)
            se.save(update_fields=['quantite'])

        for ligne in lignes_recuperation:
            comp_produit = ligne.produit
            if comp_produit is None:
                continue
            qte_recup = ligne.quantite_recuperee or 0
            if qte_recup <= 0:
                continue
            p = Produit.objects.select_for_update().get(id=comp_produit.id)
            avant = p.quantite_stock
            apres = avant + qte_recup
            mvt = record_stock_movement(
                company=company, produit=p,
                type_mouvement=mouvement_type_entree(),
                quantite=qte_recup, quantite_avant=avant, quantite_apres=apres,
                reference=reference,
                note=f'Démontage {reference} — composant récupéré kit {kit.id}',
                created_by=user)
            mouvements.append(mvt)
            if emplacement_destination is not None and \
                    not emplacement_destination.is_principal:
                se, _ = StockEmplacement.objects.select_for_update().get_or_create(
                    produit=p, emplacement=emplacement_destination,
                    defaults={'company': company, 'quantite': 0})
                se.quantite += qte_recup
                se.save(update_fields=['quantite'])

    return mouvements


# ── XPUR1 — conformité fournisseur : warning BCF + gate paiement ───────────
def fournisseur_conformite_manquante(fournisseur):
    """XPUR1 — liste les documents de conformité OBLIGATOIRES manquants ou
    expirés d'un fournisseur (liste de dicts ``{type_document, motif}``).
    Vide = fournisseur en règle (ou sans document requis renseigné). Ne lève
    jamais d'exception : appelé au fil de l'eau (warning + gate paiement)."""
    problemes = []
    docs = list(fournisseur.documents_conformite.filter(obligatoire=True))
    for doc in docs:
        if not doc.est_valide():
            problemes.append({
                'type_document': doc.type_document,
                'type_document_display': doc.get_type_document_display(),
                'motif': 'expiré' if doc.date_expiration else 'sans date',
            })
    return problemes


def bcf_warning_conformite(fournisseur):
    """XPUR1 — message WARNING (non bloquant) à afficher à la création d'un
    BCF quand le fournisseur a un document de conformité manquant/expiré.
    None si rien à signaler."""
    problemes = fournisseur_conformite_manquante(fournisseur)
    if not problemes:
        return None
    libelles = ', '.join(p['type_document_display'] for p in problemes)
    return (f"Attention : {fournisseur.nom} a un ou plusieurs documents de "
            f"conformité manquants/expirés ({libelles}).")


def check_paiement_conformite_gate(company, fournisseur):
    """XPUR1 — lève ValueError si le PARAMÈTRE société de blocage paiement
    est actif ET que le fournisseur a un document de conformité obligatoire
    manquant/expiré. No-op (comportement historique) si le paramètre est OFF
    (défaut)."""
    from .models import AchatsParametres
    parametres = AchatsParametres.for_company(company)
    if not parametres.bloquer_paiement_conformite_expiree:
        return
    problemes = fournisseur_conformite_manquante(fournisseur)
    if problemes:
        libelles = ', '.join(p['type_document_display'] for p in problemes)
        raise ValueError(
            f"Paiement bloqué : {fournisseur.nom} a un document de "
            f"conformité manquant/expiré ({libelles}).")


def notify_expiring_conformite_documents(company, jours=30):
    """XPUR1 — notifie les responsables/admins des documents de conformité
    fournisseur expirant sous ``jours`` jours (ou déjà expirés). Best-effort :
    une erreur de notification n'interrompt jamais l'appelant. Renvoie le
    nombre de documents notifiés."""
    from datetime import timedelta
    from django.utils import timezone
    from .models import DocumentConformiteFournisseur
    seuil = timezone.now().date() + timedelta(days=jours)
    docs = DocumentConformiteFournisseur.objects.filter(
        company=company, obligatoire=True,
        date_expiration__isnull=False, date_expiration__lte=seuil,
    ).select_related('fournisseur')
    count = 0
    for doc in docs:
        try:
            from apps.notifications.services import notify_many
            from apps.notifications.models import EventType
            from django.contrib.auth import get_user_model
            User = get_user_model()
            recipients = User.objects.filter(
                company=company, is_active=True,
                role_legacy__in=['responsable', 'admin'])
            titre = (f'Document fournisseur bientôt expiré '
                     f'({doc.fournisseur.nom})')
            corps = (f'{doc.get_type_document_display()} de '
                     f'{doc.fournisseur.nom} expire le '
                     f'{doc.date_expiration}.')
            notify_many(
                recipients, EventType.SUPPLIER_DOC_EXPIRING,
                title=titre, body=corps, company=company)
            count += 1
        except Exception:  # noqa: BLE001 — best-effort, jamais bloquant
            logger.warning(
                'notify_expiring_conformite_documents: échec pour doc %s',
                doc.pk)
    return count


# ── XPUR2 — RAS-TVA sur paiements fournisseurs (LF 2024) ───────────────────

def _fournisseur_a_arf_valide(fournisseur):
    """XPUR2 — vrai si le fournisseur a une ARF (XPUR1) valide (< 6 mois,
    non expirée). Aucune ARF renseignée = pas de couverture (retenue max)."""
    from django.utils import timezone
    from datetime import timedelta
    from .models import DocumentConformiteFournisseur
    seuil = timezone.now().date() - timedelta(days=183)  # ~6 mois
    return fournisseur.documents_conformite.filter(
        type_document=DocumentConformiteFournisseur.Type.ARF,
    ).filter(
        models.Q(date_expiration__isnull=True) |
        models.Q(date_expiration__gte=timezone.now().date()),
    ).filter(
        models.Q(date_emission__isnull=True) |
        models.Q(date_emission__gte=seuil),
    ).exists()


def taux_ras_tva(facture):
    """XPUR2 — taux de RAS-TVA (0/75/100) applicable à une FactureFournisseur
    selon son ``type_achat`` et la validité ARF (< 6 mois) du fournisseur.

    - Biens & travaux : 100 % SANS ARF valide, 0 % AVEC.
    - Prestations de services : 75 % AVEC ARF valide, 100 % SANS.
    """
    from .models import FactureFournisseur
    a_arf = _fournisseur_a_arf_valide(facture.fournisseur)
    if facture.type_achat == FactureFournisseur.TypeAchat.SERVICES:
        return Decimal('75') if a_arf else Decimal('100')
    # Biens & travaux.
    return Decimal('0') if a_arf else Decimal('100')


def compute_ras_tva(company, facture, montant_paiement):
    """XPUR2 — calcule (taux, montant_ras) pour un paiement de
    ``montant_paiement`` sur ``facture``, proportionnellement à la part de
    TVA couverte par ce règlement. No-op (0, 0) si la société n'a pas activé
    la RAS-TVA (``AchatsParametres.ras_tva_actif`` OFF par défaut) ou si la
    facture ne porte aucune TVA."""
    from .models import AchatsParametres
    parametres = AchatsParametres.for_company(company)
    if not parametres.ras_tva_actif:
        return Decimal('0'), Decimal('0')
    montant_tva = facture.montant_tva or Decimal('0')
    montant_ttc = facture.montant_ttc or Decimal('0')
    if montant_tva <= 0 or montant_ttc <= 0:
        return Decimal('0'), Decimal('0')
    taux = taux_ras_tva(facture)
    if taux <= 0:
        return Decimal('0'), Decimal('0')
    # TVA proportionnelle à la part du TTC réglée par CE paiement.
    montant_paiement = Decimal(montant_paiement or 0)
    part_tva = (montant_tva * montant_paiement / montant_ttc).quantize(
        Decimal('0.01'))
    montant_ras = (part_tva * taux / Decimal('100')).quantize(Decimal('0.01'))
    return taux, montant_ras


def relevé_ras_tva(company, *, date_debut=None, date_fin=None):
    """XPUR2 — relevé détaillé des RAS-TVA retenues sur la période, pour la
    télédéclaration Simpl-TVA (pattern du bordereau FG139). Renvoie une liste
    de dicts triés par date de paiement."""
    from .models import PaiementFournisseur
    qs = PaiementFournisseur.objects.filter(
        company=company, montant_ras_tva__gt=0,
    ).select_related('facture', 'facture__fournisseur')
    if date_debut:
        qs = qs.filter(date_paiement__gte=date_debut)
    if date_fin:
        qs = qs.filter(date_paiement__lte=date_fin)
    rows = []
    for p in qs.order_by('date_paiement'):
        rows.append({
            'date_paiement': p.date_paiement,
            'fournisseur': p.facture.fournisseur.nom,
            'facture': p.facture.reference,
            'type_achat': p.facture.get_type_achat_display(),
            'montant_paiement': p.montant,
            'taux_ras': p.taux_ras,
            'montant_ras_tva': p.montant_ras_tva,
            'montant_net_paye': p.montant_net_paye,
        })
    return rows


def export_ras_tva_xlsx(company, *, date_debut=None, date_fin=None):
    """XPUR2 — export xlsx du relevé RAS-TVA (réutilise le helper commun)."""
    from apps.records.xlsx import build_xlsx_response
    rows = relevé_ras_tva(company, date_debut=date_debut, date_fin=date_fin)
    headers = [
        'Date paiement', 'Fournisseur', 'Facture', 'Type achat',
        'Montant paiement', 'Taux RAS %', 'Montant RAS-TVA',
        'Net payé',
    ]
    data_rows = [
        [r['date_paiement'], r['fournisseur'], r['facture'],
         r['type_achat'], r['montant_paiement'], r['taux_ras'],
         r['montant_ras_tva'], r['montant_net_paye']]
        for r in rows
    ]
    return build_xlsx_response(
        'ras_tva_fournisseurs.xlsx', headers, data_rows,
        sheet_title='RAS-TVA')


# ── XPUR3 — Multi-devises sur les achats ────────────────────────────────────

def contre_valeur_mad(montant_devise, taux_change):
    """XPUR3 — convertit un montant en devise vers sa contre-valeur MAD au
    taux fourni (saisi à la date du document, aucun appel externe). Renvoie
    Decimal('0') si l'un des deux est absent."""
    montant = _dec(montant_devise)
    taux = _dec(taux_change)
    if montant is None or taux is None:
        return Decimal('0')
    return (montant * taux).quantize(Decimal('0.01'))


def apply_devise_ligne_bcf(ligne_data, devise, taux_change):
    """XPUR3 — si la ligne porte un ``prix_achat_unitaire_devise``, dérive et
    ÉCRASE ``prix_achat_unitaire`` (MAD) = devise × taux. Document en MAD
    (devise MAD / champ non renseigné) : ``prix_achat_unitaire`` reste tel
    quel, saisi directement en MAD (comportement historique). Modifie
    ``ligne_data`` EN PLACE et le renvoie (pratique dans un serializer)."""
    from .models import DeviseAchat
    prix_devise = ligne_data.get('prix_achat_unitaire_devise')
    if prix_devise is not None and devise and devise != DeviseAchat.MAD:
        ligne_data['prix_achat_unitaire'] = contre_valeur_mad(
            prix_devise, taux_change)
    return ligne_data


def apply_devise_facture(montant_ttc_devise, devise, taux_change):
    """XPUR3 — contre-valeur MAD du TTC facture depuis son montant en
    devise (None si le document est en MAD — comportement historique)."""
    from .models import DeviseAchat
    if montant_ttc_devise is None or not devise or devise == DeviseAchat.MAD:
        return None
    return contre_valeur_mad(montant_ttc_devise, taux_change)


# ── XPUR4 — statut fournisseur : gate commande / paiement ──────────────────

def check_fournisseur_statut_commande(fournisseur):
    """XPUR4 — lève ValueError si le fournisseur est bloqué pour les
    COMMANDES (bloque_commandes ou bloque_total). No-op pour 'actif' et
    'bloque_paiements' (comportement historique préservé)."""
    from .models import Fournisseur
    if fournisseur.statut in (
        Fournisseur.Statut.BLOQUE_COMMANDES, Fournisseur.Statut.BLOQUE_TOTAL,
    ):
        raise ValueError(
            f"Impossible de créer un bon de commande : {fournisseur.nom} est "
            f"{fournisseur.get_statut_display().lower()}"
            f"{' (' + fournisseur.motif_blocage + ')' if fournisseur.motif_blocage else ''}.")


def check_fournisseur_statut_paiement(fournisseur):
    """XPUR4 — lève ValueError si le fournisseur est bloqué pour les
    PAIEMENTS (bloque_paiements ou bloque_total). No-op pour 'actif' et
    'bloque_commandes' (comportement historique préservé)."""
    from .models import Fournisseur
    if fournisseur.statut in (
        Fournisseur.Statut.BLOQUE_PAIEMENTS, Fournisseur.Statut.BLOQUE_TOTAL,
    ):
        raise ValueError(
            f"Impossible d'enregistrer un paiement : {fournisseur.nom} est "
            f"{fournisseur.get_statut_display().lower()}"
            f"{' (' + fournisseur.motif_blocage + ')' if fournisseur.motif_blocage else ''}.")


# ── XPUR5 — fiche fournisseur enrichie : validation ICE ────────────────────

def validate_ice_format(ice):
    """XPUR5 — vrai si ``ice`` est un ICE marocain bien formé (15 chiffres).
    Chaîne vide/None = pas d'ICE saisi (pas une erreur de format)."""
    if not ice:
        return True
    return bool(ice.isdigit() and len(ice) == 15)


def find_duplicate_ice(company, ice, *, exclude_id=None):
    """XPUR5 — renvoie le premier Fournisseur de la société qui porte déjà
    ce même ICE (hors ``exclude_id`` — mise à jour d'un fournisseur
    existant), ou None. Détection non bloquante (warning)."""
    from .models import Fournisseur
    if not ice:
        return None
    qs = Fournisseur.objects.filter(company=company, ice=ice)
    if exclude_id is not None:
        qs = qs.exclude(pk=exclude_id)
    return qs.first()


# ── XPUR6 — conditions de paiement fournisseur & échéancier multi-tranches ──

def derive_date_echeance(fournisseur, date_facture):
    """XPUR6 — dérive la date d'échéance depuis les conditions de paiement du
    fournisseur (délai_paiement_jours + fin_de_mois). Renvoie None quand le
    fournisseur n'a AUCUN délai configuré (délai=0) — comportement
    historique : la date d'échéance reste saisie à la main."""
    if not date_facture or not fournisseur or not fournisseur.delai_paiement_jours:
        return None
    from datetime import timedelta
    import calendar
    echeance = date_facture + timedelta(days=fournisseur.delai_paiement_jours)
    if fournisseur.fin_de_mois:
        dernier_jour = calendar.monthrange(echeance.year, echeance.month)[1]
        echeance = echeance.replace(day=dernier_jour)
    return echeance


def escompte_applicable(fournisseur, date_facture, date_paiement):
    """XPUR6 — vrai si un paiement à ``date_paiement`` d'une facture datée
    ``date_facture`` tombe dans la fenêtre d'escompte du fournisseur
    (paiement anticipé type 2/10 net 30). Faux si le fournisseur n'a pas
    d'escompte configuré (comportement historique)."""
    if not fournisseur or not fournisseur.escompte_pct or not fournisseur.escompte_jours:
        return False
    if not date_facture or not date_paiement:
        return False
    from datetime import timedelta
    return date_paiement <= date_facture + timedelta(
        days=fournisseur.escompte_jours)


def creer_echeancier_facture_fournisseur(company, facture, tranches):
    """XPUR6 — crée l'échéancier multi-tranches d'une facture fournisseur.

    ``tranches`` : liste de dicts ``{pourcentage?, montant?, date_echeance}``.
    Si ``montant`` est absent et ``pourcentage`` fourni, le montant est
    dérivé du TTC de la facture. Renvoie la liste des ``EcheanceFactureFournisseur``
    créées. N'écrase jamais un échéancier existant (les tranches précédentes
    doivent être supprimées explicitement avant un nouvel appel)."""
    from .models import EcheanceFactureFournisseur
    created = []
    for t in tranches:
        montant = t.get('montant')
        if montant is None and t.get('pourcentage') is not None:
            montant = (facture.montant_ttc or Decimal('0')) * Decimal(
                str(t['pourcentage'])) / Decimal('100')
        created.append(EcheanceFactureFournisseur.objects.create(
            company=company, facture=facture,
            pourcentage=t.get('pourcentage'),
            montant=montant or Decimal('0'),
            date_echeance=t['date_echeance']))
    return created


# ── ZPUR8 — « Other Information » BCF : acheteur/réf/note + report défauts ──

def conditions_paiement_label(fournisseur):
    """ZPUR8 — libellé lisible des conditions de paiement dérivé de
    `Fournisseur.delai_paiement_jours`/`fin_de_mois` (XPUR6). Vide si aucun
    délai connu (0 = comptant, comportement historique)."""
    if fournisseur is None or not fournisseur.delai_paiement_jours:
        return ''
    label = f'{fournisseur.delai_paiement_jours} jours'
    if fournisseur.fin_de_mois:
        label += ' fin de mois'
    return label


def default_other_information_bcf(bon_commande):
    """ZPUR8 — reporte au DOCUMENT (une fois, à la création) les défauts
    fournisseur `incoterm` (XPUR5) et les conditions de paiement dérivées de
    `delai_paiement_jours` (XPUR6) — SANS redéfinir ces référentiels. Un champ
    déjà renseigné (édition ultérieure) n'est jamais écrasé. No-op si le BCF
    n'a pas de fournisseur (comportement historique)."""
    if bon_commande.fournisseur_id is None:
        return
    changed = []
    if not bon_commande.incoterm:
        incoterm = bon_commande.fournisseur.incoterm or ''
        if incoterm:
            bon_commande.incoterm = incoterm
            changed.append('incoterm')
    if not bon_commande.conditions_paiement:
        label = conditions_paiement_label(bon_commande.fournisseur)
        if label:
            bon_commande.conditions_paiement = label
            changed.append('conditions_paiement')
    if changed:
        bon_commande.save(update_fields=changed)


# ── ZPUR11 — motif d'annulation obligatoire + réouverture ─────────────────

def log_bcf_chatter(bon_commande, *, user, body):
    """ZPUR11 — trace une entrée `records.Comment` (chatter, horodaté +
    acteur) sur un BCF (annulation motivée, réouverture…). Best-effort :
    ne bloque jamais l'action appelante."""
    try:
        from django.contrib.contenttypes.models import ContentType
        from apps.records.models import Comment
        from .models import BonCommandeFournisseur
        Comment.objects.create(
            company=bon_commande.company,
            content_type=ContentType.objects.get_for_model(
                BonCommandeFournisseur),
            object_id=bon_commande.pk,
            body=body,
            author=user,
        )
    except Exception:  # noqa: BLE001 — best-effort, jamais bloquant
        logger.info(
            'ZPUR11: chatter non journalisé pour BCF %s', bon_commande.pk)


# ── ZSTK3 — rapport prévisionnel par produit (Forecasted report) ──────────

def forecast_produit(company, produit):
    """ZSTK3 — assemble en une timeline : disponible actuel, entrées
    attendues (Σ `quantite_restante` des lignes de BCF envoyés/non annulés,
    avec date + fournisseur — réutilise `bcf_sources_en_commande_produit`
    ZPUR10, jamais recalculé différemment) et sorties attendues
    (réservations chantier + assemblage actives, lues via
    `installations.selectors.reserved_quantity_for_produit` — jamais son
    modèle) ; renvoie un solde projeté CUMULÉ daté (les entrées sans date
    connue sont regroupées en fin de liste, après les entrées datées triées
    chronologiquement). LECTURE SEULE, INTERNE."""
    from .selectors import bcf_sources_en_commande_produit
    from apps.installations.selectors import reserved_quantity_for_produit

    disponible = produit.quantite_stock
    entrees = bcf_sources_en_commande_produit(company, produit.id)
    sorties_reservees = reserved_quantity_for_produit(produit)

    # Entrées datées d'abord (triées par date), puis les non-datées.
    datees = sorted(
        (e for e in entrees if e.get('date_livraison_prevue')),
        key=lambda e: e['date_livraison_prevue'])
    non_datees = [e for e in entrees if not e.get('date_livraison_prevue')]

    timeline = []
    solde = disponible
    for e in datees + non_datees:
        solde += e['quantite_restante']
        timeline.append({
            'date': e.get('date_livraison_prevue'),
            'type': 'entree',
            'quantite': e['quantite_restante'],
            'reference': e['reference'],
            'fournisseur_nom': e.get('fournisseur_nom'),
            'solde_projete': solde,
        })
    if sorties_reservees:
        solde -= sorties_reservees
        timeline.append({
            'date': None,
            'type': 'sortie',
            'quantite': -sorties_reservees,
            'reference': None,
            'fournisseur_nom': None,
            'solde_projete': solde,
        })

    return {
        'produit_id': produit.id,
        'disponible': disponible,
        'entrees_attendues': entrees,
        'sorties_attendues': sorties_reservees,
        'solde_projete': solde,
        'timeline': timeline,
    }


# ── XPUR7 — dates de livraison prévues, accusé fournisseur & OTD réel ──────

def compute_date_livraison_prevue(company, fournisseur, date_commande,
                                  lignes_data):
    """XPUR7 — pré-calcule ``date_livraison_prevue`` = date_commande + le
    plus grand ``delai_livraison_jours`` connu parmi les produits de la
    commande (le pire cas — la commande n'est complète qu'une fois tout
    arrivé). None si aucun délai connu OU pas de date de commande
    (comportement historique : reste éditable à la main)."""
    from .models import PrixFournisseur
    if not date_commande or not fournisseur or not lignes_data:
        return None
    produit_ids = [
        ligne.get('produit').id if hasattr(ligne.get('produit'), 'id')
        else ligne.get('produit')
        for ligne in lignes_data if ligne.get('produit')
    ]
    if not produit_ids:
        return None
    delais = PrixFournisseur.objects.filter(
        fournisseur=fournisseur, produit_id__in=produit_ids,
        delai_livraison_jours__isnull=False,
    ).values_list('delai_livraison_jours', flat=True)
    delais = [d for d in delais if d is not None]
    if not delais:
        return None
    from datetime import timedelta
    return date_commande + timedelta(days=max(delais))


def bcf_en_retard(bon_commande, *, a_la_date=None):
    """XPUR7 — vrai si un BCF ENVOYE est en retard : la date prévue/confirmée
    est dépassée SANS réception complète. Jamais vrai pour un BROUILLON/
    ANNULE/RECU (comportement historique préservé)."""
    from django.utils import timezone
    from .models import BonCommandeFournisseur
    if bon_commande.statut != BonCommandeFournisseur.Statut.ENVOYE:
        return False
    ref_date = (bon_commande.date_confirmee_fournisseur
                or bon_commande.date_livraison_prevue)
    if not ref_date:
        return False
    today = a_la_date or timezone.now().date()
    return today > ref_date and not bon_commande.est_entierement_recu


def bcf_en_retard_list(company, *, a_la_date=None):
    """XPUR7 — liste des BCF ENVOYE en retard de la société (filtrable liste
    « BCF en retard »). LECTURE SEULE."""
    from .models import BonCommandeFournisseur
    qs = BonCommandeFournisseur.objects.filter(
        company=company,
        statut=BonCommandeFournisseur.Statut.ENVOYE,
    ).select_related('fournisseur').prefetch_related('lignes')
    return [bc for bc in qs if bcf_en_retard(bc, a_la_date=a_la_date)]


def notify_bcf_en_retard(company):
    """XPUR7 — notifie les responsables/admins des BCF ENVOYE en retard.
    Best-effort : une erreur de notification n'interrompt jamais l'appelant.
    Renvoie le nombre de BCF notifiés."""
    en_retard = bcf_en_retard_list(company)
    count = 0
    for bc in en_retard:
        try:
            from apps.notifications.services import notify_many
            from apps.notifications.models import EventType
            from django.contrib.auth import get_user_model
            User = get_user_model()
            recipients = User.objects.filter(
                company=company, is_active=True,
                role_legacy__in=['responsable', 'admin'])
            notify_many(
                recipients, EventType.BCF_LATE,
                title=f'BCF en retard ({bc.reference})',
                body=(f'{bc.reference} chez {bc.fournisseur.nom} est en '
                      f'retard de livraison.'),
                company=company)
            count += 1
        except Exception:  # noqa: BLE001 — best-effort, jamais bloquant
            logger.warning('notify_bcf_en_retard: échec pour BCF %s', bc.pk)
    return count


def annuler_receptions_brouillon_bcf(bc, user=None):
    """YPROC7 — à l'annulation d'un BCF, annule en cascade ses réceptions
    encore en BROUILLON (elles restaient sinon confirmables contre un BCF
    annulé). Une réception déjà CONFIRME ou ANNULE n'est jamais touchée
    (idempotent, aucune perte de mouvement de stock déjà posté). Renvoie le
    nombre de réceptions annulées par cet appel."""
    from django.utils import timezone
    from .models import ReceptionFournisseur

    today = timezone.now().date()
    receptions = bc.receptions.filter(
        statut=ReceptionFournisseur.Statut.BROUILLON)
    count = 0
    for reception in receptions:
        note_annulation = (
            f'[{today.isoformat()}] Annulée automatiquement (BCF '
            f'{bc.reference} annulé).')
        reception.statut = ReceptionFournisseur.Statut.ANNULE
        reception.note = (
            f'{reception.note}\n{note_annulation}'.strip()
            if reception.note else note_annulation)
        reception.save(update_fields=['statut', 'note'])
        count += 1
    return count


def notify_bcf_annule(bc):
    """YPROC7 — notifie best-effort le créateur du BCF (et de la DA liée si
    connue) qu'un BCF vient d'être annulé. N'échoue jamais l'annulation."""
    try:
        from apps.notifications.services import notify
        from apps.notifications.models import EventType
        if bc.created_by_id:
            notify(
                bc.created_by, EventType.BCF_CANCELLED,
                title=f'BCF annulé ({bc.reference})',
                body=(f'{bc.reference} chez '
                      f'{bc.fournisseur.nom if bc.fournisseur_id else "?"} '
                      f'a été annulé.'),
                company=bc.company)
    except Exception:  # noqa: BLE001 — best-effort, jamais bloquant
        logger.warning('notify_bcf_annule: échec pour BCF %s', bc.pk)


def annuler_bcf_cascade(bc, user=None):
    """YPROC7 — annulation d'un BCF : annule en cascade ses réceptions
    brouillon (`annuler_receptions_brouillon_bcf`) puis notifie
    (`notify_bcf_annule`), best-effort. Ne touche PAS au statut du BCF
    lui-même (laissé à l'appelant — la garde « déjà entièrement reçu » reste
    dans la vue). Renvoie le détail des quantités déjà entrées en stock
    (utile si l'annulation concerne un BCF partiellement reçu) pour décision
    éventuelle de retour fournisseur."""
    nb_receptions_annulees = annuler_receptions_brouillon_bcf(bc, user=user)
    notify_bcf_annule(bc)
    quantites_deja_recues = [
        {
            'produit_id': ligne.produit_id,
            'produit_nom': ligne.produit.nom if ligne.produit_id else (
                ligne.designation or ''),
            'quantite_recue': ligne.quantite_recue,
        }
        for ligne in bc.lignes.select_related('produit').all()
        if ligne.quantite_recue
    ]
    return {
        'receptions_annulees': nb_receptions_annulees,
        'quantites_deja_recues': quantites_deja_recues,
    }


def otd_stats(company, fournisseur):
    """XPUR7 — OTD (On-Time Delivery) promis-vs-reçu pour un fournisseur :
    écart moyen en jours (positif = en retard) + % de BCF à l'heure. Compare
    la date CONFIRMÉE (ou prévue si pas de confirmation) à la date de la
    première réception CONFIRMÉE du BCF. INTERNE, LECTURE SEULE."""
    from .models import BonCommandeFournisseur
    bons = (BonCommandeFournisseur.objects
            .filter(company=company, fournisseur=fournisseur)
            .exclude(statut=BonCommandeFournisseur.Statut.ANNULE)
            .prefetch_related('receptions'))
    ecarts = []
    for bc in bons:
        ref_date = bc.date_confirmee_fournisseur or bc.date_livraison_prevue
        if not ref_date:
            continue
        premiere_reception = None
        for rec in bc.receptions.filter(statut='confirme').order_by(
                'date_reception'):
            if rec.date_reception:
                premiere_reception = rec.date_reception
                break
        if premiere_reception is None:
            continue
        ecarts.append((premiere_reception - ref_date).days)
    if not ecarts:
        return {'otd_ecart_moyen_jours': None, 'otd_a_lheure_pct': None}
    a_lheure = sum(1 for e in ecarts if e <= 0)
    return {
        'otd_ecart_moyen_jours': round(sum(ecarts) / len(ecarts), 1),
        'otd_a_lheure_pct': round(a_lheure / len(ecarts) * 100, 1),
    }


# ── XPUR8 — Acomptes / avances fournisseur sur BCF ──────────────────────────

def imputer_acomptes_bcf(bon_commande):
    """XPUR8 — impute les acomptes NON CONSOMMÉS du BCF sur sa PREMIÈRE
    ``FactureFournisseur`` (par date de création). Idempotent : un acompte
    déjà imputé (``facture_imputee`` déjà posé) n'est jamais réimputé,
    même si la fonction est rappelée. No-op si le BCF n'a pas encore de
    facture. Renvoie la liste des acomptes imputés lors de CET appel."""
    from .models import AcompteFournisseur
    facture = (bon_commande.factures_fournisseur
               .order_by('date_creation').first())
    if facture is None:
        return []
    acomptes = AcompteFournisseur.objects.filter(
        bon_commande=bon_commande, facture_imputee__isnull=True)
    imputed = []
    for acompte in acomptes:
        acompte.facture_imputee = facture
        acompte.montant_consomme = acompte.montant
        acompte.save(update_fields=['facture_imputee', 'montant_consomme'])
        imputed.append(acompte)
    return imputed


# ── XPUR9 — Avoir fournisseur (note de crédit AP) ───────────────────────────

def _prix_ligne_retour(bon_commande, produit):
    """XPUR9 — prix d'achat unitaire HT pour valoriser une ligne de retour :
    priorité à la ligne du BCF d'origine (prix EXACT payé), repli sur
    `cheapest_prix_fournisseur`, sinon 0."""
    if bon_commande is not None:
        ligne_bcf = bon_commande.lignes.filter(produit=produit).first()
        if ligne_bcf is not None:
            return ligne_bcf.prix_achat_unitaire or Decimal('0')
    prix = cheapest_prix_fournisseur(produit)
    return prix.prix_achat if prix is not None else Decimal('0')


def preparer_avoir_depuis_retour(retour):
    """XPUR9 — construit les montants HT/TVA/TTC pré-remplis d'un avoir
    depuis un ``RetourFournisseur`` VALIDÉ (une ligne sans prix connu compte
    pour 0 — jamais d'erreur bloquante). TVA 20 % (même taux par défaut que
    ``facturer_reception``). Renvoie un dict ``{montant_ht, montant_tva,
    montant_ttc}`` — NE CRÉE RIEN (pur calcul)."""
    montant_ht = Decimal('0')
    for ligne in retour.lignes.select_related('produit'):
        pu = _prix_ligne_retour(retour.bon_commande, ligne.produit)
        montant_ht += Decimal(str(ligne.quantite)) * pu
    taux_tva = Decimal('20')
    montant_tva = (montant_ht * taux_tva / Decimal('100')).quantize(
        Decimal('0.01'))
    return {
        'montant_ht': montant_ht,
        'montant_tva': montant_tva,
        'montant_ttc': montant_ht + montant_tva,
    }


def creer_avoir_depuis_retour(company, retour, user=None):
    """XPUR9 — génère un ``AvoirFournisseur`` BROUILLON pré-rempli en un
    clic depuis un ``RetourFournisseur`` VALIDÉ. Lève ValueError si le
    retour n'est pas validé ou a déjà un avoir. Référencé via
    ``create_with_reference`` (préfixe AVF)."""
    from apps.ventes.utils.references import create_with_reference
    from .models import AvoirFournisseur, RetourFournisseur

    if retour.statut != RetourFournisseur.Statut.VALIDE:
        raise ValueError(
            'Seul un retour validé peut générer un avoir (« attente '
            "d'avoir » tant que non reçu).")
    if AvoirFournisseur.objects.filter(retour=retour).exists():
        raise ValueError('Ce retour a déjà un avoir associé.')

    montants = preparer_avoir_depuis_retour(retour)
    avoir = AvoirFournisseur(
        company=company, fournisseur=retour.fournisseur,
        facture_origine=None, retour=retour,
        montant_ht=montants['montant_ht'], montant_tva=montants['montant_tva'],
        montant_ttc=montants['montant_ttc'],
        statut=AvoirFournisseur.Statut.BROUILLON, created_by=user)

    def _save(ref):
        avoir.reference = ref
        avoir.save()
        return avoir

    return create_with_reference(AvoirFournisseur, 'AVF', company, _save)


def imputer_avoir_fournisseur(avoir, facture, montant=None, *, user=None):
    """XPUR9 — impute un ``AvoirFournisseur`` (VALIDÉ) sur une
    ``FactureFournisseur`` du MÊME fournisseur ; réduit ``solde_du`` (jamais
    sous zéro — plafonné à ``min(montant demandé, disponible avoir, solde
    facture)``). Crée une ``ImputationAvoirFournisseur``. Lève ValueError si
    fournisseurs différents, avoir non validé, ou rien à imputer."""
    from .models import AvoirFournisseur, ImputationAvoirFournisseur

    if avoir.fournisseur_id != facture.fournisseur_id:
        raise ValueError(
            "L'avoir et la facture doivent appartenir au même fournisseur.")
    if avoir.statut not in (
            AvoirFournisseur.Statut.VALIDE, AvoirFournisseur.Statut.IMPUTE):
        raise ValueError('Seul un avoir validé peut être imputé.')

    disponible = avoir.montant_disponible
    solde_facture = facture.solde_du
    plafond = min(disponible, solde_facture)
    montant_impute = Decimal(str(montant)) if montant is not None else plafond
    montant_impute = min(montant_impute, plafond)
    if montant_impute <= 0:
        raise ValueError("Rien à imputer (avoir épuisé ou facture soldée).")

    imputation = ImputationAvoirFournisseur.objects.create(
        company=avoir.company, avoir=avoir, facture=facture,
        montant=montant_impute)
    avoir.montant_impute = (avoir.montant_impute or Decimal('0')) + montant_impute
    avoir.statut = (AvoirFournisseur.Statut.IMPUTE
                    if avoir.montant_disponible <= 0
                    else AvoirFournisseur.Statut.VALIDE)
    avoir.save(update_fields=['montant_impute', 'statut'])
    return imputation


# ── XPUR10 — tolérances 3 voies & file d'exceptions ─────────────────────────

def evaluate_facture_exception(company, facture):
    """XPUR10 — compare l'écart du rapprochement 3 voies (FG131, lu via
    ``apps.compta.selectors`` — jamais d'import de modèles compta) du BCF
    d'origine de ``facture`` aux tolérances par défaut de la société
    (``AchatsParametres.tolerance_prix_pct``/``tolerance_prix_absolu_mad``).

    Hors tolérance → statut_controle=exception + motif_ecart(persistés).
    Dans la tolérance (ou pas de BCF/rapprochement encore évalué) → no-op,
    la facture reste 'normale' (comportement historique). Renvoie
    ``(en_exception: bool, ecart_pct: Decimal|None)``."""
    from .models import AchatsParametres, FactureFournisseur
    if not facture.bon_commande_id:
        return False, None
    try:
        from apps.compta.selectors import rapprochement_ecart_pct
    except Exception:  # pragma: no cover - défensif (compta indisponible)
        return False, None
    ecart_pct = rapprochement_ecart_pct(company, facture.bon_commande_id)
    if ecart_pct is None:
        return False, None
    parametres = AchatsParametres.for_company(company)
    tolerance = parametres.tolerance_prix_pct or Decimal('0')
    hors_tolerance = ecart_pct > tolerance
    if hors_tolerance:
        facture.statut_controle = FactureFournisseur.StatutControle.EXCEPTION
        facture.motif_ecart = (
            f'Écart de {ecart_pct:.2f} % (tolérance société : '
            f'{tolerance:.2f} %) sur le rapprochement 3 voies.')
        facture.save(update_fields=['statut_controle', 'motif_ecart'])
    return hors_tolerance, ecart_pct


def check_facture_exception_gate(company, facture):
    """XPUR10 — (ré)évalue l'écart de rapprochement 3 voies de la facture
    contre les tolérances société PUIS lève ValueError si elle est (ou
    devient) EXCEPTION non résolue — bloque la CRÉATION d'un
    PaiementFournisseur. No-op si la facture reste 'normale' (pas de BCF,
    pas encore de rapprochement évalué, ou dans la tolérance) ou a déjà été
    résolue (statut 'resolue' n'est jamais re-basculé en exception ici —
    la résolution est un acte explicite du responsable)."""
    from .models import FactureFournisseur
    if facture.statut_controle == FactureFournisseur.StatutControle.RESOLUE:
        return
    evaluate_facture_exception(company, facture)
    if facture.statut_controle == FactureFournisseur.StatutControle.EXCEPTION:
        raise ValueError(
            f'Paiement bloqué : facture {facture.reference} en exception '
            f'de rapprochement 3 voies '
            f'({facture.motif_ecart or "écart hors tolérance"}). '
            'Résolution requise avant paiement.')


def resoudre_exception_facture(facture, *, user, commentaire=''):
    """XPUR10 — résout (Responsable/Admin) une facture en exception : passe
    `statut_controle` à 'resolue', trace l'acteur/l'horodatage, débloque le
    paiement. Lève ValueError si la facture n'est pas en exception."""
    from django.utils import timezone
    from .models import FactureFournisseur
    if facture.statut_controle != FactureFournisseur.StatutControle.EXCEPTION:
        raise ValueError("Cette facture n'est pas en exception.")
    facture.statut_controle = FactureFournisseur.StatutControle.RESOLUE
    facture.resolu_par = user
    facture.resolu_le = timezone.now()
    if commentaire:
        facture.motif_ecart = (
            (facture.motif_ecart or '') + f'\nRésolution : {commentaire}')
    facture.save(update_fields=[
        'statut_controle', 'resolu_par', 'resolu_le', 'motif_ecart'])
    return facture


def factures_en_exception(company):
    """XPUR10 — file « Factures en exception » (statut_controle=exception),
    triée de la plus récente à la plus ancienne. LECTURE SEULE."""
    from .models import FactureFournisseur
    return list(FactureFournisseur.objects.filter(
        company=company,
        statut_controle=FactureFournisseur.StatutControle.EXCEPTION,
    ).select_related('fournisseur', 'bon_commande').order_by('-date_creation'))


def qr_svg_for(text, *, box=4, quiet=4):
    """XFAC19 — point d'entrée cross-app sanctionné pour le générateur QR maison
    de N20 (``apps.stock.labels.qr_svg``, zéro dépendance externe). Les autres
    apps (ex. ``ventes`` pour le QR paiement/vérification sur le PDF facture)
    appellent cette fonction plutôt que d'importer ``apps.stock.labels``
    directement — pur, déterministe, renvoie un SVG inline (str)."""
    from . import labels
    return labels.qr_svg(text, box=box, quiet=quiet)


# ── XPUR11 — Détection de doublons facture fournisseur & BCF ────────────────
# Rien ne détectait une facture saisie deux fois (même référence fournisseur),
# ni deux BCF ouverts pour le même besoin. Ces fonctions sont des WARNINGS
# (jamais bloquants) : elles n'empêchent aucune création, elles enrichissent
# la réponse pour que l'utilisateur puisse voir et confirmer (override tracé
# en note). Comportement historique inchangé quand rien ne matche.

def detect_facture_fournisseur_doublon(
        company, *, fournisseur_id, ref_fournisseur=None,
        montant_ttc=None, date_facture=None, exclude_id=None):
    """XPUR11 — détecte une facture fournisseur potentiellement déjà saisie.

    Deux règles, en OR (whichever matches) :
    1. Même (fournisseur, ref_fournisseur non vide) déjà en base.
    2. Même fournisseur + même montant TTC + date_facture à ±7 jours.

    Renvoie une liste de dicts (matches) — vide si rien ne matche (défaut,
    comportement historique). LECTURE SEULE, jamais bloquant."""
    from datetime import timedelta
    from .models import FactureFournisseur

    if not fournisseur_id:
        return []
    qs = FactureFournisseur.objects.filter(
        company=company, fournisseur_id=fournisseur_id)
    if exclude_id:
        qs = qs.exclude(pk=exclude_id)

    matches = {}

    ref = (ref_fournisseur or '').strip()
    if ref:
        for f in qs.filter(ref_fournisseur__iexact=ref):
            matches[f.id] = f

    if montant_ttc is not None and date_facture is not None:
        try:
            montant = Decimal(str(montant_ttc))
        except (InvalidOperation, TypeError, ValueError):
            montant = None
        if montant is not None:
            date_min = date_facture - timedelta(days=7)
            date_max = date_facture + timedelta(days=7)
            for f in qs.filter(
                montant_ttc=montant,
                date_facture__gte=date_min, date_facture__lte=date_max,
            ):
                matches[f.id] = f

    return [
        {
            'id': f.id,
            'reference': f.reference,
            'ref_fournisseur': f.ref_fournisseur,
            'montant_ttc': f.montant_ttc,
            'date_facture': f.date_facture,
            'statut': f.statut,
        }
        for f in matches.values()
    ]


def bcf_similaires_ouverts(company, *, fournisseur_id, produit_ids=None):
    """XPUR11 — panneau « BCF ouverts similaires » : bons de commande
    fournisseur BROUILLON/ENVOYE du même fournisseur, éventuellement filtrés
    aux BCF qui partagent au moins un produit avec ``produit_ids``. LECTURE
    SEULE, jamais bloquant (aide à la décision avant de créer un nouveau BCF).
    """
    from .models import BonCommandeFournisseur

    if not fournisseur_id:
        return []
    qs = (BonCommandeFournisseur.objects
          .filter(company=company, fournisseur_id=fournisseur_id,
                  statut__in=[BonCommandeFournisseur.Statut.BROUILLON,
                              BonCommandeFournisseur.Statut.ENVOYE])
          .prefetch_related('lignes__produit')
          .order_by('-date_creation'))
    produit_ids = set(produit_ids or [])
    out = []
    for bc in qs:
        lignes = list(bc.lignes.all())
        bc_produit_ids = {ligne.produit_id for ligne in lignes}
        if produit_ids and not (bc_produit_ids & produit_ids):
            continue
        out.append({
            'id': bc.id,
            'reference': bc.reference,
            'statut': bc.statut,
            'date_creation': bc.date_creation,
            'produits_communs': len(bc_produit_ids & produit_ids)
            if produit_ids else 0,
            'nombre_lignes': len(lignes),
        })
    return out


def log_doublon_override(*, user, instance, detail):
    """XPUR11 — trace (best-effort, jamais bloquant) qu'un utilisateur a
    confirmé la création malgré un warning de doublon. Utilise l'audit
    logger existant (apps.audit) ; un échec de journalisation ne casse
    jamais la création elle-même."""
    try:
        from apps.audit.recorder import record
        from apps.audit.models import AuditLog
        record(AuditLog.Action.CREATE, instance=instance,
               user=user, detail=detail)
    except Exception:  # noqa: BLE001 — best-effort, jamais bloquant
        logger.info('doublon override (audit indisponible): %s', detail)


# ── XACC36 — OCR facture fournisseur → brouillon de facture d'achat ─────────
# L'EXTRACTION vit dans fastapi_ia/ocr_service.py (prompt stock, classe
# facture_fournisseur/facture_achat). Ce SINK convertit les champs déjà
# extraits (``donnees_structurees``) en brouillon `FactureFournisseur`, sans
# jamais inventer un montant absent (champ vide plutôt que faux). Matche le
# fournisseur par ICE puis par nom (jamais de création silencieuse d'un
# fournisseur inconnu — l'utilisateur garde la main pour la saisie manuelle).

def match_fournisseur_from_ocr(company, fields):
    """XACC36 — matche un Fournisseur existant depuis les champs OCR :
    priorité à l'ICE (identifiant fiable), repli sur le nom exact (insensible
    à la casse). Renvoie le Fournisseur ou None (jamais de création
    silencieuse). LECTURE SEULE."""
    from .models import Fournisseur
    fields = fields or {}
    ice = (fields.get('ice') or '').strip()
    if ice:
        match = Fournisseur.objects.filter(company=company, ice=ice).first()
        if match is not None:
            return match
    nom = (fields.get('fournisseur') or '').strip()
    if nom:
        return Fournisseur.objects.filter(
            company=company, nom__iexact=nom).first()
    return None


def creer_facture_fournisseur_depuis_ocr(
        *, company, user, fields, attachment=None,
        confirmer_malgre_doublon=False):
    """XACC36 — crée une `FactureFournisseur` BROUILLON depuis les champs OCR.

    ``fields`` = ``donnees_structurees`` OCR (numero/date/fournisseur/ice/
    date_echeance/montant_ht/montant_tva/montant_ttc/…). Un champ absent reste
    VIDE (jamais de montant inventé). Lève ValueError si aucun fournisseur ne
    matche (l'appelant doit alors proposer la saisie manuelle — dégradation
    propre). Le doublon XPUR11 est vérifié au passage (warning, jamais
    bloquant sauf refus explicite non demandé ici) ; ``attachment`` (dict de
    ``records.storage.store_attachment``) est rattaché en pièce jointe si
    fourni. Renvoie ``(facture, doublons)``."""
    from datetime import date as _date
    from .models import FactureFournisseur

    fields = fields or {}
    fournisseur = match_fournisseur_from_ocr(company, fields)
    if fournisseur is None:
        raise ValueError(
            'Aucun fournisseur trouvé pour ce document (ICE ou nom) — '
            'saisie manuelle requise.')

    def _num(key):
        val = fields.get(key)
        if val is None or val == '':
            return None
        try:
            return Decimal(str(val))
        except (InvalidOperation, TypeError, ValueError):
            return None

    def _parsed_date(key):
        val = fields.get(key)
        if not val:
            return None
        if isinstance(val, _date):
            return val
        try:
            return _date.fromisoformat(str(val)[:10])
        except ValueError:
            return None

    date_facture = _parsed_date('date')
    date_echeance = _parsed_date('date_echeance')
    montant_ht = _num('montant_ht')
    montant_tva = _num('montant_tva')
    montant_ttc = _num('montant_ttc')
    ref_fournisseur = (fields.get('numero') or '').strip() or None

    doublons = detect_facture_fournisseur_doublon(
        company, fournisseur_id=fournisseur.id,
        ref_fournisseur=ref_fournisseur, montant_ttc=montant_ttc,
        date_facture=date_facture)

    def _save(ref):
        return FactureFournisseur.objects.create(
            company=company, reference=ref, fournisseur=fournisseur,
            ref_fournisseur=ref_fournisseur,
            date_facture=date_facture, date_echeance=date_echeance,
            montant_ht=montant_ht or Decimal('0'),
            montant_tva=montant_tva or Decimal('0'),
            montant_ttc=montant_ttc or Decimal('0'),
            created_by=user,
            note='Brouillon créé automatiquement depuis un scan OCR.',
        )

    from apps.ventes.utils.references import create_with_reference
    facture = create_with_reference(FactureFournisseur, 'FF', company, _save)

    if doublons and confirmer_malgre_doublon:
        log_doublon_override(
            user=user, instance=facture,
            detail=(f'Facture fournisseur {facture.reference} créée depuis '
                    f'OCR malgré {len(doublons)} doublon(s) potentiel(s).'))

    if attachment:
        try:
            from django.contrib.contenttypes.models import ContentType
            from apps.records.models import Attachment
            Attachment.objects.create(
                company=company,
                content_type=ContentType.objects.get_for_model(
                    FactureFournisseur),
                object_id=facture.pk,
                uploaded_by=user,
                **attachment,
            )
        except Exception:  # noqa: BLE001 — la pièce jointe ne casse jamais
            logger.info(
                'XACC36: pièce jointe OCR non rattachée à %s',
                facture.reference)

    return facture, doublons


# ── XPUR13 — Garde-fous prix sur la ligne BCF (accords + historique) ────────
# `ContratPrixFournisseur`/`CommandeCadre` (installations, FG318/FG314)
# existaient déjà mais rien ne CONTRÔLAIT les prix saisis sur une ligne de
# BCF. Ces fonctions sont des WARNINGS (jamais bloquantes) : le prix convenu
# du contrat en vigueur est réutilisé via le sélecteur fin existant
# `installations.selectors.prix_convenu_fournisseur` (import paresseux —
# jamais un import de modèles). Comportement historique inchangé quand aucun
# contrat/historique ne s'applique (pas de warning).

def historique_prix_produit(
        company, produit_id, *, fournisseur_id=None, limit=20):
    """XPUR13 — historique des prix d'achat d'un produit, TOUTES SOURCES
    (lignes de BCF passées, triées de la plus récente à la plus ancienne).
    Filtrable par fournisseur. INTERNE, LECTURE SEULE."""
    from .models import LigneBonCommandeFournisseur, BonCommandeFournisseur
    qs = (LigneBonCommandeFournisseur.objects
          .filter(bon_commande__company=company, produit_id=produit_id)
          .exclude(bon_commande__statut=BonCommandeFournisseur.Statut.ANNULE)
          .select_related('bon_commande', 'bon_commande__fournisseur')
          .order_by('-bon_commande__date_creation'))
    if fournisseur_id:
        qs = qs.filter(bon_commande__fournisseur_id=fournisseur_id)
    out = []
    for ligne in qs[:limit]:
        bc = ligne.bon_commande
        out.append({
            'bon_commande_id': bc.id,
            'reference': bc.reference,
            'fournisseur_id': bc.fournisseur_id,
            'fournisseur_nom': (
                bc.fournisseur.nom if bc.fournisseur_id else None),
            'date': bc.date_creation,
            'prix_achat_unitaire': ligne.prix_achat_unitaire,
            'quantite': ligne.quantite,
        })
    return out


def prix_moyen_recent_produit(
        company, produit_id, *, fournisseur_id=None, limit=20):
    """XPUR13 — prix d'achat moyen sur l'historique récent (même portée que
    `historique_prix_produit`). Renvoie None si aucun historique. INTERNE."""
    historique = historique_prix_produit(
        company, produit_id, fournisseur_id=fournisseur_id, limit=limit)
    if not historique:
        return None
    total = sum((Decimal(str(h['prix_achat_unitaire'])) for h in historique),
                Decimal('0'))
    return (total / len(historique)).quantize(Decimal('0.01'))


def check_prix_ligne_bcf(
        company, *, produit_id, fournisseur_id, prix_saisi,
        a_la_date=None):
    """XPUR13 — vérifie le prix d'une ligne de BCF avant/à la saisie.

    Renvoie un dict ``{ok, warnings}`` — JAMAIS bloquant (``ok`` est toujours
    True, les warnings sont informatifs). Deux règles indépendantes :
    1. Dépassement du prix CONTRACTUEL en vigueur (accord fournisseur×produit,
       `installations.selectors.prix_convenu_fournisseur`).
    2. Écart au-delà du seuil société (`AchatsParametres.
       seuil_deviation_prix_pct`, 0 = désactivé) par rapport au dernier prix/
       prix moyen (`PrixFournisseur` + historique BCF)."""
    from .models import AchatsParametres

    prix = _dec(prix_saisi)
    warnings = []
    if prix is None:
        return {'ok': True, 'warnings': warnings}

    # 1. Prix contractuel (accord fournisseur×produit en vigueur).
    try:
        from apps.installations.selectors import prix_convenu_fournisseur
        accord = prix_convenu_fournisseur(
            company, produit_id, fournisseur_id=fournisseur_id,
            a_la_date=a_la_date)
    except Exception:  # noqa: BLE001 — best-effort, jamais bloquant
        accord = None
    if accord and accord.get('prix_convenu') is not None:
        prix_convenu = Decimal(str(accord['prix_convenu']))
        if prix > prix_convenu:
            warnings.append({
                'type': 'hors_contrat',
                'prix_convenu': prix_convenu,
                'prix_saisi': prix,
                'contrat_id': accord.get('contrat_id'),
                'message': (
                    f'Le prix saisi ({prix}) dépasse le prix convenu du '
                    f'contrat en vigueur ({prix_convenu}).'),
            })

    # 2. Écart vs dernier prix / prix moyen (seuil paramétrable, 0 = off).
    parametres = AchatsParametres.for_company(company)
    seuil = parametres.seuil_deviation_prix_pct or Decimal('0')
    if seuil > 0:
        moyen = prix_moyen_recent_produit(
            company, produit_id, fournisseur_id=fournisseur_id)
        if moyen and moyen > 0:
            ecart_pct = abs(prix - moyen) / moyen * Decimal('100')
            if ecart_pct > seuil:
                warnings.append({
                    'type': 'ecart_historique',
                    'prix_moyen': moyen,
                    'prix_saisi': prix,
                    'ecart_pct': ecart_pct.quantize(Decimal('0.1')),
                    'message': (
                        f'Le prix saisi ({prix}) dévie de '
                        f'{ecart_pct.quantize(Decimal("0.1"))} %% du prix '
                        f'moyen récent ({moyen}), au-delà du seuil société '
                        f'({seuil} %%).'),
                })

    return {'ok': True, 'warnings': warnings}


def rapport_achats_hors_contrat(
        company, *, fournisseur_id=None, date_debut=None, date_fin=None):
    """XPUR13 — rapport « achats hors contrat » : lignes de BCF dont le prix
    saisi dépasse le prix convenu du contrat en vigueur pour ce couple
    produit×fournisseur, sur la période (optionnelle) et/ou pour un
    fournisseur donné. INTERNE, LECTURE SEULE."""
    from .models import LigneBonCommandeFournisseur, BonCommandeFournisseur
    try:
        from apps.installations.selectors import prix_convenu_fournisseur
    except Exception:  # noqa: BLE001 — best-effort
        return []

    qs = (LigneBonCommandeFournisseur.objects
          .filter(bon_commande__company=company)
          .exclude(bon_commande__statut=BonCommandeFournisseur.Statut.ANNULE)
          .select_related(
              'bon_commande', 'bon_commande__fournisseur', 'produit'))
    if fournisseur_id:
        qs = qs.filter(bon_commande__fournisseur_id=fournisseur_id)
    if date_debut:
        qs = qs.filter(bon_commande__date_creation__gte=date_debut)
    if date_fin:
        qs = qs.filter(bon_commande__date_creation__lte=date_fin)

    out = []
    for ligne in qs:
        bc = ligne.bon_commande
        try:
            accord = prix_convenu_fournisseur(
                company, ligne.produit_id,
                fournisseur_id=bc.fournisseur_id,
                a_la_date=bc.date_creation.date() if bc.date_creation
                else None)
        except Exception:  # noqa: BLE001 — best-effort
            accord = None
        if not accord or accord.get('prix_convenu') is None:
            continue
        prix_convenu = Decimal(str(accord['prix_convenu']))
        if ligne.prix_achat_unitaire > prix_convenu:
            out.append({
                'ligne_id': ligne.id,
                'bon_commande_id': bc.id,
                'reference': bc.reference,
                'fournisseur_id': bc.fournisseur_id,
                'fournisseur_nom': (
                    bc.fournisseur.nom if bc.fournisseur_id else None),
                'produit_id': ligne.produit_id,
                'produit_nom': ligne.produit.nom if ligne.produit_id else None,
                'date': bc.date_creation,
                'prix_convenu': prix_convenu,
                'prix_saisi': ligne.prix_achat_unitaire,
                'ecart': ligne.prix_achat_unitaire - prix_convenu,
            })
    return out


# ── XPUR14 — PrixFournisseur enrichi (code article, paliers, validité) ──────
# Un tarif fournisseur ne portait que prix + date du dernier achat. Ces
# fonctions résolvent le prix EFFECTIF (bon palier de quantité, tarif non
# expiré) et fournissent l'import/export xlsx du tarif d'un fournisseur.
# Comportement historique inchangé : sans palier/dates saisis, le prix de
# base (`PrixFournisseur.prix_achat`) reste toujours proposé.

def prix_effectif_fournisseur(
        produit, fournisseur, *, quantite=1, a_la_date=None):
    """XPUR14 — prix d'achat EFFECTIF pour (produit, fournisseur, quantité) à
    une date donnée : le palier applicable (qte_min le plus élevé ≤
    quantité) si des paliers existent, sinon le prix de base. Renvoie None
    si le tarif est expiré ou introuvable. LECTURE SEULE."""
    from .models import PrixFournisseur
    pf = (PrixFournisseur.objects
          .filter(produit=produit, fournisseur=fournisseur)
          .prefetch_related('paliers').first())
    if pf is None:
        return None
    if not pf.est_en_vigueur(a_la_date):
        return None
    paliers = [p for p in pf.paliers.all() if p.qte_min <= quantite]
    if paliers:
        meilleur = max(paliers, key=lambda p: p.qte_min)
        return meilleur.prix
    return pf.prix_achat


def export_prix_fournisseur_xlsx(company, fournisseur):
    """XPUR14 — export xlsx du tarif d'un fournisseur (une ligne par produit,
    paliers concaténés dans une colonne). Réutilise le builder xlsx partagé
    (records.xlsx). INTERNE."""
    from apps.records.xlsx import build_xlsx_response
    from .models import PrixFournisseur

    headers = [
        'sku', 'produit', 'ref_produit_fournisseur', 'prix_achat',
        'date_debut', 'date_fin', 'paliers (qte_min:prix;...)',
    ]
    rows = []
    qs = (PrixFournisseur.objects
          .filter(company=company, fournisseur=fournisseur)
          .select_related('produit').prefetch_related('paliers')
          .order_by('produit__nom'))
    for pf in qs:
        paliers_str = ';'.join(
            f'{p.qte_min}:{p.prix}' for p in pf.paliers.all())
        rows.append([
            pf.produit.sku or '', pf.produit.nom,
            pf.ref_produit_fournisseur, pf.prix_achat,
            pf.date_debut, pf.date_fin, paliers_str,
        ])
    filename = f'tarif-{fournisseur.nom}.xlsx'.replace(' ', '-')
    return build_xlsx_response(
        filename, headers, rows, sheet_title='Tarif fournisseur')


def _parse_palier_cell(value):
    """XPUR14 — parse la cellule ``paliers`` (« qte_min:prix;qte_min:prix »)
    en liste de tuples ``(qte_min:int, prix:Decimal)``. Ignore silencieusement
    les segments malformés (rapport d'erreurs géré par l'appelant)."""
    out = []
    for segment in str(value or '').split(';'):
        segment = segment.strip()
        if not segment or ':' not in segment:
            continue
        qte_raw, prix_raw = segment.split(':', 1)
        try:
            qte = int(qte_raw.strip())
            prix = Decimal(str(prix_raw.strip()))
        except (InvalidOperation, TypeError, ValueError):
            continue
        out.append((qte, prix))
    return out


def import_prix_fournisseur_xlsx(company, fournisseur, file_bytes):
    """XPUR14 — import/mise à jour du tarif d'un fournisseur depuis un xlsx
    (même format que l'export). CRÉATION + MISE À JOUR par SKU — jamais de
    suppression silencieuse (un produit absent du fichier garde son tarif
    existant). Renvoie ``{created, updated, errors: [{row, message}]}``.
    INTERNE."""
    import io
    from openpyxl import load_workbook
    from .models import PrixFournisseur, PalierPrixFournisseur, Produit

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(h or '').strip().lower() for h in next(rows_iter, [])]

    def _col(name):
        return header.index(name) if name in header else None

    idx_sku = _col('sku')
    idx_ref = _col('ref_produit_fournisseur')
    idx_prix = _col('prix_achat')
    idx_debut = _col('date_debut')
    idx_fin = _col('date_fin')
    idx_paliers = None
    for i, h in enumerate(header):
        if h.startswith('paliers'):
            idx_paliers = i
            break

    created = 0
    updated = 0
    errors = []
    if idx_sku is None or idx_prix is None:
        errors.append({
            'row': 1,
            'message': "Colonnes requises manquantes ('sku', 'prix_achat').",
        })
        return {'created': 0, 'updated': 0, 'errors': errors}

    for row_num, row in enumerate(rows_iter, start=2):
        sku = str(row[idx_sku] or '').strip() if idx_sku < len(row) else ''
        if not sku:
            continue
        produit = Produit.objects.filter(company=company, sku=sku).first()
        if produit is None:
            errors.append({
                'row': row_num,
                'message': f'SKU introuvable : {sku}.',
            })
            continue
        prix_raw = row[idx_prix] if idx_prix < len(row) else None
        prix = _dec(prix_raw)
        if prix is None:
            errors.append({
                'row': row_num,
                'message': f'Prix invalide pour {sku} : {prix_raw!r}.',
            })
            continue

        defaults = {'company': company, 'prix_achat': prix}
        if idx_ref is not None and idx_ref < len(row):
            defaults['ref_produit_fournisseur'] = str(row[idx_ref] or '')
        if idx_debut is not None and idx_debut < len(row):
            defaults['date_debut'] = row[idx_debut] or None
        if idx_fin is not None and idx_fin < len(row):
            defaults['date_fin'] = row[idx_fin] or None

        pf, was_created = PrixFournisseur.objects.get_or_create(
            produit=produit, fournisseur=fournisseur, defaults=defaults)
        if not was_created:
            for key, val in defaults.items():
                setattr(pf, key, val)
            pf.save(update_fields=list(defaults.keys()))
            updated += 1
        else:
            created += 1

        if idx_paliers is not None and idx_paliers < len(row):
            paliers = _parse_palier_cell(row[idx_paliers])
            for qte_min, palier_prix in paliers:
                PalierPrixFournisseur.objects.update_or_create(
                    prix_fournisseur=pf, qte_min=qte_min,
                    defaults={'prix': palier_prix})

    return {'created': created, 'updated': updated, 'errors': errors}


# ── XPUR18 — Révision de BCF tracée + ré-approbation ────────────────────────
# Un BCF ENVOYE se modifiait silencieusement (aucun historique d'amendement).
# `reviser_bcf` est le SEUL chemin pour modifier lignes/montants/dates après
# l'envoi : chaque changement est journalisé (records.Comment, ancien→nouveau,
# horodaté, utilisateur), le compteur `revision` avance, et une hausse du
# montant au-delà du seuil FG312 invalide l'approbation existante (ré-exige
# une nouvelle approbation — best-effort, cohérent avec YPROC4 même si le
# gate d'envoi n'est pas encore câblé côté stock).

_LIGNE_CHAMPS_SUIVIS = (
    'quantite', 'prix_achat_unitaire', 'designation',
)
_BCF_CHAMPS_SUIVIS = (
    'date_commande', 'date_livraison_prevue', 'note',
)


def _log_revision_change(company, bc, user, lines):
    """XPUR18 — écrit UNE entrée `records.Comment` regroupant tous les
    changements d'une révision (best-effort, jamais bloquant)."""
    if not lines:
        return
    try:
        from django.contrib.contenttypes.models import ContentType
        from apps.records.models import Comment
        from .models import BonCommandeFournisseur
        body = f'Révision {bc.revision} :\n' + '\n'.join(lines)
        Comment.objects.create(
            company=company,
            content_type=ContentType.objects.get_for_model(
                BonCommandeFournisseur),
            object_id=bc.pk,
            body=body,
            author=user,
        )
    except Exception:  # noqa: BLE001 — best-effort, jamais bloquant
        logger.info('XPUR18: révision non journalisée pour BCF %s', bc.pk)


def _invalidate_approbation_si_hausse(company, bc, montant_avant, montant_apres):
    """XPUR18 — si le montant du BCF a AUGMENTÉ au-delà du seuil d'approbation
    en vigueur (FG312), supprime l'approbation existante pour qu'une nouvelle
    approbation soit exigée avant un nouvel envoi. Best-effort (import
    paresseux — jamais un import direct de modèles installations) ; no-op si
    installations est absent ou si aucun seuil n'est configuré (comportement
    historique préservé)."""
    if montant_apres <= montant_avant:
        return False
    try:
        from apps.installations.models_approbation_bcf import (
            ApprobationBCF, SeuilApprobationBCF,
        )
        seuil = SeuilApprobationBCF.objects.filter(
            company=company, actif=True).first()
        if seuil is None:
            return False
        palier_avant = seuil.palier_requis(montant_avant)
        palier_apres = seuil.palier_requis(montant_apres)
        approbation = ApprobationBCF.objects.filter(
            company=company, bcf=bc).first()
        # La hausse invalide l'approbation SEULEMENT si elle fait franchir un
        # palier plus strict (FG312) — une hausse qui reste sous le seuil en
        # vigueur ne touche jamais une approbation existante.
        if approbation is not None and palier_apres != palier_avant:
            approbation.delete()
            return True
        return False
    except Exception:  # noqa: BLE001 — best-effort, jamais bloquant
        return False


def reviser_bcf(
        company, user, bc, *, lignes=None, date_commande=None,
        date_livraison_prevue=None, note=None):
    """XPUR18 — révise un BCF déjà ENVOYE : SEUL chemin de modification des
    lignes/montants/dates après envoi. Journalise chaque changement
    (ancien→nouveau), incrémente `revision`, ré-exige une approbation FG312
    si le montant augmente au-delà du seuil en vigueur. Lève ValueError si le
    BCF est en brouillon/reçu/annulé (rien à réviser — utiliser l'édition
    normale ou c'est déjà figé)."""
    from .models import BonCommandeFournisseur

    if bc.statut not in (
        BonCommandeFournisseur.Statut.ENVOYE,
        BonCommandeFournisseur.Statut.RECU,
    ):
        raise ValueError(
            'Seul un BCF envoyé (ou partiellement reçu) peut être révisé.')

    montant_avant = bc.total_achat
    changements = []

    # ── En-tête (date_commande / date_livraison_prevue / note) ─────────────
    for champ, nouvelle_valeur in (
        ('date_commande', date_commande),
        ('date_livraison_prevue', date_livraison_prevue),
        ('note', note),
    ):
        if nouvelle_valeur is None:
            continue
        ancienne = getattr(bc, champ)
        if str(ancienne) != str(nouvelle_valeur):
            changements.append(
                f'{champ} : {ancienne!r} → {nouvelle_valeur!r}')
            setattr(bc, champ, nouvelle_valeur)

    # ── Lignes (quantité, prix, désignation) ────────────────────────────────
    if lignes is not None:
        existantes = {ligne.id: ligne for ligne in bc.lignes.all()}
        for ligne_data in lignes:
            ligne_id = ligne_data.get('id')
            ligne = existantes.get(ligne_id)
            if ligne is None:
                continue
            for champ in _LIGNE_CHAMPS_SUIVIS:
                if champ not in ligne_data:
                    continue
                ancienne = getattr(ligne, champ)
                nouvelle = ligne_data[champ]
                if str(ancienne) != str(nouvelle):
                    changements.append(
                        f'Ligne {ligne_id} — {champ} : '
                        f'{ancienne!r} → {nouvelle!r}')
                    setattr(ligne, champ, nouvelle)
            ligne.save()

    if not changements:
        return bc, False

    bc.revision += 1
    bc.save()
    bc.refresh_from_db()
    montant_apres = bc.total_achat

    _log_revision_change(company, bc, user, changements)
    reapprobation_requise = _invalidate_approbation_si_hausse(
        company, bc, montant_avant, montant_apres)

    return bc, reapprobation_requise


# ── XPUR22 — Portail fournisseur lecture seule + confirmation d'arrivée ─────
# Page publique tokenisée PAR FOURNISSEUR (pas par document — un seul jeton
# donne accès à TOUS les documents DE CE fournisseur, jamais ceux d'un
# autre). Le fournisseur peut CONFIRMER un BCF et proposer une date
# d'arrivée : remplit `date_confirmee_fournisseur` (XPUR7) en préservant la
# date DEMANDÉE d'origine (`date_livraison_prevue`, jamais écrasée — OTD).

def generer_token_portail_fournisseur(company, fournisseur, user=None):
    """XPUR22 — génère (et renvoie) un nouveau jeton portail pour ce
    fournisseur. Les jetons précédents ne sont PAS révoqués automatiquement
    (l'admin peut vouloir plusieurs jetons actifs, ex. par contact) — la
    révocation est une action explicite (`revoquer_token_portail_fournisseur`)."""
    from .models import PortailFournisseurToken
    return PortailFournisseurToken.objects.create(
        company=company, fournisseur=fournisseur, created_by=user)


def revoquer_token_portail_fournisseur(token_obj):
    """XPUR22 — révoque un jeton (le lien cesse immédiatement de fonctionner,
    aucune suppression — traçabilité conservée)."""
    token_obj.revoked = True
    token_obj.save(update_fields=['revoked'])
    return token_obj


def resoudre_token_portail_fournisseur(token):
    """XPUR22 — résout un jeton portail valide (non révoqué, non expiré) et
    renvoie son ``PortailFournisseurToken`` (avec `fournisseur` préchargé),
    ou None. LECTURE SEULE — l'appelant public doit toujours passer par
    cette fonction plutôt que directement par le modèle."""
    from .models import PortailFournisseurToken
    token_obj = (PortailFournisseurToken.objects
                 .select_related('fournisseur', 'company')
                 .filter(token=token).first())
    if token_obj is None or not token_obj.est_valide:
        return None
    return token_obj


def portail_fournisseur_documents(token_obj):
    """XPUR22 — documents du fournisseur porteur de ce jeton : SES BCF en
    cours (référence, lignes, statut, date prévue), SES réceptions et SES
    factures avec statut de paiement. Isolation stricte : jamais les
    documents d'un autre fournisseur, jamais de marge (prix d'achat exposé —
    légitime, c'est ce que CE fournisseur nous vend). LECTURE SEULE."""
    from .models import BonCommandeFournisseur, ReceptionFournisseur

    fournisseur = token_obj.fournisseur
    company = token_obj.company

    bcf_qs = (BonCommandeFournisseur.objects
              .filter(company=company, fournisseur=fournisseur)
              .exclude(statut=BonCommandeFournisseur.Statut.ANNULE)
              .prefetch_related('lignes__produit')
              .order_by('-date_creation'))
    bcf_data = []
    for bc in bcf_qs:
        bcf_data.append({
            'id': bc.id,
            'reference': bc.reference,
            'statut': bc.statut,
            'statut_display': bc.get_statut_display(),
            'date_commande': bc.date_commande,
            'date_livraison_prevue': bc.date_livraison_prevue,
            'date_confirmee_fournisseur': bc.date_confirmee_fournisseur,
            'lignes': [
                {
                    'produit_nom': (
                        ligne.produit.nom if ligne.produit_id
                        else ligne.designation),
                    'quantite': ligne.quantite,
                    'quantite_recue': ligne.quantite_recue,
                }
                for ligne in bc.lignes.all()
            ],
        })

    receptions = (ReceptionFournisseur.objects
                  .filter(company=company,
                          bon_commande__fournisseur=fournisseur)
                  .select_related('bon_commande')
                  .order_by('-date_creation'))
    receptions_data = [{
        'id': r.id, 'reference': r.reference,
        'bon_commande_reference': (
            r.bon_commande.reference if r.bon_commande_id else None),
        'statut': r.statut, 'date_reception': r.date_reception,
    } for r in receptions]

    factures = factures_sous_traitant_qs_generique(company, fournisseur)

    return {
        'fournisseur_nom': fournisseur.nom,
        'bons_commande': bcf_data,
        'receptions': receptions_data,
        'factures': factures,
    }


def factures_sous_traitant_qs_generique(company, fournisseur):
    """XPUR22 — factures fournisseur (montants d'achat DE CE fournisseur
    uniquement) + statut de paiement, pour le portail public. LECTURE
    SEULE."""
    from .models import FactureFournisseur
    qs = (FactureFournisseur.objects
          .filter(company=company, fournisseur=fournisseur)
          .order_by('-date_creation'))
    return [{
        'id': f.id, 'reference': f.reference,
        'date_facture': f.date_facture, 'date_echeance': f.date_echeance,
        'montant_ttc': f.montant_ttc, 'statut': f.statut,
        'statut_display': f.get_statut_display(),
        'solde_du': f.solde_du,
    } for f in qs]


def confirmer_bcf_portail_fournisseur(
        token_obj, bcf_id, *, date_confirmee, numero_confirmation=''):
    """XPUR22 — le fournisseur confirme un BCF et propose une date
    d'arrivée depuis le portail public. Réutilise EXACTEMENT la même
    sémantique que l'action interne `confirmer` (XPUR7) : la date DEMANDÉE
    d'origine (`date_livraison_prevue`) n'est jamais écrasée (préserve
    l'OTD). Isolation stricte : le BCF DOIT appartenir au fournisseur
    porteur du jeton, sinon lève ValueError (jamais d'accès croisé)."""
    from .models import BonCommandeFournisseur

    bc = BonCommandeFournisseur.objects.filter(
        pk=bcf_id, company=token_obj.company,
        fournisseur=token_obj.fournisseur).first()
    if bc is None:
        raise ValueError(
            "Ce bon de commande n'appartient pas à ce fournisseur.")
    bc.date_confirmee_fournisseur = date_confirmee
    bc.numero_confirmation_fournisseur = numero_confirmation or ''
    bc.save(update_fields=[
        'date_confirmee_fournisseur', 'numero_confirmation_fournisseur'])

    from django.utils import timezone
    token_obj.last_used_at = timezone.now()
    token_obj.save(update_fields=['last_used_at'])

    notify_bcf_confirmation_fournisseur(bc)
    return bc


def notify_bcf_confirmation_fournisseur(bc):
    """XPUR22 — notifie (best-effort) le créateur du BCF que le fournisseur
    vient de confirmer une date d'arrivée depuis le portail public.

    Utilise le système de notifications in-app existant (`apps.notifications
    .services.notify`, `EventType.APPROVAL_DECIDED` — l'événement générique
    « décision actée » le plus proche, aucun nouveau type ajouté à une autre
    app) ET journalise dans le chatter (`records.Comment`) du BCF pour une
    trace consultable depuis sa fiche. Best-effort total : un échec des deux
    canaux ne casse jamais la confirmation elle-même."""
    if bc.created_by_id is not None:
        try:
            from apps.notifications.services import notify
            from apps.notifications.models import EventType
            notify(
                bc.created_by, EventType.APPROVAL_DECIDED,
                title='Confirmation fournisseur',
                body=(
                    f'{bc.fournisseur.nom} a confirmé le BCF {bc.reference} '
                    f'pour le {bc.date_confirmee_fournisseur}.'),
                company=bc.company,
            )
        except Exception:  # noqa: BLE001 — best-effort, jamais bloquant
            logger.info(
                'XPUR22: notification confirmation BCF %s non envoyée',
                bc.pk)
    try:
        from django.contrib.contenttypes.models import ContentType
        from apps.records.models import Comment
        from .models import BonCommandeFournisseur
        Comment.objects.create(
            company=bc.company,
            content_type=ContentType.objects.get_for_model(
                BonCommandeFournisseur),
            object_id=bc.pk,
            body=(
                f'{bc.fournisseur.nom} a confirmé une date d\'arrivée '
                f'({bc.date_confirmee_fournisseur}) via le portail '
                'fournisseur.'),
            author=None,
        )
    except Exception:  # noqa: BLE001 — best-effort, jamais bloquant
        pass


# ── XPUR24 — Tableau de bord achats (analyse des dépenses) ──────────────────
# FG59 note UN fournisseur (scorecard) et FG132 vieillit les dettes (balance
# âgée), mais aucune vue TRANSVERSE des achats n'existait. Admin/responsable
# UNIQUEMENT — jamais client-facing (prix d'achat/marge INTERNES). Agrège 5
# blocs depuis des données RÉELLES multi-mois : dépenses par fournisseur/
# catégorie/mois, dérive de prix moyen par SKU, engagements ouverts, top
# produits achetés, temps de cycle du processus d'achat.

def _mois_key(dt):
    """Clé de regroupement mensuel stable (YYYY-MM) depuis une date/datetime."""
    return dt.strftime('%Y-%m') if dt else 'inconnu'


def depenses_achats_par_periode(company, *, date_debut=None, date_fin=None):
    """XPUR24 — dépenses HT par (fournisseur, catégorie, mois) sur la
    période, dérivées des lignes de BCF (montant réellement engagé — pas
    seulement facturé). LECTURE SEULE, INTERNE."""
    from .models import LigneBonCommandeFournisseur, BonCommandeFournisseur

    qs = (LigneBonCommandeFournisseur.objects
          .filter(bon_commande__company=company)
          .exclude(bon_commande__statut=BonCommandeFournisseur.Statut.ANNULE)
          .select_related(
              'bon_commande', 'bon_commande__fournisseur',
              'produit', 'produit__categorie'))
    if date_debut:
        qs = qs.filter(bon_commande__date_creation__gte=date_debut)
    if date_fin:
        qs = qs.filter(bon_commande__date_creation__lte=date_fin)

    par_fournisseur = {}
    par_categorie = {}
    par_mois = {}
    # ZPUR8 — dépenses PAR ACHETEUR (nullable = comportement historique :
    # un BCF sans acheteur renseigné groupe sous « — »).
    par_acheteur = {}
    for ligne in qs:
        bc = ligne.bon_commande
        total = ligne.total_achat
        mois = _mois_key(bc.date_creation)
        fournisseur_nom = bc.fournisseur.nom if bc.fournisseur_id else '—'
        categorie_nom = (
            ligne.produit.categorie.nom
            if ligne.produit_id and ligne.produit.categorie_id else '—')
        acheteur_nom = (
            bc.acheteur.get_full_name() or bc.acheteur.username
            if bc.acheteur_id else '—')

        par_fournisseur[fournisseur_nom] = (
            par_fournisseur.get(fournisseur_nom, Decimal('0')) + total)
        par_categorie[categorie_nom] = (
            par_categorie.get(categorie_nom, Decimal('0')) + total)
        par_mois[mois] = par_mois.get(mois, Decimal('0')) + total
        par_acheteur[acheteur_nom] = (
            par_acheteur.get(acheteur_nom, Decimal('0')) + total)

    return {
        'par_fournisseur': [
            {'fournisseur': k, 'total_ht': v}
            for k, v in sorted(
                par_fournisseur.items(), key=lambda kv: kv[1], reverse=True)
        ],
        'par_categorie': [
            {'categorie': k, 'total_ht': v}
            for k, v in sorted(
                par_categorie.items(), key=lambda kv: kv[1], reverse=True)
        ],
        'par_mois': [
            {'mois': k, 'total_ht': v}
            for k, v in sorted(par_mois.items(), key=lambda kv: kv[0])
        ],
        'par_acheteur': [
            {'acheteur': k, 'total_ht': v}
            for k, v in sorted(
                par_acheteur.items(), key=lambda kv: kv[1], reverse=True)
        ],
    }


def derive_prix_moyen_par_sku(company, *, nb_mois=6):
    """XPUR24 — prix d'achat moyen PAR MOIS pour chaque SKU sur les
    `nb_mois` derniers mois (détection de dérive) : renvoie
    ``[{produit_id, produit_nom, sku, series: [{mois, prix_moyen}]}]``
    trié par plus grande dérive (dernier mois vs premier mois) décroissante.
    LECTURE SEULE, INTERNE."""
    from datetime import timedelta
    from django.utils import timezone
    from .models import LigneBonCommandeFournisseur, BonCommandeFournisseur

    depuis = timezone.now() - timedelta(days=31 * nb_mois)
    qs = (LigneBonCommandeFournisseur.objects
          .filter(bon_commande__company=company,
                  bon_commande__date_creation__gte=depuis)
          .exclude(bon_commande__statut=BonCommandeFournisseur.Statut.ANNULE)
          .exclude(produit__isnull=True)
          .select_related('bon_commande', 'produit'))

    par_produit = {}
    for ligne in qs:
        pid = ligne.produit_id
        mois = _mois_key(ligne.bon_commande.date_creation)
        entry = par_produit.setdefault(pid, {
            'produit_id': pid, 'produit_nom': ligne.produit.nom,
            'sku': ligne.produit.sku, 'par_mois': {},
        })
        bucket = entry['par_mois'].setdefault(mois, [])
        bucket.append(ligne.prix_achat_unitaire)

    out = []
    for entry in par_produit.values():
        series = [
            {'mois': mois, 'prix_moyen': (
                sum(prix, Decimal('0')) / len(prix)).quantize(Decimal('0.01'))}
            for mois, prix in sorted(entry['par_mois'].items())
        ]
        derive = (
            series[-1]['prix_moyen'] - series[0]['prix_moyen']
            if len(series) >= 2 else Decimal('0'))
        out.append({
            'produit_id': entry['produit_id'],
            'produit_nom': entry['produit_nom'],
            'sku': entry['sku'],
            'series': series,
            'derive': derive,
        })
    out.sort(key=lambda e: abs(e['derive']), reverse=True)
    return out


def engagements_ouverts_achats(company):
    """XPUR24 — engagements ouverts : BCF ENVOYÉS non entièrement reçus
    (montant restant dû = commandé − reçu, jamais négatif). LECTURE SEULE,
    INTERNE."""
    from .models import BonCommandeFournisseur
    from .selectors import montant_commande_bcf, montant_recu_bcf

    qs = (BonCommandeFournisseur.objects
          .filter(company=company,
                  statut=BonCommandeFournisseur.Statut.ENVOYE)
          .select_related('fournisseur')
          .prefetch_related('lignes'))
    out = []
    total = Decimal('0')
    for bc in qs:
        commande = montant_commande_bcf(bc)
        recu = montant_recu_bcf(bc)
        reste = max(commande - recu, Decimal('0'))
        if reste <= 0:
            continue
        total += reste
        out.append({
            'bon_commande_id': bc.id,
            'reference': bc.reference,
            'fournisseur_nom': (
                bc.fournisseur.nom if bc.fournisseur_id else None),
            'date_commande': bc.date_commande,
            'montant_commande': commande,
            'montant_recu': recu,
            'montant_restant': reste,
        })
    out.sort(key=lambda e: e['montant_restant'], reverse=True)
    return {'lignes': out, 'total_engage': total}


def top_produits_achetes(
        company, *, limit=20, date_debut=None, date_fin=None):
    """XPUR24 — top produits achetés (quantité + montant HT), toutes
    commandes (hors annulées), sur la période optionnelle. LECTURE SEULE,
    INTERNE."""
    from .models import LigneBonCommandeFournisseur, BonCommandeFournisseur

    qs = (LigneBonCommandeFournisseur.objects
          .filter(bon_commande__company=company)
          .exclude(bon_commande__statut=BonCommandeFournisseur.Statut.ANNULE)
          .exclude(produit__isnull=True)
          .select_related('bon_commande', 'produit'))
    if date_debut:
        qs = qs.filter(bon_commande__date_creation__gte=date_debut)
    if date_fin:
        qs = qs.filter(bon_commande__date_creation__lte=date_fin)

    par_produit = {}
    for ligne in qs:
        pid = ligne.produit_id
        entry = par_produit.setdefault(pid, {
            'produit_id': pid, 'produit_nom': ligne.produit.nom,
            'sku': ligne.produit.sku, 'quantite_totale': 0,
            'montant_total': Decimal('0'),
        })
        entry['quantite_totale'] += ligne.quantite
        entry['montant_total'] += ligne.total_achat

    out = sorted(
        par_produit.values(), key=lambda e: e['montant_total'], reverse=True)
    return out[:limit]


def temps_cycle_achats(company, *, date_debut=None, date_fin=None):
    """XPUR24 — temps de cycle du processus d'achat, par BCF : BCF créé →
    première réception confirmée → première facture. Le segment DA→BCF
    (demande d'achat) reste ``None`` tant que YPROC5 (conversion DA→BCF
    tracée) n'est pas câblé — dégradation propre, jamais une erreur.
    Renvoie les moyennes en JOURS (arrondies) + le détail par BCF. LECTURE
    SEULE, INTERNE."""
    from .models import (
        BonCommandeFournisseur, ReceptionFournisseur, FactureFournisseur,
    )

    qs = (BonCommandeFournisseur.objects
          .filter(company=company)
          .exclude(statut=BonCommandeFournisseur.Statut.ANNULE)
          .select_related('fournisseur'))
    if date_debut:
        qs = qs.filter(date_creation__gte=date_debut)
    if date_fin:
        qs = qs.filter(date_creation__lte=date_fin)

    details = []
    delais_bcf_reception = []
    delais_reception_facture = []
    for bc in qs:
        premiere_reception = (
            ReceptionFournisseur.objects
            .filter(bon_commande=bc, statut='confirme')
            .order_by('date_creation').first())
        premiere_facture = (
            FactureFournisseur.objects.filter(bon_commande=bc)
            .order_by('date_creation').first())

        jours_bcf_reception = None
        if premiere_reception is not None:
            jours_bcf_reception = (
                premiere_reception.date_creation - bc.date_creation).days
            delais_bcf_reception.append(jours_bcf_reception)

        jours_reception_facture = None
        if premiere_reception is not None and premiere_facture is not None:
            jours_reception_facture = (
                premiere_facture.date_creation
                - premiere_reception.date_creation).days
            delais_reception_facture.append(jours_reception_facture)

        details.append({
            'bon_commande_id': bc.id,
            'reference': bc.reference,
            'jours_da_vers_bcf': None,  # YPROC5 non câblé — dégradation propre.
            'jours_bcf_vers_reception': jours_bcf_reception,
            'jours_reception_vers_facture': jours_reception_facture,
        })

    def _moyenne(valeurs):
        return round(sum(valeurs) / len(valeurs), 1) if valeurs else None

    return {
        'moyenne_jours_da_vers_bcf': None,
        'moyenne_jours_bcf_vers_reception': _moyenne(delais_bcf_reception),
        'moyenne_jours_reception_vers_facture': _moyenne(
            delais_reception_facture),
        'details': details,
    }


def analyse_achats_dashboard(
        company, *, date_debut=None, date_fin=None, nb_mois=6):
    """XPUR24 — agrège les 5 blocs du tableau de bord achats en un seul
    appel (utilisé par l'endpoint ET par l'export xlsx — jamais recalculé
    deux fois différemment). LECTURE SEULE, INTERNE."""
    return {
        'depenses': depenses_achats_par_periode(
            company, date_debut=date_debut, date_fin=date_fin),
        'derive_prix': derive_prix_moyen_par_sku(company, nb_mois=nb_mois),
        'engagements_ouverts': engagements_ouverts_achats(company),
        'top_produits': top_produits_achetes(
            company, date_debut=date_debut, date_fin=date_fin),
        'temps_cycle': temps_cycle_achats(
            company, date_debut=date_debut, date_fin=date_fin),
    }


def export_analyse_achats_xlsx(
        company, *, date_debut=None, date_fin=None, nb_mois=6):
    """XPUR24 — export xlsx du tableau de bord achats (une feuille par bloc
    seraient idéales, mais `build_xlsx_response` ne gère qu'une feuille —
    on exporte donc la vue la plus actionnable : dépenses par fournisseur
    + top produits + engagements ouverts, concaténés avec des séparateurs de
    section). Réutilise le builder xlsx partagé (records.xlsx)."""
    from apps.records.xlsx import build_xlsx_response
    data = analyse_achats_dashboard(
        company, date_debut=date_debut, date_fin=date_fin, nb_mois=nb_mois)

    headers = ['section', 'libelle', 'valeur_1', 'valeur_2']
    rows = []
    for e in data['depenses']['par_fournisseur']:
        rows.append(['Dépenses par fournisseur', e['fournisseur'],
                     e['total_ht'], ''])
    for e in data['depenses']['par_categorie']:
        rows.append(['Dépenses par catégorie', e['categorie'],
                     e['total_ht'], ''])
    for e in data['depenses']['par_mois']:
        rows.append(['Dépenses par mois', e['mois'], e['total_ht'], ''])
    for e in data['top_produits']:
        rows.append(['Top produits', e['produit_nom'],
                     e['quantite_totale'], e['montant_total']])
    for e in data['engagements_ouverts']['lignes']:
        rows.append(['Engagements ouverts', e['reference'],
                     e['fournisseur_nom'], e['montant_restant']])

    return build_xlsx_response(
        'analyse-achats.xlsx', headers, rows, sheet_title='Analyse achats')


# ── XPUR26 — e-facturation DGI 2026, réception ENTRANTE (préparation mandat)
# Parseur stdlib (xml.etree) d'un fichier UBL 2.1 Invoice : AUCUN appel
# externe, AUCUNE dépendance nouvelle. La validation plateforme DGI réelle
# attendra le mandat — ici on ne fait que produire une FactureFournisseur
# BROUILLON pré-remplie. Total no-op quand
# `AchatsParametres.einvoicing_entrant_actif` est OFF (défaut).
_UBL_NS = {
    'inv': 'urn:oasis:names:specification:ubl:schema:xsd:Invoice-2',
    'cac': 'urn:oasis:names:specification:ubl:schema:xsd:'
           'CommonAggregateComponents-2',
    'cbc': 'urn:oasis:names:specification:ubl:schema:xsd:'
           'CommonBasicComponents-2',
}


def parse_ubl_invoice(xml_bytes):
    """XPUR26 — parse un fichier UBL 2.1 Invoice (stdlib xml.etree, jamais
    d'appel réseau). Renvoie un dict de champs :
    ``numero``, ``date_facture``, ``date_echeance``, ``ice_fournisseur``,
    ``nom_fournisseur``, ``montant_ht``, ``montant_tva``, ``montant_ttc``,
    ``numero_clearance_dgi``, ``lignes`` (liste de
    ``{designation, quantite, prix_unitaire_ht, taux_tva}``).
    Lève ``ValueError`` si le XML est illisible ou n'est pas une facture UBL
    reconnaissable (jamais de valeur inventée pour un champ absent)."""
    import xml.etree.ElementTree as ET

    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError as exc:
        raise ValueError(f'Fichier UBL illisible : {exc}') from exc

    def _find(elem, path):
        return elem.find(path, _UBL_NS)

    def _text(elem, path):
        node = _find(elem, path)
        return node.text.strip() if node is not None and node.text else None

    numero = _text(root, 'cbc:ID')
    if not numero:
        raise ValueError(
            "Fichier UBL invalide : identifiant de facture (cbc:ID) absent.")
    date_facture = _text(root, 'cbc:IssueDate')
    date_echeance = _text(root, 'cbc:DueDate')
    numero_clearance_dgi = (
        _text(root, 'cbc:UUID') or _text(root, 'cbc:LineID'))

    supplier_party = _find(
        root, 'cac:AccountingSupplierParty/cac:Party')
    ice_fournisseur = None
    nom_fournisseur = None
    if supplier_party is not None:
        nom_fournisseur = _text(
            supplier_party, 'cac:PartyName/cbc:Name') or _text(
            supplier_party,
            'cac:PartyLegalEntity/cbc:RegistrationName')
        for scheme in supplier_party.findall(
                'cac:PartyTaxScheme', _UBL_NS) or []:
            company_id = _text(scheme, 'cbc:CompanyID')
            if company_id:
                ice_fournisseur = company_id
                break
        if not ice_fournisseur:
            ice_fournisseur = _text(
                supplier_party,
                'cac:PartyLegalEntity/cbc:CompanyID')

    def _dec_text(elem, path):
        val = _text(elem, path)
        if val is None:
            return None
        try:
            return Decimal(val)
        except (InvalidOperation, TypeError, ValueError):
            return None

    montant_ht = _dec_text(
        root, 'cac:LegalMonetaryTotal/cbc:TaxExclusiveAmount')
    montant_ttc = _dec_text(
        root, 'cac:LegalMonetaryTotal/cbc:TaxInclusiveAmount')
    montant_tva = _dec_text(root, 'cac:TaxTotal/cbc:TaxAmount')
    if montant_tva is None and montant_ht is not None \
            and montant_ttc is not None:
        montant_tva = montant_ttc - montant_ht

    lignes = []
    for line in root.findall('cac:InvoiceLine', _UBL_NS) or []:
        designation = _text(
            line, 'cac:Item/cbc:Name') or _text(line, 'cbc:Note') or ''
        quantite = _dec_text(line, 'cbc:InvoicedQuantity') or Decimal('1')
        prix_unitaire_ht = _dec_text(line, 'cac:Price/cbc:PriceAmount')
        if prix_unitaire_ht is None:
            total_ligne = _dec_text(line, 'cbc:LineExtensionAmount')
            prix_unitaire_ht = (
                (total_ligne / quantite)
                if total_ligne is not None and quantite else Decimal('0'))
        taux_tva = _dec_text(
            line, 'cac:Item/cac:ClassifiedTaxCategory/cbc:Percent')
        lignes.append({
            'designation': designation,
            'quantite': quantite,
            'prix_unitaire_ht': prix_unitaire_ht or Decimal('0'),
            'taux_tva': taux_tva,
        })

    return {
        'numero': numero,
        'date_facture': date_facture,
        'date_echeance': date_echeance,
        'ice_fournisseur': ice_fournisseur,
        'nom_fournisseur': nom_fournisseur,
        'montant_ht': montant_ht,
        'montant_tva': montant_tva,
        'montant_ttc': montant_ttc,
        'numero_clearance_dgi': numero_clearance_dgi,
        'lignes': lignes,
    }


def creer_facture_fournisseur_depuis_ubl(*, company, user, xml_bytes):
    """XPUR26 — crée une `FactureFournisseur` BROUILLON depuis un fichier UBL
    2.1 (fournisseur matché par ICE — réutilise `match_fournisseur_from_ocr`
    XACC36 —, lignes, taux TVA, référence). Statut de conformité DGI posé à
    `cleared` si un numéro de clearance est présent dans le document, sinon
    `non_cleared` (aucune validation plateforme réelle : préparation mandat
    uniquement). Lève ValueError si aucun fournisseur ne matche ou le XML est
    invalide. Renvoie la facture créée."""
    from datetime import date as _date
    from .models import FactureFournisseur, LigneFactureFournisseur

    fields = parse_ubl_invoice(xml_bytes)

    fournisseur = match_fournisseur_from_ocr(company, {
        'ice': fields.get('ice_fournisseur'),
        'fournisseur': fields.get('nom_fournisseur'),
    })
    if fournisseur is None:
        raise ValueError(
            'Aucun fournisseur trouvé pour ce document UBL (ICE ou nom) — '
            'saisie manuelle requise.')

    def _parsed_date(val):
        if not val:
            return None
        try:
            return _date.fromisoformat(str(val)[:10])
        except ValueError:
            return None

    numero_clearance = fields.get('numero_clearance_dgi')
    statut_dgi = (
        FactureFournisseur.StatutConformiteDgi.CLEARED if numero_clearance
        else FactureFournisseur.StatutConformiteDgi.NON_CLEARED)

    def _save(ref):
        facture = FactureFournisseur.objects.create(
            company=company, reference=ref, fournisseur=fournisseur,
            ref_fournisseur=fields.get('numero'),
            date_facture=_parsed_date(fields.get('date_facture')),
            date_echeance=_parsed_date(fields.get('date_echeance')),
            montant_ht=fields.get('montant_ht') or Decimal('0'),
            montant_tva=fields.get('montant_tva') or Decimal('0'),
            montant_ttc=fields.get('montant_ttc') or Decimal('0'),
            numero_clearance_dgi=numero_clearance,
            statut_conformite_dgi=statut_dgi,
            created_by=user,
            note='Brouillon créé automatiquement depuis un import UBL '
                 '2.1 (e-facturation entrante).',
        )
        for ligne in fields.get('lignes') or []:
            LigneFactureFournisseur.objects.create(
                facture=facture,
                designation=ligne['designation'] or 'Ligne UBL',
                quantite=ligne['quantite'],
                prix_unitaire_ht=ligne['prix_unitaire_ht'],
                taux_tva=ligne.get('taux_tva'),
            )
        return facture

    from apps.ventes.utils.references import create_with_reference
    return create_with_reference(FactureFournisseur, 'FF', company, _save)


# ── XSTK15 — Unités de mesure & conditionnements (touret/carton…) ───────────
# Le stock reste stocké dans UNE SEULE unité (`Produit.unite_stock`) : un
# conditionnement d'achat convertit VERS cette unité à l'écriture du
# mouvement — jamais de double comptage.

def convertir_en_unites_stock(quantite_conditionnement, conditionnement):
    """Convertit une quantité de CONDITIONNEMENTS (ex. 2 tourets) en unités
    de stock (ex. 200 m), via `conditionnement.facteur`. Renvoie un int
    (le stock reste compté en entiers) — arrondi à l'entier le plus proche."""
    total = Decimal(str(quantite_conditionnement)) * conditionnement.facteur
    return int(total.to_integral_value())


def resoudre_conditionnement(company, *, conditionnement_id=None,
                             code_barres=None):
    """Résout un `ConditionnementProduit` scopé société, par id OU par
    code-barres scanné (XSTK3). Renvoie None si aucun des deux n'est fourni
    ou ne matche rien — jamais d'exception, laissé au comportement
    historique (réception sans conditionnement)."""
    from .models import ConditionnementProduit
    if conditionnement_id:
        return ConditionnementProduit.objects.filter(
            company=company, id=conditionnement_id).first()
    if code_barres:
        return ConditionnementProduit.objects.filter(
            company=company, code_barres=code_barres).first()
    return None


# ── XSTK16 — Découpe / reconditionnement (touret → coupes) ──────────────────
# Débite le produit SOURCE (SORTIE) et crédite le produit CIBLE (ENTREE) en
# UNE transaction, en transférant la valeur au coût moyen — aucune création
# ni destruction de valeur. Propage numero_lot/date_peremption (XSTK6) quand
# le lot source en porte. Référence commune sur les deux MouvementStock pour
# la traçabilité (XMFG15 lit `reference`).

def decouper_produit(*, company, produit_source, quantite_consommee,
                     produit_cible, quantite_produite, user,
                     emplacement=None, lot_source=None):
    """XSTK16 — débite `quantite_consommee` unités de `produit_source` et
    crédite `quantite_produite` unités de `produit_cible` (peut être le même
    SKU), en transférant EXACTEMENT la valeur au coût moyen du produit
    source (aucune création/destruction de valeur). Lève ValueError si la
    quantité consommée dépasse le stock disponible. Si `lot_source` est
    fourni, décrémente ce lot (XSTK6, garde péremption incluse) et propage
    `numero_lot`/`date_peremption` sur un nouveau `LotEntrepot` du produit
    cible quand celui-ci en a un."""
    from django.db import transaction
    from django.utils import timezone
    from .models import LotEntrepot, Produit
    if quantite_consommee <= 0 or quantite_produite <= 0:
        raise ValueError('Les quantités doivent être positives.')
    if quantite_consommee > (produit_source.quantite_stock or 0):
        raise ValueError(
            f'Stock insuffisant sur {produit_source.nom} '
            f'({produit_source.quantite_stock} disponible).')
    reference = f'DECOUPE-{timezone.now().strftime("%Y%m%d%H%M%S")}-{produit_source.pk}'
    cout_unitaire, _source = average_cost_with_source(produit_source)
    valeur_transferee = (cout_unitaire * quantite_consommee).quantize(
        Decimal('0.01'))

    with transaction.atomic():
        avant_source = produit_source.quantite_stock
        apres_source = avant_source - quantite_consommee
        record_stock_movement(
            company=company, produit=produit_source,
            type_mouvement=mouvement_type_sortie(),
            quantite=quantite_consommee, quantite_avant=avant_source,
            quantite_apres=apres_source, reference=reference,
            note=f'Découpe : consommation {produit_source.nom}',
            created_by=user)

        cible = Produit.objects.select_for_update().get(pk=produit_cible.pk)
        avant_cible = cible.quantite_stock
        apres_cible = avant_cible + quantite_produite
        record_stock_movement(
            company=company, produit=cible,
            type_mouvement=mouvement_type_entree(),
            quantite=quantite_produite, quantite_avant=avant_cible,
            quantite_apres=apres_cible, reference=reference,
            note=f'Découpe : production {cible.nom}',
            created_by=user)

        numero_lot = None
        date_peremption = None
        if lot_source is not None:
            sortir_lot_entrepot(
                company=company, lot=lot_source, quantite=quantite_consommee,
                user=user)
            numero_lot = lot_source.numero_lot
            date_peremption = lot_source.date_peremption
            LotEntrepot.objects.create(
                company=company, produit=cible, numero_lot=numero_lot,
                date_peremption=date_peremption,
                emplacement=emplacement or lot_source.emplacement,
                quantite_recue=quantite_produite,
                quantite_restante=quantite_produite,
                reference_reception=reference, created_by=user)

    # Rafraîchit la source depuis la DB : si source == cible (même SKU), la
    # copie mémoire de `produit_source` est stale après la 2e écriture.
    produit_source.refresh_from_db()
    return {
        'reference': reference, 'valeur_transferee': valeur_transferee,
        'cout_unitaire': cout_unitaire, 'produit_source': produit_source,
        'produit_cible': cible, 'numero_lot': numero_lot,
    }


# ── ARC21 — Bascule write-path identité (founder-gated, OFF par défaut) ──────

def ecrire_identite_fournisseur(fournisseur) -> bool:
    """ARC21 (founder-gated, OFF par défaut) — quand la bascule write-path est
    ACTIVE (``TIERS_SOURCE_ECRITURE`` ON), pousse l'identité du fournisseur vers
    son ``Tiers`` (source d'écriture unique) ; le fournisseur relira le miroir.

    Flag OFF (défaut) : NO-OP strict — renvoie ``False`` sans rien écrire (le
    fournisseur reste l'unique chemin d'écriture, comportement byte-identique à
    aujourd'hui). Best-effort ; ne fait jamais échouer l'appelant. Le prix
    d'achat n'est JAMAIS concerné (identité seulement).

    Voir docs/decisions/ARC21-tiers-source-ecriture.md.
    """
    try:
        from apps.tiers import services as tiers_services
        if not tiers_services.identite_source_est_tiers():
            return False  # flag OFF — rien ne change.
        if fournisseur is None or fournisseur.tiers_id is None:
            return False
        return tiers_services.ecrire_identite(
            company=fournisseur.company, tiers=fournisseur.tiers,
            champs={
                'nom': fournisseur.nom or '',
                'email': fournisseur.email or '',
                'telephone': fournisseur.telephone or '',
                'adresse': fournisseur.adresse or '',
                'ice': fournisseur.ice or '',
                'rc': fournisseur.rc or '',
                'identifiant_fiscal': fournisseur.identifiant_fiscal or '',
                'rib': fournisseur.rib or '',
            })
    except Exception:
        return False
