"""
Tests for the seed_catalogue management command (devis-simulator catalogue).

Run:
    docker compose exec django_core python manage.py test apps.stock -v 2
"""
from decimal import Decimal
from io import StringIO

from django.core.management import call_command
from django.test import TestCase

from apps.stock.models import Produit, MouvementStock


def make_company(slug='test-cat-co'):
    from authentication.models import Company
    company, _ = Company.objects.get_or_create(
        slug=slug, defaults={'nom': 'Test Catalogue Co'},
    )
    return company


def seed(company):
    out = StringIO()
    call_command('seed_catalogue', company_slug=company.slug, stdout=out)
    return out.getvalue()


class TestSeedCatalogue(TestCase):
    def setUp(self):
        self.company = make_company()

    def test_seeds_full_catalogue(self):
        seed(self.company)
        qs = Produit.objects.filter(company=self.company)
        # 31 solaire + 9 pompage + 16 VEICHI + 11 pompes OSP
        self.assertEqual(qs.count(), 67)
        # Spot-check key items: HT price = simulator TTC / 1.2
        huawei_10t = qs.get(sku='OND-R-HUA-10T')
        self.assertEqual(huawei_10t.nom, 'Onduleur réseau Huawei 10kW Triphasé')
        self.assertEqual(huawei_10t.prix_vente, Decimal('16666.67'))  # 20 000 TTC
        # Réforme TVA : panneau à 10 % — HT dérivé pour préserver 1 400 TTC
        panneau = qs.get(sku='PAN-CS-710')
        self.assertEqual(panneau.prix_vente, Decimal('1272.73'))      # 1 400 TTC @ 10 %
        self.assertEqual(panneau.tva, Decimal('10.00'))
        bat10 = qs.get(sku='BAT-DEY-10')
        self.assertEqual(bat10.prix_vente, Decimal('25000.00'))       # 30 000 TTC
        socles = qs.get(sku='SOC-BET')
        self.assertEqual(socles.prix_vente, Decimal('66.67'))         # 80 TTC
        # Stock available so auto-fill is never blocked
        self.assertTrue(all(p.quantite_stock > 0 for p in qs))
        # Traceability: one entry movement per product
        self.assertEqual(
            MouvementStock.objects.filter(
                company=self.company, reference='SEED-CATALOGUE').count(), 67,
        )

    def test_fiches_and_pompage_seeded(self):
        seed(self.company)
        qs = Produit.objects.filter(company=self.company)
        # Fiches commerciales remplies (marque/description/garantie)
        huawei = qs.get(sku='OND-R-HUA-10T')
        self.assertEqual(huawei.marque, 'Huawei')
        self.assertIn('FusionSolar', huawei.description)
        self.assertIn('10 ans', huawei.garantie)
        panneau = qs.get(sku='PAN-CS-710')
        self.assertIn('30 ans performance', panneau.garantie)
        # Pompage : specs de dimensionnement + prix d'achat laissé vide
        pompe = qs.get(sku='PMP-IMM-5.5T')
        self.assertEqual(str(pompe.pompe_cv), '5.50')
        self.assertEqual(pompe.prix_achat, 0)
        self.assertEqual(pompe.categorie.nom, 'Pompes')
        # Prix existants jamais modifiés par la passe fiches
        self.assertEqual(huawei.prix_vente, Decimal('16666.67'))

    def test_veichi_seeded_with_real_buy_and_sell_prices(self):
        seed(self.company)
        qs = Produit.objects.filter(company=self.company)
        v75 = qs.get(sku='VEI-SI23-7.5-380')
        self.assertEqual(v75.nom, 'VARIATEUR VEICHI SI23 7.5KW 380V')
        self.assertEqual(v75.prix_vente, Decimal('3333.33'))   # 4 000 TTC public
        self.assertEqual(v75.prix_achat, Decimal('2875.00'))   # 3 450 TTC revendeur
        self.assertEqual(str(v75.pompe_kw), '7.50')
        self.assertEqual(v75.tension_v, 380)
        self.assertEqual(v75.marque, 'VEICHI')
        self.assertEqual(v75.categorie.nom, 'Variateurs')
        # L'afficheur n'a pas de kW : il ne peut jamais être pris pour le variateur
        aff = qs.get(sku='VEI-SI22-AFF')
        self.assertIsNone(aff.pompe_kw)
        self.assertEqual(aff.prix_vente, Decimal('350.00'))    # 420 TTC
        self.assertEqual(aff.prix_achat, Decimal('300.00'))    # 360 TTC

    def test_osp_pumps_seeded_with_curves_and_empty_price(self):
        seed(self.company)
        p = Produit.objects.get(company=self.company, sku='PMP-OSP-30-8')
        self.assertEqual(p.prix_vente, Decimal('0'))   # à renseigner par le fondateur
        self.assertEqual(p.prix_achat, Decimal('0'))
        self.assertEqual(str(p.pompe_cv), '10.00')
        self.assertEqual(str(p.pompe_kw), '7.50')
        self.assertEqual(p.tension_v, 380)
        self.assertEqual(p.courbe_pompe['debits_m3h'], [0, 12, 24, 30, 36, 39])
        self.assertEqual(p.courbe_pompe['hmt_m'], [91, 85, 70, 60, 43, 34])

    def test_placeholder_coffrets_archived_prices_intact(self):
        # Un ancien coffret placeholder existant est archivé par le seeder
        # (autorisation fondateur) — jamais supprimé, prix jamais modifié.
        old = Produit.objects.create(
            company=self.company, nom='Variateur pompage solaire 5.5 CV Triphasé (coffret complet)',
            sku='VFD-PMP-5.5T', prix_vente=Decimal('5416.67'), quantite_stock=20,
        )
        seed(self.company)
        old.refresh_from_db()
        self.assertTrue(old.is_archived)
        self.assertEqual(old.prix_vente, Decimal('5416.67'))
        # Et le seeder ne les recrée jamais
        self.assertEqual(
            Produit.objects.filter(
                company=self.company, sku__startswith='VFD-PMP').count(), 1)

    def test_fiches_update_is_idempotent_and_price_safe(self):
        seed(self.company)
        before = dict(Produit.objects.filter(company=self.company)
                      .values_list('sku', 'prix_vente'))
        seed(self.company)
        after = dict(Produit.objects.filter(company=self.company)
                     .values_list('sku', 'prix_vente'))
        self.assertEqual(before, after)

    def test_idempotent_second_run_creates_nothing(self):
        seed(self.company)
        count_after_first = Produit.objects.filter(company=self.company).count()
        out = seed(self.company)
        self.assertEqual(
            Produit.objects.filter(company=self.company).count(), count_after_first)
        self.assertIn('0 created, 67 already present', out)

    def test_never_overwrites_existing_product(self):
        # Pre-existing product with the same name but a different price
        existing = Produit.objects.create(
            company=self.company, nom='Structures acier', sku='STR-LEGACY',
            prix_vente=Decimal('375.00'), prix_achat=Decimal('280.00'),
            quantite_stock=10,
        )
        out = seed(self.company)
        existing.refresh_from_db()
        # Untouched, no duplicate created under the catalogue SKU
        self.assertEqual(existing.prix_vente, Decimal('375.00'))
        self.assertFalse(
            Produit.objects.filter(company=self.company, sku='STR-ACIER').exists())
        self.assertEqual(
            Produit.objects.filter(
                company=self.company, nom__iexact='Structures acier').count(), 1)
        self.assertIn('Structures acier', out)

    def test_archived_product_frees_its_name_for_the_catalogue(self):
        # Un produit démo ARCHIVÉ ne bloque plus la création de la version
        # catalogue portant le même nom (l'actif, lui, bloque toujours).
        Produit.objects.create(
            company=self.company, nom='Structures acier', sku='STR-LEGACY2',
            prix_vente=Decimal('375.00'), quantite_stock=5, is_archived=True,
        )
        seed(self.company)
        actifs = Produit.objects.filter(
            company=self.company, nom__iexact='Structures acier',
            is_archived=False)
        self.assertEqual(actifs.count(), 1)
        self.assertEqual(actifs.first().sku, 'STR-ACIER')
        self.assertEqual(actifs.first().prix_vente, Decimal('416.67'))  # 500 TTC

    def test_tva_reform_panels_10_others_20_ttc_preserved(self):
        seed(self.company)
        qs = Produit.objects.filter(company=self.company)
        # TOUS les panneaux à 10 %, TTC strictement préservé
        for p in qs.filter(nom__icontains='panneau'):
            self.assertEqual(p.tva, Decimal('10.00'), p.nom)
            ttc = p.prix_vente * Decimal('1.10')
            self.assertEqual(ttc.quantize(Decimal('1')), Decimal('1400'), p.nom)
        # Tout le reste à 20 % (onduleurs, batteries, structures, pompes…)
        for p in qs.exclude(nom__icontains='panneau'):
            self.assertEqual(p.tva, Decimal('20.00'), p.nom)
        # Idempotent : un second passage ne retouche plus les prix
        before = dict(qs.values_list('sku', 'prix_vente'))
        seed(self.company)
        after = dict(Produit.objects.filter(company=self.company)
                     .values_list('sku', 'prix_vente'))
        self.assertEqual(before, after)

    def test_tva_reform_converts_existing_panel_preserving_ttc(self):
        # Un panneau créé AVANT la réforme (HT à 20 %) est converti :
        # 1 166,67 HT @20 % (1 400 TTC) → 1 272,73 HT @10 % (1 400 TTC)
        p = Produit.objects.create(
            company=self.company, nom='Panneau Maison 550W', sku='PAN-LEGACY',
            prix_vente=Decimal('1166.67'), prix_achat=Decimal('1000.00'),
            quantite_stock=5, tva=Decimal('20.00'),
        )
        seed(self.company)
        p.refresh_from_db()
        self.assertEqual(p.tva, Decimal('10.00'))
        self.assertEqual(p.prix_vente, Decimal('1272.73'))
        self.assertEqual(p.prix_achat, Decimal('1090.91'))  # 1 200 TTC préservé

    def test_taxonomy_every_product_in_exactly_one_ordered_category(self):
        from apps.stock.models import Categorie
        seed(self.company)
        qs = Produit.objects.filter(company=self.company)
        # chaque produit a une catégorie de la taxonomie (jamais orphelin)
        noms_taxo = {
            'Panneaux photovoltaïques', 'Onduleurs réseau', 'Onduleurs hybrides',
            'Batteries', 'Structures & fixation', 'Protection & accessoires',
            'Câbles', 'Pompes', 'Variateurs', 'Services & prestations',
        }
        for p in qs:
            self.assertIsNotNone(p.categorie, p.nom)
            self.assertIn(p.categorie.nom, noms_taxo, p.nom)
        # hybrides et réseau SÉPARÉS, spot-checks de rangement
        by = {p.sku: p.categorie.nom for p in qs}
        self.assertEqual(by['OND-R-HUA-10T'], 'Onduleurs réseau')
        self.assertEqual(by['OND-H-DEY-5M'], 'Onduleurs hybrides')
        self.assertEqual(by['PAN-CS-710'], 'Panneaux photovoltaïques')
        self.assertEqual(by['VEI-SI23-7.5-380'], 'Variateurs')
        self.assertEqual(by['VEI-SI22-AFF'], 'Variateurs')
        self.assertEqual(by['PMP-OSP-30-8'], 'Pompes')
        self.assertEqual(by['STR-ACIER'], 'Structures & fixation')
        self.assertEqual(by['SOC-BET'], 'Structures & fixation')
        self.assertEqual(by['CAB-6MM-M'], 'Câbles')
        self.assertEqual(by['SMART-MET'], 'Protection & accessoires')
        self.assertEqual(by['INST-CAT'], 'Services & prestations')
        self.assertEqual(by['SUIVI-2A'], 'Services & prestations')
        # ordre délibéré : panneaux d'abord, services en dernier
        cats = list(Categorie.objects.filter(
            company=self.company, nom__in=noms_taxo).order_by('ordre'))
        self.assertEqual(cats[0].nom, 'Panneaux photovoltaïques')
        self.assertEqual(cats[-1].nom, 'Services & prestations')
        # un produit du fondateur hors seed est aussi rangé (re-catégorisation)
        perso = Produit.objects.create(
            company=self.company, nom='Onduleur hybride Growatt 6kW',
            sku='OND-H-GRW-6', prix_vente=Decimal('15000'), quantite_stock=1)
        seed(self.company)
        perso.refresh_from_db()
        self.assertEqual(perso.categorie.nom, 'Onduleurs hybrides')

    def test_stock_read_only_role_writes_rejected(self):
        """Rôle fin « Commerciale » (stock_voir uniquement) : lecture OK,
        toute écriture Stock rejetée côté serveur ; un responsable hérité
        (sans rôle fin) garde l'écriture — rien ne change pour lui."""
        from rest_framework.test import APIClient
        from rest_framework_simplejwt.tokens import AccessToken
        from django.contrib.auth import get_user_model
        from apps.roles.models import Role

        User = get_user_model()
        role = Role.objects.create(
            company=self.company, nom='Commerciale', permissions=[
                'stock_voir', 'crm_voir', 'crm_creer', 'crm_modifier',
                'ventes_voir', 'ventes_creer', 'ventes_modifier',
                'ventes_valider', 'ventes_pdf', 'reporting_voir',
                'parametres_voir', 'users_voir',
            ])
        commerciale = User.objects.create_user(
            username='test_commerciale', password='x',
            company=self.company, role=role)
        legacy = User.objects.create_user(
            username='test_resp_legacy', password='x',
            company=self.company, role_legacy='responsable')
        seed(self.company)
        produit = Produit.objects.filter(company=self.company).first()

        http = APIClient()
        http.credentials(
            HTTP_AUTHORIZATION=f'Bearer {AccessToken.for_user(commerciale)}')
        # Lecture : autorisée
        self.assertEqual(http.get('/api/django/stock/produits/').status_code, 200)
        # Écritures : toutes rejetées
        r = http.patch(f'/api/django/stock/produits/{produit.id}/',
                       {'prix_vente': '1.00'}, format='json')
        self.assertEqual(r.status_code, 403)
        r = http.post('/api/django/stock/produits/',
                      {'nom': 'X', 'prix_vente': '1'}, format='json')
        self.assertEqual(r.status_code, 403)
        r = http.post('/api/django/stock/mouvements/',
                      {'produit': produit.id, 'type_mouvement': 'entree',
                       'quantite': 1}, format='json')
        self.assertEqual(r.status_code, 403)
        produit.refresh_from_db()
        self.assertNotEqual(produit.prix_vente, 0)  # rien n'a bougé

        # Responsable hérité : l'écriture passe toujours
        http2 = APIClient()
        http2.credentials(
            HTTP_AUTHORIZATION=f'Bearer {AccessToken.for_user(legacy)}')
        r = http2.patch(f'/api/django/stock/produits/{produit.id}/',
                        {'seuil_alerte': 9}, format='json')
        self.assertEqual(r.status_code, 200)

    def test_scoped_to_target_company_only(self):
        other = make_company(slug='test-cat-other')
        seed(self.company)
        self.assertEqual(Produit.objects.filter(company=other).count(), 0)


class TestBulkAndInlineEditing(TestCase):
    """Édition groupée (T8) + édition en ligne (T4) du catalogue produits."""

    def setUp(self):
        from rest_framework.test import APIClient
        from rest_framework_simplejwt.tokens import AccessToken
        from django.contrib.auth import get_user_model
        from apps.stock.models import Categorie

        User = get_user_model()
        self.company = make_company(slug='bulk-co')
        self.other = make_company(slug='bulk-other-co')
        self.cat = Categorie.objects.create(company=self.company, nom='Onduleurs')
        self.cat_other = Categorie.objects.create(company=self.other, nom='Autres')

        self.user = User.objects.create_user(
            username='bulk_resp', password='x',
            company=self.company, role_legacy='responsable')
        self.client = APIClient()
        self.client.credentials(
            HTTP_AUTHORIZATION=f'Bearer {AccessToken.for_user(self.user)}')

        self.p1 = Produit.objects.create(
            company=self.company, nom='Onduleur A', sku='OND-A',
            prix_vente=Decimal('1000.00'), prix_achat=Decimal('700.00'),
            quantite_stock=10)
        self.p2 = Produit.objects.create(
            company=self.company, nom='Onduleur B', sku='OND-B',
            prix_vente=Decimal('2000.00'), prix_achat=Decimal('1500.00'),
            quantite_stock=5)
        # Produit d'une AUTRE entreprise : doit toujours être ignoré.
        self.foreign = Produit.objects.create(
            company=self.other, nom='Etranger', sku='FOR-1',
            prix_vente=Decimal('500.00'), prix_achat=Decimal('400.00'),
            quantite_stock=3)

    def _bulk(self, action, ids, params=None):
        return self.client.post(
            '/api/django/stock/produits/bulk/',
            {'action': action, 'ids': ids, 'params': params or {}},
            format='json')

    # ── T8 : prix de vente en pourcentage ──
    def test_bulk_price_percent(self):
        r = self._bulk('change_prix', [self.p1.id, self.p2.id],
                       {'mode': 'percent', 'valeur': '10'})
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.data['updated'], 2)
        self.p1.refresh_from_db()
        self.p2.refresh_from_db()
        self.assertEqual(self.p1.prix_vente, Decimal('1100.00'))
        self.assertEqual(self.p2.prix_vente, Decimal('2200.00'))

    # ── T8 : prix de vente en montant fixe ──
    def test_bulk_price_fixed(self):
        r = self._bulk('change_prix', [self.p1.id],
                       {'mode': 'fixed', 'valeur': '250'})
        self.assertEqual(r.status_code, 200)
        self.p1.refresh_from_db()
        self.assertEqual(self.p1.prix_vente, Decimal('1250.00'))

    def test_bulk_price_negative_floors_at_zero(self):
        r = self._bulk('change_prix', [self.p1.id],
                       {'mode': 'fixed', 'valeur': '-99999'})
        self.assertEqual(r.status_code, 200)
        self.p1.refresh_from_db()
        self.assertEqual(self.p1.prix_vente, Decimal('0.00'))

    # ── prix_achat n'est JAMAIS modifié par un changement de prix de vente ──
    def test_bulk_price_never_touches_prix_achat(self):
        before1 = self.p1.prix_achat
        before2 = self.p2.prix_achat
        self._bulk('change_prix', [self.p1.id, self.p2.id],
                   {'mode': 'percent', 'valeur': '50'})
        self.p1.refresh_from_db()
        self.p2.refresh_from_db()
        self.assertEqual(self.p1.prix_achat, before1)
        self.assertEqual(self.p2.prix_achat, before2)

    # ── T8 : multi-tenant — les ids étrangers sont ignorés ──
    def test_bulk_ignores_foreign_ids(self):
        before = self.foreign.prix_vente
        r = self._bulk('change_prix', [self.p1.id, self.foreign.id],
                       {'mode': 'percent', 'valeur': '10'})
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.data['updated'], 1)  # seul p1 compté
        self.foreign.refresh_from_db()
        self.assertEqual(self.foreign.prix_vente, before)  # intact

    def test_bulk_only_foreign_ids_rejected(self):
        r = self._bulk('change_prix', [self.foreign.id],
                       {'mode': 'percent', 'valeur': '10'})
        self.assertEqual(r.status_code, 400)
        self.foreign.refresh_from_db()
        self.assertEqual(self.foreign.prix_vente, Decimal('500.00'))

    # ── T8 : garantie ──
    def test_bulk_set_garantie(self):
        r = self._bulk('set_garantie', [self.p1.id, self.p2.id],
                       {'garantie_mois': 60, 'garantie_production_mois': 300})
        self.assertEqual(r.status_code, 200)
        self.p1.refresh_from_db()
        self.assertEqual(self.p1.garantie_mois, 60)
        self.assertEqual(self.p1.garantie_production_mois, 300)

    # ── T8 : catégorie ──
    def test_bulk_set_categorie(self):
        r = self._bulk('set_categorie', [self.p1.id], {'categorie_id': self.cat.id})
        self.assertEqual(r.status_code, 200)
        self.p1.refresh_from_db()
        self.assertEqual(self.p1.categorie_id, self.cat.id)

    def test_bulk_set_categorie_foreign_rejected(self):
        r = self._bulk('set_categorie', [self.p1.id],
                       {'categorie_id': self.cat_other.id})
        self.assertEqual(r.status_code, 400)
        self.p1.refresh_from_db()
        self.assertIsNone(self.p1.categorie_id)

    # ── T8 : marque ──
    def test_bulk_set_marque(self):
        r = self._bulk('set_marque', [self.p1.id, self.p2.id], {'marque': 'Huawei'})
        self.assertEqual(r.status_code, 200)
        self.p1.refresh_from_db()
        self.p2.refresh_from_db()
        self.assertEqual(self.p1.marque, 'Huawei')
        self.assertEqual(self.p2.marque, 'Huawei')

    # ── T8 : export .xlsx ──
    def test_bulk_export_xlsx(self):
        r = self._bulk('export_xlsx', [self.p1.id, self.p2.id])
        self.assertEqual(r.status_code, 200)
        self.assertIn('spreadsheetml', r['Content-Type'])
        self.assertIn('attachment', r['Content-Disposition'])
        self.assertTrue(len(r.content) > 0)

    def test_bulk_export_excludes_prix_achat(self):
        # Le prix d'achat ne doit JAMAIS apparaître dans l'export client.
        import io
        import openpyxl
        r = self._bulk('export_xlsx', [self.p1.id])
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        self.assertNotIn('Prix achat', headers)
        for h in headers:
            self.assertNotIn('achat', (h or '').lower())
            self.assertNotIn('marge', (h or '').lower())
        # Et aucune cellule ne contient la valeur du prix d'achat.
        values = [c.value for row in ws.iter_rows() for c in row]
        self.assertNotIn(700.0, values)

    # ── T8 : action inconnue / aucun id ──
    def test_bulk_unknown_action(self):
        r = self._bulk('drop_table', [self.p1.id])
        self.assertEqual(r.status_code, 400)

    def test_bulk_no_ids(self):
        r = self._bulk('change_prix', [], {'mode': 'percent', 'valeur': '1'})
        self.assertEqual(r.status_code, 400)

    # ── T8 : audit logging ──
    def test_bulk_creates_audit_log(self):
        from apps.stock.models import ProduitAuditLog
        self._bulk('change_prix', [self.p1.id], {'mode': 'percent', 'valeur': '10'})
        log = ProduitAuditLog.objects.filter(
            produit=self.p1, action='change_prix').first()
        self.assertIsNotNone(log)
        self.assertEqual(log.champ, 'prix_vente')
        self.assertEqual(log.company, self.company)
        self.assertEqual(log.created_by, self.user)

    # ── T8 : un autre tenant ne voit pas / ne touche pas mes produits ──
    def test_bulk_cross_tenant_user_cannot_touch(self):
        from rest_framework.test import APIClient
        from rest_framework_simplejwt.tokens import AccessToken
        from django.contrib.auth import get_user_model
        User = get_user_model()
        intruder = User.objects.create_user(
            username='intruder', password='x',
            company=self.other, role_legacy='responsable')
        cli = APIClient()
        cli.credentials(HTTP_AUTHORIZATION=f'Bearer {AccessToken.for_user(intruder)}')
        r = cli.post('/api/django/stock/produits/bulk/',
                     {'action': 'change_prix', 'ids': [self.p1.id],
                      'params': {'mode': 'percent', 'valeur': '99'}},
                     format='json')
        # Aucun produit de l'intrus ne correspond → 400, p1 intact.
        self.assertEqual(r.status_code, 400)
        self.p1.refresh_from_db()
        self.assertEqual(self.p1.prix_vente, Decimal('1000.00'))

    # ── T4 : édition en ligne, PATCH d'un seul champ, validé serveur ──
    def test_inline_patch_prix_vente(self):
        r = self.client.patch(
            f'/api/django/stock/produits/{self.p1.id}/',
            {'prix_vente': '1234.50'}, format='json')
        self.assertEqual(r.status_code, 200)
        self.p1.refresh_from_db()
        self.assertEqual(self.p1.prix_vente, Decimal('1234.50'))
        self.assertEqual(self.p1.prix_achat, Decimal('700.00'))  # inchangé

    def test_inline_patch_quantite(self):
        r = self.client.patch(
            f'/api/django/stock/produits/{self.p1.id}/',
            {'quantite_stock': 42}, format='json')
        self.assertEqual(r.status_code, 200)
        self.p1.refresh_from_db()
        self.assertEqual(self.p1.quantite_stock, 42)

    def test_inline_patch_categorie(self):
        r = self.client.patch(
            f'/api/django/stock/produits/{self.p1.id}/',
            {'categorie_id': self.cat.id}, format='json')
        self.assertEqual(r.status_code, 200)
        self.p1.refresh_from_db()
        self.assertEqual(self.p1.categorie_id, self.cat.id)

    def test_inline_patch_invalid_prix_rejected(self):
        r = self.client.patch(
            f'/api/django/stock/produits/{self.p1.id}/',
            {'prix_vente': 'pas-un-nombre'}, format='json')
        self.assertEqual(r.status_code, 400)
        self.p1.refresh_from_db()
        self.assertEqual(self.p1.prix_vente, Decimal('1000.00'))

    def test_inline_patch_foreign_categorie_rejected(self):
        r = self.client.patch(
            f'/api/django/stock/produits/{self.p1.id}/',
            {'categorie_id': self.cat_other.id}, format='json')
        self.assertEqual(r.status_code, 400)

    def test_inline_patch_cross_tenant_product_404(self):
        r = self.client.patch(
            f'/api/django/stock/produits/{self.foreign.id}/',
            {'prix_vente': '1.00'}, format='json')
        self.assertEqual(r.status_code, 404)
        self.foreign.refresh_from_db()
        self.assertEqual(self.foreign.prix_vente, Decimal('500.00'))
