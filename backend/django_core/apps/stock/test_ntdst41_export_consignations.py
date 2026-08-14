"""NTDST41 — export XLSX du relevé de consignation multi-clients.

Critère d'acceptation testé : le total « restant » de l'export correspond
EXACTEMENT à la somme des ``quantite_restante`` des dépôts actifs au moment de
l'export.

Run :
    python manage.py test apps.stock.test_ntdst41_export_consignations -v 2
"""
import datetime
import io
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import TestCase
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import AccessToken

from apps.stock.models import DepotConsignation, Produit
from apps.stock.services_consignation import (
    creer_depot_consignation, declarer_consommation,
)

User = get_user_model()

JOUR = datetime.date(2026, 3, 1)
CONSO = datetime.date(2026, 3, 20)
URL = '/api/django/stock/consignations/export-xlsx/'


def make_company(slug, nom):
    from authentication.models import Company
    company, _ = Company.objects.get_or_create(slug=slug, defaults={'nom': nom})
    return company


def auth(user):
    api = APIClient()
    api.credentials(HTTP_AUTHORIZATION=f'Bearer {AccessToken.for_user(user)}')
    return api


def _feuille(reponse):
    from openpyxl import load_workbook
    return load_workbook(io.BytesIO(reponse.content)).active


class Ntdst41Base(TestCase):
    def setUp(self):
        from apps.crm.models import Client

        self.company = make_company('ntdst41-co', 'NTDST41 Co')
        self.autre = make_company('ntdst41-autre', 'NTDST41 Autre')
        self.admin = User.objects.create_user(
            username='ntdst41_admin', password='x', role_legacy='admin',
            company=self.company)
        self.normal = User.objects.create_user(
            username='ntdst41_normal', password='x', role_legacy='normal',
            company=self.company)
        self.client_a = Client.objects.create(
            company=self.company, nom='Client A NTDST41')
        self.client_b = Client.objects.create(
            company=self.company, nom='Client B NTDST41')
        self.produit = Produit.objects.create(
            company=self.company, nom='Panneau 550 Wc', sku='PAN-NTDST41',
            prix_achat=Decimal('900'), prix_vente=Decimal('1200'),
            quantite_stock=500)

        self.depot_a = creer_depot_consignation(
            company=self.company, user=self.admin,
            client_id=self.client_a.id, produit_id=self.produit.id,
            quantite=40, date_depot=JOUR, adresse_site='Site A')
        self.depot_b = creer_depot_consignation(
            company=self.company, user=self.admin,
            client_id=self.client_b.id, produit_id=self.produit.id,
            quantite=25, date_depot=JOUR, adresse_site='Site B')
        declarer_consommation(depot=self.depot_a, user=self.admin,
                              quantite=15, date_declaration=CONSO)
        self.depot_a.refresh_from_db()
        self.depot_b.refresh_from_db()


class Ntdst41ExportTests(Ntdst41Base):
    def test_le_total_correspond_a_la_somme_des_restants(self):
        attendu = (self.depot_a.quantite_restante
                   + self.depot_b.quantite_restante)
        self.assertEqual(attendu, 50)  # (40 − 15) + 25

        res = auth(self.admin).get(URL)
        self.assertEqual(res.status_code, 200)
        feuille = _feuille(res)
        derniere = list(feuille.iter_rows(values_only=True))[-1]
        self.assertEqual(derniere[0], 'TOTAL')
        self.assertEqual(derniere[6], attendu)

    def test_lexport_porte_une_ligne_par_depot_plus_le_total(self):
        feuille = _feuille(auth(self.admin).get(URL))
        lignes = list(feuille.iter_rows(values_only=True))
        # en-tête + 2 dépôts + total
        self.assertEqual(len(lignes), 4)
        self.assertEqual(lignes[0][0], 'Client')

    def test_le_filtre_statut_actif_exclut_les_depots_clos(self):
        declarer_consommation(depot=self.depot_b, user=self.admin,
                              quantite=25, date_declaration=CONSO)
        self.depot_b.refresh_from_db()
        self.assertEqual(self.depot_b.statut, DepotConsignation.Statut.CLOS)

        feuille = _feuille(auth(self.admin).get(URL, {'statut': 'actif'}))
        lignes = list(feuille.iter_rows(values_only=True))
        self.assertEqual(len(lignes), 3)  # en-tête + dépôt A + total
        self.assertEqual(lignes[-1][6], self.depot_a.quantite_restante)

    def test_lexport_ne_contient_aucun_depot_dune_autre_societe(self):
        from apps.crm.models import Client

        voisin = Client.objects.create(company=self.autre, nom='Voisin')
        autre_produit = Produit.objects.create(
            company=self.autre, nom='Voisin', sku='VOISIN-DST41',
            prix_achat=Decimal('1'), prix_vente=Decimal('2'),
            quantite_stock=99)
        creer_depot_consignation(
            company=self.autre, user=self.admin, client_id=voisin.id,
            produit_id=autre_produit.id, quantite=9, date_depot=JOUR)

        feuille = _feuille(auth(self.admin).get(URL))
        contenu = str(list(feuille.iter_rows(values_only=True)))
        self.assertNotIn('VOISIN-DST41', contenu)

    def test_export_refuse_a_un_role_normal(self):
        self.assertEqual(auth(self.normal).get(URL).status_code, 403)

    def test_export_refuse_lanonyme(self):
        self.assertEqual(APIClient().get(URL).status_code, 401)
