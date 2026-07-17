"""Vues de la Comptabilité générale (toutes scopées société, admin-gated).

La compta est INTERNE : aucune donnée n'est exposée côté client. L'accès est
réservé au palier Administrateur/Responsable (``IsResponsableOrAdmin``).
Les viewsets filtrent par ``request.user.company`` (TenantMixin) et posent la
société côté serveur ; aucun prix d'achat ni marge n'apparaît ici.
"""
import csv
import io

from django.core.exceptions import ValidationError as DjangoValidationError
from django.db import IntegrityError
from django.http import HttpResponse

from rest_framework import filters, generics, status, viewsets
from rest_framework.decorators import (
    action, api_view, permission_classes, throttle_classes,
)
from rest_framework.exceptions import ValidationError
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.throttling import SimpleRateThrottle
from rest_framework.views import APIView

from django.utils import timezone

from authentication.mixins import TenantMixin
from authentication.permissions import HasPermissionOrLegacy, IsResponsableOrAdmin

from . import selectors, services
from .models import (
    AppelTelephonique,
    BaremeIndemnite, BordereauRemise, Budget, BudgetLigne, Caisse,
    Campagne, CautionBancaire, CentreCout, CessionImmobilisation, CodePromotion,
    CommissionPayoutRun, EnvoiCampagne,
    CompteComptable, CompteTresorerie, ContratAvancement, DeclarationTVA,
    DemandeApprobationConfig,
    DotationAmortissement, ECatalogue, EcritureComptable, Effet,
    EntiteConsolidation, EtapeSequence, InscriptionSequence,
    ListeDiffusion, AbonnementListe, SegmentMarketing,
    ExerciceComptable, FormulaireIntake, Immobilisation, IndemniteChantier,
    Journal,
    LignePrevisionnelTresorerie, LigneReleve, MessageWhatsAppEntrant,
    ModeleDevis, MouvementCaisse, NoteFrais, OuverturePartage, PostSocial,
    PaymentRun, PeriodeComptable, PlafondNoteFrais,
    PlanComptable, PouvoirBancaire, Provision,
    ProvisionCreance, RapportNoteFrais,
    Rapprochement, RapprochementBancaire, RelanceDevisAbandonne,
    RetenueGarantie, RetenueSource, SequenceRelance, SessionGuidedSelling,
    TimbreFiscal, TravauxEnCours, VirementInterne,
    DocumentProposition, SimulationPublique, SimulationFinancement,
    OffreFinancement, LigneIncitation, EcheancierPaiement, TranchePaiement,
    AppelOffre, BordereauPrix, LigneBordereau, CautionSoumission,
    DossierSoumission, PieceSoumission, EcheanceAO, ResultatAO,
    ComptePortailClient, AcceptationDevisPortail, PaiementFacturePortail,
    DocumentClientPortail, JalonChantierPortail, DemandeTicketPortail,
    Partenaire, SoumissionLeadPartenaire, CommissionPartenaire,
    TerritoireCommercial, EnqueteNPS, AvisClient,
    CompteFidelite, MouvementFidelite, RegleUpsell,
    AbonnementMonitoring,
    MappingCompte, CompteAuxiliaire, PieceJustificative,
    PisteAuditComptable,
    ModeleRapprochement,
    ObligationFiscale,
    FamilleTvaNonDeductible,
    Compensation,
    ApprobationEnvoiCampagne,
    Enquete, ReponseEnquete,
    EvenementMarketing, InscriptionEvenement,
    SupportOffline,
    DomaineEnvoi,
    TypeEvenement,
    BilletEvenement,
    QuestionEvenement,
    CommunicationEvenement,
)
from .serializers import (
    AppelTelephoniqueSerializer, AvancementRevenuSerializer,
    BaremeIndemniteSerializer,
    BordereauRemiseSerializer, BudgetSerializer, CaisseSerializer,
    CampagneSerializer, CautionBancaireSerializer, CentreCoutSerializer,
    CessionImmobilisationSerializer, ClotureCaisseSerializer,
    CodePromotionSerializer, EnvoiCampagneSerializer,
    ApprobationEnvoiCampagneSerializer,
    EnqueteSerializer,
    EvenementMarketingSerializer, InscriptionEvenementSerializer,
    SupportOfflineSerializer,
    DomaineEnvoiSerializer,
    TypeEvenementSerializer,
    BilletEvenementSerializer,
    QuestionEvenementSerializer,
    CommunicationEvenementSerializer,
    PostSocialSerializer,
    CommissionPayoutRunSerializer, CompteComptableSerializer,
    CompteTresorerieSerializer, ContratAvancementSerializer,
    DeclarationTVASerializer, DemandeApprobationConfigSerializer,
    DotationAmortissementSerializer, ECatalogueSerializer,
    EcritureComptableSerializer, EffetSerializer, EntiteConsolidationSerializer,
    EtapeSequenceSerializer, InscriptionSequenceSerializer,
    ListeDiffusionSerializer, AbonnementListeSerializer,
    SegmentMarketingSerializer,
    ExerciceComptableSerializer, FormulaireIntakeSerializer,
    ImmobilisationSerializer,
    IndemniteChantierSerializer, JournalSerializer,
    LignePrevisionnelTresorerieSerializer, LigneReleveSerializer,
    MessageWhatsAppEntrantSerializer, ModeleDevisSerializer,
    MouvementCaisseSerializer, NoteFraisSerializer, OuverturePartageSerializer,
    ParametresTresorerieSerializer, PaymentRunSerializer,
    PouvoirBancaireSerializer,
    PeriodeComptableSerializer, PlanAmortissementSerializer,
    PlafondNoteFraisSerializer,
    PlanComptableSerializer, ProvisionSerializer, ProvisionCreanceSerializer,
    RapportNoteFraisSerializer,
    RapprochementBancaireSerializer, RapprochementSerializer,
    RelanceDevisAbandonneSerializer,
    RetenueGarantieSerializer, RetenueSourceSerializer,
    SequenceRelanceSerializer, SessionGuidedSellingSerializer,
    TimbreFiscalSerializer,
    TravauxEnCoursSerializer, VirementInterneSerializer,
    DocumentPropositionSerializer, SimulationPubliqueSerializer,
    SimulationFinancementSerializer, OffreFinancementSerializer,
    LigneIncitationSerializer, EcheancierPaiementSerializer,
    TranchePaiementSerializer, AppelOffreSerializer, BordereauPrixSerializer,
    LigneBordereauSerializer, CautionSoumissionSerializer,
    DossierSoumissionSerializer, PieceSoumissionSerializer,
    EcheanceAOSerializer, ResultatAOSerializer, ComptePortailClientSerializer,
    AcceptationDevisPortailSerializer, PaiementFacturePortailSerializer,
    DocumentClientPortailSerializer, JalonChantierPortailSerializer,
    DemandeTicketPortailSerializer,
    PartenaireSerializer, SoumissionLeadPartenaireSerializer,
    CommissionPartenaireSerializer, TerritoireCommercialSerializer,
    EnqueteNPSSerializer, AvisClientSerializer,
    CompteFideliteSerializer, MouvementFideliteSerializer,
    RegleUpsellSerializer, AbonnementMonitoringSerializer,
    MappingCompteSerializer, CompteAuxiliaireSerializer,
    PieceJustificativeSerializer,
    PisteAuditComptableSerializer,
    ModeleRapprochementSerializer,
    ObligationFiscaleSerializer,
    FamilleTvaNonDeductibleSerializer,
    CompensationSerializer,
)


class _ComptaBaseViewSet(TenantMixin, viewsets.ModelViewSet):
    """Base : société scopée + accès Administrateur/Responsable uniquement."""
    permission_classes = [IsResponsableOrAdmin]


def _peut(user, code):
    """COMPTA40 — L'utilisateur porte-t-il la permission comptable ``code`` ?

    Repli historique (comme ``HasPermissionOrLegacy``) : un compte SANS rôle fin
    garde l'accès Responsable/Admin d'avant — aucune régression pour les comptes
    hérités. Un compte AVEC rôle fin est jugé sur ses permissions granulaires.
    """
    if not (user and user.is_authenticated):
        return False
    if user.is_superuser:
        return True
    if getattr(user, 'role_id', None):
        return user.has_erp_permission(code)
    return user.is_responsable


class PlanComptableViewSet(_ComptaBaseViewSet):
    """Plan(s) comptable(s) de la société (FG107). Action ``seed`` pour amorcer
    le plan CGNC + les journaux standards (idempotent)."""
    queryset = PlanComptable.objects.all()
    serializer_class = PlanComptableSerializer

    @action(detail=False, methods=['post'])
    def seed(self, request):
        company = request.user.company
        plan = services.seed_plan_comptable(company)
        services.seed_journaux(company)
        return Response(PlanComptableSerializer(plan).data)


class CompteComptableViewSet(_ComptaBaseViewSet):
    """Comptes du plan comptable (FG107). Recherche par numéro/intitulé."""
    queryset = CompteComptable.objects.select_related('plan').all()
    serializer_class = CompteComptableSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['numero', 'intitule']
    ordering_fields = ['numero', 'classe']

    def get_queryset(self):
        qs = super().get_queryset()
        classe = self.request.query_params.get('classe')
        if classe:
            qs = qs.filter(classe=classe)
        return qs


class JournalViewSet(_ComptaBaseViewSet):
    """Journaux comptables (FG108)."""
    queryset = Journal.objects.select_related('compte_contrepartie').all()
    serializer_class = JournalSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['code']


class EcritureComptableViewSet(_ComptaBaseViewSet):
    """Écritures en partie double (FG108). Filtrable par journal/période."""
    queryset = EcritureComptable.objects.select_related('journal').prefetch_related(
        'lignes__compte').all()
    serializer_class = EcritureComptableSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_ecriture', 'id']

    def get_queryset(self):
        qs = super().get_queryset()
        params = self.request.query_params
        journal = params.get('journal')
        if journal:
            qs = qs.filter(journal_id=journal)
        date_debut = params.get('date_debut')
        if date_debut:
            qs = qs.filter(date_ecriture__gte=date_debut)
        date_fin = params.get('date_fin')
        if date_fin:
            qs = qs.filter(date_ecriture__lte=date_fin)
        return qs

    def perform_create(self, serializer):
        serializer.save(
            company=self.request.user.company, created_by=self.request.user)

    @action(detail=True, methods=['post'])
    def extourner(self, request, pk=None):
        """COMPTA11 — Passe l'écriture d'extourne (contre-passation).

        Ne supprime JAMAIS l'écriture d'origine : crée l'écriture inverse.
        Idempotent. Corps optionnel : ``{'date_extourne': 'YYYY-MM-DD'}``.
        """
        ecriture = self.get_object()
        date_extourne = request.data.get('date_extourne') or None
        try:
            extourne = services.extourner_ecriture(
                ecriture, date_extourne=date_extourne, user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(
            EcritureComptableSerializer(extourne).data,
            status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def valider(self, request, pk=None):
        """COMPTA40 — Valide l'écriture (second regard, séparation des tâches).

        Le saisisseur (``created_by``) ne peut JAMAIS valider sa propre
        écriture : la garde est posée côté service. Requiert la permission
        ``compta_valider`` (repli historique : Responsable/Admin pour les
        comptes sans rôle fin). En cas de violation de la séparation ou d'une
        écriture déjà validée, renvoie 400 avec un message explicite.
        """
        if not _peut(request.user, 'compta_valider'):
            return Response(
                {'detail': "Vous n'êtes pas habilité à valider une écriture."},
                status=status.HTTP_403_FORBIDDEN)
        ecriture = self.get_object()
        try:
            services.valider_ecriture(ecriture, user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(EcritureComptableSerializer(ecriture).data)


class CompteTresorerieViewSet(_ComptaBaseViewSet):
    """Référentiel des comptes bancaires & caisses (FG121)."""
    queryset = CompteTresorerie.objects.select_related('compte_comptable').all()
    serializer_class = CompteTresorerieSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['type_compte', 'libelle']

    @action(detail=True, methods=['get'])
    def solde(self, request, pk=None):
        """Solde courant du compte de trésorerie (solde initial + GL)."""
        treso = self.get_object()
        from decimal import Decimal
        mouvements = selectors.solde_compte(
            request.user.company, treso.compte_comptable)
        total = Decimal(treso.solde_initial) + mouvements
        return Response({
            'solde_initial': treso.solde_initial,
            'mouvements': mouvements,
            'solde': total,
        })


class EtatsComptablesViewSet(viewsets.ViewSet):
    """États de synthèse en LECTURE SEULE (FG110-114) : grand livre, balance,
    CPC, bilan. Admin/Responsable uniquement, scopés société côté selector."""
    permission_classes = [IsResponsableOrAdmin]

    def get_permissions(self):
        # YRBAC13 — toutes les actions de ce viewset sont des rapports en
        # LECTURE SEULE (GET) ; garde explicite par action pour le scanner
        # YRBAC4, comportement STRICTEMENT inchangé (IsResponsableOrAdmin).
        return [IsResponsableOrAdmin()]

    def _periode(self, request):
        params = request.query_params
        return {
            'date_debut': params.get('date_debut') or None,
            'date_fin': params.get('date_fin') or None,
            'validees_seulement': params.get('validees') == '1',
        }

    def _company_profile(self, request):
        """ZACC1 — ``CompanyProfile`` (entête PDF) ; ``None`` si absent —
        JAMAIS bloquant (profil optionnel, foundation app exemptée)."""
        try:
            from apps.parametres.models_company import CompanyProfile
            return CompanyProfile.get(company=request.user.company)
        except Exception:  # pragma: no cover - profil optionnel.
            return None

    def _pdf_response(self, pdf_bytes, filename):
        resp = HttpResponse(pdf_bytes, content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp

    def _pdf_or_503(self, render_fn):
        """ZACC1 — invoque un ``render_*_pdf`` de ``pdf_etats``, ou renvoie un
        503 explicite si WeasyPrint est indisponible (jamais un crash)."""
        try:
            return render_fn()
        except RuntimeError as exc:
            return Response(
                {'detail': str(exc)}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

    def _comparatif_kwargs(self, request):
        """ZACC2 — kwargs du mode comparatif N-1 (``comparer=1`` ou bornes
        explicites ``date_debut_n1``/``date_fin_n1``). Défaut = comportement
        actuel intact (``comparer=False``)."""
        params = request.query_params
        comparer = (params.get('comparer') == '1'
                    or bool(params.get('date_debut_n1'))
                    or bool(params.get('date_fin_n1')))
        return {
            'comparer': comparer,
            'date_debut_n1': params.get('date_debut_n1') or None,
            'date_fin_n1': params.get('date_fin_n1') or None,
        }

    @action(detail=False, methods=['get'])
    def grand_livre(self, request):
        company = request.user.company
        periode = self._periode(request)
        compte = None
        numero = request.query_params.get('compte')
        if numero:
            compte = CompteComptable.objects.filter(
                company=company, numero=numero).first()
        data = selectors.grand_livre(company, compte=compte, **periode)
        if request.query_params.get('export') == 'pdf':
            from .pdf_etats import render_grand_livre_pdf
            result = self._pdf_or_503(lambda: render_grand_livre_pdf(
                data, self._company_profile(request),
                date_debut=periode['date_debut'], date_fin=periode['date_fin']))
            if isinstance(result, Response):
                return result
            return self._pdf_response(result, 'grand_livre.pdf')
        return Response(data)

    @action(detail=False, methods=['get'])
    def balance(self, request):
        periode = self._periode(request)
        data = selectors.balance_generale(
            request.user.company, **periode,
            **self._comparatif_kwargs(request))
        if request.query_params.get('export') == 'pdf':
            from .pdf_etats import render_balance_pdf
            result = self._pdf_or_503(lambda: render_balance_pdf(
                data, self._company_profile(request),
                date_debut=periode['date_debut'], date_fin=periode['date_fin']))
            if isinstance(result, Response):
                return result
            return self._pdf_response(result, 'balance.pdf')
        return Response(data)

    @action(detail=False, methods=['get'])
    def cpc(self, request):
        periode = self._periode(request)
        data = selectors.cpc(
            request.user.company, **periode,
            **self._comparatif_kwargs(request))
        if request.query_params.get('export') == 'pdf':
            from .pdf_etats import render_cpc_pdf
            result = self._pdf_or_503(lambda: render_cpc_pdf(
                data, self._company_profile(request),
                date_debut=periode['date_debut'], date_fin=periode['date_fin']))
            if isinstance(result, Response):
                return result
            return self._pdf_response(result, 'cpc.pdf')
        return Response(data)

    @action(detail=False, methods=['get'])
    def bilan(self, request):
        periode = self._periode(request)
        comparatif = self._comparatif_kwargs(request)
        data = selectors.bilan(
            request.user.company, date_fin=periode['date_fin'],
            validees_seulement=periode['validees_seulement'],
            comparer=comparatif['comparer'],
            date_fin_n1=comparatif['date_fin_n1'])
        if request.query_params.get('export') == 'pdf':
            from .pdf_etats import render_bilan_pdf
            result = self._pdf_or_503(lambda: render_bilan_pdf(
                data, self._company_profile(request),
                date_fin=periode['date_fin']))
            if isinstance(result, Response):
                return result
            return self._pdf_response(result, 'bilan.pdf')
        return Response(data)

    @action(detail=False, methods=['get'])
    def esg(self, request):
        """ESG — état des soldes de gestion (SIG) CGNC (COMPTA29).

        Cascade marge → valeur ajoutée → EBE → résultat courant → résultat net,
        déduite du grand livre (comptes de gestion classes 6 & 7). Paramètres :
        ``date_debut``/``date_fin`` (bornes), ``validees`` (1 → validées),
        ``comparer`` (1 → colonne N-1, ZACC2). Lecture seule, scopée société,
        Admin/Responsable.
        """
        data = selectors.esg(
            request.user.company, **self._periode(request),
            **self._comparatif_kwargs(request))
        return Response(data)

    @action(detail=False, methods=['get'])
    def etic(self, request):
        """ETIC — état des informations complémentaires CGNC (COMPTA29).

        Paquet des tableaux annexes (principes & méthodes, immobilisations,
        provisions, engagements hors-bilan, rappel du résultat) pour un exercice,
        assemblé sans recalcul depuis les sélecteurs existants. Paramètres :
        ``exercice`` (id, requis), ``validees`` (1 → validées). Lecture seule,
        scopée société, Admin/Responsable.
        """
        company = request.user.company
        exercice_id = request.query_params.get('exercice')
        if not exercice_id:
            return Response(
                {'detail': "Le paramètre 'exercice' est requis."},
                status=status.HTTP_400_BAD_REQUEST)
        exercice = ExerciceComptable.objects.filter(
            company=company, pk=exercice_id).first()
        if exercice is None:
            return Response(
                {'detail': 'Exercice introuvable pour cette société.'},
                status=status.HTTP_404_NOT_FOUND)
        data = selectors.etic(
            company, exercice,
            validees_seulement=request.query_params.get('validees') == '1')
        return Response(data)

    @action(detail=False, methods=['get'], url_path='position-tresorerie')
    def position_tresorerie(self, request):
        """Position de trésorerie consolidée + projection nette (FG122).

        Solde par compte/caisse + total (depuis les comptes de trésorerie et le
        grand livre), enrichi d'une projection nette indicative (AR/AP/paie/TVA).
        Lecture seule, scopée société, Admin/Responsable uniquement.
        """
        periode = self._periode(request)
        company = request.user.company
        position = selectors.position_tresorerie(
            company, date_fin=periode['date_fin'],
            validees_seulement=periode['validees_seulement'])
        projection = selectors.projection_tresorerie(
            company, date_fin=periode['date_fin'],
            validees_seulement=periode['validees_seulement'])
        return Response({
            'comptes': position['comptes'],
            'total': position['total'],
            'projection': projection,
        })

    @action(detail=False, methods=['get'], url_path='frais-bancaires')
    def frais_bancaires(self, request):
        """NTTRE9 — Analyse des frais bancaires par compte et par mois.

        Query ``?debut=YYYY-MM-DD&fin=YYYY-MM-DD``. Agrège les écritures GL des
        comptes de frais (réglage NTTRE28) par compte de trésorerie et par mois.
        Lecture seule, scopée société, Admin/Responsable.
        """
        params = request.query_params
        data = selectors.analyse_frais_bancaires(
            request.user.company,
            debut=params.get('debut') or None,
            fin=params.get('fin') or None)
        return Response(data)

    @action(detail=False, methods=['get'], url_path='previsionnel-tresorerie')
    def previsionnel_tresorerie(self, request):
        """Prévisionnel de trésorerie roulant 13 semaines (FG126).

        Empile les lignes prévues éditables (crédits, leasing, salaires, IS…) et
        les effets ouverts (FG127/FG128) au-dessus de la position actuelle, semaine
        par semaine. Paramètres : ``date_debut`` (défaut aujourd'hui),
        ``nb_semaines`` (défaut 13). Lecture seule, Admin/Responsable.
        """
        company = request.user.company
        nb = request.query_params.get('nb_semaines')
        try:
            nb_semaines = int(nb) if nb else 13
        except (TypeError, ValueError):
            nb_semaines = 13
        data = selectors.previsionnel_tresorerie(
            company,
            date_debut=request.query_params.get('date_debut') or None,
            nb_semaines=max(1, min(nb_semaines, 52)))
        return Response(data)

    @action(detail=False, methods=['get'], url_path='balance-agee-fournisseurs')
    def balance_agee_fournisseurs(self, request):
        """Balance âgée fournisseurs (FG132) : encours dû par fournisseur,
        bucketé 0–30 / 31–60 / 61–90 / 90+ jours depuis le grand livre (compte
        4411). Miroir AP de la balance âgée clients. Paramètres :
        ``date_reference`` (défaut aujourd'hui), ``validees`` (1 → écritures
        validées seulement). Lecture seule, scopée société, Admin/Responsable.
        """
        date_reference = request.query_params.get('date_reference') or None
        data = selectors.balance_agee_fournisseurs(
            request.user.company, date_reference=date_reference,
            validees_seulement=request.query_params.get('validees') == '1')
        if request.query_params.get('export') == 'pdf':
            from .pdf_etats import render_balance_agee_pdf
            result = self._pdf_or_503(lambda: render_balance_agee_pdf(
                data, self._company_profile(request),
                date_reference=date_reference,
                titre='Balance âgée fournisseurs'))
            if isinstance(result, Response):
                return result
            return self._pdf_response(result, 'balance_agee_fournisseurs.pdf')
        return Response(data)

    @action(detail=False, methods=['get'],
            url_path='releve-fournisseur/(?P<tiers_id>[0-9]+)')
    def releve_fournisseur(self, request, tiers_id=None):
        """Relevé de compte d'un fournisseur (FG132) : mouvements chronologiques
        du compte 4411 pour l'auxiliaire ``tiers_id`` + solde dû. Miroir AP du
        relevé client. Paramètres : ``date_debut`` / ``date_fin`` / ``validees``.
        Lecture seule, scopée société, Admin/Responsable.
        """
        periode = self._periode(request)
        data = selectors.releve_fournisseur(
            request.user.company, int(tiers_id),
            date_debut=periode['date_debut'], date_fin=periode['date_fin'],
            validees_seulement=periode['validees_seulement'])
        return Response(data)

    @action(detail=False, methods=['get'], url_path='releve-deductions-tva')
    def releve_deductions_tva(self, request):
        """Relevé de déductions détaillé — annexe TVA exigée par la DGI (FG138).

        Liste ligne par ligne chaque pièce ouvrant droit à déduction de TVA sur
        la période (date, référence/pièce, journal, tiers fournisseur, base HT,
        TVA, taux), déduite du grand livre (comptes 3455…). La somme des TVA
        reconcilie avec la TVA déductible de la déclaration (FG137). Paramètres :
        ``date_debut`` / ``date_fin`` (la période de la déclaration), ``validees``
        (1 → écritures validées seulement) et ``export=csv`` pour l'export CSV.
        Lecture seule, scopée société, Admin/Responsable.
        """
        periode = self._periode(request)
        company = request.user.company
        data = selectors.releve_deductions_tva(
            company, date_debut=periode['date_debut'],
            date_fin=periode['date_fin'],
            validees_seulement=periode['validees_seulement'])
        if request.query_params.get('export') == 'csv':
            return self._export_deductions_csv(data)
        return Response(data)

    @staticmethod
    def _export_deductions_csv(data):
        """Sérialise le relevé de déductions TVA (FG138) en CSV (DGI)."""
        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=';')
        writer.writerow(['Relevé de déductions de TVA (annexe DGI)'])
        writer.writerow(
            ['Période', f"{data['date_debut'] or ''} → {data['date_fin'] or ''}"])
        writer.writerow([])
        writer.writerow(
            ['Date', 'Référence', 'Journal', 'Libellé', 'Tiers',
             'Base HT', 'TVA', 'Taux %', 'Prorata'])
        for ligne in data['lignes']:
            taux = ligne['taux']
            writer.writerow([
                ligne['date'], ligne['reference'], ligne['journal'],
                ligne['libelle'], ligne['tiers'], ligne['base_ht'],
                ligne['tva'], '' if taux is None else taux,
                'Oui' if ligne.get('prorata_applique') else ''])
        writer.writerow([])
        writer.writerow(
            ['Totaux', '', '', '', '', data['totaux']['base_ht'],
             data['totaux']['tva'], '', ''])
        resp = HttpResponse(
            buffer.getvalue(), content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = (
            'attachment; filename="releve_deductions_tva_'
            f"{data['date_debut'] or 'periode'}_{data['date_fin'] or ''}.csv\"")
        return resp

    @action(detail=False, methods=['get'], url_path='declaration-honoraires')
    def declaration_honoraires(self, request):
        """Déclaration annuelle des honoraires — état 9421 (FG143).

        Agrège par bénéficiaire les paiements aux tiers (honoraires /
        prestations) de l'année civile, depuis les retenues à la source
        enregistrées (FG139). Chaque ligne porte l'identité fiscale du
        bénéficiaire (IF/ICE), le montant brut versé, la retenue à la source
        pratiquée et le net payé. Paramètres : ``annee`` (défaut année
        courante), ``type_prestation`` (filtre optionnel) et ``export=csv``
        pour le CSV (jamais ``?format=``). Lecture seule, scopée société,
        Admin/Responsable.
        """
        from django.utils import timezone
        params = request.query_params
        annee_param = params.get('annee')
        try:
            annee = int(annee_param) if annee_param else timezone.now().year
        except (TypeError, ValueError):
            return Response(
                {'detail': "Le paramètre 'annee' doit être une année valide."},
                status=status.HTTP_400_BAD_REQUEST)
        data = selectors.declaration_honoraires(
            request.user.company, annee,
            type_prestation=params.get('type_prestation') or None)
        if params.get('export') == 'csv':
            return self._export_declaration_honoraires_csv(data)
        return Response(data)

    @staticmethod
    def _export_declaration_honoraires_csv(data):
        """Sérialise la déclaration des honoraires (état 9421, FG143) en CSV."""
        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=';')
        writer.writerow(['Déclaration des honoraires (état 9421)'])
        writer.writerow(['Année', data['annee']])
        writer.writerow([])
        writer.writerow(
            ['Bénéficiaire', 'Identifiant fiscal (IF/ICE)', 'Nb pièces',
             'Montant brut', 'Retenue à la source', 'Net payé'])
        for ligne in data['lignes']:
            writer.writerow([
                ligne['tiers_nom'], ligne['identifiant_fiscal'],
                ligne['nb_pieces'], ligne['brut'], ligne['retenue'],
                ligne['net']])
        writer.writerow([])
        writer.writerow(
            ['Totaux', '', data['totaux']['nb_pieces'],
             data['totaux']['brut'], data['totaux']['retenue'],
             data['totaux']['net']])
        resp = HttpResponse(
            buffer.getvalue(), content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = (
            'attachment; filename="declaration_honoraires_9421_'
            f"{data['annee']}.csv\"")
        return resp

    @action(detail=False, methods=['get'], url_path='aide-is')
    def aide_is(self, request):
        """Aide au calcul de l'IS — estimation + acomptes + régularisation (FG140).

        Estime l'IS dû d'un exercice depuis le CPC (résultat fiscal × barème
        progressif marocain, plancher de la cotisation minimale), l'échéancier
        des 4 acomptes provisionnels (25 % de l'IS de référence chacun, échus
        aux 3e/6e/9e/12e mois) et la régularisation (IS dû − acomptes). C'est
        une aide indicative et lecture seule (aucune écriture). Paramètres :
        ``exercice`` (id, requis), ``reintegrations`` / ``deductions``
        (ajustements extra-comptables, défaut 0), ``is_reference`` (IS N-1 pour
        les acomptes ; défaut = IS estimé courant), ``validees`` (1 → écritures
        validées seulement) et ``export=csv``. Scopée société, Admin/Responsable.
        """
        company = request.user.company
        params = request.query_params
        exercice_id = params.get('exercice')
        if not exercice_id:
            return Response(
                {'detail': "Le paramètre 'exercice' est requis."},
                status=status.HTTP_400_BAD_REQUEST)
        exercice = ExerciceComptable.objects.filter(
            company=company, pk=exercice_id).first()
        if exercice is None:
            return Response(
                {'detail': 'Exercice introuvable pour cette société.'},
                status=status.HTTP_404_NOT_FOUND)
        kwargs = {
            'reintegrations': self._decimal_param(params, 'reintegrations'),
            'deductions': self._decimal_param(params, 'deductions'),
            'is_reference': self._decimal_param(params, 'is_reference'),
            'validees_seulement': params.get('validees') == '1',
        }
        data = selectors.aide_calcul_is(company, exercice, **kwargs)
        if params.get('export') == 'csv':
            return self._export_aide_is_csv(exercice, data)
        return Response(data)

    @staticmethod
    def _decimal_param(params, key):
        """Parse un paramètre décimal optionnel (None si absent/invalide)."""
        from decimal import Decimal, InvalidOperation
        raw = params.get(key)
        if raw in (None, ''):
            return None
        try:
            return Decimal(str(raw))
        except (InvalidOperation, TypeError, ValueError):
            return None

    @staticmethod
    def _export_aide_is_csv(exercice, data):
        """Sérialise l'aide au calcul de l'IS (FG140) en CSV."""
        estimation = data['estimation']
        echeancier = data['echeancier_acomptes']
        regul = data['regularisation']
        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=';')
        writer.writerow(['Aide au calcul de l\'IS'])
        writer.writerow(
            ['Exercice', f"{exercice.date_debut} → {exercice.date_fin}"])
        writer.writerow([])
        writer.writerow(['Résultat comptable', estimation['resultat_comptable']])
        writer.writerow(['Réintégrations', estimation['reintegrations']])
        writer.writerow(['Déductions', estimation['deductions']])
        writer.writerow(['Résultat fiscal', estimation['resultat_fiscal']])
        writer.writerow(['IS au barème', estimation['bareme']['is_bareme']])
        writer.writerow(
            ['Cotisation minimale', estimation['cotisation_minimale']['cm']])
        writer.writerow(
            ['IS dû', estimation['is_du'], f"(base : {estimation['base_retenue']})"])
        writer.writerow([])
        writer.writerow(['Acomptes provisionnels', 'Échéance', 'Montant'])
        for acompte in echeancier['acomptes']:
            writer.writerow(
                [f"Acompte {acompte['numero']}", acompte['date_echeance'],
                 acompte['montant']])
        writer.writerow(['Total acomptes', '', echeancier['total_acomptes']])
        writer.writerow([])
        writer.writerow(
            ['Régularisation', regul['regularisation'], regul['sens']])
        writer.writerow(
            ['Date limite de paiement', regul['date_limite_paiement']])
        resp = HttpResponse(
            buffer.getvalue(), content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = (
            f'attachment; filename="aide_is_exercice_{exercice.pk}.csv"')
        return resp

    @action(detail=False, methods=['get'], url_path='export-fec')
    def export_fec(self, request):
        """Export FEC — Fichier des Écritures Comptables au format DGI (FG141).

        Restitue, pour un exercice, l'ensemble ORDONNÉ des écritures du grand
        livre (une ligne par ``LigneEcriture``, triée par date d'écriture puis
        numéro de pièce) au format auditable normalisé. Paramètres : ``exercice``
        (id, requis), ``validees`` (1 → écritures validées seulement) et
        ``export`` qui choisit le rendu — ``export=fec`` (texte tabulé, le
        format DGI), ``export=csv`` (point-virgule) ; sans ``export``, renvoie le
        JSON (colonnes + lignes + totaux). On utilise ``export=`` et JAMAIS le
        ``format=`` de DRF (qui répond 404). Lecture seule, scopée société,
        Admin/Responsable.
        """
        company = request.user.company
        exercice_id = request.query_params.get('exercice')
        if not exercice_id:
            return Response(
                {'detail': "Le paramètre 'exercice' est requis."},
                status=status.HTTP_400_BAD_REQUEST)
        exercice = ExerciceComptable.objects.filter(
            company=company, pk=exercice_id).first()
        if exercice is None:
            return Response(
                {'detail': 'Exercice introuvable pour cette société.'},
                status=status.HTTP_404_NOT_FOUND)
        data = selectors.export_fec(
            company, exercice,
            validees_seulement=request.query_params.get('validees') == '1')
        export = request.query_params.get('export')
        if export in ('fec', 'csv'):
            return self._export_fec_file(exercice, data, export)
        return Response(data)

    @staticmethod
    def _export_fec_file(exercice, data, export):
        """Sérialise le FEC (FG141) en fichier délimité (tabulé DGI ou CSV)."""
        # 'fec' → tabulation (format DGI officiel) ; 'csv' → point-virgule.
        delimiter = '\t' if export == 'fec' else ';'
        extension = 'txt' if export == 'fec' else 'csv'
        buffer = io.StringIO()
        writer = csv.writer(
            buffer, delimiter=delimiter, lineterminator='\r\n')
        writer.writerow(data['columns'])
        for ligne in data['lignes']:
            writer.writerow([ligne[col] for col in data['columns']])
        resp = HttpResponse(
            buffer.getvalue(),
            content_type=(
                'text/plain; charset=utf-8' if export == 'fec'
                else 'text/csv; charset=utf-8'))
        resp['Content-Disposition'] = (
            'attachment; filename='
            f'"FEC_exercice_{exercice.pk}_{exercice.date_debut}'
            f'_{exercice.date_fin}.{extension}"')
        return resp

    @action(detail=False, methods=['get'], url_path='liasse-fiscale')
    def liasse_fiscale(self, request):
        """Trousse liasse fiscale — états de synthèse en un paquet (FG142).

        Assemble en un seul appel, pour un exercice, les états de synthèse déjà
        produits par les sélecteurs existants (bilan + CPC + balance + annexe TVA
        des déductions), sans rien recalculer. C'est le paquet remis au
        fiduciaire / à la DGI. Paramètres : ``exercice`` (id, requis), ``validees``
        (1 → écritures validées seulement) et ``export`` qui choisit le rendu —
        ``export=csv`` produit un CSV multi-sections (une section après l'autre :
        bilan, CPC, balance, annexe TVA) ; sans ``export``, renvoie le JSON
        structuré. On utilise ``export=`` et JAMAIS le ``format=`` de DRF (qui
        répond 404). Lecture seule, scopée société, Admin/Responsable.
        """
        company = request.user.company
        exercice_id = request.query_params.get('exercice')
        if not exercice_id:
            return Response(
                {'detail': "Le paramètre 'exercice' est requis."},
                status=status.HTTP_400_BAD_REQUEST)
        exercice = ExerciceComptable.objects.filter(
            company=company, pk=exercice_id).first()
        if exercice is None:
            return Response(
                {'detail': 'Exercice introuvable pour cette société.'},
                status=status.HTTP_404_NOT_FOUND)
        data = selectors.liasse_fiscale(
            company, exercice,
            validees_seulement=request.query_params.get('validees') == '1')
        export = request.query_params.get('export')
        if export == 'csv':
            return self._export_liasse_csv(exercice, data)
        if export == 'pdf':
            from .pdf_etats import render_liasse_pdf
            result = self._pdf_or_503(lambda: render_liasse_pdf(
                data, self._company_profile(request), exercice=exercice))
            if isinstance(result, Response):
                return result
            return self._pdf_response(
                result, f'liasse_fiscale_exercice_{exercice.pk}.pdf')
        return Response(data)

    @staticmethod
    def _export_liasse_csv(exercice, data):
        """Sérialise la liasse fiscale (FG142) en CSV multi-sections (DGI)."""
        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=';')
        writer.writerow(['Liasse fiscale — États de synthèse'])
        writer.writerow(
            ['Exercice', f"{data['date_debut']} → {data['date_fin']}"])
        writer.writerow(['Résultat de l\'exercice', data['resultat']])
        writer.writerow(['Équilibré', 'oui' if data['equilibre'] else 'non'])
        writer.writerow([])

        # Section BILAN (actif / passif + résultat).
        bilan = data['bilan']
        writer.writerow(['BILAN'])
        writer.writerow(['Actif', 'Numéro', 'Intitulé', 'Montant'])
        for poste in bilan['actif']:
            writer.writerow(
                ['', poste['numero'], poste['intitule'], poste['montant']])
        writer.writerow(['Total actif', '', '', bilan['total_actif']])
        writer.writerow(['Passif', 'Numéro', 'Intitulé', 'Montant'])
        for poste in bilan['passif']:
            writer.writerow(
                ['', poste['numero'], poste['intitule'], poste['montant']])
        writer.writerow(['Total passif', '', '', bilan['total_passif']])
        writer.writerow(['Résultat (porté au passif)', '', '', bilan['resultat']])
        writer.writerow([])

        # Section CPC (produits / charges + résultat).
        cpc = data['cpc']
        writer.writerow(['CPC (Compte de Produits et Charges)'])
        writer.writerow(['Produits', 'Numéro', 'Intitulé', 'Montant'])
        for poste in cpc['produits']:
            writer.writerow(
                ['', poste['numero'], poste['intitule'], poste['montant']])
        writer.writerow(['Total produits', '', '', cpc['total_produits']])
        writer.writerow(['Charges', 'Numéro', 'Intitulé', 'Montant'])
        for poste in cpc['charges']:
            writer.writerow(
                ['', poste['numero'], poste['intitule'], poste['montant']])
        writer.writerow(['Total charges', '', '', cpc['total_charges']])
        writer.writerow(['Résultat', '', '', cpc['resultat']])
        writer.writerow([])

        # Section BALANCE générale (trial balance).
        balance = data['balance']
        writer.writerow(['BALANCE GÉNÉRALE'])
        writer.writerow(
            ['Numéro', 'Intitulé', 'Débit', 'Crédit', 'Solde débiteur',
             'Solde créditeur'])
        for ligne in balance['lignes']:
            writer.writerow([
                ligne['numero'], ligne['intitule'], ligne['debit'],
                ligne['credit'], ligne['solde_debiteur'],
                ligne['solde_crediteur']])
        writer.writerow(
            ['Totaux', '', balance['total_debit'], balance['total_credit'],
             '', ''])
        writer.writerow([])

        # Section ANNEXE TVA (relevé de déductions, FG138).
        annexe = data['annexe_tva']
        writer.writerow(['ANNEXE — Relevé de déductions de TVA'])
        writer.writerow(
            ['Date', 'Référence', 'Journal', 'Libellé', 'Tiers', 'Base HT',
             'TVA', 'Taux %'])
        for ligne in annexe['lignes']:
            taux = ligne['taux']
            writer.writerow([
                ligne['date'], ligne['reference'], ligne['journal'],
                ligne['libelle'], ligne['tiers'], ligne['base_ht'],
                ligne['tva'], '' if taux is None else taux])
        writer.writerow(
            ['Totaux', '', '', '', '', annexe['totaux']['base_ht'],
             annexe['totaux']['tva'], ''])

        resp = HttpResponse(
            buffer.getvalue(), content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = (
            'attachment; filename='
            f'"liasse_fiscale_exercice_{exercice.pk}_{exercice.date_debut}'
            f'_{exercice.date_fin}.csv"')
        return resp

    def _resolve_exercice(self, request):
        """Résout ``?exercice=<id>`` scopé société : (exercice, erreur_response
        ou None). Réutilisé par tous les états portant un paramètre exercice
        (ZACC3, ZACC12, ZACC16)."""
        company = request.user.company
        exercice_id = request.query_params.get('exercice')
        if not exercice_id:
            return None, Response(
                {'detail': "Le paramètre 'exercice' est requis."},
                status=status.HTTP_400_BAD_REQUEST)
        exercice = ExerciceComptable.objects.filter(
            company=company, pk=exercice_id).first()
        if exercice is None:
            return None, Response(
                {'detail': 'Exercice introuvable pour cette société.'},
                status=status.HTTP_404_NOT_FOUND)
        return exercice, None

    @action(detail=False, methods=['get'], url_path='tableau-flux')
    def tableau_flux(self, request):
        """ZACC3 — Tableau de financement / des flux de trésorerie CGNC
        (méthode indirecte).

        5e état CGNC (standard Odoo « Cash Flow Statement ») jamais couvert
        avant : réconcilie le résultat de l'exercice à la variation RÉELLE de
        trésorerie via 3 sections (exploitation/investissement/financement).
        Paramètres : ``exercice`` (id, requis), ``validees`` (1 → validées),
        ``export`` (``csv``/``pdf`` via ZACC1). Lecture seule, scopée société,
        Admin/Responsable.
        """
        exercice, err = self._resolve_exercice(request)
        if err is not None:
            return err
        data = selectors.tableau_flux_tresorerie(
            request.user.company, exercice,
            validees_seulement=request.query_params.get('validees') == '1')
        export = request.query_params.get('export')
        if export == 'csv':
            return self._export_tableau_flux_csv(exercice, data)
        if export == 'pdf':
            result = self._pdf_or_503(
                lambda: self._render_tableau_flux_pdf(request, data))
            if isinstance(result, Response):
                return result
            return self._pdf_response(
                result, f'tableau_flux_exercice_{exercice.pk}.pdf')
        return Response(data)

    def _render_tableau_flux_pdf(self, request, data):
        from .pdf_etats import _entete_societe_html, _fmt, _wrap
        from datetime import date as _date
        entete = _entete_societe_html(self._company_profile(request))
        corps = f"""
        <table><tbody>
        <tr class="total-row"><td colspan="2">Exploitation</td></tr>
        <tr><td>Résultat net</td><td class="montant">
        {_fmt(data['exploitation']['resultat_net'])}</td></tr>
        <tr><td>+ Dotations</td><td class="montant">
        {_fmt(data['exploitation']['dotations'])}</td></tr>
        <tr><td>- Reprises</td><td class="montant">
        {_fmt(data['exploitation']['reprises'])}</td></tr>
        <tr><td>= Capacité d'autofinancement</td><td class="montant">
        {_fmt(data['exploitation']['capacite_autofinancement'])}</td></tr>
        <tr><td>± Variation du BFR</td><td class="montant">
        {_fmt(data['exploitation']['variation_bfr'])}</td></tr>
        <tr class="total-row"><td>Flux net d'exploitation</td>
        <td class="montant">{_fmt(data['exploitation']['flux_net'])}</td></tr>
        <tr class="total-row"><td colspan="2">Investissement</td></tr>
        <tr class="total-row"><td>Flux net d'investissement</td>
        <td class="montant">
        {_fmt(data['investissement']['flux_net'])}</td></tr>
        <tr class="total-row"><td colspan="2">Financement</td></tr>
        <tr class="total-row"><td>Flux net de financement</td>
        <td class="montant">{_fmt(data['financement']['flux_net'])}</td></tr>
        <tr class="total-row"><td>Variation nette de trésorerie</td>
        <td class="montant">
        {_fmt(data['variation_nette_tresorerie'])}</td></tr>
        <tr><td>Trésorerie à l'ouverture</td><td class="montant">
        {_fmt(data['tresorerie_ouverture'])}</td></tr>
        <tr><td>Trésorerie à la clôture</td><td class="montant">
        {_fmt(data['tresorerie_cloture'])}</td></tr>
        </tbody></table>"""
        html = _wrap(
            entete, 'Tableau de financement / des flux de trésorerie',
            f"Exercice du {data['date_debut']} au {data['date_fin']}",
            corps, _date.today())
        from .pdf_etats import _html_to_pdf
        return _html_to_pdf(html)

    @staticmethod
    def _export_tableau_flux_csv(exercice, data):
        """Sérialise le tableau de financement (ZACC3) en CSV."""
        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=';')
        writer.writerow(['Tableau de financement / des flux de trésorerie'])
        writer.writerow(
            ['Exercice', f"{data['date_debut']} → {data['date_fin']}"])
        writer.writerow([])
        writer.writerow(['EXPLOITATION'])
        expl = data['exploitation']
        writer.writerow(['Résultat net', expl['resultat_net']])
        writer.writerow(['Dotations', expl['dotations']])
        writer.writerow(['Reprises', expl['reprises']])
        writer.writerow(
            ["Capacité d'autofinancement", expl['capacite_autofinancement']])
        writer.writerow(['Variation du BFR', expl['variation_bfr']])
        writer.writerow(['Flux net exploitation', expl['flux_net']])
        writer.writerow([])
        writer.writerow(['INVESTISSEMENT'])
        writer.writerow(
            ['Flux net investissement', data['investissement']['flux_net']])
        writer.writerow([])
        writer.writerow(['FINANCEMENT'])
        writer.writerow(
            ['Flux net financement', data['financement']['flux_net']])
        writer.writerow([])
        writer.writerow(
            ['Variation nette de trésorerie',
             data['variation_nette_tresorerie']])
        writer.writerow(['Trésorerie ouverture', data['tresorerie_ouverture']])
        writer.writerow(['Trésorerie clôture', data['tresorerie_cloture']])
        resp = HttpResponse(
            buffer.getvalue(), content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = (
            'attachment; filename='
            f'"tableau_flux_exercice_{exercice.pk}.csv"')
        return resp

    @action(detail=False, methods=['get'], url_path='tableau-immobilisations')
    def tableau_immobilisations(self, request):
        """ZACC12 — Rapport des immobilisations (tableau CGNC B2/B2bis).

        Par immobilisation : valeur brute ouverture/acquisitions/cessions/
        clôture, cumul d'amortissement ouverture/dotations/reprises/clôture,
        VNC. Paramètres : ``exercice`` (id, requis), ``validees`` (1 →
        validées), ``export`` (``csv``/``pdf`` via ZACC1). Lecture seule,
        scopée société, Admin/Responsable.
        """
        exercice, err = self._resolve_exercice(request)
        if err is not None:
            return err
        data = selectors.tableau_immobilisations(
            request.user.company, exercice,
            validees_seulement=request.query_params.get('validees') == '1')
        export = request.query_params.get('export')
        if export == 'csv':
            return self._export_tableau_immobilisations_csv(exercice, data)
        if export == 'pdf':
            result = self._pdf_or_503(
                lambda: self._render_tableau_immobilisations_pdf(
                    request, data))
            if isinstance(result, Response):
                return result
            return self._pdf_response(
                result, f'tableau_immobilisations_exercice_{exercice.pk}.pdf')
        return Response(data)

    def _render_tableau_immobilisations_pdf(self, request, data):
        from datetime import date as _date
        from html import escape
        from .pdf_etats import _entete_societe_html, _fmt, _wrap, _html_to_pdf
        entete = _entete_societe_html(self._company_profile(request))
        rows = ''.join(
            f"<tr><td>{escape(li['libelle'])}</td>"
            f"<td class=\"montant\">{_fmt(li['brut_ouverture'])}</td>"
            f"<td class=\"montant\">{_fmt(li['acquisitions'])}</td>"
            f"<td class=\"montant\">{_fmt(li['cessions'])}</td>"
            f"<td class=\"montant\">{_fmt(li['brut_cloture'])}</td>"
            f"<td class=\"montant\">{_fmt(li['amort_ouverture'])}</td>"
            f"<td class=\"montant\">{_fmt(li['dotations'])}</td>"
            f"<td class=\"montant\">{_fmt(li['reprises'])}</td>"
            f"<td class=\"montant\">{_fmt(li['amort_cloture'])}</td>"
            f"<td class=\"montant\">"
            f"{_fmt(li['valeur_nette_comptable'])}</td></tr>"
            for li in data['lignes']
        )
        totaux = data['totaux']
        corps = f"""
        <table><thead><tr><th>Immobilisation</th>
        <th class="montant">Brut ouv.</th><th class="montant">Acquis.</th>
        <th class="montant">Cessions</th><th class="montant">Brut clôt.</th>
        <th class="montant">Amort. ouv.</th>
        <th class="montant">Dotations</th><th class="montant">Reprises</th>
        <th class="montant">Amort. clôt.</th><th class="montant">VNC</th>
        </tr></thead><tbody>{rows}
        <tr class="total-row"><td>Totaux</td>
        <td class="montant">{_fmt(totaux['brut_ouverture'])}</td>
        <td class="montant">{_fmt(totaux['acquisitions'])}</td>
        <td class="montant">{_fmt(totaux['cessions'])}</td>
        <td class="montant">{_fmt(totaux['brut_cloture'])}</td>
        <td class="montant">{_fmt(totaux['amort_ouverture'])}</td>
        <td class="montant">{_fmt(totaux['dotations'])}</td>
        <td class="montant">{_fmt(totaux['reprises'])}</td>
        <td class="montant">{_fmt(totaux['amort_cloture'])}</td>
        <td class="montant">{_fmt(totaux['valeur_nette_comptable'])}</td>
        </tr></tbody></table>"""
        html = _wrap(
            entete, 'Tableau des immobilisations & amortissements',
            f"Exercice du {data['date_debut']} au {data['date_fin']}",
            corps, _date.today())
        return _html_to_pdf(html)

    @staticmethod
    def _export_tableau_immobilisations_csv(exercice, data):
        """Sérialise le tableau des immobilisations (ZACC12) en CSV."""
        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=';')
        writer.writerow(['Tableau des immobilisations & amortissements'])
        writer.writerow(
            ['Exercice', f"{data['date_debut']} → {data['date_fin']}"])
        writer.writerow([])
        writer.writerow([
            'Immobilisation', 'Brut ouverture', 'Acquisitions', 'Cessions',
            'Brut clôture', 'Amort. ouverture', 'Dotations', 'Reprises',
            'Amort. clôture', 'VNC'])
        for li in data['lignes']:
            writer.writerow([
                li['libelle'], li['brut_ouverture'], li['acquisitions'],
                li['cessions'], li['brut_cloture'], li['amort_ouverture'],
                li['dotations'], li['reprises'], li['amort_cloture'],
                li['valeur_nette_comptable']])
        totaux = data['totaux']
        writer.writerow([
            'Totaux', totaux['brut_ouverture'], totaux['acquisitions'],
            totaux['cessions'], totaux['brut_cloture'],
            totaux['amort_ouverture'], totaux['dotations'],
            totaux['reprises'], totaux['amort_cloture'],
            totaux['valeur_nette_comptable']])
        resp = HttpResponse(
            buffer.getvalue(), content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = (
            'attachment; filename='
            f'"tableau_immobilisations_exercice_{exercice.pk}.csv"')
        return resp

    @action(detail=False, methods=['get'], url_path='dossier-cloture')
    def dossier_cloture(self, request):
        """ZACC16 — Dossier de clôture xlsx multi-onglets (bilan/CPC/
        balance/grand-livre/balance-âgée/tableau-immos ZACC12/tableau-flux
        ZACC3) en UN seul fichier remis au fiduciaire.

        Assemble les sélecteurs EXISTANTS (aucun recalcul). Paramètres :
        ``exercice`` (id, requis), ``validees`` (1 → validées),
        ``export=xlsx`` (seul format supporté — sinon 400 explicite : ce
        dossier n'a pas de rendu JSON, il n'existe qu'en xlsx). Company-
        scopé, Admin/Responsable.
        """
        exercice, err = self._resolve_exercice(request)
        if err is not None:
            return err
        if request.query_params.get('export') != 'xlsx':
            return Response(
                {'detail': "Seul 'export=xlsx' est supporté pour le "
                           "dossier de clôture."},
                status=status.HTTP_400_BAD_REQUEST)
        company = request.user.company
        validees = request.query_params.get('validees') == '1'
        date_debut = exercice.date_debut.isoformat()
        date_fin = exercice.date_fin.isoformat()

        bilan = selectors.bilan(
            company, date_fin=date_fin, validees_seulement=validees)
        cpc_data = selectors.cpc(
            company, date_debut=date_debut, date_fin=date_fin,
            validees_seulement=validees)
        balance = selectors.balance_generale(
            company, date_debut=date_debut, date_fin=date_fin,
            validees_seulement=validees)
        grand_livre_data = selectors.grand_livre(
            company, date_debut=date_debut, date_fin=date_fin,
            validees_seulement=validees)
        balance_agee = selectors.balance_agee_fournisseurs(
            company, date_reference=date_fin, validees_seulement=validees)
        tableau_immos = selectors.tableau_immobilisations(
            company, exercice, validees_seulement=validees)
        tableau_flux = selectors.tableau_flux_tresorerie(
            company, exercice, validees_seulement=validees)

        from openpyxl import Workbook
        from openpyxl.styles import Font
        from apps.records.xlsx import coerce_cell, XLSX_CONTENT_TYPE

        def _feuille(wb, titre, headers, rows, first=False):
            ws = wb.active if first else wb.create_sheet()
            ws.title = titre[:31]
            ws.append(list(headers))
            bold = Font(bold=True)
            for cell in ws[1]:
                cell.font = bold
            for row in rows:
                ws.append([coerce_cell(v) for v in row])
            return ws

        wb = Workbook()
        _feuille(
            wb, 'Bilan', ['Section', 'N°', 'Intitulé', 'Montant'],
            [['Actif', p['numero'], p['intitule'], p['montant']]
             for p in bilan['actif']]
            + [['Passif', p['numero'], p['intitule'], p['montant']]
               for p in bilan['passif']]
            + [['Résultat', '', '', bilan['resultat']]],
            first=True)
        _feuille(
            wb, 'CPC', ['Section', 'N°', 'Intitulé', 'Montant'],
            [['Produit', p['numero'], p['intitule'], p['montant']]
             for p in cpc_data['produits']]
            + [['Charge', p['numero'], p['intitule'], p['montant']]
               for p in cpc_data['charges']]
            + [['Résultat', '', '', cpc_data['resultat']]])
        _feuille(
            wb, 'Balance',
            ['N°', 'Intitulé', 'Débit', 'Crédit', 'Solde débiteur',
             'Solde créditeur'],
            [[li['numero'], li['intitule'], li['debit'], li['credit'],
              li['solde_debiteur'], li['solde_crediteur']]
             for li in balance['lignes']])
        _feuille(
            wb, 'Grand livre',
            ['Compte', 'Date', 'Journal', 'Référence', 'Libellé', 'Débit',
             'Crédit'],
            [[compte['numero'], li['date'], li['journal'], li['reference'],
              li['libelle'], li['debit'], li['credit']]
             for compte in grand_livre_data for li in compte['lignes']])
        _feuille(
            wb, 'Balance âgée fournisseurs',
            ['Fournisseur', '0-30j', '31-60j', '61-90j', '90j+', 'Total'],
            [[li['fournisseur_nom'], li['b0_30'], li['b31_60'],
              li['b61_90'], li['b90_plus'], li['total']]
             for li in balance_agee])
        _feuille(
            wb, 'Tableau immobilisations',
            ['Immobilisation', 'Brut ouverture', 'Acquisitions', 'Cessions',
             'Brut clôture', 'Amort. ouverture', 'Dotations', 'Reprises',
             'Amort. clôture', 'VNC'],
            [[li['libelle'], li['brut_ouverture'], li['acquisitions'],
              li['cessions'], li['brut_cloture'], li['amort_ouverture'],
              li['dotations'], li['reprises'], li['amort_cloture'],
              li['valeur_nette_comptable']]
             for li in tableau_immos['lignes']])
        _feuille(
            wb, 'Tableau des flux',
            ['Section', 'Poste', 'Montant'],
            [['Exploitation', k, v]
             for k, v in tableau_flux['exploitation'].items()]
            + [['Investissement', k, v]
               for k, v in tableau_flux['investissement'].items()]
            + [['Financement', k, v]
               for k, v in tableau_flux['financement'].items()]
            + [['Synthèse', 'variation_nette_tresorerie',
                tableau_flux['variation_nette_tresorerie']]])

        response = HttpResponse(content_type=XLSX_CONTENT_TYPE)
        response['Content-Disposition'] = (
            'attachment; filename='
            f'"dossier_cloture_exercice_{exercice.pk}.xlsx"')
        wb.save(response)
        return response

    @action(detail=False, methods=['get'], url_path='export-fiduciaire')
    def export_fiduciaire(self, request):
        """Export fiduciaire Sage/CEGID des écritures d'un exercice (COMPTA37).

        Restitue, pour un exercice, l'ensemble ORDONNÉ des écritures reprojetées
        dans le jeu de colonnes d'échange fiduciaire (code journal, date, compte
        général/auxiliaire, référence de pièce, libellé, sens D/C, montant) —
        le pivot que les logiciels de tenue Sage et CEGID savent réimporter.
        100 % OFFLINE : un fichier téléchargeable, jamais d'appel externe.
        Paramètres : ``exercice`` (id, requis), ``validees`` (1 → écritures
        validées seulement) et ``export=csv`` (point-virgule) pour le fichier ;
        sans ``export``, renvoie le JSON (colonnes + lignes + synthèse liasse).
        On utilise ``export=`` et JAMAIS le ``format=`` de DRF (qui répond 404).
        Lecture seule, scopée société, Admin/Responsable.
        """
        company = request.user.company
        exercice_id = request.query_params.get('exercice')
        if not exercice_id:
            return Response(
                {'detail': "Le paramètre 'exercice' est requis."},
                status=status.HTTP_400_BAD_REQUEST)
        exercice = ExerciceComptable.objects.filter(
            company=company, pk=exercice_id).first()
        if exercice is None:
            return Response(
                {'detail': 'Exercice introuvable pour cette société.'},
                status=status.HTTP_404_NOT_FOUND)
        data = selectors.export_fiduciaire(
            company, exercice,
            validees_seulement=request.query_params.get('validees') == '1')
        if request.query_params.get('export') == 'csv':
            return self._export_fiduciaire_file(exercice, data)
        return Response(data)

    @staticmethod
    def _export_fiduciaire_file(exercice, data):
        """Sérialise l'export fiduciaire (COMPTA37) en CSV Sage/CEGID.

        Journal d'import délimité point-virgule : l'entête de colonnes, une ligne
        par mouvement, puis une synthèse de liasse (produits/charges/résultat).
        """
        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=';', lineterminator='\r\n')
        writer.writerow(data['columns'])
        for ligne in data['lignes']:
            writer.writerow([ligne[col] for col in data['columns']])
        writer.writerow([])
        writer.writerow(['SYNTHESE LIASSE'])
        synthese = data['synthese']
        writer.writerow(['Total produits', synthese['total_produits']])
        writer.writerow(['Total charges', synthese['total_charges']])
        writer.writerow(['Resultat', synthese['resultat']])
        resp = HttpResponse(
            buffer.getvalue(), content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = (
            'attachment; filename='
            f'"fiduciaire_sage_cegid_exercice_{exercice.pk}'
            f'_{exercice.date_debut}_{exercice.date_fin}.csv"')
        return resp

    @action(detail=False, methods=['get'], url_path='provisions')
    def provisions(self, request):
        """État récapitulatif des dotations/reprises de provisions (XACC26)."""
        params = request.query_params
        data = selectors.etat_provisions(
            request.user.company,
            date_debut=params.get('date_debut') or None,
            date_fin=params.get('date_fin') or None,
            nature=params.get('nature') or None)
        # Decimal n'est pas JSON-sérialisable nativement ; DRF Response le gère
        # via son encoder, mais les clés dict (nature enum) doivent rester str.
        return Response({str(k): v for k, v in data.items()})

    @action(detail=False, methods=['get'], url_path='continuite-sequences')
    def continuite_sequences(self, request):
        """XACC29 — Rapport de continuité des séquences (gap detection).

        Balaie factures/avoirs/pièces comptables et liste les trous de
        numérotation par radical. ``?export=csv`` télécharge le rapport ;
        sans le paramètre, renvoie le JSON (liste vide = tout continu)."""
        rapport = selectors.trous_sequences(request.user.company)
        if request.query_params.get('export') == 'csv':
            buffer = io.StringIO()
            writer = csv.writer(buffer, delimiter=';', lineterminator='\r\n')
            writer.writerow(['source', 'journal', 'de', 'a', 'manquants'])
            for entree in rapport:
                writer.writerow([
                    entree['source'], entree['journal'],
                    entree['plage'][0], entree['plage'][1],
                    ' '.join(str(n) for n in entree['manquants']),
                ])
            resp = HttpResponse(
                buffer.getvalue(), content_type='text/csv; charset=utf-8')
            resp['Content-Disposition'] = (
                'attachment; filename="continuite_sequences.csv"')
            return resp
        return Response(rapport)

    def _rapprochement_csv(self, rapport, filename):
        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=';', lineterminator='\r\n')
        writer.writerow([
            'tiers_id', 'nom', 'encours_documentaire', 'solde_gl', 'ecart',
            'references'])
        for ligne in rapport['lignes']:
            writer.writerow([
                ligne['tiers_id'], ligne['nom'],
                ligne['encours_documentaire'], ligne['solde_gl'],
                ligne['ecart'], ' '.join(ligne['references'])])
        resp = HttpResponse(
            buffer.getvalue(), content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp

    @action(detail=False, methods=['get'], url_path='rapprochement-clients')
    def rapprochement_clients(self, request):
        """YLEDG13 — tie-out AR : encours documentaire (ventes) vs solde GL
        3421 non lettré, par client. ``?date=YYYY-MM-DD`` borne le GL ;
        ``?export=csv`` télécharge. Sur un jeu sain, écart global = 0."""
        date = request.query_params.get('date') or None
        rapport = selectors.rapprochement_auxiliaire_clients(
            request.user.company, date=date)
        if request.query_params.get('export') == 'csv':
            return self._rapprochement_csv(
                rapport, 'rapprochement_clients.csv')
        return Response(rapport)

    @action(detail=False, methods=['get'],
            url_path='rapprochement-fournisseurs')
    def rapprochement_fournisseurs(self, request):
        """YLEDG13 — tie-out AP : encours documentaire (stock) vs solde GL
        4411 non lettré, par fournisseur. Miroir de ``rapprochement_clients``."""
        date = request.query_params.get('date') or None
        rapport = selectors.rapprochement_auxiliaire_fournisseurs(
            request.user.company, date=date)
        if request.query_params.get('export') == 'csv':
            return self._rapprochement_csv(
                rapport, 'rapprochement_fournisseurs.csv')
        return Response(rapport)

    @action(detail=False, methods=['get'], url_path='loi-69-21')
    def loi_69_21(self, request):
        """XFAC2 — Conformité loi 69-21 : factures fournisseur impayées au
        delà du délai légal (60 j défaut, 120 j max), avec amende estimée.
        ``?periode=YYYY-MM`` borne au trimestre civil (déclaration DGI) ;
        ``?export=csv`` télécharge."""
        periode = request.query_params.get('periode') or None
        rapport = selectors.exposition_69_21(
            request.user.company, periode=periode)
        if request.query_params.get('export') == 'csv':
            return self._loi_69_21_csv(rapport)
        return Response(rapport)

    @staticmethod
    def _loi_69_21_csv(rapport):
        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=';', lineterminator='\r\n')
        writer.writerow([
            'facture_id', 'reference', 'fournisseur_id', 'fournisseur_nom',
            'date_emission', 'delai_legal_jours', 'date_echeance_legale',
            'jours_depassement', 'montant_du', 'amende_estimee'])
        for ligne in rapport['lignes']:
            writer.writerow([
                ligne['facture_id'], ligne['reference'],
                ligne['fournisseur_id'], ligne['fournisseur_nom'],
                ligne['date_emission'], ligne['delai_legal_jours'],
                ligne['date_echeance_legale'], ligne['jours_depassement'],
                ligne['montant_du'], ligne['amende_estimee']])
        resp = HttpResponse(
            buffer.getvalue(), content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = 'attachment; filename="loi_69_21.csv"'
        return resp

    @action(detail=False, methods=['get'], url_path='journal-items')
    def journal_items(self, request):
        """ZACC4 — Vue « Journal Items » : ledger PLAT ligne-à-ligne, toutes
        écritures confondues, filtrable par ``journal``/``compte``/
        ``tiers_type``+``tiers_id``/``date_debut``/``date_fin``/
        ``lettrage`` (``lettrees``/``non_lettrees``)/``validees`` (``1``/
        ``0``). Paginé (``?limit=``/``?offset=``, défaut 200/0) ;
        ``?export=csv``/``?export=xlsx`` télécharge TOUT (non paginé)."""
        params = request.query_params
        validees = params.get('validees')
        if validees == '1':
            validees_bool = True
        elif validees == '0':
            validees_bool = False
        else:
            validees_bool = None
        lignes = selectors.journal_items(
            request.user.company,
            journal=params.get('journal') or None,
            compte=params.get('compte') or None,
            tiers_type=params.get('tiers_type') or None,
            tiers_id=params.get('tiers_id') or None,
            date_debut=params.get('date_debut') or None,
            date_fin=params.get('date_fin') or None,
            lettrage=params.get('lettrage') or None,
            validees=validees_bool,
        )
        export = params.get('export')
        if export in ('csv', 'xlsx'):
            return self._journal_items_export(lignes, export)
        try:
            limit = int(params.get('limit', 200))
        except (TypeError, ValueError):
            limit = 200
        try:
            offset = int(params.get('offset', 0))
        except (TypeError, ValueError):
            offset = 0
        limit = max(1, min(limit, 1000))
        offset = max(0, offset)
        page = lignes[offset:offset + limit]
        return Response({'count': len(lignes), 'results': page})

    @staticmethod
    def _journal_items_export(lignes, export):
        headers = [
            'id', 'date_ecriture', 'journal_code', 'ecriture_reference',
            'compte_numero', 'compte_intitule', 'libelle', 'tiers_type',
            'tiers_id', 'debit', 'credit', 'lettrage', 'statut',
        ]
        rows = [[ligne[h] for h in headers] for ligne in lignes]
        if export == 'xlsx':
            from apps.records.xlsx import build_xlsx_response
            return build_xlsx_response(
                'journal-items.xlsx', headers, rows,
                sheet_title='Journal Items')
        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=';', lineterminator='\r\n')
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        resp = HttpResponse(
            buffer.getvalue(), content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = (
            'attachment; filename="journal-items.csv"')
        return resp

    @action(detail=False, methods=['get'], url_path='controle-ice')
    def controle_ice(self, request):
        """ZACC14 — Tiers (clients entreprise + fournisseurs) à ICE manquant
        ou de format invalide (≠ 15 chiffres). ``?export=csv`` télécharge."""
        rapport = selectors.controle_identifiants_tiers(request.user.company)
        if request.query_params.get('export') == 'csv':
            return self._controle_ice_csv(rapport)
        return Response(rapport)

    @staticmethod
    def _controle_ice_csv(rapport):
        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=';', lineterminator='\r\n')
        writer.writerow(['type_tiers', 'id', 'nom', 'ice', 'if_fiscal', 'motif'])
        for tiers in rapport['clients']:
            writer.writerow([
                'client', tiers['id'], tiers['nom'], tiers['ice'],
                tiers['if_fiscal'], tiers['motif']])
        for tiers in rapport['fournisseurs']:
            writer.writerow([
                'fournisseur', tiers['id'], tiers['nom'], tiers['ice'],
                tiers['if_fiscal'], tiers['motif']])
        resp = HttpResponse(
            buffer.getvalue(), content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = (
            'attachment; filename="controle_ice.csv"')
        return resp


# ── YLEDG6 — Lettrage / délettrage (FG112) ──────────────────────────────────

class LettrageViewSet(viewsets.ViewSet):
    """Lettrage manuel + délettrage (FG112/YLEDG6). L'auto-lettrage à
    l'encaissement (YLEDG6) se pose déjà seul via
    ``compta.receivers``/``services.auto_lettrer_facture_soldee`` — cette
    vue couvre la correction manuelle : lettrer un lot choisi, ou délettrer
    un code posé par erreur (rouvre le lot ; balance âgée/encours
    ré-incluent les lignes). Admin/Responsable uniquement."""
    permission_classes = [IsResponsableOrAdmin]

    @action(detail=False, methods=['post'])
    def lettrer(self, request):
        ligne_ids = request.data.get('ligne_ids')
        code = (request.data.get('code') or '').strip()
        if not isinstance(ligne_ids, list) or not ligne_ids:
            return Response(
                {'detail': "'ligne_ids' (liste non vide) requis."},
                status=status.HTTP_400_BAD_REQUEST)
        company = request.user.company
        if not code:
            compte_id = request.data.get('compte')
            compte = CompteComptable.objects.filter(
                company=company, pk=compte_id).first() if compte_id else None
            if compte is None:
                return Response(
                    {'detail': "'code' ou 'compte' (pour en générer un) "
                               'requis.'},
                    status=status.HTTP_400_BAD_REQUEST)
            code = selectors.prochain_code_lettrage(company, compte)
        try:
            nb = selectors.lettrer(company, ligne_ids, code)
        except ValueError as exc:
            return Response(
                {'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response({'code': code, 'lignes_lettrees': nb})

    @action(detail=False, methods=['post'])
    def delettrer(self, request):
        code = (request.data.get('code') or '').strip()
        if not code:
            return Response(
                {'detail': "'code' requis."},
                status=status.HTTP_400_BAD_REQUEST)
        nb = selectors.delettrer(request.user.company, code)
        return Response({'code': code, 'lignes_delettrees': nb})


# ── FG115 — Périodes comptables verrouillables ─────────────────────────────

class PeriodeComptableViewSet(_ComptaBaseViewSet):
    """Périodes comptables (FG115) : clôture/réouverture (verrouillage).

    Actions ``cloturer`` / ``rouvrir`` figent ou libèrent la période ; une fois
    verrouillée, les écritures & factures de l'intervalle deviennent immuables.
    """
    queryset = PeriodeComptable.objects.select_related('exercice').all()
    serializer_class = PeriodeComptableSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_debut', 'date_fin']

    @action(detail=True, methods=['post'])
    def cloturer(self, request, pk=None):
        # COMPTA40 — la clôture est une action de gouvernance dédiée.
        if not _peut(request.user, 'compta_cloturer'):
            return Response(
                {'detail': "Vous n'êtes pas habilité à clôturer une période."},
                status=status.HTTP_403_FORBIDDEN)
        periode = self.get_object()
        services.cloturer_periode(periode, user=request.user)
        return Response(self.get_serializer(periode).data)

    @action(detail=True, methods=['post'])
    def rouvrir(self, request, pk=None):
        periode = self.get_object()
        try:
            services.rouvrir_periode(periode)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(periode).data)

    @action(detail=True, methods=['get'])
    def checklist(self, request, pk=None):
        """Checklist de clôture calculée depuis les données réelles (XACC10).

        INFORMATIF SEULEMENT : n'empêche jamais ``cloturer`` (warning, pas de
        blocage dur — cf. ``cloturer`` ci-dessus, inchangée).
        """
        periode = self.get_object()  # scopé société par TenantMixin.
        return Response(selectors.checklist_cloture_periode(periode))

    @action(detail=True, methods=['post'], url_path='solder-tva')
    def solder_tva(self, request, pk=None):
        """Poste l'écriture de solde TVA de la période (XACC10)."""
        periode = self.get_object()  # scopé société par TenantMixin.
        try:
            ecriture = services.solder_tva_periode(periode, user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        if ecriture is None:
            return Response(
                {'detail': 'Rien à solder (pas de TVA nette due sur la '
                           'période).', 'ecriture_id': None})
        return Response({'ecriture_id': ecriture.id,
                         'reference': ecriture.reference})


# ── FG116 / FG117 — Exercices comptables (clôture, OD, à-nouveaux) ──────────

class ExerciceComptableViewSet(_ComptaBaseViewSet):
    """Exercices comptables (FG117) : clôture, réouverture, report d'à-nouveaux.

    Porte aussi l'action ``ecriture-od`` (FG116) : saisie d'une écriture de
    régularisation manuelle (sans document source), refusée si la période est
    verrouillée.
    """
    queryset = ExerciceComptable.objects.prefetch_related('periodes').all()
    serializer_class = ExerciceComptableSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_debut', 'date_fin']

    @action(detail=True, methods=['post'])
    def cloturer(self, request, pk=None):
        # COMPTA40 — la clôture est une action de gouvernance dédiée.
        if not _peut(request.user, 'compta_cloturer'):
            return Response(
                {'detail': "Vous n'êtes pas habilité à clôturer un exercice."},
                status=status.HTTP_403_FORBIDDEN)
        exercice = self.get_object()
        services.cloturer_exercice(exercice, user=request.user)
        return Response(self.get_serializer(exercice).data)

    @action(detail=True, methods=['post'])
    def rouvrir(self, request, pk=None):
        exercice = self.get_object()
        services.rouvrir_exercice(exercice)
        return Response(self.get_serializer(exercice).data)

    @action(detail=True, methods=['post'], url_path='reporter-a-nouveaux')
    def reporter_a_nouveaux(self, request, pk=None):
        """Reporte les soldes de bilan de CET exercice dans ``exercice_cible``.

        Corps : ``{"exercice_cible": <id>}``. L'exercice source (celui de l'URL)
        doit être clôturé ; le report est idempotent.
        """
        exercice_clos = self.get_object()
        cible_id = request.data.get('exercice_cible')
        cible = ExerciceComptable.objects.filter(
            company=request.user.company, id=cible_id).first()
        if cible is None:
            return Response(
                {'detail': 'Exercice cible inconnu.'},
                status=status.HTTP_400_BAD_REQUEST)
        try:
            ecriture = services.reporter_a_nouveaux(
                exercice_clos, cible, user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response({
            'exercice': self.get_serializer(cible).data,
            'ecriture_id': ecriture.id if ecriture else None,
        })

    @action(detail=False, methods=['post'], url_path='ecriture-od')
    def ecriture_od(self, request):
        """Crée une écriture de régularisation manuelle (OD) — FG116.

        Corps : ``{date_ecriture, libelle, reference?, lignes:[{compte, debit,
        credit, libelle?}]}``. Σ débit = Σ crédit exigé ; refusée si la période
        est verrouillée.
        """
        company = request.user.company
        data = request.data
        lignes_in = data.get('lignes') or []
        lignes = []
        for lig in lignes_in:
            compte = CompteComptable.objects.filter(
                company=company, id=lig.get('compte')).first()
            if compte is None:
                return Response(
                    {'detail': 'Compte inconnu.'},
                    status=status.HTTP_400_BAD_REQUEST)
            lignes.append({
                'compte': compte,
                'debit': lig.get('debit') or 0,
                'credit': lig.get('credit') or 0,
                'libelle': lig.get('libelle', '') or '',
            })
        try:
            ecriture = services.creer_ecriture_od(
                company, data.get('date_ecriture'),
                data.get('libelle', '') or 'Régularisation', lignes,
                reference=data.get('reference', '') or '',
                created_by=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(
            EcritureComptableSerializer(ecriture).data,
            status=status.HTTP_201_CREATED)


# ── FG118 — Registre des immobilisations ───────────────────────────────────

class ImmobilisationViewSet(_ComptaBaseViewSet):
    """Registre des immobilisations (FG118).

    Inventaire des actifs immobilisés (véhicules, outillage, matériel…) avec
    coût, date d'acquisition, catégorie et TVA. Société scopée + accès
    Administrateur/Responsable. Filtrable par catégorie, recherche sur
    libellé/référence. Pas d'amortissement (hors périmètre) — juste le registre.
    """
    queryset = Immobilisation.objects.all()
    serializer_class = ImmobilisationSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['libelle', 'reference']
    ordering_fields = ['date_acquisition', 'cout', 'libelle']

    def get_queryset(self):
        qs = super().get_queryset()
        categorie = self.request.query_params.get('categorie')
        if categorie:
            qs = qs.filter(categorie=categorie)
        return qs

    @action(detail=True, methods=['get', 'post'], url_path='plan-amortissement')
    def plan_amortissement(self, request, pk=None):
        """Plan d'amortissement de l'immobilisation (FG119).

        GET : renvoie le plan + son calendrier de dotations (ou 404 si aucun).
        POST : (re)génère le plan. Corps : ``{mode, duree_annees,
        base_amortissable?, date_debut?, coefficient_degressif?}``. Idempotent —
        les dotations déjà postées sont préservées. La société est posée côté
        serveur (jamais du corps).
        """
        immo = self.get_object()  # déjà scopé société par TenantMixin.
        if request.method == 'GET':
            plan = getattr(immo, 'plan_amortissement', None)
            if plan is None:
                return Response(
                    {'detail': "Aucun plan d'amortissement."},
                    status=status.HTTP_404_NOT_FOUND)
            return Response(PlanAmortissementSerializer(
                plan, context={'request': request}).data)
        data = request.data
        try:
            plan = services.generer_plan_amortissement(
                immo,
                mode=data.get('mode') or None,
                duree_annees=data.get('duree_annees') or None,
                base_amortissable=data.get('base_amortissable'),
                date_debut=data.get('date_debut') or None,
                coefficient_degressif=data.get('coefficient_degressif'),
            )
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(
            PlanAmortissementSerializer(
                plan, context={'request': request}).data,
            status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def ceder(self, request, pk=None):
        """Enregistre et poste la cession / mise au rebut de l'actif (FG120).

        Corps : ``{type_cession?, date_cession, prix_cession?}`` —
        ``type_cession`` (vente/rebut) déduit du prix si absent (0 → rebut). La
        VNC, le cumul d'amortissement et le résultat de cession sont figés côté
        serveur, l'écriture de sortie est postée au grand livre (refusée si la
        période est verrouillée) et l'immobilisation est marquée inactive. La
        société est posée côté serveur (jamais du corps). Renvoie la cession.
        """
        immo = self.get_object()  # déjà scopée société par TenantMixin.
        data = request.data
        try:
            cession = services.enregistrer_cession(
                immo,
                date_cession=data.get('date_cession'),
                prix_cession=data.get('prix_cession'),
                type_cession=data.get('type_cession') or None,
                user=request.user,
            )
            services.poster_cession(cession, user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        cession.refresh_from_db()
        return Response(
            CessionImmobilisationSerializer(
                cession, context={'request': request}).data,
            status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'],
            url_path='depuis-facture-fournisseur')
    def depuis_facture_fournisseur(self, request):
        """XACC33 — Capitalise une ligne de facture fournisseur (bouton
        « Immobiliser » sur l'écran facture fournisseur).

        Corps : ``{facture_id, ligne_id, categorie?, duree?, mode?}``. La
        ligne est résolue via ``apps.stock.selectors.
        ligne_facture_fournisseur_scoped`` (company-scopée, jamais un import
        de ``stock.models``) — introuvable ou cross-company → 404. Anti-
        doublon : re-capitaliser la même ligne → 400. Crée l'Immobilisation
        pré-remplie + son plan d'amortissement en un geste.
        """
        data = request.data
        facture_id = data.get('facture_id')
        ligne_id = data.get('ligne_id')
        if not facture_id or not ligne_id:
            return Response(
                {'detail': 'facture_id et ligne_id sont obligatoires.'},
                status=status.HTTP_400_BAD_REQUEST)
        try:
            immo = services.capitaliser_ligne_facture_fournisseur(
                request.user.company,
                facture_id=facture_id, ligne_id=ligne_id,
                categorie=data.get('categorie') or None,
                duree_annees=data.get('duree') or 5,
                mode=data.get('mode') or None,
                user=request.user,
            )
        except DjangoValidationError as exc:
            message = exc.messages[0] if exc.messages else str(exc)
            if 'introuvable' in message.lower():
                return Response({'detail': message}, status=status.HTTP_404_NOT_FOUND)
            return Response({'detail': message}, status=status.HTTP_400_BAD_REQUEST)
        return Response(
            ImmobilisationSerializer(immo, context={'request': request}).data,
            status=status.HTTP_201_CREATED)


class CessionImmobilisationViewSet(_ComptaBaseViewSet):
    """Cessions / mises au rebut d'immobilisations (FG120) — lecture seule.

    Les cessions sont enregistrées et postées via l'action ``ceder`` de
    l'immobilisation. Ce viewset les restitue (liste/détail) ; l'action
    ``poster`` re-poste une cession non encore postée (idempotente, refusée en
    période close). Société scopée.
    """
    http_method_names = ['get', 'post', 'head', 'options']
    queryset = CessionImmobilisation.objects.select_related(
        'immobilisation', 'ecriture').all()
    serializer_class = CessionImmobilisationSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_cession', 'id']

    def get_queryset(self):
        qs = super().get_queryset()
        immobilisation = self.request.query_params.get('immobilisation')
        if immobilisation:
            qs = qs.filter(immobilisation_id=immobilisation)
        type_cession = self.request.query_params.get('type_cession')
        if type_cession:
            qs = qs.filter(type_cession=type_cession)
        return qs

    @action(detail=True, methods=['post'])
    def poster(self, request, pk=None):
        """Poste la cession au grand livre (FG120). Refusée en période close."""
        cession = self.get_object()  # scopée société par TenantMixin.
        try:
            ecriture = services.poster_cession(cession, user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        cession.refresh_from_db()
        return Response({
            'cession': self.get_serializer(cession).data,
            'ecriture_id': ecriture.id if ecriture else None,
        })


class DotationAmortissementViewSet(_ComptaBaseViewSet):
    """Dotations d'amortissement (FG119) — lecture seule + action ``poster``.

    Les dotations sont calculées par le service (jamais saisies à la main) ;
    l'action ``poster`` passe la dotation au grand livre (débit classe 6 / crédit
    classe 28), refusée si la période est verrouillée. Société scopée.
    """
    http_method_names = ['get', 'post', 'head', 'options']
    queryset = DotationAmortissement.objects.select_related(
        'plan__immobilisation', 'ecriture').all()
    serializer_class = DotationAmortissementSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['annee', 'plan']

    def get_queryset(self):
        qs = super().get_queryset()
        plan = self.request.query_params.get('plan')
        if plan:
            qs = qs.filter(plan_id=plan)
        immobilisation = self.request.query_params.get('immobilisation')
        if immobilisation:
            qs = qs.filter(plan__immobilisation_id=immobilisation)
        return qs

    @action(detail=True, methods=['post'])
    def poster(self, request, pk=None):
        """Poste la dotation au grand livre (FG119). Refusée en période close."""
        dotation = self.get_object()  # scopée société par TenantMixin.
        try:
            ecriture = services.poster_dotation(dotation, user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        dotation.refresh_from_db()
        return Response({
            'dotation': self.get_serializer(dotation).data,
            'ecriture_id': ecriture.id if ecriture else None,
        })


# ── FG123 — Rapprochement bancaire (relevé ↔ écritures) ────────────────────

class RapprochementBancaireViewSet(_ComptaBaseViewSet):
    """Rapprochements bancaires (FG123) : pointer relevé ↔ grand livre.

    Crée un rapprochement par compte de trésorerie/période, ajoute des lignes de
    relevé, pointe chaque ligne contre des lignes du grand livre jusqu'à
    concordance, et expose la synthèse (solde relevé vs solde GL vs écart). C'est
    DISTINCT de l'import de paiements clients (FG42) : aucune écriture n'est
    créée. Société scopée, posée côté serveur ; Admin/Responsable uniquement.
    """
    queryset = RapprochementBancaire.objects.select_related(
        'compte_tresorerie').prefetch_related(
        'lignes_releve__lignes_gl').all()
    serializer_class = RapprochementBancaireSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_fin', 'date_debut', 'id']

    def get_queryset(self):
        qs = super().get_queryset()
        compte = self.request.query_params.get('compte_tresorerie')
        if compte:
            qs = qs.filter(compte_tresorerie_id=compte)
        statut = self.request.query_params.get('statut')
        if statut:
            qs = qs.filter(statut=statut)
        return qs

    def perform_create(self, serializer):
        serializer.save(
            company=self.request.user.company, created_by=self.request.user)

    def get_permissions(self):
        # YRBAC13 — fine-grain les @action au-delà du grossier
        # IsResponsableOrAdmin de base : lecture (rapports/synthèse) reste
        # ouverte à Admin/Responsable ; saisie (import/ligne de relevé) exige
        # ``compta_saisir`` ; validation (pointage/acceptation/clôture) exige
        # ``compta_valider`` — les deux sont déjà octroyés par défaut au
        # Responsable (COMPTA40), donc AUCUNE régression pour les rôles par
        # défaut ; seul un rôle fin explicitement privé du code est affecté.
        if self.action in ('lignes_gl', 'resume', 'suggestions'):
            return [IsResponsableOrAdmin()]
        if self.action in ('ligne_releve', 'ocr_import', 'import_releve'):
            return [HasPermissionOrLegacy('compta_saisir')()]
        if self.action in ('pointer', 'accepter_suggestions', 'cloturer'):
            return [HasPermissionOrLegacy('compta_valider')()]
        return super().get_permissions()

    @action(detail=True, methods=['get'], url_path='lignes-gl')
    def lignes_gl(self, request, pk=None):
        """Lignes du grand livre pointables sur la période (FG123)."""
        rapprochement = self.get_object()  # scopé société par TenantMixin.
        return Response(selectors.lignes_gl_pointables(rapprochement))

    @action(detail=True, methods=['get'])
    def resume(self, request, pk=None):
        """Synthèse : solde relevé vs solde GL vs écart (FG123)."""
        rapprochement = self.get_object()  # scopé société par TenantMixin.
        return Response(selectors.resume_rapprochement(rapprochement))

    @action(detail=True, methods=['post'], url_path='ligne-releve')
    def ligne_releve(self, request, pk=None):
        """Ajoute une ligne de relevé bancaire (FG123).

        Corps : ``{date_operation, libelle, montant, reference?}``. ``montant``
        signé (+ entrée, − sortie). Société héritée du rapprochement.
        """
        rapprochement = self.get_object()  # scopé société par TenantMixin.
        data = request.data
        try:
            ligne = services.ajouter_ligne_releve(
                rapprochement,
                date_operation=data.get('date_operation'),
                libelle=data.get('libelle', '') or '',
                montant=data.get('montant'),
                reference=data.get('reference', '') or '',
            )
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(
            LigneReleveSerializer(ligne).data,
            status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], url_path='pointer')
    def pointer(self, request, pk=None):
        """Pointe une ligne de relevé contre des lignes du grand livre (FG123).

        Corps : ``{ligne_releve: <id>, lignes_gl: [<id>, ...]}``. Remplace les
        lignes GL appariées ; la ligne devient ``rapprochee`` si l'écart est nul.
        Toutes les lignes restent scopées à la société du rapprochement.
        """
        rapprochement = self.get_object()  # scopé société par TenantMixin.
        data = request.data
        ligne = LigneReleve.objects.filter(
            company=request.user.company, rapprochement=rapprochement,
            id=data.get('ligne_releve')).first()
        if ligne is None:
            return Response(
                {'detail': 'Ligne de relevé inconnue.'},
                status=status.HTTP_400_BAD_REQUEST)
        try:
            ligne = services.pointer_ligne_releve(
                ligne, data.get('lignes_gl') or [])
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(LigneReleveSerializer(ligne).data)

    @action(detail=True, methods=['get'])
    def suggestions(self, request, pk=None):
        """Suggestions d'appariement relevé↔GL, notées par confiance (XACC3)."""
        rapprochement = self.get_object()  # scopé société par TenantMixin.
        return Response(selectors.suggestions_rapprochement(rapprochement))

    @action(detail=True, methods=['post'], url_path='accepter-suggestions')
    def accepter_suggestions(self, request, pk=None):
        """Pointe en un clic les suggestions non ambiguës (XACC3).

        Ne pointe JAMAIS silencieusement une ligne ambiguë (≥2 candidats au
        même score) — celles-ci restent listées dans ``ignorees`` pour
        arbitrage manuel via l'action ``pointer`` existante.
        """
        rapprochement = self.get_object()  # scopé société par TenantMixin.
        resultat = services.accepter_suggestions_rapprochement(rapprochement)
        return Response(resultat)

    @action(detail=True, methods=['post'])
    def cloturer(self, request, pk=None):
        """Clôture le rapprochement quand tout concorde (FG123)."""
        rapprochement = self.get_object()  # scopé société par TenantMixin.
        try:
            services.cloturer_rapprochement(rapprochement)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        rapprochement.refresh_from_db()
        return Response(self.get_serializer(rapprochement).data)

    @action(detail=True, methods=['post'], url_path='ocr-import',
            parser_classes=[MultiPartParser, FormParser])
    def ocr_import(self, request, pk=None):
        """XACC30 — OCR d'un relevé (PDF/scan) → lignes proposées (gated).

        Corps multipart : ``releve`` (fichier). Réponse : les lignes
        extraites + le contrôle de solde (jamais d'intégration silencieuse).
        Avec ``accepter=1`` dans le corps, injecte les lignes ``lignes``
        fournies (JSON, normalement celles proposées par un appel précédent)
        dans le rapprochement FG123 via ``ligne-releve``. Sans clé OCR
        configurée : 503 explicite, rien d'autre ne change.
        """
        rapprochement = self.get_object()  # scopé société par TenantMixin.
        accepter = str(request.data.get('accepter', '')).lower() in (
            '1', 'true', 'yes')
        if accepter:
            import json
            lignes_brutes = request.data.get('lignes') or '[]'
            if isinstance(lignes_brutes, str):
                try:
                    lignes_brutes = json.loads(lignes_brutes)
                except ValueError:
                    lignes_brutes = []
            try:
                creees = services.accepter_lignes_releve_ocr(
                    rapprochement, lignes_brutes)
            except DjangoValidationError as exc:
                return Response(
                    {'detail': exc.messages[0] if exc.messages else str(exc)},
                    status=status.HTTP_400_BAD_REQUEST)
            return Response(
                {'lignes_creees': LigneReleveSerializer(creees, many=True).data},
                status=status.HTTP_201_CREATED)

        releve = request.FILES.get('releve')
        if releve is None:
            return Response(
                {'releve': "Le fichier 'releve' est obligatoire."},
                status=status.HTTP_400_BAD_REQUEST)
        try:
            champs_bruts = services.extraire_releve_bancaire(
                releve.read(), mime=getattr(releve, 'content_type', '') or '')
        except RuntimeError as exc:
            return Response(
                {'detail': str(exc)}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        return Response(services.controler_solde_releve_ocr(champs_bruts))

    @action(detail=True, methods=['post'], url_path='import-releve',
            parser_classes=[MultiPartParser, FormParser])
    def import_releve(self, request, pk=None):
        """NTTRE1-3 — Importe un relevé bancaire au format normalisé.

        Query ``?format=cfonb120|mt940|camt053`` ; corps multipart ``releve``
        (fichier). Parse le fichier via ``bank_formats`` puis matérialise chaque
        ligne parsée en ``LigneReleve`` via ``services.ajouter_ligne_releve``
        (aucun nouveau modèle). Rejette avec un message FR clair si le format est
        inconnu ou le fichier mal formé.
        """
        from .bank_formats import (
            parser_camt053, parser_cfonb120, parser_mt940)

        rapprochement = self.get_object()  # scopé société par TenantMixin.
        parseurs = {
            'cfonb120': parser_cfonb120,
            'mt940': parser_mt940,
            'camt053': parser_camt053,
        }
        fmt = (request.query_params.get('format')
               or request.data.get('format') or '').lower()
        parseur = parseurs.get(fmt)
        if parseur is None:
            return Response(
                {'detail': "Format inconnu : utilisez « cfonb120 », "
                           "« mt940 » ou « camt053 »."},
                status=status.HTTP_400_BAD_REQUEST)
        releve = request.FILES.get('releve')
        if releve is None:
            return Response(
                {'releve': "Le fichier 'releve' est obligatoire."},
                status=status.HTTP_400_BAD_REQUEST)
        try:
            lignes_parsees = parseur(releve.read())
        except ValueError as exc:
            return Response(
                {'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        creees = []
        try:
            for ligne in lignes_parsees:
                creees.append(services.ajouter_ligne_releve(
                    rapprochement,
                    date_operation=ligne.get('date_operation'),
                    libelle=ligne.get('libelle', '') or '',
                    montant=ligne.get('montant'),
                    reference=ligne.get('reference', '') or '',
                ))
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(
            {'lignes_creees': LigneReleveSerializer(creees, many=True).data,
             'nombre': len(creees)},
            status=status.HTTP_201_CREATED)


# ── FG124 — Caisse / petty cash (journal d'espèces) ────────────────────────

class CaisseViewSet(_ComptaBaseViewSet):
    """Caisses d'espèces (petty cash) pour les achats terrain (FG124).

    Tient un journal d'espèces par caisse : entrées/sorties (avec justificatif
    et pièce) via l'action ``mouvement``, journal + solde courant via
    ``journal``/``resume``, et clôture de caisse (comptage physique, écart) via
    ``cloturer``. Le ``compte_tresorerie`` lié doit être de type caisse. Société
    scopée, posée côté serveur ; Admin/Responsable uniquement.
    """
    queryset = Caisse.objects.select_related('compte_tresorerie').all()
    serializer_class = CaisseSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['libelle', 'date_creation']

    def get_queryset(self):
        qs = super().get_queryset()
        actif = self.request.query_params.get('actif')
        if actif in ('0', '1'):
            qs = qs.filter(actif=(actif == '1'))
        return qs

    def perform_create(self, serializer):
        serializer.save(company=self.request.user.company)

    @action(detail=True, methods=['get', 'post'])
    def mouvement(self, request, pk=None):
        """Liste ou enregistre un mouvement d'espèces de la caisse (FG124).

        GET : renvoie le journal d'espèces (mouvements + solde cumulé), filtrable
        par ``date_debut``/``date_fin``.
        POST : enregistre une entrée/sortie. Corps : ``{sens, montant,
        date_mouvement, motif, justificatif?, piece?, compte_contrepartie?,
        poster?}``. ``montant`` strictement positif ; refusé à une date clôturée.
        Si ``poster`` est vrai, l'écriture de caisse est passée au grand livre
        (refusée si la période comptable est verrouillée). Société côté serveur.
        """
        caisse = self.get_object()  # scopée société par TenantMixin.
        if request.method == 'GET':
            params = request.query_params
            return Response(selectors.journal_caisse(
                caisse,
                date_debut=params.get('date_debut') or None,
                date_fin=params.get('date_fin') or None))
        data = request.data
        contrepartie = None
        cp_id = data.get('compte_contrepartie')
        if cp_id:
            contrepartie = CompteComptable.objects.filter(
                company=request.user.company, id=cp_id).first()
            if contrepartie is None:
                return Response(
                    {'detail': 'Compte de contrepartie inconnu.'},
                    status=status.HTTP_400_BAD_REQUEST)
        try:
            mouvement = services.enregistrer_mouvement_caisse(
                caisse,
                sens=data.get('sens'),
                montant=data.get('montant'),
                date_mouvement=data.get('date_mouvement'),
                motif=data.get('motif', '') or '',
                justificatif=data.get('justificatif', '') or '',
                piece=data.get('piece', '') or '',
                compte_contrepartie=contrepartie,
                poster=bool(data.get('poster')),
                user=request.user,
            )
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(
            MouvementCaisseSerializer(mouvement).data,
            status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], url_path='poster-mouvement')
    def poster_mouvement(self, request, pk=None):
        """Poste un mouvement existant au grand livre (FG124).

        Corps : ``{mouvement: <id>}``. Refusé si la période comptable est
        verrouillée. Idempotent. Le mouvement doit appartenir à la caisse.
        """
        caisse = self.get_object()  # scopée société par TenantMixin.
        mouvement = MouvementCaisse.objects.filter(
            company=request.user.company, caisse=caisse,
            id=request.data.get('mouvement')).first()
        if mouvement is None:
            return Response(
                {'detail': 'Mouvement inconnu.'},
                status=status.HTTP_400_BAD_REQUEST)
        try:
            ecriture = services.poster_mouvement_caisse(
                mouvement, user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        mouvement.refresh_from_db()
        return Response({
            'mouvement': MouvementCaisseSerializer(mouvement).data,
            'ecriture_id': ecriture.id if ecriture else None,
        })

    @action(detail=True, methods=['get'])
    def resume(self, request, pk=None):
        """Synthèse de la caisse : solde initial/entrées/sorties/courant (FG124)."""
        caisse = self.get_object()  # scopée société par TenantMixin.
        date_fin = request.query_params.get('date_fin') or None
        return Response(selectors.resume_caisse(caisse, date_fin=date_fin))

    @action(detail=True, methods=['get', 'post'])
    def cloturer(self, request, pk=None):
        """Liste les clôtures ou clôture la caisse (cash count) — FG124.

        GET : renvoie l'historique des clôtures de la caisse.
        POST : effectue une clôture. Corps : ``{date_cloture, solde_compte,
        commentaire?}``. Le solde théorique et l'écart sont figés côté serveur ;
        les mouvements ≤ ``date_cloture`` deviennent immuables. La société est
        posée côté serveur (jamais du corps).
        """
        caisse = self.get_object()  # scopée société par TenantMixin.
        if request.method == 'GET':
            return Response(ClotureCaisseSerializer(
                caisse.clotures.all(), many=True).data)
        data = request.data
        try:
            cloture = services.cloturer_caisse(
                caisse,
                date_cloture=data.get('date_cloture'),
                solde_compte=data.get('solde_compte'),
                commentaire=data.get('commentaire', '') or '',
                user=request.user,
            )
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(
            ClotureCaisseSerializer(cloture).data,
            status=status.HTTP_201_CREATED)


# ── FG125 — Virements internes entre comptes de trésorerie ─────────────────

class VirementInterneViewSet(_ComptaBaseViewSet):
    """Virements internes entre comptes de trésorerie (FG125).

    Banque↔banque/caisse en écriture à deux jambes. La création valide les
    comptes (mêmes société, différents) ; l'action ``poster`` passe l'écriture
    équilibrée au grand livre (refusée en période close). Société scopée.
    """
    queryset = VirementInterne.objects.select_related(
        'compte_source', 'compte_destination', 'ecriture').all()
    serializer_class = VirementInterneSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_virement', 'id']

    def get_queryset(self):
        qs = super().get_queryset()
        posted = self.request.query_params.get('posted')
        if posted in ('0', '1'):
            qs = qs.filter(posted=(posted == '1'))
        return qs

    def perform_create(self, serializer):
        serializer.save(
            company=self.request.user.company, created_by=self.request.user)

    @action(detail=True, methods=['post'])
    def poster(self, request, pk=None):
        """Poste le virement au grand livre (FG125). Refusé en période close."""
        virement = self.get_object()  # scopé société par TenantMixin.
        try:
            ecriture = services.poster_virement(virement, user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        virement.refresh_from_db()
        return Response({
            'virement': self.get_serializer(virement).data,
            'ecriture_id': ecriture.id if ecriture else None,
        })


# ── FG126 — Prévisionnel de trésorerie roulant 13 semaines ─────────────────

class LignePrevisionnelTresorerieViewSet(_ComptaBaseViewSet):
    """Lignes prévues éditables du prévisionnel de trésorerie (FG126).

    Saisie manuelle des flux attendus (crédits, leasing, salaires, acomptes IS…)
    empilés au-dessus de la projection AR/AP. La projection consolidée roulante
    13 semaines est servie par ``etats/previsionnel-tresorerie``. Société scopée.
    """
    queryset = LignePrevisionnelTresorerie.objects.all()
    serializer_class = LignePrevisionnelTresorerieSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_prevue', 'id']

    def get_queryset(self):
        qs = super().get_queryset()
        categorie = self.request.query_params.get('categorie')
        if categorie:
            qs = qs.filter(categorie=categorie)
        return qs

    def perform_create(self, serializer):
        serializer.save(company=self.request.user.company)


# ── FG127 / FG128 — Effets (chèques / traites) ─────────────────────────────

class EffetViewSet(_ComptaBaseViewSet):
    """Portefeuille d'effets à recevoir (FG127) & à payer (FG128).

    Chèques/traites avec échéance/banque/statut (portefeuille→remis→encaissé→
    impayé pour les effets à recevoir ; portefeuille→payé→impayé pour les effets
    à payer). Actions ``encaisser``/``payer``/``rejeter`` font évoluer le statut
    et passent l'écriture au grand livre (refusée en période close). Société
    scopée, posée côté serveur.
    """
    queryset = Effet.objects.select_related('bordereau').all()
    serializer_class = EffetSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['numero', 'tireur', 'banque']
    ordering_fields = ['date_echeance', 'date_emission', 'montant', 'id']

    def get_queryset(self):
        qs = super().get_queryset()
        params = self.request.query_params
        sens = params.get('sens')
        if sens:
            qs = qs.filter(sens=sens)
        statut = params.get('statut')
        if statut:
            qs = qs.filter(statut=statut)
        return qs

    def perform_create(self, serializer):
        serializer.save(
            company=self.request.user.company, created_by=self.request.user)

    def get_permissions(self):
        # YRBAC13 — toutes les actions font évoluer le statut d'un effet ET
        # passent une écriture au grand livre (docstring de la classe) : ce
        # sont des saisies comptables, gardées par ``compta_saisir`` (déjà
        # octroyé au Responsable par défaut — comportement inchangé pour les
        # rôles existants, resserré pour un rôle fin qui en serait privé).
        if self.action in ('encaisser', 'payer', 'rejeter', 'escompter',
                           'apurer_escompte', 'endosser'):
            return [HasPermissionOrLegacy('compta_saisir')()]
        return super().get_permissions()

    @action(detail=True, methods=['post'])
    def encaisser(self, request, pk=None):
        """Encaisse un effet à recevoir : remis/portefeuille → encaissé (FG127)."""
        effet = self.get_object()  # scopé société par TenantMixin.
        try:
            effet = services.encaisser_effet(
                effet,
                date_encaissement=request.data.get('date_encaissement') or None,
                user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(effet).data)

    @action(detail=True, methods=['post'])
    def payer(self, request, pk=None):
        """Paie un effet à payer fournisseur : portefeuille → payé (FG128)."""
        effet = self.get_object()  # scopé société par TenantMixin.
        try:
            effet = services.payer_effet(
                effet,
                date_paiement=request.data.get('date_paiement') or None,
                user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(effet).data)

    @action(detail=True, methods=['post'])
    def rejeter(self, request, pk=None):
        """Constate l'impayé / rejet d'un effet (FG130).

        Corps : ``{date_rejet?, frais_rejet?, commentaire?}``. Rouvre le montant
        dû et comptabilise les frais de rejet. Refusé en période close.
        """
        effet = self.get_object()  # scopé société par TenantMixin.
        try:
            effet = services.rejeter_effet(
                effet,
                date_rejet=request.data.get('date_rejet') or None,
                frais_rejet=request.data.get('frais_rejet'),
                commentaire=request.data.get('commentaire', '') or '',
                user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(effet).data)

    @action(detail=True, methods=['post'])
    def escompter(self, request, pk=None):
        """XACC34 — Remet un effet à recevoir à l'escompte (mobilisation).

        Corps : ``{compte_tresorerie, agios?, interets?, date_escompte?}``.
        Poste le net (débit trésorerie) + agios/intérêts (débit charge) /
        crédit 5520 « crédits d'escompte » du montant brut. Refusé en période
        close ; transitions illégales -> 400.
        """
        effet = self.get_object()  # scopé société par TenantMixin.
        treso = CompteTresorerie.objects.filter(
            company=request.user.company,
            id=request.data.get('compte_tresorerie')).first()
        if treso is None:
            return Response(
                {'detail': 'Compte de trésorerie inconnu.'},
                status=status.HTTP_400_BAD_REQUEST)
        try:
            effet = services.escompter_effet(
                effet, compte_tresorerie=treso,
                agios=request.data.get('agios'),
                interets=request.data.get('interets'),
                date_escompte=request.data.get('date_escompte') or None,
                user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(effet).data)

    @action(detail=True, methods=['post'], url_path='apurer-escompte')
    def apurer_escompte(self, request, pk=None):
        """XACC34 — Apure le crédit d'escompte à l'échéance (5520 → 3425)."""
        effet = self.get_object()  # scopé société par TenantMixin.
        try:
            effet = services.apurer_escompte_effet(
                effet,
                date_apurement=request.data.get('date_apurement') or None,
                user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(effet).data)

    @action(detail=True, methods=['post'])
    def endosser(self, request, pk=None):
        """XACC34 — Endosse un effet à recevoir à un tiers bénéficiaire.

        Corps : ``{beneficiaire, date_endossement?}``. Transitions illégales
        (effet déjà soldé/escompté/impayé) -> 400.
        """
        effet = self.get_object()  # scopé société par TenantMixin.
        try:
            effet = services.endosser_effet(
                effet, beneficiaire=request.data.get('beneficiaire', '') or '',
                date_endossement=request.data.get('date_endossement') or None,
                user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(effet).data)

    @action(detail=True, methods=['post'], url_path='constater-protet')
    def constater_protet(self, request, pk=None):
        """NTTRE7 — Constate un protêt (constat d'huissier) sur un effet impayé.

        Corps : ``{frais_protet?, date_protet?}``. L'effet doit être ``impaye``.
        Distinct du simple rejet FG130 : trace les frais + la date de protêt.
        """
        effet = self.get_object()  # scopé société par TenantMixin.
        try:
            effet = services.constater_protet(
                effet,
                frais_protet=request.data.get('frais_protet'),
                date_protet=request.data.get('date_protet') or None,
                user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(effet).data)


# ── FG129 — Bordereau de remise en banque ──────────────────────────────────

class BordereauRemiseViewSet(_ComptaBaseViewSet):
    """Bordereaux de remise en banque d'effets à recevoir (FG129).

    Regroupe des effets pour un dépôt groupé + écriture de remise. La création
    rattache les effets (corps ``{compte_tresorerie, date_remise, reference?,
    effet_ids: [...]}``) ; l'action ``poster`` passe l'écriture et fait basculer
    les effets en ``remis``. Société scopée, posée côté serveur.
    """
    http_method_names = ['get', 'post', 'head', 'options']
    queryset = BordereauRemise.objects.select_related(
        'compte_tresorerie', 'ecriture').prefetch_related('effets').all()
    serializer_class = BordereauRemiseSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_remise', 'id']

    def get_queryset(self):
        qs = super().get_queryset()
        statut = self.request.query_params.get('statut')
        if statut:
            qs = qs.filter(statut=statut)
        return qs

    def create(self, request, *args, **kwargs):
        data = request.data
        company = request.user.company
        treso = CompteTresorerie.objects.filter(
            company=company, id=data.get('compte_tresorerie')).first()
        if treso is None:
            return Response(
                {'detail': 'Compte de trésorerie inconnu.'},
                status=status.HTTP_400_BAD_REQUEST)
        try:
            bordereau = services.creer_bordereau(
                company, treso,
                date_remise=data.get('date_remise'),
                effet_ids=data.get('effet_ids') or [],
                reference=data.get('reference', '') or '',
                user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(
            self.get_serializer(bordereau).data,
            status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def poster(self, request, pk=None):
        """Poste le bordereau au grand livre (FG129). Refusé en période close."""
        bordereau = self.get_object()  # scopé société par TenantMixin.
        try:
            ecriture = services.poster_bordereau(bordereau, user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        bordereau.refresh_from_db()
        return Response({
            'bordereau': self.get_serializer(bordereau).data,
            'ecriture_id': ecriture.id if ecriture else None,
        })


# ── FG131 — Rapprochement 3 voies (BC ↔ réception ↔ facture fournisseur) ────

class RapprochementViewSet(_ComptaBaseViewSet):
    """Rapprochements 3 voies avant paiement (FG131).

    Confronte commandé (BC) ↔ reçu (réception) ↔ facturé (facture fournisseur)
    — les trois documents vivent dans ``apps.stock`` et sont lus via ses
    sélecteurs (jamais dupliqués). La création (corps ``{bon_commande,
    tolerance?, note?}``) passe par le service qui valide le BCF de la société
    et évalue immédiatement l'écart. ``evaluer`` rafraîchit les montants ;
    ``valider`` pose le bon-à-payer (refusé tant qu'un écart bloquant subsiste).
    Société scopée, posée côté serveur.
    """
    queryset = Rapprochement.objects.select_related(
        'bon_commande', 'bon_commande__fournisseur', 'valide_par').all()
    serializer_class = RapprochementSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_creation', 'date_evaluation', 'ecart', 'id']

    def get_queryset(self):
        qs = super().get_queryset()
        params = self.request.query_params
        statut = params.get('statut')
        if statut:
            qs = qs.filter(statut=statut)
        bon_commande = params.get('bon_commande')
        if bon_commande:
            qs = qs.filter(bon_commande_id=bon_commande)
        return qs

    def create(self, request, *args, **kwargs):
        data = request.data
        company = request.user.company
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        try:
            rapp = services.creer_rapprochement_3voies(
                company,
                bon_commande_id=serializer.validated_data['bon_commande'].id,
                tolerance=serializer.validated_data.get('tolerance'),
                note=serializer.validated_data.get('note', '') or '',
                user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(
            self.get_serializer(rapp).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def evaluer(self, request, pk=None):
        """Rafraîchit les trois montants et recalcule l'écart/statut (FG131)."""
        rapp = self.get_object()  # scopé société par TenantMixin.
        try:
            rapp = services.evaluer_rapprochement(rapp)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(rapp).data)

    @action(detail=True, methods=['post'])
    def valider(self, request, pk=None):
        """Pose le bon-à-payer (FG131). Refusé tant qu'un écart bloque."""
        rapp = self.get_object()  # scopé société par TenantMixin.
        try:
            rapp = services.valider_rapprochement(
                rapp, user=request.user,
                commentaire=request.data.get('commentaire', '') or '')
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(rapp).data)


# ── FG133 / FG134 — Campagnes de règlement fournisseurs + fichier virement ──

class PaymentRunViewSet(_ComptaBaseViewSet):
    """Campagnes de règlement fournisseurs (FG133) + export virement (FG134).

    Regroupe une sélection de dettes fournisseur dues à régler en un lot. La
    création (corps ``{date_paiement, mode_paiement?, compte_tresorerie?,
    reference?, note?, lignes: [{tiers_id, montant, reference?, date_echeance?,
    beneficiaire?, rib?, iban?}]}``) passe par le service, qui complète les
    bénéficiaires/coordonnées depuis le sélecteur de stock et calcule le total.
    ``figer`` fige la proposition (brouillon → proposée) ; ``poster`` passe
    l'écriture EN LOT (débit 4411 par ligne / crédit 5141 banque) et solde les
    dettes ; ``fichier-virement`` exporte l'ordre de virement bancaire (CSV).
    Société scopée, posée côté serveur ; suppression interdite une fois postée.
    """
    http_method_names = ['get', 'post', 'delete', 'head', 'options']
    queryset = PaymentRun.objects.select_related(
        'compte_tresorerie', 'ecriture').prefetch_related('lignes').all()
    serializer_class = PaymentRunSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_paiement', 'date_creation', 'total', 'id']

    def get_queryset(self):
        qs = super().get_queryset()
        params = self.request.query_params
        statut = params.get('statut')
        if statut:
            qs = qs.filter(statut=statut)
        mode = params.get('mode_paiement')
        if mode:
            qs = qs.filter(mode_paiement=mode)
        return qs

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        vd = serializer.validated_data
        compte = vd.get('compte_tresorerie')
        try:
            run = services.creer_payment_run(
                request.user.company,
                date_paiement=vd['date_paiement'],
                mode_paiement=vd.get('mode_paiement'),
                compte_tresorerie=compte,
                reference=vd.get('reference', '') or '',
                note=vd.get('note', '') or '',
                lignes=vd.get('lignes') or [],
                user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(
            self.get_serializer(run).data, status=status.HTTP_201_CREATED)

    def destroy(self, request, *args, **kwargs):
        run = self.get_object()  # scopé société par TenantMixin.
        if run.posted:
            return Response(
                {'detail': 'Une campagne postée ne peut être supprimée.'},
                status=status.HTTP_400_BAD_REQUEST)
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['post'])
    def proposer(self, request, pk=None):
        """YLEDG8 — Remplit la campagne BROUILLON depuis les échéances
        fournisseur dues (``?date_limite=YYYY-MM-DD`` optionnel). Idempotent :
        n'ajoute jamais deux fois la même facture fournisseur."""
        run = self.get_object()  # scopé société par TenantMixin.
        try:
            services.proposer_lignes_payment_run(
                run, date_limite=request.data.get('date_limite') or None)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        run.refresh_from_db()
        return Response(self.get_serializer(run).data)

    @action(detail=True, methods=['post'])
    def figer(self, request, pk=None):
        """Fige la proposition de règlement (brouillon → proposée, FG133)."""
        run = self.get_object()  # scopé société par TenantMixin.
        try:
            run = services.figer_payment_run(run)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(run).data)

    @action(detail=True, methods=['post'])
    def poster(self, request, pk=None):
        """Poste la campagne au grand livre EN LOT (FG133). Refusé en période
        close ; idempotent. Renvoie la campagne postée."""
        run = self.get_object()  # scopé société par TenantMixin.
        try:
            services.poster_payment_run(run, user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        run.refresh_from_db()
        return Response(self.get_serializer(run).data)

    @action(detail=True, methods=['post'])
    def approuver(self, request, pk=None):
        """NTTRE5 — Première approbation (approbateur ≠ créateur).

        Contrôle à 4 yeux : le premier approbateur ne peut pas être le créateur.
        Fige la campagne en ``en_attente_approbation``.
        """
        run = self.get_object()  # scopé société par TenantMixin.
        try:
            run = services.approuver_payment_run(run, request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(run).data)

    @action(detail=True, methods=['post'], url_path='approuver-final')
    def approuver_final(self, request, pk=None):
        """NTTRE5 — Seconde approbation (approbateur ≠ premier ≠ créateur).

        Rend la campagne éligible au posting (contrôle à 4 yeux complet).
        """
        run = self.get_object()  # scopé société par TenantMixin.
        try:
            run = services.approuver_final_payment_run(run, request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(run).data)

    @action(detail=True, methods=['get'], url_path='fichier-virement')
    def fichier_virement(self, request, pk=None):
        """Exporte l'ordre de virement bancaire de la campagne (FG134).

        Fichier CSV (un virement par ligne : bénéficiaire, RIB/IBAN, montant,
        devise, référence, motif). Refusé si la campagne n'est pas en virement
        ou si une ligne n'a aucune coordonnée bancaire. Lecture seule.
        """
        run = self.get_object()  # scopé société par TenantMixin.
        try:
            data = services.fichier_virement(run)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=';')
        writer.writerow(data['headers'])
        writer.writerows(data['rows'])
        resp = HttpResponse(
            buffer.getvalue(), content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = (
            f'attachment; filename="virements_{run.reference or run.id}.csv"')
        return resp


# ── NTTRE6 — Registre des pouvoirs bancaires (signataires autorisés) ────────

class PouvoirBancaireViewSet(_ComptaBaseViewSet):
    """Pouvoirs bancaires / signataires autorisés (NTTRE6).

    CRUD des habilitations de signature par compte de trésorerie, consommées par
    le workflow à 4 yeux (NTTRE5). Société scopée, posée côté serveur ;
    Admin/Responsable. La lecture reste ouverte ; la modification des signataires
    exige ``compta_gerer_pouvoirs_bancaires`` (NTTRE32) — octroyé par défaut au
    Responsable/Admin via le repli légal, restreint pour un rôle explicitement
    privé du code.
    """
    queryset = PouvoirBancaire.objects.select_related(
        'compte_tresorerie', 'utilisateur').all()
    serializer_class = PouvoirBancaireSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['titulaire_nom', 'id']

    def get_queryset(self):
        qs = super().get_queryset()
        params = self.request.query_params
        compte = params.get('compte_tresorerie')
        if compte:
            qs = qs.filter(compte_tresorerie_id=compte)
        statut = params.get('statut')
        if statut:
            qs = qs.filter(statut=statut)
        return qs

    def get_permissions(self):
        # NTTRE32 — la modification des signataires exige une permission dédiée
        # (distincte de la gestion trésorerie générique) ; la lecture reste
        # ouverte à Admin/Responsable.
        if self.action in (
                'create', 'update', 'partial_update', 'destroy', 'revoquer'):
            return [HasPermissionOrLegacy('compta_gerer_pouvoirs_bancaires')()]
        return super().get_permissions()

    @action(detail=True, methods=['post'])
    def revoquer(self, request, pk=None):
        """NTTRE6/42 — Révoque un pouvoir bancaire (statut → révoqué, audité)."""
        pouvoir = self.get_object()  # scopé société par TenantMixin.
        pouvoir.statut = PouvoirBancaire.Statut.REVOQUE
        pouvoir.save(update_fields=['statut', 'updated_at'])
        services._auditer_action_sensible(
            pouvoir, request.user, 'pouvoir_bancaire.revocation')
        return Response(self.get_serializer(pouvoir).data)


# ── NTTRE27 — Réglages trésorerie par société (singleton) ───────────────────

class ParametresTresorerieView(generics.RetrieveUpdateAPIView):
    """Réglages trésorerie de la société (NTTRE27), singleton auto-créé.

    ``GET`` renvoie le réglage (créé avec les valeurs par défaut à la première
    lecture — aucune régression) ; ``PATCH`` le met à jour. Société scopée, posée
    côté serveur ; Admin/Responsable.
    """
    http_method_names = ['get', 'patch', 'head', 'options']
    serializer_class = ParametresTresorerieSerializer
    permission_classes = [IsResponsableOrAdmin]

    def get_object(self):
        return services.get_parametres_tresorerie(self.request.user.company)


# ── FG135 — Notes de frais & remboursements employés ───────────────────────

class NoteFraisViewSet(_ComptaBaseViewSet):
    """Notes de frais & remboursements employés (FG135).

    Saisie d'une dépense avancée par un employé avec justificatif photo, puis
    cycle de validation/remboursement par les actions de service : ``soumettre``
    (brouillon → soumise), ``valider`` (poste la charge : débit classe 6 /
    crédit 4432 personnel-créditeur), ``rejeter`` (motif figé) et ``rembourser``
    (poste le paiement : débit 4432 / crédit trésorerie). Société scopée, posée
    côté serveur ; Admin/Responsable uniquement. Le justificatif accepte un
    upload de fichier (multipart) ; aucune écriture de statut directe par le
    corps.
    """
    queryset = NoteFrais.objects.select_related(
        'employe', 'compte_charge', 'compte_tresorerie',
        'ecriture_charge', 'ecriture_remboursement').all()
    serializer_class = NoteFraisSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['reference', 'motif']
    ordering_fields = ['date_frais', 'montant', 'statut', 'id']

    def get_queryset(self):
        qs = super().get_queryset()
        params = self.request.query_params
        statut = params.get('statut')
        if statut:
            qs = qs.filter(statut=statut)
        employe = params.get('employe')
        if employe:
            qs = qs.filter(employe_id=employe)
        categorie = params.get('categorie')
        if categorie:
            qs = qs.filter(categorie=categorie)
        return qs

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        vd = serializer.validated_data
        # XACC27 — doublon (même employé + date + montant) : WARNING joint à
        # la réponse, jamais bloquant (l'utilisateur peut confirmer volontairement).
        doublon = services.note_frais_doublon_possible(
            request.user.company, employe=vd['employe'],
            date_frais=vd['date_frais'], montant=vd['montant']).exists()
        try:
            note = services.creer_note_frais(
                request.user.company,
                employe=vd['employe'],
                date_frais=vd['date_frais'],
                montant=vd['montant'],
                motif=vd.get('motif', '') or '',
                categorie=vd.get('categorie'),
                justificatif=vd.get('justificatif'),
                compte_charge=vd.get('compte_charge'),
                refacturable=vd.get('refacturable') or False,
                taux_marge=vd.get('taux_marge'),
                client_refacturation_id=vd.get('client_refacturation_id'),
                chantier_refacturation=vd.get('chantier_refacturation', '') or '',
                user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        data = self.get_serializer(note).data
        data['doublon_possible'] = doublon
        return Response(data, status=status.HTTP_201_CREATED)

    def get_permissions(self):
        # YRBAC13 — cycle de validation décrit par la docstring de la classe :
        # ``soumettre``/``refacturer``/``ocr`` sont des saisies
        # (``compta_saisir``) ; ``valider``/``rejeter``/``rembourser`` postent
        # une écriture (charge ou paiement) et exigent ``compta_valider`` — le
        # même palier que le reste du cycle de validation comptable (COMPTA40).
        # Lecture (rapports/reçu/analyse) reste IsResponsableOrAdmin, inchangée.
        if self.action in ('refacturables', 'recu_pdf', 'analyse'):
            return [IsResponsableOrAdmin()]
        if self.action in ('soumettre', 'refacturer', 'ocr'):
            return [HasPermissionOrLegacy('compta_saisir')()]
        if self.action in ('valider', 'rejeter', 'rembourser'):
            return [HasPermissionOrLegacy('compta_valider')()]
        return super().get_permissions()

    @action(detail=False, methods=['get'], url_path='refacturables')
    def refacturables(self, request):
        """XACC28 — Notes refacturables VALIDÉES pas encore refacturées."""
        client_id = request.query_params.get('client')
        qs = selectors.frais_refacturables_non_factures(
            request.user.company, client_id=client_id)
        return Response(self.get_serializer(qs, many=True).data)

    @action(detail=False, methods=['post'], url_path='refacturer')
    def refacturer(self, request):
        """XACC28 — Génère des lignes de refacturation sur une facture existante.

        Corps : ``{facture_id, note_frais_ids: [...]}``. La facture doit
        appartenir à la société courante (résolue via
        ``apps.ventes.services.get_facture_or_none``, jamais un import de
        ``ventes.models``).
        """
        from apps.ventes.services import get_facture_or_none

        facture_id = request.data.get('facture_id')
        note_frais_ids = request.data.get('note_frais_ids') or []
        facture = get_facture_or_none(
            company=request.user.company, facture_id=facture_id)
        if facture is None:
            return Response(
                {'detail': 'Facture introuvable pour cette société.'},
                status=status.HTTP_404_NOT_FOUND)
        try:
            services.refacturer_frais_client(
                request.user.company, facture=facture,
                note_frais_ids=note_frais_ids, user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response({'facture_id': facture.id, 'refacture': True})

    @action(detail=False, methods=['post'])
    def ocr(self, request):
        """XACC27 — OCR du justificatif → champs pré-remplis (gated).

        Accepte une photo (``request.FILES['justificatif']``, multipart) et
        renvoie les champs extraits (``montant``, ``date_frais``, ``motif``)
        pour pré-remplir le formulaire — ne crée JAMAIS de note de frais et
        n'écrase JAMAIS une saisie manuelle déjà présente (fusion côté
        frontend). Sans clé OCR configurée : 503 explicite, la saisie
        manuelle reste intacte.
        """
        from .services import (
            extraire_justificatif_note_frais,
            mapper_justificatif_vers_note_frais,
        )

        photo = request.FILES.get('justificatif')
        if photo is None:
            return Response(
                {'justificatif': "Le fichier 'justificatif' est obligatoire."},
                status=status.HTTP_400_BAD_REQUEST)
        try:
            champs_bruts = extraire_justificatif_note_frais(
                photo.read(), mime=getattr(photo, 'content_type', '') or '')
        except RuntimeError as exc:
            return Response(
                {'detail': str(exc)}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        return Response(
            {'champs': mapper_justificatif_vers_note_frais(champs_bruts)})

    @action(detail=True, methods=['post'])
    def soumettre(self, request, pk=None):
        """Soumet la note pour validation (brouillon → soumise) — FG135."""
        note = self.get_object()  # scopée société par TenantMixin.
        try:
            note = services.soumettre_note_frais(note)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(note).data)

    @action(detail=True, methods=['post'])
    def valider(self, request, pk=None):
        """Valide la note et poste la charge au grand livre (FG135).

        Corps : ``{compte_charge?}``. Refusé en période close. Idempotent.
        """
        note = self.get_object()  # scopée société par TenantMixin.
        compte_charge = None
        cc_id = request.data.get('compte_charge')
        if cc_id:
            compte_charge = CompteComptable.objects.filter(
                company=request.user.company, id=cc_id).first()
            if compte_charge is None:
                return Response(
                    {'detail': 'Compte de charge inconnu.'},
                    status=status.HTTP_400_BAD_REQUEST)
        try:
            note = services.valider_note_frais(
                note, user=request.user, compte_charge=compte_charge)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(note).data)

    @action(detail=True, methods=['post'])
    def rejeter(self, request, pk=None):
        """Rejette une note soumise (soumise → rejetée), motif figé — FG135."""
        note = self.get_object()  # scopée société par TenantMixin.
        try:
            note = services.rejeter_note_frais(
                note, motif_rejet=request.data.get('motif_rejet', '') or '',
                user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(note).data)

    @action(detail=True, methods=['post'])
    def rembourser(self, request, pk=None):
        """Rembourse la note validée et poste le paiement (FG135).

        Corps : ``{compte_tresorerie, date_remboursement?, mode_remboursement?}``.
        Le compte de trésorerie payeur doit appartenir à la société. Refusé en
        période close. Idempotent.
        """
        note = self.get_object()  # scopée société par TenantMixin.
        treso = CompteTresorerie.objects.filter(
            company=request.user.company,
            id=request.data.get('compte_tresorerie')).first()
        if treso is None:
            return Response(
                {'detail': 'Compte de trésorerie inconnu.'},
                status=status.HTTP_400_BAD_REQUEST)
        try:
            note = services.rembourser_note_frais(
                note, compte_tresorerie=treso,
                date_remboursement=request.data.get('date_remboursement')
                or None,
                mode_remboursement=request.data.get('mode_remboursement')
                or None,
                user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(note).data)

    @action(detail=True, methods=['get'], url_path='recu-pdf')
    def recu_pdf(self, request, pk=None):
        """ZACC8 — Reçu PDF de remboursement de la note (entête société,
        détail, total en chiffres ET en lettres, mode/date, référence
        écriture). Une note NON remboursée -> 400 explicite. Company-scopée
        (404 cross-company via ``get_object``)."""
        note = self.get_object()  # scopée société par TenantMixin.
        if note.statut != NoteFrais.Statut.REMBOURSEE:
            return Response(
                {'detail': "Cette note de frais n'est pas encore "
                           "remboursée : aucun reçu à générer."},
                status=status.HTTP_400_BAD_REQUEST)
        from .pdf_notes_frais import render_recu_note_frais_pdf
        try:
            pdf_bytes = render_recu_note_frais_pdf(
                note, self._company_profile_for(request))
        except RuntimeError as exc:
            return Response(
                {'detail': str(exc)}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        resp = HttpResponse(pdf_bytes, content_type='application/pdf')
        resp['Content-Disposition'] = (
            f'attachment; filename="recu_note_frais_'
            f'{note.reference or note.id}.pdf"')
        return resp

    def _company_profile_for(self, request):
        try:
            from apps.parametres.models_company import CompanyProfile
            return CompanyProfile.get(company=request.user.company)
        except Exception:  # pragma: no cover - profil optionnel.
            return None

    @action(detail=False, methods=['get'])
    def analyse(self, request):
        """ZACC7 — Pivot des frais par employé/catégorie/mois
        (``?group_by=employe|categorie|mois``, défaut employe) sur la
        période ``?date_debut=``/``?date_fin=``. ``?export=xlsx`` télécharge."""
        params = request.query_params
        group_by = params.get('group_by') or 'employe'
        if group_by not in ('employe', 'categorie', 'mois'):
            group_by = 'employe'
        rapport = selectors.analyse_notes_frais(
            request.user.company,
            date_debut=params.get('date_debut') or None,
            date_fin=params.get('date_fin') or None,
            group_by=group_by)
        if params.get('export') == 'xlsx':
            from apps.records.xlsx import build_xlsx_response
            headers = [
                'cle', 'libelle', 'total', 'nombre', 'hors_politique_total']
            rows = [[ligne[h] for h in headers]
                    for ligne in rapport['lignes']]
            return build_xlsx_response(
                'analyse-notes-frais.xlsx', headers, rows,
                sheet_title='Analyse frais')
        return Response(rapport)


class RapportNoteFraisViewSet(_ComptaBaseViewSet):
    """Rapport regroupant N notes de frais d'un employé (ZACC6).

    ``creer`` (POST) crée le rapport ET rattache les notes désignées en un
    appel ; ``soumettre``/``valider``/``rembourser`` font suivre le cycle du
    RAPPORT (une seule écriture agrégée à la validation, un seul paiement au
    remboursement). Les notes ISOLÉES (sans rapport) gardent leur cycle
    individuel intact via ``NoteFraisViewSet``. Société scopée, posée côté
    serveur ; Admin/Responsable uniquement.
    """
    queryset = RapportNoteFrais.objects.select_related(
        'employe', 'compte_tresorerie', 'ecriture_charge',
        'ecriture_remboursement').prefetch_related('notes').all()
    serializer_class = RapportNoteFraisSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['reference', 'libelle']
    ordering_fields = ['date_creation', 'statut', 'id']

    def get_queryset(self):
        qs = super().get_queryset()
        statut = self.request.query_params.get('statut')
        if statut:
            qs = qs.filter(statut=statut)
        employe = self.request.query_params.get('employe')
        if employe:
            qs = qs.filter(employe_id=employe)
        return qs

    def create(self, request, *args, **kwargs):
        """Crée le rapport ET rattache les notes désignées en un appel.

        Corps : ``{employe, libelle?, note_frais_ids: [...]}``. Les notes
        doivent appartenir au même employé/société, être en brouillon/
        rejetées et ne pas déjà porter un rapport.
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        vd = serializer.validated_data
        note_frais_ids = request.data.get('note_frais_ids') or []
        try:
            rapport = services.creer_rapport_note_frais(
                request.user.company, employe=vd['employe'],
                note_frais_ids=note_frais_ids,
                libelle=vd.get('libelle', '') or '', user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(
            self.get_serializer(rapport).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def soumettre(self, request, pk=None):
        """Soumet le rapport (brouillon → soumis), soumet les notes rattachées
        encore en brouillon/rejetées."""
        rapport = self.get_object()  # scopée société par TenantMixin.
        try:
            rapport = services.soumettre_rapport_note_frais(rapport)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(rapport).data)

    @action(detail=True, methods=['post'])
    def valider(self, request, pk=None):
        """Valide le rapport et poste UNE écriture AGRÉGÉE (Σ des charges par
        compte / crédit 4432 unique). Idempotent."""
        rapport = self.get_object()  # scopée société par TenantMixin.
        try:
            rapport = services.valider_rapport_note_frais(
                rapport, user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(rapport).data)

    @action(detail=True, methods=['post'])
    def rembourser(self, request, pk=None):
        """Rembourse le rapport validé en UN SEUL paiement agrégé.

        Corps : ``{compte_tresorerie, date_remboursement?,
        mode_remboursement?}``. Refusé en période close. Idempotent : un
        rapport déjà remboursé n'est jamais re-postable.
        """
        rapport = self.get_object()  # scopée société par TenantMixin.
        treso = CompteTresorerie.objects.filter(
            company=request.user.company,
            id=request.data.get('compte_tresorerie')).first()
        if treso is None:
            return Response(
                {'detail': 'Compte de trésorerie inconnu.'},
                status=status.HTTP_400_BAD_REQUEST)
        try:
            rapport = services.rembourser_rapport_note_frais(
                rapport, compte_tresorerie=treso,
                date_remboursement=request.data.get('date_remboursement')
                or None,
                mode_remboursement=request.data.get('mode_remboursement')
                or None,
                user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(rapport).data)

    @action(detail=True, methods=['get'], url_path='recu-pdf')
    def recu_pdf(self, request, pk=None):
        """ZACC8 — Reçu PDF de remboursement du RAPPORT (détail de chaque
        note rattachée + total en chiffres ET en lettres). Un rapport NON
        remboursé -> 400 explicite. Company-scopé (404 cross-company via
        ``get_object``)."""
        rapport = self.get_object()  # scopée société par TenantMixin.
        if rapport.statut != RapportNoteFrais.Statut.REMBOURSE:
            return Response(
                {'detail': "Ce rapport de notes de frais n'est pas encore "
                           "remboursé : aucun reçu à générer."},
                status=status.HTTP_400_BAD_REQUEST)
        from .pdf_notes_frais import render_recu_rapport_note_frais_pdf
        try:
            from apps.parametres.models_company import CompanyProfile
            profile = CompanyProfile.get(company=request.user.company)
        except Exception:  # pragma: no cover - profil optionnel.
            profile = None
        try:
            pdf_bytes = render_recu_rapport_note_frais_pdf(rapport, profile)
        except RuntimeError as exc:
            return Response(
                {'detail': str(exc)}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        resp = HttpResponse(pdf_bytes, content_type='application/pdf')
        resp['Content-Disposition'] = (
            f'attachment; filename="recu_rapport_note_frais_'
            f'{rapport.reference or rapport.id}.pdf"')
        return resp


class PlafondNoteFraisViewSet(_ComptaBaseViewSet):
    """Plafonds de notes de frais par catégorie (XACC27).

    ``company`` posée côté serveur ; une seule ligne par catégorie
    (``get_or_create`` implicite via la contrainte d'unicité, 400 explicite
    en cas de doublon) ; Admin/Responsable uniquement."""
    queryset = PlafondNoteFrais.objects.all()
    serializer_class = PlafondNoteFraisSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['categorie', 'montant_max', 'id']

    def create(self, request, *args, **kwargs):
        try:
            return super().create(request, *args, **kwargs)
        except IntegrityError:
            return Response(
                {'detail': 'Un plafond existe déjà pour cette catégorie.'},
                status=status.HTTP_400_BAD_REQUEST)


class BaremeIndemniteViewSet(_ComptaBaseViewSet):
    """Barèmes d'indemnités km/per-diem chantier (FG136).

    Société scopée, posée côté serveur ; Admin/Responsable uniquement. Un seul
    barème peut être marqué « par défaut » actif par société : le poser sur un
    barème démote automatiquement l'ancien défaut (jamais d'erreur de
    contrainte).
    """
    queryset = BaremeIndemnite.objects.all()
    serializer_class = BaremeIndemniteSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['libelle', 'defaut', 'id']

    def _demote_other_defaults(self, company, keep_id=None):
        qs = BaremeIndemnite.objects.filter(
            company=company, defaut=True, actif=True)
        if keep_id is not None:
            qs = qs.exclude(id=keep_id)
        qs.update(defaut=False)

    def perform_create(self, serializer):
        company = self.request.user.company
        if serializer.validated_data.get('defaut') and \
                serializer.validated_data.get('actif', True):
            self._demote_other_defaults(company)
        serializer.save(company=company)

    def perform_update(self, serializer):
        company = self.request.user.company
        will_default = serializer.validated_data.get(
            'defaut', serializer.instance.defaut)
        will_actif = serializer.validated_data.get(
            'actif', serializer.instance.actif)
        if will_default and will_actif:
            self._demote_other_defaults(
                company, keep_id=serializer.instance.id)
        serializer.save(company=company)


class IndemniteChantierViewSet(_ComptaBaseViewSet):
    """Indemnités kilométriques & per-diem chantier (FG136).

    Saisie d'un déplacement chantier (employé, barème, GPS départ/chantier,
    jours, aller-retour) : la distance (haversine) et les montants sont calculés
    AUTOMATIQUEMENT côté serveur. Cycle de validation/remboursement par les
    actions de service : ``soumettre``, ``valider`` (poste la charge : débit
    classe 6 / crédit 4432), ``rejeter`` (motif figé), ``rembourser`` (poste le
    paiement). Société scopée, posée côté serveur ; Admin/Responsable
    uniquement ; aucune écriture de statut/montant directe par le corps.
    """
    queryset = IndemniteChantier.objects.select_related(
        'employe', 'bareme', 'compte_charge', 'compte_tresorerie',
        'ecriture_charge', 'ecriture_remboursement').all()
    serializer_class = IndemniteChantierSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['reference', 'libelle_chantier']
    ordering_fields = ['date_deplacement', 'montant_total', 'statut', 'id']

    def get_queryset(self):
        qs = super().get_queryset()
        params = self.request.query_params
        statut = params.get('statut')
        if statut:
            qs = qs.filter(statut=statut)
        employe = params.get('employe')
        if employe:
            qs = qs.filter(employe_id=employe)
        return qs

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        vd = serializer.validated_data
        try:
            indem = services.creer_indemnite_chantier(
                request.user.company,
                employe=vd['employe'],
                date_deplacement=vd['date_deplacement'],
                bareme=vd.get('bareme'),
                site_lat=vd.get('site_lat'),
                site_lng=vd.get('site_lng'),
                depart_lat=vd.get('depart_lat'),
                depart_lng=vd.get('depart_lng'),
                aller_retour=vd.get('aller_retour', True),
                nombre_jours=vd.get('nombre_jours', 1),
                libelle_chantier=vd.get('libelle_chantier', '') or '',
                user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(
            self.get_serializer(indem).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def soumettre(self, request, pk=None):
        """Soumet l'indemnité pour validation (brouillon → soumise) — FG136."""
        indem = self.get_object()  # scopée société par TenantMixin.
        try:
            indem = services.soumettre_indemnite_chantier(indem)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(indem).data)

    @action(detail=True, methods=['post'])
    def valider(self, request, pk=None):
        """Valide l'indemnité et poste la charge au grand livre (FG136).

        Corps : ``{compte_charge?}``. Refusé en période close. Idempotent.
        """
        indem = self.get_object()  # scopée société par TenantMixin.
        compte_charge = None
        cc_id = request.data.get('compte_charge')
        if cc_id:
            compte_charge = CompteComptable.objects.filter(
                company=request.user.company, id=cc_id).first()
            if compte_charge is None:
                return Response(
                    {'detail': 'Compte de charge inconnu.'},
                    status=status.HTTP_400_BAD_REQUEST)
        try:
            indem = services.valider_indemnite_chantier(
                indem, user=request.user, compte_charge=compte_charge)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(indem).data)

    @action(detail=True, methods=['post'])
    def rejeter(self, request, pk=None):
        """Rejette une indemnité soumise (soumise → rejetée) — FG136."""
        indem = self.get_object()  # scopée société par TenantMixin.
        try:
            indem = services.rejeter_indemnite_chantier(
                indem, motif_rejet=request.data.get('motif_rejet', '') or '',
                user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(indem).data)

    @action(detail=True, methods=['post'])
    def rembourser(self, request, pk=None):
        """Rembourse l'indemnité validée et poste le paiement (FG136).

        Corps : ``{compte_tresorerie, date_remboursement?}``. Le compte payeur
        doit appartenir à la société. Refusé en période close. Idempotent.
        """
        indem = self.get_object()  # scopée société par TenantMixin.
        treso = CompteTresorerie.objects.filter(
            company=request.user.company,
            id=request.data.get('compte_tresorerie')).first()
        if treso is None:
            return Response(
                {'detail': 'Compte de trésorerie inconnu.'},
                status=status.HTTP_400_BAD_REQUEST)
        try:
            indem = services.rembourser_indemnite_chantier(
                indem, compte_tresorerie=treso,
                date_remboursement=request.data.get('date_remboursement')
                or None,
                user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(indem).data)


# ── FG137 — Préparation de la déclaration de TVA ───────────────────────────

class DeclarationTVAViewSet(_ComptaBaseViewSet):
    """Préparation de la déclaration de TVA (FG137).

    Liste/consulte les déclarations préparées et expose deux actions :
    ``preparer`` (POST) calcule la TVA collectée − déductible sur une période
    depuis le grand livre et FIGE un snapshot ``DeclarationTVA`` (régime
    mensuel/trimestriel, méthode débit/encaissement, crédit antérieur), et
    ``export`` (GET) renvoie le détail en CSV. Société scopée, posée côté
    serveur ; Admin/Responsable uniquement. La création directe (POST sur la
    collection) passe par ``preparer`` afin que les montants soient toujours
    dérivés du GL (le corps ne peut jamais les imposer).
    """
    queryset = DeclarationTVA.objects.select_related('created_by').all()
    serializer_class = DeclarationTVASerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['reference', 'libelle']
    ordering_fields = ['date_fin', 'date_debut', 'tva_a_declarer', 'id']

    def get_queryset(self):
        qs = super().get_queryset()
        params = self.request.query_params
        regime = params.get('regime')
        if regime:
            qs = qs.filter(regime=regime)
        statut = params.get('statut')
        if statut:
            qs = qs.filter(statut=statut)
        return qs

    def _preparer(self, request):
        """Valide le corps et fige la déclaration depuis le GL (FG137)."""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        vd = serializer.validated_data
        declaration = services.preparer_declaration_tva(
            request.user.company,
            date_debut=vd['date_debut'],
            date_fin=vd['date_fin'],
            regime=vd.get('regime') or DeclarationTVA.Regime.MENSUEL,
            methode=vd.get('methode') or DeclarationTVA.Methode.DEBIT,
            credit_anterieur=vd.get('credit_anterieur') or 0,
            libelle=vd.get('libelle', '') or '',
            validees_seulement=request.query_params.get('validees') == '1',
            user=request.user)
        return declaration

    def create(self, request, *args, **kwargs):
        """POST sur la collection = préparation (montants dérivés du GL)."""
        try:
            declaration = self._preparer(request)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(
            self.get_serializer(declaration).data,
            status=status.HTTP_201_CREATED)

    def get_permissions(self):
        # YRBAC13 — ``preparer`` calcule/fige un snapshot depuis le GL (une
        # saisie, ``compta_saisir``) ; ``deposer`` est le dépôt OFFICIEL de la
        # déclaration (action irréversible côté administration fiscale),
        # gardée par ``compta_valider`` (même palier que le reste du cycle de
        # validation, COMPTA40). Lecture (export/comparatif/bordereau) reste
        # IsResponsableOrAdmin, inchangée.
        if self.action in ('export', 'comparatif', 'bordereau_pdf'):
            return [IsResponsableOrAdmin()]
        if self.action == 'preparer':
            return [HasPermissionOrLegacy('compta_saisir')()]
        if self.action == 'deposer':
            return [HasPermissionOrLegacy('compta_valider')()]
        return super().get_permissions()

    @action(detail=False, methods=['post'])
    def preparer(self, request):
        """Prépare et fige une déclaration de TVA sur une période (FG137).

        Corps : ``{date_debut, date_fin, regime?, methode?, credit_anterieur?,
        libelle?}``. La TVA collectée/déductible et le montant à déclarer sont
        calculés depuis le grand livre côté serveur. Renvoie la déclaration figée.
        """
        try:
            declaration = self._preparer(request)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(
            self.get_serializer(declaration).data,
            status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['get'])
    def export(self, request, pk=None):
        """Exporte la déclaration de TVA en CSV (FG137).

        Une ligne par poste (TVA collectée, TVA déductible, crédit antérieur,
        TVA à déclarer, crédit reportable) + l'en-tête de période/régime/méthode.
        Lecture seule, scopée société.
        """
        decl = self.get_object()  # scopée société par TenantMixin.
        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=';')
        writer.writerow(['Déclaration de TVA', decl.reference or decl.id])
        writer.writerow(['Période', f'{decl.date_debut} → {decl.date_fin}'])
        writer.writerow(['Régime', decl.get_regime_display()])
        writer.writerow(['Méthode', decl.get_methode_display()])
        writer.writerow([])
        writer.writerow(['Poste', 'Montant'])
        writer.writerow(['TVA collectée', decl.tva_collectee])
        writer.writerow(['TVA déductible', decl.tva_deductible])
        writer.writerow(['Crédit de TVA antérieur', decl.credit_anterieur])
        writer.writerow(['TVA à déclarer', decl.tva_a_declarer])
        writer.writerow(['Crédit de TVA reportable', decl.credit_reportable])
        resp = HttpResponse(
            buffer.getvalue(), content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = (
            f'attachment; filename="declaration_tva_'
            f'{decl.reference or decl.id}.csv"')
        return resp

    @action(detail=True, methods=['post'])
    def deposer(self, request, pk=None):
        """Dépose la déclaration : DEPOSEE + son obligation fiscale (XACC9)."""
        decl = self.get_object()  # scopée société par TenantMixin.
        decl = services.deposer_declaration_tva(decl)
        return Response(self.get_serializer(decl).data)

    @action(detail=True, methods=['get'])
    def comparatif(self, request, pk=None):
        """ZACC10 — Comparatif M-1 de la déclaration (collecté/déductible/net
        vs période précédente, écart %). Recalculé LIVE depuis le GL sur la
        même période que la déclaration figée (jamais un recalcul de la
        déclaration elle-même) ; ``defaut`` sans ce paramètre reste inchangé —
        l'action existe précisément pour exposer le mode ``comparer=1``.
        """
        decl = self.get_object()  # scopée société par TenantMixin.
        data = selectors.preparer_declaration_tva(
            request.user.company, date_debut=decl.date_debut,
            date_fin=decl.date_fin, regime=decl.regime, methode=decl.methode,
            credit_anterieur=decl.credit_anterieur, comparer=True,
            date_debut_m1=request.query_params.get('date_debut_m1') or None,
            date_fin_m1=request.query_params.get('date_fin_m1') or None)
        return Response(data)

    @action(detail=True, methods=['get'], url_path='bordereau-pdf')
    def bordereau_pdf(self, request, pk=None):
        """ZACC10 — Bordereau PDF de la déclaration de TVA (entête société
        IF/ICE, régime débit/encaissement, TVA collectée par taux, TVA
        déductible sur biens/services/immo, crédit reporté, net à payer).
        Reprend le snapshot FIGÉ de la déclaration — AUCUNE télétransmission
        (le gate DGI reste XFAC29). Company-scopée (404 cross-company)."""
        from datetime import date as _date
        from html import escape
        decl = self.get_object()  # scopée société par TenantMixin.
        from .pdf_notes_frais import _entete_societe_html, _fmt
        from .pdf_etats import _wrap, _html_to_pdf
        try:
            from apps.parametres.models_company import CompanyProfile
            profile = CompanyProfile.get(company=request.user.company)
        except Exception:  # pragma: no cover - profil optionnel.
            profile = None
        entete = _entete_societe_html(profile)
        corps = f"""
        <table><tbody>
        <tr><td>Régime</td><td class="montant">
        {escape(decl.get_regime_display())}</td></tr>
        <tr><td>Méthode</td><td class="montant">
        {escape(decl.get_methode_display())}</td></tr>
        <tr><td>TVA collectée</td><td class="montant">
        {_fmt(decl.tva_collectee)}</td></tr>
        <tr><td>TVA déductible</td><td class="montant">
        {_fmt(decl.tva_deductible)}</td></tr>
        <tr><td>Crédit de TVA antérieur</td><td class="montant">
        {_fmt(decl.credit_anterieur)}</td></tr>
        <tr class="total-row"><td>TVA à déclarer</td><td class="montant">
        {_fmt(decl.tva_a_declarer)}</td></tr>
        <tr><td>Crédit de TVA reportable</td><td class="montant">
        {_fmt(decl.credit_reportable)}</td></tr>
        </tbody></table>"""
        html = _wrap(
            entete, 'Bordereau de déclaration de TVA',
            f'Période du {decl.date_debut} au {decl.date_fin}', corps,
            _date.today())
        try:
            pdf_bytes = _html_to_pdf(html)
        except RuntimeError as exc:
            return Response(
                {'detail': str(exc)}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        resp = HttpResponse(pdf_bytes, content_type='application/pdf')
        resp['Content-Disposition'] = (
            f'attachment; filename="bordereau_tva_'
            f'{decl.reference or decl.id}.pdf"')
        return resp


class RetenueSourceViewSet(_ComptaBaseViewSet):
    """Retenue à la source (RAS) sur honoraires/prestations (FG139).

    Enregistre/consulte les retenues à la source et expose le bordereau de
    versement : ``bordereau`` (GET) regroupe les RAS d'une période par
    prestataire et donne le total à reverser au Trésor (``export=csv`` pour le
    CSV), ``export`` (GET, détail) liste les retenues de la période en CSV, et
    ``verser`` (POST) marque une retenue comme versée. La création POST calcule
    le montant retenu côté serveur (base × taux %) ; le corps ne peut jamais
    l'imposer. Société scopée, posée côté serveur ; Admin/Responsable uniquement.
    """
    queryset = RetenueSource.objects.select_related('created_by').all()
    serializer_class = RetenueSourceSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['reference', 'piece', 'tiers_nom', 'identifiant_fiscal']
    ordering_fields = ['date_piece', 'montant', 'base', 'id']

    def get_queryset(self):
        qs = super().get_queryset()
        params = self.request.query_params
        statut = params.get('statut')
        if statut:
            qs = qs.filter(statut=statut)
        type_prestation = params.get('type_prestation')
        if type_prestation:
            qs = qs.filter(type_prestation=type_prestation)
        date_debut = params.get('date_debut')
        if date_debut:
            qs = qs.filter(date_piece__gte=date_debut)
        date_fin = params.get('date_fin')
        if date_fin:
            qs = qs.filter(date_piece__lte=date_fin)
        return qs

    def create(self, request, *args, **kwargs):
        """POST = enregistre une RAS (montant retenu dérivé côté serveur)."""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        vd = serializer.validated_data
        try:
            ras = services.enregistrer_retenue_source(
                request.user.company,
                date_piece=vd['date_piece'],
                base=vd.get('base') or 0,
                taux=vd.get('taux'),
                type_prestation=vd.get('type_prestation'),
                tiers_type=vd.get('tiers_type', '') or '',
                tiers_id=vd.get('tiers_id'),
                tiers_nom=vd.get('tiers_nom', '') or '',
                identifiant_fiscal=vd.get('identifiant_fiscal', '') or '',
                piece=vd.get('piece', '') or '',
                libelle=vd.get('libelle', '') or '',
                user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(
            self.get_serializer(ras).data, status=status.HTTP_201_CREATED)

    def get_permissions(self):
        # YRBAC13 — ``verser`` marque un versement RÉEL au Trésor (paiement),
        # gardée par ``compta_valider`` (même palier que les autres actions de
        # paiement/validation du cycle comptable, COMPTA40). Les rapports en
        # lecture (bordereau/export/attestations) restent IsResponsableOrAdmin,
        # inchangés.
        if self.action in ('bordereau', 'export', 'attestation',
                           'attestation_annuelle'):
            return [IsResponsableOrAdmin()]
        if self.action == 'verser':
            return [HasPermissionOrLegacy('compta_valider')()]
        return super().get_permissions()

    @action(detail=True, methods=['post'])
    def verser(self, request, pk=None):
        """Marque la retenue comme versée au Trésor (FG139)."""
        ras = self.get_object()  # scopée société par TenantMixin.
        services.marquer_ras_versee(ras)
        return Response(self.get_serializer(ras).data)

    def _periode(self, request):
        params = request.query_params
        return {
            'date_debut': params.get('date_debut') or None,
            'date_fin': params.get('date_fin') or None,
            'statut': params.get('statut') or None,
        }

    @action(detail=False, methods=['get'])
    def bordereau(self, request):
        """Bordereau de versement de la RAS : totaux par prestataire (FG139).

        Regroupe les retenues de la période ``[date_debut ; date_fin]`` (bornées
        sur la date de pièce) par prestataire et donne le total à reverser au
        Trésor. Paramètres : ``date_debut`` / ``date_fin`` / ``statut`` et
        ``export=csv`` pour le CSV. Lecture seule, scopée société,
        Admin/Responsable.
        """
        periode = self._periode(request)
        data = selectors.bordereau_versement_ras(
            request.user.company, **periode)
        if request.query_params.get('export') == 'csv':
            return self._export_bordereau_csv(data)
        return Response(data)

    @staticmethod
    def _export_bordereau_csv(data):
        """Sérialise le bordereau de versement de la RAS (FG139) en CSV."""
        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=';')
        writer.writerow(['Bordereau de versement — Retenue à la source'])
        writer.writerow(
            ['Période', f"{data['date_debut'] or ''} → {data['date_fin'] or ''}"])
        writer.writerow([])
        writer.writerow(
            ['Prestataire', 'Identifiant fiscal', 'Nb pièces', 'Base',
             'Montant retenu'])
        for ligne in data['lignes']:
            writer.writerow([
                ligne['tiers_nom'], ligne['identifiant_fiscal'],
                ligne['nb_pieces'], ligne['base'], ligne['montant']])
        writer.writerow([])
        writer.writerow(
            ['Totaux', '', data['totaux']['nb_pieces'],
             data['totaux']['base'], data['totaux']['montant']])
        writer.writerow(['Total à verser', '', '', '', data['total_a_verser']])
        resp = HttpResponse(
            buffer.getvalue(), content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = (
            'attachment; filename="bordereau_versement_ras_'
            f"{data['date_debut'] or 'periode'}_{data['date_fin'] or ''}.csv\"")
        return resp

    @action(detail=False, methods=['get'])
    def export(self, request):
        """Exporte le détail des RAS d'une période en CSV (FG139).

        Une ligne par retenue (date, pièce, prestataire, IF, type, base, taux,
        montant retenu, net à payer). Paramètres : ``date_debut`` / ``date_fin``
        / ``statut``. Lecture seule, scopée société, Admin/Responsable.
        """
        periode = self._periode(request)
        data = selectors.retenues_source_periode(
            request.user.company, **periode)
        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=';')
        writer.writerow(['Retenues à la source (détail)'])
        writer.writerow(
            ['Période', f"{data['date_debut'] or ''} → {data['date_fin'] or ''}"])
        writer.writerow([])
        writer.writerow(
            ['Date', 'Référence', 'Pièce', 'Prestataire', 'Identifiant fiscal',
             'Type', 'Base', 'Taux %', 'Montant retenu', 'Net à payer'])
        for ligne in data['lignes']:
            writer.writerow([
                ligne['date_piece'], ligne['reference'], ligne['piece'],
                ligne['tiers_nom'], ligne['identifiant_fiscal'],
                ligne['type_prestation'], ligne['base'], ligne['taux'],
                ligne['montant'], ligne['net_a_payer']])
        writer.writerow([])
        writer.writerow(
            ['Totaux', '', '', '', '', '', data['totaux']['base'], '',
             data['totaux']['montant'], data['totaux']['net']])
        resp = HttpResponse(
            buffer.getvalue(), content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = (
            'attachment; filename="retenues_source_'
            f"{data['date_debut'] or 'periode'}_{data['date_fin'] or ''}.csv\"")
        return resp

    @action(detail=True, methods=['get'])
    def attestation(self, request, pk=None):
        """XACC35 — Attestation PDF de RAS pour CE versement (par prestataire).

        Rendu WeasyPrint sobre (moteur légal existant, PAS le quote engine) :
        identité société (N27), tiers (IF/ICE), base, taux, montant retenu,
        période — chiffres identiques au snapshot figé (ledger FG139)."""
        from .pdf_ras import render_attestation_retenue_pdf

        ras = self.get_object()  # scopée société par TenantMixin.
        try:
            from apps.parametres.models_company import CompanyProfile
            profile = CompanyProfile.get(company=request.user.company)
        except Exception:  # pragma: no cover - profil optionnel, jamais bloquant.
            profile = None
        try:
            pdf_bytes = render_attestation_retenue_pdf(ras, profile)
        except RuntimeError as exc:
            return Response(
                {'detail': str(exc)}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        resp = HttpResponse(pdf_bytes, content_type='application/pdf')
        resp['Content-Disposition'] = (
            f'attachment; filename="attestation_ras_{ras.reference or ras.id}.pdf"')
        return resp

    @action(detail=False, methods=['get'], url_path='attestation-annuelle')
    def attestation_annuelle(self, request):
        """XACC35 — Cumul annuel de RAS d'un prestataire (PDF).

        Paramètres : ``?tiers=<tiers_id>&annee=<YYYY>``. Le cumul reprend
        TOUTES les ``RetenueSource`` du tiers pour l'année, tous types de
        prestation confondus, company-scopées."""
        from .pdf_ras import render_attestation_annuelle_pdf

        tiers_id = request.query_params.get('tiers')
        annee = request.query_params.get('annee')
        if not tiers_id or not annee:
            return Response(
                {'detail': 'Les paramètres tiers et annee sont obligatoires.'},
                status=status.HTTP_400_BAD_REQUEST)
        try:
            annee_int = int(annee)
        except ValueError:
            return Response(
                {'detail': 'annee doit être une année valide.'},
                status=status.HTTP_400_BAD_REQUEST)
        retenues = list(RetenueSource.objects.filter(
            company=request.user.company, tiers_id=tiers_id,
            date_piece__year=annee_int,
        ).order_by('date_piece', 'id'))
        if not retenues:
            return Response(
                {'detail': 'Aucune retenue trouvée pour ce prestataire/année.'},
                status=status.HTTP_404_NOT_FOUND)
        try:
            from apps.parametres.models_company import CompanyProfile
            profile = CompanyProfile.get(company=request.user.company)
        except Exception:  # pragma: no cover - profil optionnel, jamais bloquant.
            profile = None
        try:
            pdf_bytes = render_attestation_annuelle_pdf(
                retenues, retenues[0].tiers_nom, annee_int, profile)
        except RuntimeError as exc:
            return Response(
                {'detail': str(exc)}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
        resp = HttpResponse(pdf_bytes, content_type='application/pdf')
        resp['Content-Disposition'] = (
            f'attachment; filename="attestation_ras_annuelle_{tiers_id}_'
            f'{annee_int}.pdf"')
        return resp


class TimbreFiscalViewSet(_ComptaBaseViewSet):
    """Droit de timbre sur les encaissements en espèces (FG144).

    Enregistre/consulte les droits de timbre de quittance : la création POST
    calcule le ``montant`` côté serveur (max(base × taux %, minimum)) — le corps
    ne peut jamais l'imposer — et N'ENREGISTRE un timbre QUE pour un règlement en
    espèces ; tout autre mode est EXONÉRÉ et renvoie 400. ``verser`` (POST) marque
    un timbre comme versé ; ``export`` (GET, détail) liste les timbres de la
    période en CSV (``?export=csv`` jamais ``?format=``). Société scopée, posée
    côté serveur ; Admin/Responsable uniquement.
    """
    queryset = TimbreFiscal.objects.select_related('created_by').all()
    serializer_class = TimbreFiscalSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['reference', 'facture_ref', 'tiers_nom']
    ordering_fields = ['date_encaissement', 'montant', 'base', 'id']

    def get_queryset(self):
        qs = super().get_queryset()
        params = self.request.query_params
        statut = params.get('statut')
        if statut:
            qs = qs.filter(statut=statut)
        date_debut = params.get('date_debut')
        if date_debut:
            qs = qs.filter(date_encaissement__gte=date_debut)
        date_fin = params.get('date_fin')
        if date_fin:
            qs = qs.filter(date_encaissement__lte=date_fin)
        return qs

    def create(self, request, *args, **kwargs):
        """POST = enregistre un droit de timbre (montant dérivé côté serveur).

        Un règlement non espèces est exonéré : on renvoie 400 sans créer de
        timbre."""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        vd = serializer.validated_data
        try:
            timbre = services.enregistrer_timbre_fiscal(
                request.user.company,
                date_encaissement=vd['date_encaissement'],
                base=vd.get('base') or 0,
                mode_reglement=vd.get('mode_reglement') or services.MODE_ESPECES,
                taux=vd.get('taux'),
                minimum=vd.get('minimum'),
                paiement_id=vd.get('paiement_id'),
                facture_ref=vd.get('facture_ref', '') or '',
                tiers_type=vd.get('tiers_type', '') or '',
                tiers_id=vd.get('tiers_id'),
                tiers_nom=vd.get('tiers_nom', '') or '',
                libelle=vd.get('libelle', '') or '',
                user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        if timbre is None:
            return Response(
                {'detail': "Le droit de timbre ne s'applique qu'aux "
                           "encaissements en espèces."},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(
            self.get_serializer(timbre).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def verser(self, request, pk=None):
        """Marque le droit de timbre comme versé au Trésor (FG144)."""
        timbre = self.get_object()  # scopé société par TenantMixin.
        services.marquer_timbre_verse(timbre)
        return Response(self.get_serializer(timbre).data)

    @action(detail=False, methods=['get'])
    def export(self, request):
        """Exporte le détail des droits de timbre d'une période en CSV (FG144).

        Une ligne par timbre (date, référence, facture, payeur, base, taux,
        minimum, droit de timbre). Paramètres : ``date_debut`` / ``date_fin`` /
        ``statut``. Lecture seule, scopée société, Admin/Responsable.
        """
        params = request.query_params
        data = selectors.timbres_fiscaux_periode(
            request.user.company,
            date_debut=params.get('date_debut') or None,
            date_fin=params.get('date_fin') or None,
            statut=params.get('statut') or None)
        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=';')
        writer.writerow(['Droits de timbre sur encaissements espèces (détail)'])
        writer.writerow(
            ['Période', f"{data['date_debut'] or ''} → {data['date_fin'] or ''}"])
        writer.writerow([])
        writer.writerow(
            ['Date', 'Référence', 'Facture', 'Payeur', 'Base', 'Taux %',
             'Minimum', 'Droit de timbre'])
        for ligne in data['lignes']:
            writer.writerow([
                ligne['date_encaissement'], ligne['reference'],
                ligne['facture_ref'], ligne['tiers_nom'], ligne['base'],
                ligne['taux'], ligne['minimum'], ligne['montant']])
        writer.writerow([])
        writer.writerow(
            ['Totaux', '', '', '', data['totaux']['base'], '', '',
             data['totaux']['montant']])
        writer.writerow(
            ['Total à verser', '', '', '', '', '', '', data['total_a_verser']])
        resp = HttpResponse(
            buffer.getvalue(), content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = (
            'attachment; filename="timbres_fiscaux_'
            f"{data['date_debut'] or 'periode'}_{data['date_fin'] or ''}.csv\"")
        return resp


class RetenueGarantieViewSet(_ComptaBaseViewSet):
    """Retenue de garantie (RG / bonne fin) sur les marchés (FG145).

    Enregistre/consulte les retenues de garantie prélevées sur les décomptes :
    la création POST calcule le ``montant`` retenu côté serveur (base × taux %) —
    le corps ne peut jamais l'imposer. ``liberer`` (POST) libère la RG à sa levée
    (statut « libérée » + date de libération) ; ``echeances`` (GET) liste les RG
    dont la levée prévue arrive sous ``jours`` (défaut 30, ``?export=csv`` pour le
    CSV). Filtres ``statut`` ; société scopée posée côté serveur ;
    Admin/Responsable uniquement.
    """
    queryset = RetenueGarantie.objects.select_related('created_by').all()
    serializer_class = RetenueGarantieSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['reference', 'marche_ref', 'facture_ref', 'tiers_nom']
    ordering_fields = [
        'date_constitution', 'date_levee_prevue', 'montant', 'base', 'id']

    def get_queryset(self):
        qs = super().get_queryset()
        params = self.request.query_params
        statut = params.get('statut')
        if statut:
            qs = qs.filter(statut=statut)
        date_debut = params.get('date_debut')
        if date_debut:
            qs = qs.filter(date_constitution__gte=date_debut)
        date_fin = params.get('date_fin')
        if date_fin:
            qs = qs.filter(date_constitution__lte=date_fin)
        return qs

    def create(self, request, *args, **kwargs):
        """POST = enregistre une RG (montant retenu dérivé côté serveur)."""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        vd = serializer.validated_data
        try:
            rg = services.enregistrer_retenue_garantie(
                request.user.company,
                date_constitution=vd['date_constitution'],
                base=vd.get('base') or 0,
                taux=vd.get('taux'),
                marche_ref=vd.get('marche_ref', '') or '',
                facture_id=vd.get('facture_id'),
                facture_ref=vd.get('facture_ref', '') or '',
                tiers_type=vd.get('tiers_type', '') or '',
                tiers_id=vd.get('tiers_id'),
                tiers_nom=vd.get('tiers_nom', '') or '',
                date_levee_prevue=vd.get('date_levee_prevue'),
                libelle=vd.get('libelle', '') or '',
                user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(
            self.get_serializer(rg).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def liberer(self, request, pk=None):
        """Libère (restitue) la retenue de garantie à sa levée (FG145)."""
        rg = self.get_object()  # scopée société par TenantMixin.
        date_liberation = request.data.get('date_liberation') or None
        services.liberer_retenue_garantie(
            rg, date_liberation=date_liberation)
        return Response(self.get_serializer(rg).data)

    @action(detail=False, methods=['get'])
    def echeances(self, request):
        """RG dont la levée prévue arrive à échéance sous ``jours`` (FG145).

        Paramètre ``jours`` (défaut 30) ; ``?export=csv`` pour le CSV. Lecture
        seule, scopée société, Admin/Responsable.
        """
        try:
            jours = int(request.query_params.get('jours') or 30)
        except (TypeError, ValueError):
            jours = 30
        data = selectors.retenues_garantie_a_echeance(
            request.user.company, jours=jours)
        if request.query_params.get('export') == 'csv':
            buffer = io.StringIO()
            writer = csv.writer(buffer, delimiter=';')
            writer.writerow(
                ['Retenues de garantie arrivant à échéance',
                 f"≤ {data['jours']} j"])
            writer.writerow([])
            writer.writerow(
                ['Référence', 'Marché', 'Facture', 'Maître d\'ouvrage', 'Base',
                 'Taux %', 'Montant retenu', 'Constitution', 'Levée prévue',
                 'En retard'])
            for ligne in data['lignes']:
                writer.writerow([
                    ligne['reference'], ligne['marche_ref'],
                    ligne['facture_ref'], ligne['tiers_nom'], ligne['base'],
                    ligne['taux'], ligne['montant'],
                    ligne['date_constitution'], ligne['date_levee_prevue'],
                    'Oui' if ligne['en_retard'] else 'Non'])
            writer.writerow([])
            writer.writerow(['Total', '', '', '', '', '', data['total_montant']])
            resp = HttpResponse(
                buffer.getvalue(), content_type='text/csv; charset=utf-8')
            resp['Content-Disposition'] = (
                'attachment; filename="retenues_garantie_echeances.csv"')
            return resp
        return Response(data)


class CautionBancaireViewSet(_ComptaBaseViewSet):
    """Cautions / garanties bancaires émises sur les marchés (FG145).

    Enregistre/consulte les cautions (provisoire/définitive/retenue de
    garantie/restitution) : ``company`` / ``reference`` / ``statut`` sont posés
    côté serveur. ``mainlevee`` (POST) lève la caution (mainlevée, ou restitution
    si ``restituee=true``) avec la date de mainlevée ; ``echeances`` (GET) liste
    les cautions actives arrivant à échéance sous ``jours`` (défaut 30,
    ``?export=csv`` pour le CSV). Filtres ``statut`` / ``type_caution`` ; société
    scopée ; Admin/Responsable uniquement.
    """
    queryset = CautionBancaire.objects.select_related('created_by').all()
    serializer_class = CautionBancaireSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['reference', 'marche_ref', 'tiers_nom', 'banque']
    ordering_fields = [
        'date_emission', 'date_echeance', 'montant', 'id']

    def get_queryset(self):
        qs = super().get_queryset()
        params = self.request.query_params
        statut = params.get('statut')
        if statut:
            qs = qs.filter(statut=statut)
        type_caution = params.get('type_caution')
        if type_caution:
            qs = qs.filter(type_caution=type_caution)
        date_debut = params.get('date_debut')
        if date_debut:
            qs = qs.filter(date_emission__gte=date_debut)
        date_fin = params.get('date_fin')
        if date_fin:
            qs = qs.filter(date_emission__lte=date_fin)
        return qs

    def create(self, request, *args, **kwargs):
        """POST = enregistre une caution bancaire (référence côté serveur)."""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        vd = serializer.validated_data
        try:
            caution = services.enregistrer_caution_bancaire(
                request.user.company,
                type_caution=vd.get('type_caution'),
                date_emission=vd['date_emission'],
                montant=vd.get('montant') or 0,
                banque=vd.get('banque', '') or '',
                marche_ref=vd.get('marche_ref', '') or '',
                tiers_nom=vd.get('tiers_nom', '') or '',
                date_echeance=vd.get('date_echeance'),
                libelle=vd.get('libelle', '') or '',
                user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(
            self.get_serializer(caution).data,
            status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def mainlevee(self, request, pk=None):
        """Lève (mainlevée) ou restitue la caution bancaire (FG145)."""
        caution = self.get_object()  # scopée société par TenantMixin.
        date_mainlevee = request.data.get('date_mainlevee') or None
        restituee = bool(request.data.get('restituee'))
        services.mainlevee_caution_bancaire(
            caution, date_mainlevee=date_mainlevee, restituee=restituee)
        return Response(self.get_serializer(caution).data)

    @action(detail=False, methods=['get'])
    def echeances(self, request):
        """Cautions actives arrivant à échéance sous ``jours`` (FG145).

        Paramètre ``jours`` (défaut 30) ; ``?export=csv`` pour le CSV. Lecture
        seule, scopée société, Admin/Responsable.
        """
        try:
            jours = int(request.query_params.get('jours') or 30)
        except (TypeError, ValueError):
            jours = 30
        data = selectors.cautions_a_echeance(
            request.user.company, jours=jours)
        if request.query_params.get('export') == 'csv':
            buffer = io.StringIO()
            writer = csv.writer(buffer, delimiter=';')
            writer.writerow(
                ['Cautions bancaires arrivant à échéance',
                 f"≤ {data['jours']} j"])
            writer.writerow([])
            writer.writerow(
                ['Référence', 'Type', 'Marché', 'Bénéficiaire', 'Banque',
                 'Montant', 'Émission', 'Échéance', 'En retard'])
            for ligne in data['lignes']:
                writer.writerow([
                    ligne['reference'], ligne['type_caution'],
                    ligne['marche_ref'], ligne['tiers_nom'], ligne['banque'],
                    ligne['montant'], ligne['date_emission'],
                    ligne['date_echeance'],
                    'Oui' if ligne['en_retard'] else 'Non'])
            writer.writerow([])
            writer.writerow(['Total', '', '', '', '', data['total_montant']])
            resp = HttpResponse(
                buffer.getvalue(), content_type='text/csv; charset=utf-8')
            resp['Content-Disposition'] = (
                'attachment; filename="cautions_bancaires_echeances.csv"')
            return resp
        return Response(data)


class ContratAvancementViewSet(_ComptaBaseViewSet):
    """Contrats reconnus au pourcentage d'avancement (FG146).

    Crée/consulte les contrats pluri-tranches ; ``company`` / ``reference`` /
    ``statut`` sont posés côté serveur. ``constater`` (POST) ajoute un constat
    d'avancement et reconnaît le CA cumulé (écriture OD) ; ``avancement``
    (GET) renvoie la synthèse d'avancement/marge. Société scopée ;
    Admin/Responsable uniquement.
    """
    queryset = ContratAvancement.objects.select_related('created_by').all()
    serializer_class = ContratAvancementSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['reference', 'libelle', 'chantier_ref', 'marche_ref',
                     'client_nom']
    ordering_fields = ['date_creation', 'revenu_total', 'id']

    def get_queryset(self):
        qs = super().get_queryset()
        statut = self.request.query_params.get('statut')
        if statut:
            qs = qs.filter(statut=statut)
        return qs

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        vd = serializer.validated_data
        try:
            contrat = services.creer_contrat_avancement(
                request.user.company,
                revenu_total=vd.get('revenu_total') or 0,
                cout_total_estime=vd.get('cout_total_estime') or 0,
                methode=vd.get('methode'),
                libelle=vd.get('libelle', '') or '',
                chantier_ref=vd.get('chantier_ref', '') or '',
                marche_ref=vd.get('marche_ref', '') or '',
                client_id=vd.get('client_id'),
                client_nom=vd.get('client_nom', '') or '',
                date_debut=vd.get('date_debut'),
                date_fin_prevue=vd.get('date_fin_prevue'),
                user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(
            self.get_serializer(contrat).data,
            status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def constater(self, request, pk=None):
        """Ajoute un constat d'avancement et reconnaît le CA cumulé (FG146)."""
        contrat = self.get_object()  # scopé société par TenantMixin.
        date_arrete = request.data.get('date_arrete')
        if not date_arrete:
            return Response(
                {'detail': "La date d'arrêté est obligatoire."},
                status=status.HTTP_400_BAD_REQUEST)
        poster = request.data.get('poster', True)
        if isinstance(poster, str):
            poster = poster.lower() not in ('false', '0', 'no', '')
        try:
            constat = services.constater_avancement(
                contrat, date_arrete=date_arrete,
                pourcentage=request.data.get('pourcentage'),
                cout_engage_cumule=request.data.get('cout_engage_cumule'),
                libelle=request.data.get('libelle', '') or '',
                poster=bool(poster), user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(
            AvancementRevenuSerializer(constat).data,
            status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['get'])
    def avancement(self, request, pk=None):
        """Synthèse d'avancement et de marge du contrat (FG146)."""
        contrat = self.get_object()
        return Response(selectors.avancement_contrat(
            request.user.company, contrat))


class TravauxEnCoursViewSet(_ComptaBaseViewSet):
    """Régularisations de cut-off : PCA / WIP (FG147).

    Crée/consulte les régularisations ; ``company`` / ``reference`` / ``statut``
    sont posés côté serveur. La création passe l'écriture OD de constat (sauf
    ``poster=false``) ; ``reprendre`` (POST) extourne à l'ouverture suivante.
    Filtres ``nature`` / ``statut`` ; société scopée ; Admin/Responsable.
    """
    queryset = TravauxEnCours.objects.select_related('created_by').all()
    serializer_class = TravauxEnCoursSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['reference', 'libelle', 'chantier_ref']
    ordering_fields = ['date_arrete', 'montant', 'id']

    def get_queryset(self):
        qs = super().get_queryset()
        params = self.request.query_params
        nature = params.get('nature')
        if nature:
            qs = qs.filter(nature=nature)
        statut = params.get('statut')
        if statut:
            qs = qs.filter(statut=statut)
        return qs

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        vd = serializer.validated_data
        poster = request.data.get('poster', True)
        if isinstance(poster, str):
            poster = poster.lower() not in ('false', '0', 'no', '')
        try:
            reg = services.constater_regularisation(
                request.user.company,
                nature=vd.get('nature'),
                montant=vd.get('montant') or 0,
                date_arrete=vd['date_arrete'],
                libelle=vd.get('libelle', '') or '',
                chantier_ref=vd.get('chantier_ref', '') or '',
                contrat_id=vd.get('contrat_id'),
                poster=bool(poster), user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(
            self.get_serializer(reg).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def reprendre(self, request, pk=None):
        """Reprend (extourne) la régularisation à l'ouverture suivante."""
        reg = self.get_object()
        date_reprise = request.data.get('date_reprise') or None
        services.reprendre_regularisation(
            reg, date_reprise=date_reprise, user=request.user)
        return Response(self.get_serializer(reg).data)


class CommissionPayoutRunViewSet(_ComptaBaseViewSet):
    """Campagnes de versement des commissions (FG148).

    ``company`` / ``reference`` / ``statut`` / ``total`` posés côté serveur. La
    création accepte un bloc ``lignes`` (commerciaux + montants). ``valider``
    (POST) gèle les montants ; ``poster`` (POST) passe l'écriture OD au grand
    livre. Filtre ``statut`` ; société scopée ; Admin/Responsable.
    """
    queryset = CommissionPayoutRun.objects.select_related(
        'created_by').prefetch_related('lignes').all()
    serializer_class = CommissionPayoutRunSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['reference', 'libelle', 'periode']
    ordering_fields = ['date_run', 'total', 'id']

    def get_queryset(self):
        qs = super().get_queryset()
        statut = self.request.query_params.get('statut')
        if statut:
            qs = qs.filter(statut=statut)
        return qs

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        vd = serializer.validated_data
        run = services.creer_commission_run(
            request.user.company,
            date_run=vd['date_run'],
            periode=vd.get('periode', '') or '',
            libelle=vd.get('libelle', '') or '',
            lignes=request.data.get('lignes') or [],
            user=request.user)
        return Response(
            self.get_serializer(run).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def valider(self, request, pk=None):
        """Valide le run (gèle les montants)."""
        run = self.get_object()
        try:
            services.valider_commission_run(run)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(run).data)

    @action(detail=True, methods=['post'])
    def poster(self, request, pk=None):
        """Poste le run au grand livre (écriture OD)."""
        run = self.get_object()
        try:
            services.poster_commission_run(run, user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(run).data)


class BudgetViewSet(_ComptaBaseViewSet):
    """Budgets annuels & suivi budget-vs-réalisé (FG149).

    ``company`` posée côté serveur. La création accepte un bloc ``lignes``
    (compte/centre de coût + 12 mois). ``vs_realise`` (GET) renvoie la variance
    budget-vs-réalisé lue du grand livre (``?export=csv`` pour le CSV). Société
    scopée ; Admin/Responsable.
    """
    queryset = Budget.objects.select_related(
        'created_by').prefetch_related('lignes__compte').all()
    serializer_class = BudgetSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['annee', 'id']

    def get_queryset(self):
        qs = super().get_queryset()
        annee = self.request.query_params.get('annee')
        if annee:
            qs = qs.filter(annee=annee)
        return qs

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        vd = serializer.validated_data
        lignes_in = request.data.get('lignes') or []
        company = request.user.company
        lignes = []
        for lig in lignes_in:
            compte = CompteComptable.objects.filter(
                company=company, id=lig.get('compte')).first()
            if compte is None:
                return Response(
                    {'detail': 'Compte inconnu.'},
                    status=status.HTTP_400_BAD_REQUEST)
            centre = None
            if lig.get('centre_cout'):
                centre = CentreCout.objects.filter(
                    company=company, id=lig.get('centre_cout')).first()
            entry = {'compte': compte, 'centre_cout': centre,
                     'libelle': lig.get('libelle', '') or ''}
            for m in BudgetLigne.MOIS:
                entry[m] = lig.get(m) or 0
            lignes.append(entry)
        budget = services.creer_budget(
            company, annee=vd['annee'], libelle=vd.get('libelle', '') or '',
            lignes=lignes, user=request.user)
        return Response(
            self.get_serializer(budget).data,
            status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['get'])
    def vs_realise(self, request, pk=None):
        """Variance budget-vs-réalisé lue du grand livre (FG149)."""
        budget = self.get_object()
        data = selectors.budget_vs_realise(request.user.company, budget)
        if request.query_params.get('export') == 'csv':
            buffer = io.StringIO()
            writer = csv.writer(buffer, delimiter=';')
            writer.writerow(
                ['Budget vs réalisé', data['annee'], data['libelle']])
            writer.writerow([])
            writer.writerow(
                ['Compte', 'Intitulé', 'Centre de coût', 'Budget', 'Réalisé',
                 'Variance', 'Consommation %'])
            for ligne in data['lignes']:
                writer.writerow([
                    ligne['compte_numero'], ligne['compte_intitule'],
                    ligne['centre_cout'], ligne['budget'], ligne['realise'],
                    ligne['variance'], ligne['taux_consommation']])
            writer.writerow([])
            writer.writerow(
                ['Total', '', '', data['total_budget'],
                 data['total_realise'], data['total_variance']])
            resp = HttpResponse(
                buffer.getvalue(), content_type='text/csv; charset=utf-8')
            resp['Content-Disposition'] = (
                'attachment; filename="budget_vs_realise.csv"')
            return resp
        return Response(data)


class CentreCoutViewSet(_ComptaBaseViewSet):
    """Référentiel des centres de coût / axes analytiques (FG150).

    ``company`` posée côté serveur. ``resultat`` (GET, detail=False) renvoie le
    résultat ventilé par centre de coût. Filtre ``axe`` / ``actif`` ; société
    scopée ; Admin/Responsable.
    """
    queryset = CentreCout.objects.all()
    serializer_class = CentreCoutSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['code', 'libelle']
    ordering_fields = ['code', 'id']

    def get_queryset(self):
        qs = super().get_queryset()
        params = self.request.query_params
        axe = params.get('axe')
        if axe:
            qs = qs.filter(axe=axe)
        actif = params.get('actif')
        if actif in ('0', '1'):
            qs = qs.filter(actif=(actif == '1'))
        return qs

    @action(detail=False, methods=['get'])
    def resultat(self, request):
        """Résultat (produits − charges) ventilé par centre de coût (FG150)."""
        params = request.query_params
        data = selectors.resultat_analytique(
            request.user.company,
            date_debut=params.get('date_debut') or None,
            date_fin=params.get('date_fin') or None,
            validees_seulement=params.get('validees') == '1')
        return Response(data)


class ProvisionCreanceViewSet(_ComptaBaseViewSet):
    """Provisions pour créances douteuses (FG152).

    La création calcule la ``dotation`` côté serveur (base × taux %) et passe
    l'écriture OD (sauf ``poster=false``). ``reprendre`` (POST) reprend la
    provision. ``company`` / ``reference`` / ``statut`` posés côté serveur ;
    filtre ``statut`` ; société scopée ; Admin/Responsable.
    """
    queryset = ProvisionCreance.objects.select_related('created_by').all()
    serializer_class = ProvisionCreanceSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['reference', 'tiers_nom']
    ordering_fields = ['date_dotation', 'dotation', 'base', 'id']

    def get_queryset(self):
        qs = super().get_queryset()
        statut = self.request.query_params.get('statut')
        if statut:
            qs = qs.filter(statut=statut)
        return qs

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        vd = serializer.validated_data
        poster = request.data.get('poster', True)
        if isinstance(poster, str):
            poster = poster.lower() not in ('false', '0', 'no', '')
        try:
            prov = services.enregistrer_provision_creance(
                request.user.company,
                date_dotation=vd['date_dotation'],
                base=vd.get('base') or 0,
                taux=vd.get('taux'),
                tiers_type=vd.get('tiers_type', '') or '',
                tiers_id=vd.get('tiers_id'),
                tiers_nom=vd.get('tiers_nom', '') or '',
                anciennete_jours=vd.get('anciennete_jours') or 0,
                libelle=vd.get('libelle', '') or '',
                poster=bool(poster), user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(
            self.get_serializer(prov).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def reprendre(self, request, pk=None):
        """Reprend la provision (créance recouvrée/soldée)."""
        prov = self.get_object()
        date_reprise = request.data.get('date_reprise') or None
        services.reprendre_provision_creance(
            prov, date_reprise=date_reprise, user=request.user)
        return Response(self.get_serializer(prov).data)


class ProvisionViewSet(_ComptaBaseViewSet):
    """Provisions risques & charges / dépréciation stock / immo (XACC26).

    La création poste l'écriture OD de dotation (sauf ``poster=false``).
    ``reprendre`` (POST, ``montant`` optionnel = solde) reprend tout ou partie
    de la provision. ``company`` / ``reference`` / ``montant_repris`` posés
    côté serveur ; filtre ``nature`` ; société scopée ; Admin/Responsable.
    """
    queryset = Provision.objects.select_related('created_by').all()
    serializer_class = ProvisionSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['reference', 'motif']
    ordering_fields = ['date_dotation', 'montant_dotation', 'id']

    def get_queryset(self):
        qs = super().get_queryset()
        nature = self.request.query_params.get('nature')
        if nature:
            qs = qs.filter(nature=nature)
        return qs

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        vd = serializer.validated_data
        poster = request.data.get('poster', True)
        if isinstance(poster, str):
            poster = poster.lower() not in ('false', '0', 'no', '')
        try:
            prov = services.enregistrer_provision(
                request.user.company,
                nature=vd['nature'],
                date_dotation=vd['date_dotation'],
                montant=vd.get('montant_dotation') or 0,
                motif=vd.get('motif', '') or '',
                date_echeance_revue=vd.get('date_echeance_revue'),
                poster=bool(poster), user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(
            self.get_serializer(prov).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def reprendre(self, request, pk=None):
        """Reprend tout ou partie de la provision (``montant`` optionnel)."""
        prov = self.get_object()
        montant = request.data.get('montant')
        date_reprise = request.data.get('date_reprise') or None
        try:
            services.reprendre_provision(
                prov, montant=montant, date_reprise=date_reprise,
                user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(prov).data)


class EntiteConsolidationViewSet(_ComptaBaseViewSet):
    """Périmètre de consolidation multi-entités (FG153).

    ``company`` (tête de groupe) posée côté serveur ; ``entite`` est une autre
    société du groupe. ``cpc_consolide`` (GET, detail=False) renvoie le CPC
    consolidé du périmètre. Société scopée ; Admin/Responsable.
    """
    queryset = EntiteConsolidation.objects.select_related('entite').all()
    serializer_class = EntiteConsolidationSerializer

    @action(detail=False, methods=['get'])
    def cpc_consolide(self, request):
        """CPC consolidé du périmètre de la société tête de groupe (FG153)."""
        params = request.query_params
        data = selectors.cpc_consolide(
            request.user.company,
            date_debut=params.get('date_debut') or None,
            date_fin=params.get('date_fin') or None)
        return Response(data)


class PilotageViewSet(viewsets.ViewSet):
    """Tableau de bord financier directeur (FG151) — LECTURE SEULE.

    Distinct de FG45 (quote-to-cash) : agrège résultat du mois, position de
    trésorerie, DSO/DPO, marge brute % et top encours depuis le grand livre.
    Admin/Responsable uniquement, scopé société côté selector.
    """
    permission_classes = [IsResponsableOrAdmin]

    @action(detail=False, methods=['get'])
    def cockpit(self, request):
        params = request.query_params
        data = selectors.pilotage_financier(
            request.user.company,
            date_debut=params.get('date_debut') or None,
            date_fin=params.get('date_fin') or None)
        return Response(data)


# ── FG201 — Campagnes email & SMS ──────────────────────────────────────────

class CampagneViewSet(_ComptaBaseViewSet):
    """Campagnes email/SMS (FG201). L'action ``envoyer`` déclenche l'envoi
    groupé — NO-OP gated (Brevo) tant que l'intégration est désactivée."""
    queryset = Campagne.objects.all()
    serializer_class = CampagneSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['nom', 'objet']
    # ZMKT2 — tri par les mesures agrégées (les taux dérivés se recalculent
    # depuis ces champs bruts — DRF ne peut trier que sur des champs réels).
    ordering_fields = [
        'date_creation', 'nom', 'envoyee_le', 'nb_envois', 'nb_ouvertures',
        'nb_clics', 'nb_destinataires',
    ]

    def get_queryset(self):
        qs = super().get_queryset()
        # ZMKT2 — Group By statut/canal/mois d'envoi (réutilisé par le
        # frontend pour le regroupement de liste).
        groupby = self.request.query_params.get('groupby')
        if groupby in ('statut', 'canal'):
            qs = qs.order_by(groupby, '-date_creation')
        return qs

    def get_permissions(self):
        # YRBAC13 — marketing-en-compta (ré-exportée par apps/marketing/
        # views.py, ODX10) : lecture (aperçus/KPI/reporting/kanban/modèles)
        # reste IsResponsableOrAdmin, inchangée. ``envoyer``/``envoyer-test``
        # déclenchent un envoi RÉEL (même sous garde d'approbation XMKT23) →
        # ``compta_valider``. Les actions de gestion courante du cycle de vie
        # (dupliquer/annuler/renvoyer les échecs/rattacher/cloner un modèle)
        # restent ``compta_saisir`` — les deux déjà octroyés au Responsable
        # par défaut (COMPTA40), donc aucune régression pour les rôles par
        # défaut.
        if self.action in (
                'apercu_fusion', 'precheck', 'cout_sms', 'clics_par_lien',
                'roi', 'roi_leads_sources', 'kpi_mere', 'kanban', 'reporting',
                'reporting_export', 'modeles', 'rendu_lead'):
            return [IsResponsableOrAdmin()]
        if self.action in ('envoyer', 'envoyer_test'):
            return [HasPermissionOrLegacy('compta_valider')()]
        if self.action in ('creer_depuis_modele', 'dupliquer', 'annuler',
                           'renvoyer_echecs', 'rattacher'):
            return [HasPermissionOrLegacy('compta_saisir')()]
        return super().get_permissions()

    @action(detail=True, methods=['post'])
    def envoyer(self, request, pk=None):
        campagne = self.get_object()
        destinataires = request.data.get('destinataires') or []
        # XMKT23 — au-delà du seuil société, l'envoi reste bloqué en attente
        # d'approbation (comportement inchangé sous le seuil).
        _campagne, approbation = services.demander_ou_envoyer_campagne(
            campagne, destinataires=destinataires, user=request.user)
        data = CampagneSerializer(campagne).data
        if approbation is not None:
            data['approbation_requise'] = True
            data['approbation_id'] = approbation.id
        return Response(data)

    @action(detail=True, methods=['get'])
    def apercu_fusion(self, request, pk=None):
        """XMKT8 — Aperçu du corps fusionné pour un lead d'exemple
        (``?lead_id=``), fallback appliqué par variable si le champ est vide."""
        campagne = self.get_object()
        lead_id = request.query_params.get('lead_id')
        if not lead_id:
            return Response({'detail': 'lead_id requis.'}, status=400)
        try:
            rendu = services.rendre_variables_fusion(
                campagne.corps, request.user.company, lead_id)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=400)
        return Response({'corps_fusionne': rendu})

    @action(detail=True, methods=['post'], url_path='envoyer-test')
    def envoyer_test(self, request, pk=None):
        """XMKT13 — Envoi de test (jamais vers de vrais destinataires)."""
        campagne = self.get_object()
        adresses_seed = request.data.get('adresses_seed') or []
        lead_id_exemple = request.data.get('lead_id_exemple')
        resultat = services.envoyer_test_campagne(
            campagne, adresses_seed=adresses_seed,
            lead_id_exemple=lead_id_exemple)
        return Response(resultat)

    @action(detail=True, methods=['get'], url_path='precheck')
    def precheck(self, request, pk=None):
        """XMKT13 — Pré-check bloquant/avertissant avant l'envoi de masse."""
        campagne = self.get_object()
        verifier_liens = request.query_params.get('verifier_liens') == '1'
        rapport = services.precheck_sante_campagne(
            campagne, verifier_liens=verifier_liens)
        return Response(rapport)

    @action(detail=True, methods=['get'], url_path='cout-sms')
    def cout_sms(self, request, pk=None):
        """XMKT15 — Segments GSM-7/UCS-2 + coût multi-part en direct."""
        campagne = self.get_object()
        prix_unitaire = request.query_params.get('prix_unitaire_mad')
        kwargs = {}
        if prix_unitaire:
            kwargs['prix_unitaire_mad'] = prix_unitaire
        nb_destinataires = int(
            request.query_params.get('nb_destinataires') or 1)
        estimation = services.estimer_cout_sms(
            campagne.corps, nb_destinataires=nb_destinataires, **kwargs)
        return Response(estimation)

    @action(detail=False, methods=['get'], url_path='generer-ia-disponible',
            permission_classes=[IsResponsableOrAdmin])
    def generer_ia_disponible(self, request):
        """XMKT34 — probe UI : la génération IA est-elle configurée ?

        Ne déclenche JAMAIS d'appel LLM (lit seulement le registre
        ``core.ai``) — le frontend s'en sert pour masquer complètement le
        bouton « Générer avec l'IA » sans clé (aucune trace UI)."""
        from core.ai.registry import is_capability_configured
        return Response({'configured': is_capability_configured('llm')})

    @action(detail=False, methods=['post'], url_path='generer-ia',
            permission_classes=[IsResponsableOrAdmin])
    def generer_ia(self, request):
        """XMKT34 — Génère (suggestion éditable) un objet + corps de campagne.

        NO-OP-safe : sans clé LLM configurée (Groq/Zhipu via
        ``settings.AI_PROVIDERS``), renvoie ``configured: false`` (le
        frontend masque le bouton). Le contenu généré n'est JAMAIS
        auto-appliqué à la campagne — l'utilisateur relit/édite avant de
        sauvegarder. Ne reçoit que du texte libre (segment/offre/consigne) :
        aucun champ interne (prix_achat/marge) n'est jamais envoyé."""
        from core.ai.services import draft_campaign_content
        data = request.data
        draft = draft_campaign_content(
            segment_label=str(data.get('segment_label') or '')[:500],
            offre=str(data.get('offre') or '')[:1000],
            instruction=str(data.get('instruction') or '')[:500],
            langue=str(data.get('langue') or 'fr')[:5],
            longueur=str(data.get('longueur') or '')[:100],
        )
        return Response({
            'configured': draft.configured,
            'ok': draft.ok,
            'objet': draft.objet,
            'corps': draft.corps,
            'langue': draft.langue,
            'source': draft.source,
        })

    @action(detail=True, methods=['get'], url_path='clics-par-lien')
    def clics_par_lien(self, request, pk=None):
        """XMKT9 — Page « clics par lien » du détail campagne."""
        campagne = self.get_object()
        return Response(services.clics_par_lien(campagne))

    @action(detail=True, methods=['get'], url_path='roi')
    def roi(self, request, pk=None):
        """XMKT17 — ROI MAD : dépensé vs revenu signé attribué + coût/lead."""
        campagne = self.get_object()
        return Response(services.roi_campagne(campagne))

    @action(detail=True, methods=['get'], url_path='roi/leads-sources')
    def roi_leads_sources(self, request, pk=None):
        """XMKT17 — Drill-down vers les leads sources du ROI."""
        campagne = self.get_object()
        return Response(services.leads_source_roi(campagne))

    @action(detail=True, methods=['get'], url_path='kpi-mere')
    def kpi_mere(self, request, pk=None):
        """XMKT31 — agrège KPI/coûts/ROI de tous les enfants d'une campagne
        mère (conteneur multi-canal)."""
        campagne = self.get_object()
        return Response(services.kpi_campagne_mere(campagne))

    @action(detail=False, methods=['get'])
    def kanban(self, request):
        """ZMKT1 — campagnes groupées par statut (pipeline Odoo-style)."""
        return Response(services.campagnes_par_statut(request.user.company))

    @action(detail=False, methods=['get'])
    def reporting(self, request):
        """ZMKT8 — reporting multi-vue (Graph/Pivot/Cohorte) : mesures
        délivrés/ouverts/cliqués/rebonds/désinscrits + CTR/CTOR/
        délivrabilité, groupable par ``?groupby=canal|mois|campagne``."""
        groupby = request.query_params.get('groupby', 'canal')
        return Response(
            services.reporting_campagnes(request.user.company, groupby=groupby))

    @action(detail=False, methods=['get'], url_path='reporting/export')
    def reporting_export(self, request):
        """ZMKT8 — export XLSX du reporting multi-vue."""
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from apps.records.xlsx import coerce_cell, XLSX_CONTENT_TYPE

        groupby = request.query_params.get('groupby', 'canal')
        lignes = services.reporting_campagnes(request.user.company, groupby=groupby)
        wb = Workbook()
        ws = wb.active
        ws.title = 'Reporting campagnes'
        headers = [
            'Groupe', 'Délivrés', 'Ouverts', 'Clics', 'Rebonds',
            'Désinscrits', 'CTR %', 'CTOR %', 'Délivrabilité %',
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for ligne in lignes:
            ws.append([coerce_cell(v) for v in [
                ligne['groupe'], ligne['delivres'], ligne['ouverts'],
                ligne['cliques'], ligne['rebonds'], ligne['desinscrits'],
                ligne['ctr_pct'], ligne['ctor_pct'], ligne['delivrabilite_pct'],
            ]])
        import io
        buf = io.BytesIO()
        wb.save(buf)
        resp = HttpResponse(buf.getvalue(), content_type=XLSX_CONTENT_TYPE)
        resp['Content-Disposition'] = 'attachment; filename="reporting_campagnes.xlsx"'
        return resp

    @action(detail=False, methods=['get'], url_path='modeles')
    def modeles(self, request):
        """ZMKT3 — liste des modèles company-scopés."""
        qs = self.get_queryset().filter(est_modele=True)
        return Response(CampagneSerializer(qs, many=True).data)

    @action(detail=True, methods=['post'], url_path='creer-depuis-modele')
    def creer_depuis_modele(self, request, pk=None):
        """ZMKT3 — clone un modèle en une nouvelle campagne brouillon."""
        modele = self.get_object()
        if not modele.est_modele:
            return Response(
                {'detail': "Cette campagne n'est pas un modèle."}, status=400)
        clone = services.creer_depuis_modele(modele)
        return Response(CampagneSerializer(clone).data, status=201)

    @action(detail=True, methods=['post'])
    def dupliquer(self, request, pk=None):
        """ZMKT4 — duplique une campagne en brouillon indépendant."""
        campagne = self.get_object()
        clone = services.dupliquer_campagne(campagne)
        return Response(CampagneSerializer(clone).data, status=201)

    @action(detail=True, methods=['post'])
    def annuler(self, request, pk=None):
        """ZMKT4 — annule une campagne en file/en cours d'envoi."""
        campagne = self.get_object()
        services.annuler_campagne(campagne)
        campagne.refresh_from_db()
        return Response(CampagneSerializer(campagne).data)

    @action(detail=True, methods=['post'], url_path='renvoyer-echecs')
    def renvoyer_echecs(self, request, pk=None):
        """ZMKT4 — recrée l'envoi vers les destinataires en échec
        récupérable uniquement."""
        campagne = self.get_object()
        nouvelles = services.renvoyer_echecs_campagne(campagne)
        return Response(
            {'campagnes_creees': [c.id for c in nouvelles]}, status=201)

    @action(detail=True, methods=['post'], url_path='rattacher')
    def rattacher(self, request, pk=None):
        """XMKT31 — rattache un objet (séquence/formulaire/code promo/
        événement) à cette campagne mère."""
        campagne = self.get_object()
        type_objet = request.data.get('type')
        objet_id = request.data.get('id')
        if not type_objet or not objet_id:
            return Response({'detail': 'type et id requis.'}, status=400)
        services.rattacher_a_campagne_mere(
            campagne, type_objet=type_objet, objet_id=objet_id)
        return Response(CampagneSerializer(campagne).data)

    @action(detail=True, methods=['get'], url_path='rendu-lead')
    def rendu_lead(self, request, pk=None):
        """XMKT11 — Rendu final (variante de langue + fusion) pour un lead
        donné (``?lead_id=``)."""
        campagne = self.get_object()
        lead_id = request.query_params.get('lead_id')
        if not lead_id:
            return Response({'detail': 'lead_id requis.'}, status=400)
        try:
            rendu = services.rendre_pour_lead(
                campagne, request.user.company, lead_id)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=400)
        return Response(rendu)


# ── XMKT35 — Posts réseaux sociaux (calendrier de contenu, gated) ───────────

class PostSocialViewSet(_ComptaBaseViewSet):
    """Posts réseaux sociaux planifiés (XMKT35). La publication réelle Meta
    Graph est gated (défaut OFF) : sans jeton, le post dû devient un rappel
    manuel notifié (texte prêt à coller) — voir services.traiter_posts_
    sociaux_dus. L'action ``planifier`` pose la date + le statut planifié."""
    queryset = PostSocial.objects.all()
    serializer_class = PostSocialSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['texte']
    ordering_fields = ['date_planifiee', 'date_creation', 'statut']

    def get_queryset(self):
        qs = super().get_queryset()
        statut = self.request.query_params.get('statut')
        if statut:
            qs = qs.filter(statut=statut)
        return qs

    def perform_create(self, serializer):
        serializer.save(company=self.request.user.company,
                        created_by=self.request.user)

    @action(detail=True, methods=['post'],
            permission_classes=[IsResponsableOrAdmin])
    def planifier(self, request, pk=None):
        """Planifie le post à ``date_planifiee`` (brouillon → planifié)."""
        post = self.get_object()
        date_planifiee = request.data.get('date_planifiee')
        if not date_planifiee:
            return Response({'detail': 'date_planifiee requise.'}, status=400)
        from django.utils.dateparse import parse_datetime
        quand = parse_datetime(str(date_planifiee))
        if quand is None:
            return Response(
                {'detail': 'date_planifiee invalide (ISO attendu).'},
                status=400)
        if timezone.is_naive(quand):
            quand = timezone.make_aware(quand)
        services.planifier_post_social(post, date_planifiee=quand)
        return Response(PostSocialSerializer(post).data)


# ── XMKT30 (partiel) — Calendrier marketing unifié ──────────────────────────
# Endpoint agrégé consommé par features/marketing/MarketingCalendarScreen.jsx.
# Sources servies AUJOURD'HUI : campagnes (planifiee_le, XMKT7) et posts
# sociaux (date_planifiee, XMKT35). Les sources étapes-de-séquence/relances/
# événements restent à brancher par la tâche XMKT30 backend (même contrat de
# réponse {events: [...]}) — additif, aucun contrat cassé.

class CalendrierMarketingView(APIView):
    permission_classes = [IsResponsableOrAdmin]

    def get(self, request):
        company = request.user.company
        date_from = request.query_params.get('from') or ''
        date_to = request.query_params.get('to') or ''
        events = []
        campagnes = Campagne.objects.filter(
            company=company, planifiee_le__isnull=False)
        if date_from:
            campagnes = campagnes.filter(planifiee_le__date__gte=date_from)
        if date_to:
            campagnes = campagnes.filter(planifiee_le__date__lte=date_to)
        for c in campagnes:
            events.append({
                'id': f'campagne-{c.id}', 'obj_id': c.id,
                'source': 'campagne', 'link_type': 'campagne',
                'date': c.planifiee_le.date().isoformat(),
                'title': c.nom, 'channel': c.canal,
                'editable': c.statut in (
                    Campagne.Statut.BROUILLON, Campagne.Statut.EN_FILE),
            })
        posts = PostSocial.objects.filter(
            company=company, date_planifiee__isnull=False)
        if date_from:
            posts = posts.filter(date_planifiee__date__gte=date_from)
        if date_to:
            posts = posts.filter(date_planifiee__date__lte=date_to)
        for p in posts:
            events.append({
                'id': f'post_social-{p.id}', 'obj_id': p.id,
                'source': 'post_social', 'link_type': 'post_social',
                'date': p.date_planifiee.date().isoformat(),
                'title': f'{p.get_reseau_display()} — {(p.texte or "")[:60]}',
                'channel': '', 'editable': False,
            })
        return Response({'events': events})


class CalendrierMarketingRescheduleView(APIView):
    permission_classes = [IsResponsableOrAdmin]

    def post(self, request):
        """Replanifie une CAMPAGNE non partie par glisser-déposer (XMKT30).
        Seule source déplaçable aujourd'hui (contrat frontend isDraggable)."""
        company = request.user.company
        source = request.data.get('source')
        obj_id = request.data.get('id')
        date_str = str(request.data.get('date') or '')
        if source != 'campagne':
            return Response({'detail': 'Source non replanifiable.'}, status=400)
        from django.utils.dateparse import parse_date
        cible = parse_date(date_str)
        if cible is None:
            return Response({'detail': 'date invalide.'}, status=400)
        campagne = Campagne.objects.filter(
            company=company, id=obj_id).first()
        if campagne is None:
            return Response({'detail': 'Campagne inconnue.'}, status=404)
        if campagne.statut not in (
                Campagne.Statut.BROUILLON, Campagne.Statut.EN_FILE):
            return Response(
                {'detail': 'Campagne déjà partie — non replanifiable.'},
                status=400)
        ancienne = campagne.planifiee_le or timezone.now()
        campagne.planifiee_le = ancienne.replace(
            year=cible.year, month=cible.month, day=cible.day)
        campagne.save(update_fields=['planifiee_le'])
        return Response({'ok': True})


# ── XMKT2 — Journal d'envoi par destinataire (drill-down) ───────────────────

class EnvoiCampagneViewSet(_ComptaBaseViewSet):
    """Trace d'envoi par destinataire (XMKT2), lecture seule pour le
    drill-down depuis un KPI de campagne. Filtrable par ``campagne`` et
    ``statut``."""
    http_method_names = ['get', 'head', 'options']
    queryset = EnvoiCampagne.objects.select_related('campagne').all()
    serializer_class = EnvoiCampagneSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_creation']

    def get_queryset(self):
        qs = super().get_queryset()
        campagne_id = self.request.query_params.get('campagne')
        if campagne_id:
            qs = qs.filter(campagne_id=campagne_id)
        statut = self.request.query_params.get('statut')
        if statut:
            qs = qs.filter(statut=statut)
        return qs


# ── XMKT23 — Approbation avant envoi de masse + journal d'audit ────────────

class ApprobationEnvoiCampagneViewSet(_ComptaBaseViewSet):
    """Demandes d'approbation d'envoi de masse (XMKT23)."""
    http_method_names = ['get', 'post', 'head', 'options']
    queryset = ApprobationEnvoiCampagne.objects.select_related('campagne').all()
    serializer_class = ApprobationEnvoiCampagneSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_creation']

    @action(detail=True, methods=['post'])
    def approuver(self, request, pk=None):
        approbation = self.get_object()
        services.approuver_envoi_campagne(approbation, user=request.user)
        approbation.refresh_from_db()
        return Response(ApprobationEnvoiCampagneSerializer(approbation).data)

    @action(detail=True, methods=['post'])
    def rejeter(self, request, pk=None):
        approbation = self.get_object()
        motif = request.data.get('motif', '')
        services.rejeter_envoi_campagne(approbation, motif=motif, user=request.user)
        approbation.refresh_from_db()
        return Response(ApprobationEnvoiCampagneSerializer(approbation).data)

    @action(detail=False, methods=['get'], url_path='journal-audit')
    def journal_audit(self, request):
        return Response(
            services.journal_audit_envois(request.user.company))


# ── XMKT27 — Constructeur d'enquêtes avec logique conditionnelle ───────────

class EnqueteViewSet(_ComptaBaseViewSet):
    """CRUD des enquêtes (XMKT27)."""
    queryset = Enquete.objects.all()
    serializer_class = EnqueteSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['titre']
    ordering_fields = ['date_creation', 'titre']

    def perform_create(self, serializer):
        import uuid
        serializer.save(
            company=self.request.user.company, token=uuid.uuid4().hex)

    def get_permissions(self):
        # YRBAC13 — marketing-en-compta (ré-exportée par apps/marketing/
        # views.py). Lecture (résultats/aperçu/QR/participations/export)
        # reste IsResponsableOrAdmin, inchangée. ``emettre_jeton_invite``/
        # ``inviter`` déclenchent un envoi RÉEL vers des destinataires
        # (segment/liste) → ``compta_valider`` (déjà octroyé au Responsable
        # par défaut — comportement inchangé pour les rôles par défaut).
        if self.action in ('resultats', 'tester', 'qr', 'participations',
                           'resultats_export'):
            return [IsResponsableOrAdmin()]
        if self.action in ('emettre_jeton_invite', 'inviter'):
            return [HasPermissionOrLegacy('compta_valider')()]
        return super().get_permissions()

    @action(detail=True, methods=['get'])
    def resultats(self, request, pk=None):
        """XMKT27 — analytics agrégées par question + taux de complétion."""
        enquete = self.get_object()
        return Response(services.analytics_enquete(enquete))

    @action(detail=True, methods=['get'])
    def tester(self, request, pk=None):
        """ZMKT11 — aperçu SANS enregistrer de réponse."""
        enquete = self.get_object()
        return Response(services.tester_enquete(enquete))

    @action(detail=True, methods=['post'], url_path='emettre-jeton-invite')
    def emettre_jeton_invite(self, request, pk=None):
        """ZMKT11 — émet un jeton d'invitation (mode invités-seulement)."""
        enquete = self.get_object()
        jeton = services.emettre_jeton_invite(enquete)
        return Response({'jeton': jeton})

    @action(detail=True, methods=['get'])
    def qr(self, request, pk=None):
        """ZMKT12 — QR SVG téléchargeable du lien public de l'enquête."""
        enquete = self.get_object()
        svg = services.qr_svg_enquete(enquete)
        return HttpResponse(svg, content_type='image/svg+xml')

    @action(detail=True, methods=['get'])
    def participations(self, request, pk=None):
        """ZMKT13 — liste des soumissions individuelles, filtrable
        réussi/échoué (``?reussi=true|false``)."""
        enquete = self.get_object()
        reussi_param = request.query_params.get('reussi')
        reussi = None
        if reussi_param is not None:
            reussi = reussi_param.lower() == 'true'
        return Response(services.participations_enquete(enquete, reussi=reussi))

    @action(detail=True, methods=['get'], url_path='resultats/export')
    def resultats_export(self, request, pk=None):
        """ZMKT13 — export XLSX des participations."""
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from apps.records.xlsx import coerce_cell, XLSX_CONTENT_TYPE
        import io

        enquete = self.get_object()
        participations = services.participations_enquete(enquete)
        wb = Workbook()
        ws = wb.active
        ws.title = 'Participations'
        ws.append(['Contact', 'Score %', 'Réussi', 'Date'])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for p in participations:
            ws.append([coerce_cell(v) for v in [
                p['contact'], p['score_pct'], p['reussi'], p['date_creation']]])
        buf = io.BytesIO()
        wb.save(buf)
        resp = HttpResponse(buf.getvalue(), content_type=XLSX_CONTENT_TYPE)
        resp['Content-Disposition'] = 'attachment; filename="participations.xlsx"'
        return resp

    @action(detail=True, methods=['post'])
    def inviter(self, request, pk=None):
        """ZMKT12 — invitation email vers un segment (XMKT6) ou une liste
        (XMKT5), consentement + suppression respectés."""
        enquete = self.get_object()
        segment_id = request.data.get('segment_id')
        liste_id = request.data.get('liste_id')
        segment = (SegmentMarketing.objects.filter(
            id=segment_id, company=request.user.company).first()
            if segment_id else None)
        liste = (ListeDiffusion.objects.filter(
            id=liste_id, company=request.user.company).first()
            if liste_id else None)
        resultat = services.inviter_enquete(enquete, segment=segment, liste=liste)
        return Response(resultat)


# ── XMKT28 — Événements marketing légers ────────────────────────────────────

class EvenementMarketingViewSet(_ComptaBaseViewSet):
    """CRUD des événements marketing (XMKT28)."""
    queryset = EvenementMarketing.objects.all()
    serializer_class = EvenementMarketingSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['nom']
    ordering_fields = ['date_debut', 'date_creation']

    def get_permissions(self):
        # YRBAC13 — marketing-en-compta (ré-exportée par apps/marketing/
        # views.py). Lecture (borne/badges/reporting/kanban) reste
        # IsResponsableOrAdmin, inchangée. ``cloturer_presences`` fige les
        # présences de l'événement (action de clôture) → ``compta_valider`` ;
        # ``avancer_etape`` est une saisie courante du cycle de vie →
        # ``compta_saisir`` — les deux déjà octroyés au Responsable par
        # défaut (COMPTA40), aucune régression pour les rôles par défaut.
        if self.action in ('borne', 'badges', 'reporting', 'reporting_export',
                           'kanban'):
            return [IsResponsableOrAdmin()]
        if self.action == 'cloturer_presences':
            return [HasPermissionOrLegacy('compta_valider')()]
        if self.action == 'avancer_etape':
            return [HasPermissionOrLegacy('compta_saisir')()]
        return super().get_permissions()

    @action(detail=True, methods=['post'], url_path='cloturer-presences')
    def cloturer_presences(self, request, pk=None):
        evenement = self.get_object()
        nb = services.cloturer_presences_evenement(evenement)
        return Response({'absents_marques': nb})

    @action(detail=True, methods=['get'])
    def borne(self, request, pk=None):
        """ZMKT18 — recherche par nom/email parmi les inscrits."""
        evenement = self.get_object()
        terme = request.query_params.get('q', '')
        return Response(services.rechercher_inscrits_borne(evenement, terme))

    @action(detail=True, methods=['get'])
    def badges(self, request, pk=None):
        """ZMKT19 — impression en lot des badges (PDF multi-pages)."""
        evenement = self.get_object()
        pdf_bytes = services.generer_badges_pdf_lot(evenement)
        resp = HttpResponse(pdf_bytes, content_type='application/pdf')
        resp['Content-Disposition'] = 'inline; filename="badges.pdf"'
        return resp

    @action(detail=False, methods=['get'])
    def reporting(self, request):
        """ZMKT20 — reporting événement (participants & billetterie),
        groupable par ``?groupby=type|mois``."""
        groupby = request.query_params.get('groupby')
        return Response(
            services.reporting_evenements(request.user.company, groupby=groupby))

    @action(detail=False, methods=['get'], url_path='reporting/export')
    def reporting_export(self, request):
        """ZMKT20 — export XLSX du reporting événement."""
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from apps.records.xlsx import coerce_cell, XLSX_CONTENT_TYPE
        import io

        lignes = services.reporting_evenements(request.user.company)
        wb = Workbook()
        ws = wb.active
        ws.title = 'Reporting événements'
        headers = [
            'Événement', 'Type', 'Inscrits', 'Confirmés', 'Présents',
            'Absents', 'Taux présence %', 'Recette théorique MAD', 'Leads',
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for ligne in lignes:
            ws.append([coerce_cell(v) for v in [
                ligne['nom'], ligne['type_evenement'], ligne['nb_inscrits'],
                ligne['nb_confirmes'], ligne['nb_presents'],
                ligne['nb_absents'], ligne['taux_presence_pct'],
                ligne['recette_theorique_mad'], ligne['nb_leads'],
            ]])
        buf = io.BytesIO()
        wb.save(buf)
        resp = HttpResponse(buf.getvalue(), content_type=XLSX_CONTENT_TYPE)
        resp['Content-Disposition'] = 'attachment; filename="reporting_evenements.xlsx"'
        return resp

    @action(detail=False, methods=['get'])
    def kanban(self, request):
        """ZMKT14 — Kanban par étape configurable."""
        return Response(services.evenements_par_etape(request.user.company))

    @action(detail=True, methods=['post'], url_path='avancer-etape')
    def avancer_etape(self, request, pk=None):
        evenement = self.get_object()
        nouvelle_etape = request.data.get('etape')
        services.avancer_etape_evenement(evenement, nouvelle_etape)
        evenement.refresh_from_db()
        return Response(EvenementMarketingSerializer(evenement).data)


class TypeEvenementViewSet(_ComptaBaseViewSet):
    """Modèles réutilisables d'événement (ZMKT14)."""
    queryset = TypeEvenement.objects.all()
    serializer_class = TypeEvenementSerializer

    @action(detail=True, methods=['post'], url_path='creer-evenement')
    def creer_evenement(self, request, pk=None):
        type_evenement = self.get_object()
        nom = request.data.get('nom')
        date_debut = request.data.get('date_debut')
        if not nom or not date_debut:
            return Response({'detail': 'nom et date_debut requis.'}, status=400)
        evenement = services.creer_evenement_depuis_type(
            type_evenement, nom=nom, date_debut=date_debut)
        return Response(EvenementMarketingSerializer(evenement).data, status=201)


class BilletEvenementViewSet(_ComptaBaseViewSet):
    """Billets d'événement (ZMKT15)."""
    queryset = BilletEvenement.objects.select_related('evenement').all()
    serializer_class = BilletEvenementSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        evenement_id = self.request.query_params.get('evenement')
        if evenement_id:
            qs = qs.filter(evenement_id=evenement_id)
        return qs


class QuestionEvenementViewSet(_ComptaBaseViewSet):
    """Questions d'inscription par événement (ZMKT16)."""
    queryset = QuestionEvenement.objects.select_related('evenement').all()
    serializer_class = QuestionEvenementSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        evenement_id = self.request.query_params.get('evenement')
        if evenement_id:
            qs = qs.filter(evenement_id=evenement_id)
        return qs


class CommunicationEvenementViewSet(_ComptaBaseViewSet):
    """Communications programmées d'événement (ZMKT17)."""
    queryset = CommunicationEvenement.objects.select_related('evenement').all()
    serializer_class = CommunicationEvenementSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        evenement_id = self.request.query_params.get('evenement')
        if evenement_id:
            qs = qs.filter(evenement_id=evenement_id)
        return qs


class InscriptionEvenementViewSet(_ComptaBaseViewSet):
    """Inscriptions à un événement (XMKT28)."""
    queryset = InscriptionEvenement.objects.select_related('evenement').all()
    serializer_class = InscriptionEvenementSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_creation']

    def get_queryset(self):
        qs = super().get_queryset()
        evenement_id = self.request.query_params.get('evenement')
        if evenement_id:
            qs = qs.filter(evenement_id=evenement_id)
        return qs

    @action(detail=True, methods=['post'])
    def pointer(self, request, pk=None):
        inscription = self.get_object()
        services.pointer_presence(inscription)
        inscription.refresh_from_db()
        return Response(InscriptionEvenementSerializer(inscription).data)

    @action(detail=True, methods=['get'])
    def badge(self, request, pk=None):
        """ZMKT19 — badge PDF imprimable d'un inscrit."""
        inscription = self.get_object()
        pdf_bytes = services.generer_badge_pdf(inscription)
        resp = HttpResponse(pdf_bytes, content_type='application/pdf')
        resp['Content-Disposition'] = 'inline; filename="badge.pdf"'
        return resp


class _MarketingPublicThrottle(SimpleRateThrottle):
    """Débit des endpoints marketing PUBLICS (inscriptions événement,
    désinscription/double opt-in/enquête tokenisés, webhooks entrants) par IP
    — anti-abus/brute-force du jeton (YRBAC9). Même patron que
    ``_PortailComptaThrottle``."""
    scope = 'compta_marketing_public'
    rate = '30/minute'

    def get_rate(self):
        return self.rate

    def get_cache_key(self, request, view):
        ident = self.get_ident(request)
        return self.cache_format % {'scope': self.scope, 'ident': ident}


@api_view(['POST'])
@permission_classes([AllowAny])
@throttle_classes([_MarketingPublicThrottle])
def evenement_inscription_publique(request, evenement_id):
    """XMKT28 — Inscription publique à un événement (aucune auth)."""
    evenement = EvenementMarketing.objects.filter(id=evenement_id).first()
    if not evenement:
        return Response({'detail': 'Événement introuvable.'}, status=404)
    nom = (request.data.get('nom') or '').strip()
    if not nom:
        return Response({'detail': 'nom requis.'}, status=400)
    billet = None
    billet_id = request.data.get('billet_id')
    if billet_id:
        billet = BilletEvenement.objects.filter(
            id=billet_id, evenement=evenement).first()
        if not billet:
            return Response({'detail': 'Billet introuvable.'}, status=404)
    try:
        inscription = services.inscrire_evenement(
            evenement, nom=nom,
            email=request.data.get('email', ''),
            telephone=request.data.get('telephone', ''), billet=billet,
            reponses_questions=request.data.get('reponses_questions'))
    except ValueError as exc:
        return Response({'detail': str(exc)}, status=400)
    return Response(
        {'id': inscription.id, 'qr_token': inscription.qr_token}, status=201)


# ── XMKT29 — Ponts QR pour supports offline ─────────────────────────────────

class SupportOfflineViewSet(_ComptaBaseViewSet):
    """CRUD des supports offline avec QR téléchargeable (XMKT29)."""
    queryset = SupportOffline.objects.select_related('lien_tracke').all()
    serializer_class = SupportOfflineSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['nom']
    ordering_fields = ['date_creation']

    def perform_create(self, serializer):
        nom = serializer.validated_data.get('nom')
        url_cible = serializer.validated_data.get('url_cible')
        support = services.creer_support_offline(
            self.request.user.company, nom=nom, url_cible=url_cible)
        serializer.instance = support

    @action(detail=True, methods=['get'])
    def qr(self, request, pk=None):
        support = self.get_object()
        svg = services.qr_svg_support_offline(support)
        if svg is None:
            return Response({'detail': 'QR indisponible.'}, status=404)
        return HttpResponse(svg, content_type='image/svg+xml')

    @action(detail=False, methods=['get'], url_path='scans-par-support')
    def scans_par_support(self, request):
        return Response(
            services.tableau_scans_par_support(request.user.company))


# ── XMKT33 — Assistant d'authentification du domaine d'envoi ──────────────

class DomaineEnvoiViewSet(_ComptaBaseViewSet):
    """Page Paramètres « Domaine d'envoi » (XMKT33)."""
    queryset = DomaineEnvoi.objects.all()
    serializer_class = DomaineEnvoiSerializer

    @action(detail=True, methods=['get'], url_path='enregistrements-attendus')
    def enregistrements_attendus(self, request, pk=None):
        domaine_envoi = self.get_object()
        return Response(
            services.enregistrements_dns_attendus(domaine_envoi.domaine))

    @action(detail=True, methods=['post'], url_path='verifier')
    def verifier(self, request, pk=None):
        domaine_envoi = self.get_object()
        services.verifier_domaine_envoi(domaine_envoi)
        domaine_envoi.refresh_from_db()
        return Response(DomaineEnvoiSerializer(domaine_envoi).data)


# ── XMKT5 — Listes de diffusion nommées + abonnements ───────────────────────

class ListeDiffusionViewSet(_ComptaBaseViewSet):
    """CRUD des listes de diffusion nommées (XMKT5)."""
    queryset = ListeDiffusion.objects.all()
    serializer_class = ListeDiffusionSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['nom']
    ordering_fields = ['nom', 'date_creation']

    @action(detail=True, methods=['post'])
    def importer(self, request, pk=None):
        """Import CSV/XLSX déjà mappé côté client : ``lignes`` = liste de
        ``{'destinataire': ..., 'contact_ref': ...}``."""
        liste = self.get_object()
        lignes = request.data.get('lignes') or []
        rapport = services.importer_abonnements_liste(liste, lignes)
        return Response(rapport)

    @action(detail=True, methods=['get'])
    def abonnes(self, request, pk=None):
        liste = self.get_object()
        qs = liste.abonnements.all()
        statut = request.query_params.get('statut')
        if statut:
            qs = qs.filter(statut=statut)
        return Response(AbonnementListeSerializer(qs, many=True).data)


class AbonnementListeViewSet(_ComptaBaseViewSet):
    """Abonnements individuels à une liste (XMKT5)."""
    queryset = AbonnementListe.objects.select_related('liste').all()
    serializer_class = AbonnementListeSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        liste_id = self.request.query_params.get('liste')
        if liste_id:
            qs = qs.filter(liste_id=liste_id)
        return qs


# ── XMKT6 — Segments dynamiques enregistrés et réutilisables ────────────────

class SegmentMarketingViewSet(_ComptaBaseViewSet):
    """Segments nommés réutilisables (XMKT6), auto-actualisés à chaque
    usage — ``previsualiser`` renvoie le compte exact + un échantillon."""
    queryset = SegmentMarketing.objects.all()
    serializer_class = SegmentMarketingSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['nom']
    ordering_fields = ['nom', 'date_creation']

    @action(detail=True, methods=['get'])
    def previsualiser(self, request, pk=None):
        segment = self.get_object()
        try:
            data = services.previsualiser_segment(segment)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=400)
        return Response(data)

    @action(detail=True, methods=['post'], url_path='exporter-audience-meta',
            permission_classes=[IsResponsableOrAdmin])
    def exporter_audience_meta(self, request, pk=None):
        """XMKT36 — [DECISION] Synchronise le segment comme audience Meta.

        Identifiants hashés SHA-256 côté serveur, consentement XMKT4 exigé,
        clients signés en liste d'exclusion. GATED : sans jeton
        (``META_ADS_ENABLED``/``META_ADS_TOKEN``/``META_AD_ACCOUNT_ID``),
        aucun appel réseau — le résumé (compteurs, configured=false) est
        renvoyé pour l'UI. AUCUNE campagne publicitaire créée (règle n°3)."""
        segment = self.get_object()
        try:
            resume = services.exporter_segment_audience_meta(segment)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=400)
        return Response(resume)


# ── FG202 — Séquences de relance automatisées ──────────────────────────────

class SequenceRelanceViewSet(_ComptaBaseViewSet):
    """Séquences de relance drip/nurture (FG202). ``planifier`` calcule le
    calendrier des étapes sans rien envoyer (envoi gated)."""
    queryset = SequenceRelance.objects.prefetch_related('etapes').all()
    serializer_class = SequenceRelanceSerializer

    @action(detail=True, methods=['get'])
    def planifier(self, request, pk=None):
        sequence = self.get_object()
        plan = services.planifier_etapes_sequence(sequence)
        return Response({'etapes': plan})

    @action(detail=True, methods=['get'])
    def traces(self, request, pk=None):
        """ZMKT5 — traces filtrables par étape et statut."""
        sequence = self.get_object()
        etape_id = request.query_params.get('etape')
        statut_trace = request.query_params.get('statut')
        return Response(services.traces_sequence(
            sequence, etape_id=etape_id, statut_trace=statut_trace))

    @action(detail=True, methods=['get'], url_path='compteurs-par-etape')
    def compteurs_par_etape(self, request, pk=None):
        """ZMKT5 — compteurs Succès/Rejeté/Envoyé par étape."""
        sequence = self.get_object()
        return Response(services.compteurs_par_etape(sequence))

    @action(detail=True, methods=['get'])
    def participants(self, request, pk=None):
        """ZMKT6 — liste des participants (nœud courant + prochaine
        échéance), filtrable par statut, + compteur actifs."""
        sequence = self.get_object()
        statut = request.query_params.get('statut')
        return Response({
            'participants': services.participants_sequence(sequence, statut=statut),
            'nb_actifs': services.nb_participants_actifs(sequence),
        })


class EtapeSequenceViewSet(_ComptaBaseViewSet):
    """Étapes d'une séquence de relance (FG202)."""
    queryset = EtapeSequence.objects.select_related('sequence').all()
    serializer_class = EtapeSequenceSerializer


# ── XMKT1 — Inscriptions & exécution réelle des séquences ──────────────────

class InscriptionSequenceViewSet(_ComptaBaseViewSet):
    """Inscriptions de leads dans une séquence de relance (XMKT1) : trace par
    participant (quel nœud, quand, quoi exécuté) via ``executions``."""
    queryset = InscriptionSequence.objects.select_related(
        'sequence', 'etape_courante').prefetch_related('executions').all()
    serializer_class = InscriptionSequenceSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['declenchee_le']

    def get_queryset(self):
        qs = super().get_queryset()
        sequence_id = self.request.query_params.get('sequence')
        if sequence_id:
            qs = qs.filter(sequence_id=sequence_id)
        lead_id = self.request.query_params.get('lead_id')
        if lead_id:
            qs = qs.filter(lead_id=lead_id)
        return qs

    @action(detail=False, methods=['post'])
    def inscrire(self, request):
        sequence_id = request.data.get('sequence')
        lead_id = request.data.get('lead_id')
        if not sequence_id or not lead_id:
            return Response(
                {'detail': 'sequence et lead_id requis.'}, status=400)
        sequence = SequenceRelance.objects.filter(
            id=sequence_id, company=request.user.company).first()
        if not sequence:
            return Response({'detail': 'Séquence introuvable.'}, status=404)
        inscription = services.inscrire_lead_sequence(
            request.user.company, sequence, lead_id=lead_id,
            lead_reference=request.data.get('lead_reference', ''))
        return Response(
            InscriptionSequenceSerializer(inscription).data, status=201)

    @action(detail=True, methods=['post'])
    def sortir(self, request, pk=None):
        inscription = self.get_object()
        services.sortir_inscription(
            inscription, motif=request.data.get('motif', 'manuel'))
        inscription.refresh_from_db()
        return Response(InscriptionSequenceSerializer(inscription).data)


# ── XMKT2 — Webhook Brevo (gated, public, aucune auth) ──────────────────────

@api_view(['POST'])
@permission_classes([AllowAny])
@throttle_classes([_MarketingPublicThrottle])
def webhook_brevo_campagne(request):
    """Réception d'un événement webhook Brevo (XMKT2/XMKT12) : delivered/
    opened/click/bounce/unsubscribed/complaint. Résout la société depuis la
    ``Campagne`` référencée (aucune session utilisateur côté webhook
    externe). Payload minimal attendu : ``campagne_id``, ``destinataire``,
    ``event``, et optionnellement ``reason`` (raison SMTP) + ``bounce_type``
    (``hard``/``soft``, XMKT12).
    """
    data = request.data or {}
    campagne_id = data.get('campagne_id')
    destinataire = (data.get('destinataire') or '').strip()
    evenement = data.get('event') or ''
    if not campagne_id or not destinataire or not evenement:
        return Response({'detail': 'payload incomplet'}, status=400)
    campagne = Campagne.objects.filter(id=campagne_id).first()
    if not campagne:
        return Response({'detail': 'campagne introuvable'}, status=404)
    envoi = services.webhook_brevo_evenement(
        campagne.company, campagne_id=campagne.id,
        destinataire=destinataire, evenement=evenement,
        raison_smtp=data.get('reason', ''),
        bounce_type=data.get('bounce_type', ''))
    if not envoi:
        return Response({'detail': 'destinataire introuvable'}, status=404)
    return Response({'statut': envoi.statut})


# ── XMKT15 — Webhook agrégateur SMS : mot-clé STOP entrant (gated, public) ──

@api_view(['POST'])
@permission_classes([AllowAny])
@throttle_classes([_MarketingPublicThrottle])
def webhook_sms_stop(request):
    """Réception d'un SMS entrant STOP (XMKT15, gated/no-op sans intégration
    d'agrégateur active). Payload minimal attendu : ``company_id``,
    ``numero``. Désinscrit immédiatement le numéro (XMKT3).
    """
    from authentication.models import Company

    data = request.data or {}
    company_id = data.get('company_id')
    numero = (data.get('numero') or '').strip()
    if not company_id or not numero:
        return Response({'detail': 'payload incomplet'}, status=400)
    company = Company.objects.filter(id=company_id).first()
    if not company:
        return Response({'detail': 'société introuvable'}, status=404)
    supprime = services.traiter_stop_entrant(company, numero)
    if not supprime:
        return Response({'detail': 'numéro invalide'}, status=400)
    return Response({'desinscrit': True, 'destinataire': supprime.destinataire})


# ── XMKT3 — Désinscription un clic (public, tokenisé, aucune auth) ─────────

@api_view(['GET', 'POST'])
@permission_classes([AllowAny])
@throttle_classes([_MarketingPublicThrottle])
def desinscription_publique(request, token):
    """Page/endpoint public de désinscription un clic (XMKT3, RFC 8058).

    Un jeton signé (``services.generer_token_desinscription``) porte la
    société + le destinataire — aucune authentification requise. GET et POST
    font la même chose (RFC 8058 recommande un POST simple sans confirmation
    pour les clients mail qui suivent ``List-Unsubscribe``).
    """
    ok, resultat = services.desinscrire_via_token(token)
    if not ok:
        return Response({'detail': resultat}, status=400)
    return Response({'desinscrit': True, 'destinataire': resultat})


# ── XMKT4 — Confirmation double opt-in (public, tokenisé, aucune auth) ─────

@api_view(['GET', 'POST'])
@permission_classes([AllowAny])
@throttle_classes([_MarketingPublicThrottle])
def double_optin_confirmer(request, token):
    """Clic de confirmation du double opt-in (XMKT4, loi 09-08).

    Pose un ``core.ConsentRecord`` accordé pour la finalité marketing, preuve
    IP + horodatage. Jeton invalide → 400 sans effet.
    """
    ok, resultat = services.confirmer_double_optin_via_token(token)
    if not ok:
        return Response({'detail': resultat}, status=400)
    return Response({'confirme': True, 'destinataire': resultat})


# ── XMKT9 — Redirection tokenisée (public, tracking de clics) ──────────────

@api_view(['GET'])
@permission_classes([AllowAny])
@throttle_classes([_MarketingPublicThrottle])
def redirection_lien_tracke(request, token):
    """Redirige vers l'URL cible d'un ``LienTrackee`` (XMKT9), en comptant le
    clic (par lien + par destinataire via ``?d=`` si fourni par l'appelant
    email/SMS). Jeton invalide → 404 (aucune fuite d'existence)."""
    from django.http import HttpResponseRedirect

    destinataire = (request.GET.get('d') or '').strip()
    ok, resultat = services.traiter_clic_lien(token, destinataire=destinataire)
    if not ok:
        return Response({'detail': resultat}, status=404)
    return HttpResponseRedirect(resultat)


# ── XMKT27 — Enquêtes (public, tokenisé, aucune auth) ───────────────────────

@api_view(['GET'])
@permission_classes([AllowAny])
@throttle_classes([_MarketingPublicThrottle])
def enquete_publique(request, token):
    """Récupère les questions VISIBLES d'une enquête via son lien public
    (XMKT27). Jeton invalide/enquête inactive → 404 (aucune fuite
    d'existence). Les réponses déjà données (``?reponses=`` JSON encodé) sont
    prises en compte pour la logique conditionnelle."""
    import json as _json

    enquete = Enquete.objects.filter(token=token, actif=True).first()
    if not enquete:
        return Response({'detail': 'Enquête introuvable.'}, status=404)
    jeton_invite = request.GET.get('invite')
    if not services.acces_enquete_autorise(enquete, jeton_invite=jeton_invite):
        return Response({'detail': 'Enquête introuvable.'}, status=404)
    reponses_partielles = {}
    brut = request.GET.get('reponses')
    if brut:
        try:
            reponses_partielles = _json.loads(brut)
        except (ValueError, TypeError):
            reponses_partielles = {}
    rendu = services.rendre_enquete_publique(enquete, reponses_partielles)
    return Response(rendu)


@api_view(['POST'])
@permission_classes([AllowAny])
@throttle_classes([_MarketingPublicThrottle])
def enquete_soumettre(request, token):
    """Soumission publique d'une enquête (XMKT27), aucune authentification.

    ZMKT9 — ``debute_le`` (ISO datetime, optionnel) permet de vérifier la
    limite de temps de l'enquête ; sans elle, aucune vérification (illimité,
    comportement actuel)."""
    from django.utils.dateparse import parse_datetime

    enquete = Enquete.objects.filter(token=token, actif=True).first()
    if not enquete:
        return Response({'detail': 'Enquête introuvable.'}, status=404)
    debute_le_brut = request.data.get('debute_le')
    if debute_le_brut:
        debute_le = parse_datetime(debute_le_brut)
        if debute_le and services.limite_temps_depassee(enquete, debute_le=debute_le):
            return Response({'detail': 'Temps limite dépassé.'}, status=400)
    reponses = request.data.get('reponses') or {}
    contact_ref = request.data.get('contact_ref', '')
    try:
        reponse = services.soumettre_reponse_enquete(
            enquete, reponses=reponses, contact_ref=contact_ref,
            nom_repondant=request.data.get('nom_repondant', ''))
    except ValueError as exc:
        return Response({'detail': str(exc)}, status=400)
    return Response({
        'id': reponse.id,
        'score_pct': reponse.score_pct,
        'reussi': reponse.reussi,
        'certificat_genere': reponse.certificat_genere,
    }, status=201)


@api_view(['GET'])
@permission_classes([AllowAny])
@throttle_classes([_MarketingPublicThrottle])
def enquete_certificat_pdf(request, reponse_id):
    """ZMKT10 — téléchargement du certificat PDF (répondant), 404 si non
    certifié/échoué (aucune fuite d'existence)."""
    reponse = ReponseEnquete.objects.filter(id=reponse_id).first()
    if not reponse or not reponse.certificat_genere:
        return Response({'detail': 'Certificat indisponible.'}, status=404)
    pdf_bytes = services.generer_certificat_pdf(reponse)
    if pdf_bytes is None:
        return Response({'detail': 'Certificat indisponible.'}, status=404)
    resp = HttpResponse(pdf_bytes, content_type='application/pdf')
    resp['Content-Disposition'] = 'inline; filename="certificat.pdf"'
    return resp


# ── XFAC26/27 — Portail client self-service : relevé + contestation ───────
# Le client s'identifie par le token du portail EXISTANT
# (``ComptePortailClient.token_acces``, FG228) — jamais une 2ᵉ auth. Les
# données de facturation (relevé, factures) sont lues via
# ``apps.ventes.selectors`` (jamais un import de ``apps.ventes.models``).

class _PortailComptaThrottle(SimpleRateThrottle):
    """Débit du portail compta par IP (même patron que
    ``contrats.public_views.ContratsPortailThrottle`` — sans dépendance
    externe)."""
    scope = 'compta_portail'
    rate = '30/minute'

    def get_rate(self):
        return self.rate

    def get_cache_key(self, request, view):
        ident = self.get_ident(request)
        return self.cache_format % {'scope': self.scope, 'ident': ident}


def _portail_noindex(response):
    response['X-Robots-Tag'] = 'noindex, nofollow, noarchive'
    return response


def _portail_not_found():
    return _portail_noindex(Response(
        {'detail': "Ce lien de portail est invalide ou n'existe pas."},
        status=status.HTTP_404_NOT_FOUND,
    ))


def _resoudre_compte_portail(token):
    return selectors.compte_portail_par_token(token)


@api_view(['GET'])
@permission_classes([AllowAny])
@throttle_classes([_PortailComptaThrottle])
def portail_mon_releve(request, token):
    """XFAC26 — Relevé de compte self-service : postes ouverts, solde
    courant, mini balance âgée (0-30/31-60/61-90/90+).

    GET /api/django/compta/portail/<token>/mon-releve/

    Résout le compte portail par token (404 si invalide/inconnu, sans fuite
    d'existence) puis lit le relevé via ``apps.ventes.selectors`` — jamais un
    import de ``apps.ventes.models``. Le client ne voit JAMAIS le compte
    d'un autre (le sélecteur est borné à ``compte.client_id``)."""
    compte = _resoudre_compte_portail(token)
    if compte is None:
        return _portail_not_found()

    from apps.ventes import selectors as ventes_selectors
    client = compte.client
    if client is None or client.company_id != compte.company_id:
        return _portail_not_found()

    data = ventes_selectors.releve_client_portail(client)
    return _portail_noindex(Response(data))


@api_view(['GET'])
@permission_classes([AllowAny])
@throttle_classes([_PortailComptaThrottle])
def portail_mon_releve_pdf(request, token):
    """XFAC26 — Téléchargement du relevé de compte PDF (même rendu que
    l'écran interne)."""
    compte = _resoudre_compte_portail(token)
    if compte is None:
        return _portail_not_found()

    from apps.ventes import selectors as ventes_selectors
    client = compte.client
    if client is None or client.company_id != compte.company_id:
        return _portail_not_found()

    try:
        pdf_bytes = ventes_selectors.releve_client_pdf_bytes(client)
    except Exception as exc:
        return _portail_noindex(Response(
            {'detail': f'PDF indisponible : {exc}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR))
    resp = HttpResponse(pdf_bytes, content_type='application/pdf')
    resp['Content-Disposition'] = (
        f'inline; filename="Releve_{client.nom}.pdf"')
    return _portail_noindex(resp)


@api_view(['POST'])
@permission_classes([AllowAny])
@throttle_classes([_PortailComptaThrottle])
def portail_contester_facture(request, token, facture_id):
    """XFAC27 — Le client conteste UNE de ses factures depuis le portail.

    POST /api/django/compta/portail/<token>/factures/<facture_id>/contester/
    body: {"motif": "montant"|"prestation"|"deja_payee"|"autre",
           "commentaire": str (optionnel)}

    Crée une ``litiges.Reclamation`` (type FINANCIER, ``bloque_relances``
    True — LITIGE3 suspend automatiquement les relances de cette facture),
    trace la contestation sur le chatter de la facture (``apps.ventes``) et
    notifie best-effort le créateur de la facture. La facture DOIT
    appartenir au client résolu par le token (sinon 404, aucune fuite)."""
    compte = _resoudre_compte_portail(token)
    if compte is None:
        return _portail_not_found()

    from apps.ventes import selectors as ventes_selectors
    from apps.ventes import services as ventes_services

    facture = ventes_selectors.get_facture_scoped(compte.company, facture_id)
    if facture is None or facture.client_id != compte.client_id:
        return _portail_not_found()

    motifs = {
        'montant': 'Montant contesté',
        'prestation': 'Prestation contestée',
        'deja_payee': 'Facture déjà payée',
        'autre': 'Autre motif',
    }
    motif = (request.data.get('motif') or '').strip()
    motif_label = motifs.get(motif, motifs['autre'])
    commentaire = (request.data.get('commentaire') or '').strip()

    reclamation = services.creer_reclamation_portail(
        facture, motif_label=motif_label, commentaire=commentaire)

    ventes_services.enregistrer_contestation_portail(
        facture, motif_label=motif_label, commentaire=commentaire)

    return _portail_noindex(Response(
        {'reclamation_id': reclamation.id, 'reference': reclamation.reference},
        status=status.HTTP_201_CREATED))


# ── FG203 — Récupération des devis abandonnés ──────────────────────────────

class RelanceDevisAbandonneViewSet(_ComptaBaseViewSet):
    """Journal des relances sur devis abandonnés (FG203). Référence le devis
    par id/référence opaques (ventes) — jamais d'import de ses modèles."""
    queryset = RelanceDevisAbandonne.objects.all()
    serializer_class = RelanceDevisAbandonneSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_relance', 'jours_sans_reponse']


# ── FG205 — Tracking d'ouverture des ShareLink ─────────────────────────────

class OuverturePartageViewSet(_ComptaBaseViewSet):
    """Index des ouvertures de liens de partage devis/facture (FG205).
    ``ping`` horodate une ouverture (idempotent, incrémente le compteur)."""
    queryset = OuverturePartage.objects.all()
    serializer_class = OuverturePartageSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['dernier_vu_le', 'nb_ouvertures']

    @action(detail=False, methods=['post'])
    def ping(self, request):
        token = request.data.get('token')
        if not token:
            return Response(
                {'detail': 'token requis.'},
                status=status.HTTP_400_BAD_REQUEST)
        obj = services.enregistrer_ouverture_partage(
            request.user.company,
            token=token,
            cible=request.data.get('cible') or 'devis',
            cible_reference=request.data.get('cible_reference') or '')
        return Response(OuverturePartageSerializer(obj).data)


# ── FG206 — Formulaires d'intake / landing pages ───────────────────────────

class FormulaireIntakeViewSet(_ComptaBaseViewSet):
    """Formulaires d'intake multiples / landing pages (FG206), pré-taguant le
    lead (pompage agricole, régularisation 82-21…)."""
    queryset = FormulaireIntake.objects.all()
    serializer_class = FormulaireIntakeSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['nom', 'slug', 'tag_prefill']
    ordering_fields = ['nom', 'date_creation']


# ── FG207 — Messages WhatsApp entrants (lecture seule) ─────────────────────

class MessageWhatsAppEntrantViewSet(TenantMixin, viewsets.ReadOnlyModelViewSet):
    """Messages WhatsApp entrants capturés (FG207). LECTURE SEULE : la capture
    réelle passe par le webhook gated (NO-OP tant que Meta n'est pas
    provisionné). Admin/Responsable uniquement."""
    permission_classes = [IsResponsableOrAdmin]
    queryset = MessageWhatsAppEntrant.objects.all()
    serializer_class = MessageWhatsAppEntrantSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_reception']


# ── FG208 — Journal d'appels & click-to-call ───────────────────────────────

class AppelTelephoniqueViewSet(_ComptaBaseViewSet):
    """Journal d'appels (FG208). L'auteur est posé côté serveur."""
    queryset = AppelTelephonique.objects.select_related('auteur').all()
    serializer_class = AppelTelephoniqueSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['numero', 'note']
    ordering_fields = ['date_appel', 'duree_secondes']

    def perform_create(self, serializer):
        serializer.save(
            company=self.request.user.company, auteur=self.request.user)


# ── FG209 — Codes de promotion ─────────────────────────────────────────────

class CodePromotionViewSet(_ComptaBaseViewSet):
    """Codes de remise datés applicables au devis (FG209), traçables au ROI."""
    queryset = CodePromotion.objects.all()
    serializer_class = CodePromotionSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['code', 'libelle']
    ordering_fields = ['date_creation', 'date_debut', 'date_fin']


# ── FG210 — Bibliothèque de modèles de devis ───────────────────────────────

class ModeleDevisViewSet(_ComptaBaseViewSet):
    """Modèles de devis réutilisables par marché (FG210)."""
    queryset = ModeleDevis.objects.all()
    serializer_class = ModeleDevisSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['nom', 'description']
    ordering_fields = ['marche', 'nom', 'date_creation']


# ── FG211 — Configurateur d'options guidé ──────────────────────────────────

class SessionGuidedSellingViewSet(_ComptaBaseViewSet):
    """Assistant pas-à-pas guided selling (FG211). ``evaluer`` valide la
    cohérence des réponses et renvoie une composition proposée."""
    queryset = SessionGuidedSelling.objects.select_related('auteur').all()
    serializer_class = SessionGuidedSellingSerializer

    def perform_create(self, serializer):
        serializer.save(
            company=self.request.user.company, auteur=self.request.user)

    @action(detail=True, methods=['post'])
    def evaluer(self, request, pk=None):
        session = self.get_object()
        composition, complet, alertes = (
            services.evaluer_session_guided_selling(session.reponses))
        session.composition = composition
        session.complet = complet
        session.save(update_fields=['composition', 'complet'])
        data = SessionGuidedSellingSerializer(session).data
        data['alertes'] = alertes
        return Response(data)


# ── FG212 — Comparateur de versions de devis ───────────────────────────────

class ComparateurDevisViewSet(viewsets.ViewSet):
    """Comparateur côte-à-côte de deux versions de devis (FG212), LECTURE
    SEULE. Lit les cartes devis via les selectors de ventes (jamais ses
    modèles) et renvoie un diff champ-à-champ pour l'affichage UI."""
    permission_classes = [IsResponsableOrAdmin]

    @action(detail=False, methods=['get'])
    def comparer(self, request):
        company = request.user.company
        a_id = request.query_params.get('a')
        b_id = request.query_params.get('b')
        if not a_id or not b_id:
            return Response(
                {'detail': 'Paramètres a et b (ids de devis) requis.'},
                status=status.HTTP_400_BAD_REQUEST)
        from apps.ventes import selectors as ventes_selectors
        try:
            carte_a = ventes_selectors.devis_card(int(a_id), company)
            carte_b = ventes_selectors.devis_card(int(b_id), company)
        except (ValueError, TypeError):
            return Response(
                {'detail': 'Identifiants de devis invalides.'},
                status=status.HTTP_400_BAD_REQUEST)
        if carte_a is None or carte_b is None:
            return Response(
                {'detail': 'Devis introuvable pour cette société.'},
                status=status.HTTP_404_NOT_FOUND)
        diff = {}
        cles = set(carte_a) | set(carte_b)
        for cle in cles:
            va, vb = carte_a.get(cle), carte_b.get(cle)
            if va != vb:
                diff[cle] = {'a': va, 'b': vb}
        return Response({'a': carte_a, 'b': carte_b, 'diff': diff})


# ── FG213 — Routage d'approbation des configs non-standard ─────────────────

class DemandeApprobationConfigViewSet(_ComptaBaseViewSet):
    """Workflow d'approbation des compositions non-standard (FG213).
    ``approuver`` / ``refuser`` clôturent la demande (idempotent)."""
    queryset = DemandeApprobationConfig.objects.select_related(
        'demandeur', 'decideur').all()
    serializer_class = DemandeApprobationConfigSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_creation', 'statut']

    def perform_create(self, serializer):
        serializer.save(
            company=self.request.user.company, demandeur=self.request.user)

    @action(detail=True, methods=['post'])
    def approuver(self, request, pk=None):
        demande = self.get_object()
        services.decider_approbation_config(
            demande, approuver=True, user=request.user,
            commentaire=request.data.get('commentaire') or '')
        return Response(DemandeApprobationConfigSerializer(demande).data)

    @action(detail=True, methods=['post'])
    def refuser(self, request, pk=None):
        demande = self.get_object()
        services.decider_approbation_config(
            demande, approuver=False, user=request.user,
            commentaire=request.data.get('commentaire') or '')
        return Response(DemandeApprobationConfigSerializer(demande).data)


# ── FG214 — E-catalogue à prix publics (tokenisé) ──────────────────────────

class ECatalogueViewSet(_ComptaBaseViewSet):
    """E-catalogues publics tokenisés (FG214). Le token est généré côté serveur.
    Le rendu public n'expose JAMAIS le prix d'achat ni de marge — uniquement le
    prix public TTC (filtré au rendu)."""
    queryset = ECatalogue.objects.all()
    serializer_class = ECatalogueSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_creation']

    def perform_create(self, serializer):
        import secrets
        serializer.save(
            company=self.request.user.company,
            token=secrets.token_urlsafe(32))


# ── FG215 — Bibliothèque de documents de proposition ───────────────────────

class DocumentPropositionViewSet(_ComptaBaseViewSet):
    """Annexes réutilisables attachables au PDF de proposition (FG215)."""
    queryset = DocumentProposition.objects.all()
    serializer_class = DocumentPropositionSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['titre', 'contenu']
    ordering_fields = ['type_document', 'ordre', 'date_creation']


# ── FG216 — Simulateur public « configurez votre kit » → lead ──────────────

class SimulationPubliqueViewSet(_ComptaBaseViewSet):
    """Simulations publiques de kit → lead pré-rempli (FG216). L'action
    ``creer_lead`` est GATED (NO-OP par défaut) : elle ne crée un lead via le
    service crm que si ``PUBLIC_SIM_LEAD_ENABLED`` est activé."""
    queryset = SimulationPublique.objects.all()
    serializer_class = SimulationPubliqueSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['nom_prospect', 'email', 'telephone']
    ordering_fields = ['date_creation', 'puissance_kwc']

    @action(detail=True, methods=['post'])
    def creer_lead(self, request, pk=None):
        simulation = self.get_object()
        lead_id = services.creer_lead_depuis_simulation(
            simulation, user=request.user)
        data = SimulationPubliqueSerializer(simulation).data
        data['lead_id'] = lead_id
        data['gated'] = not services.leads_depuis_simulation_actif()
        return Response(data)


# ── FG217 — Simulation de financement dans le devis ────────────────────────

class SimulationFinancementViewSet(_ComptaBaseViewSet):
    """Bloc mensualités crédit/leasing rattaché à un devis (FG217). La
    mensualité et le coût total sont calculés côté serveur au save."""
    queryset = SimulationFinancement.objects.all()
    serializer_class = SimulationFinancementSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_creation']

    def perform_create(self, serializer):
        obj = serializer.save(company=self.request.user.company)
        services.recalculer_simulation_financement(obj)
        obj.save(update_fields=['mensualite', 'cout_total_credit'])

    def perform_update(self, serializer):
        obj = serializer.save(company=self.request.user.company)
        services.recalculer_simulation_financement(obj)
        obj.save(update_fields=['mensualite', 'cout_total_credit'])


# ── FG218 — Offres de banques/partenaires de financement ───────────────────

class OffreFinancementViewSet(_ComptaBaseViewSet):
    """Catalogue d'offres de financement sélectionnables sur un devis (FG218)."""
    queryset = OffreFinancement.objects.all()
    serializer_class = OffreFinancementSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['partenaire', 'libelle']
    ordering_fields = ['partenaire', 'taux_annuel', 'date_creation']


# ── FG219 — Ligne d'incitation / subvention ────────────────────────────────

class LigneIncitationViewSet(_ComptaBaseViewSet):
    """Incitations/subventions (Tatwir/MASEN…) déductibles affichées sur un
    devis (FG219). Encart informatif — n'altère pas le total du devis."""
    queryset = LigneIncitation.objects.all()
    serializer_class = LigneIncitationSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_creation']


# ── FG220 — Paiement échelonné (type Tayssir) sur facture ──────────────────

class EcheancierPaiementViewSet(_ComptaBaseViewSet):
    """Échéanciers de tranches sur facture (FG220) avec suivi des versements."""
    queryset = EcheancierPaiement.objects.prefetch_related('tranches').all()
    serializer_class = EcheancierPaiementSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_creation']


class TranchePaiementViewSet(_ComptaBaseViewSet):
    """Tranches d'un échéancier de paiement (FG220). ``regler`` marque une
    tranche payée (idempotent)."""
    queryset = TranchePaiement.objects.all()
    serializer_class = TranchePaiementSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['numero', 'date_echeance']

    @action(detail=True, methods=['post'])
    def regler(self, request, pk=None):
        from django.utils import timezone
        tranche = self.get_object()
        montant = request.data.get('montant')
        tranche.montant_regle = montant if montant is not None else tranche.montant
        tranche.paye = True
        tranche.date_reglement = timezone.now().date()
        tranche.save(update_fields=[
            'montant_regle', 'paye', 'date_reglement'])
        return Response(TranchePaiementSerializer(tranche).data)


# ── FG221 — Comparateur cash vs financement ────────────────────────────────

class ComparateurCashFinancementViewSet(viewsets.ViewSet):
    """Encart comparateur cash vs financement (FG221), LECTURE/CALCUL seul.
    Aucun stockage : renvoie coûts totaux + payback pour lever l'objection
    prix côté client."""
    permission_classes = [IsResponsableOrAdmin]

    @action(detail=False, methods=['get'])
    def comparer(self, request):
        from decimal import Decimal, InvalidOperation
        params = request.query_params
        try:
            montant = Decimal(params.get('montant') or '0')
            duree = int(params.get('duree_mois') or '0')
            taux = Decimal(params.get('taux_annuel') or '0')
            economie = Decimal(params.get('economie_annuelle') or '0')
        except (InvalidOperation, ValueError, TypeError):
            return Response(
                {'detail': 'Paramètres numériques invalides.'},
                status=status.HTTP_400_BAD_REQUEST)
        resultat = services.comparer_cash_vs_financement(
            montant, duree, taux, economie_annuelle=economie)
        return Response(resultat)


# ── FG222 — Gestion des appels d'offres ────────────────────────────────────

class AppelOffreViewSet(_ComptaBaseViewSet):
    """Objets appels d'offres public/privé (FG222)."""
    queryset = AppelOffre.objects.all()
    serializer_class = AppelOffreSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['reference', 'objet', 'acheteur', 'lot']
    ordering_fields = ['date_creation', 'date_limite', 'statut']


# ── FG223 — Bordereau des prix (BOQ) ───────────────────────────────────────

class BordereauPrixViewSet(_ComptaBaseViewSet):
    """Bordereaux des prix (BOQ) d'AO (FG223), séparés du devis client."""
    queryset = BordereauPrix.objects.prefetch_related('lignes').all()
    serializer_class = BordereauPrixSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_creation']


class LigneBordereauViewSet(_ComptaBaseViewSet):
    """Lignes chiffrées d'un BOQ (FG223)."""
    queryset = LigneBordereau.objects.all()
    serializer_class = LigneBordereauSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['numero']


# ── FG224 — Cautions & garanties de soumission ─────────────────────────────

class CautionSoumissionViewSet(_ComptaBaseViewSet):
    """Cautions de soumission (provisoires/définitives) d'AO (FG224)."""
    queryset = CautionSoumission.objects.all()
    serializer_class = CautionSoumissionSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_creation', 'date_echeance', 'statut']


# ── FG225 — Dossier de soumission (pièces administratives) ─────────────────

class DossierSoumissionViewSet(_ComptaBaseViewSet):
    """Dossiers de soumission d'AO (FG225) : checklist des pièces."""
    queryset = DossierSoumission.objects.prefetch_related('pieces').all()
    serializer_class = DossierSoumissionSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_creation']


class PieceSoumissionViewSet(_ComptaBaseViewSet):
    """Pièces administratives d'un dossier de soumission (FG225)."""
    queryset = PieceSoumission.objects.all()
    serializer_class = PieceSoumissionSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['libelle']


# ── FG226 — Échéancier & alertes de deadline d'AO ──────────────────────────

class EcheanceAOViewSet(_ComptaBaseViewSet):
    """Dates clés d'un AO avec rappels (FG226). L'action ``dues`` liste les
    échéances dont le rappel est échu et non traité."""
    queryset = EcheanceAO.objects.all()
    serializer_class = EcheanceAOSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_echeance', 'date_creation']

    @action(detail=False, methods=['get'])
    def dues(self, request):
        dues = services.echeances_ao_dues(request.user.company)
        return Response(EcheanceAOSerializer(dues, many=True).data)


# ── FG227 — Analyse gagné/perdu des appels d'offres ────────────────────────

class ResultatAOViewSet(_ComptaBaseViewSet):
    """Résultats d'AO pour l'analyse gagné/perdu (FG227). L'action ``stats``
    renvoie le taux de réussite consolidé."""
    queryset = ResultatAO.objects.all()
    serializer_class = ResultatAOSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_creation', 'date_resultat']

    @action(detail=False, methods=['get'])
    def stats(self, request):
        return Response(services.taux_reussite_ao(request.user.company))


# ── FG228 — Portail self-service client ────────────────────────────────────

class ComptePortailClientViewSet(_ComptaBaseViewSet):
    """Comptes d'accès au portail self-service client (FG228). Le token est
    généré côté serveur ; le compte se lie au client par id (résolu via le
    service crm) et ne duplique aucune donnée métier (DC32)."""
    queryset = ComptePortailClient.objects.all()
    serializer_class = ComptePortailClientSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_creation']

    def perform_create(self, serializer):
        import secrets
        # DC32 — le client est lié PAR FK ; on vérifie qu'il est bien dans la
        # société de l'utilisateur (jamais un client d'un autre tenant).
        client = serializer.validated_data.get('client')
        company = self.request.user.company
        if client is not None and getattr(
                client, 'company_id', None) != getattr(company, 'id', None):
            raise ValidationError(
                {'client': 'Client inconnu pour cette société.'})
        serializer.save(
            company=company,
            token_acces=secrets.token_urlsafe(32))


class AcceptationDevisPortailViewSet(_ComptaBaseViewSet):
    """Acceptations / e-signatures de devis depuis le portail (FG229). La
    société est posée côté serveur ; l'action ``signer`` horodate la signature
    et capture l'IP (preuve légère, loi 53-05)."""
    queryset = AcceptationDevisPortail.objects.all()
    serializer_class = AcceptationDevisPortailSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_creation', 'signe_le']

    @action(detail=True, methods=['post'])
    def signer(self, request, pk=None):
        acceptation = self.get_object()
        nom = request.data.get('nom_signataire') or None
        ip = request.META.get('REMOTE_ADDR')
        services.signer_acceptation_devis(acceptation, nom=nom, ip=ip)
        return Response(self.get_serializer(acceptation).data)


class PaiementFacturePortailViewSet(_ComptaBaseViewSet):
    """Intentions de paiement en ligne d'une facture depuis le portail (FG230).
    La société est posée côté serveur ; ``initier`` pose une référence (NO-OP
    tant que CMI est OFF) et ``rapprocher`` marque le paiement comme payé
    (rapprochement auto webhook CMI ou manuel virement)."""
    queryset = PaiementFacturePortail.objects.all()
    serializer_class = PaiementFacturePortailSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_creation', 'paye_le']

    def perform_create(self, serializer):
        paiement = serializer.save(company=self.request.user.company)
        services.initier_paiement_facture(paiement)

    @action(detail=True, methods=['post'])
    def rapprocher(self, request, pk=None):
        paiement = self.get_object()
        reference = request.data.get('reference') or None
        services.rapprocher_paiement_facture(paiement, reference=reference)
        return Response(self.get_serializer(paiement).data)


class DocumentClientPortailViewSet(_ComptaBaseViewSet):
    """Documents (factures ONEE…) téléversés par le client depuis le portail
    (FG231). La société est posée côté serveur ; ``marquer_traite`` signale
    qu'un document a été intégré à l'étude."""
    queryset = DocumentClientPortail.objects.all()
    serializer_class = DocumentClientPortailSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_depot']

    @action(detail=True, methods=['post'])
    def marquer_traite(self, request, pk=None):
        doc = self.get_object()
        if not doc.traite:
            doc.traite = True
            doc.save(update_fields=['traite'])
        return Response(self.get_serializer(doc).data)


class JalonChantierPortailViewSet(_ComptaBaseViewSet):
    """Jalons d'avancement de chantier exposés au client (FG232). La société est
    posée côté serveur ; ``marquer_atteint`` avance un jalon (côté interne). Le
    client lit la timeline en lecture-seule côté portail."""
    queryset = JalonChantierPortail.objects.all()
    serializer_class = JalonChantierPortailSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['ordre', 'date_jalon', 'chantier_id']

    @action(detail=True, methods=['post'])
    def marquer_atteint(self, request, pk=None):
        jalon = self.get_object()
        if not jalon.atteint:
            from django.utils import timezone
            jalon.atteint = True
            if not jalon.date_jalon:
                jalon.date_jalon = timezone.localdate()
            jalon.save(update_fields=['atteint', 'date_jalon'])
        return Response(self.get_serializer(jalon).data)


class DemandeTicketPortailViewSet(_ComptaBaseViewSet):
    """Demandes de ticket SAV ouvertes par le client depuis le portail (FG233).
    La société est posée côté serveur ; ``prendre_en_charge`` avance la demande
    et référence le ticket SAV créé (par id — le vrai ticket vit dans l'app sav,
    créé via son service, jamais importée ici)."""
    queryset = DemandeTicketPortail.objects.all()
    serializer_class = DemandeTicketPortailSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_creation']

    @action(detail=True, methods=['post'])
    def prendre_en_charge(self, request, pk=None):
        demande = self.get_object()
        ticket_id = request.data.get('ticket_id')
        if demande.statut == DemandeTicketPortail.Statut.SOUMISE:
            demande.statut = DemandeTicketPortail.Statut.PRISE_EN_CHARGE
            if ticket_id:
                demande.ticket_id = ticket_id
            demande.save(update_fields=['statut', 'ticket_id'])
        return Response(self.get_serializer(demande).data)

    # ── XSAV22 — Déflection KB sur le formulaire d'ouverture de ticket ─────
    # Lit/écrit UNIQUEMENT via ``apps.kb.selectors``/``apps.kb.services``
    # (jamais ``apps.kb.models``, frontière cross-app CLAUDE.md). Actions
    # ``detail=False`` : appelables pendant la SAISIE, avant toute création
    # de ``DemandeTicketPortail``.

    @action(detail=False, methods=['get'], url_path='suggestions-kb',
            permission_classes=[IsResponsableOrAdmin])
    def suggestions_kb(self, request):
        """Articles KB (publiés + ``visible_portail``) suggérés pendant la
        saisie du sujet, avant soumission de la demande."""
        from apps.kb.selectors import suggestions_portail
        texte = request.query_params.get('q', '')
        suggestions = suggestions_portail(request.user.company, texte)
        return Response({'suggestions': suggestions})

    @action(detail=False, methods=['post'], url_path='consulter-article-kb',
            permission_classes=[IsResponsableOrAdmin])
    def consulter_article_kb(self, request):
        """Journalise la consultation d'un article suggéré (déflection) —
        appelée quand le client ouvre/lit une suggestion depuis le
        formulaire, avant (ou sans) soumettre sa demande."""
        from apps.kb.services import enregistrer_consultation_portail
        article_id = request.data.get('article_id')
        if not article_id:
            return Response(
                {'detail': "article_id requis."}, status=400)
        ok = enregistrer_consultation_portail(
            request.user.company, article_id)
        return Response({'enregistre': ok})


class PartenaireViewSet(_ComptaBaseViewSet):
    """Partenaires commerciaux (apporteurs / sous-revendeurs / installateurs,
    FG234/FG237). La société et le token d'accès sont posés côté serveur."""
    queryset = Partenaire.objects.all()
    serializer_class = PartenaireSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['nom', 'date_creation']

    def perform_create(self, serializer):
        import secrets
        serializer.save(
            company=self.request.user.company,
            token_acces=secrets.token_urlsafe(32))

    @action(detail=True, methods=['post'])
    def activer(self, request, pk=None):
        """Active (agrée) un partenaire installateur (FG237) — pose le statut
        ``agree``, l'actif et la date d'activation."""
        from django.utils import timezone
        partenaire = self.get_object()
        partenaire.statut_onboarding = 'agree'
        partenaire.actif = True
        if not partenaire.date_activation:
            partenaire.date_activation = timezone.localdate()
        partenaire.save(update_fields=[
            'statut_onboarding', 'actif', 'date_activation'])
        return Response(self.get_serializer(partenaire).data)


class SoumissionLeadPartenaireViewSet(_ComptaBaseViewSet):
    """Leads soumis par un partenaire via le portail (FG234). La société est
    posée côté serveur ; ``qualifier`` avance la soumission et référence le lead
    créé (par id — le vrai lead vit dans crm, jamais importé ici)."""
    queryset = SoumissionLeadPartenaire.objects.all()
    serializer_class = SoumissionLeadPartenaireSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_soumission']

    @action(detail=True, methods=['post'])
    def qualifier(self, request, pk=None):
        soumission = self.get_object()
        lead_id = request.data.get('lead_id')
        if soumission.statut == SoumissionLeadPartenaire.Statut.SOUMIS:
            soumission.statut = SoumissionLeadPartenaire.Statut.QUALIFIE
            if lead_id:
                soumission.lead_id = lead_id
            soumission.save(update_fields=['statut', 'lead_id'])
        return Response(self.get_serializer(soumission).data)


class CommissionPartenaireViewSet(_ComptaBaseViewSet):
    """Commissions dues aux partenaires (FG235). Le montant est calculé côté
    serveur (base×taux) ; ``marquer_payee`` solde une commission ; ``releve``
    agrège le dû/payé par partenaire (relevé)."""
    queryset = CommissionPartenaire.objects.all()
    serializer_class = CommissionPartenaireSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_creation']

    def perform_create(self, serializer):
        commission = serializer.save(company=self.request.user.company)
        services.enregistrer_commission(commission)
        commission.save(update_fields=['taux', 'montant'])

    def perform_update(self, serializer):
        commission = serializer.save(company=self.request.user.company)
        services.enregistrer_commission(commission)
        commission.save(update_fields=['taux', 'montant'])

    @action(detail=True, methods=['post'])
    def marquer_payee(self, request, pk=None):
        commission = self.get_object()
        if commission.statut == CommissionPartenaire.Statut.DUE:
            from django.utils import timezone
            commission.statut = CommissionPartenaire.Statut.PAYEE
            commission.paye_le = timezone.localdate()
            commission.save(update_fields=['statut', 'paye_le'])
        return Response(self.get_serializer(commission).data)

    @action(detail=False, methods=['get'])
    def releve(self, request):
        """Relevé agrégé par partenaire : dû / payé / total."""
        from django.db.models import Sum
        rows = (
            self.get_queryset()
            .values('partenaire', 'partenaire__nom', 'statut')
            .annotate(total=Sum('montant'))
            .order_by('partenaire')
        )
        releve = {}
        for r in rows:
            pid = r['partenaire']
            entry = releve.setdefault(pid, {
                'partenaire': pid, 'nom': r['partenaire__nom'],
                'due': 0, 'payee': 0, 'total': 0,
            })
            montant = float(r['total'] or 0)
            if r['statut'] == CommissionPartenaire.Statut.DUE:
                entry['due'] += montant
            elif r['statut'] == CommissionPartenaire.Statut.PAYEE:
                entry['payee'] += montant
            if r['statut'] != CommissionPartenaire.Statut.ANNULEE:
                entry['total'] += montant
        return Response(list(releve.values()))


class TerritoireCommercialViewSet(_ComptaBaseViewSet):
    """Territoires / zones commerciales (FG236). La société est posée côté
    serveur ; ``affecter`` résout le territoire (et son commercial) matchant une
    ville donnée, pour l'affectation auto d'un lead."""
    queryset = TerritoireCommercial.objects.all()
    serializer_class = TerritoireCommercialSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['priorite', 'nom']

    @action(detail=False, methods=['get'])
    def affecter(self, request):
        ville = request.query_params.get('ville', '')
        territoire = services.affecter_territoire(
            request.user.company, ville)
        if territoire is None:
            return Response({'territoire': None, 'owner_user_id': None})
        return Response({
            'territoire': territoire.id,
            'nom': territoire.nom,
            'owner_user_id': territoire.owner_user_id,
        })


class EnqueteNPSViewSet(_ComptaBaseViewSet):
    """Enquêtes NPS / satisfaction post-installation (FG238). La société est
    posée côté serveur ; l'envoi réel est gated Brevo (NO-OP par défaut).
    ``repondre`` enregistre la note client ; ``score`` renvoie le NPS
    consolidé."""
    queryset = EnqueteNPS.objects.all()
    serializer_class = EnqueteNPSSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['envoyee_le']

    def perform_create(self, serializer):
        enquete = serializer.save(company=self.request.user.company)
        services.envoyer_enquete_nps(enquete)

    @action(detail=True, methods=['post'])
    def repondre(self, request, pk=None):
        enquete = self.get_object()
        score = request.data.get('score')
        if score is None:
            return Response(
                {'detail': 'score requis (0–10).'},
                status=status.HTTP_400_BAD_REQUEST)
        commentaire = request.data.get('commentaire')
        services.repondre_enquete_nps(
            enquete, score=score, commentaire=commentaire)
        return Response(self.get_serializer(enquete).data)

    @action(detail=False, methods=['get'])
    def score(self, request):
        return Response(services.score_nps(request.user.company))


class AvisClientViewSet(_ComptaBaseViewSet):
    """Avis / témoignages clients + push Google Reviews (FG239). La société est
    posée côté serveur ; ``recevoir`` enregistre la note/témoignage ;
    ``pousser_google`` route l'avis vers Google (NO-OP si l'URL société n'est
    pas configurée — aucune API payante)."""
    queryset = AvisClient.objects.all()
    serializer_class = AvisClientSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_creation']

    @action(detail=True, methods=['post'])
    def recevoir(self, request, pk=None):
        avis = self.get_object()
        if avis.statut == AvisClient.Statut.SOLLICITE:
            note = request.data.get('note')
            temoignage = request.data.get('temoignage')
            if note is not None:
                avis.note = max(1, min(5, int(note)))
            if temoignage is not None:
                avis.temoignage = temoignage
            avis.statut = AvisClient.Statut.RECU
            avis.save(update_fields=['note', 'temoignage', 'statut'])
        return Response(self.get_serializer(avis).data)

    @action(detail=True, methods=['post'])
    def pousser_google(self, request, pk=None):
        avis = self.get_object()
        services.pousser_avis_google(avis)
        return Response(self.get_serializer(avis).data)


class CompteFideliteViewSet(_ComptaBaseViewSet):
    """Comptes de fidélité clients (points + paliers, FG240). La société est
    posée côté serveur ; points/palier sont recalculés côté serveur depuis les
    mouvements. ``crediter`` ajoute des points (parrainage étendu, achat…)."""
    queryset = CompteFidelite.objects.all()
    serializer_class = CompteFideliteSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['points', 'date_creation']

    @action(detail=True, methods=['post'])
    def crediter(self, request, pk=None):
        compte = self.get_object()
        points = request.data.get('points')
        if points is None:
            return Response(
                {'detail': 'points requis.'},
                status=status.HTTP_400_BAD_REQUEST)
        motif = request.data.get('motif', '')
        services.appliquer_mouvement_fidelite(
            compte, points=int(points), motif=motif)
        compte.refresh_from_db()
        return Response(self.get_serializer(compte).data)


class MouvementFideliteViewSet(_ComptaBaseViewSet):
    """Mouvements de points de fidélité (FG240). La création recalcule le solde
    et le palier du compte côté serveur (jamais depuis le corps)."""
    queryset = MouvementFidelite.objects.all()
    serializer_class = MouvementFideliteSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_creation']

    def perform_create(self, serializer):
        # On ne persiste PAS via serializer.save() : le service crée le
        # mouvement ET recalcule le compte de façon atomique.
        data = serializer.validated_data
        mouvement = services.appliquer_mouvement_fidelite(
            data['compte'], points=data['points'],
            motif=data.get('motif', ''))
        serializer.instance = mouvement


class RegleUpsellViewSet(_ComptaBaseViewSet):
    """Règles d'upsell / cross-sell (FG241). La société est posée côté serveur ;
    ``suggestions`` évalue un contexte client (drapeaux) et renvoie les
    suggestions actives déclenchées, triées par priorité."""
    queryset = RegleUpsell.objects.all()
    serializer_class = RegleUpsellSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['priorite', 'date_creation']

    @action(detail=False, methods=['post'])
    def suggestions(self, request):
        contexte = request.data.get('contexte') or {}
        if not isinstance(contexte, dict):
            return Response(
                {'detail': 'contexte doit être un objet de drapeaux.'},
                status=status.HTTP_400_BAD_REQUEST)
        regles = services.suggestions_upsell(request.user.company, contexte)
        return Response(self.get_serializer(regles, many=True).data)


class AbonnementMonitoringViewSet(_ComptaBaseViewSet):
    """Abonnements de monitoring (supervision récurrente, FG244). La société est
    posée côté serveur ; la 1re échéance est calculée à la création ;
    ``renouveler`` avance SEULEMENT l'échéance (YSUBS3 : découplé de la
    facturation) ; ``facturer`` émet la facture standard de la période due
    (YSUBS3) ; ``suspendre`` / ``resilier`` passent par les transitions
    gardées de service (YSUBS4 — jamais un PATCH direct de ``statut``) ;
    ``a_echeance`` liste les abonnements arrivant à échéance."""
    queryset = AbonnementMonitoring.objects.all()
    serializer_class = AbonnementMonitoringSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['prochaine_echeance', 'date_creation']

    def perform_create(self, serializer):
        abonnement = serializer.save(company=self.request.user.company)
        services.renouveler_abonnement_monitoring(abonnement)

    def get_permissions(self):
        # YRBAC13 — ``facturer`` émet une facture réelle (impact GL) et
        # ``suspendre``/``resilier`` sont des transitions gardées service
        # (irréversibles côté abonnement) → ``compta_valider``. ``renouveler``
        # avance seulement l'échéance (routine, découplée de la facturation,
        # YSUBS3) → ``compta_saisir``. ``a_echeance`` (liste) reste
        # IsResponsableOrAdmin, inchangée.
        if self.action == 'a_echeance':
            return [IsResponsableOrAdmin()]
        if self.action == 'renouveler':
            return [HasPermissionOrLegacy('compta_saisir')()]
        if self.action in ('facturer', 'suspendre', 'resilier'):
            return [HasPermissionOrLegacy('compta_valider')()]
        return super().get_permissions()

    @action(detail=True, methods=['post'])
    def renouveler(self, request, pk=None):
        abonnement = self.get_object()
        services.renouveler_abonnement_monitoring(abonnement)
        return Response(self.get_serializer(abonnement).data)

    @action(detail=True, methods=['post'])
    def facturer(self, request, pk=None):
        """YSUBS3 — Émet la facture standard de la période due (garde
        d'idempotence par ``derniere_facturation`` — refuse de re-facturer
        la même période)."""
        abonnement = self.get_object()
        try:
            facture = services.facturer_abonnement_monitoring(
                abonnement, user=request.user)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(
            {'facture_id': facture.id, 'reference': facture.reference,
             'montant_ttc': str(facture.montant_ttc)},
            status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def suspendre(self, request, pk=None):
        abonnement = self.get_object()
        services.suspendre_abonnement_monitoring(abonnement)
        return Response(self.get_serializer(abonnement).data)

    @action(detail=True, methods=['post'])
    def resilier(self, request, pk=None):
        abonnement = self.get_object()
        try:
            services.resilier_abonnement_monitoring(
                abonnement, motif=request.data.get('motif', ''))
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(abonnement).data)

    @action(detail=False, methods=['get'])
    def a_echeance(self, request):
        """Abonnements actifs dont l'échéance tombe dans ``within`` jours."""
        from datetime import timedelta
        from django.utils import timezone
        try:
            within = int(request.query_params.get('within', 30))
        except (TypeError, ValueError):
            within = 30
        today = timezone.localdate()
        limite = today + timedelta(days=max(0, within))
        qs = (self.get_queryset()
              .filter(statut=AbonnementMonitoring.Statut.ACTIF,
                      prochaine_echeance__isnull=False,
                      prochaine_echeance__lte=limite)
              .order_by('prochaine_echeance'))
        page = self.paginate_queryset(qs)
        if page is not None:
            return self.get_paginated_response(
                self.get_serializer(page, many=True).data)
        return Response(self.get_serializer(qs, many=True).data)


# ── COMPTA2 — Mapping document → compte ────────────────────────────────────

class MappingCompteViewSet(_ComptaBaseViewSet):
    """Mappings document→compte de la société (COMPTA2). Filtrable par type de
    clef. Action ``seed`` pour semer les correspondances par défaut."""
    queryset = MappingCompte.objects.select_related('compte').all()
    serializer_class = MappingCompteSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['type_clef', 'clef']

    def get_queryset(self):
        qs = super().get_queryset()
        type_clef = self.request.query_params.get('type_clef')
        if type_clef:
            qs = qs.filter(type_clef=type_clef)
        return qs

    @action(detail=False, methods=['post'])
    def seed(self, request):
        mappings = services.seed_mappings_defaut(request.user.company)
        return Response(
            MappingCompteSerializer(mappings, many=True).data)


# ── XACC4 — Modèles de rapprochement (règles de contrepartie automatique) ──

class ModeleRapprochementViewSet(_ComptaBaseViewSet):
    """Règles de contrepartie automatique pour le rapprochement bancaire
    (XACC4). CRUD company-scopé + action ``appliquer`` (bouton un-clic)."""
    queryset = ModeleRapprochement.objects.select_related(
        'compte_contrepartie').all()
    serializer_class = ModeleRapprochementSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['priorite', 'libelle']

    @action(detail=True, methods=['post'])
    def appliquer(self, request, pk=None):
        """Applique CE modèle à une ligne de relevé (corps : ``ligne_releve``)."""
        modele = self.get_object()  # scopé société par TenantMixin.
        ligne = LigneReleve.objects.filter(
            company=request.user.company,
            id=request.data.get('ligne_releve')).first()
        if ligne is None:
            return Response(
                {'detail': 'Ligne de relevé inconnue.'},
                status=status.HTTP_400_BAD_REQUEST)
        try:
            ecriture = services.appliquer_modele_rapprochement(ligne, modele)
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response(
            {'ecriture_id': ecriture.id, 'reference': ecriture.reference})


# ── XFAC14 — Compensation AR/AP (netting) ──────────────────────────────────

class CompensationViewSet(_ComptaBaseViewSet):
    """Compensation AR/AP pour un tiers à la fois client et fournisseur
    (XFAC14). La création passe par ``services.creer_compensation`` (garde-
    fous de sur-compensation) ; ``valider`` poste l'écriture 4411/3421 +
    enregistre les règlements croisés. Société scopée ; Admin/Responsable."""
    queryset = Compensation.objects.prefetch_related('lignes').all()
    serializer_class = CompensationSerializer
    http_method_names = ['get', 'post', 'head', 'options']

    def create(self, request, *args, **kwargs):
        data = request.data
        try:
            compensation = services.creer_compensation(
                request.user.company,
                client_id=data.get('client_id'),
                fournisseur_id=data.get('fournisseur_id'),
                lignes=data.get('lignes') or [],
                user=request.user,
            )
        except (services.CompensationError, DjangoValidationError) as exc:
            detail = exc.messages[0] if getattr(exc, 'messages', None) else str(exc)
            return Response({'detail': detail}, status=status.HTTP_400_BAD_REQUEST)
        return Response(
            self.get_serializer(compensation).data,
            status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def valider(self, request, pk=None):
        """Valide la compensation : poste l'écriture équilibrée + les
        règlements croisés (idempotent — déjà validée = no-op)."""
        compensation = self.get_object()
        try:
            services.valider_compensation(compensation, user=request.user)
        except (services.CompensationError, DjangoValidationError) as exc:
            detail = exc.messages[0] if getattr(exc, 'messages', None) else str(exc)
            return Response({'detail': detail}, status=status.HTTP_400_BAD_REQUEST)
        return Response(self.get_serializer(compensation).data)


# ── COMPTA3 — Comptes auxiliaires tiers ────────────────────────────────────

class CompteAuxiliaireViewSet(_ComptaBaseViewSet):
    """Comptes auxiliaires clients/fournisseurs (COMPTA3). Le ``code`` est posé
    côté serveur ; création via l'action ``assurer`` (idempotente, validée par
    les sélecteurs crm/stock)."""
    queryset = CompteAuxiliaire.objects.select_related(
        'compte_collectif').all()
    serializer_class = CompteAuxiliaireSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['type_tiers', 'code']

    def get_queryset(self):
        qs = super().get_queryset()
        type_tiers = self.request.query_params.get('type_tiers')
        if type_tiers:
            qs = qs.filter(type_tiers=type_tiers)
        return qs

    @action(detail=False, methods=['post'])
    def assurer(self, request):
        """Crée (ou récupère) l'auxiliaire d'un tiers.

        Corps : ``{'type_tiers': 'client'|'fournisseur', 'tiers_id': <int>}``.
        Le tiers est validé scopé société par les sélecteurs crm/stock.
        """
        company = request.user.company
        type_tiers = request.data.get('type_tiers')
        tiers_id = request.data.get('tiers_id')
        if not tiers_id:
            return Response(
                {'detail': 'tiers_id requis.'},
                status=status.HTTP_400_BAD_REQUEST)
        if type_tiers == CompteAuxiliaire.TypeTiers.CLIENT:
            aux = services.assurer_compte_auxiliaire_client(company, tiers_id)
        elif type_tiers == CompteAuxiliaire.TypeTiers.FOURNISSEUR:
            aux = services.assurer_compte_auxiliaire_fournisseur(
                company, tiers_id)
        else:
            return Response(
                {'detail': "type_tiers doit être 'client' ou 'fournisseur'."},
                status=status.HTTP_400_BAD_REQUEST)
        if aux is None:
            return Response(
                {'detail': 'Tiers introuvable pour cette société.'},
                status=status.HTTP_404_NOT_FOUND)
        return Response(CompteAuxiliaireSerializer(aux).data)


# ── COMPTA10 — Pièces justificatives sur écriture ──────────────────────────

class PieceJustificativeViewSet(_ComptaBaseViewSet):
    """Pièces justificatives attachées aux écritures (COMPTA10). Filtrable par
    écriture. ``ajoute_par`` est posé côté serveur."""
    queryset = PieceJustificative.objects.select_related('ecriture').all()
    serializer_class = PieceJustificativeSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_creation']

    def get_queryset(self):
        qs = super().get_queryset()
        ecriture = self.request.query_params.get('ecriture')
        if ecriture:
            qs = qs.filter(ecriture_id=ecriture)
        return qs

    def perform_create(self, serializer):
        serializer.save(
            company=self.request.user.company,
            ajoute_par=self.request.user)


class PisteAuditComptableViewSet(TenantMixin, viewsets.ReadOnlyModelViewSet):
    """Piste d'audit comptable inaltérable, hash-chaînée (COMPTA39).

    LECTURE SEULE : les maillons se créent par scellement d'écritures via le
    service (jamais par POST direct sur un maillon). Deux actions :
      * ``verifier`` — recalcule toute la chaîne de la société et signale la
        première rupture (altération d'écriture ou de maillon) éventuelle ;
      * ``sceller`` — scelle une écriture (id dans le corps) dans la chaîne,
        idempotent. Admin/Responsable, scopé société.
    """
    queryset = PisteAuditComptable.objects.select_related('ecriture').all()
    serializer_class = PisteAuditComptableSerializer
    permission_classes = [IsResponsableOrAdmin]
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['sequence', 'date_creation']

    @action(detail=False, methods=['get'])
    def verifier(self, request):
        data = services.verifier_integrite_piste(request.user.company)
        return Response(data)

    @action(detail=False, methods=['post'])
    def sceller(self, request):
        ecriture_id = request.data.get('ecriture')
        if not ecriture_id:
            return Response(
                {'detail': "Le champ 'ecriture' est requis."},
                status=status.HTTP_400_BAD_REQUEST)
        ecriture = EcritureComptable.objects.filter(
            company=request.user.company, pk=ecriture_id).first()
        if ecriture is None:
            return Response(
                {'detail': 'Écriture introuvable pour cette société.'},
                status=status.HTTP_404_NOT_FOUND)
        maillon = services.enregistrer_piste_audit(ecriture)
        return Response(
            PisteAuditComptableSerializer(maillon).data,
            status=status.HTTP_201_CREATED)


# ── XACC2 — Import de la balance d'ouverture (reprise des existants) ────────

class BalanceOuvertureViewSet(viewsets.ViewSet):
    """Import guidé de la balance d'ouverture (COMPTA3, migration tooling).

    ``gabarit`` télécharge le fichier modèle CSV ; ``importer`` valide puis
    poste une écriture AN unique équilibrée (rejet détaillé ligne à ligne si
    invalide, idempotent par exercice). Admin/Responsable, scopé société."""
    permission_classes = [IsResponsableOrAdmin]

    @action(detail=False, methods=['get'])
    def gabarit(self, request):
        data = services.gabarit_import_balance_ouverture()
        resp = HttpResponse(data, content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = (
            'attachment; filename="gabarit_balance_ouverture.csv"')
        return resp

    @action(detail=False, methods=['post'],
            parser_classes=[MultiPartParser, FormParser])
    def importer(self, request):
        company = request.user.company
        f = request.FILES.get('file')
        exercice_id = request.data.get('exercice')
        if f is None:
            return Response(
                {'detail': 'Aucun fichier fourni.'},
                status=status.HTTP_400_BAD_REQUEST)
        if not exercice_id:
            return Response(
                {'detail': "Le champ 'exercice' est requis."},
                status=status.HTTP_400_BAD_REQUEST)
        exercice = ExerciceComptable.objects.filter(
            company=company, pk=exercice_id).first()
        if exercice is None:
            return Response(
                {'detail': 'Exercice introuvable pour cette société.'},
                status=status.HTTP_404_NOT_FOUND)
        # ARC13 — lecture déléguée à ``apps.dataimport.parsing.iter_rows``
        # (parseur générique partagé) au lieu d'un ``csv.DictReader`` local ;
        # comportement inchangé (mêmes en-têtes, mêmes clés de lignes).
        from apps.dataimport.parsing import iter_rows
        try:
            _headers, rows = iter_rows(f.read(), f.name)
        except Exception:
            return Response(
                {'detail': 'Fichier illisible (encodage invalide).'},
                status=status.HTTP_400_BAD_REQUEST)
        result = services.importer_balance_ouverture(
            company, rows, exercice=exercice, user=request.user)
        if not result['ok']:
            return Response(
                {'detail': 'Fichier invalide.', 'erreurs': result['erreurs']},
                status=status.HTTP_400_BAD_REQUEST)
        ecriture = result['ecriture']
        return Response({
            'ok': True,
            'deja_importee': result['deja_importee'],
            'ecriture_id': ecriture.id if ecriture else None,
            'reference': ecriture.reference if ecriture else '',
            'total': str(ecriture.total_debit) if ecriture else '0',
        }, status=status.HTTP_201_CREATED)


# ── XACC7 — Provisions FNP / FAE de fin de période ──────────────────────────

class ProvisionsPeriodeViewSet(viewsets.ViewSet):
    """Provisions de fin de période FNP/FAE (XACC7). ``items`` (corps :
    réceptions/avancements non facturés déjà résolus côté appelant) est posté
    en une écriture OD par item + extourne automatique. Admin/Responsable,
    scopé société côté serveur."""
    permission_classes = [IsResponsableOrAdmin]

    @action(detail=False, methods=['post'], url_path='generer-fnp')
    def generer_fnp(self, request):
        data = request.data
        try:
            resultats = services.generer_provisions_fnp(
                request.user.company,
                date_periode=data.get('date_periode'),
                items=data.get('items') or [],
                date_extourne=data.get('date_extourne'),
                user=request.user,
            )
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response({'postees': resultats}, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'], url_path='generer-fae')
    def generer_fae(self, request):
        data = request.data
        try:
            resultats = services.generer_provisions_fae(
                request.user.company,
                date_periode=data.get('date_periode'),
                items=data.get('items') or [],
                date_extourne=data.get('date_extourne'),
                user=request.user,
            )
        except DjangoValidationError as exc:
            return Response(
                {'detail': exc.messages[0] if exc.messages else str(exc)},
                status=status.HTTP_400_BAD_REQUEST)
        return Response({'postees': resultats}, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'])
    def rapport(self, request):
        params = request.query_params
        data = services.rapport_provisions_periode(
            request.user.company,
            date_debut=params.get('date_debut'),
            date_fin=params.get('date_fin'),
            type_provision=params.get('type') or None,
        )
        return Response(data)

    @action(detail=False, methods=['get'], url_path='export-csv')
    def export_csv(self, request):
        params = request.query_params
        data = services.export_provisions_periode_csv(
            request.user.company,
            date_debut=params.get('date_debut'),
            date_fin=params.get('date_fin'),
            type_provision=params.get('type') or None,
        )
        resp = HttpResponse(data, content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = (
            'attachment; filename="provisions_fnp_fae.csv"')
        return resp


# ── XACC9 — Calendrier des obligations fiscales ─────────────────────────────

class ObligationFiscaleViewSet(TenantMixin, viewsets.ReadOnlyModelViewSet):
    """Calendrier des échéances fiscales de l'exercice (XACC9). LECTURE
    SEULE (générées par ``generer``) + action ``rappels`` (J-7). Admin/
    Responsable, scopé société."""
    queryset = ObligationFiscale.objects.all()
    serializer_class = ObligationFiscaleSerializer
    permission_classes = [IsResponsableOrAdmin]
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date_limite', 'id']

    def get_queryset(self):
        qs = super().get_queryset()
        statut = self.request.query_params.get('statut')
        if statut:
            qs = qs.filter(statut=statut)
        return qs

    @action(detail=False, methods=['post'])
    def generer(self, request):
        """Génère (idempotent) le calendrier fiscal d'un exercice (XACC9).

        Corps : ``{exercice: <id>, regime_tva?}``.
        """
        exercice_id = request.data.get('exercice')
        exercice = ExerciceComptable.objects.filter(
            company=request.user.company, pk=exercice_id).first()
        if exercice is None:
            return Response(
                {'detail': 'Exercice introuvable pour cette société.'},
                status=status.HTTP_404_NOT_FOUND)
        obligations = services.generer_calendrier_fiscal(
            request.user.company, exercice,
            regime_tva=request.data.get('regime_tva') or None)
        return Response(
            ObligationFiscaleSerializer(obligations, many=True).data,
            status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'])
    def rappels(self, request):
        """Envoie les rappels J-7 des obligations « à préparer » (XACC9)."""
        notifiees = services.envoyer_rappels_j7(request.user.company)
        return Response(
            ObligationFiscaleSerializer(notifiees, many=True).data)


# ── XACC11 — Référentiel des familles de charge à TVA non déductible ───────

class FamilleTvaNonDeductibleViewSet(_ComptaBaseViewSet):
    """CRUD du référentiel des familles à TVA non déductible (XACC11,
    véhicules de tourisme, missions/réceptions…). Company-scopé."""
    queryset = FamilleTvaNonDeductible.objects.all()
    serializer_class = FamilleTvaNonDeductibleSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['famille']
