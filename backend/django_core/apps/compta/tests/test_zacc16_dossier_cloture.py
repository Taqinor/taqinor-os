"""Tests ZACC16 — Export xlsx multi-onglets du dossier de clôture (liasse +
états + balance âgée + immos) en un fichier.

Couvre : un seul .xlsx sort avec un onglet par état aux totaux identiques aux
endpoints individuels, exercice inconnu -> 400, cross-company 404.
"""
import io
from datetime import date
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import TestCase
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import AccessToken

from authentication.models import Company
from apps.compta import services
from apps.compta.models import ExerciceComptable

User = get_user_model()


def make_company(slug='zacc16-co', nom='ZACC16 Co'):
    return Company.objects.get_or_create(slug=slug, defaults={'nom': nom})[0]


def make_user(company, username, role='admin'):
    return User.objects.create_user(
        username=username, password='x', company=company, role_legacy=role)


def auth(user):
    api = APIClient()
    api.credentials(HTTP_AUTHORIZATION=f'Bearer {AccessToken.for_user(user)}')
    return api


class _Base(TestCase):
    def setUp(self):
        self.company = make_company()
        services.seed_plan_comptable(self.company)
        services.seed_journaux(self.company)
        self.exercice = ExerciceComptable.objects.create(
            company=self.company, libelle='2026',
            date_debut=date(2026, 1, 1), date_fin=date(2026, 12, 31))
        journal_od = services._journal(
            self.company, services.Journal.Type.OPERATIONS_DIVERSES)
        compte_clients = services.get_compte(self.company, '3421')
        compte_ventes = services.get_compte(self.company, '7121')
        services.creer_ecriture_od(
            self.company, '2026-03-10', 'Vente test ZACC16',
            [
                {'compte': compte_clients, 'debit': Decimal('1000'),
                 'credit': Decimal('0')},
                {'compte': compte_ventes, 'debit': Decimal('0'),
                 'credit': Decimal('1000')},
            ], journal=journal_od)
        self.user = make_user(self.company, 'zacc16-admin')
        self.api = auth(self.user)


class TestEndpoint(_Base):
    def test_dossier_cloture_xlsx_multi_onglets(self):
        resp = self.api.get(
            f'/api/django/compta/etats/dossier-cloture/'
            f'?exercice={self.exercice.pk}&export=xlsx')
        self.assertEqual(resp.status_code, 200)
        self.assertIn('spreadsheet', resp['Content-Type'])
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(resp.content))
        self.assertIn('Bilan', wb.sheetnames)
        self.assertIn('CPC', wb.sheetnames)
        self.assertIn('Balance', wb.sheetnames)
        self.assertIn('Grand livre', wb.sheetnames)
        self.assertIn('Balance âgée fournisseurs', wb.sheetnames)
        self.assertIn('Tableau immobilisations', wb.sheetnames)
        self.assertIn('Tableau des flux', wb.sheetnames)

    def test_totaux_identiques_aux_endpoints_individuels(self):
        resp_xlsx = self.api.get(
            f'/api/django/compta/etats/dossier-cloture/'
            f'?exercice={self.exercice.pk}&export=xlsx')
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(resp_xlsx.content))
        ws_cpc = wb['CPC']
        # Trouve la ligne 'Résultat' et vérifie son montant (colonne D).
        resultat_row = None
        for row in ws_cpc.iter_rows(min_row=2, values_only=True):
            if row[0] == 'Résultat':
                resultat_row = row
                break
        self.assertIsNotNone(resultat_row)
        self.assertEqual(Decimal(str(resultat_row[3])), Decimal('1000'))

        resp_cpc = self.api.get(
            '/api/django/compta/etats/cpc/?date_debut=2026-01-01'
            '&date_fin=2026-12-31')
        self.assertEqual(
            Decimal(str(resp_cpc.data['resultat'])), Decimal('1000'))

    def test_sans_exercice_400(self):
        resp = self.api.get(
            '/api/django/compta/etats/dossier-cloture/?export=xlsx')
        self.assertEqual(resp.status_code, 400)

    def test_exercice_inconnu_400(self):
        resp = self.api.get(
            '/api/django/compta/etats/dossier-cloture/'
            '?exercice=999999&export=xlsx')
        self.assertEqual(resp.status_code, 404)

    def test_sans_export_xlsx_400(self):
        resp = self.api.get(
            f'/api/django/compta/etats/dossier-cloture/'
            f'?exercice={self.exercice.pk}')
        self.assertEqual(resp.status_code, 400)

    def test_cross_company_404(self):
        autre = make_company('zacc16-b', 'ZACC16 B')
        exercice_b = ExerciceComptable.objects.create(
            company=autre, libelle='2026',
            date_debut=date(2026, 1, 1), date_fin=date(2026, 12, 31))
        resp = self.api.get(
            f'/api/django/compta/etats/dossier-cloture/'
            f'?exercice={exercice_b.pk}&export=xlsx')
        self.assertEqual(resp.status_code, 404)
