"""Tests du builder .xlsx PARTAGÉ (L879) — apps.records.xlsx.

Garantit que le seul builder produit un classeur cohérent (en-têtes en gras,
coercition fr-MA des valeurs) et que les deux anciens points d'entrée
(crm.exports.build_xlsx_response, dataimport.exporters.export_xlsx) délèguent
bien à ce builder unique.
"""
import datetime
import io
from decimal import Decimal

from django.test import SimpleTestCase

from apps.records.xlsx import (
    XLSX_CONTENT_TYPE, build_workbook, build_xlsx_response, coerce_cell,
    workbook_bytes,
)


class CoerceCellTest(SimpleTestCase):
    def test_none_becomes_empty_string(self):
        self.assertEqual(coerce_cell(None), '')

    def test_decimal_becomes_float(self):
        self.assertEqual(coerce_cell(Decimal('12.50')), 12.5)
        self.assertIsInstance(coerce_cell(Decimal('1')), float)

    def test_date_iso(self):
        self.assertEqual(
            coerce_cell(datetime.date(2026, 6, 19)), '2026-06-19')

    def test_bool_kept(self):
        self.assertIs(coerce_cell(True), True)

    def test_passthrough(self):
        self.assertEqual(coerce_cell('Bennani'), 'Bennani')
        self.assertEqual(coerce_cell(42), 42)


class BuildWorkbookTest(SimpleTestCase):
    def test_header_bold_and_values(self):
        wb = build_workbook(
            ['Référence', 'Total TTC'],
            [['DEV-1', Decimal('1500.00')], ['DEV-2', None]],
            sheet_title='Devis')
        ws = wb.active
        self.assertEqual(ws.title, 'Devis')
        self.assertEqual([c.value for c in ws[1]], ['Référence', 'Total TTC'])
        self.assertTrue(ws[1][0].font.bold)
        self.assertEqual(ws.cell(row=2, column=2).value, 1500.0)
        # None -> cellule vide.
        self.assertEqual(ws.cell(row=3, column=2).value, '')

    def test_sheet_title_truncated_to_31(self):
        wb = build_workbook(['A'], [], sheet_title='x' * 50)
        self.assertEqual(len(wb.active.title), 31)


class WorkbookBytesTest(SimpleTestCase):
    def test_returns_openable_xlsx_bytes(self):
        from openpyxl import load_workbook
        data = workbook_bytes(['Nom'], [['Karim'], ['Sami']], sheet_title='T')
        self.assertTrue(data.startswith(b'PK'))  # zip/xlsx magic
        ws = load_workbook(io.BytesIO(data)).active
        self.assertEqual(ws.cell(row=1, column=1).value, 'Nom')
        self.assertEqual(ws.cell(row=2, column=1).value, 'Karim')


class ResponseTest(SimpleTestCase):
    def test_response_headers(self):
        resp = build_xlsx_response(
            'devis.xlsx', ['Référence'], [['DEV-1']], sheet_title='Devis')
        self.assertEqual(resp['Content-Type'], XLSX_CONTENT_TYPE)
        self.assertIn('devis.xlsx', resp['Content-Disposition'])


class DelegationTest(SimpleTestCase):
    """Les deux anciens points d'entrée pointent vers le builder partagé."""

    def test_crm_exports_reexports_shared(self):
        from apps.crm.exports import build_xlsx_response as crm_builder
        self.assertIs(crm_builder, build_xlsx_response)

    def test_dataimport_export_xlsx_uses_shared(self):
        # export_xlsx(spec, company) délègue à workbook_bytes : un faux spec
        # suffit à prouver le format partagé (en-têtes + lignes coercés).
        from openpyxl import load_workbook
        from apps.dataimport.exporters import export_xlsx

        class _Spec:
            key = 'demo'

            def header(self):
                return ['Réf', 'Montant']

            def rows(self, company):
                return [['A', Decimal('10.00')]]

        data = export_xlsx(_Spec(), company=None)
        ws = load_workbook(io.BytesIO(data)).active
        self.assertEqual([c.value for c in ws[1]], ['Réf', 'Montant'])
        self.assertTrue(ws[1][0].font.bold)  # même mise en forme partagée
        self.assertEqual(ws.cell(row=2, column=2).value, 10.0)
