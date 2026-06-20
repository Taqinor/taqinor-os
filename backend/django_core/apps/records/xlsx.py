"""Constructeur .xlsx PARTAGÉ (L879) — un seul builder pour tous les exports.

Avant, deux constructeurs divergeaient :
  * ``apps.crm.exports.build_xlsx_response`` (réponse HTTP, en-têtes en gras,
    largeurs de colonnes) — utilisé par les exports de listes (devis, factures,
    chantiers, équipements, tickets) et par les exports leads/clients ;
  * ``apps.dataimport.exporters.export_xlsx`` (octets, sans mise en forme) —
    utilisé par l'export/sauvegarde configurable + le bundle ZIP.

Résultat : en-têtes, locale, encodage et coercition des cellules variaient d'un
export à l'autre. Ce module centralise la construction du classeur pour que
CHAQUE export de liste partage le même format fr-MA cohérent (en-têtes en gras,
largeurs lisibles, coercition identique des valeurs).

``openpyxl`` est une dépendance pré-approuvée ; l'import reste local pour ne
charger la lib que lors d'un export.
"""
import datetime
from decimal import Decimal

XLSX_CONTENT_TYPE = (
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
)


def coerce_cell(value):
    """Valeur sûre pour une cellule openpyxl, identique partout (fr-MA).

    - None         -> '' (cellule vide)
    - bool         -> booléen natif
    - date/datetime-> ISO 8601 (texte, locale-stable)
    - Decimal      -> float (Excel n'accepte pas Decimal)
    - int/float/str-> tels quels
    - autre        -> str(value)
    """
    if value is None:
        return ''
    if isinstance(value, bool):
        return value
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (int, float, str)):
        return value
    return str(value)


def build_workbook(headers, rows, sheet_title='Export'):
    """Construit un ``openpyxl.Workbook`` : en-têtes gras + largeurs lisibles.

    headers : itérable de chaînes (1re ligne, en gras).
    rows    : itérable de listes (valeurs brutes ; coercition appliquée ici).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    headers = list(headers)
    wb = Workbook()
    ws = wb.active
    ws.title = (str(sheet_title)[:31] or 'Export')

    ws.append(headers)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold

    # Largeurs lisibles : on borne à la valeur la plus longue par colonne.
    widest = [len(str(h)) for h in headers]
    for row in rows:
        cells = [coerce_cell(v) for v in row]
        ws.append(cells)
        for idx, cell in enumerate(cells):
            if idx < len(widest):
                widest[idx] = max(widest[idx], len(str(cell)))
    for idx in range(len(headers)):
        col = ws.cell(row=1, column=idx + 1).column_letter
        ws.column_dimensions[col].width = min(max(widest[idx] + 2, 10), 50)
    return wb


def workbook_bytes(headers, rows, sheet_title='Export'):
    """Sérialise le classeur partagé en ``bytes`` (export ZIP / non-HTTP)."""
    import io

    wb = build_workbook(headers, rows, sheet_title=sheet_title)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ERR11 — caractères qui, en tête d'une cellule, déclenchent une FORMULE à
# l'ouverture du classeur dans Excel/LibreOffice (injection CSV/Excel).
_RISKY_LEADING = ('=', '+', '-', '@')


def _neutralize_cell(value):
    """Préfixe une apostrophe aux chaînes texte commençant par = + - @.

    La cellule reste lisible (texte) et n'exécute jamais de formule. Ne
    concerne QUE les chaînes — nombres, booléens, dates passent intacts.
    """
    if isinstance(value, str) and value[:1] in _RISKY_LEADING:
        return "'" + value
    return value


def build_xlsx_response(filename, headers, rows, sheet_title='Export'):
    """Réponse HTTP .xlsx construite avec le builder partagé.

    ERR11 — chaque cellule TEXTE commençant par ``= + - @`` est neutralisée
    (apostrophe) pour qu'aucun export TÉLÉCHARGÉ n'exécute de formule à
    l'ouverture. Le chemin octets (``workbook_bytes``, sauvegardes/restaurations)
    n'est volontairement PAS neutralisé, pour préserver les valeurs au
    round-trip (ex. un téléphone « +212… »).
    """
    from django.http import HttpResponse

    safe_rows = [[_neutralize_cell(v) for v in row] for row in rows]
    wb = build_workbook(headers, safe_rows, sheet_title=sheet_title)
    response = HttpResponse(content_type=XLSX_CONTENT_TYPE)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
