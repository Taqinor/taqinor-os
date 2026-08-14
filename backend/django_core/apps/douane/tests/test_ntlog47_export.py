"""NTLOG47 — export CSV/xlsx dossiers-export pour l'expert-comptable (volet
``dossiers-export/`` seulement — ``dossiers-import/`` n'existe pas dans
cette app, NTLOG10 reste BLOCKED, voir ``apps/douane/apps.py``).

Couvre le critère d'acceptation : le fichier exporté contient une colonne
« estimation non contractuelle » explicite en en-tête, réutilise
``apps.records.xlsx.build_xlsx_response`` (aucune nouvelle dépendance —
``openpyxl`` déjà pré-approuvé) pour le xlsx et le module standard ``csv``
pour le CSV (motif ``compta.views.export_fiduciaire``).

Run :
    python manage.py test apps.douane.tests.test_ntlog47_export -v2
"""
import csv
import io
import itertools

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.utils import timezone
from rest_framework import status
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import AccessToken

from apps.douane.models import DossierExport

User = get_user_model()
_seq = itertools.count(1)
BASE = '/api/django/douane'


def make_company():
    from authentication.models import Company
    n = next(_seq)
    company, _ = Company.objects.get_or_create(
        slug=f'ntlog47-co-{n}', defaults={'nom': f'NTLOG47 Co {n}'})
    return company


def auth(user):
    api = APIClient()
    api.credentials(HTTP_AUTHORIZATION=f'Bearer {AccessToken.for_user(user)}')
    return api


def make_user(company, role='responsable'):
    return User.objects.create_user(
        username=f'ntlog47-{next(_seq)}', password='x',
        role_legacy=role, company=company)


class TestExportCsv(TestCase):
    def setUp(self):
        self.company = make_company()
        self.user = make_user(self.company)
        self.api = auth(self.user)
        DossierExport.objects.create(
            company=self.company, numero='EXP-NTLOG47-1',
            statut=DossierExport.Statut.LEVE, incoterm='fob', devise='EUR',
            valeur_marchandise_devise='1500.00', pays_destinataire='France')

    def test_defaut_csv_avec_colonne_estimation_non_contractuelle(self):
        r = self.api.get(f'{BASE}/dossiers-export/export/')
        self.assertEqual(r.status_code, status.HTTP_200_OK, r.content)
        self.assertIn('text/csv', r['Content-Type'])
        content = r.content.decode('utf-8')
        rows = list(csv.reader(io.StringIO(content), delimiter=';'))
        header = rows[0]
        self.assertTrue(
            any('estimation non contractuelle' in h for h in header), header)

    def test_ligne_dossier_presente(self):
        r = self.api.get(f'{BASE}/dossiers-export/export/?format=csv')
        content = r.content.decode('utf-8')
        rows = list(csv.reader(io.StringIO(content), delimiter=';'))
        numeros = [row[0] for row in rows[1:]]
        self.assertIn('EXP-NTLOG47-1', numeros)

    def test_format_invalide_400(self):
        r = self.api.get(f'{BASE}/dossiers-export/export/?format=pdf')
        self.assertEqual(r.status_code, status.HTTP_400_BAD_REQUEST, r.data)

    def test_periode_invalide_400(self):
        r = self.api.get(f'{BASE}/dossiers-export/export/?periode=notadate')
        self.assertEqual(r.status_code, status.HTTP_400_BAD_REQUEST, r.data)

    def test_periode_filtre_mois_courant_inclut(self):
        periode = timezone.now().strftime('%Y-%m')
        r = self.api.get(f'{BASE}/dossiers-export/export/?periode={periode}')
        content = r.content.decode('utf-8')
        self.assertIn('EXP-NTLOG47-1', content)

    def test_periode_filtre_autre_mois_exclut(self):
        r = self.api.get(f'{BASE}/dossiers-export/export/?periode=2019-01')
        content = r.content.decode('utf-8')
        self.assertNotIn('EXP-NTLOG47-1', content)


class TestExportXlsx(TestCase):
    def setUp(self):
        self.company = make_company()
        self.user = make_user(self.company)
        self.api = auth(self.user)
        DossierExport.objects.create(
            company=self.company, numero='EXP-NTLOG47-XL', devise='USD',
            valeur_marchandise_devise='250.00')

    def test_xlsx_contenu_type_et_entetes(self):
        import openpyxl

        r = self.api.get(f'{BASE}/dossiers-export/export/?format=xlsx')
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        self.assertIn('spreadsheetml', r['Content-Type'])
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        self.assertTrue(
            any('estimation non contractuelle' in (h or '') for h in header), header)
        numeros = [row[0].value for row in ws.iter_rows(min_row=2)]
        self.assertIn('EXP-NTLOG47-XL', numeros)


class TestExportIsolationSociete(TestCase):
    def test_dossier_autre_societe_absent(self):
        company_a = make_company()
        company_b = make_company()
        DossierExport.objects.create(company=company_b, numero='EXP-ETRANGER')
        user_a = make_user(company_a)
        api_a = auth(user_a)
        r = api_a.get(f'{BASE}/dossiers-export/export/')
        content = r.content.decode('utf-8')
        self.assertNotIn('EXP-ETRANGER', content)
