"""NTEDU37 (export liste de classe PDF/Excel) et NTEDU38 (trombinoscope)."""
import io
from datetime import date

from django.contrib.auth import get_user_model
from django.test import TestCase
from rest_framework.test import APIClient

from authentication.models import Company

from .models import AnneeScolaire, Classe, Eleve, Famille, Niveau

User = get_user_model()


class ExportsTrombinoscopeTestCaseMixin:
    def setUp(self):
        super().setUp()
        self.company, _ = Company.objects.get_or_create(
            slug='ecole-export-test', defaults={'nom': 'École Export Test'})
        self.user = User.objects.create_user(
            username='admin@ecole-export-test.ma', password='x',
            company=self.company)
        self.client = APIClient()
        self.client.force_authenticate(self.user)

        self.annee = AnneeScolaire.objects.create(
            company=self.company, libelle='2026-2027',
            date_debut=date(2026, 9, 1), date_fin=date(2027, 6, 30))
        self.niveau = Niveau.objects.create(
            company=self.company, nom='CP', cycle=Niveau.Cycle.PRIMAIRE, ordre=1)
        self.classe = Classe.objects.create(
            company=self.company, annee_scolaire=self.annee,
            niveau=self.niveau, nom='CP A', capacite_max=30)
        self.famille = Famille.objects.create(
            company=self.company, nom='Bennani',
            parent1_nom='Karim Bennani', parent1_telephone='+212600000000',
            parent1_email='karim@example.com')
        self.eleve_avec_photo = Eleve.objects.create(
            company=self.company, famille=self.famille, nom='Bennani',
            prenom='Yasmine', classe=self.classe, photo_id=None)
        self.eleve_sans_photo = Eleve.objects.create(
            company=self.company, famille=self.famille, nom='Bennani',
            prenom='Karim Jr', classe=self.classe)


class NTEDU37ExportClasseTests(ExportsTrombinoscopeTestCaseMixin, TestCase):
    def test_export_xlsx_ouvrable_avec_entetes_corrects(self):
        resp = self.client.get(
            f'/api/django/education/classes/{self.classe.id}/export/'
            '?format=xlsx')
        self.assertEqual(resp.status_code, 200)
        self.assertIn('spreadsheetml', resp['Content-Type'])

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        self.assertEqual(headers[0], 'Nom')
        self.assertIn('Parent 1 — téléphone', headers)
        # 2 élèves + 1 ligne d'en-tête.
        self.assertEqual(ws.max_row, 3)

    def test_export_pdf_genere_octets_pdf(self):
        resp = self.client.get(
            f'/api/django/education/classes/{self.classe.id}/export/'
            '?format=pdf')
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp['Content-Type'], 'application/pdf')

    def test_format_invalide_rejete(self):
        resp = self.client.get(
            f'/api/django/education/classes/{self.classe.id}/export/'
            '?format=doc')
        self.assertEqual(resp.status_code, 400)


class NTEDU38TrombinoscopeTests(ExportsTrombinoscopeTestCaseMixin, TestCase):
    def test_eleve_sans_photo_a_photo_url_null(self):
        resp = self.client.get(
            f'/api/django/education/classes/{self.classe.id}/'
            'trombinoscope/')
        self.assertEqual(resp.status_code, 200)
        rows = {r['id']: r for r in resp.data['results']}
        self.assertIsNone(rows[self.eleve_sans_photo.id]['photo_url'])
        self.assertIsNone(rows[self.eleve_avec_photo.id]['photo_url'])

    def test_trombinoscope_liste_tous_les_eleves_de_la_classe(self):
        resp = self.client.get(
            f'/api/django/education/classes/{self.classe.id}/'
            'trombinoscope/')
        self.assertEqual(resp.data['count'], 2)
